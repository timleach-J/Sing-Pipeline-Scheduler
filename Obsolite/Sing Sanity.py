
/
Sing Pipeline
Sing Pipeline
Created by you
·
Private
A pipeline for scheduling, executing and extracting deliverables for a complex project




How can I help you today?


All chats are private unless shared
Project files organization
6 minutes ago
GCP Google Sheets API integration setup for mouse harvest pipeline
7 minutes ago
Sing Scheduling Script
11 minutes ago
Daily reminders
5 days ago
TGS accountability
Jul 29
Scheduler with breeding calc
Jul 24
JCMS climb translation
Jul 16
Automated Google Sheets comparison for logging
Jul 16
Automation strategy and objectives
Jul 13
Journal club discussion
Jun 5
Obsolete - Building the sing scheduler
Jun 4
AI transcription device
May 28
Continuing a Python project
May 12
Python expert skill for Sing Scheduler
May 12
Presentation preparation
May 11
Daily Schedule
May 1
Protocol formatting and reagent list
Apr 28
PFA aliquoting storage best practices
Apr 24
Project improvement suggestions
Apr 21
Building a breeding calculator in Python
Apr 15
Show more
Instructions
Add instructions to tailor Claude’s responses

Memory
Only you
Purpose & context Tim Leach is a Research Assistant at The Jackson Laboratory (JAX) in Bar Harbor, Maine, working in Dr. Vivek Kumar's lab on the SING project (Scalable and Systematic Neurobiology of NPD Risk Genes) — a 5-year NIH-funded study (2024–2029) processing ~114 mouse strains carrying neurodevelopmental disorder risk gene mutations. The project coordinates scheduled mouse harvests, snRNAseq batching, MERFISH, RNA-Seq, and tissue collection across JAX, Penn State (Paul Lab), and NYU Langone. Tim manages colony scheduling, harvest coordination, genotype validation, sample tracking, and pipeline development. He self-describes as a "vibe coder" who works iteratively with AI assistance and is actively building Python tooling to automate manual workflows. His working style is direct and task-focused: he prefers targeted single-file changes, explicit checkpoints before code execution, real-data testing, and strict scope enforcement on AI coding sessions. Key colleagues & contacts: Marina Santos (Marina.Santos@jax.org) — MERFISH, RNA-Seq, Envision tag attachment Sean Deats (Sean.Deats@jax.org) — JAX Single Cell Biology core harvester Tuan Nguyen — data analysis Fionna Kennedy — departed ~June 1; Tim now performs Envision tag attachment (Marina also does it) Alex Berger — Kumar Lab GCP contact (Slack: UGJ9F86TE) Alec (Instem) — Climb LIMS support; OAuth transition deadline November 3, 2026 Camille Klotz — replacement sample payment structure, Climb API Ingvild — NYU collaborator, owns Animal and Sample Tracking sheet (ingvieb@gmail.com) Sophie Adelmann — primary F29 TGS submitter Sandy / Greg Perry — JAX Single Cell core snRNAseq shipments Vijay Roy / Anirban Paul / Madhumita Rajagopalan — Penn State contacts Dan Bancalari (Dan.Bancalari@jax.org) — JAX AI platform chargebacks Dr. Vivek Kumar — PI --- Current state SING Pipeline Scheduler (singpipelinev2): The primary active codebase. Current pushed version is v2.3.0; v2.4.0 is ready to push pending GCP going live. Active pipeline scripts and Google integration scripts live in Z:\kumarlab-new\Tim Leach\Scripts\API Pipeline\. Standalone scripts (Labelgenerator, DeliverablesSheetExport, ClimbtoEnvisionTranslation, SampleCreation) and singcommon.py live in Z:\kumarlab-new\Tim Leach\Scripts\Pipeline Scripts Shared Logic\. Recent major pipeline features completed: SHANK3 split into SHANK3-Het and SHANK3-Hom throughout all outputs Climb API integration (singclimb.py) using animalId (not materialKey) getnextsamplenumber() searches all samples globally (not filtered by study name) MAGEL2 HET2-only scheduling via STRAINGENOTYPERESTRICTIONS; HET1→Wild, HET2→Het globally in canonicalizegenotype; Inconclusive→Wild throughout Wednesday capacity overflow dialog (18→36, user chooses Tuesday/Thursday) Behavior cap system reading Female/Male counts from Summary Sheet CSV (max 9 per sex) P14/P56 Complete! flag support with mismatch warnings Birth-date cohort cap (max 3 animals per birth date per behavior session) Blocking GUI genotype mismatch dialog before scheduling proceeds OFA Testing Sheets auto-generated per behavior Wednesday with correct Computer# color cycling and S4/S3/S2 arena correspondence STRAINDISPLAYNAMES config (e.g., Cacna1g KO displays as Cacna1gKO in Envision, Cacna1g on labels) singgoogle.py created with snapshot→write→verify safeguard infrastructure Genotype validation fixed to compare gene name from Gene<allele> -/+ format against Line (Short) Google Sheets / GCP integration: GCP project jax-kumar-sing-api-01-prod provisioned (JAX ticket RITM0348414). Service account: sing-pipeline@jax-kumar-sing-api-01-prod.iam.gserviceaccount.com. Auth path is blocked pending resolution from Alex Berger — JAX IT policy blocks both service account JSON key download (iam.disableServiceAccountKeyCreation) and gcloud auth application-default login for Sheets/Drive scopes. All terminal commands must run in Command Prompt (cmd), not PowerShell (execution policy block on Tim's Windows machine). Deliverables sheet: Only confirmed-harvest uploads, never forecasts — collaborators (Penn State, NYU) were receiving unfulfilled sample expectations. Deliverables sheet ID not yet configured in singgoogle.py (placeholder). Dravet syndrome colony: Being added to SING. Two sub-colonies: Feeder/Unaffected — JR 034129 × JR 002448 (129 background suppresses seizures) Experimental/Affected — JR 040920 × B6 (B6 background activates seizures) Affected animals begin seizing at P18, high attrition through ~6 weeks — must overproduce; do not schedule harvest until confirmed past 6 weeks. All animals genotyped with JR 034129 assay. Ordering 2F+2M JR 002448 from JAX surplus to refresh aging feeder colony; B6 colony also needs refreshing. Testing Discovery racks (Envision) for seizure monitoring during P18–6wk window. --- On the horizon Resolve GCP auth path with Alex Berger; complete singgoogle.py implementation and push v2.4.0 singgoogle.py planned self-verification loop: (1) write new harvested animals to Harvest Sheet Worksheet tab via API, (2) read Summary Sheet and verify harvest counts (cols 3–14) increased by correct delta against the sheet's COUNTIF totals Climb OAuth transition deadline: November 3, 2026 (current clientcredentials grant may already be compliant — confirm with Alec) Running totals tracker: unscheduled animals from Climb, scheduled from Harvest Sheet, strain targets from Summary tab WT control animals in behavior scheduling: WT littermates from het×WT or het×het crosses — manual selection per Wednesday, not counting toward capacity, flagged as WT Control, not harvested yet (not yet implemented) September 7–11 vacation (Labor Day week); 9/9 harvest covered by Sean (perfusions) and Marina (MERFISH/RNA-Seq) — do not flag as conflict --- Key learnings & principles Output file dating: Every output file must be dated with the date changes were actually made, not the calendar download date. Tim has corrected this rule explicitly and repeatedly — never ask again. Deliverables sheet is confirmed-harvest only: Never write forecasts or expectations to the Deliverables sheet; collaborators act on that data. Genotype classification nuances: Two intentional duplications in the codebase — genotypetosymbol in pipeline (uses canonicalizegenotype) vs. singcommon.py regex version; cleangenotypebase pipeline vs. Envision variants — these differ on purpose, do not sync. TGS monitoring no longer Tim's responsibility as of July 2026. If TGS reports are run, use: toe clip window = P0–P6; toe clips = letter-first IDs (R1, R2, L1, L4); ear notch = single letter or number-first (R, 2L, 2R1L); exclude GET strains + Tom Sproule + Timothy Leach; scheduler run days = Tue+Thu; report actual clip days from data (Mon+Wed is a recommendation only, not assumed). Email/Teams sweep rule: Only flag messages where Tim is directly named or @mentioned — not CC'd threads or broader project threads he isn't responsible for. Interpersonal coordination: Tim prefers to handle in person rather than via drafted emails. Harvest scheduling rules: Perfusion timing calculated per round (max 6 animals/round); P14 and P56 perfusions always separate rounds; Marina's MERFISH/RNA-Seq runs in parallel (timing driven by perfusion rounds only); P14 perfusion must START within 2 hours of leaving the mother; calendar location = "Kumar Fume Hood" (with perfusions) or "Kumar Wet Lab" (without); Marina Santos and Sean Deats are optional attendees on all harvest events. --- Approach & patterns Weekly schedule: Monday: Animal room + colony + scheduling review Tuesday: Protected writing (2hr AM) Wednesday: Paperwork (30min AM), harvest, animal room litter check (2–3PM) Thursday: Comms + meetings + catchup Friday: Animal room litter check (1hr), enter harvest data, protected writing, EOD week-ahead review + 2-week forward calendar build (30min EOD) Friday workflow sequence: (1) Animal room litter check, (2) enter harvest data for the week, (3) protected writing block, (4) EOD: week-ahead review, enter meetings/tasks, build 2-week forward Outlook calendar export as .ics. Daily session routine: (1) Search email for action items where Tim is directly named, (2) search Teams for direct mentions, (3) check Outlook calendar for OOO on harvest dates, (4) read Harvest Sheet and verify all calendar events are accurate — push updated .ics for any changes. Task duration rules: Friday look-ahead = always 30min (never ask) Animal room litter check Wed/Fri = always 1hr (never ask) Wednesday harvest avg done = 1:28PM Enter harvest data = ask each Friday; calculate per-animal avg from Toggl Apply per-unit averaging to all quantifiable tasks If Tim says "always X" — never ask again; otherwise report last budgeted, actual from Toggl, and predicted avg Development approach: Test before pushing to GitHub; keep original pipeline working as fallback while testing new versions in separate folder; one file at a time with real-data testing between each; no unprompted refactoring; Tim picks the next thread. Git workflow: git add → git commit → git tag → git push → git push origin <tag>. Repo: github.com/timleach-J/Sing-Pipeline-Scheduler (private). --- Tools & resources Climb LIMS — primary animal/sample tracking system; API via singclimb.py Google Sheets — Sing Harvest Sheet (1eAYEEgB65wmOpmqtzpbVjNyidhJuxFhefC0aRN89jH8), Animal and Sample Tracking (1VTSB91dpedVsSMwRcNq-uodZ1NikYHZBBjXE0--g0), PSU MERFISH Tracker (1YzzMxKIN16ujyO4j8-PXnsZEUv4dscjISSacRyBzRKY) GCP project: jax-kumar-sing-api-01-prod; service account sing-pipeline@jax-kumar-sing-api-01-prod.iam.gserviceaccount.com Envision — phenotyping/tagging system; output files follow Envision sort order Toggl — time tracking; used each Friday to compare actual vs. planned task durations Outlook / .ics imports — calendar management; Tim imports Claude-generated .ics files with proper METHOD:PUBLISH and SEQUENCE numbering to prevent duplicates Python 3.13 on Windows — all scripts double-clickable; all terminal commands in cmd (not PowerShell) singcommon.py — shared utilities (genotypetosymbol, combinesamplenumbers, naturalsortkey); imported by Labelgenerator, Deliverables, and Envision scripts; SampleCreation.py has own copies — sync manually GitHub — version control; private repo timleach-J/Sing-Pipeline-Scheduler Microsoft 365 (Outlook, Teams) and Google Drive — daily operations and communication

Last updated 2 days ago

Context
10% of project capacity used
Search mode

test_google_connection_20260806.py
90 lines

py



Label_generator_20260806.py
479 lines

py



Deliverables_Sheet_Export_20260806.py
661 lines

py



Climb_to_Envision_Translation_20260806.py
498 lines

py



sing_climb_20260806.py
605 lines

py



Sing_Sanity_20260810.py
657 lines

py



Sample_Creation_20260807.py
554 lines

py



test_google_connection_20260806.py
90 lines

py



sing_google_20260806.py
608 lines

py



sing_pipeline_v2_20260528.py
9,164 lines

py



SING_Task_Inventory.docx
149 lines

docx



SING_Pipeline_Documentation.docx
317 lines

docx



SING_Pipeline_Flowchart.html
357 lines

html


Kumar Lab Sing Perfusion Protocol 2026_05_01.pdf
pdf


Sing_Sanity_20260810.py
"""
compare_trackers.py
───────────────────
Pre-shipment sanity check. Compares the Sing Harvest Sheet against the
three tracker Google Sheets (downloaded as CSV) to verify:
 
  1. Every animal is logged in the correct tracker
  2. Key data fields match (Sex, Line, Birth Date, Harvest Date)
 
TRACKER ROUTING (from the Protocol column in Harvest Sheet):
  Protocol contains "MERFISH"  → MERFISH tracker
  Protocol contains "RNA-Seq"  → RNA-Seq tracker
  Anything else (real sample)  → LSFM/MRI tracker
 
SKIPPED ROWS (not checked):
  Sample Number is: Fail, QC Fail, Extra, Extra NB, or blank
 
HOW TO USE:
  1. Make sure you are connected to the internet
  2. Double-click to run — sheets are downloaded automatically
  3. Open the Comparison_Report_*.xlsx in G:\\My Drive\\SING Sanity Script\
 
Requirements: pip install pandas openpyxl
"""
 
import os
import sys
import glob
from datetime import datetime
 
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
 
# ── Folder to search for CSV files ──────────────────────────────────────────────
# Set this to your Google Drive synced folder path.
# Leave as None to search in the same folder as this script.
SEARCH_FOLDER = r"C:\Users\tjleach\My Drive\Sing Sanity Script"
 
# ── File search keywords (matched against filenames case-insensitively) ─────────
# The script looks for any CSV or XLSX in the same folder whose name contains
# ALL of the keywords listed. Edit these if your filenames change.
# Keys are matched against the end of the filename (before the extension),
# case-insensitively. Edit if your Google Sheet tab names change.
FILE_SUFFIXES = {
    "Harvest":  "sing_harvest_sheet",
    "LSFM/MRI": "animal_and_sample_tracking",
    "MERFISH":  "merfish-rnaseq_sampletracker",
    "RNA-Seq":  "merfish-rnaseq_sampletracker",
}
 
# ── Sample numbers that mean the animal should be skipped ─────────────────────
SKIP_SAMPLE_NUMBERS = {"fail", "qc fail", "extra", "extra nb"}
 
# ── Per-tracker animal IDs to skip (known special cases) ─────────────────────
# Add animal IDs here if they are logged correctly but won't match automatically
# e.g. tubes swapped/mislabelled at harvest
SKIP_ANIMALS = {
    "LSFM/MRI": set(),
    "MERFISH":  set(),
    "RNA-Seq":  {"103498", "103615"},
}
 
# ── Colours ───────────────────────────────────────────────────────────────────
C_HEADER   = "1F4E79"   # dark navy    – column headers
C_MISSING  = "FFB3B3"   # soft red     – missing animals
C_MISMATCH = "FFE699"   # amber        – data mismatch
C_OK       = "C6EFCE"   # soft green   – all good
C_WHITE    = "FFFFFF"
C_SECTION  = "D6E4F0"   # light blue   – section label rows
 
# ── Line name normalisations for tracker comparison ──────────────────────────
# The pipeline now writes SHANK3-Het / SHANK3-Hom in the harvest sheet, but
# the tracker may still have SHANK3. Map the harvest-side name to the
# tracker-friendly name so these don't generate false mismatch flags.
# Update this dict whenever a new strain split is introduced in the pipeline.
LINE_NORMALIZATIONS = {
    "shank3-het": "shank3",
    "shank3-hom": "shank3",
}
 
 
def _norm_line(val: str) -> str:
    """Lowercase and apply LINE_NORMALIZATIONS for tracker comparison."""
    s = str(val).strip().lower()
    return LINE_NORMALIZATIONS.get(s, s)
 
 
COMMON_FIELDS = [
    ("Sex",  "Sex",          "Sex",        False),
    ("Line", "Line (Short)", "Line",       False),
    ("BD",   "Birth Date",   "Birth Date", True),
]
 
# Harvest date column differs per tracker
HARVEST_DATE_COLS = {
    "LSFM/MRI":  "Harvest Date",
    "MERFISH":   "Dissect Date",
    "RNA-Seq":   "Dissect Date",
}
 
 
 
# ── Per-tracker column definitions for report output ─────────────────────────
# (display_header, result_key, mismatch_field_name)
# result_key = None means blank for missing animals
TRACKER_COLS = {
    "LSFM/MRI": [
        ("Status",           "status_icon",  ""),
        ("Name_sample",      "sample",       ""),
        ("Harvest Date",     "harvest_date", "harvest date"),
        ("Age (weeks)",      "age_weeks",    ""),
        ("Name_subject",     "animal",       ""),
        ("Sex",              "sex",          "sex"),
        ("Line_subject",     "line",         "line"),
        ("Line (Short)",     "line_short",   "line"),
        ("Line (Stock)",     "line_stock",   ""),
        ("Species_subject",  "species",      ""),
        ("Genotype",         "genotype",     ""),
        ("Birth Date",       "birth_date",   "birth date"),
        ("Wean Date",        "wean_date",    ""),
        ("Harvest Timepoint","timepoint",    ""),
        ("Perfusion Quality","",             ""),
        ("Lectin Injection", "",             ""),
        ("Ship to",          "",             ""),
        ("Shipped from JAX", "",             ""),
        ("Received at NYU",  "",             ""),
        ("Imaged MRI",       "",             ""),
        ("MRI-related Note", "",             ""),
        ("Rigid aligned/masking","",         ""),
        ("Shipped from NYU", "",             ""),
        ("Received at PSU",  "",             ""),
        ("Dissected",        "",             ""),
        ("Cleared",          "",             ""),
        ("Labelled",         "",             ""),
        ("Imaged LSFM",      "",             ""),
        ("Lectin Quality",   "",             ""),
        ("QC LSFM",          "",             ""),
        ("Other comments",   "",             ""),
        ("Mismatch Details", "mismatch_text",""),
    ],
    "MERFISH": [
        ("Status",           "status_icon",  ""),
        ("Name_sample",      "sample",       ""),
        ("Line (Short)",     "line_short",   "line"),
        ("Age (weeks)",      "age_weeks",    ""),
        ("Sex",              "sex",          "sex"),
        ("Name_subject",     "animal",       ""),
        ("Line_subject",     "line",         "line"),
        ("Line (Stock)",     "line_stock",   ""),
        ("Species_subject",  "species",      ""),
        ("Genotype",         "genotype",     ""),
        ("Birth Date",       "birth_date",   "birth date"),
        ("Wean Date",        "wean_date",    ""),
        ("Dissect Date",     "harvest_date", "harvest date"),
        ("Shipped from JAX", "",             ""),
        ("Received at PSU",  "",             ""),
        ("Box ID",           "",             ""),
        ("Intermediate Box", "",             ""),
        ("RIN score",        "",             ""),
        ("Mounted",          "",             ""),
        ("MERSCOPE Run ID",  "",             ""),
        ("Tissue QC",        "",             ""),
        ("Transcript/Cell QC","",            ""),
        ("Pass/Fail",        "",             ""),
        ("Request for additional samples","",""),
        ("Remarks",          "",             ""),
        ("Mismatch Details", "mismatch_text",""),
    ],
    "RNA-Seq": [
        ("Status",           "status_icon",  ""),
        ("Name_sample",      "sample",       ""),
        ("Age (weeks)",      "age_weeks",    ""),
        ("Name_subject",     "animal",       ""),
        ("Sex",              "sex",          "sex"),
        ("Line_subject",     "line",         "line"),
        ("Line (Short)",     "line_short",   "line"),
        ("Line (Stock)",     "line_stock",   ""),
        ("Species_subject",  "species",      ""),
        ("Genotype",         "genotype",     ""),
        ("Birth Date",       "birth_date",   "birth date"),
        ("Wean Date",        "wean_date",    ""),
        ("Dissect Date",     "harvest_date", "harvest date"),
        ("Shipped from JAX Maine","",        ""),
        ("Received at JAX GenomeCore","",    ""),
        ("SamplePrep Step",  "",             ""),
        ("RIN score",        "",             ""),
        ("Amplification",    "",             ""),
        ("Run ID",           "",             ""),
        ("Run QC",           "",             ""),
        ("Remarks",          "",             ""),
        ("Mismatch Details", "mismatch_text",""),
    ],
}
 
# ── Styling helpers ───────────────────────────────────────────────────────────
 
def _border():
    s = Side(style="thin")
    return Border(left=s, right=s, top=s, bottom=s)
 
 
def _hdr(ws, row, col, text, font_size=11):
    c = ws.cell(row=row, column=col, value=text)
    c.font      = Font(bold=True, color=C_WHITE, size=font_size, name="Arial")
    c.fill      = PatternFill(start_color=C_HEADER, end_color=C_HEADER, fill_type="solid")
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border    = _border()
    return c
 
 
def _cell(ws, row, col, text, bg=C_WHITE, bold=False, wrap=False):
    c = ws.cell(row=row, column=col, value=str(text) if text is not None else "")
    c.font      = Font(name="Arial", size=10, bold=bold)
    c.fill      = PatternFill(start_color=bg, end_color=bg, fill_type="solid")
    c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=wrap)
    c.border    = _border()
    return c
 
 
def _autofit(ws, min_w=10, max_w=45):
    from openpyxl.cell.cell import MergedCell
    for col in ws.columns:
        best = max((len(str(c.value)) if c.value and not isinstance(c, MergedCell) else 0) for c in col)
        first = next((c for c in col if not isinstance(c, MergedCell)), None)
        if first:
            ws.column_dimensions[first.column_letter].width = max(min_w, min(best + 3, max_w))
 
 
# ── Date normalisation ────────────────────────────────────────────────────────
 
def _norm_date(val) -> str:
    """Parse a date and return M/D/YYYY with no leading zeroes, for consistent comparison."""
    if val is None or (isinstance(val, float) and val != val):
        return ""
    s = str(val).strip()
    if not s or s.lower() in ("nan", "none", ""):
        return ""
    for fmt in ("%m/%d/%Y", "%m/%d/%y", "%Y-%m-%d", "%d/%m/%Y", "%-m/%-d/%Y"):
        try:
            dt = datetime.strptime(s, fmt)
            # Format without leading zeroes, cross-platform
            return f"{dt.month}/{dt.day}/{dt.year}"
        except ValueError:
            pass
    # Fallback: strip leading zeroes from each part if slash-separated
    parts = s.split("/")
    if len(parts) == 3:
        try:
            return f"{int(parts[0])}/{int(parts[1])}/{int(parts[2])}"
        except ValueError:
            pass
    return s   # return as-is if completely unparseable
 
 
def _norm_str(val) -> str:
    return str(val).strip().lower() if val is not None else ""
 
 
# ── Protocol → tracker routing ────────────────────────────────────────────────
 
def route_protocol(protocol: str) -> str:
    """Return 'MERFISH', 'RNA-Seq', or 'LSFM/MRI'.
    Handles both 'RNA-Seq' (tracker style) and 'RNAseq' (pipeline harvest type).
    """
    p = str(protocol).strip()
    pl = p.lower()
    if "rna-seq" in pl or "rnaseq" in pl:
        return "RNA-Seq"
    if "merfish" in pl:
        return "MERFISH"
    return "LSFM/MRI"
 
 
def should_skip(sample_number: str) -> bool:
    return sample_number.strip().lower() in SKIP_SAMPLE_NUMBERS
 
 
# ── Load CSVs or XLSX ────────────────────────────────────────────────────────
 
def find_file(folder: str, suffix: str) -> str:
    """Find a CSV or XLSX whose name ends with the given suffix (before extension).
    Ignores spaces, underscores, and case when matching."""
    def normalise(s):
        return s.lower().replace(" ", "").replace("_", "").replace("-", "")
    norm_suffix = normalise(suffix)
    for fname in os.listdir(folder):
        base, ext = os.path.splitext(fname)
        if ext.lower() not in (".csv", ".xlsx"):
            continue
        if normalise(base).endswith(norm_suffix):
            return os.path.join(folder, fname)
    candidates = [f for f in os.listdir(folder) if f.endswith((".csv", ".xlsx"))]
    raise FileNotFoundError(
        f"Could not find a file ending with '{suffix}' in:\n  {folder}\n"
        f"Files present: {candidates}"
    )
 
 
# ── Specific sheet/tab to read from multi-tab Excel files ────────────────────
EXCEL_SHEET_NAMES = {
    "Harvest":  "Harvest Worksheet",
    "LSFM/MRI": "LSFM-MRI",
    "MERFISH":  "MERFISH",
    "RNA-Seq":  "RNA-Seq",
}
 
def load_file(folder: str, suffix: str, tracker_key: str = "") -> pd.DataFrame:
    path = find_file(folder, suffix)
    print(f"    Found: {os.path.basename(path)}")
    if path.lower().endswith(".xlsx"):
        sheet_name = EXCEL_SHEET_NAMES.get(tracker_key, 0)
        df = pd.read_excel(path, sheet_name=sheet_name, dtype=str).fillna("")
    else:
        df = pd.read_csv(path, dtype=str).fillna("")
    df = df.apply(lambda col: col.str.strip() if col.dtype == object else col)
    return df
 
 
# ── Build lookup: animal_id → list of rows, for each tracker ─────────────────
 
def build_tracker_index(df: pd.DataFrame) -> dict:
    """Index by Name_subject (animal ID)."""
    index = {}
    for _, row in df.iterrows():
        key = str(row.get("Name_subject", "")).strip()
        if key:
            index.setdefault(key, []).append(row)
    return index
 
 
def _p_days_to_weeks(val: str) -> str:
    """Convert P-prefixed day string (e.g. 'P56') to whole weeks (e.g. '8')."""
    if not val:
        return ""
    import re
    m = re.match(r'P(\d+)', val.strip(), re.IGNORECASE)
    if m:
        return str(round(int(m.group(1)) / 7))
    return val  # return as-is if not P-format
 
 
# ── Compare one harvest row against a tracker ─────────────────────────────────
 
def compare_row(h_row, tracker_index: dict, tracker_name: str) -> dict:
    """
    Returns a result dict:
        status:   'ok' | 'missing' | 'mismatch'
        animal:   str
        sample:   str
        line:     str
        sex:      str
        harvest_date: str
        mismatches: list of (field, harvest_val, tracker_val)
        tracker:  str
    """
    animal_id_raw = str(h_row["Name"]).strip()
    animal_id     = animal_id_raw.upper().replace(" NB", "").strip()  # strip NB suffix for lookup
    sample_num = str(h_row["Sample Number"]).strip()
    harvest_date_col = HARVEST_DATE_COLS[tracker_name]
 
    result = {
        "status":       "missing",
        "animal":       animal_id_raw,  # show original name (with NB) in report
        "sample":       sample_num,
        "line":         str(h_row.get("Line", "")).strip(),
        "line_short":   str(h_row.get("Line", "")).strip(),
        "line_stock":   str(h_row.get("Line (Stock)", "")).strip(),
        "sex":          str(h_row.get("Sex", "")).strip(),
        "harvest_date": _norm_date(h_row.get("Harvest Date", "")),
        "birth_date":   _norm_date(h_row.get("BD", "")),
        "wean_date":    _norm_date(h_row.get("Wean Date", "")),
        "age_weeks":    _p_days_to_weeks(str(h_row.get("Age (Days)", "")).strip()),
        "species":      "Mouse",
        "genotype":     str(h_row.get("Genotype", "")).strip(),
        "timepoint":    str(h_row.get("Age (Days)", "")).strip(),
        "protocol":     str(h_row.get("Protocol", "")).strip(),
        "mismatches":   [],
        "mismatch_text": "",
        "tracker":      tracker_name,
    }
 
    if animal_id not in tracker_index:
        return result   # missing
 
    tracker_rows = tracker_index[animal_id]
    result["status"] = "ok"
    mismatches = []
 
    # Pick the first tracker row for this animal (could be multiple samples)
    t_row = tracker_rows[0]
 
    # Check common fields
    for h_col, t_col, label, is_date in COMMON_FIELDS:
        if is_date:
            norm = _norm_date
        elif label == "Line":
            norm = _norm_line   # use line normalisation for strain name comparison
        else:
            norm = _norm_str
        h_val = norm(h_row.get(h_col, ""))
        t_val = norm(t_row.get(t_col, ""))
        if h_val and t_val and h_val != t_val:
            mismatches.append((label,
                               str(h_row.get(h_col, "")).strip(),
                               str(t_row.get(t_col, "")).strip()))
 
    # Check harvest date
    h_date = _norm_date(h_row.get("Harvest Date", ""))
    t_date = _norm_date(t_row.get(harvest_date_col, ""))
    if h_date and t_date and h_date != t_date:
        mismatches.append(("Harvest Date", h_date, t_date))
 
    if mismatches:
        result["status"]     = "mismatch"
        result["mismatches"] = mismatches
 
    return result
 
 
# ── Main comparison ───────────────────────────────────────────────────────────
 
def run_comparison(folder: str) -> dict:
    print("\nLoading files...")
    harvest_df = load_file(folder, FILE_SUFFIXES["Harvest"],  "Harvest")
    lsfm_df    = load_file(folder, FILE_SUFFIXES["LSFM/MRI"], "LSFM/MRI")
    merfish_df = load_file(folder, FILE_SUFFIXES["MERFISH"],  "MERFISH")
    rnaseq_df  = load_file(folder, FILE_SUFFIXES["RNA-Seq"],  "RNA-Seq")
 
    trackers = {
        "LSFM/MRI": build_tracker_index(lsfm_df),
        "MERFISH":  build_tracker_index(merfish_df),
        "RNA-Seq":  build_tracker_index(rnaseq_df),
    }
 
    results = {"LSFM/MRI": [], "MERFISH": [], "RNA-Seq": []}
    skipped = []
 
    print("Comparing rows...")
    for _, h_row in harvest_df.iterrows():
        sample_num = str(h_row.get("Sample Number", "")).strip()
 
        protocol = str(h_row.get("Protocol", "")).strip().lower()
        tracker_name_pre = route_protocol(str(h_row.get("Protocol", "")))
        animal_pre = str(h_row.get("Name", "")).strip().upper().replace(" NB", "").strip()
        if (should_skip(sample_num)
                or "found dead" in protocol
                or animal_pre in SKIP_ANIMALS.get(tracker_name_pre, set())):
            skipped.append(str(h_row.get("Name", "")).strip())
            continue
 
        tracker_name = route_protocol(str(h_row.get("Protocol", "")))
        result = compare_row(h_row, trackers[tracker_name], tracker_name)
        results[tracker_name].append(result)
 
    print(f"  Skipped {len(skipped)} rows (Fail/Extra/QC Fail)")
    for name, rows in results.items():
        missing  = sum(1 for r in rows if r["status"] == "missing")
        mismatch = sum(1 for r in rows if r["status"] == "mismatch")
        ok       = sum(1 for r in rows if r["status"] == "ok")
        print(f"  {name}: {ok} OK, {missing} missing, {mismatch} mismatch")
 
    return results, skipped
 
 
# ── Write report ──────────────────────────────────────────────────────────────
 
def write_report(results: dict, skipped: list, folder: str) -> str:
    timestamp   = datetime.now().strftime("%Y%m%d_%H%M%S")
    report_path = os.path.join(folder, f"Comparison_Report_{timestamp}.xlsx")
    wb = Workbook()
 
    # ── Summary sheet ─────────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Summary"
    ws.row_dimensions[1].height = 14
 
    # Title
    title = ws.cell(row=1, column=1, value="SING Pre-Shipment Tracker Comparison")
    title.font      = Font(bold=True, size=14, name="Arial")
    title.alignment = Alignment(horizontal="left")
    ws.merge_cells("A1:G1")
 
    subtitle = ws.cell(row=2, column=1,
                       value=f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}   "
                             f"| Harvest rows skipped (Fail/Extra): {len(skipped)}")
    subtitle.font      = Font(italic=True, size=10, name="Arial", color="555555")
    subtitle.alignment = Alignment(horizontal="left")
    ws.merge_cells("A2:G2")
 
    # Header row
    headers = ["Tracker", "Total Checked", "✅ OK", "⚠ Mismatch", "❌ Missing", "Status", "Detail Sheet"]
    for col, h in enumerate(headers, 1):
        _hdr(ws, 4, col, h)
 
    row = 5
    for tracker_name, rows in results.items():
        total    = len(rows)
        ok       = sum(1 for r in rows if r["status"] == "ok")
        mismatch = sum(1 for r in rows if r["status"] == "mismatch")
        missing  = sum(1 for r in rows if r["status"] == "missing")
 
        if missing > 0:
            status_text = f"❌ {missing} MISSING"
            bg = C_MISSING
        elif mismatch > 0:
            status_text = f"⚠ {mismatch} MISMATCH"
            bg = C_MISMATCH
        else:
            status_text = "✅ All logged & matched"
            bg = C_OK
 
        _cell(ws, row, 1, tracker_name,  bg, bold=True)
        _cell(ws, row, 2, total,         bg)
        _cell(ws, row, 3, ok,            bg)
        _cell(ws, row, 4, mismatch,      bg)
        _cell(ws, row, 5, missing,       bg)
        _cell(ws, row, 6, status_text,   bg, bold=True)
        _cell(ws, row, 7, tracker_name,  bg)  # matches detail sheet tab name
        row += 1
 
    _autofit(ws)
 
    # ── Detail sheet per tracker ───────────────────────────────────────────────
    for tracker_name, rows in results.items():
        ws_d = wb.create_sheet(title=tracker_name.replace("/", "-"))
 
        # Get column definitions for this tracker
        tab_cols = TRACKER_COLS.get(tracker_name, TRACKER_COLS["LSFM/MRI"])
        n_cols = len(tab_cols)
 
        # Column headers
        for col_i, (header, _, _) in enumerate(tab_cols, 1):
            _hdr(ws_d, 1, col_i, header)
 
        # Only show problems — skip OK rows
        order = {"missing": 0, "mismatch": 1}
        sorted_rows = sorted(
            [r for r in rows if r["status"] != "ok"],
            key=lambda r: (order[r["status"]], r["animal"])
        )
 
        current_section = None
        data_row = 2
 
        if not sorted_rows:
            no_issues = ws_d.cell(row=data_row, column=1, value="✅ No issues — all animals logged and data matched.")
            no_issues.font = Font(bold=True, size=11, name="Arial", color="375623")
            no_issues.fill = PatternFill(start_color=C_OK, end_color=C_OK, fill_type="solid")
            no_issues.alignment = Alignment(horizontal="left")
            ws_d.merge_cells(f"A{data_row}:A{data_row}")
 
        for r in sorted_rows:
            section = r["status"]
 
            # Section label row
            if section != current_section:
                current_section = section
                labels = {"missing": "❌ MISSING", "mismatch": "⚠ DATA MISMATCH"}
                sec_cell = ws_d.cell(row=data_row, column=1, value=labels[section])
                sec_cell.font      = Font(bold=True, size=11, name="Arial")
                sec_cell.fill      = PatternFill(start_color=C_SECTION, end_color=C_SECTION, fill_type="solid")
                sec_cell.alignment = Alignment(horizontal="left")
                if n_cols > 1:
                    ws_d.merge_cells(f"A{data_row}:{get_column_letter(n_cols)}{data_row}")
                data_row += 1
 
            bg = {"missing": C_MISSING, "mismatch": C_MISMATCH}[section]
            status_icons = {"missing": "❌ Missing", "mismatch": "⚠ Mismatch"}
            mismatched_fields = {f.lower() for f, _, _ in r["mismatches"]}
 
            # Build mismatch detail text and store on r
            if r["mismatches"]:
                parts = [f"{f}: harvest='{hv}' → tracker='{tv}'"
                         for f, hv, tv in r["mismatches"]]
                r["mismatch_text"] = "  |  ".join(parts)
            else:
                r["mismatch_text"] = ""
 
            # Resolve status_icon
            r["status_icon"] = status_icons[section]
 
            for col_i, (header, key, mismatch_field) in enumerate(tab_cols, 1):
                if key == "status_icon":
                    val = r["status_icon"]
                    _cell(ws_d, data_row, col_i, val, bg, bold=True)
                elif key == "mismatch_text":
                    _cell(ws_d, data_row, col_i, r.get("mismatch_text", ""), bg, wrap=True)
                elif key == "":
                    _cell(ws_d, data_row, col_i, "", bg)
                else:
                    val = r.get(key, "")
                    is_mm = mismatch_field and mismatch_field.lower() in mismatched_fields
                    if is_mm:
                        cell_bg = "FF4C4C"
                        c = ws_d.cell(row=data_row, column=col_i, value=str(val) if val is not None else "")
                        c.font      = Font(name="Arial", size=10, bold=True, color=C_WHITE)
                        c.fill      = PatternFill(start_color=cell_bg, end_color=cell_bg, fill_type="solid")
                        c.alignment = Alignment(horizontal="left", vertical="center")
                        c.border    = _border()
                    else:
                        _cell(ws_d, data_row, col_i, val, bg)
 
            data_row += 1
 
        _autofit(ws_d)
        ws_d.row_dimensions[1].height = 30   # header row taller
 
    wb.save(report_path)
    return report_path
 
 
# ── Entry point ───────────────────────────────────────────────────────────────
 
def main():
    print("=" * 60)
    print("  SING Pre-Shipment Tracker Comparison")
    print("=" * 60)
 
    folder = SEARCH_FOLDER if SEARCH_FOLDER else os.path.dirname(os.path.abspath(__file__))
 
    try:
        results, skipped = run_comparison(folder)
        print("\nWriting report...")
        report_path = write_report(results, skipped, folder)
 
        print("\n" + "=" * 60)
        print(f"  Report: {os.path.basename(report_path)}")
        print("=" * 60)
 
        any_issues = False
        for tracker_name, rows in results.items():
            missing  = sum(1 for r in rows if r["status"] == "missing")
            mismatch = sum(1 for r in rows if r["status"] == "mismatch")
            if missing:
                print(f"  ❌ {tracker_name}: {missing} animals MISSING from tracker")
                any_issues = True
            if mismatch:
                print(f"  ⚠  {tracker_name}: {mismatch} animals have DATA MISMATCHES")
                any_issues = True
        if not any_issues:
            print("  ✅ All animals logged and data matches across all trackers!")
 
    except FileNotFoundError as e:
        print(f"\n❌  {e}")
    except Exception as e:
        print(f"\n❌  Unexpected error: {e}")
        raise
 
    print()
    input("Press Enter to close...")
 
 
if __name__ == "__main__":
    main()
 
