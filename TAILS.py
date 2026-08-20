# =============================================================================
# TAILS — Tracking Animal Inventory, Logging Shipments
# (formerly the SING Pipeline Scheduler)
# Version: 2.0.0
# Updated: 2026-03-24
#
# Changes from v1.6:
#   - Narrowed warnings.filterwarnings — no longer silences all warnings globally
#   - Fixed bare except clauses in auto_size_columns and _run_again queue clearing
#   - Removed redundant `date as date_type` alias — all hints now use `date`
#   - Removed unused `timezone` import
#   - Removed top-level `import unittest` (lazy-loaded in test block only)
#   - Added `Any` to typing imports
#   - Documented unused `sex` param in get_strain_breeding_type
#   - Replaced convoluted argsort sort with sort_values(key=...) in
#     build_births_sexing_schedule
#   - Simplified process_large_dataset — removed unnecessary line-count pre-pass
#   - Vectorized filter_animals_by_use excluded record construction (no iterrows)
#   - Added docstrings to to_date, validate_config_advanced, auto_size_columns,
#     get_strain_breeding_type, process_large_dataset, filter_animals_by_use
#   - Replaced len(df) == 0 checks with df.empty throughout (~30 sites)
# =============================================================================
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Side
import sys
import traceback
import glob
from datetime import datetime, timedelta, date
import warnings
import os
import re
import copy
from collections import defaultdict
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font
import unicodedata
import logging
from logging.handlers import RotatingFileHandler
from typing import Any, Dict, List, Optional, Tuple, Union

# ── Support files live in lib\ ────────────────────────────────────────────────
# sing_common.py, sing_climb*.py and sing_credentials.json belong in a "lib"
# subfolder. The script folder itself is kept on the path as a fallback so a
# flattened copy (everything in one folder) still runs.
_SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
_LIB_DIR    = os.path.join(_SCRIPT_DIR, 'lib')

# Bumped on every change — printed at startup so you can confirm which build ran.
PIPELINE_VERSION = '2.9.4  (2026-08-20)'
for _d in (_LIB_DIR, _SCRIPT_DIR):
    if os.path.isdir(_d) and _d not in sys.path:
        sys.path.insert(0, _d)

from sing_common import combine_sample_numbers, natural_sort_key

warnings.filterwarnings('ignore', category=pd.errors.PerformanceWarning)
warnings.filterwarnings('ignore', message='.*chained assignment.*')
warnings.filterwarnings('ignore', message='.*DataFrame.groupby.*')

# Try to import tqdm for progress bars
try:
    from tqdm import tqdm
    HAS_TQDM = True
except ImportError:
    HAS_TQDM = False
    class tqdm:
        def __init__(self, iterable=None, total=None, desc=None, **kwargs):
            self.iterable = iterable
            self.total = total or (len(iterable) if hasattr(iterable, '__len__') else None)
            self.desc = desc
            self.current = 0
            if desc and self.total:
                print(f"{desc}... (0/{self.total})")

        def __iter__(self):
            for item in self.iterable:
                self.current += 1
                if self.desc and self.total and self.total > 0 and self.current % max(1, self.total // 10) == 0:
                    print(f"{self.desc}... ({self.current}/{self.total})")
                yield item

        def __enter__(self):
            return self

        def __exit__(self, *args):
            if self.desc:
                print(f"{self.desc}... Complete!")

        def update(self, n=1):
            self.current += n

# ============================================================================
# CONFIGURATION
# ============================================================================
CONFIG = {
    # Before scheduling, upload any TGS genotype calls sitting in the
    # 'genotypes' subfolder. Animals that already have that assay are skipped.
    # Reports named in a tgs_genotype_upload_*.json receipt are skipped too —
    # that receipt is the record of what has already been uploaded.
    'UPLOAD_TGS_GENOTYPES': True,

    # After scheduling, move Wild and Inconclusive animals out of the Sing pool
    # to 'Available' — they can never be scheduled, so they'd sit in Sing
    # Inventory forever. Matched on Line (Short), case-insensitive substring.
    'RELEASE_UNUSABLE': True,
    'RELEASE_GENOTYPES': ['Wild', 'Inconclusive'],
    'RELEASE_EXCLUDE_LINES': ['Shank3', 'Bcl11b', 'Scn1a'],   # Scn1a = Dravet

    # After harvest assignments are confirmed, set each scheduled animal's Use
    # in Climb: P14 -> 'Sing - P14', adults -> 'Sing - P56'.
    'UPDATE_ANIMAL_USE': True,

    # Add scheduled animals to cohorts: 'P14 <harvest date>' / 'P56 <behavior
    # date>'. Cohorts cannot be created via the API, so the run pauses and asks
    # you to make any missing ones in Climb. Skippable.
    'ASSIGN_COHORTS': True,

    # When the Envision export is created, write the assigned RapID tags back
    # into Climb as "<original marker>, <tag>" with marker type RapID.
    # Only happens on a complete list — see _push_rapid_markers_to_climb.
    'PUSH_RAPID_MARKERS_TO_CLIMB': True,

    'INPUT_ANIMAL_FILE': 'animals.csv',
    'INPUT_TRACKING_FILE': 'Sing Harvest Sheet.xlsx',
    'INPUT_BIRTHS_FILE': 'births.csv',

    'WEDNESDAY_CAPACITY': 9999,  # Cap removed — no limit on Wednesday slots
    'CAGE_SIZE': 3,
    # Harvest sheet auto-fill. Perfusion protocols use PFA_PER_MOUSE_ML each;
    # the day's batch is split by PFA_MIX_RATIOS on the first row of that date.
    'PFA_PER_MOUSE_ML': 50,
    'PFA_MIX_RATIOS': {'Distilled Water': 0.25, '2xPBS': 0.50, '16% PFA': 0.25},

    'P14_VALID_DAYS': [0, 1, 2, 3, 4],  # Monday=0 through Friday=4
    'P14_HARVEST_AGE_DAYS': 14,         # P14 harvest age — used for cohort born dates
    'P56_BEHAVIOR_START_DAY': 42,
    'P56_BEHAVIOR_END_DAY': 49,
    'P56_BEHAVIOR_DAY_OF_WEEK': 2,  # Wednesday=2
    'P56_HARVEST_DAYS_AFTER_BEHAVIOR': 14,

    'SEXING_OFFSET_DAYS': 9,

    'B6_MIN_PER_MONTH': 3,
    'B6_STRAINS': ['B6J', 'B6NJ'],

    'DATE_VALIDATION': {
        'MAX_FUTURE_DAYS': 365,
        'MAX_PAST_DAYS': 730
    },

    'CHUNK_SIZE': 10000,
    'ENABLE_PROGRESS_BARS': True,

    'LOG_LEVEL': 'INFO',
    'LOG_TO_FILE': True,
    'LOG_TO_CONSOLE': True,

    'DEBUG_MODE': False,
    'RUN_TESTS': False,

    'HARVEST_TARGETS': {
        'Perfusion': 5,
        'MERFISH': 1,
        'RNAseq': 1
    },

    'REQUIRED_ANIMAL_COLUMNS': [
        'Name', 'Birth Date', 'Sex', 'Line (Short)',
        'Genotype', 'Use', 'Status', 'Birth ID', 'Marker Type'
    ],

    'REQUIRED_BIRTHS_COLUMNS': [
        'Birth ID', 'Status', 'Birth Date', 'Live Count', '# of Pups', 'Line (Short)', 'Dam', 'Sire'
    ],

    'SUPER_PRIORITY_STRAINS': [
        'ARID1B', 'CACNA1G', 'CHD8', 'CNTNAP2', 'CTCF',
        'CTNNB1', 'DLL1', 'FMR1', 'GABRA1', 'KMT2C',
        'SCN2A', 'SHANK3', 'SMARCC2'
    ],

    'PRIORITY_STRAINS': {
        'AFF3': 'All', 'AP2S1': 'Half', 'ARID1B': 'Half', 'ASXL3': 'All',
        'ATP6V0A1': 'Half', 'AUTS2': 'Half', 'B6J': 'All', 'B6NJ': 'All',
        'BAP1': 'Half', 'BCL11B': 'Half', 'C3': 'All', 'CACNA1A': 'Half',
        'CACNA1C': 'Half', 'CACNA1G': 'Half', 'CAMK2B': 'Half', 'CASKIN1': 'All',
        'CDKL5': 'All', 'CERT1': 'Half', 'CHAMP1': 'Half', 'CHD2': 'All',
        'CHD8': 'Half', 'CNTNAP2': 'All', 'CTCF': 'All', 'CTNNB1': 'All',
        'CYFIP2': 'All', 'DDX23': 'Half', 'DEAF1': 'All', 'DHDDS': 'Half',
        'DLG4': 'All', 'DLL1': 'Half', 'DNMT3A': 'Half', 'DYRK1A': 'All',
        'EBF3': 'Half', 'EHMT1': 'Half', 'EIF5A': 'All', 'EP300': 'Half',
        'ERF': 'Half', 'FAM120A': 'All', 'FBN1': 'Half', 'FMR1': 'All',
        'FOXP1': 'Half', 'GABRA1': 'All', 'GET204': 'All', 'GABRG2': 'Half', 'GRIA2': 'Half',
        'GRIN2A': 'Half', 'GRIN2B': 'Half', 'GRN': 'All', 'HECW2': 'Half',
        'HERC1': 'All', 'IQSEC2': 'All', 'ITPR1': 'All', 'KAT6B': 'Half',
        'KBTBD7': 'All', 'KCNB1': 'All', 'KCND3': 'All', 'KCNMA1': 'All',
        'KCNT1': 'All', 'KCNT2': 'All', 'KDM5B': 'Half', 'KDM6B': 'Half',
        'KMT2C': 'All', 'KMT2E': 'Half', 'MAGEL2': 'Half', 'MECP2': 'Half',
        'MED13L': 'All', 'MED23': 'Half', 'MTOR': 'Half', 'MYT1L': 'Half',
        'NAA10': 'All', 'NALCN': 'Half', 'NFIX': 'All', 'NRXN1': 'All',
        'PACS2': 'All', 'PAH': 'All', 'PAX5': 'Half', 'POGZ': 'All',
        'POLR3B': 'Half', 'PREP': 'Half', 'PTEN': 'Half', 'PTPRD': 'Half',
        'RAC1': 'Half', 'RALA': 'Half', 'RB1CC1': 'Half', 'RBOBTB2': 'All',
        'RYR2': 'Half', 'SATB1': 'Half', 'SATB2': 'Half', 'SCN1A': 'Half',
        'SCN2A': 'Half', 'SETD1A': 'Half', 'SETD2': 'All', 'SETD5': 'Half',
        'SHANK3': 'Half', 'SLC6A1': 'Half', 'SMARCC2': 'Half', 'SMARCE1': 'Half',
        'SOX2': 'Half', 'SPAST': 'All', 'STXBP1': 'Half', 'SYNCRIP': 'Half',
        'SYNGAP1': 'Half', 'TAOK1': 'Half', 'TBR1': 'Half', 'TCF20': 'All',
        'TCF4': 'Half', 'TCF7L2': 'Half', 'TFAP4': 'All', 'TOP2B': 'Half',
        'TRIO': 'Half', 'U2AF2': 'Half', 'UBE3A': 'Half', 'VPS13B': 'All',
        'WAC': 'Half', 'XPO1': 'Half', 'ZBTB10': 'Half', 'ZBTB21': 'All',
        'ZFHX4': 'Half', 'ZMYM2': 'Half', 'ZNF292': 'All'
    },

    # Strains bred Het×Het — yield is 3/4 usable (1/4 Hom + 1/2 Het), 1/4 Wild
    # Hom animals from these crosses get absolute highest scheduling priority
    'HETXHET_STRAINS': [
        'SHANK3',
    ],
}

DAYS_IN_WEEK = 7
P14_OFFSET_DAYS = 14
P56_HARVEST_OFFSET_FROM_BEHAVIOR = 14

# ============================================================================
# CUSTOM EXCEPTIONS
# ============================================================================

class SchedulerError(Exception):
    pass

class DataValidationError(SchedulerError):
    pass

class SchedulingError(SchedulerError):
    pass

# ============================================================================
# LOGGING SETUP
# ============================================================================

def setup_logging(output_dir: str, level: str = 'INFO') -> logging.Logger:
    log_dir = os.path.join(output_dir, 'logs')
    os.makedirs(log_dir, exist_ok=True)

    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    log_file = os.path.join(log_dir, f'scheduler_{timestamp}.log')

    formatter = logging.Formatter(
        '%(asctime)s - %(levelname)s - %(funcName)s:%(lineno)d - %(message)s',
        datefmt='%Y-%m-%d %H:%M:%S'
    )

    logger_instance = logging.getLogger()
    logger_instance.setLevel(getattr(logging, level.upper()))
    logger_instance.handlers.clear()

    if CONFIG['LOG_TO_FILE']:
        file_handler = RotatingFileHandler(
            log_file, maxBytes=10*1024*1024, backupCount=5, encoding='utf-8')
        file_handler.setLevel(getattr(logging, level.upper()))
        file_handler.setFormatter(formatter)
        logger_instance.addHandler(file_handler)

    if CONFIG['LOG_TO_CONSOLE']:
        import sys as _sys
        stream = _sys.stdout
        # On Windows, wrap stream to handle Unicode safely
        if hasattr(stream, 'reconfigure'):
            try:
                stream.reconfigure(encoding='utf-8', errors='replace')
            except Exception:
                pass
        console_handler = logging.StreamHandler(stream)
        console_handler.setLevel(logging.INFO)
        console_handler.setFormatter(formatter)
        logger_instance.addHandler(console_handler)

    logger_instance.info(f"Logging initialized: {log_file}")
    return logger_instance

logger = logging.getLogger(__name__)

# ============================================================================
# CONFIGURATION VALIDATION
# ============================================================================

def validate_config_advanced() -> bool:
    """Validate all CONFIG values before the pipeline runs.

    Checks ranges, required keys, and cross-field consistency
    (e.g. WEDNESDAY_CAPACITY must be a multiple of CAGE_SIZE).

    Returns:
        True if all checks pass.

    Raises:
        ValueError: If any critical configuration error is found.
    """
    errors = []
    warnings_list = []

    if CONFIG['CAGE_SIZE'] <= 0:
        errors.append(f"CAGE_SIZE must be positive, got {CONFIG['CAGE_SIZE']}")
    elif CONFIG['CAGE_SIZE'] > 10:
        warnings_list.append(f"CAGE_SIZE ({CONFIG['CAGE_SIZE']}) is unusually large")

    if CONFIG['WEDNESDAY_CAPACITY'] <= 0:
        errors.append("WEDNESDAY_CAPACITY must be positive")
    elif CONFIG['WEDNESDAY_CAPACITY'] % CONFIG['CAGE_SIZE'] != 0:
        warnings_list.append(
            f"WEDNESDAY_CAPACITY ({CONFIG['WEDNESDAY_CAPACITY']}) is not a multiple of "
            f"CAGE_SIZE ({CONFIG['CAGE_SIZE']})"
        )

    for harvest_type, target in CONFIG['HARVEST_TARGETS'].items():
        if target < 0:
            errors.append(f"HARVEST_TARGETS['{harvest_type}'] must be non-negative")

    if not CONFIG['P14_VALID_DAYS']:
        errors.append("P14_VALID_DAYS cannot be empty")
    elif not all(0 <= day <= 6 for day in CONFIG['P14_VALID_DAYS']):
        errors.append("P14_VALID_DAYS must contain values 0-6")

    if CONFIG.get('SEXING_OFFSET_DAYS', 9) <= 0:
        errors.append("SEXING_OFFSET_DAYS must be positive")

    if CONFIG.get('B6_MIN_PER_MONTH', 3) < 0:
        errors.append("B6_MIN_PER_MONTH must be non-negative")

    priority_strains = set(s.upper() for s in CONFIG['PRIORITY_STRAINS'].keys())
    super_priority_strains = set(s.upper() for s in CONFIG['SUPER_PRIORITY_STRAINS'])
    missing_from_priority = super_priority_strains - priority_strains
    if missing_from_priority:
        warnings_list.append(
            f"SUPER_PRIORITY_STRAINS contains strains not in PRIORITY_STRAINS: "
            f"{', '.join(sorted(missing_from_priority))}"
        )

    if errors:
        error_msg = "Configuration errors:\n  - " + "\n  - ".join(errors)
        raise ValueError(error_msg)

    return True

try:
    validate_config_advanced()
except ValueError as e:
    print(f"❌ Configuration Error: {e}")
    raise

_PRIORITY_STRAINS_UPPER = {s.upper(): v for s, v in CONFIG['PRIORITY_STRAINS'].items()}
_SUPER_PRIORITY_STRAINS_UPPER = frozenset(s.upper() for s in CONFIG['SUPER_PRIORITY_STRAINS'])
_B6_STRAINS_UPPER = frozenset(s.upper() for s in CONFIG.get('B6_STRAINS', ['B6J', 'B6NJ']))
_HETXHET_STRAINS_UPPER = frozenset(s.upper() for s in CONFIG.get('HETXHET_STRAINS', []))

# ============================================================================
# CANONICAL GENOTYPE LABELS
# ============================================================================

GENOTYPE_WILD   = 'Wild'
GENOTYPE_HET    = 'Het'
GENOTYPE_HOM    = 'Hom'
GENOTYPE_HEMI   = 'Hemi'
GENOTYPE_INBRED = 'Inbred'
GENOTYPE_BLANK  = 'Blank'           # never genotyped — no record exists
GENOTYPE_INCONC = 'Inconclusive'    # genotyped, but the assay gave no usable call

_CANONICAL_GENOTYPES = frozenset([
    GENOTYPE_WILD, GENOTYPE_HET, GENOTYPE_HOM,
    GENOTYPE_HEMI, GENOTYPE_INBRED, GENOTYPE_BLANK,
    GENOTYPE_INCONC,
])

# Neither can be scheduled — both need a genotype before the animal is usable.
_NEEDS_GENOTYPE = frozenset([GENOTYPE_BLANK, GENOTYPE_INCONC])

# ============================================================================
# UTILITY FUNCTIONS
# ============================================================================

def to_date(date_obj: Union[date, datetime, pd.Timestamp, None]) -> Optional[date]:
    """Coerce any date-like object to a plain Python date.

    Handles pd.Timestamp, datetime, date, and string-parseable objects.
    Returns None for None, NaT, or unparseable inputs.

    Args:
        date_obj: Any date-like value — date, datetime, Timestamp, or None.

    Returns:
        A Python date, or None if conversion fails.

    Example:
        >>> to_date(pd.Timestamp('2024-01-15'))
        datetime.date(2024, 1, 15)
        >>> to_date(None)
        None
    """
    if date_obj is None:
        return None
    if pd.isna(date_obj):
        return None
    if isinstance(date_obj, pd.Timestamp):
        return date_obj.date()
    if isinstance(date_obj, datetime):
        return date_obj.date()
    if isinstance(date_obj, date):
        return date_obj
    try:
        ts = pd.to_datetime(date_obj, errors='coerce')
        if pd.isna(ts):
            return None
        return ts.date()
    except Exception:
        return None


def normalize_genotype(genotype: str) -> str:
    if pd.isna(genotype):
        return genotype
    geno_str = str(genotype)
    replacements = {
        '‹': '<', '›': '>', '«': '<', '»': '>',
        '⟨': '<', '⟩': '>', '〈': '<', '〉': '>'
    }
    for old, new in replacements.items():
        geno_str = geno_str.replace(old, new)
    geno_str = unicodedata.normalize('NFKC', geno_str)
    geno_str = ' '.join(geno_str.split())
    return geno_str


def canonicalize_genotype(genotype, strain: str = '') -> str:
    """
    Normalize any genotype string to one of seven canonical labels:
      Wild, Het, Hom, Hemi, Inbred, Blank, Inconclusive

    Blank        = no genotype on record
    Inconclusive = genotyped, but the assay returned no usable call
    """
    if isinstance(genotype, str) and genotype in _CANONICAL_GENOTYPES:
        return genotype

    if genotype is None:
        return GENOTYPE_BLANK
    try:
        if pd.isna(genotype):
            return GENOTYPE_BLANK
    except (TypeError, ValueError):
        pass
    geno_str = str(genotype).strip()
    if geno_str == '' or geno_str.lower() in ('nan', 'none', 'n/a', 'na', '-'):
        return GENOTYPE_BLANK

    if strain:
        strain_upper = str(strain).strip().upper()
        if strain_upper in _B6_STRAINS_UPPER:
            return GENOTYPE_INBRED

    geno_norm = normalize_genotype(geno_str)
    gl = geno_norm.lower()

    # 'Inconclusive' is the only no-call value in Climb's genotypeSymbol
    # vocabulary (key 11), so that's all we need to match.
    if 'inconclusive' in gl:
        return GENOTYPE_INCONC

    hemi_patterns = [
        r'hem[i]?', r'tg/\+', r'\+/tg', r'tg/-', r'-/y', r'[a-z]/y',
    ]
    if any(re.search(p, gl) for p in hemi_patterns):
        return GENOTYPE_HEMI

    wild_patterns = [
        r'\+/\+', r'\+/y', r'\bwt\b', r'\bwild.?type\b', r'\bwildtype\b',
        r'cre.ncar', r'generic.cre', r'cre \+/\+', r'cre \+/y',
    ]
    if any(re.search(p, gl) for p in wild_patterns):
        if not (re.search(r'-/\+', geno_norm) or re.search(r'\+/-', geno_norm)):
            return GENOTYPE_WILD

    hom_patterns = [
        r'-/-', r'\bhom\b', r'\bhomozygous\b', r'mut/mut', r'ko/ko',
        r'\bhom\d',  # matches HOM1, HOM2 etc
    ]
    if any(re.search(p, gl) for p in hom_patterns):
        return GENOTYPE_HOM

    het_patterns = [
        r'-/\+', r'\+/-', r'\bhet\b', r'\bheterozygous\b', r'\bcarrier\b',
        r'\bhet\d',  # matches HET1, HET2 etc (Climb numbering style)
    ]
    if any(re.search(p, gl) for p in het_patterns):
        return GENOTYPE_HET

    if is_wildtype_cre_only(geno_norm):
        return GENOTYPE_WILD

    logger.debug(f"canonicalize_genotype: unrecognised genotype '{geno_str}' — returning Blank")
    return GENOTYPE_BLANK


def is_heterozygous(genotype: str) -> bool:
    if pd.isna(genotype):
        return False
    geno_str = str(genotype).strip()
    if geno_str in (GENOTYPE_HET, GENOTYPE_HEMI):
        return True
    if geno_str in _CANONICAL_GENOTYPES:
        return False
    if '-/+' in geno_str or '+/-' in geno_str:
        return True
    if 'HET' in geno_str.upper():
        return True
    return False


def is_wildtype_cre_only(genotype: str) -> bool:
    if pd.isna(genotype):
        return False
    geno_str = str(genotype).strip()
    if geno_str in _CANONICAL_GENOTYPES:
        return geno_str == GENOTYPE_WILD
    if geno_str == '':
        return False
    geno_lower = geno_str.lower()
    cre_patterns = ['cre ncar', 'cre-ncar', 'generic cre', 'cre +/+', 'cre +/y']
    has_cre_pattern = any(p in geno_lower for p in cre_patterns)
    if not has_cre_pattern:
        return False
    if is_heterozygous(geno_str):
        return False
    if '-/-' in geno_str:
        return False
    return True


def is_priority_strain(strain: str) -> bool:
    if pd.isna(strain):
        return False
    return str(strain).strip().upper() in _PRIORITY_STRAINS_UPPER


def is_super_priority_strain(strain: str) -> bool:
    if pd.isna(strain):
        return False
    return str(strain).strip().upper() in _SUPER_PRIORITY_STRAINS_UPPER


def is_b6_strain(strain: str) -> bool:
    if pd.isna(strain):
        return False
    return str(strain).strip().upper() in _B6_STRAINS_UPPER


def get_strain_breeding_type(strain: str, sex: Optional[str] = None) -> str:
    """Return the breeding type for a given strain.

    Returns:
        'HetxHet' if bred Het×Het (3/4 usable, Hom+Het),
        'All'     if all pups usable (Hom×Hom / inbred),
        'Half'    if ~50% usable (Het×WT), or default for unknown.
    """
    if pd.isna(strain):
        return 'Unknown'
    strain_upper = str(strain).strip().upper()
    if strain_upper in _HETXHET_STRAINS_UPPER:
        return 'HetxHet'
    return _PRIORITY_STRAINS_UPPER.get(strain_upper, 'Half')


def has_toe_clip(marker_type: str) -> bool:
    if pd.isna(marker_type):
        return False
    return 'Toe Clip' in str(marker_type)


def parse_date(date_str: str) -> Optional[date]:
    if not date_str or date_str.strip() == '':
        return None
    try:
        return datetime.strptime(date_str.strip(), '%Y-%m-%d').date()
    except ValueError as e:
        logger.error(f"Invalid date format '{date_str}': {e}")
        return None


def parse_multiple_dates(input_str: str) -> List[date]:
    if not input_str or input_str.strip() == '':
        return []
    dates = []
    for date_str in input_str.split(','):
        date_obj = parse_date(date_str.strip())
        if date_obj:
            dates.append(date_obj)
    return dates


def is_valid_p14_day(date_obj: date) -> bool:
    date_obj = to_date(date_obj)
    if date_obj is None:
        return False
    return date_obj.weekday() in CONFIG['P14_VALID_DAYS']


def next_wednesday(target_date: date) -> Optional[date]:
    target_date = to_date(target_date)
    if target_date is None:
        return None
    days_ahead = CONFIG['P56_BEHAVIOR_DAY_OF_WEEK'] - target_date.weekday()
    if days_ahead < 0:
        days_ahead += DAYS_IN_WEEK
    return target_date + timedelta(days=days_ahead)


def calculate_schedule_dates(birth_date: Union[date, datetime, pd.Timestamp]) -> Optional[Dict[str, date]]:
    birth_date = to_date(birth_date)
    if birth_date is None:
        return None

    today = datetime.now().date()
    max_future = today + timedelta(days=CONFIG['DATE_VALIDATION']['MAX_FUTURE_DAYS'])

    if birth_date > max_future:
        return None

    try:
        p14_harvest = birth_date + timedelta(days=P14_OFFSET_DAYS)
        behavior_start_min = birth_date + timedelta(days=CONFIG['P56_BEHAVIOR_START_DAY'])
        behavior_start_max = birth_date + timedelta(days=CONFIG['P56_BEHAVIOR_END_DAY'])
        sexing_date = birth_date + timedelta(days=CONFIG['SEXING_OFFSET_DAYS'])

        return {
            'birth_date': birth_date,
            'sexing_date': sexing_date,
            'p14_harvest': p14_harvest,
            'p56_behavior_window_start': behavior_start_min,
            'p56_behavior_window_end': behavior_start_max,
        }
    except (OverflowError, ValueError) as e:
        logger.warning(f"Error calculating dates for birth {birth_date}: {e}")
        return None


def get_next_wednesdays(n: int = 6, from_date: Optional[date] = None) -> List[date]:
    if from_date is None:
        from_date = datetime.now().date()

    wednesdays = []
    current = next_wednesday(from_date)

    if current == from_date:
        wednesdays.append(current)
        current = current + timedelta(days=DAYS_IN_WEEK)
    else:
        wednesdays.append(current)
        current = current + timedelta(days=DAYS_IN_WEEK)

    while len(wednesdays) < n:
        wednesdays.append(current)
        current = current + timedelta(days=DAYS_IN_WEEK)

    return wednesdays


def get_p56_behavior_wednesday(birth_date_obj: Optional[date]) -> Optional[date]:
    """
    Return the first Wednesday that falls in the P42-P49 window for a given
    birth date, or None if the birth date is invalid / window has no Wednesday.

    This is used to group blank-genotype animals that share the same behavior
    session for P56 worth-it assessment.
    """
    if birth_date_obj is None:
        return None
    dates = calculate_schedule_dates(birth_date_obj)
    if dates is None:
        return None
    first_wed = next_wednesday(dates['p56_behavior_window_start'])
    if first_wed is None:
        return None
    if first_wed > dates['p56_behavior_window_end']:
        return None  # No Wednesday falls inside P42-P49
    return first_wed


def prompt_wednesday_capacity() -> Tuple[List[date], Optional[List[date]]]:
    wednesdays = get_next_wednesdays(6)
    capacity = CONFIG['WEDNESDAY_CAPACITY']

    print("\n" + "=" * 70)
    print("WEDNESDAY P56 BEHAVIOR CAPACITY CHECK")
    print("=" * 70)
    print(f"Maximum capacity per Wednesday: {capacity} animals")
    print(f"Enter how many animals are ALREADY SCHEDULED for each Wednesday.")
    print(f"Press Enter to skip (assumes 0 scheduled).\n")

    scheduled_counts = {}
    full_dates = []

    for i, wed in enumerate(wednesdays, 1):
        day_label = wed.strftime('%A, %Y-%m-%d')
        while True:
            raw = input(f"  {i}. {day_label} — Already scheduled: ").strip()
            if raw == '':
                scheduled_counts[wed] = 0
                break
            try:
                count = int(raw)
                if count < 0:
                    print(f"     ⚠️  Please enter a number >= 0")
                    continue
                if count > capacity:
                    print(f"     ⚠️  That's already over capacity ({capacity})!")
                scheduled_counts[wed] = count
                break
            except ValueError:
                print(f"     ⚠️  Please enter a whole number (e.g., 0, 6, 18)")

    print(f"\n  {'Wednesday':<28} {'Scheduled':>10} {'Remaining':>10} {'Status':>12}")
    print(f"  {'-'*28} {'-'*10} {'-'*10} {'-'*12}")

    for wed in wednesdays:
        count = scheduled_counts[wed]
        remaining = capacity - count
        if remaining <= 0:
            status = '🔴 FULL'
            full_dates.append(wed)
        elif remaining <= 3:
            status = '🟡 LOW'
        else:
            status = '🟢 OPEN'
        print(f"  {wed.strftime('%A, %Y-%m-%d'):<28} {count:>10} {remaining:>10} {status:>12}")

    if full_dates:
        print(f"\n  ⚠️  {len(full_dates)} Wednesday(s) at capacity")
    else:
        print(f"\n  ✓ All Wednesdays have available capacity")

    return wednesdays, full_dates if full_dates else None


def auto_size_columns(worksheet) -> None:
    """Auto-fit column widths in an openpyxl worksheet based on cell content.

    Iterates all columns and sets width to the longest cell value,
    capped at 60 characters to prevent excessively wide columns.

    Args:
        worksheet: An openpyxl Worksheet object to resize in place.
    """
    for column in worksheet.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if cell.value:
                    cell_length = len(str(cell.value))
                    if cell_length > max_length:
                        max_length = cell_length
            except Exception:
                pass
        adjusted_width = min(max_length + 2, 60)
        worksheet.column_dimensions[column_letter].width = adjusted_width


def truncate_sheet_name(name: str, max_length: int = 31) -> str:
    if len(name) <= max_length:
        return name
    return name[:max_length - 3] + '...'


# ============================================================================
# FILE READING AND VALIDATION
# ============================================================================

def validate_animal_file(df: pd.DataFrame) -> bool:
    missing = [col for col in CONFIG['REQUIRED_ANIMAL_COLUMNS'] if col not in df.columns]
    if missing:
        raise DataValidationError(f"Missing required columns in animal file: {missing}")
    return True


def validate_births_file(df: pd.DataFrame) -> bool:
    core_required = ['Birth ID', 'Status', 'Birth Date']
    missing_core = [col for col in core_required if col not in df.columns]
    if missing_core:
        raise DataValidationError(f"Missing required columns in births file: {missing_core}")
    return True


def process_large_dataset(animal_file: str, chunk_size: int = None) -> pd.DataFrame:
    """Read animal CSV and filter to alive animals only.

    Args:
        animal_file: Path to the animals CSV file.
        chunk_size: Unused; kept for backwards compatibility.

    Returns:
        DataFrame containing only animals with Status == 'Alive'.
    """
    df = pd.read_csv(animal_file)
    return df[df['Status'] == 'Alive'].copy()


def read_animal_data(filename: str) -> pd.DataFrame:
    if not os.path.exists(filename):
        raise FileNotFoundError(f"Animal file not found: {filename}")
    df = process_large_dataset(filename)
    validate_animal_file(df)
    df['Birth Date'] = pd.to_datetime(df['Birth Date'], errors='coerce')
    df = df[df['Status'] == 'Alive'].copy()
    df['Birth ID'] = df['Birth ID'].astype(str)

    df['Genotype'] = df['Genotype'].apply(normalize_genotype)
    df['Raw_Genotype'] = df['Genotype'].copy()  # preserve before canonicalization
    df['Genotype'] = df.apply(
        lambda row: canonicalize_genotype(row['Genotype'], row.get('Line (Short)', '')),
        axis=1
    )

    logger.info(f"Loaded {len(df)} alive animals")
    logger.info(f"Genotype breakdown:\n{df['Genotype'].value_counts().to_string()}")
    return df


def read_births_data(filename: str) -> Optional[pd.DataFrame]:
    if filename is None or not os.path.exists(filename):
        logger.warning(f"Births file not found: {filename}")
        return None
    try:
        df = pd.read_csv(filename)
    except Exception as e:
        logger.warning(f"Error reading births file: {e}")
        return None
    try:
        validate_births_file(df)
    except DataValidationError as e:
        logger.warning(f"{e}")
        return None
    df['Birth Date'] = pd.to_datetime(df['Birth Date'], errors='coerce')
    df['Birth ID'] = df['Birth ID'].astype(str)
    logger.info(f"Loaded {len(df)} birth records")
    return df


def read_tracking_data(filename: str) -> Optional[pd.DataFrame]:
    if filename is None or not os.path.exists(filename):
        logger.warning(f"Tracking file not found: {filename}")
        return None
    try:
        if filename.lower().endswith('.xlsx'):
            df = pd.read_excel(filename, sheet_name='Summary Sheet')
        else:
            df = pd.read_csv(filename)
        logger.info(f"Loaded tracking file: {len(df)} rows")
        return df
    except Exception as e:
        logger.warning(f"Error reading tracking file: {e}")
        return None


# ============================================================================
# DIAGNOSTIC HELPER
# ============================================================================

def diagnose_animal_file(df: pd.DataFrame) -> None:
    print("\n" + "=" * 70)
    print("DIAGNOSTIC: ANIMAL FILE CONTENTS")
    print("=" * 70)
    print(f"  Total rows loaded:     {len(df):,}")
    print(f"  Total columns:         {len(df.columns)}")
    print(f"\n  Column names:")
    for col in df.columns.tolist():
        print(f"    - {repr(col)}")

    if 'Status' in df.columns:
        print(f"\n  'Status' value counts:")
        for val, cnt in df['Status'].value_counts(dropna=False).items():
            print(f"    {repr(val)}: {cnt}")
        alive_count = len(df[df['Status'] == 'Alive'])
        print(f"\n  Animals with Status == 'Alive': {alive_count}")
    else:
        print("\n  ⚠️  'Status' column NOT FOUND")

    if 'Use' in df.columns:
        print(f"\n  'Use' value counts (top 10):")
        for val, cnt in df['Use'].value_counts(dropna=False).head(10).items():
            print(f"    {repr(val)}: {cnt}")
        sing_mask = df['Use'].str.contains('Sing Inventory', na=False, case=False)
        print(f"\n  Animals matching 'Sing Inventory' in Use: {sing_mask.sum()}")
    else:
        print("\n  ⚠️  'Use' column NOT FOUND")

    if 'Genotype' in df.columns:
        print(f"\n  'Genotype' canonical value counts:")
        for val, cnt in df['Genotype'].value_counts(dropna=False).items():
            print(f"    {repr(val)}: {cnt}")
    else:
        print("\n  ⚠️  'Genotype' column NOT FOUND")

    if 'Line (Short)' in df.columns:
        print(f"\n  'Line (Short)' (strain) value counts (top 15):")
        for val, cnt in df['Line (Short)'].value_counts(dropna=False).head(15).items():
            print(f"    {repr(val)}: {cnt}")
    else:
        print("\n  ⚠️  'Line (Short)' column NOT FOUND")

    print("=" * 70 + "\n")


# ============================================================================
# BIRTHS ANALYSIS
# ============================================================================

def calculate_sexing_date(birth_date: Union[date, datetime, pd.Timestamp]) -> Optional[date]:
    bd = to_date(birth_date)
    if bd is None:
        return None
    return bd + timedelta(days=CONFIG['SEXING_OFFSET_DAYS'])


def build_births_sexing_schedule(
    births_df: pd.DataFrame,
    animals_df: Optional[pd.DataFrame] = None
) -> pd.DataFrame:
    """
    Build a sexing schedule for births that have not yet been sexed.
    Any Birth ID that already has animals in animals_df is excluded.
    """
    if births_df is None or births_df.empty:
        return pd.DataFrame()

    today = datetime.now().date()

    sing = births_df[
        births_df['Status'].str.contains('Sing Inventory', na=False, case=False)
    ].copy()

    if sing.empty:
        return pd.DataFrame()

    already_sexed_birth_ids = set()
    if animals_df is not None and len(animals_df) > 0 and 'Birth ID' in animals_df.columns:
        already_sexed_birth_ids = set(
            animals_df['Birth ID'].astype(str).unique()
        )
        logger.info(
            f"build_births_sexing_schedule: {len(already_sexed_birth_ids)} "
            f"Birth IDs already have animals entered (already sexed)"
        )

    rows = []
    skipped_already_sexed = 0

    for _, birth in sing.iterrows():
        birth_id = str(birth.get('Birth ID', 'N/A'))

        if birth_id in already_sexed_birth_ids:
            skipped_already_sexed += 1
            continue

        birth_date_obj = to_date(birth['Birth Date'])
        strain = birth.get('Line (Short)', 'N/A')
        dam = birth.get('Dam', 'N/A')
        sire = birth.get('Sire', 'N/A')
        num_pups = birth.get('# of Pups', birth.get('Live Count', 'N/A'))

        if birth_date_obj is None:
            rows.append({
                'Birth_ID': birth_id,
                'Strain': strain if pd.notna(strain) else 'N/A',
                'Dam': dam if pd.notna(dam) else 'N/A',
                'Sire': sire if pd.notna(sire) else 'N/A',
                'Birth_Date': 'N/A',
                'Num_Pups': num_pups if pd.notna(num_pups) else 'N/A',
                'Sexing_Date': 'N/A',
                'Day_of_Week': 'N/A',
                'Days_Until_Sexing': 'N/A',
                'Sexing_Status': '❓ Unknown — No birth date',
                'P14_Expected_Date': 'N/A',
                'P14_Day_of_Week': 'N/A',
            })
            continue

        sexing_date = birth_date_obj + timedelta(days=CONFIG['SEXING_OFFSET_DAYS'])
        p14_date = birth_date_obj + timedelta(days=P14_OFFSET_DAYS)
        days_until = (sexing_date - today).days

        if days_until < 0:
            status = f'✅ Done (was {sexing_date.strftime("%Y-%m-%d")})'
        elif days_until == 0:
            status = '🔴 TODAY — Sex pups now!'
        elif days_until == 1:
            status = '🟠 TOMORROW — Prepare'
        elif days_until <= 3:
            status = f'🟡 SOON — {days_until} days'
        else:
            status = f'🟢 Upcoming — {days_until} days'

        rows.append({
            'Birth_ID': birth_id,
            'Strain': strain if pd.notna(strain) else 'N/A',
            'Dam': dam if pd.notna(dam) else 'N/A',
            'Sire': sire if pd.notna(sire) else 'N/A',
            'Birth_Date': birth_date_obj.strftime('%Y-%m-%d'),
            'Num_Pups': int(num_pups) if pd.notna(num_pups) else 'N/A',
            'Sexing_Date': sexing_date.strftime('%Y-%m-%d'),
            'Day_of_Week': sexing_date.strftime('%A'),
            'Days_Until_Sexing': days_until,
            'Sexing_Status': status,
            'P14_Expected_Date': p14_date.strftime('%Y-%m-%d'),
            'P14_Day_of_Week': p14_date.strftime('%A'),
        })

    if skipped_already_sexed > 0:
        logger.info(
            f"build_births_sexing_schedule: skipped {skipped_already_sexed} "
            f"births already sexed (animals entered in system)"
        )

    df = pd.DataFrame(rows)
    if df.empty:
        return df

    def _sexing_sort_key(val: Any) -> tuple:
        """Sort key: upcoming dates first (asc), then past dates, then unknowns."""
        if isinstance(val, int):
            return (0 if val >= 0 else 1, val if val >= 0 else -val)
        return (2, 0)

    df = df.sort_values(
        by='Days_Until_Sexing',
        key=lambda col: col.map(_sexing_sort_key)
    )
    df = df.reset_index(drop=True)
    return df


def analyze_birth_scheduling_potential(birth: pd.Series, requirements: Dict,
                                       remaining_needs: Dict, today: date) -> Dict:
    birth_date = birth['Birth Date']
    if pd.isna(birth_date):
        return {
            'P14_Potential': 'Unknown', 'P14_Reason': 'No birth date',
            'P14_Expected_Date': 'N/A', 'P14_Day_of_Week': 'N/A',
            'P56_Potential': 'Unknown', 'P56_Reason': 'No birth date',
            'P56_Expected_Behavior_Date': 'N/A', 'P56_Expected_Harvest_Date': 'N/A',
            'Quota_Status': 'Unknown', 'Priority_Strain': 'Unknown', 'Age_Today_Days': 'N/A',
            'Sexing_Date': 'N/A', 'Sexing_Day_of_Week': 'N/A',
        }

    birth_date_obj = to_date(birth_date)
    strain = birth.get('Line (Short)', '')

    if birth_date_obj is None:
        return {
            'P14_Potential': 'Unknown', 'P14_Reason': 'Invalid birth date',
            'P14_Expected_Date': 'N/A', 'P14_Day_of_Week': 'N/A',
            'P56_Potential': 'Unknown', 'P56_Reason': 'Invalid birth date',
            'P56_Expected_Behavior_Date': 'N/A', 'P56_Expected_Harvest_Date': 'N/A',
            'Quota_Status': 'Unknown',
            'Priority_Strain': 'YES' if is_priority_strain(strain) else 'No',
            'Age_Today_Days': 'N/A', 'Sexing_Date': 'N/A', 'Sexing_Day_of_Week': 'N/A',
        }

    dates = calculate_schedule_dates(birth_date_obj)

    if dates is None:
        return {
            'P14_Potential': 'Unknown', 'P14_Reason': 'Invalid birth date',
            'P14_Expected_Date': 'N/A', 'P14_Day_of_Week': 'N/A',
            'P56_Potential': 'Unknown', 'P56_Reason': 'Invalid birth date',
            'P56_Expected_Behavior_Date': 'N/A', 'P56_Expected_Harvest_Date': 'N/A',
            'Quota_Status': 'Unknown',
            'Priority_Strain': 'YES' if is_priority_strain(strain) else 'No',
            'Age_Today_Days': 'N/A', 'Sexing_Date': 'N/A', 'Sexing_Day_of_Week': 'N/A',
        }

    p14_harvest = dates['p14_harvest']
    behavior_window_start = dates['p56_behavior_window_start']
    behavior_window_end = dates['p56_behavior_window_end']
    sexing_date = dates['sexing_date']

    p14_valid = is_valid_p14_day(p14_harvest)
    p14_in_future = p14_harvest > today

    if not p14_in_future:
        p14_potential = 'Past'
        if p14_harvest == today:
            p14_reason = f'P14 date is today ({p14_harvest}) — too late to schedule'
        else:
            p14_reason = f'P14 date ({p14_harvest}) has passed'
    elif not p14_valid:
        p14_potential = 'No'
        p14_reason = f'P14 falls on {p14_harvest.strftime("%A")} (invalid day)'
    else:
        p14_potential = 'Yes'
        p14_reason = f'Could schedule on {p14_harvest.strftime("%A, %Y-%m-%d")}'

    first_wednesday = next_wednesday(behavior_window_start)
    p56_harvest_date = None

    if first_wednesday is None:
        p56_potential = 'No'
        p56_reason = 'Cannot calculate P56 behavior date'
    elif first_wednesday > behavior_window_end:
        p56_potential = 'No'
        p56_reason = 'No Wednesday in P42-49 window'
    elif first_wednesday < today:
        p56_potential = 'Past'
        p56_reason = f'P56 window ({first_wednesday}) has passed'
    else:
        p56_potential = 'Yes'
        p56_reason = f'Could schedule behavior on {first_wednesday.strftime("%A, %Y-%m-%d")}'
        p56_harvest_date = first_wednesday + timedelta(days=P56_HARVEST_OFFSET_FROM_BEHAVIOR)

    quota_status = 'Unknown'
    quota_details = []
    is_priority = is_priority_strain(strain)

    if remaining_needs and pd.notna(strain):
        strain_key = str(strain).strip().upper()
        if strain_key in remaining_needs:
            p14_needs = remaining_needs[strain_key]['P14']
            p56_needs = remaining_needs[strain_key]['P56']
            for timepoint, needs_dict in [('P14', p14_needs), ('P56', p56_needs)]:
                for sex in ['Male', 'Female']:
                    total = sum(needs_dict[sex][ht]['needed'] for ht in ['Perfusion', 'MERFISH', 'RNAseq'])
                    if total > 0:
                        quota_details.append(f"{timepoint} {sex}: {total} needed")
            quota_status = 'NEEDED - ' + '; '.join(quota_details) if quota_details else 'Quota Complete'
        else:
            quota_status = 'Not tracked in requirements'

    age_days = (today - birth_date_obj).days

    return {
        'P14_Potential': p14_potential,
        'P14_Reason': p14_reason,
        'P14_Expected_Date': p14_harvest.strftime('%Y-%m-%d'),
        'P14_Day_of_Week': p14_harvest.strftime('%A'),
        'P56_Potential': p56_potential,
        'P56_Reason': p56_reason,
        'P56_Expected_Behavior_Date': first_wednesday.strftime('%Y-%m-%d') if p56_potential != 'No' and first_wednesday else 'N/A',
        'P56_Expected_Harvest_Date': p56_harvest_date.strftime('%Y-%m-%d') if p56_harvest_date else 'N/A',
        'Quota_Status': quota_status,
        'Priority_Strain': 'YES' if is_priority else 'No',
        'Age_Today_Days': age_days,
        'Sexing_Date': sexing_date.strftime('%Y-%m-%d') if sexing_date else 'N/A',
        'Sexing_Day_of_Week': sexing_date.strftime('%A') if sexing_date else 'N/A',
    }


def estimate_expected_animals(birth: pd.Series) -> Dict:
    num_pups = None
    count_source = 'No count field'

    if 'Live Count' in birth.index:
        num_pups = birth.get('Live Count', None)
        if pd.notna(num_pups):
            count_source = 'Live Count'

    if num_pups is None and '# of Pups' in birth.index:
        num_pups = birth.get('# of Pups', None)
        if pd.notna(num_pups):
            count_source = '# of Pups'

    strain = birth.get('Line (Short)', '')
    breeding_type = get_strain_breeding_type(strain)

    if pd.isna(num_pups):
        return {
            'Expected_Total_Born': 'Unknown', 'Expected_Usable': 'Unknown',
            'Expected_Usable_Males': 'Unknown', 'Expected_Usable_Females': 'Unknown',
            'Breeding_Type': breeding_type,
            'Estimation_Note': 'No pup count in birth record'
        }

    try:
        total_pups = int(num_pups)
    except (ValueError, TypeError):
        return {
            'Expected_Total_Born': 'Unknown', 'Expected_Usable': 'Unknown',
            'Expected_Usable_Males': 'Unknown', 'Expected_Usable_Females': 'Unknown',
            'Breeding_Type': breeding_type,
            'Estimation_Note': f'Invalid pup count: {num_pups}'
        }

    if breeding_type == 'Half':
        expected_usable = total_pups // 2
        expected_usable_males = expected_usable // 2
        expected_usable_females = expected_usable - expected_usable_males
        note = f'Het×WT: ~50% usable ({expected_usable} of {total_pups}) [from {count_source}]'
    elif breeding_type == 'All':
        expected_usable = total_pups
        expected_usable_males = total_pups // 2
        expected_usable_females = total_pups - expected_usable_males
        note = f'Hom×Hom: All usable ({expected_usable} of {total_pups}) [from {count_source}]'
    else:
        expected_usable = total_pups // 2
        expected_usable_males = expected_usable // 2
        expected_usable_females = expected_usable - expected_usable_males
        note = f'Unknown strain: Het×WT cross — ~50%) [from {count_source}]'

    return {
        'Expected_Total_Born': total_pups,
        'Expected_Usable': expected_usable,
        'Expected_Usable_Males': f'~{expected_usable_males}',
        'Expected_Usable_Females': f'~{expected_usable_females}',
        'Breeding_Type': breeding_type,
        'Estimation_Note': note
    }


def determine_action_required(potential: Dict, expectations: Dict, age_days) -> str:
    actions = []
    if potential['P14_Potential'] == 'Yes' or potential['P56_Potential'] == 'Yes':
        actions.append('🔍 VERIFY animals exist and have correct Birth ID')
        if potential['Quota_Status'].startswith('NEEDED'):
            actions.append('⚠️ URGENT: Quota needs exist - locate animals immediately')
    if age_days is not None and age_days != 'N/A':
        if age_days > 56:
            actions.append('❌ Too old for P56 - consider P14 retrospective or exclude')
        elif age_days > 49:
            actions.append('⏰ P56 window closing - urgent genotyping needed')
        elif age_days >= 42:
            actions.append('📋 P56 window open - genotype and schedule behavior')
        elif age_days > 14:
            actions.append('⏰ P14 window passed - plan for P56')
        elif age_days >= 10:
            actions.append('📋 Genotype for P14 scheduling')
        else:
            actions.append('⏳ Monitor - too young for scheduling')
    if expectations.get('Expected_Total_Born') == 0:
        actions.append('ℹ️ Birth shows 0 pups - verify and update status')
    if not actions:
        actions.append('📧 Contact lab manager for clarification')
    return ' | '.join(actions)


def find_unmatched_births_enhanced(births_df: Optional[pd.DataFrame], animals_df: pd.DataFrame,
                                    requirements: Dict, remaining_needs: Dict) -> pd.DataFrame:
    if births_df is None or births_df.empty:
        return pd.DataFrame()

    today = datetime.now().date()
    logger.info("Analyzing unmatched births...")

    sing_inventory_births = births_df[
        births_df['Status'].str.contains('Sing Inventory', na=False, case=False)
    ].copy()

    if sing_inventory_births.empty:
        return pd.DataFrame()

    animal_birth_ids = set(animals_df['Birth ID'].astype(str).unique())
    unmatched_births = []

    for idx, birth in sing_inventory_births.iterrows():
        birth_id = str(birth['Birth ID'])
        if birth_id == 'nan' or birth_id.strip() == '':
            continue
        if birth_id not in animal_birth_ids:
            birth_date = birth['Birth Date']
            birth_date_str = to_date(birth_date).strftime('%Y-%m-%d') if pd.notna(birth_date) else 'N/A'
            strain = birth.get('Line (Short)', 'N/A')
            dam = birth.get('Dam', 'N/A')
            sire = birth.get('Sire', 'N/A')
            num_pups = birth.get('# of Pups', 'N/A')

            potential = analyze_birth_scheduling_potential(birth, requirements, remaining_needs, today)
            expectations = estimate_expected_animals(birth)
            age_days = potential.get('Age_Today_Days', 'N/A')

            if age_days != 'N/A':
                if age_days > 56:
                    urgency = '🔴 URGENT - Past P56'
                elif age_days > 42:
                    urgency = '🟡 HIGH - In P56 window'
                elif age_days > 14:
                    urgency = '🟢 MEDIUM - Past P14'
                elif age_days >= 10:
                    urgency = '🟢 LOW - Approaching P14'
                else:
                    urgency = '⚪ INFO - Too young'
            else:
                urgency = '❓ UNKNOWN - No birth date'

            possible_reasons = []
            if pd.notna(num_pups) and num_pups == 0:
                possible_reasons.append('Birth record shows 0 pups')
            elif pd.notna(birth_date) and age_days != 'N/A' and age_days < 5:
                possible_reasons.append('Birth too recent - animals may not be entered yet')
            else:
                possible_reasons.append('Animals not found/entered in Climb')
                possible_reasons.append('Animals may have been culled')
                possible_reasons.append('Birth ID mismatch possible')

            unmatched_births.append({
                'Urgency': urgency,
                'Birth_ID': birth_id,
                'Birth_Date': birth_date_str,
                'Age_Days': age_days,
                'Strain': strain if pd.notna(strain) else 'N/A',
                'Priority_Strain': potential.get('Priority_Strain', 'Unknown'),
                'Dam': dam if pd.notna(dam) else 'N/A',
                'Sire': sire if pd.notna(sire) else 'N/A',
                'Num_Pups_Recorded': num_pups if pd.notna(num_pups) else 'N/A',
                'Status': birth['Status'],
                **expectations,
                'Sexing_Date': potential.get('Sexing_Date', 'N/A'),
                'Sexing_Day_of_Week': potential.get('Sexing_Day_of_Week', 'N/A'),
                'P14_Potential': potential['P14_Potential'],
                'P14_Expected_Date': potential['P14_Expected_Date'],
                'P14_Day_of_Week': potential['P14_Day_of_Week'],
                'P14_Analysis': potential['P14_Reason'],
                'P56_Potential': potential['P56_Potential'],
                'P56_Expected_Behavior_Date': potential['P56_Expected_Behavior_Date'],
                'P56_Expected_Harvest_Date': potential['P56_Expected_Harvest_Date'],
                'P56_Analysis': potential['P56_Reason'],
                'Quota_Status': potential['Quota_Status'],
                'Possible_Reasons': ' | '.join(possible_reasons),
                'Action_Required': determine_action_required(potential, expectations, age_days)
            })

    unmatched_df = pd.DataFrame(unmatched_births)

    if len(unmatched_df) > 0:
        urgency_order = {
            '🔴 URGENT - Past P56': 0, '🟡 HIGH - In P56 window': 1,
            '🟢 MEDIUM - Past P14': 2, '🟢 LOW - Approaching P14': 3,
            '⚪ INFO - Too young': 4, '❓ UNKNOWN - No birth date': 5
        }
        unmatched_df['_urgency_sort'] = unmatched_df['Urgency'].map(urgency_order)
        unmatched_df = unmatched_df.sort_values(['_urgency_sort', 'Birth_Date'])
        unmatched_df = unmatched_df.drop(columns=['_urgency_sort'])

    return unmatched_df


# ============================================================================
# REQUIREMENTS PARSING
# ============================================================================

def parse_requirements(tracking_df: Optional[pd.DataFrame]) -> Dict:
    if tracking_df is None or tracking_df.empty:
        return {}

    logger.info("Parsing tracking file")

    column_indices = {
        'P14': {'Male': {'Perfusion': 3, 'MERFISH': 7, 'RNAseq': 11},
                'Female': {'Perfusion': 4, 'MERFISH': 8, 'RNAseq': 12}},
        'P56': {'Male': {'Perfusion': 5, 'MERFISH': 9, 'RNAseq': 13},
                'Female': {'Perfusion': 6, 'MERFISH': 10, 'RNAseq': 14}}
    }

    if len(tracking_df.columns) < 15:
        logger.warning(f"Expected at least 15 columns, found {len(tracking_df.columns)}")
        return {}

    requirements = {}
    seen_first_data = False
    for idx, row in tracking_df.iterrows():
        strain = row.iloc[0]
        strain_str_raw = str(strain).strip() if not pd.isna(strain) else ''

        # Stop when we hit a second 'Lines'/'Line' header — marks end of strain data
        if strain_str_raw in ('Lines', 'Line'):
            if seen_first_data:
                logger.info(f"parse_requirements: stopping at repeat header row {idx}")
                break
            continue

        if pd.isna(strain) or strain_str_raw == '':
            continue

        seen_first_data = True
        strain_str = strain_str_raw
        strain_key = strain_str.upper()

        completed = {
            'P14': {'Male': {'Perfusion': 0, 'MERFISH': 0, 'RNAseq': 0},
                    'Female': {'Perfusion': 0, 'MERFISH': 0, 'RNAseq': 0}},
            'P56': {'Male': {'Perfusion': 0, 'MERFISH': 0, 'RNAseq': 0},
                    'Female': {'Perfusion': 0, 'MERFISH': 0, 'RNAseq': 0}}
        }

        try:
            for timepoint in ['P14', 'P56']:
                for sex in ['Male', 'Female']:
                    for harvest_type in ['Perfusion', 'MERFISH', 'RNAseq']:
                        col_idx = column_indices[timepoint][sex][harvest_type]
                        if col_idx < len(row):
                            value = row.iloc[col_idx]
                            completed[timepoint][sex][harvest_type] = int(value) if pd.notna(value) and str(value).strip() != '' else 0
        except Exception as e:
            logger.warning(f"Could not parse row for strain '{strain}': {e}")
            continue

        targets = {
            'P14': {'Male': dict(CONFIG['HARVEST_TARGETS']), 'Female': dict(CONFIG['HARVEST_TARGETS'])},
            'P56': {'Male': dict(CONFIG['HARVEST_TARGETS']), 'Female': dict(CONFIG['HARVEST_TARGETS'])}
        }

        requirements[strain_key] = {
            'original_name': strain_str,
            'completed': completed,
            'targets': targets,
            'is_priority': is_priority_strain(strain_str),
            'genotyped': str(row.iloc[1]).strip().upper() != 'NO',
        }

    logger.info(f"Parsed {len(requirements)} strains")
    return requirements


def calculate_remaining_needs(requirements: Dict) -> Dict:
    if not requirements:
        return {}

    remaining = {}
    for strain_key, data in requirements.items():
        remaining[strain_key] = {}
        for timepoint in ['P14', 'P56']:
            remaining[strain_key][timepoint] = {}
            for sex in ['Male', 'Female']:
                remaining[strain_key][timepoint][sex] = {}
                for harvest_type in ['Perfusion', 'MERFISH', 'RNAseq']:
                    completed = data['completed'][timepoint][sex][harvest_type]
                    target = data['targets'][timepoint][sex][harvest_type]
                    needed = max(0, target - completed)
                    remaining[strain_key][timepoint][sex][harvest_type] = {
                        'completed': completed, 'target': target, 'needed': needed
                    }
    return remaining


def resolve_strain_key(strain: str, genotype: str, remaining_needs: Dict) -> str:
    """Return the requirements key for a strain+genotype combo.

    Tries 'STRAIN-GENOTYPE' first (e.g. 'SHANK3-HET'), falls back to
    'STRAIN' (e.g. 'SHANK3') if the composite key isn't in the tracking sheet.
    """
    base = str(strain).strip().upper()
    geno = str(genotype).strip().upper()
    composite = f"{base}-{geno}"
    if composite in remaining_needs:
        return composite
    return base


def group_has_quota(strain: str, sex: str, timepoint: str, remaining_needs: Dict) -> bool:
    strain_upper = str(strain).strip().upper()
    if strain_upper in _B6_STRAINS_UPPER:
        return True
    if not remaining_needs:
        return True

    strain_key = strain_upper
    if strain_key not in remaining_needs:
        return True

    if timepoint not in remaining_needs.get(strain_key, {}):
        return True
    if sex not in remaining_needs[strain_key].get(timepoint, {}):
        return True
    needs = remaining_needs[strain_key][timepoint][sex]
    total_needed = needs['MERFISH']['needed'] + needs['RNAseq']['needed'] + needs['Perfusion']['needed']
    return total_needed >= 1


# ============================================================================
# ANIMAL FILTERING
# ============================================================================

def filter_animals_by_use(animals_df: pd.DataFrame) -> Tuple[pd.DataFrame, pd.DataFrame]:
    """Filter animals to only those with 'Sing Inventory' in Use column.

    Args:
        animals_df: Full animal DataFrame.

    Returns:
        Tuple of (filtered_df, excluded_df) where excluded_df documents
        all animals that did not match 'Sing Inventory'.
    """
    if 'Use' not in animals_df.columns:
        return animals_df, pd.DataFrame()

    mask = animals_df['Use'].str.strip().str.lower() == 'sing inventory'
    filtered = animals_df[mask].copy()
    excluded = animals_df[~mask].copy()

    if excluded.empty:
        return filtered, pd.DataFrame()

    excluded_df = pd.DataFrame({
        'Animal_Name': excluded.get('Name', pd.Series(['Unknown'] * len(excluded), index=excluded.index)),
        'Birth_ID':    excluded.get('Birth ID', pd.Series(['N/A'] * len(excluded), index=excluded.index)),
        'Strain':      excluded.get('Line (Short)', pd.Series(['N/A'] * len(excluded), index=excluded.index)),
        'Genotype':    excluded.get('Genotype', pd.Series(['N/A'] * len(excluded), index=excluded.index)),
        'Sex':         excluded.get('Sex', pd.Series(['N/A'] * len(excluded), index=excluded.index)),
        'Reason':      "Use = '" + excluded['Use'].fillna('N/A') + "' — not 'Sing Inventory'",
    })

    return filtered, excluded_df


def filter_animals_by_genotype_first_pass(
    animals_df: pd.DataFrame,
    no_geno_strains: Optional[frozenset] = None,
) -> Tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    """
    First-pass genotype filter.

    Returns
    -------
    filtered_df   : animals with usable genotypes (Het, Hom, Hemi, Inbred)
    excluded_df   : animals definitively excluded (Wild, Cre-only Wild, Het-in-hom-only-strain)
    blank_df      : animals with Blank genotype — held for second pass

    Parameters
    ----------
    no_geno_strains : upper-cased strain keys where column B == 'No' in the
                      tracking sheet (inbred / hom-only crosses).  Het animals
                      from these strains are excluded — only Hom is scheduled.
                      Combined with _HETXHET_STRAINS_UPPER so both rules apply.
    """
    if no_geno_strains is None:
        no_geno_strains = frozenset()
    # All strains where Het animals must be excluded
    het_excluded_strains = no_geno_strains | _HETXHET_STRAINS_UPPER

    excluded_records = []
    blank_records    = []
    keep_mask        = []

    for _, row in animals_df.iterrows():
        geno   = row.get('Genotype', GENOTYPE_BLANK)
        strain = row.get('Line (Short)', '')
        name   = row.get('Name', 'Unknown')

        if geno in _NEEDS_GENOTYPE:
            blank_records.append(row)
            keep_mask.append(False)

        elif geno == GENOTYPE_WILD:
            excluded_records.append({
                'Animal_Name': name,
                'Birth_ID':    row.get('Birth ID', 'N/A'),
                'Strain':      strain,
                'Genotype':    geno,
                'Sex':         row.get('Sex', 'N/A'),
                'Birth_Date':  (
                    to_date(row.get('Birth Date')).strftime('%Y-%m-%d')
                    if to_date(row.get('Birth Date')) else 'N/A'
                ),
                'Reason': 'Wild genotype — not usable for harvest',
            })
            keep_mask.append(False)

        elif geno in (GENOTYPE_HET, GENOTYPE_HOM, GENOTYPE_HEMI, GENOTYPE_INBRED):
            # Het animals from hom-only or Het×Het strains are not scheduled
            strain_upper = str(strain).strip().upper()
            if geno == GENOTYPE_HET and strain_upper in het_excluded_strains:
                excluded_records.append({
                    'Animal_Name': name,
                    'Birth_ID':    row.get('Birth ID', 'N/A'),
                    'Strain':      strain,
                    'Genotype':    geno,
                    'Sex':         row.get('Sex', 'N/A'),
                    'Birth_Date':  (
                        to_date(row.get('Birth Date')).strftime('%Y-%m-%d')
                        if to_date(row.get('Birth Date')) else 'N/A'
                    ),
                    'Reason': f'Het excluded — {strain} only schedules Hom animals',
                })
                keep_mask.append(False)
            else:
                keep_mask.append(True)

        else:
            # Unrecognised canonical value — treat as blank
            blank_records.append(row)
            keep_mask.append(False)

    filtered_df = animals_df[keep_mask].copy()
    excluded_df = pd.DataFrame(excluded_records)
    blank_df    = (
        pd.DataFrame(blank_records)
        if blank_records
        else pd.DataFrame(columns=animals_df.columns)
    )

    logger.info(
        f"Genotype first pass: {len(filtered_df)} kept, "
        f"{len(excluded_df)} excluded, {len(blank_df)} blanks"
    )
    return filtered_df, excluded_df, blank_df


def filter_animals_by_dates(animals_df: pd.DataFrame,
                            birth_date_start: Optional[date] = None,
                            birth_date_end: Optional[date] = None,
                            behavior_date_start: Optional[date] = None,
                            behavior_date_end: Optional[date] = None) -> Tuple[pd.DataFrame, pd.DataFrame]:
    if not any([birth_date_start, birth_date_end, behavior_date_start, behavior_date_end]):
        return animals_df, pd.DataFrame()

    df = animals_df.copy()
    df['_birth_date_obj'] = df['Birth Date'].apply(to_date)
    mask = pd.Series(True, index=df.index)
    reasons = pd.Series('', index=df.index)

    if birth_date_start:
        too_early = df['_birth_date_obj'].apply(lambda d: d is not None and d < birth_date_start)
        mask &= ~too_early
        reasons = reasons.where(~too_early, reasons + f'Birth date before {birth_date_start} | ')

    if birth_date_end:
        too_late = df['_birth_date_obj'].apply(lambda d: d is not None and d > birth_date_end)
        mask &= ~too_late
        reasons = reasons.where(~too_late, reasons + f'Birth date after {birth_date_end} | ')

    if behavior_date_start or behavior_date_end:
        def calc_first_wednesday(birth_d):
            if birth_d is None:
                return None
            sched = calculate_schedule_dates(birth_d)
            if sched is None:
                return None
            return next_wednesday(sched['p56_behavior_window_start'])

        df['_first_wed'] = df['_birth_date_obj'].apply(calc_first_wednesday)

        if behavior_date_start:
            too_early_b = df['_first_wed'].apply(lambda d: d is not None and d < behavior_date_start)
            mask &= ~too_early_b
            reasons = reasons.where(~too_early_b, reasons + f'P56 behavior before {behavior_date_start} | ')

        if behavior_date_end:
            too_late_b = df['_first_wed'].apply(lambda d: d is not None and d > behavior_date_end)
            mask &= ~too_late_b
            reasons = reasons.where(~too_late_b, reasons + f'P56 behavior after {behavior_date_end} | ')

    filtered = animals_df.loc[mask].copy()

    excluded_indices = df.index[~mask]
    if len(excluded_indices) > 0:
        excluded_records = []
        for idx in excluded_indices:
            row = animals_df.loc[idx]
            birth_d = to_date(row.get('Birth Date'))
            excluded_records.append({
                'Animal_Name': row.get('Name', 'Unknown'),
                'Birth_ID': row.get('Birth ID', 'N/A'),
                'Strain': row.get('Line (Short)', 'N/A'),
                'Sex': row.get('Sex', 'N/A'),
                'Birth_Date': birth_d.strftime('%Y-%m-%d') if birth_d else 'N/A',
                'Reason': reasons.loc[idx].rstrip(' | ')
            })
        excluded_df = pd.DataFrame(excluded_records)
    else:
        excluded_df = pd.DataFrame()

    return filtered, excluded_df


# ============================================================================
# ELIGIBILITY CHECKING
# ============================================================================

def check_eligibility(animals_df: pd.DataFrame,
                      full_behavior_dates: Optional[List[date]] = None) -> pd.DataFrame:
    today = datetime.now().date()
    full_dates_set = set(full_behavior_dates) if full_behavior_dates else set()

    logger.info("Checking animal eligibility...")

    if animals_df.empty:
        logger.warning("check_eligibility received empty DataFrame — returning empty result")
        return pd.DataFrame()

    original_columns = animals_df.columns.tolist()
    eligibility = []

    iterator = animals_df.iterrows()
    if CONFIG['ENABLE_PROGRESS_BARS']:
        iterator = tqdm(list(animals_df.iterrows()), total=len(animals_df), desc="Checking eligibility")

    for idx, row in iterator:
        animal_name = row.get('Name', 'Unknown')
        birth_date = to_date(row.get('Birth Date'))
        strain = row.get('Line (Short)', 'N/A')
        genotype = row.get('Genotype')
        sex = row.get('Sex')
        marker_type = row.get('Marker Type', '')
        birth_id = row.get('Birth ID', 'N/A')

        original_data = {col: row.get(col) for col in original_columns}

        base_record = {
            **original_data,
            'Animal_Name': animal_name,
            'Birth_ID': str(birth_id),
            'Strain': strain,
            'Genotype': genotype if pd.notna(genotype) else GENOTYPE_BLANK,
            'Sex': sex,
            'Marker_Type': marker_type,
        }

        if birth_date is None:
            eligibility.append({
                **base_record,
                'Birth_Date': 'N/A',
                'Age_Today_Days': None,
                'P14_Eligible': False,
                'P14_Reason': 'No birth date',
                'P14_Too_Old': False,
                'P14_Date': None,
                'P14_Age_At_Harvest_Days': None,
                'P14_Age_At_Harvest_Months': None,
                'P56_Eligible': False,
                'P56_Reason': 'No birth date',
                'P56_Too_Old': False,
                'P56_Behavior_Date': None,
                'P56_Harvest_Date': None,
                'P56_Age_At_Behavior_Days': None,
                'P56_Age_At_Behavior_Months': None,
                'P56_Age_At_Harvest_Days': None,
                'P56_Age_At_Harvest_Months': None,
                'Unusable_Note': '',
            })
            continue

        dates = calculate_schedule_dates(birth_date)

        if dates is None:
            eligibility.append({
                **base_record,
                'Birth_Date': birth_date.strftime('%Y-%m-%d'),
                'Age_Today_Days': (today - birth_date).days,
                'P14_Eligible': False,
                'P14_Reason': 'Invalid birth date',
                'P14_Too_Old': False,
                'P14_Date': None,
                'P14_Age_At_Harvest_Days': None,
                'P14_Age_At_Harvest_Months': None,
                'P56_Eligible': False,
                'P56_Reason': 'Invalid birth date',
                'P56_Too_Old': False,
                'P56_Behavior_Date': None,
                'P56_Harvest_Date': None,
                'P56_Age_At_Behavior_Days': None,
                'P56_Age_At_Behavior_Months': None,
                'P56_Age_At_Harvest_Days': None,
                'P56_Age_At_Harvest_Months': None,
                'Unusable_Note': '',
            })
            continue

        p14_harvest = dates['p14_harvest']
        behavior_window_start = dates['p56_behavior_window_start']
        behavior_window_end = dates['p56_behavior_window_end']
        age_today = (today - birth_date).days

        # P14 eligibility
        p14_age_at_harvest_days = (p14_harvest - birth_date).days
        p14_age_at_harvest_months = round(p14_age_at_harvest_days / 30.44)
        p14_too_old = p14_harvest < today

        if p14_too_old:
            p14_eligible = False
            days_past = (today - p14_harvest).days
            if days_past == 0:
                p14_reason = (
                    f'❌ TOO LATE FOR P14 — P14 date is today '
                    f'({p14_harvest.strftime("%Y-%m-%d")}) — '
                    f'harvest must be scheduled in advance'
                )
            else:
                p14_reason = (
                    f'❌ TOO OLD FOR P14 — P14 date was '
                    f'{p14_harvest.strftime("%Y-%m-%d")} '
                    f'({days_past} days ago, animal is {age_today}d old)'
                )
        elif not is_valid_p14_day(p14_harvest):
            p14_eligible = False
            p14_reason = (
                f'P14 falls on {p14_harvest.strftime("%A")} '
                f'({p14_harvest.strftime("%Y-%m-%d")}) — not a valid harvest day'
            )
        else:
            p14_eligible = True
            p14_reason = f'Eligible: {p14_harvest.strftime("%A, %Y-%m-%d")}'

        # P56 eligibility
        p56_eligible = False
        p56_reason = ''
        p56_too_old = False
        behavior_suggested = None
        p56_age_at_behavior_days = None
        p56_age_at_behavior_months = None
        p56_age_at_harvest_days = None
        p56_age_at_harvest_months = None
        p56_window_passed = behavior_window_end < today

        if has_toe_clip(marker_type):
            p56_reason = 'Has Toe Clip marker — not allowed for P56 behavior'
        elif p56_window_passed:
            p56_too_old = True
            days_past_p56 = (today - behavior_window_end).days
            p56_reason = (
                f'❌ TOO OLD FOR P56 — P56 behavior window ended '
                f'{behavior_window_end.strftime("%Y-%m-%d")} '
                f'({days_past_p56} days ago, animal is {age_today}d old). '
                f'Unusable for P56.'
            )
        else:
            first_wednesday = next_wednesday(behavior_window_start)

            if first_wednesday is None:
                p56_reason = 'Cannot calculate P56 behavior date'
            elif first_wednesday > behavior_window_end:
                p56_reason = 'No Wednesday falls within the P42–P49 window'
            elif first_wednesday < today:
                p56_too_old = True
                p56_reason = (
                    f'❌ TOO OLD FOR P56 — P56 window '
                    f'({first_wednesday.strftime("%Y-%m-%d")}) has passed '
                    f'(animal is {age_today}d old). Unusable for P56.'
                )
            elif first_wednesday in full_dates_set:
                p56_reason = (
                    f'Wednesday {first_wednesday.strftime("%Y-%m-%d")} '
                    f'is at capacity — cannot schedule P56'
                )
            else:
                p56_eligible = True
                p56_reason = f'Eligible: {first_wednesday.strftime("%A, %Y-%m-%d")}'
                behavior_suggested = first_wednesday

            if first_wednesday is not None and first_wednesday <= behavior_window_end:
                p56_age_at_behavior_days = (first_wednesday - birth_date).days
                p56_age_at_behavior_months = round(p56_age_at_behavior_days / 30.44)
                p56_harvest_calc = first_wednesday + timedelta(days=P56_HARVEST_OFFSET_FROM_BEHAVIOR)
                p56_age_at_harvest_days = (p56_harvest_calc - birth_date).days
                p56_age_at_harvest_months = round(p56_age_at_harvest_days / 30.44)

        # Combined unusable note
        if p14_too_old and p56_too_old:
            combined_unusable_note = (
                f'⛔ UNUSABLE FOR BOTH TIMEPOINTS — '
                f'Too old for P14 (was due {p14_harvest.strftime("%Y-%m-%d")}) '
                f'AND too old for P56 (window ended {behavior_window_end.strftime("%Y-%m-%d")}). '
                f'Animal is {age_today} days old.'
            )
        elif p14_too_old and not p56_too_old:
            combined_unusable_note = 'Too old for P14 only — P56 may still be viable'
        elif p56_too_old and not p14_too_old:
            combined_unusable_note = 'Too old for P56 only — P14 still viable'
        else:
            combined_unusable_note = ''

        eligibility.append({
            **base_record,
            'Birth_Date': birth_date.strftime('%Y-%m-%d'),
            'Age_Today_Days': age_today,
            'P14_Eligible': p14_eligible,
            'P14_Too_Old': p14_too_old,
            'P14_Reason': p14_reason,
            'P14_Date': p14_harvest if p14_eligible else None,
            'P14_Age_At_Harvest_Days': p14_age_at_harvest_days,
            'P14_Age_At_Harvest_Months': p14_age_at_harvest_months,
            'P56_Eligible': p56_eligible,
            'P56_Too_Old': p56_too_old,
            'P56_Reason': p56_reason,
            'P56_Behavior_Date': behavior_suggested if p56_eligible else None,
            'P56_Harvest_Date': (
                behavior_suggested + timedelta(days=P56_HARVEST_OFFSET_FROM_BEHAVIOR)
                if p56_eligible and behavior_suggested else None
            ),
            'P56_Age_At_Behavior_Days': p56_age_at_behavior_days,
            'P56_Age_At_Behavior_Months': p56_age_at_behavior_months,
            'P56_Age_At_Harvest_Days': p56_age_at_harvest_days,
            'P56_Age_At_Harvest_Months': p56_age_at_harvest_months,
            'Unusable_Note': combined_unusable_note,
        })

    logger.info(f"Eligibility check complete: {len(eligibility)} animals processed")
    return pd.DataFrame(eligibility)


# ============================================================================
# ANIMAL ASSIGNMENT
# ============================================================================

def assign_animals_smart(eligibility_df: pd.DataFrame, remaining_needs: Dict) -> pd.DataFrame:
    logger.info("Assigning animals to timepoints...")

    if eligibility_df.empty:
        logger.warning("assign_animals_smart: eligibility_df is empty — no animals to assign")
        print("\n⚠️  WARNING: No animals passed eligibility checks.")
        print("    Check the diagnostic output above to see where animals were filtered.")
        return pd.DataFrame()

    required_cols = ['Strain', 'Sex', 'Genotype', 'Birth_Date', 'Animal_Name']
    missing_cols = [c for c in required_cols if c not in eligibility_df.columns]
    if missing_cols:
        logger.error(
            f"assign_animals_smart: eligibility_df missing columns: {missing_cols}. "
            f"Available: {eligibility_df.columns.tolist()}"
        )
        raise KeyError(
            f"eligibility_df is missing required columns: {missing_cols}. "
            f"Available columns: {eligibility_df.columns.tolist()}"
        )

    eligibility_df = eligibility_df.sort_values(
        ['Strain', 'Sex', 'Genotype', 'Birth_Date', 'Animal_Name']
    ).reset_index(drop=True)

    eligibility_df['breeding_type'] = eligibility_df['Strain'].apply(get_strain_breeding_type)

    is_super = eligibility_df['Strain'].apply(is_super_priority_strain)
    is_prio = eligibility_df['Strain'].apply(is_priority_strain)
    bt = eligibility_df['breeding_type']

    # Het×Het Hom: rarest (1/4 yield) — absolute highest priority
    is_hxh_hom = (bt == 'HetxHet') & eligibility_df['Genotype'].apply(
        lambda g: canonicalize_genotype(g) == GENOTYPE_HOM
    )
    # Het×Het Het: second highest (1/2 yield from Het×Het cross)
    is_hxh_het = (bt == 'HetxHet') & ~is_hxh_hom

    tier_hxh_hom = eligibility_df[is_hxh_hom].copy()                                          # #1: HetxHet Hom
    tier_hxh_het = eligibility_df[is_hxh_het].copy()                                          # #2: HetxHet Het
    tier0a = eligibility_df[is_super & (bt == 'Half')  & ~is_hxh_hom & ~is_hxh_het].copy()   # #3: Super Half
    tier0b = eligibility_df[is_super & (bt == 'All')   & ~is_hxh_hom & ~is_hxh_het].copy()   # #4: Super All
    tier1  = eligibility_df[is_prio & ~is_super & (bt == 'Half')  & ~is_hxh_hom & ~is_hxh_het].copy()
    tier2  = eligibility_df[is_prio & ~is_super & (bt == 'All')   & ~is_hxh_hom & ~is_hxh_het].copy()
    tier3  = eligibility_df[~is_prio & (bt == 'Half')  & ~is_hxh_hom & ~is_hxh_het].copy()
    tier4  = eligibility_df[~is_prio & (bt == 'All')   & ~is_hxh_hom & ~is_hxh_het].copy()
    tier5  = eligibility_df[~is_prio & (bt == 'Unknown') & ~is_hxh_hom & ~is_hxh_het].copy()

    all_assignments = []
    # Track the +1 flex Perfusion slot per strain per timepoint (either sex)
    # Once both sexes hit their Perfusion quota, one extra animal of either sex is allowed
    perfusion_flex_used: Dict[tuple, bool] = {}  # (strain_key, timepoint) → used

    tier_names = [
        "🔴 HET×HET HOM — #1 PRIORITY (1/4 yield)",
        "🔴 HET×HET HET — #2 PRIORITY (1/2 yield from Het×Het)",
        "🔴 SUPER PRIORITY - Half (Het×WT)",
        "🔴 SUPER PRIORITY - All (Hom×Hom)",
        "Priority - Half (Het×WT)",
        "Priority - All (Hom×Hom)",
        "Standard - Half (Het×WT)",
        "Standard - All (Hom×Hom)",
        "Standard - Unknown"
    ]

    for tier_num, (tier_name, animals_batch) in enumerate(
        zip(tier_names, [tier_hxh_hom, tier_hxh_het, tier0a, tier0b, tier1, tier2, tier3, tier4, tier5])
    ):
        if animals_batch.empty:
            continue

        logger.info(f"Tier {tier_num}: {tier_name} — {len(animals_batch)} animals")

        animals_batch = animals_batch.copy()
        animals_batch['is_het'] = animals_batch['Genotype'].apply(is_heterozygous)
        animals_batch = animals_batch.sort_values('is_het', ascending=False)

        p56_candidates = animals_batch[animals_batch['P56_Eligible']].copy()
        p56_blocked_by_full_date = animals_batch[
            (~animals_batch['P56_Eligible']) &
            (animals_batch['P56_Reason'].str.contains('capacity', na=False, case=False))
        ].copy()

        p56_assignments = []
        p56_fallback = []

        if len(p56_candidates) > 0:
            p56_candidates['Genotype'] = p56_candidates['Genotype'].apply(
                lambda g: g if g in _CANONICAL_GENOTYPES else canonicalize_genotype(g)
            )

        grouped = (
            p56_candidates.groupby(['Strain', 'Genotype', 'Sex', 'P56_Behavior_Date'])
            if len(p56_candidates) > 0 else []
        )

        sorted_groups = sorted(
            grouped,
            key=lambda x: (
                not is_heterozygous(x[0][1]),
                str(x[0][0]) if x[0][0] is not None else '',
                str(x[0][2]) if x[0][2] is not None else '',
                to_date(x[0][3]) if x[0][3] is not None else date.min,
                str(x[0][1]) if x[0][1] is not None else ''
            )
        ) if len(p56_candidates) > 0 else []

        unschedulable = []

        for (strain, genotype, sex, behavior_date), group in sorted_groups:
            group_sorted = group.sort_values('Animal_Name').copy()
            animals = group_sorted.to_dict('records')

            if not group_has_quota(strain, sex, 'P56', remaining_needs):
                strain_key_flex = resolve_strain_key(strain, '', remaining_needs)
                flex_key = (strain_key_flex, 'P56')
                # Check if flex already consumed in a previous run (completed > target)
                already_over = False
                if remaining_needs and strain_key_flex in remaining_needs and strain_key_flex not in _B6_STRAINS_UPPER:
                    perf_m = remaining_needs[strain_key_flex]['P56']['Male']['Perfusion']
                    perf_f = remaining_needs[strain_key_flex]['P56']['Female']['Perfusion']
                    total_completed = perf_m.get('completed', 0) + perf_f.get('completed', 0)
                    total_target    = perf_m.get('target', 5)    + perf_f.get('target', 5)
                    # Flex is used if total meets the +1 cap, OR if either sex already
                    # exceeds its per-sex base target (meaning the flex went to that sex)
                    male_over   = perf_m.get('completed', 0) > perf_m.get('target', 5)
                    female_over = perf_f.get('completed', 0) > perf_f.get('target', 5)
                    already_over = total_completed >= (total_target + 1) or male_over or female_over
                # Allow flex only if not B6, not already over, and not yet used this run
                if (strain_key_flex not in _B6_STRAINS_UPPER
                        and not already_over
                        and not perfusion_flex_used.get(flex_key, False)):
                    perfusion_flex_used[flex_key] = True
                    flex_animals = animals[:CONFIG['CAGE_SIZE']]
                    for animal in flex_animals:
                        animal['_flex_slot'] = True
                        p56_assignments.append({
                            **animal,
                            'Assigned_Timepoint': 'P56',
                            'Assignment_Reason': 'P56 flex slot (+1 over quota)',
                        })
                    for animal in animals[CONFIG['CAGE_SIZE']:]:
                        animal['_quota_limited_complete_group'] = True
                        p56_fallback.append(animal)
                else:
                    for animal in animals:
                        animal['_quota_limited_complete_group'] = True
                        animal['_incomplete_group'] = False
                        p56_fallback.append(animal)
                continue

            num_complete_groups = len(animals) // CONFIG['CAGE_SIZE']

            # Batch-effect guard: only ONE cage of a given strain + sex + age may
            # run in a behavior round. A second cage matching all three is a
            # confound, so cap at one. The first cage survives; the rest fall into
            # `leftover` below, become Unschedulable, and are released to
            # 'Available' in Climb by release_unusable_to_available().
            # 3 males and 3 females of the same age is fine — different sex.
            if num_complete_groups > 1:
                print(f"    Batch-effect cap: {strain} {sex} had "
                      f"{num_complete_groups} cages this round — keeping 1, "
                      f"releasing {(num_complete_groups - 1) * CONFIG['CAGE_SIZE']} to Available")
                num_complete_groups = 1

            strain_key = str(strain).strip().upper()
            if remaining_needs and strain_key in remaining_needs and strain_key not in _B6_STRAINS_UPPER:
                needs = remaining_needs[strain_key]['P56'][sex]
                total_needed = needs['MERFISH']['needed'] + needs['RNAseq']['needed'] + needs['Perfusion']['needed']
                max_groups_by_quota = (total_needed + CONFIG['CAGE_SIZE'] - 1) // CONFIG['CAGE_SIZE']
                num_complete_groups = min(num_complete_groups, max_groups_by_quota)

            for i in range(num_complete_groups * CONFIG['CAGE_SIZE']):
                p56_assignments.append({
                    **animals[i],
                    'Assigned_Timepoint': 'P56',
                    'Assignment_Reason': 'Complete cage group of 3',
                })

            leftover = animals[num_complete_groups * CONFIG['CAGE_SIZE']:]
            if leftover:
                birth_groups = defaultdict(list)
                for animal in leftover:
                    birth_groups[animal.get('Birth_ID', 'Unknown')].append(animal)
                for bid, ba in birth_groups.items():
                    is_complete = len(ba) >= CONFIG['CAGE_SIZE']
                    for animal in ba:
                        animal['_quota_limited_complete_group'] = is_complete
                        animal['_incomplete_group'] = not is_complete
                        if not is_complete:
                            # Incomplete group — unschedulable, no P14 fallback
                            unschedulable.append({
                                **animal,
                                'Assigned_Timepoint': 'Unschedulable',
                                'Assignment_Reason': f'Incomplete P56 group ({len(ba)} of {CONFIG["CAGE_SIZE"]})',
                            })
                        else:
                            p56_fallback.append(animal)

        # Handle P56 blocked by full date
        if len(p56_blocked_by_full_date) > 0:
            p56_blocked_by_full_date = p56_blocked_by_full_date.copy()
            p56_blocked_by_full_date['Genotype'] = p56_blocked_by_full_date['Genotype'].apply(
                lambda g: g if g in _CANONICAL_GENOTYPES else canonicalize_genotype(g)
            )

            def get_p56_behavior_date(row):
                birth_d = to_date(row['Birth_Date'])
                if birth_d is None:
                    return None
                sched_dates = calculate_schedule_dates(birth_d)
                if sched_dates is None:
                    return None
                return next_wednesday(sched_dates['p56_behavior_window_start'])

            p56_blocked_by_full_date['P56_Behavior_Date_Calc'] = p56_blocked_by_full_date.apply(
                get_p56_behavior_date, axis=1
            )

            for (strain, genotype, sex, behavior_date), group in p56_blocked_by_full_date.groupby(
                ['Strain', 'Genotype', 'Sex', 'P56_Behavior_Date_Calc']
            ):
                animals = group.sort_values('Animal_Name').to_dict('records')
                num_complete_groups = len(animals) // CONFIG['CAGE_SIZE']

                if num_complete_groups > 0:
                    for animal in animals:
                        if animal['P14_Eligible']:
                            animal['_full_date_complete'] = True
                            animal['_incomplete_group'] = False
                            animal['_quota_limited_complete_group'] = False
                            p56_fallback.append(animal)
                        else:
                            unschedulable.append({
                                **animal,
                                'Assigned_Timepoint': 'Unschedulable',
                                'Assignment_Reason': (
                                    f'P56 date at capacity '
                                    f'({behavior_date.strftime("%Y-%m-%d") if behavior_date else "?"}). '
                                    f'P14 unavailable: {animal["P14_Reason"]}'
                                ),
                            })
                else:
                    for animal in animals:
                        unschedulable.append({
                            **animal,
                            'Assigned_Timepoint': 'Unschedulable',
                            'Assignment_Reason': (
                                f'Incomplete P56 group ({len(animals)} of {CONFIG["CAGE_SIZE"]}); '
                                f'P14 unavailable: {animal["P14_Reason"]}'
                                if not animal['P14_Eligible']
                                else f'Incomplete P56 group ({len(animals)} of {CONFIG["CAGE_SIZE"]})'
                            ),
                        })

        # P14 fallback
        p14_assignments = []
        p14_quota_used: Dict[tuple, int] = {}

        for animal in sorted(p56_fallback, key=lambda x: not is_heterozygous(x.get('Genotype', ''))):
            # Incomplete groups never get reassigned to P14 — just unschedulable
            if animal.get('_incomplete_group'):
                unschedulable.append({
                    **animal,
                    'Assigned_Timepoint': 'Unschedulable',
                    'Assignment_Reason': 'Incomplete P56 group — fewer than 3 animals available',
                })
                continue

            if not animal.get('P14_Eligible', False):
                if animal.get('_quota_limited_complete_group'):
                    reason_prefix = 'P56 quota filled (complete cage not needed)'
                elif animal.get('_full_date_complete'):
                    reason_prefix = 'P56 date at capacity'
                else:
                    reason_prefix = 'Incomplete P56 group'
                unschedulable.append({
                    **animal,
                    'Assigned_Timepoint': 'Unschedulable',
                    'Assignment_Reason': f'{reason_prefix}; P14 unavailable: {animal.get("P14_Reason", "Unknown")}',
                })
                continue

            strain   = animal.get('Strain', '')
            sex      = animal.get('Sex', '')
            genotype = animal.get('Genotype', '')
            strain_key = resolve_strain_key(strain, genotype, remaining_needs)

            # Check quota with decrement tracking
            has_quota = False
            if not remaining_needs or strain_key not in remaining_needs:
                has_quota = True  # untracked / B6 — always allow
            else:
                needs = remaining_needs[strain_key]['P14'][sex]
                total_needed = (needs['Perfusion']['needed'] +
                                needs['MERFISH']['needed'] +
                                needs['RNAseq']['needed'])
                used = p14_quota_used.get((strain_key, sex), 0)
                if used < total_needed:
                    has_quota = True
                    p14_quota_used[(strain_key, sex)] = used + 1
                else:
                    # Check flex slot — triggers as soon as this sex's quota is full
                    flex_key = (strain_key, 'P14')
                    already_over = False
                    if strain_key in remaining_needs:
                        perf_m = remaining_needs[strain_key]['P14']['Male']['Perfusion']
                        perf_f = remaining_needs[strain_key]['P14']['Female']['Perfusion']
                        total_completed = perf_m.get('completed', 0) + perf_f.get('completed', 0)
                        total_target = perf_m.get('target', 5) + perf_f.get('target', 5)
                        already_over = total_completed >= total_target
                    if not already_over and not perfusion_flex_used.get(flex_key, False):
                        perfusion_flex_used[flex_key] = True
                        has_quota = True

            if has_quota:
                if animal.get('_quota_limited_complete_group'):
                    reason = 'P56 quota filled for strain — reassigned to P14'
                elif animal.get('_full_date_complete'):
                    reason = 'P56 date at capacity — reassigned to P14'
                else:
                    reason = 'Incomplete P56 group — reassigned to P14'
                p14_assignments.append({
                    **animal,
                    'Assigned_Timepoint': 'P14',
                    'Assignment_Reason': reason,
                })
            else:
                if animal.get('_quota_limited_complete_group'):
                    unsched = 'P56 quota filled; P14 quota also filled'
                elif animal.get('_full_date_complete'):
                    unsched = 'P56 date at capacity; P14 quota also filled'
                else:
                    unsched = 'Incomplete P56 group; P14 quota also filled'
                unschedulable.append({
                    **animal,
                    'Assigned_Timepoint': 'Unschedulable',
                    'Assignment_Reason': unsched,
                })

        # P14-only animals — stop assigning when quota is met
        p14_only = animals_batch[
            animals_batch['P14_Eligible'] &
            ~animals_batch['P56_Eligible'] &
            ~animals_batch['P56_Reason'].str.contains('capacity', na=False, case=False)
        ].copy().sort_values('is_het', ascending=False)

        for idx2, animal in p14_only.iterrows():
            strain   = animal['Strain']
            sex      = animal['Sex']
            genotype = animal.get('Genotype', '')
            strain_key = resolve_strain_key(strain, genotype, remaining_needs)

            if not remaining_needs or strain_key not in remaining_needs:
                # B6 controls or untracked — always schedule
                p14_assignments.append({
                    **animal.to_dict(),
                    'Assigned_Timepoint': 'P14',
                    'Assignment_Reason': f'P14 only (P56: {animal["P56_Reason"]})',
                })
                continue

            needs = remaining_needs[strain_key]['P14'][sex]
            total_needed = (needs['Perfusion']['needed'] +
                            needs['MERFISH']['needed'] +
                            needs['RNAseq']['needed'])
            used = p14_quota_used.get((strain_key, sex), 0)

            if used < total_needed:
                p14_quota_used[(strain_key, sex)] = used + 1
                p14_assignments.append({
                    **animal.to_dict(),
                    'Assigned_Timepoint': 'P14',
                    'Assignment_Reason': f'P14 only (P56: {animal["P56_Reason"]})',
                })
            else:
                # Check flex slot — triggers as soon as this sex's quota is full
                flex_key = (strain_key, 'P14')
                already_over = False
                if remaining_needs and strain_key in remaining_needs:
                    perf_m = remaining_needs[strain_key]['P14']['Male']['Perfusion']
                    perf_f = remaining_needs[strain_key]['P14']['Female']['Perfusion']
                    total_completed = perf_m.get('completed', 0) + perf_f.get('completed', 0)
                    total_target = perf_m.get('target', 5) + perf_f.get('target', 5)
                    already_over = total_completed >= total_target
                if (remaining_needs and strain_key in remaining_needs
                        and not already_over
                        and not perfusion_flex_used.get(flex_key, False)):
                    perfusion_flex_used[flex_key] = True
                    p14_assignments.append({
                        **animal.to_dict(),
                        'Assigned_Timepoint': 'P14',
                        'Assignment_Reason': 'P14 flex slot (+1 over quota)',
                    })
                else:
                    unschedulable.append({
                        **animal.to_dict(),
                        'Assigned_Timepoint': 'Unschedulable',
                        'Assignment_Reason': f'P14 quota filled for {strain} {sex}',
                    })

        # Neither eligible
        neither = animals_batch[
            ~animals_batch['P14_Eligible'] &
            ~animals_batch['P56_Eligible'] &
            ~animals_batch['P56_Reason'].str.contains('capacity', na=False, case=False)
        ].copy()

        for idx2, animal in neither.iterrows():
            unschedulable.append({
                **animal.to_dict(),
                'Assigned_Timepoint': 'Unschedulable',
                'Assignment_Reason': f'P14: {animal["P14_Reason"]}; P56: {animal["P56_Reason"]}',
            })

        all_assignments.extend(p56_assignments + p14_assignments + unschedulable)

    logger.info(f"Assignment complete: {len(all_assignments)} animals")
    return pd.DataFrame(all_assignments)


# ============================================================================
# B6/B6N MONTHLY MINIMUM ENFORCEMENT
# ============================================================================

def enforce_b6_monthly_minimum(assignments_df: pd.DataFrame,
                                eligibility_df: pd.DataFrame,
                                remaining_needs: Dict) -> pd.DataFrame:
    min_per_month = CONFIG.get('B6_MIN_PER_MONTH', 3)
    if min_per_month <= 0:
        return assignments_df

    if assignments_df.empty:
        return assignments_df

    logger.info(f"Enforcing B6/B6N minimum of {min_per_month}/month after quota...")

    scheduled = assignments_df[
        assignments_df['Assigned_Timepoint'].isin(['P14', 'P56'])
    ].copy()

    def harvest_month(row):
        tp = row.get('Assigned_Timepoint', '')
        if tp == 'P14':
            d = to_date(row.get('P14_Date'))
        elif tp == 'P56':
            d = to_date(row.get('P56_Harvest_Date'))
        else:
            d = None
        return (d.year, d.month) if d else None

    scheduled['_harvest_month'] = scheduled.apply(harvest_month, axis=1)
    scheduled = scheduled[scheduled['_harvest_month'].notna()]

    all_harvest_months = sorted(scheduled['_harvest_month'].unique())
    if not all_harvest_months:
        logger.info("No scheduled harvest months found — skipping B6/B6N minimum check")
        return assignments_df

    b6_scheduled = scheduled[scheduled['Strain'].apply(is_b6_strain)].copy()
    b6_per_month: Dict[Tuple, int] = {}
    for month in all_harvest_months:
        b6_per_month[month] = int((b6_scheduled['_harvest_month'] == month).sum())

    logger.info("B6/B6N current scheduled counts by month:")
    for month, count in sorted(b6_per_month.items()):
        logger.info(f"  {month[0]}-{month[1]:02d}: {count} (min required: {min_per_month})")

    already_scheduled_names = set(
        assignments_df[
            assignments_df['Assigned_Timepoint'].isin(['P14', 'P56'])
        ]['Animal_Name'].tolist()
    )

    if len(eligibility_df) > 0:
        b6_pool = eligibility_df[
            eligibility_df['Strain'].apply(is_b6_strain)
        ].copy()
        b6_pool = b6_pool[~b6_pool['Animal_Name'].isin(already_scheduled_names)].copy()
    else:
        b6_pool = pd.DataFrame()

    b6_unschedulable = assignments_df[
        (assignments_df['Assigned_Timepoint'] == 'Unschedulable') &
        assignments_df['Strain'].apply(is_b6_strain)
    ].copy()

    logger.info(f"B6/B6N pool: {len(b6_pool)} eligible not yet scheduled, "
                f"{len(b6_unschedulable)} currently unschedulable")

    new_rows: List[Dict] = []
    added_names: set = set()

    for month in all_harvest_months:
        current_count = b6_per_month.get(month, 0)
        shortfall = min_per_month - current_count

        if shortfall <= 0:
            logger.info(f"  {month[0]}-{month[1]:02d}: already has {current_count} >= {min_per_month} — OK")
            continue

        logger.info(f"  {month[0]}-{month[1]:02d}: needs {shortfall} more B6/B6N (has {current_count})")
        added_this_month = 0

        if len(b6_pool) > 0:
            p14_candidates = b6_pool[b6_pool['P14_Eligible'] == True].copy()

            for _, candidate in p14_candidates.iterrows():
                if added_this_month >= shortfall:
                    break
                name = candidate['Animal_Name']
                if name in added_names:
                    continue

                p14_date = to_date(candidate.get('P14_Date'))
                if p14_date is None:
                    continue

                candidate_month = (p14_date.year, p14_date.month)
                if candidate_month != month:
                    continue

                new_rows.append({
                    **candidate.to_dict(),
                    'Assigned_Timepoint': 'P14',
                    'Assignment_Reason': (
                        f'B6/B6N monthly minimum top-up '
                        f'(month {month[0]}-{month[1]:02d} had only {current_count}, '
                        f'min={min_per_month})'
                    ),
                    'Harvest_Type': 'Perfusion',
                    'Priority': 'B6_MIN',
                    'Strain_Priority': 'B6/B6N Control',
                    'Genotype_Priority': 'B6/B6N',
                })
                added_names.add(name)
                added_this_month += 1

            if added_this_month < shortfall:
                p56_candidates = b6_pool[b6_pool['P56_Eligible'] == True].copy()

                # Group by behavior date — only add complete groups of CAGE_SIZE
                p56_by_date = {}
                for _, candidate in p56_candidates.iterrows():
                    name = candidate['Animal_Name']
                    if name in added_names:
                        continue
                    p56_harvest = to_date(candidate.get('P56_Harvest_Date'))
                    if p56_harvest is None:
                        bhv = to_date(candidate.get('P56_Behavior_Date'))
                        if bhv:
                            p56_harvest = bhv + timedelta(days=P56_HARVEST_OFFSET_FROM_BEHAVIOR)
                    if p56_harvest is None:
                        continue
                    candidate_month = (p56_harvest.year, p56_harvest.month)
                    if candidate_month != month:
                        continue
                    bhv_date = str(candidate.get('P56_Behavior_Date', ''))
                    if bhv_date not in p56_by_date:
                        p56_by_date[bhv_date] = []
                    p56_by_date[bhv_date].append(candidate)

                for bhv_date, group in p56_by_date.items():
                    if added_this_month >= shortfall:
                        break
                    if len(group) < CONFIG['CAGE_SIZE']:
                        logger.info(f"  Skipping B6/B6N P56 group on {bhv_date} — only {len(group)} animals (need {CONFIG['CAGE_SIZE']})")
                        continue
                    for candidate in group[:CONFIG['CAGE_SIZE']]:
                        if added_this_month >= shortfall:
                            break
                        name = candidate['Animal_Name']
                        bhv = to_date(candidate.get('P56_Behavior_Date'))
                        p56_harvest = to_date(candidate.get('P56_Harvest_Date'))
                        if p56_harvest is None and bhv:
                            p56_harvest = bhv + timedelta(days=P56_HARVEST_OFFSET_FROM_BEHAVIOR)
                        new_rows.append({
                            **candidate.to_dict(),
                            'Assigned_Timepoint': 'P56',
                            'Assignment_Reason': (
                                f'B6/B6N monthly minimum top-up '
                                f'(month {month[0]}-{month[1]:02d} had only {current_count}, '
                                f'min={min_per_month})'
                            ),
                            'Harvest_Type': 'Perfusion',
                            'Priority': 'B6_MIN',
                            'Strain_Priority': 'B6/B6N Control',
                            'Genotype_Priority': 'B6/B6N',
                            'P56_Behavior_Date': bhv,
                            'P56_Harvest_Date': p56_harvest,
                        })
                        added_names.add(name)
                        added_this_month += 1

        if added_this_month < shortfall and len(b6_unschedulable) > 0:
            for _, candidate in b6_unschedulable.iterrows():
                if added_this_month >= shortfall:
                    break
                name = candidate['Animal_Name']
                if name in added_names:
                    continue

                if candidate.get('P14_Eligible', False):
                    p14_date = to_date(candidate.get('P14_Date'))
                    if p14_date and (p14_date.year, p14_date.month) == month:
                        new_rows.append({
                            **candidate.to_dict(),
                            'Assigned_Timepoint': 'P14',
                            'Assignment_Reason': (
                                f'B6/B6N monthly minimum top-up from unschedulable pool '
                                f'(month {month[0]}-{month[1]:02d}, min={min_per_month})'
                            ),
                            'Harvest_Type': 'Perfusion',
                            'Priority': 'B6_MIN',
                            'Strain_Priority': 'B6/B6N Control',
                            'Genotype_Priority': 'B6/B6N',
                        })
                        added_names.add(name)
                        added_this_month += 1
                        continue

                if candidate.get('P56_Eligible', False):
                    p56_harvest = to_date(candidate.get('P56_Harvest_Date'))
                    if p56_harvest and (p56_harvest.year, p56_harvest.month) == month:
                        new_rows.append({
                            **candidate.to_dict(),
                            'Assigned_Timepoint': 'P56',
                            'Assignment_Reason': (
                                f'B6/B6N monthly minimum top-up from unschedulable pool '
                                f'(month {month[0]}-{month[1]:02d}, min={min_per_month})'
                            ),
                            'Harvest_Type': 'Perfusion',
                            'Priority': 'B6_MIN',
                            'Strain_Priority': 'B6/B6N Control',
                            'Genotype_Priority': 'B6/B6N',
                        })
                        added_names.add(name)
                        added_this_month += 1

        if added_this_month > 0:
            logger.info(f"  → Added {added_this_month} B6/B6N top-up animals for {month[0]}-{month[1]:02d}")
        else:
            logger.warning(
                f"  ⚠️ Could not find enough B6/B6N animals for {month[0]}-{month[1]:02d} "
                f"(added {added_this_month}/{shortfall})"
            )

    if new_rows:
        top_up_df = pd.DataFrame(new_rows)
        updated_assignments = assignments_df[
            ~(
                (assignments_df['Animal_Name'].isin(added_names)) &
                (assignments_df['Assigned_Timepoint'] == 'Unschedulable')
            )
        ].copy()
        all_cols = updated_assignments.columns.tolist()
        for col in all_cols:
            if col not in top_up_df.columns:
                top_up_df[col] = None
        updated_assignments = pd.concat(
            [updated_assignments, top_up_df[all_cols]], ignore_index=True
        )
        logger.info(f"B6/B6N minimum enforcement: added {len(new_rows)} top-up animals")
        return updated_assignments

    logger.info("B6/B6N minimum enforcement: no top-up needed or no animals available")
    return assignments_df


# ============================================================================
# HARVEST ASSIGNMENT GUI
# ============================================================================

# ============================================================================
# HARVEST ASSIGNMENT GUI
# ============================================================================

import tkinter as tk
from tkinter import ttk, messagebox
import copy as _copy


def _compute_auto_types(schedulable_df, remaining_needs):
    """
    Suggest harvest type per animal based on remaining quota needs.
    - MERFISH/RNAseq/Perfusion: fills the remaining quota for that type
    - Perfusion/MERFISH/RNAseq NB: incomplete group — no behavior, direct harvest
    - Perfusion: first over-quota animal per strain/timepoint (flex slot, 5+1 rule)
    - Extra: all quotas filled AND flex slot already used
    B6/B6NJ check quota needs before defaulting to Perfusion.
    """
    working = _copy.deepcopy(remaining_needs)
    result = {}
    flex_used = {}  # (strain_key, timepoint) → bool

    # Pre-compute group sizes by (Strain, Sex, P56_Behavior_Date) for NB flagging
    p56_group_sizes: Dict[tuple, int] = {}
    for _, row in schedulable_df[schedulable_df['Assigned_Timepoint'] == 'P56'].iterrows():
        key = (
            str(row.get('Strain', '')).strip(),
            str(row.get('Sex', '')).strip(),
            str(row.get('P56_Behavior_Date', '')).strip(),
        )
        p56_group_sizes[key] = p56_group_sizes.get(key, 0) + 1

    sorted_df = schedulable_df.sort_values(
        ['Assigned_Timepoint', 'Strain', 'Sex', 'Animal_Name']
    ).reset_index(drop=True)

    for _, row in sorted_df.iterrows():
        name       = str(row.get('Animal_Name', '')).strip()
        strain     = row.get('Strain', '')
        sex        = row.get('Sex', '')
        timepoint  = row.get('Assigned_Timepoint', '')
        genotype   = row.get('Genotype', '')
        strain_key = resolve_strain_key(strain, genotype, working)

        # Check if this P56 animal is in an incomplete group
        nb_flag = False
        if timepoint == 'P56':
            grp_key = (
                str(strain).strip(),
                str(sex).strip(),
                str(row.get('P56_Behavior_Date', '')).strip(),
            )
            if p56_group_sizes.get(grp_key, 0) < CONFIG['CAGE_SIZE']:
                nb_flag = True
        elif timepoint == 'P14' and row.get('_dns_nb_candidate', False):
            # P14 animal DNS'd with no P56 path — bypass quota, always Perfusion NB
            result[name] = 'Perfusion NB'
            continue

        # B6/B6NJ — check quota first, then default to Perfusion
        if strain_key in _B6_STRAINS_UPPER:
            if strain_key in working:
                needs = working[strain_key].get(timepoint, {}).get(sex, {})
                if needs.get('MERFISH', {}).get('needed', 0) > 0:
                    result[name] = 'MERFISH'
                    needs['MERFISH']['needed'] -= 1
                    continue
                elif needs.get('RNAseq', {}).get('needed', 0) > 0:
                    result[name] = 'RNAseq'
                    needs['RNAseq']['needed'] -= 1
                    continue
            result[name] = 'Perfusion'
            continue

        if strain_key not in working:
            result[name] = 'Perfusion NB' if nb_flag else 'Extra'
            continue

        needs = working[strain_key].get(timepoint, {}).get(sex, {})
        if not needs:
            result[name] = 'Extra'
            continue

        if needs['MERFISH']['needed'] > 0:
            base = 'MERFISH'
            needs['MERFISH']['needed'] -= 1
        elif needs['RNAseq']['needed'] > 0:
            base = 'RNAseq'
            needs['RNAseq']['needed'] -= 1
        elif needs['Perfusion']['needed'] > 0:
            base = 'Perfusion'
            needs['Perfusion']['needed'] -= 1
        else:
            # Quota filled for this sex — offer flex slot (5+1 rule)
            flex_key = (strain_key, timepoint)
            already_over = False
            if strain_key in working and timepoint in working.get(strain_key, {}):
                perf_m = working[strain_key][timepoint]['Male']['Perfusion']
                perf_f = working[strain_key][timepoint]['Female']['Perfusion']
                total_completed = perf_m.get('completed', 0) + perf_f.get('completed', 0)
                total_target = perf_m.get('target', 5) + perf_f.get('target', 5)
                # Flex is consumed only when the +1 cap is met, or when one sex
                # has already gone past its base target (i.e. the flex went there).
                # 5M + 5F leaves the flex OPEN — neither sex is at 6 yet.
                # Must match the rule in create_p56_schedule; they disagreed before.
                male_over   = perf_m.get('completed', 0) > perf_m.get('target', 5)
                female_over = perf_f.get('completed', 0) > perf_f.get('target', 5)
                already_over = (total_completed >= (total_target + 1)
                                or male_over or female_over)
            if not already_over and not flex_used.get(flex_key, False):
                flex_used[flex_key] = True
                base = 'Perfusion'  # flex slot
            else:
                base = 'Extra'

        # Extras follow the pen: if no animal in the group gets behavior,
        # everything in it is NB — including the Extras ('Extra NB').
        result[name] = f'{base} NB' if nb_flag else base

    return result


def _compute_quota_status(selections, schedulable_df, remaining_needs):
    """
    Compare current GUI selections against remaining_needs.
    Returns a list of (strain, timepoint, sex, harvest_type, needed, selected, status_str).
    """
    if not remaining_needs:
        return []

    # Count selected per (strain_key, timepoint, sex, harvest_type)
    counts = {}
    df_map = {
        str(r.get('Animal_Name', '')).strip(): r
        for _, r in schedulable_df.iterrows()
    }

    for name, htype in selections.items():
        if htype in ('Do Not Schedule', 'Extra'):
            continue
        row = df_map.get(name)
        if row is None:
            continue
        # Strip NB suffix for quota counting purposes
        base_htype = htype.replace(' NB', '').strip()
        strain_key = resolve_strain_key(row.get('Strain', ''), row.get('Genotype', ''), remaining_needs)
        timepoint  = str(row.get('Assigned_Timepoint', '')).strip()
        sex        = str(row.get('Sex', '')).strip()
        key = (strain_key, timepoint, sex, base_htype)
        counts[key] = counts.get(key, 0) + 1

    present_combos = set()
    for name, htype in selections.items():
        if htype in ('Do Not Schedule', 'Extra'):
            continue
        row = df_map.get(name)
        if row is None:
            continue
        sk = resolve_strain_key(row.get('Strain', ''), row.get('Genotype', ''), remaining_needs)
        tp = str(row.get('Assigned_Timepoint', '')).strip()
        sx = str(row.get('Sex', '')).strip()
        if tp in ('P14', 'P56'):
            present_combos.add((sk, tp, sx))

    rows = []
    for strain_key, timepoints in remaining_needs.items():
        for timepoint, sexes in timepoints.items():
            for sex, htypes in sexes.items():
                # Skip entirely if no animals from this group are in this run
                if (strain_key, timepoint, sex) not in present_combos:
                    continue
                for htype, info in htypes.items():
                    needed   = info['needed']
                    selected = counts.get((strain_key, timepoint, sex, htype), 0)
                    if needed == 0 and selected == 0:
                        continue  # not interesting
                    if selected == needed:
                        status = '✓ Match'
                    elif selected > needed:
                        status = f'↑ {selected - needed} over'
                    else:
                        status = f'↓ {needed - selected} short'
                    rows.append((
                        strain_key, timepoint, sex, htype,
                        needed, selected, status
                    ))
    return rows


def prompt_harvest_assignments_gui(assignments_df, remaining_needs, prior_selections=None):
    """
    Block the pipeline and show a GUI letting the user review and confirm
    harvest type assignments for every scheduled animal.

    Returns a dict { animal_name: harvest_type_or_None }
    where None / missing means 'auto-assign as normal'.
    'Do Not Schedule' is returned as the string 'DO_NOT_SCHEDULE'
    so the caller can act on it.

    If tkinter is unavailable the function returns {} and the pipeline
    continues with auto-assignment.
    """
    try:
        import tkinter as _tk_test
        _tk_test.Tk().destroy()
    except Exception:
        print("  ⚠ tkinter not available — skipping harvest assignment GUI.")
        return {}

    BASE_TYPES = ['Perfusion', 'MERFISH', 'RNAseq']
    OPTION_COLORS   = {
        'P14 Perfusion':    '#d4edda',
        'P56 Perfusion':    '#d4edda',
        'P14 MERFISH':      '#cce5ff',
        'P56 MERFISH':      '#cce5ff',
        'P14 RNAseq':       '#fff3cd',
        'P56 RNAseq':       '#fff3cd',
        'Extra':            '#e8d5f5',
        'NB Perfusion':     '#a8d5b5',
        'NB MERFISH':       '#7ab8f5',
        'NB RNAseq':        '#f5d76e',
        'NB Extra':         '#d9bce8',
        'Do Not Schedule':  '#f8d7da',
    }

    # Display-label <-> internal harvest-type translation
    _DISPLAY_TO_INTERNAL = {
        'P14 Perfusion':   'Perfusion',
        'P56 Perfusion':   'Perfusion',
        'NB Perfusion':    'Perfusion NB',
        'P14 MERFISH':     'MERFISH',
        'P56 MERFISH':     'MERFISH',
        'NB MERFISH':      'MERFISH NB',
        'P14 RNAseq':      'RNAseq',
        'P56 RNAseq':      'RNAseq',
        'NB RNAseq':       'RNAseq NB',
        'Extra':           'Extra',
        'NB Extra':        'Extra NB',
        'Do Not Schedule': 'Do Not Schedule',
    }

    def _display_to_internal(display):
        return _DISPLAY_TO_INTERNAL.get(display, display)

    def _internal_to_display(internal, timepoint):
        mapping = {
            'Perfusion':    f'{timepoint} Perfusion',
            'MERFISH':      f'{timepoint} MERFISH',
            'RNAseq':       f'{timepoint} RNAseq',
            'Perfusion NB': 'NB Perfusion',
            'MERFISH NB':   'NB MERFISH',
            'RNAseq NB':    'NB RNAseq',
            'Extra':        'Extra',
            'DO_NOT_SCHEDULE': 'Do Not Schedule',
        }
        return mapping.get(internal, internal)
    STATUS_COLORS = {
        '✓ Match':  '#c3e6cb',
        '↑':        '#ffeeba',
        '↓':        '#f5c6cb',
    }

    def _animal_options(strain_key: str, timepoint: str, sex: str, nb_flag: bool) -> List[str]:
        """Return the ordered dropdown options for one animal.

        Includes only harvest types where quota is still needed, plus NB variants
        when the animal is in an incomplete group or is a P14 animal.
        Extra and Do Not Schedule are always present.

        Args:
            strain_key: Upper-cased strain key used in remaining_needs.
            timepoint:  'P14' or 'P56'.
            sex:        'Male' or 'Female'.
            nb_flag:    True when the animal's group is incomplete (P56) or
                        the timepoint is P14 (no behaviour session exists at P14).

        Returns:
            Ordered list of option strings for the Combobox.
        """
        if strain_key in _B6_STRAINS_UPPER:
            # B6/B6NJ always fills — offer all regular types
            available = list(BASE_TYPES)
        elif strain_key in remaining_needs:
            needs = remaining_needs.get(strain_key, {}).get(timepoint, {}).get(sex, {})
            available = [t for t in BASE_TYPES if needs.get(t, {}).get('needed', 0) > 0]
            if not available:
                # All quota filled — still offer Perfusion for the flex slot
                available = ['Perfusion']
        else:
            # Untracked strain — offer all regular types
            available = list(BASE_TYPES)

        # Regular options include the timepoint prefix
        options = [f'{timepoint} {t}' for t in available]

        # NB options only for incomplete P56 groups — P14 never gets NB variants
        if nb_flag and timepoint == 'P56':
            for t in available:
                options.append(f'NB {t}')

        if nb_flag and timepoint == 'P56':
            options += ['NB Extra']
        options += ['Extra', 'Do Not Schedule']
        return options

    # Only show schedulable animals (P14 / P56), skip Unschedulable
    schedulable = assignments_df[
        assignments_df['Assigned_Timepoint'].isin(['P14', 'P56'])
    ].copy()

    if len(schedulable) == 0:
        return {}

    # Compute auto-suggested types
    auto_types = _compute_auto_types(schedulable, remaining_needs)

    # ── Build the window ──────────────────────────────────────────────────────
    root = tk.Tk()
    root.title("Harvest Assignment Review")
    root.configure(bg='#f0f0f0')
    root.resizable(True, True)

    # Keep track of the user's final decision
    confirmed = {'result': None}

    # ── Header ────────────────────────────────────────────────────────────────
    header_frame = tk.Frame(root, bg='#2c3e50', pady=12)
    header_frame.pack(fill='x')
    tk.Label(
        header_frame,
        text="Harvest Assignment Review",
        font=('Helvetica', 16, 'bold'),
        bg='#2c3e50', fg='white'
    ).pack()
    tk.Label(
        header_frame,
        text=f"{len(schedulable)} animals ready to schedule  •  Review assignments below, make any changes, then confirm.",
        font=('Helvetica', 10),
        bg='#2c3e50', fg='#bdc3c7'
    ).pack()

    # ── Main pane (left = animal table, right = quota panel) ──────────────────
    main_frame = tk.Frame(root, bg='#f0f0f0')
    main_frame.pack(fill='both', expand=True, padx=12, pady=8)

    # ── LEFT: Animal table ────────────────────────────────────────────────────
    left_frame = tk.LabelFrame(
        main_frame, text=" Animals to Schedule ",
        font=('Helvetica', 11, 'bold'),
        bg='#f0f0f0', fg='#2c3e50', padx=6, pady=6
    )
    left_frame.pack(side='left', fill='both', expand=True, padx=(0, 6))

    # Column headers
    headers = ['Animal Name', 'Strain', 'Sex', 'Date', 'Group', 'Harvest Type']
    col_widths = [18, 12, 8, 10, 8, 20]

    hdr_row = tk.Frame(left_frame, bg='#2c3e50')
    hdr_row.pack(fill='x')
    for h, w in zip(headers, col_widths):
        tk.Label(
            hdr_row, text=h, width=w, anchor='w',
            font=('Helvetica', 9, 'bold'),
            bg='#2c3e50', fg='white', padx=4, pady=4
        ).pack(side='left')

    # ── Color key ─────────────────────────────────────────────────────────────
    key_frame = tk.Frame(left_frame, bg='#e8e8e8', pady=3)
    key_frame.pack(fill='x')
    tk.Label(
        key_frame, text='Row color = selected harvest type:',
        font=('Helvetica', 8, 'italic'), bg='#e8e8e8', fg='#555555', padx=6
    ).pack(side='left')
    for label, color in OPTION_COLORS.items():
        swatch = tk.Frame(key_frame, bg=color, width=12, height=12,
                          relief='solid', bd=1)
        swatch.pack(side='left', padx=(4, 1), pady=2)
        swatch.pack_propagate(False)
        tk.Label(
            key_frame, text=label,
            font=('Helvetica', 8), bg='#e8e8e8', fg='#333333', padx=2
        ).pack(side='left')

    # Scrollable rows
    canvas = tk.Canvas(left_frame, bg='#f0f0f0', highlightthickness=0)
    scrollbar = ttk.Scrollbar(left_frame, orient='vertical', command=canvas.yview)
    rows_frame = tk.Frame(canvas, bg='#f0f0f0')

    rows_frame.bind(
        '<Configure>',
        lambda e: canvas.configure(scrollregion=canvas.bbox('all'))
    )
    canvas.create_window((0, 0), window=rows_frame, anchor='nw')
    canvas.configure(yscrollcommand=scrollbar.set)

    canvas.pack(side='left', fill='both', expand=True)
    scrollbar.pack(side='right', fill='y')

    # Mouse-wheel scroll
    def _on_mousewheel(event):
        canvas.yview_scroll(int(-1 * (event.delta / 120)), 'units')
    canvas.bind_all('<MouseWheel>', _on_mousewheel)

    # Sort: timepoint → date → strain → sex → name (keeps housing groups together)
    def _harvest_sort_key(row):
        tp  = str(row.get('Assigned_Timepoint', ''))
        d   = str(row.get('P14_Date' if tp == 'P14' else 'P56_Harvest_Date', '') or '')
        sex = str(row.get('Sex', ''))
        return (tp, d, str(row.get('Strain', '')), sex, str(row.get('Animal_Name', '')))

    sorted_rows = sorted(schedulable.to_dict('records'), key=_harvest_sort_key)

    # Pre-compute P56 group sizes by (Strain, Sex, P56_Behavior_Date)
    p56_group_sizes_gui: Dict[tuple, int] = {}
    for row in sorted_rows:
        if str(row.get('Assigned_Timepoint', '')).strip() == 'P56':
            key = (
                str(row.get('Strain', '')).strip(),
                str(row.get('Sex', '')).strip(),
                str(row.get('P56_Behavior_Date', '')).strip(),
            )
            p56_group_sizes_gui[key] = p56_group_sizes_gui.get(key, 0) + 1

    # Store StringVars so we can read them later
    selection_vars   = {}   # name → StringVar
    selection_values = {}   # name → current value (always in sync, avoids tkinter canvas StringVar decouple bug)
    selection_menus  = {}   # name → Combobox widget (for direct .get() at confirm time)
    row_frames       = {}   # name → tk.Frame (for recoloring)

    def _on_type_change_cb(name, combobox, frame):
        """Called on <<ComboboxSelected>> — reads directly from combobox widget."""
        val = combobox.get()
        selection_values[name] = val   # store in plain dict (reliable)
        color = OPTION_COLORS.get(val, '#ffffff')
        frame.configure(bg=color)
        for w in frame.winfo_children():
            try:
                w.configure(bg=color)
            except Exception:
                pass
        _refresh_quota_panel()

    def _on_type_change(name, var, frame):
        val = var.get()
        selection_values[name] = val
        color = OPTION_COLORS.get(val, '#ffffff')
        frame.configure(bg=color)
        for w in frame.winfo_children():
            try:
                w.configure(bg=color)
            except Exception:
                pass
        _refresh_quota_panel()

    for i, row in enumerate(sorted_rows):
        name      = str(row.get('Animal_Name', '')).strip()
        strain    = str(row.get('Strain', '')).strip()
        sex       = str(row.get('Sex', '')).strip()
        timepoint = str(row.get('Assigned_Timepoint', '')).strip()
        genotype  = row.get('Genotype', '')
        strain_key = resolve_strain_key(strain, genotype, remaining_needs)

        # nb_flag: incomplete P56 group, or any P14 animal (no behaviour at P14)
        nb_flag = False
        if timepoint == 'P56':
            grp_key_nb = (str(strain).strip(), str(sex).strip(),
                          str(row.get('P56_Behavior_Date', '')).strip())
            if p56_group_sizes_gui.get(grp_key_nb, 0) < CONFIG['CAGE_SIZE']:
                nb_flag = True
        # P14 animals never get NB variants — P14 IS the no-behaviour harvest

        animal_options = _animal_options(strain_key, timepoint, sex, nb_flag)

        # Translate auto-suggested internal string to display label
        auto_internal = auto_types.get(name, 'Perfusion')
        default = _internal_to_display(auto_internal, timepoint)
        if default not in animal_options:
            animal_options.insert(0, default)

        var = tk.StringVar(value=default)
        selection_vars[name] = var
        selection_values[name] = default   # seed plain-dict copy

        bg = '#ffffff' if i % 2 == 0 else '#f7f7f7'
        frame = tk.Frame(rows_frame, bg=bg)
        frame.pack(fill='x')
        row_frames[name] = frame

        # Pick the relevant date: P14 -> harvest date, P56 -> behavior date
        # A missing date here means a scheduling logic error — flag it clearly
        if timepoint == 'P14':
            raw_date = str(row.get('P14_Date', '') or '')
        else:
            raw_date = str(row.get('P56_Behavior_Date', '') or '')
        try:
            from datetime import datetime as _dt
            display_date = _dt.strptime(raw_date, '%Y-%m-%d').strftime('%m/%d/%y')
        except Exception:
            display_date = '⚠ NO DATE'
            logger.warning(f"Animal {name} ({timepoint}) is scheduled but has no date — check scheduling logic")

        # Group size indicator
        if timepoint == 'P56':
            grp_key = (str(strain).strip(), str(sex).strip(), raw_date)
            grp_size = p56_group_sizes_gui.get(grp_key, 0)
            if grp_size >= CONFIG['CAGE_SIZE']:
                group_label = f'✓ {grp_size}'
            elif grp_size == 0:
                group_label = '⚠ 0'   # genuine problem — no animals found for this group
            else:
                group_label = str(grp_size)  # incomplete but normal — NB options handle it
        else:
            group_label = '—'  # P14 doesn't need groups

        for val, w in zip([name, strain, sex, display_date, group_label], col_widths[:5]):
            lbl = tk.Label(
                frame, text=val, width=w, anchor='w',
                font=('Helvetica', 9), bg=bg, padx=4, pady=3
            )
            lbl.pack(side='left')

        menu = ttk.Combobox(
            frame, textvariable=var,
            values=animal_options,
            state='readonly', width=col_widths[5] - 2
        )
        menu.pack(side='left', padx=2, pady=2)
        menu.set(default)  # explicitly set display value after pack
        selection_menus[name] = menu  # store widget ref for direct read at confirm
        menu.bind('<<ComboboxSelected>>',
                  lambda e, n=name, f=frame, m=menu: _on_type_change_cb(n, m, f))


        # Apply initial color
        c = OPTION_COLORS.get(default, bg)
        frame.configure(bg=c)
        for w in frame.winfo_children():
            try:
                w.configure(bg=c)
            except Exception:
                pass

    # ── RIGHT: Quota comparison panel ─────────────────────────────────────────
    right_frame = tk.LabelFrame(
        main_frame, text=" Quota Comparison ",
        font=('Helvetica', 11, 'bold'),
        bg='#f0f0f0', fg='#2c3e50', padx=6, pady=6
    )
    right_frame.pack(side='right', fill='y', padx=(6, 0))
    right_frame.pack_propagate(False)
    right_frame.configure(width=340)

    quota_inner = tk.Frame(right_frame, bg='#f0f0f0')
    quota_inner.pack(fill='both', expand=True)

    quota_header_cols = ['Strain', 'TP', 'Sex', 'Type', 'Need', 'Sel', 'Status']
    quota_col_widths  = [10, 4,  5,  8, 5, 5, 10]

    qhdr = tk.Frame(quota_inner, bg='#2c3e50')
    qhdr.pack(fill='x')
    for h, w in zip(quota_header_cols, quota_col_widths):
        tk.Label(
            qhdr, text=h, width=w, anchor='w',
            font=('Helvetica', 8, 'bold'),
            bg='#2c3e50', fg='white', padx=2, pady=3
        ).pack(side='left')

    quota_rows_frame = tk.Frame(quota_inner, bg='#f0f0f0')
    quota_rows_frame.pack(fill='both', expand=True)

    def _refresh_quota_panel():
        for w in quota_rows_frame.winfo_children():
            w.destroy()

        current = dict(selection_values)  # use plain-dict copy, not StringVar (avoids canvas decouple bug)
        quota_data = _compute_quota_status(current, schedulable, remaining_needs)

        if not quota_data:
            tk.Label(
                quota_rows_frame,
                text="No quota tracking data\navailable.",
                font=('Helvetica', 9), bg='#f0f0f0', fg='#7f8c8d',
                justify='center'
            ).pack(pady=20)
            return

        all_match = all(r[6] == '✓ Match' for r in quota_data)
        summary_color = '#c3e6cb' if all_match else '#ffeeba'
        summary_text  = '✓ All quotas satisfied' if all_match else '⚠ Some quotas need attention'
        tk.Label(
            quota_rows_frame,
            text=summary_text,
            font=('Helvetica', 9, 'bold'),
            bg=summary_color, fg='#155724' if all_match else '#856404',
            pady=4
        ).pack(fill='x', pady=(0, 4))

        for j, (strain_k, tp, sex, htype, needed, selected, status) in enumerate(quota_data):
            bg = '#ffffff' if j % 2 == 0 else '#f7f7f7'
            # Tint by status
            if status == '✓ Match':
                bg = '#eafaf1'
            elif status.startswith('↑'):
                bg = '#fef9e7'
            elif status.startswith('↓'):
                bg = '#fdf0ef'

            qrow = tk.Frame(quota_rows_frame, bg=bg)
            qrow.pack(fill='x')
            for val, w in zip(
                [strain_k, tp, sex, htype, str(needed), str(selected), status],
                quota_col_widths
            ):
                tk.Label(
                    qrow, text=val, width=w, anchor='w',
                    font=('Helvetica', 8), bg=bg, padx=2, pady=2
                ).pack(side='left')

    _refresh_quota_panel()

    # ── Footer buttons ────────────────────────────────────────────────────────
    footer = tk.Frame(root, bg='#ecf0f1', pady=8)
    footer.pack(fill='x', padx=12)

    def _reset_to_auto():
        for name, var in selection_vars.items():
            internal = auto_types.get(name, 'Perfusion')
            row_match = schedulable[schedulable['Animal_Name'].astype(str).str.strip() == name]
            row_tp = str(row_match['Assigned_Timepoint'].iloc[0]).strip() if not row_match.empty else 'P56'
            val = _internal_to_display(internal, row_tp)
            var.set(val)
            selection_values[name] = val  # keep plain-dict in sync
            frame = row_frames[name]
            c = OPTION_COLORS.get(var.get(), '#ffffff')
            frame.configure(bg=c)
            for w in frame.winfo_children():
                try:
                    w.configure(bg=c)
                except Exception:
                    pass
        _refresh_quota_panel()

    def _confirm():
        # Build current from selection_values (kept in sync by _on_type_change_cb)
        # Fall back to menu.get() for any animal not yet interacted with
        current = {}
        for name in selection_menus:
            val = selection_values.get(name)
            if not val:
                val = selection_menus[name].get() or 'P56 Perfusion'
            current[name] = val
        # Translate display labels to internal strings for all downstream logic
        current = {n: _display_to_internal(v) for n, v in current.items()}
        # Debug log all selections
        for _n, _h in sorted(current.items()):
            logger.info(f"CONFIRM: {_n} → {_h}")

        # ── Quota check — only warn on genuine over-quota (not B6, not flex slot) ────
        quota_data = _compute_quota_status(current, schedulable, remaining_needs)
        mismatches = [
            r for r in quota_data
            if '↑' in r[6]
            and str(r[0]).strip().upper() not in _B6_STRAINS_UPPER
            and r[5] > r[4] + 1  # more than 1 over quota (flex slot is allowed)
        ]

        if mismatches:
            lines = '\n'.join(
                f"  {r[0]} {r[1]} {r[2]} {r[3]}: need {r[4]}, selected {r[5]} ({r[6]})"
                for r in mismatches[:8]
            )
            if len(mismatches) > 8:
                lines += f"\n  ... and {len(mismatches) - 8} more"
            proceed = messagebox.askyesno(
                "Quota Mismatch",
                f"The following assignments don't match the tracking sheet:\n\n"
                f"{lines}\n\n"
                f"Proceed anyway?",
                icon='warning'
            )
            if not proceed:
                return

        # ── Group-of-3 check (Harvest + Extra count per Birth_ID, P56 only) ──
        # Build a map of animal name -> row for quick lookup
        df_map = {
            str(r.get('Animal_Name', '')).strip(): r
            for _, r in schedulable.iterrows()
        }
        # Group animals by Birth_ID, P56 only
        birth_groups = {}
        for name, htype in current.items():
            row = df_map.get(name)
            if row is None:
                continue
            if str(row.get('Assigned_Timepoint', '')).strip() != 'P56':
                continue
            birth_id = str(row.get('Birth_ID', 'Unknown')).strip()
            if birth_id not in birth_groups:
                birth_groups[birth_id] = []
            birth_groups[birth_id].append((name, htype))

        confirmed['result'] = current
        root.destroy()

    def _cancel():
        if messagebox.askyesno(
            "Skip Review",
            "Skip the harvest review and use auto-assignments for all animals?",
            icon='question'
        ):
            confirmed['result'] = {}
            root.destroy()

    tk.Button(
        footer, text="↺  Reset to Suggested",
        command=_reset_to_auto,
        font=('Helvetica', 10), bg='#95a5a6', fg='white',
        relief='flat', padx=12, pady=6, cursor='hand2'
    ).pack(side='left', padx=(0, 8))

    tk.Button(
        footer, text="Skip / Use Auto-Assignments",
        command=_cancel,
        font=('Helvetica', 10), bg='#bdc3c7', fg='#2c3e50',
        relief='flat', padx=12, pady=6, cursor='hand2'
    ).pack(side='left')

    tk.Button(
        footer, text="Confirm Assignments  →",
        command=_confirm,
        font=('Helvetica', 11, 'bold'), bg='#27ae60', fg='white',
        relief='flat', padx=16, pady=6, cursor='hand2'
    ).pack(side='right')

    # Size and center
    root.update_idletasks()
    w = min(root.winfo_screenwidth() - 80, 1100)
    h = min(root.winfo_screenheight() - 80, 700)
    x = (root.winfo_screenwidth()  - w) // 2
    y = (root.winfo_screenheight() - h) // 2
    root.geometry(f"{w}x{h}+{x}+{y}")
    root.minsize(700, 400)

    root.mainloop()

    # ── Process result ────────────────────────────────────────────────────────
    if confirmed['result'] is None:
        # Window was closed without confirming — treat as auto
        return {}

    # Convert 'Do Not Schedule' to sentinel; strip unchanged auto-assignments
    final = {}
    for name, htype in confirmed['result'].items():
        if htype == 'Do Not Schedule':
            final[name] = 'DO_NOT_SCHEDULE'
        else:
            final[name] = htype

    dns_count  = sum(1 for v in final.values() if v == 'DO_NOT_SCHEDULE')
    over_count = sum(1 for v in final.values() if v != 'DO_NOT_SCHEDULE')
    print(f"\n  ✓ Harvest review confirmed: {over_count} scheduled, {dns_count} skipped")
    return final


def assign_harvest_types(assignments_df: pd.DataFrame,
                         remaining_needs: Dict,
                         requirements: Dict,
                         harvest_overrides: Optional[Dict[str, str]] = None) -> pd.DataFrame:
    if assignments_df.empty:
        return assignments_df

    if harvest_overrides is None:
        harvest_overrides = {}

    if not remaining_needs:
        assignments_df = assignments_df.copy()
        assignments_df['Harvest_Type'] = assignments_df['Animal_Name'].map(harvest_overrides).fillna('Not Assigned')
        assignments_df['Priority'] = assignments_df['Animal_Name'].apply(
            lambda n: 'MANUAL' if n in harvest_overrides else 'Unknown'
        )
        assignments_df['Strain_Priority'] = 'Unknown'
        assignments_df['Genotype_Priority'] = 'Unknown'
        return assignments_df

    logger.info("Assigning harvest types...")
    working_needs = copy.deepcopy(remaining_needs)
    assignments_with_types = []

    sorted_df = assignments_df.sort_values(
        ['Strain', 'Sex', 'Assigned_Timepoint', 'Animal_Name']
    ).reset_index(drop=True)

    for idx, row in sorted_df.iterrows():
        strain    = row.get('Strain')
        sex       = row.get('Sex')
        genotype  = row.get('Genotype')
        timepoint = row.get('Assigned_Timepoint')
        name      = str(row.get('Animal_Name', '')).strip()

        is_prio = is_priority_strain(strain)
        strain_priority  = 'PRIORITY STRAIN' if is_prio else 'Standard'
        is_het           = is_heterozygous(genotype)
        genotype_priority = 'Het' if is_het else str(genotype) if genotype else 'Other'

        # ── Manual override — always wins, no quota consumed ─────────────────
        if name in harvest_overrides:
            manual_type = harvest_overrides[name]
            assignments_with_types.append({
                **row.to_dict(),
                'Harvest_Type':      manual_type,
                'Priority':          'MANUAL',
                'Strain_Priority':   strain_priority,
                'Genotype_Priority': genotype_priority,
            })
            logger.debug(f"Manual override applied: {name!r} → {manual_type}")
            continue

        if row.get('Priority') == 'B6_MIN':
            assignments_with_types.append({
                **row.to_dict(),
                'Strain_Priority':   row.get('Strain_Priority', 'B6/B6N Control'),
                'Genotype_Priority': row.get('Genotype_Priority', 'B6/B6N'),
            })
            continue

        if timepoint == 'Unschedulable':
            assignments_with_types.append({
                **row.to_dict(),
                'Harvest_Type':      'N/A',
                'Priority':          'N/A',
                'Strain_Priority':   strain_priority,
                'Genotype_Priority': genotype_priority,
            })
            continue

        strain_key = resolve_strain_key(strain, genotype, working_needs)

        if strain_key in _B6_STRAINS_UPPER:
            assignments_with_types.append({
                **row.to_dict(),
                'Harvest_Type':      'Perfusion',
                'Priority':          'B6_CONTROL',
                'Strain_Priority':   'B6/B6N Control',
                'Genotype_Priority': genotype_priority,
            })
            continue

        if strain_key not in working_needs:
            assignments_with_types.append({
                **row.to_dict(),
                'Harvest_Type':      'Not Tracked',
                'Priority':          'Unknown',
                'Strain_Priority':   strain_priority,
                'Genotype_Priority': genotype_priority,
            })
            continue

        needs = working_needs[strain_key].get(timepoint, {}).get(sex, {})

        if not needs:
            # Timepoint not found in working_needs (e.g. Unschedulable slipped through)
            assignments_with_types.append({
                **row.to_dict(),
                'Harvest_Type':      'N/A',
                'Priority':          'N/A',
                'Strain_Priority':   strain_priority,
                'Genotype_Priority': genotype_priority,
            })
            continue

        if needs['MERFISH']['needed'] > 0:
            harvest_type = 'MERFISH'
            priority = 'HIGH'
            needs['MERFISH']['needed'] -= 1
        elif needs['RNAseq']['needed'] > 0:
            harvest_type = 'RNAseq'
            priority = 'HIGH'
            needs['RNAseq']['needed'] -= 1
        elif needs['Perfusion']['needed'] > 0:
            harvest_type = 'Perfusion'
            priority = 'MEDIUM'
            needs['Perfusion']['needed'] -= 1
        else:
            harvest_type = 'COMPLETE (Quota Filled)'
            priority = 'NONE'

        assignments_with_types.append({
            **row.to_dict(),
            'Harvest_Type':      harvest_type,
            'Priority':          priority,
            'Strain_Priority':   strain_priority,
            'Genotype_Priority': genotype_priority,
        })

    overridden_count = sum(1 for r in assignments_with_types if r.get('Priority') == 'MANUAL')
    logger.info(f"Harvest types assigned: {len(assignments_with_types)} animals "
                f"({overridden_count} manually overridden)")
    return pd.DataFrame(assignments_with_types)


def check_capacity_and_reassign(assignments_df: pd.DataFrame,
                                remaining_needs: Dict) -> pd.DataFrame:
    if assignments_df.empty:
        return assignments_df

    logger.info("Checking Wednesday capacity...")

    p56_assigned = assignments_df[assignments_df['Assigned_Timepoint'] == 'P56'].copy()
    if p56_assigned.empty:
        return assignments_df

    p56_assigned = p56_assigned[p56_assigned['P56_Behavior_Date'].notna()].copy()
    try:
        p56_assigned['P56_Behavior_Date'] = p56_assigned['P56_Behavior_Date'].apply(to_date)
    except Exception as e:
        logger.warning(f"Error converting P56_Behavior_Date: {e}")
        return assignments_df

    p56_assigned = p56_assigned[p56_assigned['P56_Behavior_Date'].notna()].copy()
    if p56_assigned.empty:
        return assignments_df

    wednesday_counts = p56_assigned.groupby('P56_Behavior_Date').size().to_dict()
    overflow_animals = []
    kept_animals = []

    for behavior_date, group in p56_assigned.groupby('P56_Behavior_Date'):
        animals = group.to_dict('records')
        count = wednesday_counts[behavior_date]

        if count <= CONFIG['WEDNESDAY_CAPACITY']:
            kept_animals.extend(animals)
        else:
            logger.warning(f"Wednesday {behavior_date} over capacity: {count} > {CONFIG['WEDNESDAY_CAPACITY']}")
            animals_sorted = sorted(animals, key=lambda x: not is_heterozygous(x.get('Genotype', '')))
            kept_animals.extend(animals_sorted[:CONFIG['WEDNESDAY_CAPACITY']])
            overflow_animals.extend(animals_sorted[CONFIG['WEDNESDAY_CAPACITY']:])

    reassigned = []
    still_unschedulable = []

    for animal in overflow_animals:
        if not animal.get('P14_Eligible'):
            animal['Assigned_Timepoint'] = 'Unschedulable'
            animal['Assignment_Reason'] = (
                f'Wednesday over capacity (>{CONFIG["WEDNESDAY_CAPACITY"]}); '
                f'P14 unavailable: {animal.get("P14_Reason", "Unknown")}'
            )
            still_unschedulable.append(animal)
            continue

        strain = animal.get('Strain')
        sex = animal.get('Sex')
        if group_has_quota(strain, sex, 'P14', remaining_needs):
            animal['Assigned_Timepoint'] = 'P14'
            animal['Assignment_Reason'] = (
                f'Wednesday over capacity (>{CONFIG["WEDNESDAY_CAPACITY"]}) — reassigned to P14'
            )
            reassigned.append(animal)
        else:
            animal['Assigned_Timepoint'] = 'Unschedulable'
            animal['Assignment_Reason'] = (
                f'Wednesday over capacity (>{CONFIG["WEDNESDAY_CAPACITY"]}); '
                f'P14 quota also filled'
            )
            still_unschedulable.append(animal)

    other_assignments = assignments_df[assignments_df['Assigned_Timepoint'] != 'P56']
    all_cols = assignments_df.columns.tolist()

    frames = [pd.DataFrame(kept_animals), other_assignments,
              pd.DataFrame(reassigned), pd.DataFrame(still_unschedulable)]
    for i, frame in enumerate(frames):
        if len(frame) > 0:
            for col in all_cols:
                if col not in frame.columns:
                    frame[col] = None

    final = pd.concat(frames, ignore_index=True)
    logger.info(f"Capacity check: {len(reassigned)} reassigned to P14, {len(still_unschedulable)} unschedulable")
    return final


# ============================================================================
# GENOTYPE BLANK ANALYSIS
# ============================================================================

def _assess_genotype_worth_it(
    num_blanks: int,
    breeding_type: str,
    p14_available: bool,
    p56_available: bool,
    is_schedulable: bool,
    p14_date,
    p56_date,
    today: date,
    remaining_needs: Dict,
    strain: str,
    # ── new parameter ────────────────────────────────────────────────────────
    p56_group_size: int = 0,    # total blank animals sharing this P56 Wednesday
) -> Dict:
    """
    Assess whether genotyping blank animals is worth doing for P14 / P56.

    P14 worth  — uses num_blanks (the animals in *this* exact-birth-date group).
    P56 worth  — uses p56_group_size (all blanks that share the same behavior
                 Wednesday, which may span several birth dates).  Falls back to
                 num_blanks when p56_group_size is not supplied (0).

    Returns a dict with keys:
        P14_Worth_Genotyping : str
        P56_Worth_Genotyping : str
    """
    min_cage = CONFIG['CAGE_SIZE']  # typically 3

    # ── quota / B6 helpers ────────────────────────────────────────────────────
    def _quota_met(timepoint: str) -> bool:
        """Return True when all needs are zero AND the flex slot is already used."""
        if is_b6_strain(strain):
            return False            # B6/B6N never considered quota-met
        strain_upper = str(strain).strip().upper()
        if not remaining_needs or strain_upper not in remaining_needs:
            return False
        if timepoint not in remaining_needs[strain_upper]:
            return False
        tp_needs = remaining_needs[strain_upper][timepoint]
        all_zero = all(
            tp_needs[sex][ht]['needed'] == 0
            for sex in ['Male', 'Female']
            for ht in ['Perfusion', 'MERFISH', 'RNAseq']
        )
        if not all_zero:
            return False
        # Check if flex slot is still available (total completed < target*2 + 1)
        perf_m = tp_needs['Male']['Perfusion']
        perf_f = tp_needs['Female']['Perfusion']
        total_completed = perf_m.get('completed', 0) + perf_f.get('completed', 0)
        total_target = perf_m.get('target', 5) + perf_f.get('target', 5)
        flex_still_available = total_completed < total_target + 1
        return not flex_still_available  # only truly met if flex is also used

    # ── generic worth evaluator ───────────────────────────────────────────────
    def _worth(available: bool, timepoint: str, group_n: int) -> str:
        """
        Evaluate whether genotyping is worth it for one timepoint.

        Parameters
        ----------
        available : bool   — is the scheduling window still open?
        timepoint : str    — 'P14' or 'P56'
        group_n   : int    — number of blank animals in the relevant group
                             (birth-date group for P14, Wednesday group for P56)
        """
        if not available:
            return '❌ NO — window not available'

        if is_b6_strain(strain):
            if breeding_type == 'All':
                return '✅ YES — All usable (B6/B6N control)'
            return '✅ YES — B6/B6N control'

        if _quota_met(timepoint):
            return '⚠️ QUOTA MET — genotyping low priority'

        # Check if only the flex slot remains
        strain_upper = str(strain).strip().upper()
        if remaining_needs and strain_upper in remaining_needs and timepoint in remaining_needs[strain_upper]:
            tp_needs = remaining_needs[strain_upper][timepoint]
            all_zero = all(
                tp_needs[sex][ht]['needed'] == 0
                for sex in ['Male', 'Female']
                for ht in ['Perfusion', 'MERFISH', 'RNAseq']
            )
            if all_zero:
                # Normal quota full but flex slot still open — last chance to harvest
                return '🔴 HIGH PRIORITY — only flex slot remaining (last harvest opportunity)'

        # Expected usable animals from Mendelian ratios
        if breeding_type == 'All':
            expected = group_n
        elif breeding_type == 'Half':
            expected = group_n * 0.5
        else:
            expected = group_n * 0.5     # Unknown → assume Het×WT

        # For P56 the cage-size threshold matters: need ≥ CAGE_SIZE usable
        if timepoint == 'P56':
            if expected >= min_cage:
                return (
                    f'✅ YES — ~{int(round(expected))} usable expected '
                    f'from {group_n} animals in window '
                    f'(≥{min_cage} needed for a full cage)'
                )
            elif expected >= 1.0:
                return (
                    f'🟡 MAYBE — ~{int(round(expected))} usable expected '
                    f'from {group_n} animals in window '
                    f'(need {min_cage} for a full cage)'
                )
            else:
                return (
                    f'❌ UNLIKELY — <1 usable expected from {group_n} animals '
                    f'in window ({breeding_type} mating, need {min_cage} for a cage)'
                )
        else:
            # P14: simpler threshold — at least 1 usable is sufficient
            if expected >= 2.0:
                return f'✅ YES — ~{int(round(expected))} usable expected'
            elif expected >= 1.0:
                return f'🟡 MAYBE — ~{int(round(expected))} usable expected'
            else:
                return (
                    f'❌ UNLIKELY — <1 usable expected '
                    f'({group_n} blanks, {breeding_type} mating)'
                )

    # ── evaluate each timepoint with its own group size ───────────────────────
    effective_p56_n = p56_group_size if p56_group_size > 0 else num_blanks

    p14_worth = _worth(p14_available, 'P14', num_blanks)
    p56_worth = _worth(p56_available, 'P56', effective_p56_n)

    return {
        'P14_Worth_Genotyping': p14_worth,
        'P56_Worth_Genotyping': p56_worth,
    }


def analyze_blank_genotype_for_scheduling(
    blank_animals_group: List[Dict],
    strain: str,
    sex: str,
    birth_date: str,
    full_behavior_dates: Optional[List[date]] = None,
    remaining_needs: Optional[Dict] = None,
    p56_group_size: int = 0,    # total blank animals sharing this P56 Wednesday window
) -> Dict:
    """
    Analyse a group of blank-genotype animals from the same birth/strain/sex
    and predict scheduling viability.

    p56_group_size — when supplied, overrides num_blanks for the P56 worth-it
                     assessment so that animals with different birth dates that
                     map to the same Wednesday are evaluated together.
    """
    today               = datetime.now().date()
    num_blanks          = len(blank_animals_group)
    breeding_type       = get_strain_breeding_type(strain)
    birth_date_obj      = None
    scheduling_window   = 'Unknown'
    genotype_needed_by  = None
    p14_date            = None
    p56_behavior_date   = None
    is_schedulable      = False
    p14_available       = False
    p56_available       = False
    p14_valid           = False
    p56_valid           = False

    if blank_animals_group and 'Birth Date' in blank_animals_group[0]:
        birth_date_raw = blank_animals_group[0]['Birth Date']
        birth_date_obj = to_date(birth_date_raw)

    if birth_date_obj:
        dates = calculate_schedule_dates(birth_date_obj)
        if dates is None:
            scheduling_window = "Invalid birth date - cannot calculate scheduling windows"
            is_schedulable = False
        else:
            p14_date              = dates['p14_harvest']
            behavior_window_start = dates['p56_behavior_window_start']
            behavior_window_end   = dates['p56_behavior_window_end']
            p56_behavior_date     = next_wednesday(behavior_window_start)

            p14_valid     = is_valid_p14_day(p14_date)
            p14_in_future = p14_date > today
            p14_available = p14_valid and p14_in_future

            if p56_behavior_date is not None:
                p56_valid     = p56_behavior_date <= behavior_window_end
                p56_in_future = p56_behavior_date >= today
                p56_not_full  = True
                if full_behavior_dates and p56_behavior_date in full_behavior_dates:
                    p56_not_full = False
                p56_available = p56_valid and p56_in_future and p56_not_full
            else:
                p56_available = False

            possible_dates = []
            if p14_available:
                possible_dates.append(('P14', p14_date))
            if p56_available:
                possible_dates.append(('P56', p56_behavior_date))

            if possible_dates:
                earliest_type, earliest_date = min(possible_dates, key=lambda x: x[1])
                genotype_needed_by = earliest_date - timedelta(days=1)
                if len(possible_dates) == 2:
                    scheduling_window = (
                        f"P14 on {p14_date.strftime('%Y-%m-%d')} "
                        f"or P56 on {p56_behavior_date.strftime('%Y-%m-%d')}"
                    )
                else:
                    scheduling_window = (
                        f"{earliest_type} on {earliest_date.strftime('%Y-%m-%d')}"
                    )
            else:
                reasons = []
                if p14_date and p14_date <= today:
                    reasons.append(f"P14 window passed ({p14_date.strftime('%Y-%m-%d')})")
                elif p14_date and not p14_valid:
                    reasons.append(f"P14 on invalid day ({p14_date.strftime('%A')})")
                if p56_behavior_date and p56_behavior_date < today:
                    reasons.append(
                        f"P56 window passed ({p56_behavior_date.strftime('%Y-%m-%d')})"
                    )
                elif p56_behavior_date and not p56_valid:
                    reasons.append("P56 no valid Wednesday in window")
                elif (
                    p56_behavior_date and full_behavior_dates
                    and p56_behavior_date in full_behavior_dates
                ):
                    reasons.append(
                        f"P56 date full ({p56_behavior_date.strftime('%Y-%m-%d')})"
                    )
                elif p56_behavior_date is None:
                    reasons.append("Cannot calculate P56 behavior date")
                scheduling_window = (
                    "; ".join(reasons) if reasons else "No scheduling windows available"
                )
                genotype_needed_by = None

            is_schedulable = p14_available or p56_available
    else:
        is_schedulable = False

    worth_it = _assess_genotype_worth_it(
        num_blanks      = num_blanks,
        breeding_type   = breeding_type,
        p14_available   = p14_available,
        p56_available   = p56_available,
        is_schedulable  = is_schedulable,
        p14_date        = p14_date,
        p56_date        = p56_behavior_date,
        today           = today,
        remaining_needs = remaining_needs or {},
        strain          = strain,
        p56_group_size  = p56_group_size,   # ← Wednesday-level group size
    )

    prediction      = 'UNKNOWN'
    reason          = ''
    expected_usable = 0

    if breeding_type == 'All':
        expected_usable = num_blanks
        is_b6 = str(strain).strip().upper() in _B6_STRAINS_UPPER
        mating_desc = 'Inbred colony' if is_b6 else 'Hom×Hom cross'
        if not is_schedulable:
            prediction = 'NOT SCHEDULABLE'
            reason = (
                f"'All' STRAIN — {num_blanks} blank genotype(s) from birth {birth_date}. "
                f"ALL {num_blanks} expected usable ({mating_desc}) "
                f"BUT NOT SCHEDULABLE. Reason: {scheduling_window}"
            )
        elif genotype_needed_by and genotype_needed_by >= today:
            days_until = (genotype_needed_by - today).days
            urgency    = "URGENT" if days_until <= 7 else "HIGH PRIORITY"
            prediction = 'LIKELY USABLE'
            reason = (
                f"'All' STRAIN — {num_blanks} blank genotype(s) from birth {birth_date}. "
                f"ALL {num_blanks} expected usable ({mating_desc}). "
                f"{urgency}: Genotype by "
                f"{genotype_needed_by.strftime('%Y-%m-%d')} ({days_until} days) "
                f"for {scheduling_window}"
            )
        elif genotype_needed_by and genotype_needed_by < today:
            prediction = 'DEADLINE PASSED'
            reason = (
                f"[GENOTYPE DEADLINE PASSED] "
                f"'All' STRAIN — {num_blanks} blank genotype(s) from birth {birth_date}. "
                f"ALL expected usable ({mating_desc}) but genotyping deadline passed "
                f"({genotype_needed_by.strftime('%Y-%m-%d')}). {scheduling_window}"
            )
        else:
            prediction = 'LIKELY USABLE'
            reason = (
                f"'All' STRAIN — {num_blanks} blank genotype(s) from birth {birth_date}. "
                f"ALL {num_blanks} expected usable ({mating_desc}). "
                f"Genotype ASAP! {scheduling_window}"
            )

    elif breeding_type == 'Half':
        expected_hets = num_blanks * 0.5

        if not is_schedulable:
            prediction      = 'NOT SCHEDULABLE'
            expected_usable = 0
            reason = (
                f"'Half' STRAIN — {num_blanks} blank genotype(s) from birth {birth_date}. "
                f"~{int(round(expected_hets))} Hets expected (50% — Het×WT cross) "
                f"BUT NOT SCHEDULABLE. Reason: {scheduling_window}"
            )
        elif expected_hets >= 2.0:
            prediction      = 'LIKELY USABLE'
            expected_usable = int(round(expected_hets))
            if genotype_needed_by and genotype_needed_by >= today:
                days_until = (genotype_needed_by - today).days
                urgency    = "URGENT" if days_until <= 7 else "RECOMMEND"
                reason = (
                    f"'Half' STRAIN — {num_blanks} blank genotype(s) from birth {birth_date}. "
                    f"~{int(round(expected_hets))} of {num_blanks} expected Het "
                    f"(Het×WT cross — 50% usable). {urgency}: Genotype by "
                    f"{genotype_needed_by.strftime('%Y-%m-%d')} ({days_until} days) "
                    f"for {scheduling_window}"
                )
            elif genotype_needed_by and genotype_needed_by < today:
                prediction = 'DEADLINE PASSED'
                reason = (
                    f"[GENOTYPE DEADLINE PASSED] "
                    f"'Half' STRAIN — {num_blanks} blank genotype(s) from birth {birth_date}. "
                    f"~{int(round(expected_hets))} Hets expected but genotyping deadline "
                    f"passed ({genotype_needed_by.strftime('%Y-%m-%d')}). {scheduling_window}"
                )
            else:
                reason = (
                    f"'Half' STRAIN — {num_blanks} blank genotype(s) from birth {birth_date}. "
                    f"~{int(round(expected_hets))} of {num_blanks} expected Het "
                    f"(Het×WT cross — 50% usable). Genotype for scheduling! {scheduling_window}"
                )
        elif expected_hets >= 1.0:
            prediction      = 'POSSIBLY USABLE' if is_schedulable else 'NOT SCHEDULABLE'
            expected_usable = int(round(expected_hets)) if is_schedulable else 0
            if genotype_needed_by and genotype_needed_by >= today:
                days_until = (genotype_needed_by - today).days
                reason = (
                    f"'Half' STRAIN — {num_blanks} blank genotype(s) from birth {birth_date}. "
                    f"May contain ~{int(round(expected_hets))} Het "
                    f"(Het×WT cross — 50% usable). Consider genotyping by "
                    f"{genotype_needed_by.strftime('%Y-%m-%d')} ({days_until} days). "
                    f"{scheduling_window}"
                )
            elif genotype_needed_by and genotype_needed_by < today:
                prediction = 'DEADLINE PASSED'
                reason = (
                    f"[GENOTYPE DEADLINE PASSED] "
                    f"'Half' STRAIN — {num_blanks} blank genotype(s) from birth {birth_date}. "
                    f"May have had ~{int(round(expected_hets))} Het but genotyping deadline "
                    f"passed ({genotype_needed_by.strftime('%Y-%m-%d')}). {scheduling_window}"
                )
            else:
                reason = (
                    f"'Half' STRAIN — {num_blanks} blank genotype(s) from birth {birth_date}. "
                    f"May contain ~{int(round(expected_hets))} Het. "
                    f"Consider genotyping. {scheduling_window}"
                )
        else:
            prediction      = 'LIKELY WILD' if is_schedulable else 'NOT SCHEDULABLE'
            expected_usable = 0
            if genotype_needed_by and genotype_needed_by >= today:
                days_until = (genotype_needed_by - today).days
                reason = (
                    f"'Half' STRAIN — {num_blanks} blank genotype(s) from birth {birth_date}. "
                    f"Het×WT cross — ~50% usable. "
                    f"Low statistical likelihood of usable animals. "
                    f"Deadline: {genotype_needed_by.strftime('%Y-%m-%d')} ({days_until} days). "
                    f"{scheduling_window}"
                )
            elif genotype_needed_by and genotype_needed_by < today:
                prediction = 'DEADLINE PASSED'
                reason = (
                    f"[GENOTYPE DEADLINE PASSED] "
                    f"'Half' STRAIN — {num_blanks} blank genotype(s) from birth {birth_date}. "
                    f"50%/50% Het/Wild expected. Genotyping deadline passed "
                    f"({genotype_needed_by.strftime('%Y-%m-%d')}). {scheduling_window}"
                )
            else:
                reason = (
                    f"'Half' STRAIN — {num_blanks} blank genotype(s) from birth "
                    f"{birth_date}. 50% chance Het, 50% chance Wild. "
                    f"Low likelihood. {scheduling_window}"
                )
    elif breeding_type == 'HetxHet':
        expected_usable_all = int(round(num_blanks * 0.75))  # 3/4 usable (Hom + Het)
        expected_homs_hxh   = int(round(num_blanks * 0.25))  # 1/4 Hom
        expected_hets_hxh   = int(round(num_blanks * 0.50))  # 1/2 Het

        if not is_schedulable:
            prediction      = 'NOT SCHEDULABLE'
            expected_usable = 0
            reason = (
                f"Het×Het STRAIN — {num_blanks} blank genotype(s) from birth {birth_date}. "
                f"~{expected_usable_all} usable expected "
                f"(Hom:{expected_homs_hxh} + Het:{expected_hets_hxh}) "
                f"BUT NOT SCHEDULABLE. Reason: {scheduling_window}"
            )
        elif genotype_needed_by and genotype_needed_by >= today:
            days_until      = (genotype_needed_by - today).days
            urgency         = "URGENT" if days_until <= 7 else "HIGH PRIORITY"
            prediction      = 'LIKELY USABLE'
            expected_usable = expected_usable_all
            reason = (
                f"Het×Het STRAIN — {num_blanks} blank genotype(s) from birth {birth_date}. "
                f"~{expected_usable_all} usable expected "
                f"(Hom:{expected_homs_hxh} + Het:{expected_hets_hxh}). "
                f"{urgency}: Genotype by "
                f"{genotype_needed_by.strftime('%Y-%m-%d')} ({days_until} days) "
                f"for {scheduling_window}"
            )
        elif genotype_needed_by and genotype_needed_by < today:
            prediction      = 'DEADLINE PASSED'
            expected_usable = 0
            reason = (
                f"[GENOTYPE DEADLINE PASSED] "
                f"Het×Het STRAIN — {num_blanks} blank genotype(s) from birth {birth_date}. "
                f"~{expected_usable_all} usable expected but deadline passed "
                f"({genotype_needed_by.strftime('%Y-%m-%d')}). {scheduling_window}"
            )
        else:
            prediction      = 'LIKELY USABLE'
            expected_usable = expected_usable_all
            reason = (
                f"Het×Het STRAIN — {num_blanks} blank genotype(s) from birth {birth_date}. "
                f"~{expected_usable_all} usable expected "
                f"(Hom:{expected_homs_hxh} + Het:{expected_hets_hxh}). "
                f"Genotype ASAP — Hom animals are #1 priority! {scheduling_window}"
            )

    else:
        expected_hets = num_blanks * 0.5
        prediction    = 'UNKNOWN' if is_schedulable else 'NOT SCHEDULABLE'
        if not is_schedulable:
            expected_usable = 0
            reason = (
                f"UNKNOWN STRAIN — {num_blanks} blank genotype(s) from birth {birth_date}. "
                f"Not in priority list. NOT SCHEDULABLE. {scheduling_window}"
            )
        elif genotype_needed_by and genotype_needed_by >= today:
            days_until      = (genotype_needed_by - today).days
            expected_usable = int(round(expected_hets))
            reason = (
                f"UNKNOWN STRAIN — {num_blanks} blank genotype(s) from birth {birth_date}. "
                f"Het×WT cross — ~{int(round(expected_hets))} Hets). "
                f"Genotype by {genotype_needed_by.strftime('%Y-%m-%d')} ({days_until} days). "
                f"{scheduling_window}"
            )
        elif genotype_needed_by and genotype_needed_by < today:
            prediction      = 'DEADLINE PASSED'
            expected_usable = 0
            reason = (
                f"[GENOTYPE DEADLINE PASSED] "
                f"UNKNOWN STRAIN — {num_blanks} blank genotype(s) from birth {birth_date}. "
                f"Genotyping deadline passed "
                f"({genotype_needed_by.strftime('%Y-%m-%d')}). {scheduling_window}"
            )
        else:
            expected_usable = int(round(expected_hets))
            reason = (
                f"UNKNOWN STRAIN — {num_blanks} blank genotype(s) from birth {birth_date}. "
                f"Het×WT cross — ~{int(round(expected_hets))} usable. {scheduling_window}"
            )

    return {
        'prediction':           prediction,
        'reason':               reason,
        'expected_usable':      expected_usable,
        'breeding_type':        breeding_type,
        'genotype_needed_by':   genotype_needed_by.strftime('%Y-%m-%d') if genotype_needed_by else 'N/A',
        'days_until_deadline':  (
            (genotype_needed_by - today).days
            if genotype_needed_by and genotype_needed_by >= today else None
        ),
        'scheduling_window':    scheduling_window,
        'p14_date':             p14_date.strftime('%Y-%m-%d') if p14_date else 'N/A',
        'p56_date':             p56_behavior_date.strftime('%Y-%m-%d') if p56_behavior_date else 'N/A',
        'p14_available':        p14_available,
        'p56_available':        p56_available,
        'is_schedulable':       is_schedulable,
        'P14_Worth_Genotyping': worth_it['P14_Worth_Genotyping'],
        'P56_Worth_Genotyping': worth_it['P56_Worth_Genotyping'],
    }


def analyze_blank_genotypes_second_pass(
    blank_genotypes_df: pd.DataFrame,
    full_behavior_dates: List[date],
    remaining_needs: Optional[Dict] = None,
) -> pd.DataFrame:
    """
    Second pass: analyse blank-genotype animals with capacity info.

    Grouping strategy
    -----------------
    * P14 worth-it  — assessed per exact (birth_date, strain, sex) group,
                      because P14 scheduling is birth-date-specific.
    * P56 worth-it  — assessed per (p56_wednesday, strain, sex) group,
                      because all animals whose birth dates map to the same
                      Wednesday behavior session compete for the same cages.
                      Example: 18 females born 3/17-3/24 all map to the
                      5/6/26 Wednesday → evaluated as one group of 18,
                      yielding ~9 expected Hets which fills 3 cages → YES.

    NOTE: This function operates exclusively on blank-genotype animals.
          Genotyped animals (Het, Hom, Hemi, Inbred) are never passed here
          and are completely unaffected by this logic.
    """
    if blank_genotypes_df.empty:
        return pd.DataFrame()

    logger.info("Second pass: Analyzing blank genotypes with capacity info...")

    # ── Step 1: compute P56 Wednesday for every animal ───────────────────────
    def _p56_wednesday(birth_date_val) -> Optional[date]:
        bd = to_date(birth_date_val) if pd.notna(birth_date_val) else None
        return get_p56_behavior_wednesday(bd)

    blank_genotypes_df = blank_genotypes_df.copy()
    blank_genotypes_df['_p56_wed'] = blank_genotypes_df['Birth Date'].apply(_p56_wednesday)

    # ── Step 2: pre-compute P56 group sizes (wednesday, strain, sex) ─────────
    # A "P56 group" = all blank animals of the same strain & sex whose birth
    # dates map to the same behavior Wednesday.
    p56_group_sizes: Dict[Tuple, int] = {}
    for (p56_wed, strain, sex), grp in blank_genotypes_df.groupby(
        ['_p56_wed', 'Line (Short)', 'Sex'], dropna=False
    ):
        key = (p56_wed, str(strain), str(sex))
        p56_group_sizes[key] = len(grp)

    # ── Step 3: iterate over exact (birth_date, strain, sex) sub-groups ──────
    excluded = []
    grouped = list(
        blank_genotypes_df.groupby(['Birth Date', 'Line (Short)', 'Sex'], dropna=False)
    )
    if CONFIG['ENABLE_PROGRESS_BARS']:
        grouped = tqdm(grouped, desc="Analyzing blank genotype groups")

    for (birth_date_val, strain, sex), group in grouped:
        group_animals = group.to_dict('records')

        birth_date_str = (
            to_date(birth_date_val).strftime('%Y-%m-%d')
            if pd.notna(birth_date_val) and to_date(birth_date_val) is not None
            else 'Unknown Date'
        )

        # P56 group size: look up by the Wednesday this birth date maps to
        p56_wed = _p56_wednesday(birth_date_val)
        p56_key = (p56_wed, str(strain), str(sex))
        p56_group_n = p56_group_sizes.get(p56_key, len(group_animals))

        analysis = analyze_blank_genotype_for_scheduling(
            group_animals,
            strain,
            sex,
            birth_date_str,
            full_behavior_dates=full_behavior_dates,
            remaining_needs=remaining_needs,
            p56_group_size=p56_group_n,     # ← Wednesday-level group size
        )

        # ── build one output row per animal ───────────────────────────────────
        for animal in group_animals:
            excluded.append({
                'Animal_Name':              animal['Name'],
                'Birth_ID':                 animal.get('Birth ID', 'N/A'),
                'Birth_Date':               birth_date_str,
                'Strain':                   strain,
                'Sex':                      sex,
                'Genotype':                 GENOTYPE_BLANK,
                'Breeding_Type':            analysis['breeding_type'],
                'Prediction':               analysis['prediction'],
                'Expected_Usable_In_Group': analysis['expected_usable'],
                # ── P56 window group info (new columns) ────────────────────
                'P56_Window_Group_Size':    p56_group_n,
                'P56_Window_Wednesday':     (
                    p56_wed.strftime('%Y-%m-%d') if p56_wed else 'N/A'
                ),
                # ── deadline / scheduling ──────────────────────────────────
                'Genotype_Needed_By':       analysis['genotype_needed_by'],
                'Days_Until_Deadline':      (
                    analysis['days_until_deadline']
                    if analysis['days_until_deadline'] is not None
                    else 'N/A'
                ),
                'P14_Date':                 analysis['p14_date'],
                'P14_Worth_Genotyping':     analysis['P14_Worth_Genotyping'],
                'P56_Date':                 analysis['p56_date'],
                'P56_Worth_Genotyping':     analysis['P56_Worth_Genotyping'],
                'Scheduling_Window':        analysis['scheduling_window'],
                'Reason':                   analysis['reason'],
            })

    logger.info(f"Analyzed {len(excluded)} animals with blank genotypes")
    return pd.DataFrame(excluded)


# ============================================================================
# UNSCHEDULABLE REASON PARSING
# ============================================================================

def parse_unschedulable_reason(reason) -> Dict:
    """
    Parse a raw assignment reason string into structured fields for the
    Unschedulable report.
    """
    if reason is None:
        return {
            'Primary_Reason': 'Unknown',
            'P14_Status':     'Unknown',
            'P56_Status':     'Unknown',
            'Too_Old_For_P14': 'NO',
            'Too_Old_For_P56': 'NO',
            'Unusable_Both':   'NO',
            'Detail':          '',
        }

    r  = str(reason).strip()
    rl = r.lower()

    too_old_p14   = 'NO'
    too_old_p56   = 'NO'
    unusable_both = 'NO'
    primary       = 'Unknown'
    p14_status    = 'See detail'
    p56_status    = 'See detail'

    # ── [GENOTYPE DEADLINE PASSED] sentinel — must be checked FIRST ──────────
    if '[GENOTYPE DEADLINE PASSED]' in r:
        primary    = '🧬 Genotype Deadline Passed'
        p14_status = '🧬 Deadline passed'
        p56_status = '🧬 Deadline passed'
        return {
            'Primary_Reason':  primary,
            'P14_Status':      p14_status,
            'P56_Status':      p56_status,
            'Too_Old_For_P14': too_old_p14,
            'Too_Old_For_P56': too_old_p56,
            'Unusable_Both':   unusable_both,
            'Detail':          r,
        }

    # ── Inconclusive genotype — Climb's no-call symbol ───────────────────────
    if r == GENOTYPE_INCONC or 'inconclusive' in rl:
        return {
            'Primary_Reason':  '\U0001f9ec Inconclusive Genotype \u2014 Released to Available',
            'P14_Status':      '\U0001f9ec Released',
            'P56_Status':      '\U0001f9ec Released',
            'Too_Old_For_P14': too_old_p14,
            'Too_Old_For_P56': too_old_p56,
            'Unusable_Both':   unusable_both,
            'Detail':          r,
        }

    # ── Blank genotype — never genotyped — check BEFORE wild ─────────────────
    _blank_reason_indicators = (
        r == GENOTYPE_BLANK,
        rl == 'blank',
        r.startswith('Blank'),
        'no genotype' in rl,
        'genotype not available' in rl,
        "'half' strain" in rl and 'blank genotype' in rl,
        "'all' strain" in rl and 'blank genotype' in rl,
        'unknown strain' in rl and 'blank genotype' in rl,
        'blank genotype' in rl,
    )
    if any(_blank_reason_indicators):
        primary    = '\U0001f9ec Blank Genotype \u2014 Genotype Needed'
        p14_status = '\U0001f9ec Genotype'
        p56_status = '\U0001f9ec Genotype'
        return {
            'Primary_Reason':  primary,
            'P14_Status':      p14_status,
            'P56_Status':      p56_status,
            'Too_Old_For_P14': too_old_p14,
            'Too_Old_For_P56': too_old_p56,
            'Unusable_Both':   unusable_both,
            'Detail':          r,
        }

    # ── Unusable for BOTH ─────────────────────────────────────────────────────
    if '⛔' in r or 'unusable for both' in rl:
        too_old_p14   = 'YES'
        too_old_p56   = 'YES'
        unusable_both = 'YES'
        primary       = '⛔ Unusable for BOTH Timepoints (Too Old)'
        p14_status    = '❌ Too old'
        p56_status    = '❌ Too old'

    # ── Too late / too old for P14 ────────────────────────────────────────────
    elif 'too late for p14' in rl or ('too late' in rl and 'p14' in rl):
        too_old_p14 = 'YES'
        primary     = '❌ Too Late For P14 (Today)'
        p14_status  = '❌ Too late (today)'
        p56_status  = 'See detail'

    elif 'too old for p14' in rl:
        too_old_p14 = 'YES'
        primary     = '❌ Too Old For P14'
        p14_status  = '❌ Too old'
        if 'too old for p56' in rl:
            too_old_p56   = 'YES'
            unusable_both = 'YES'
            primary       = '⛔ Unusable for BOTH Timepoints (Too Old)'
            p56_status    = '❌ Too old'
        else:
            p56_status = 'See detail'

    # ── Too old for P56 ───────────────────────────────────────────────────────
    elif 'too old for p56' in rl:
        too_old_p56 = 'YES'
        primary     = '❌ Too Old For P56'
        p56_status  = '❌ Too old'
        p14_status  = 'See detail'

    # ── Quota filled ──────────────────────────────────────────────────────────
    elif 'quota' in rl and ('filled' in rl or 'complete' in rl or 'met' in rl):
        primary    = '✅ Quota Filled / Complete'
        p14_status = '✅ Quota met'
        p56_status = '✅ Quota met'

    # ── Incomplete cage group ─────────────────────────────────────────────────
    elif 'incomplete' in rl and ('group' in rl or 'cage' in rl):
        primary    = f'⚠️ Incomplete P56 Group (< {CONFIG["CAGE_SIZE"]} animals)'
        p56_status = f'⚠️ Incomplete group < {CONFIG["CAGE_SIZE"]}'
        if 'p14 unavailable' in rl:
            p14_status = '❌ Unavailable'
        elif 'p14' in rl and ('passed' in rl or 'too old' in rl):
            too_old_p14 = 'YES'
            p14_status  = '❌ Too old / passed'
        elif 'quota' in rl:
            p14_status  = '✅ Quota met'
        else:
            p14_status  = 'See detail'

    # ── Capacity / overflow ───────────────────────────────────────────────────
    elif 'capacity' in rl or 'over capacity' in rl:
        primary    = '🔴 P56 Date at Capacity / Overflow'
        p56_status = '🔴 Date full'
        if 'p14 unavailable' in rl:
            p14_status = '❌ Unavailable'
        elif 'p14 quota' in rl:
            p14_status = '✅ Quota met'
        else:
            p14_status = 'See detail'

    # ── Invalid harvest day ───────────────────────────────────────────────────
    elif 'invalid day' in rl or 'falls on' in rl or 'valid harvest day' in rl:
        primary    = '⚠️ P14 Falls on Invalid Day (Weekend)'
        p14_status = '⚠️ Invalid day'
        if 'p56' in rl:
            if 'passed' in rl:
                p56_status = '❌ Window passed'
            elif 'no wednesday' in rl:
                p56_status = '❌ No Wed in window'
            elif 'toe clip' in rl:
                p56_status = '❌ Toe clip'
            else:
                p56_status = 'See detail'
        else:
            p56_status = 'N/A'

    # ── Both windows passed ───────────────────────────────────────────────────
    elif 'p14' in rl and 'passed' in rl and 'p56' in rl and 'passed' in rl:
        primary    = '❌ Both Windows Have Passed'
        p14_status = '❌ Window passed'
        p56_status = '❌ Window passed'

    # ── P14 window passed only ────────────────────────────────────────────────
    elif 'p14' in rl and 'passed' in rl and 'p56' not in rl:
        primary    = '❌ P14 Window Passed'
        p14_status = '❌ Window passed'
        p56_status = 'N/A'

    # ── P56 window passed only ────────────────────────────────────────────────
    elif 'p56' in rl and 'passed' in rl and 'p14' not in rl:
        primary    = '❌ P56 Window Passed'
        p14_status = 'N/A'
        p56_status = '❌ Window passed'

    # ── No Wednesday in P56 window ────────────────────────────────────────────
    elif 'no wednesday' in rl or 'p42' in rl or 'p42–p49' in rl:
        primary    = '❌ No Wednesday in P56 Window (P42–P49)'
        p56_status = '❌ No Wed in window'
        if 'p14' in rl:
            p14_status = 'See detail'

    # ── Toe clip ──────────────────────────────────────────────────────────────
    elif 'toe clip' in rl:
        primary    = '🚫 Toe Clip — Not Allowed for P56 Behavior'
        p56_status = '🚫 Toe clip'
        if 'p14' in rl:
            if 'passed' in rl or 'too old' in rl or 'too late' in rl:
                too_old_p14 = 'YES'
                p14_status  = '❌ Too old / passed'
            elif 'quota' in rl:
                p14_status = '✅ Quota met'
            elif 'invalid' in rl or 'falls on' in rl:
                p14_status = '⚠️ Invalid day'
            else:
                p14_status = 'See detail'
        else:
            p14_status = 'N/A'

    # ── Sing inventory ────────────────────────────────────────────────────────
    elif 'sing inventory' in rl:
        primary    = '🔒 Not Assigned to Sing Inventory'
        p14_status = 'N/A'
        p56_status = 'N/A'

    # ── No birth date ─────────────────────────────────────────────────────────
    elif 'no birth date' in rl:
        primary    = '❓ No Birth Date Recorded'
        p14_status = '❓ No birth date'
        p56_status = '❓ No birth date'

    # ── Invalid birth date ────────────────────────────────────────────────────
    elif 'invalid birth date' in rl:
        primary    = '❓ Invalid Birth Date'
        p14_status = '❓ Invalid date'
        p56_status = '❓ Invalid date'

    # ── Wild genotype — STRICT matching only ─────────────────────────────────
    elif (
        r == GENOTYPE_WILD
        or rl == 'wild genotype — not usable for harvest'
        or rl == 'wild genotype - not usable for harvest'
        or rl.startswith('wild genotype')
        or ('+/+' in r and len(r) < 20)
    ):
        primary    = '🧬 Wild Genotype — Not Usable'
        p14_status = '🧬 Wild excluded'
        p56_status = '🧬 Wild excluded'

    # ── Cre-only wild ─────────────────────────────────────────────────────────
    elif 'cre' in rl and ('wildtype' in rl or 'ncar' in rl or 'cre-only' in rl):
        primary    = '🧬 Cre-Only Wild — No Mutation of Interest'
        p14_status = '🧬 Cre-only Wild'
        p56_status = '🧬 Cre-only Wild'

    # ── Wednesday over capacity (overflow) ───────────────────────────────────
    elif 'wednesday over capacity' in rl:
        primary    = '🔴 Wednesday Over Capacity — Overflow'
        p56_status = '🔴 Over capacity'
        if 'p14 unavailable' in rl:
            p14_status = '❌ Unavailable'
        elif 'p14 quota' in rl:
            p14_status = '✅ Quota met'
        else:
            p14_status = 'See detail'

    # ── Fallback ──────────────────────────────────────────────────────────────
    else:
        primary    = r[:80] + ('…' if len(r) > 80 else '')
        p14_status = 'See detail'
        p56_status = 'See detail'

    return {
        'Primary_Reason':  primary,
        'P14_Status':      p14_status,
        'P56_Status':      p56_status,
        'Too_Old_For_P14': too_old_p14,
        'Too_Old_For_P56': too_old_p56,
        'Unusable_Both':   unusable_both,
        'Detail':          r,
    }


# ============================================================================
# UNSCHEDULABLE REPORT
# ============================================================================

def create_unschedulable_report(assignments_df: pd.DataFrame,
                                use_excluded_df: pd.DataFrame,
                                genotype_excluded_df: pd.DataFrame,
                                date_excluded_df: Optional[pd.DataFrame] = None) -> pd.DataFrame:
    logger.info("Creating unschedulable report...")
    report_rows = []

    def _resolve_genotype(row: pd.Series) -> str:
        """
        Safely resolve the Genotype field to a canonical label.
        Blank/NaN/empty → GENOTYPE_BLANK. Never coerces blank to Wild.
        """
        geno = row.get('Genotype', None)

        if geno is None:
            return GENOTYPE_BLANK
        try:
            if pd.isna(geno):
                return GENOTYPE_BLANK
        except (TypeError, ValueError):
            pass

        geno_str = str(geno).strip()

        if geno_str == '' or geno_str.lower() in ('nan', 'none', 'n/a', 'na', '-'):
            return GENOTYPE_BLANK

        if geno_str in _CANONICAL_GENOTYPES:
            return geno_str

        strain = row.get('Strain', row.get('Line (Short)', ''))
        return canonicalize_genotype(geno_str, strain)

    # ── From unschedulable assignments ────────────────────────────────────────
    if len(assignments_df) > 0:
        unschedulable = assignments_df[
            assignments_df['Assigned_Timepoint'] == 'Unschedulable'
        ].copy()

        for _, row in unschedulable.iterrows():
            raw_reason    = row.get('Assignment_Reason', '')
            unusable_note = row.get('Unusable_Note', '')

            combined_reason = raw_reason
            if unusable_note and str(unusable_note) not in str(raw_reason):
                combined_reason = f"{unusable_note} | {raw_reason}"

            parsed = parse_unschedulable_reason(combined_reason)

            report_rows.append({
                'Animal_Name':              row.get('Animal_Name', 'Unknown'),
                'Birth_ID':                 row.get('Birth_ID', 'N/A'),
                'Strain':                   row.get('Strain', 'N/A'),
                'Genotype':                 _resolve_genotype(row),
                'Sex':                      row.get('Sex', 'N/A'),
                'Marker_Type':              row.get('Marker_Type', 'N/A'),
                'Birth_Date':               row.get('Birth_Date', 'N/A'),
                'Age_Today_Days':           row.get('Age_Today_Days', 'N/A'),
                'Primary_Reason':           parsed['Primary_Reason'],
                'P14_Status':               parsed['P14_Status'],
                'P56_Status':               parsed['P56_Status'],
                'Too_Old_For_P14':          parsed['Too_Old_For_P14'],
                'Too_Old_For_P56':          parsed['Too_Old_For_P56'],
                'Unusable_Both_Timepoints': parsed['Unusable_Both'],
                'Full_Detail':              parsed['Detail'],
            })

    # ── From use exclusions ───────────────────────────────────────────────────
    if len(use_excluded_df) > 0:
        for _, row in use_excluded_df.iterrows():
            report_rows.append({
                'Animal_Name':              row.get('Animal_Name', 'Unknown'),
                'Birth_ID':                 row.get('Birth_ID', 'N/A'),
                'Strain':                   row.get('Strain', 'N/A'),
                'Genotype':                 _resolve_genotype(row),
                'Sex':                      row.get('Sex', 'N/A'),
                'Marker_Type':              'N/A',
                'Birth_Date':               'N/A',
                'Age_Today_Days':           'N/A',
                'Primary_Reason':           '🔒 Not Assigned to Sing Inventory',
                'P14_Status':               'N/A',
                'P56_Status':               'N/A',
                'Too_Old_For_P14':          '',
                'Too_Old_For_P56':          '',
                'Unusable_Both_Timepoints': '',
                'Full_Detail':              row.get('Reason', ''),
            })

    # ── From genotype exclusions ──────────────────────────────────────────────
    if len(genotype_excluded_df) > 0:
        for _, row in genotype_excluded_df.iterrows():
            raw_reason = row.get('Reason', '')
            parsed = parse_unschedulable_reason(raw_reason)
            report_rows.append({
                'Animal_Name':              row.get('Animal_Name', 'Unknown'),
                'Birth_ID':                 row.get('Birth_ID', 'N/A'),
                'Strain':                   row.get('Strain', 'N/A'),
                'Genotype':                 _resolve_genotype(row),
                'Sex':                      row.get('Sex', 'N/A'),
                'Marker_Type':              'N/A',
                'Birth_Date':               row.get('Birth_Date', 'N/A'),
                'Age_Today_Days':           'N/A',
                'Primary_Reason':           parsed['Primary_Reason'],
                'P14_Status':               parsed['P14_Status'],
                'P56_Status':               parsed['P56_Status'],
                'Too_Old_For_P14':          parsed['Too_Old_For_P14'],
                'Too_Old_For_P56':          parsed['Too_Old_For_P56'],
                'Unusable_Both_Timepoints': parsed['Unusable_Both'],
                'Full_Detail':              raw_reason,
            })

    # ── From date exclusions ──────────────────────────────────────────────────
    if date_excluded_df is not None and len(date_excluded_df) > 0:
        for _, row in date_excluded_df.iterrows():
            raw_reason = row.get('Reason', '')
            report_rows.append({
                'Animal_Name':              row.get('Animal_Name', 'Unknown'),
                'Birth_ID':                 row.get('Birth_ID', 'N/A'),
                'Strain':                   row.get('Strain', 'N/A'),
                'Genotype':                 _resolve_genotype(row),
                'Sex':                      row.get('Sex', 'N/A'),
                'Marker_Type':              'N/A',
                'Birth_Date':               row.get('Birth_Date', 'N/A'),
                'Age_Today_Days':           'N/A',
                'Primary_Reason':           '📅 Outside Date Filter Range',
                'P14_Status':               'Filtered out',
                'P56_Status':               'Filtered out',
                'Too_Old_For_P14':          '',
                'Too_Old_For_P56':          '',
                'Unusable_Both_Timepoints': '',
                'Full_Detail':              raw_reason,
            })

    report = pd.DataFrame(report_rows)

    if len(report) > 0:
        priority_map = {
            '⛔': 0, '❌': 1, '🔴': 2, '⚠️': 3,
            '✅': 4, '🔒': 5, '🧬': 6, '📅': 7, '❓': 8,
        }

        def sort_key(val):
            for emoji, rank in priority_map.items():
                if str(val).startswith(emoji):
                    return rank
            return 9

        report['_sort'] = report['Primary_Reason'].apply(sort_key)
        report = report.sort_values(['_sort', 'Strain', 'Animal_Name'])
        report = report.drop(columns=['_sort'])

    logger.info(f"Unschedulable report: {len(report)} animals")
    return report


# ============================================================================
# SCHEDULE CREATION
# ============================================================================

def create_p14_schedule(assignments_df: pd.DataFrame) -> pd.DataFrame:
    if assignments_df.empty:
        return pd.DataFrame()

    p14 = assignments_df[assignments_df['Assigned_Timepoint'] == 'P14'].copy()
    if p14.empty:
        return pd.DataFrame()

    p14 = p14[~p14['Harvest_Type'].isin(['COMPLETE (Quota Filled)', 'Extra'])].copy()
    if p14.empty:
        return pd.DataFrame()

    p14['Day_of_Week'] = pd.to_datetime(p14['P14_Date']).dt.day_name()

    p14 = p14.sort_values(
        ['P14_Date', 'Genotype_Priority', 'Sex', 'Genotype', 'Animal_Name'],
        ascending=[True, False, True, True, True]
    )

    desired_cols = [
        'P14_Date', 'Day_of_Week',
        'P14_Age_At_Harvest_Days', 'P14_Age_At_Harvest_Months',
        'Animal_Name', 'Strain', 'Strain_Priority',
        'Genotype', 'Genotype_Priority',
        'Sex', 'Marker_Type', 'Harvest_Type', 'Priority',
        'Birth_Date', 'Birth_ID', 'Assignment_Reason'
    ]
    available_cols = [c for c in desired_cols if c in p14.columns]

    logger.info(f"P14 schedule: {len(p14)} animals")
    return p14[available_cols]


def create_p56_schedule(assignments_df: pd.DataFrame) -> pd.DataFrame:
    if assignments_df.empty:
        return pd.DataFrame()

    p56 = assignments_df[assignments_df['Assigned_Timepoint'] == 'P56'].copy()
    if p56.empty:
        return pd.DataFrame()

    kept_animals = []
    for (behavior_date, strain, genotype, sex), group in p56.groupby(
        ['P56_Behavior_Date', 'Strain', 'Genotype', 'Sex']
    ):
        animals = group.to_dict('records')

        # NB and Extra animals always kept — they don't need a full group
        special_animals = [a for a in animals
                          if str(a.get('Harvest_Type', '')).endswith(' NB')
                          or str(a.get('Harvest_Type', '')).strip() == 'Extra']
        regular_animals = [a for a in animals
                          if not str(a.get('Harvest_Type', '')).endswith(' NB')
                          and str(a.get('Harvest_Type', '')).strip() != 'Extra']
        kept_animals.extend(special_animals)

        for i in range(0, len(regular_animals), CONFIG['CAGE_SIZE']):
            cage_group = regular_animals[i:i + CONFIG['CAGE_SIZE']]
            # Skip incomplete regular groups — minimum CAGE_SIZE required
            if len(cage_group) < CONFIG['CAGE_SIZE']:
                logger.info(f"P56 schedule: skipping incomplete group {strain} {sex} {behavior_date} ({len(cage_group)} animals)")
                continue
            all_filled = all(
                a.get('Harvest_Type') == 'COMPLETE (Quota Filled)'
                for a in cage_group
            )
            if not all_filled:
                kept_animals.extend(cage_group)

    if not kept_animals:
        return pd.DataFrame()

    p56_filtered = pd.DataFrame(kept_animals)
    p56_filtered['Day_of_Week'] = 'Wednesday'

    p56_filtered = p56_filtered.sort_values(
        ['P56_Harvest_Date', 'Genotype_Priority', 'Sex', 'Genotype', 'Animal_Name'],
        ascending=[True, False, True, True, True]
    )

    desired_cols = [
        'P56_Behavior_Date', 'P56_Harvest_Date', 'Day_of_Week',
        'P56_Age_At_Behavior_Days', 'P56_Age_At_Behavior_Months',
        'P56_Age_At_Harvest_Days', 'P56_Age_At_Harvest_Months',
        'Animal_Name', 'Strain', 'Strain_Priority',
        'Genotype', 'Genotype_Priority',
        'Sex', 'Marker_Type', 'Harvest_Type', 'Priority',
        'Birth_Date', 'Birth_ID', 'Assignment_Reason'
    ]
    available_cols = [c for c in desired_cols if c in p56_filtered.columns]

    logger.info(f"P56 schedule: {len(p56_filtered)} animals")
    return p56_filtered[available_cols]


def create_b6_monthly_summary(assignments_df: pd.DataFrame) -> pd.DataFrame:
    if assignments_df.empty:
        return pd.DataFrame()

    min_per_month = CONFIG.get('B6_MIN_PER_MONTH', 3)

    scheduled = assignments_df[
        assignments_df['Assigned_Timepoint'].isin(['P14', 'P56'])
    ].copy()

    def harvest_month_label(row):
        tp = row.get('Assigned_Timepoint', '')
        if tp == 'P14':
            d = to_date(row.get('P14_Date'))
        elif tp == 'P56':
            d = to_date(row.get('P56_Harvest_Date'))
        else:
            d = None
        return d.strftime('%Y-%m') if d else None

    scheduled['Harvest_Month'] = scheduled.apply(harvest_month_label, axis=1)
    scheduled = scheduled[scheduled['Harvest_Month'].notna()]

    if scheduled.empty:
        return pd.DataFrame()

    all_months = sorted(scheduled['Harvest_Month'].unique())

    rows = []
    for month in all_months:
        month_data = scheduled[scheduled['Harvest_Month'] == month]
        b6_data = month_data[month_data['Strain'].apply(is_b6_strain)]

        total_in_month = len(month_data)
        b6_count = len(b6_data)
        b6_topup = (
            len(b6_data[b6_data['Priority'] == 'B6_MIN'])
            if 'Priority' in b6_data.columns else 0
        )

        meets_min = b6_count >= min_per_month
        status = (
            '✅ Meets Minimum' if meets_min
            else f'⚠️ Below Minimum (need {min_per_month - b6_count} more)'
        )

        rows.append({
            'Harvest_Month':           month,
            'Total_Animals_Scheduled': total_in_month,
            'B6_B6N_Count':            b6_count,
            'B6_B6N_TopUp_Count':      b6_topup,
            'Minimum_Required':        min_per_month,
            'Shortfall':               max(0, min_per_month - b6_count),
            'Status':                  status,
        })

    return pd.DataFrame(rows)


# ============================================================================
# MAIN SCHEDULING FUNCTION
# ============================================================================

def create_complete_schedule(animal_file: str, tracking_file: str, births_file: str,
                             output_dir: Optional[str] = None,
                             birth_date_start: Optional[date] = None,
                             birth_date_end: Optional[date] = None,
                             behavior_date_start: Optional[date] = None,
                             behavior_date_end: Optional[date] = None,
                             full_behavior_dates: Optional[List[date]] = None) -> str:
    logger.info("=" * 70)
    logger.info("COMPREHENSIVE ANIMAL SCHEDULER")
    logger.info("=" * 70)
    print("=" * 70)
    print("COMPREHENSIVE ANIMAL SCHEDULER")
    print("=" * 70)

    if output_dir is None:
        output_dir = os.path.dirname(os.path.abspath(animal_file))

    if full_behavior_dates:
        full_behavior_dates = [to_date(d) for d in full_behavior_dates if d is not None]
        full_behavior_dates = [d for d in full_behavior_dates if d is not None]

    # Read data
    animals_df = read_animal_data(animal_file)
    total_alive_count = len(animals_df)

    tracking_df = read_tracking_data(tracking_file) if tracking_file else None
    births_df = read_births_data(births_file) if births_file else None

    print(f"\nTotal alive animals loaded: {total_alive_count:,}")

    diagnose_animal_file(animals_df)

    # Parse requirements
    requirements = parse_requirements(tracking_df)
    remaining_needs = calculate_remaining_needs(requirements)

    # Animal filtering — apply Use filter first so animals_df_raw is Sing Inventory only
    print("\n" + "=" * 70)
    print("ANIMAL FILTERING")
    print("=" * 70)

    animals_df, use_excluded = filter_animals_by_use(animals_df)
    print(f"After 'Sing Inventory' filter: {len(animals_df):,} animals remain")
    if animals_df.empty:
        print("  ⚠️  ALL animals were excluded by the Use filter.")
        print("  Check the 'Use' column values in your CSV.")

    # Capture raw Sing Inventory animals — used for All Animals output and births sexing
    animals_df_raw = animals_df.copy()

    # Births analysis
    print("\n" + "=" * 70)
    print("BIRTHS ANALYSIS")
    print("=" * 70)

    sexing_schedule_df = pd.DataFrame()
    if births_df is not None:
        sexing_schedule_df = build_births_sexing_schedule(births_df, animals_df_raw)
        upcoming = (
            sexing_schedule_df[
                sexing_schedule_df['Days_Until_Sexing'].apply(
                    lambda x: isinstance(x, int) and 0 <= x <= 7
                )
            ] if len(sexing_schedule_df) > 0 else pd.DataFrame()
        )
        print(f"  Births needing sexing (not yet entered): {len(sexing_schedule_df)}")
        if len(upcoming) > 0:
            print(f"  ⚠️  {len(upcoming)} litter(s) need sexing within the next 7 days!")

    unmatched_births_df = find_unmatched_births_enhanced(
        births_df, animals_df, requirements, remaining_needs
    )

    no_geno_strains = frozenset(
        k for k, v in requirements.items() if not v.get('genotyped', True)
    )
    animals_df, genotype_excluded_pass1, blank_genotypes = filter_animals_by_genotype_first_pass(
        animals_df, no_geno_strains=no_geno_strains
    )
    print(f"After genotype first pass:     {len(animals_df):,} animals remain")
    print(f"  Excluded (Wild, Cre-only Wild, Inconclusive): {len(genotype_excluded_pass1)}")
    if len(blank_genotypes) > 0 and 'Genotype' in blank_genotypes.columns:
        _n_inconc = int((blank_genotypes['Genotype'] == GENOTYPE_INCONC).sum())
        _n_blank  = len(blank_genotypes) - _n_inconc
        print(f"  Blank genotypes (pending 2nd pass):           {_n_blank}")
        print(f"  Inconclusive genotypes (released):            {_n_inconc}")
    else:
        print(f"  Blank genotypes (pending 2nd pass):           {len(blank_genotypes)}")
    if animals_df.empty:
        print("  ⚠️  ALL animals were excluded by genotype filtering.")

    animals_df, date_excluded = filter_animals_by_dates(
        animals_df, birth_date_start, birth_date_end,
        behavior_date_start, behavior_date_end
    )
    if len(date_excluded) > 0:
        print(f"After date filtering:          {len(animals_df):,} animals remain "
              f"(excluded {len(date_excluded)})")

    print(f"\nAnimals entering eligibility check: {len(animals_df):,}")

    # Eligibility
    print("\nChecking eligibility...")
    eligibility = check_eligibility(animals_df, full_behavior_dates)

    print(f"Eligibility results: {len(eligibility):,} animals processed")
    if len(eligibility) > 0:
        p14_elig_count = eligibility['P14_Eligible'].sum() if 'P14_Eligible' in eligibility.columns else 0
        p56_elig_count = eligibility['P56_Eligible'].sum() if 'P56_Eligible' in eligibility.columns else 0
        print(f"  P14 eligible: {p14_elig_count}")
        print(f"  P56 eligible: {p56_elig_count}")

    het_count_df = (
        eligibility[eligibility['Genotype'].apply(is_heterozygous)]
        if len(eligibility) > 0 and 'Genotype' in eligibility.columns
        else pd.DataFrame()
    )

    # Assignment
    print("\nAssigning animals to timepoints...")
    assignments = assign_animals_smart(eligibility, remaining_needs)

    if len(assignments) > 0:
        assignments = check_capacity_and_reassign(assignments, remaining_needs)
    else:
        print("  ⚠️  No assignments to process.")

    # Determine actually-full dates
    if len(assignments) > 0:
        p56_assigned = assignments[assignments['Assigned_Timepoint'] == 'P56'].copy()
        if len(p56_assigned) > 0:
            p56_assigned = p56_assigned[p56_assigned['P56_Behavior_Date'].notna()].copy()
            p56_assigned['P56_Behavior_Date'] = p56_assigned['P56_Behavior_Date'].apply(to_date)
            p56_assigned = p56_assigned[p56_assigned['P56_Behavior_Date'].notna()]
            if len(p56_assigned) > 0:
                wed_counts = p56_assigned.groupby('P56_Behavior_Date').size()
                actual_full = wed_counts[wed_counts >= CONFIG['WEDNESDAY_CAPACITY']].index.tolist()
                all_full_dates = list(set((full_behavior_dates or []) + actual_full))
            else:
                all_full_dates = full_behavior_dates or []
        else:
            all_full_dates = full_behavior_dates or []
    else:
        all_full_dates = full_behavior_dates or []

    # Second pass: blank genotypes
    print("Second pass: Analyzing blank genotypes...")
    genotype_excluded_pass2 = analyze_blank_genotypes_second_pass(
        blank_genotypes, all_full_dates, remaining_needs
    )
    genotype_excluded = pd.concat(
        [genotype_excluded_pass1, genotype_excluded_pass2], ignore_index=True
    )

    # Harvest type assignment
    print("\n" + "=" * 70)
    print("HARVEST ASSIGNMENT REVIEW")
    print("=" * 70)

    if len(assignments) > 0:
        # Single-pass GUI — all harvest options (including NB types) available per animal.
        # User picks the final harvest type directly; no second pass.
        gui_selections = prompt_harvest_assignments_gui(assignments, remaining_needs)

        # Log exactly what the GUI returned
        print(f"\n  GUI returned {len(gui_selections)} selections:")
        for _n, _h in sorted(gui_selections.items()):
            print(f"    {_n}: {_h}")

        # Remove Do Not Schedule animals from assignments and eligibility
        do_not_schedule = {
            name for name, htype in gui_selections.items()
            if htype == 'DO_NOT_SCHEDULE'
        }
        if do_not_schedule:
            print(f"  ⚠ {len(do_not_schedule)} animal(s) marked 'Do Not Schedule' — excluded.")
            logger.info(f"Do Not Schedule: {sorted(do_not_schedule)}")
            dns_names_str = {str(n) for n in do_not_schedule}
            assignments = assignments[
                ~assignments['Animal_Name'].astype(str).isin(dns_names_str)
            ].copy()
            eligibility = eligibility[
                ~eligibility['Animal_Name'].astype(str).isin(dns_names_str)
            ].copy()

        # Build final override dict (exclude DO_NOT_SCHEDULE sentinels)
        harvest_overrides = {
            name: htype
            for name, htype in gui_selections.items()
            if htype != 'DO_NOT_SCHEDULE'
        }

        assignments = assign_harvest_types(
            assignments, remaining_needs, requirements, harvest_overrides
        )
    else:
        assignments = pd.DataFrame()
        harvest_overrides = {}

    # B6/B6N monthly minimum
    # B6/B6N monthly minimum enforcement disabled — managed manually
    # if len(assignments) > 0:
    #     assignments = enforce_b6_monthly_minimum(assignments, eligibility, remaining_needs)

    # Build output sheets
    print("Creating schedule sheets...")
    p14_schedule = create_p14_schedule(assignments)
    p56_schedule = create_p56_schedule(assignments)
    unschedulable = create_unschedulable_report(
        assignments if len(assignments) > 0 else pd.DataFrame(),
        use_excluded,
        genotype_excluded,
        date_excluded
    )

    # Counts for summary
    p14_count = len(p14_schedule) if len(p14_schedule) > 0 else 0
    p56_count = len(p56_schedule) if len(p56_schedule) > 0 else 0
    p56_cages = p56_count // CONFIG['CAGE_SIZE'] if CONFIG['CAGE_SIZE'] > 0 else 0

    genotype_critical_count = (
        len(genotype_excluded[genotype_excluded['Reason'].str.contains('⚠️', na=False)])
        if len(genotype_excluded) > 0 and 'Reason' in genotype_excluded.columns
        else 0
    )
    high_priority_count = (
        len(assignments[assignments['Priority'] == 'HIGH'])
        if len(assignments) > 0 and 'Priority' in assignments.columns else 0
    )
    b6_topup_count = (
        len(assignments[assignments['Priority'] == 'B6_MIN'])
        if len(assignments) > 0 and 'Priority' in assignments.columns else 0
    )

    toe_clip_excluded = (
        eligibility[eligibility['P56_Reason'].str.contains('Toe Clip', na=False)]
        if len(eligibility) > 0 and 'P56_Reason' in eligibility.columns
        else pd.DataFrame()
    )
    full_date_excluded = (
        eligibility[eligibility['P56_Reason'].str.contains('capacity', na=False, case=False)]
        if len(eligibility) > 0 and 'P56_Reason' in eligibility.columns
        else pd.DataFrame()
    )
    unusable_both = (
        len(unschedulable[unschedulable['Unusable_Both_Timepoints'] == 'YES'])
        if len(unschedulable) > 0 and 'Unusable_Both_Timepoints' in unschedulable.columns
        else 0
    )

    unmatched_p14_count = (
        len(unmatched_births_df[unmatched_births_df['P14_Potential'] == 'Yes'])
        if len(unmatched_births_df) > 0 and 'P14_Potential' in unmatched_births_df.columns else 0
    )
    unmatched_p56_count = (
        len(unmatched_births_df[unmatched_births_df['P56_Potential'] == 'Yes'])
        if len(unmatched_births_df) > 0 and 'P56_Potential' in unmatched_births_df.columns else 0
    )
    unmatched_priority_count = (
        len(unmatched_births_df[unmatched_births_df['Priority_Strain'] == 'YES'])
        if len(unmatched_births_df) > 0 and 'Priority_Strain' in unmatched_births_df.columns else 0
    )
    unmatched_quota_count = (
        len(unmatched_births_df[unmatched_births_df['Quota_Status'].str.contains('NEEDED', na=False)])
        if len(unmatched_births_df) > 0 and 'Quota_Status' in unmatched_births_df.columns else 0
    )

    upcoming_sexing_count = 0
    if len(sexing_schedule_df) > 0 and 'Days_Until_Sexing' in sexing_schedule_df.columns:
        upcoming_sexing_count = len(sexing_schedule_df[
            sexing_schedule_df['Days_Until_Sexing'].apply(
                lambda x: isinstance(x, int) and 0 <= x <= 7
            )
        ])

    summary_data = {
        'Metric': [
            '── ANIMAL COUNTS ──',
            'Total Alive Animals',
            'Excluded (Not Sing Inventory)',
            'Excluded (Genotype)',
            'Excluded (Genotype - CRITICAL)',
            'Excluded (Date Filters)',
            'Excluded (Toe Clip for P56)',
            'Excluded (Full P56 Dates)',
            'Unusable for BOTH Timepoints',
            'Heterozygous (Het) Animals',
            'Animals Processed',
            '── SCHEDULE ──',
            'P14 Assigned',
            'P56 Assigned',
            'P56 Complete Cages',
            'Unschedulable',
            'HIGH Priority Animals',
            '── B6/B6N ──',
            'B6/B6N Monthly Minimum Required',
            'B6/B6N Top-Up Animals Added',
            '── BIRTHS / SEXING ──',
            'Births Needing Sexing (not yet entered)',
            'Sexing Due Within 7 Days',
            'Unmatched Births (Sing Inventory)',
            'Unmatched - Can Schedule P14',
            'Unmatched - Can Schedule P56',
            'Unmatched - Priority Strains',
            'Unmatched - With Quota Needs',
            '── SETTINGS ──',
            'Wednesday Capacity',
            'Sexing Day Offset (days)',
            'Birth Date Filter Start',
            'Birth Date Filter End',
            'Behavior Date Filter Start',
            'Behavior Date Filter End',
            'Full P56 Behavior Dates',
            'Generated On',
            'Animal File',
            'Tracking File',
            'Births File',
        ],
        'Value': [
            '',
            total_alive_count,
            len(use_excluded),
            len(genotype_excluded),
            genotype_critical_count,
            len(date_excluded) if len(date_excluded) > 0 else 0,
            len(toe_clip_excluded),
            len(full_date_excluded),
            unusable_both,
            len(het_count_df),
            len(assignments) if len(assignments) > 0 else 0,
            '',
            p14_count,
            p56_count,
            p56_cages,
            len(unschedulable),
            high_priority_count,
            '',
            CONFIG.get('B6_MIN_PER_MONTH', 3),
            b6_topup_count,
            '',
            len(sexing_schedule_df),
            upcoming_sexing_count,
            len(unmatched_births_df),
            unmatched_p14_count,
            unmatched_p56_count,
            unmatched_priority_count,
            unmatched_quota_count,
            '',
            CONFIG['WEDNESDAY_CAPACITY'],
            CONFIG.get('SEXING_OFFSET_DAYS', 9),
            str(birth_date_start) if birth_date_start else 'None',
            str(birth_date_end) if birth_date_end else 'None',
            str(behavior_date_start) if behavior_date_start else 'None',
            str(behavior_date_end) if behavior_date_end else 'None',
            ', '.join([str(d) for d in full_behavior_dates]) if full_behavior_dates else 'None',
            datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            os.path.basename(animal_file),
            os.path.basename(tracking_file) if tracking_file else 'N/A',
            os.path.basename(births_file) if births_file else 'N/A',
        ]
    }

    # Build all_animals — raw input file with scheduling results merged in
    sched_cols = ['Animal_Name', 'Assigned_Timepoint', 'Harvest_Type', 'Assignment_Reason',
                  'P14_Date', 'P56_Behavior_Date', 'P56_Harvest_Date',
                  'Strain_Priority', 'Genotype_Priority', 'Priority']
    if len(assignments) > 0:
        sched_subset = assignments[[c for c in sched_cols if c in assignments.columns]].copy()
        raw_name_col = next((c for c in ['Name', 'Animal ID', 'Animal_Name'] if c in animals_df_raw.columns), None)
        if raw_name_col:
            raw_copy = animals_df_raw.copy()
            raw_copy[raw_name_col] = raw_copy[raw_name_col].astype(str)
            sched_subset['Animal_Name'] = sched_subset['Animal_Name'].astype(str)
            if raw_name_col != 'Animal_Name':
                sched_subset = sched_subset.rename(columns={'Animal_Name': raw_name_col})
            all_animals_merged = raw_copy.merge(sched_subset, on=raw_name_col, how='left')
        else:
            all_animals_merged = animals_df_raw.copy()
    else:
        all_animals_merged = animals_df_raw.copy()

    # Build a name → exclusion_reason lookup from all filter stages
    exclusion_reasons: Dict[str, str] = {}
    if len(use_excluded) > 0:
        name_col = next((c for c in ['Name', 'Animal_Name', 'Animal ID'] if c in use_excluded.columns), None)
        use_col  = next((c for c in ['Use'] if c in use_excluded.columns), None)
        if name_col:
            for _, r in use_excluded.iterrows():
                use_val = str(r.get(use_col, '')).strip() if use_col else ''
                exclusion_reasons[str(r[name_col])] = f'Not Sing Inventory (Use: {use_val or "blank"})'
    if len(genotype_excluded_pass1) > 0 and 'Animal_Name' in genotype_excluded_pass1.columns:
        for _, r in genotype_excluded_pass1.iterrows():
            exclusion_reasons[str(r['Animal_Name'])] = f'Excluded — {r.get("Reason", "genotype issue")}'
    if len(genotype_excluded) > 0 and 'Animal_Name' in genotype_excluded.columns:
        for _, r in genotype_excluded.iterrows():
            n = str(r['Animal_Name'])
            if n not in exclusion_reasons:
                exclusion_reasons[n] = f'Excluded — {r.get("Reason", "genotype issue")}'
    if len(date_excluded) > 0:
        name_col = next((c for c in ['Name', 'Animal_Name'] if c in date_excluded.columns), None)
        if name_col:
            for _, r in date_excluded.iterrows():
                exclusion_reasons[str(r[name_col])] = 'Excluded — outside date filter window'

    # Fill in Assignment_Reason for every animal
    raw_name_col_fill = next((c for c in ['Name', 'Animal_Name'] if c in all_animals_merged.columns), None)

    def _fill_reason(row):
        reason = row.get('Assignment_Reason')
        if pd.notna(reason) and str(reason).strip() not in ('', 'nan'):
            return reason
        name = str(row.get(raw_name_col_fill, '')).strip() if raw_name_col_fill else ''
        if name in exclusion_reasons:
            return exclusion_reasons[name]
        tp = row.get('Assigned_Timepoint')
        if pd.notna(tp) and str(tp).strip() not in ('', 'nan'):
            return f'Scheduled — {tp}'
        p14_reason = str(row.get('P14_Reason', '') or '').strip()
        p56_reason = str(row.get('P56_Reason', '') or '').strip()
        if p14_reason or p56_reason:
            parts = []
            if p14_reason:
                parts.append(f'P14: {p14_reason}')
            if p56_reason:
                parts.append(f'P56: {p56_reason}')
            return ' | '.join(parts)
        return 'Not scheduled — no eligible timepoint found'

    all_animals_merged['Assignment_Reason'] = all_animals_merged.apply(_fill_reason, axis=1)

    # Fill Assigned_Timepoint — every animal must have one
    def _fill_timepoint(row):
        tp = row.get('Assigned_Timepoint')
        if pd.notna(tp) and str(tp).strip() not in ('', 'nan'):
            return str(tp).strip()
        return 'Unschedulable'

    all_animals_merged['Assigned_Timepoint'] = all_animals_merged.apply(_fill_timepoint, axis=1)

    # Fill Harvest_Type — every animal must have one
    def _fill_harvest_type(row):
        ht = row.get('Harvest_Type')
        if pd.notna(ht) and str(ht).strip() not in ('', 'nan'):
            return str(ht).strip()
        return 'N/A'

    # Ensure scheduling columns exist even if no assignments were made
    for col in ['Assigned_Timepoint', 'Harvest_Type', 'Assignment_Reason']:
        if col not in all_animals_merged.columns:
            all_animals_merged[col] = None

    all_animals_merged['Harvest_Type'] = all_animals_merged.apply(_fill_harvest_type, axis=1)

    # P14_Date — show "Too Old" if the animal missed the P14 window
    if 'P14_Too_Old' in all_animals_merged.columns and 'P14_Date' in all_animals_merged.columns:
        all_animals_merged['P14_Date'] = all_animals_merged.apply(
            lambda r: 'Too Old' if r.get('P14_Too_Old') == True and (
                pd.isna(r.get('P14_Date')) or str(r.get('P14_Date', '')).strip() in ('', 'nan', 'NaT')
            ) else r.get('P14_Date'),
            axis=1
        )

    # Column order: key SING columns first, scheduling data next, raw Climb extras at right
    priority_cols = [
        'Name', 'Line (Short)', 'Sex', 'Genotype', 'Birth Date', 'Age (days)',
        'Assigned_Timepoint', 'Harvest_Type', 'Assignment_Reason',
        'P14_Date', 'P56_Behavior_Date', 'P56_Harvest_Date',
        'Strain_Priority', 'Genotype_Priority', 'Priority',
        'Marker', 'Marker Type', 'Housing ID', 'Birth ID',
    ]
    remaining_cols = [c for c in all_animals_merged.columns if c not in priority_cols]
    ordered_cols = [c for c in priority_cols if c in all_animals_merged.columns] + remaining_cols
    all_animals_merged = all_animals_merged[ordered_cols]

    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    output_filename = f'Complete_Schedule_{timestamp}.xlsx'
    output_path = os.path.join(output_dir, output_filename)

    print(f"\nWriting Excel: {output_filename}")

    try:
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:

            pd.DataFrame(summary_data).to_excel(writer, sheet_name='Summary', index=False)

            if len(p14_schedule) > 0:
                p14_schedule.to_excel(writer, sheet_name='P14 Schedule', index=False)

            if len(p56_schedule) > 0:
                p56_schedule.to_excel(writer, sheet_name='P56 Schedule', index=False)

            if len(genotype_excluded) > 0:
                genotype_excluded.to_excel(
                    writer,
                    sheet_name=truncate_sheet_name('Genotype Excluded Details'),
                    index=False
                )

            # All Animals tab
            if len(assignments) > 0:
                _INTERNAL_COLUMNS = {
                    '_quota_limited_complete_group', '_incomplete_group',
                    '_full_date_complete', '_urgency_sort', '_birth_date_obj',
                    '_first_wed', 'is_het', 'breeding_type'
                }
                _EXCLUDED_ORIGINAL_AGE_COLUMNS = {'Age (days)', 'Age (weeks)', 'Age (months)'}
                _REDUNDANT_WITH_COMPUTED = {
                    'Name': 'Animal_Name', 'Birth ID': 'Birth_ID',
                    'Line (Short)': 'Strain', 'Birth Date': 'Birth_Date',
                    'Marker Type': 'Marker_Type',
                }

                computed_cols_front = [
                    'Animal_Name', 'Birth_ID', 'Strain', 'Strain_Priority',
                    'Genotype', 'Genotype_Priority', 'Sex', 'Marker_Type',
                    'Birth_Date', 'Age_Today_Days',
                    'Assigned_Timepoint', 'Harvest_Type', 'Priority',
                    'P14_Eligible', 'P14_Too_Old', 'P14_Date', 'P14_Reason',
                    'P14_Age_At_Harvest_Days', 'P14_Age_At_Harvest_Months',
                    'P56_Eligible', 'P56_Too_Old', 'P56_Behavior_Date',
                    'P56_Harvest_Date', 'P56_Reason',
                    'P56_Age_At_Behavior_Days', 'P56_Age_At_Behavior_Months',
                    'P56_Age_At_Harvest_Days', 'P56_Age_At_Harvest_Months',
                    'Unusable_Note', 'Assignment_Reason',
                ]

                all_available = assignments.columns.tolist()
                ordered_cols = []
                seen = set()

                for col in computed_cols_front:
                    if col in all_available and col not in seen:
                        ordered_cols.append(col)
                        seen.add(col)

                for col in all_available:
                    if col in seen:
                        continue
                    if col in _INTERNAL_COLUMNS:
                        continue
                    if col in _EXCLUDED_ORIGINAL_AGE_COLUMNS:
                        continue
                    if col in _REDUNDANT_WITH_COMPUTED:
                        if _REDUNDANT_WITH_COMPUTED[col] in seen:
                            continue
                    ordered_cols.append(col)
                    seen.add(col)

                all_animals_merged.to_excel(writer, sheet_name='All Animals', index=False)

            # ── Formatting ────────────────────────────────────────────────────
            wb = writer.book

            if 'Sexing Schedule' in wb.sheetnames:
                ws = wb['Sexing Schedule']
                headers = [cell.value for cell in ws[1]]
                status_col = headers.index('Sexing_Status') + 1 if 'Sexing_Status' in headers else None

                for row_idx in range(2, ws.max_row + 1):
                    if status_col:
                        cell = ws.cell(row=row_idx, column=status_col)
                        val = str(cell.value) if cell.value else ''
                        color = None
                        if 'TODAY' in val:
                            color = 'FF0000'
                            cell.font = Font(bold=True, color='FFFFFF')
                        elif 'TOMORROW' in val:
                            color = 'FF8C00'
                            cell.font = Font(bold=True)
                        elif 'SOON' in val:
                            color = 'FFD700'
                        elif 'Upcoming' in val:
                            color = 'A8E6CF'
                        elif 'Done' in val:
                            color = 'D3D3D3'
                        if color:
                            cell.fill = PatternFill(
                                start_color=color, end_color=color, fill_type='solid'
                            )



            geno_sheet = truncate_sheet_name('Genotype Excluded Details')
            if geno_sheet in wb.sheetnames:
                ws = wb[geno_sheet]
                headers = [cell.value for cell in ws[1]]
                pred_col      = headers.index('Prediction')           + 1 if 'Prediction'           in headers else None
                days_col      = headers.index('Days_Until_Deadline')  + 1 if 'Days_Until_Deadline'  in headers else None
                p14_worth_col = headers.index('P14_Worth_Genotyping') + 1 if 'P14_Worth_Genotyping' in headers else None
                p56_worth_col = headers.index('P56_Worth_Genotyping') + 1 if 'P56_Worth_Genotyping' in headers else None

                for row_idx in range(2, ws.max_row + 1):
                    if pred_col:
                        cell = ws.cell(row=row_idx, column=pred_col)
                        val = str(cell.value) if cell.value else ''
                        if 'LIKELY USABLE' in val:
                            cell.fill = PatternFill(start_color='C6E0B4', end_color='C6E0B4', fill_type='solid')
                            cell.font = Font(bold=True, color='00AA00')
                        elif 'POSSIBLY USABLE' in val:
                            cell.fill = PatternFill(start_color='FFE699', end_color='FFE699', fill_type='solid')
                        elif 'NOT SCHEDULABLE' in val:
                            cell.fill = PatternFill(start_color='FFB6B6', end_color='FFB6B6', fill_type='solid')
                        elif 'DEADLINE PASSED' in val:
                            cell.fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')

                    if days_col:
                        cell = ws.cell(row=row_idx, column=days_col)
                        try:
                            val = cell.value
                            if val != 'N/A' and isinstance(val, (int, float)):
                                if val <= 7:
                                    cell.fill = PatternFill(start_color='FF6B6B', end_color='FF6B6B', fill_type='solid')
                                    cell.font = Font(bold=True, color='FFFFFF')
                                elif val <= 14:
                                    cell.fill = PatternFill(start_color='FFD93D', end_color='FFD93D', fill_type='solid')
                        except Exception:
                            pass

                    for worth_col in [p14_worth_col, p56_worth_col]:
                        if worth_col:
                            cell = ws.cell(row=row_idx, column=worth_col)
                            val = str(cell.value) if cell.value else ''
                            if '✅ YES' in val:
                                cell.fill = PatternFill(start_color='C6E0B4', end_color='C6E0B4', fill_type='solid')
                                cell.font = Font(bold=True, color='006400')
                            elif '🟡' in val:
                                cell.fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
                            elif '❌' in val:
                                cell.fill = PatternFill(start_color='FFB6B6', end_color='FFB6B6', fill_type='solid')
                            elif '⚠️ QUOTA' in val:
                                cell.fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')

            for sheet_name in wb.sheetnames:
                auto_size_columns(wb[sheet_name])

        logger.info("✓ Excel written successfully")
        print("✓ Excel written successfully")

    except Exception as e:
        logger.error(f"Excel write failed: {e}", exc_info=True)
        print(f"\n⚠️ Excel write failed: {e}")
        print("✓ Backup CSVs were saved — you can open those directly.")
        raise

    # Console summary
    print("\n" + "=" * 70)
    print("ASSIGNMENT SUMMARY")
    print("=" * 70)
    print(f"  P14 assigned:          {p14_count:>6}")
    print(f"  P56 assigned:          {p56_count:>6}  ({p56_cages} complete cages)")
    print(f"  Unschedulable:         {len(unschedulable):>6}")
    if unusable_both > 0:
        print(f"  ⛔ Unusable for both:  {unusable_both:>6}  (too old for P14 AND P56)")
    if b6_topup_count > 0:
        print(f"  B6/B6N top-up added:  {b6_topup_count:>6}  (to meet {CONFIG['B6_MIN_PER_MONTH']}/month minimum)")
    if len(sexing_schedule_df) > 0:
        print(f"\n  Births needing sexing:    {len(sexing_schedule_df)}")
        if upcoming_sexing_count > 0:
            print(f"  ⚠️  Sexing due ≤7 days:    {upcoming_sexing_count}")
    if len(unmatched_births_df) > 0:
        print(f"\n  ⚠️  Unmatched births:       {len(unmatched_births_df)}")
    if genotype_critical_count > 0:
        print(f"\n  ⚠️  CRITICAL genotype issues: {genotype_critical_count}")

    print("\n" + "=" * 70)
    print(f"✓ Schedule saved to:\n  {output_path}")
    print("=" * 70)

    return output_path, assignments


# ============================================================================
# UNIT TESTS
# ============================================================================
import unittest  # lazy import — only needed when CONFIG['RUN_TESTS'] is True

class TestSchedulerFunctions(unittest.TestCase):

    def setUp(self):
        self.test_date = date(2025, 11, 15)

    def test_to_date_conversion(self):
        self.assertEqual(to_date(self.test_date), self.test_date)
        self.assertEqual(to_date(datetime(2025, 11, 15, 10, 30)), self.test_date)
        self.assertEqual(to_date(pd.Timestamp('2025-11-15')), self.test_date)
        self.assertIsNone(to_date(None))
        self.assertIsNone(to_date(pd.NaT))

    def test_is_heterozygous(self):
        self.assertTrue(is_heterozygous('-/+'))
        self.assertTrue(is_heterozygous('+/-'))
        self.assertTrue(is_heterozygous('HET'))
        self.assertTrue(is_heterozygous('Het1'))
        self.assertFalse(is_heterozygous('+/+'))
        self.assertFalse(is_heterozygous(''))
        self.assertFalse(is_heterozygous(None))
        self.assertTrue(is_heterozygous('Het'))
        self.assertFalse(is_heterozygous('Wild'))
        self.assertFalse(is_heterozygous('Hom'))
        self.assertFalse(is_heterozygous('Blank'))
        self.assertFalse(is_heterozygous('Inbred'))
        self.assertFalse(is_heterozygous('Hemi'))

    def test_canonicalize_genotype_het(self):
        for raw in ['-/+', '+/-', 'HET', 'het', 'Heterozygous', 'carrier']:
            self.assertEqual(
                canonicalize_genotype(raw), GENOTYPE_HET,
                f"Expected Het for '{raw}'"
            )

    def test_canonicalize_genotype_wild(self):
        for raw in ['+/+', '+/Y', 'WT', 'wildtype', 'wild-type',
                    'Cre ncar', 'Generic Cre', 'cre +/+']:
            self.assertEqual(
                canonicalize_genotype(raw), GENOTYPE_WILD,
                f"Expected Wild for '{raw}'"
            )

    def test_canonicalize_genotype_hom(self):
        for raw in ['-/-', 'HOM', 'homozygous', 'mut/mut', 'KO/KO']:
            self.assertEqual(
                canonicalize_genotype(raw), GENOTYPE_HOM,
                f"Expected Hom for '{raw}'"
            )

    def test_canonicalize_genotype_hemi(self):
        for raw in ['hemi', 'hemizygous', 'tg/+', '+/tg', '-/Y']:
            self.assertEqual(
                canonicalize_genotype(raw), GENOTYPE_HEMI,
                f"Expected Hemi for '{raw}'"
            )

    def test_canonicalize_genotype_inbred(self):
        self.assertEqual(
            canonicalize_genotype('+/+', strain='B6J'), GENOTYPE_INBRED
        )
        self.assertEqual(
            canonicalize_genotype('', strain='B6NJ'), GENOTYPE_INBRED
        )

    def test_canonicalize_genotype_blank(self):
        for raw in [None, '', 'nan', 'N/A']:
            self.assertEqual(
                canonicalize_genotype(raw), GENOTYPE_BLANK,
                f"Expected Blank for '{raw}'"
            )

    def test_canonicalize_inconclusive(self):
        """Climb's Inconclusive symbol is its own label, not Blank."""
        for raw in ['Inconclusive', 'inconclusive', 'Kdm5b Inconclusive']:
            self.assertEqual(
                canonicalize_genotype(raw), GENOTYPE_INCONC,
                f"Expected Inconclusive for '{raw}'"
            )

    def test_canonicalize_already_canonical(self):
        for label in ['Wild', 'Het', 'Hom', 'Hemi', 'Inbred', 'Blank',
                      'Inconclusive']:
            self.assertEqual(canonicalize_genotype(label), label)

    def test_is_wildtype_cre_only(self):
        self.assertTrue(is_wildtype_cre_only('Cre ncar'))
        self.assertTrue(is_wildtype_cre_only('Generic Cre'))
        self.assertTrue(is_wildtype_cre_only('Cre-ncar'))
        self.assertFalse(is_wildtype_cre_only('Cre ncar; -/+'))
        self.assertFalse(is_wildtype_cre_only('Cre ncar HET'))
        self.assertFalse(is_wildtype_cre_only('Cre ncar; -/-'))
        self.assertFalse(is_wildtype_cre_only('+/+'))
        self.assertFalse(is_wildtype_cre_only('-/+'))
        self.assertFalse(is_wildtype_cre_only(''))
        self.assertFalse(is_wildtype_cre_only(None))
        self.assertTrue(is_wildtype_cre_only(GENOTYPE_WILD))
        self.assertFalse(is_wildtype_cre_only(GENOTYPE_HET))

    def test_is_b6_strain(self):
        self.assertTrue(is_b6_strain('B6J'))
        self.assertTrue(is_b6_strain('b6j'))
        self.assertTrue(is_b6_strain('B6NJ'))
        self.assertTrue(is_b6_strain('b6nj'))
        self.assertFalse(is_b6_strain('SHANK3'))
        self.assertFalse(is_b6_strain('CHD8'))
        self.assertFalse(is_b6_strain(None))
        self.assertFalse(is_b6_strain(''))

    def test_calculate_sexing_date(self):
        bd = date(2025, 11, 1)
        expected = date(2025, 11, 10)
        self.assertEqual(calculate_sexing_date(bd), expected)
        self.assertIsNone(calculate_sexing_date(None))
        self.assertIsNone(calculate_sexing_date(pd.NaT))

    def test_calculate_sexing_date_pd_timestamp(self):
        ts = pd.Timestamp('2025-11-01')
        expected = date(2025, 11, 10)
        self.assertEqual(calculate_sexing_date(ts), expected)

    def test_sexing_date_in_schedule_dates(self):
        bd = date(2025, 11, 1)
        dates = calculate_schedule_dates(bd)
        self.assertIsNotNone(dates)
        self.assertIn('sexing_date', dates)
        self.assertEqual(dates['sexing_date'], date(2025, 11, 10))

    def test_sexing_date_offset_configurable(self):
        original = CONFIG['SEXING_OFFSET_DAYS']
        try:
            CONFIG['SEXING_OFFSET_DAYS'] = 7
            bd = date(2025, 11, 1)
            result = calculate_sexing_date(bd)
            self.assertEqual(result, date(2025, 11, 8))
        finally:
            CONFIG['SEXING_OFFSET_DAYS'] = original

    def test_next_wednesday(self):
        self.assertEqual(next_wednesday(date(2025, 11, 14)), date(2025, 11, 19))
        self.assertEqual(next_wednesday(date(2025, 11, 19)), date(2025, 11, 19))
        self.assertEqual(next_wednesday(date(2025, 11, 20)), date(2025, 11, 26))
        self.assertEqual(next_wednesday(date(2025, 11, 17)), date(2025, 11, 19))

    def test_calculate_schedule_dates(self):
        bd = date(2025, 11, 1)
        dates = calculate_schedule_dates(bd)
        self.assertIsNotNone(dates)
        self.assertEqual(dates['birth_date'], bd)
        self.assertEqual(dates['p14_harvest'], date(2025, 11, 15))
        self.assertEqual(dates['p56_behavior_window_start'], date(2025, 12, 13))
        self.assertEqual(dates['p56_behavior_window_end'], date(2025, 12, 20))
        self.assertEqual(dates['sexing_date'], date(2025, 11, 10))

    def test_is_valid_p14_day(self):
        self.assertTrue(is_valid_p14_day(date(2025, 11, 17)))
        self.assertTrue(is_valid_p14_day(date(2025, 11, 18)))
        self.assertTrue(is_valid_p14_day(date(2025, 11, 19)))
        self.assertTrue(is_valid_p14_day(date(2025, 11, 20)))
        self.assertTrue(is_valid_p14_day(date(2025, 11, 21)))
        self.assertFalse(is_valid_p14_day(date(2025, 11, 22)))
        self.assertFalse(is_valid_p14_day(date(2025, 11, 23)))
        self.assertFalse(is_valid_p14_day(None))

    def test_is_priority_strain(self):
        self.assertTrue(is_priority_strain('SHANK3'))
        self.assertTrue(is_priority_strain('shank3'))
        self.assertTrue(is_priority_strain('B6J'))
        self.assertTrue(is_priority_strain('CHD8'))
        self.assertFalse(is_priority_strain('NONEXISTENT_STRAIN'))
        self.assertFalse(is_priority_strain(None))
        self.assertFalse(is_priority_strain(''))

    def test_is_super_priority_strain(self):
        self.assertTrue(is_super_priority_strain('SHANK3'))
        self.assertTrue(is_super_priority_strain('shank3'))
        self.assertTrue(is_super_priority_strain('CHD8'))
        self.assertTrue(is_super_priority_strain('FMR1'))
        self.assertFalse(is_super_priority_strain('B6J'))
        self.assertFalse(is_super_priority_strain(None))
        self.assertFalse(is_super_priority_strain('NONEXISTENT'))

    def test_get_next_wednesdays_count(self):
        for n in [1, 3, 6, 10]:
            result = get_next_wednesdays(n, from_date=date(2025, 11, 17))
            self.assertEqual(len(result), n)

    def test_get_next_wednesdays_all_on_wednesday(self):
        wednesdays = get_next_wednesdays(6, from_date=date(2025, 11, 15))
        for wed in wednesdays:
            self.assertEqual(wed.weekday(), 2)

    def test_get_next_wednesdays_spacing(self):
        wednesdays = get_next_wednesdays(6, from_date=date(2025, 11, 15))
        for i in range(1, len(wednesdays)):
            delta = (wednesdays[i] - wednesdays[i - 1]).days
            self.assertEqual(delta, 7)

    def test_get_next_wednesdays_from_wednesday_includes_today(self):
        start = date(2025, 11, 19)
        wednesdays = get_next_wednesdays(6, from_date=start)
        self.assertEqual(wednesdays[0], start)

    def test_get_next_wednesdays_from_saturday(self):
        start = date(2025, 11, 15)
        wednesdays = get_next_wednesdays(6, from_date=start)
        self.assertEqual(wednesdays[0], date(2025, 11, 19))
        self.assertEqual(wednesdays[5], date(2025, 12, 24))

    # ── P14 today-is-too-late tests ───────────────────────────────────────────

    def test_p14_today_is_too_late(self):
        """P14 date == today should be ineligible."""
        today = datetime.now().date()
        birth = today - timedelta(days=14)
        test_data = pd.DataFrame({
            'Name': ['Animal1'],
            'Birth Date': [pd.Timestamp(birth)],
            'Sex': ['Male'],
            'Line (Short)': ['SHANK3'],
            'Genotype': ['Het'],
            'Use': ['Sing Inventory'],
            'Status': ['Alive'],
            'Birth ID': ['B001'],
            'Marker Type': ['Ear Punch'],
        })
        result = check_eligibility(test_data, full_behavior_dates=[])
        row = result.iloc[0]
        self.assertFalse(row['P14_Eligible'], "P14 date == today should NOT be eligible")
        self.assertTrue(row['P14_Too_Old'],    "P14 date == today should set Too_Old flag")
        self.assertIn('TOO LATE', row['P14_Reason'].upper())

    def test_p14_tomorrow_is_eligible(self):
        """P14 date == tomorrow should be eligible if a valid weekday."""
        today = datetime.now().date()
        tomorrow = today + timedelta(days=1)
        birth = tomorrow - timedelta(days=14)
        test_data = pd.DataFrame({
            'Name': ['Animal1'],
            'Birth Date': [pd.Timestamp(birth)],
            'Sex': ['Male'],
            'Line (Short)': ['SHANK3'],
            'Genotype': ['Het'],
            'Use': ['Sing Inventory'],
            'Status': ['Alive'],
            'Birth ID': ['B001'],
            'Marker Type': ['Ear Punch'],
        })
        result = check_eligibility(test_data, full_behavior_dates=[])
        row = result.iloc[0]
        if tomorrow.weekday() in CONFIG['P14_VALID_DAYS']:
            self.assertTrue(row['P14_Eligible'])
            self.assertFalse(row['P14_Too_Old'])
        else:
            self.assertFalse(row['P14_Eligible'])
            self.assertIn('valid harvest day', row['P14_Reason'].lower())

    def test_p14_yesterday_is_too_old(self):
        """P14 date == yesterday should be ineligible."""
        today = datetime.now().date()
        birth = today - timedelta(days=15)
        test_data = pd.DataFrame({
            'Name': ['Animal1'],
            'Birth Date': [pd.Timestamp(birth)],
            'Sex': ['Male'],
            'Line (Short)': ['SHANK3'],
            'Genotype': ['Het'],
            'Use': ['Sing Inventory'],
            'Status': ['Alive'],
            'Birth ID': ['B001'],
            'Marker Type': ['Ear Punch'],
        })
        result = check_eligibility(test_data, full_behavior_dates=[])
        row = result.iloc[0]
        self.assertFalse(row['P14_Eligible'])
        self.assertTrue(row['P14_Too_Old'])
        self.assertIn('TOO OLD', row['P14_Reason'].upper())

    # ── Age column tests ──────────────────────────────────────────────────────

    def test_eligibility_p56_age_values(self):
        test_data = pd.DataFrame({
            'Name': ['Animal1'],
            'Birth Date': [pd.Timestamp('2025-10-01')],
            'Sex': ['Male'],
            'Line (Short)': ['SHANK3'],
            'Genotype': ['Het'],
            'Use': ['Sing Inventory'],
            'Status': ['Alive'],
            'Birth ID': ['B001'],
            'Marker Type': ['Ear Punch'],
        })
        result = check_eligibility(test_data, full_behavior_dates=[])
        row = result.iloc[0]
        if row['P56_Age_At_Behavior_Days'] is not None:
            self.assertGreaterEqual(row['P56_Age_At_Behavior_Days'], 42)
            self.assertLessEqual(row['P56_Age_At_Behavior_Days'], 49)
            self.assertIsInstance(row['P56_Age_At_Behavior_Months'], int)
            self.assertEqual(
                row['P56_Age_At_Harvest_Days'],
                row['P56_Age_At_Behavior_Days'] + 14
            )
            self.assertIsInstance(row['P56_Age_At_Harvest_Months'], int)

    def test_eligibility_too_old_both_flags(self):
        old_birth = date(2020, 1, 1)
        test_data = pd.DataFrame({
            'Name': ['OldAnimal'],
            'Birth Date': [pd.Timestamp(old_birth)],
            'Sex': ['Male'],
            'Line (Short)': ['SHANK3'],
            'Genotype': ['Het'],
            'Use': ['Sing Inventory'],
            'Status': ['Alive'],
            'Birth ID': ['B001'],
            'Marker Type': ['Ear Punch'],
        })
        result = check_eligibility(test_data, full_behavior_dates=[])
        row = result.iloc[0]
        self.assertTrue(row['P14_Too_Old'], "Should be flagged too old for P14")
        self.assertTrue(row['P56_Too_Old'], "Should be flagged too old for P56")
        self.assertIn('TOO OLD', row['P14_Reason'].upper())
        self.assertIn('TOO OLD', row['P56_Reason'].upper())
        self.assertIn('Unusable_Note', result.columns)
        self.assertIn('UNUSABLE FOR BOTH', row['Unusable_Note'].upper())

    def test_eligibility_not_too_old_fresh_animal(self):
        recent_birth = datetime.now().date() - timedelta(days=3)
        test_data = pd.DataFrame({
            'Name': ['YoungAnimal'],
            'Birth Date': [pd.Timestamp(recent_birth)],
            'Sex': ['Female'],
            'Line (Short)': ['CHD8'],
            'Genotype': ['Het'],
            'Use': ['Sing Inventory'],
            'Status': ['Alive'],
            'Birth ID': ['B002'],
            'Marker Type': ['Ear Punch'],
        })
        result = check_eligibility(test_data, full_behavior_dates=[])
        row = result.iloc[0]
        self.assertFalse(row['P14_Too_Old'])
        self.assertFalse(row['P56_Too_Old'])

    def test_eligibility_age_today_days_present(self):
        today = datetime.now().date()
        birth = today - timedelta(days=20)
        test_data = pd.DataFrame({
            'Name': ['Animal1'],
            'Birth Date': [pd.Timestamp(birth)],
            'Sex': ['Male'],
            'Line (Short)': ['SHANK3'],
            'Genotype': ['Het'],
            'Use': ['Sing Inventory'],
            'Status': ['Alive'],
            'Birth ID': ['B003'],
            'Marker Type': ['Ear Punch'],
        })
        result = check_eligibility(test_data, full_behavior_dates=[])
        row = result.iloc[0]
        self.assertIn('Age_Today_Days', result.columns)
        self.assertEqual(row['Age_Today_Days'], 20)

    def test_eligibility_strain_column_present(self):
        test_data = pd.DataFrame({
            'Name': ['Animal1'],
            'Birth Date': [pd.Timestamp('2025-10-01')],
            'Sex': ['Male'],
            'Line (Short)': ['SHANK3'],
            'Genotype': ['Het'],
            'Use': ['Sing Inventory'],
            'Status': ['Alive'],
            'Birth ID': ['B001'],
            'Marker Type': ['Ear Punch'],
        })
        result = check_eligibility(test_data, full_behavior_dates=[])
        self.assertIn('Strain', result.columns)
        self.assertEqual(result.iloc[0]['Strain'], 'SHANK3')

    def test_eligibility_birth_date_column_present(self):
        test_data = pd.DataFrame({
            'Name': ['Animal1'],
            'Birth Date': [pd.Timestamp('2025-10-01')],
            'Sex': ['Male'],
            'Line (Short)': ['SHANK3'],
            'Genotype': ['Het'],
            'Use': ['Sing Inventory'],
            'Status': ['Alive'],
            'Birth ID': ['B001'],
            'Marker Type': ['Ear Punch'],
        })
        result = check_eligibility(test_data, full_behavior_dates=[])
        self.assertIn('Birth_Date', result.columns)

    def test_eligibility_animal_name_column_present(self):
        test_data = pd.DataFrame({
            'Name': ['MyAnimal'],
            'Birth Date': [pd.Timestamp('2025-10-01')],
            'Sex': ['Male'],
            'Line (Short)': ['SHANK3'],
            'Genotype': ['Het'],
            'Use': ['Sing Inventory'],
            'Status': ['Alive'],
            'Birth ID': ['B001'],
            'Marker Type': ['Ear Punch'],
        })
        result = check_eligibility(test_data, full_behavior_dates=[])
        self.assertIn('Animal_Name', result.columns)
        self.assertEqual(result.iloc[0]['Animal_Name'], 'MyAnimal')

    def test_eligibility_genotype_is_canonical(self):
        test_data = pd.DataFrame({
            'Name': ['Animal1'],
            'Birth Date': [pd.Timestamp('2025-10-01')],
            'Sex': ['Male'],
            'Line (Short)': ['SHANK3'],
            'Genotype': ['Het'],
            'Use': ['Sing Inventory'],
            'Status': ['Alive'],
            'Birth ID': ['B001'],
            'Marker Type': ['Ear Punch'],
        })
        result = check_eligibility(test_data, full_behavior_dates=[])
        geno = result.iloc[0]['Genotype']
        self.assertIn(geno, _CANONICAL_GENOTYPES,
                      f"'{geno}' is not a canonical genotype label")

    # ── Assignment tests ──────────────────────────────────────────────────────

    def test_assign_animals_smart_empty_input(self):
        result = assign_animals_smart(pd.DataFrame(), {})
        self.assertIsInstance(result, pd.DataFrame)
        self.assertEqual(len(result), 0)

    def test_assign_animals_smart_missing_strain_column(self):
        bad_df = pd.DataFrame({
            'Animal_Name': ['A1'],
            'Sex': ['Male'],
            'Genotype': ['Het'],
            'Birth_Date': ['2025-10-01'],
        })
        with self.assertRaises(KeyError):
            assign_animals_smart(bad_df, {})

    def test_assign_animals_smart_required_columns_present(self):
        today = datetime.now().date()
        birth = today - timedelta(days=20)
        test_data = pd.DataFrame({
            'Name': ['Animal1'],
            'Birth Date': [pd.Timestamp(birth)],
            'Sex': ['Male'],
            'Line (Short)': ['SHANK3'],
            'Genotype': ['Het'],
            'Use': ['Sing Inventory'],
            'Status': ['Alive'],
            'Birth ID': ['B001'],
            'Marker Type': ['Ear Punch'],
        })
        eligibility = check_eligibility(test_data, full_behavior_dates=[])
        result = assign_animals_smart(eligibility, {})
        self.assertIsInstance(result, pd.DataFrame)

    # ── parse_unschedulable_reason tests ─────────────────────────────────────

    def test_parse_unschedulable_genotype_deadline_passed(self):
        reason = (
            "[GENOTYPE DEADLINE PASSED] "
            "'Half' STRAIN — 2 blank genotype(s) from birth 2026-02-09. "
            "May have had ~1 Het but genotyping deadline passed (2026-02-22). "
            "P14 on 2026-02-23 or P56 on 2026-03-25"
        )
        parsed = parse_unschedulable_reason(reason)
        self.assertIn('Genotype Deadline', parsed['Primary_Reason'])
        self.assertEqual(parsed['Too_Old_For_P14'], 'NO')
        self.assertEqual(parsed['Too_Old_For_P56'], 'NO')
        self.assertEqual(parsed['Unusable_Both'],   'NO')
        self.assertNotIn('Window passed', parsed['P14_Status'])
        self.assertNotIn('Window passed', parsed['P56_Status'])
        self.assertIn('🧬', parsed['Primary_Reason'])

    def test_parse_unschedulable_windows_actually_passed(self):
        reason = "P14: TOO OLD FOR P14 — P14 date was 2025-01-01; P56: TOO OLD FOR P56"
        parsed = parse_unschedulable_reason(reason)
        self.assertNotIn('Genotype Deadline', parsed['Primary_Reason'])

    def test_parse_unschedulable_unusable_both(self):
        reason = '⛔ UNUSABLE FOR BOTH TIMEPOINTS — Too old for P14 AND P56. Animal is 1200d old.'
        parsed = parse_unschedulable_reason(reason)
        self.assertEqual(parsed['Too_Old_For_P14'], 'YES')
        self.assertEqual(parsed['Too_Old_For_P56'], 'YES')
        self.assertEqual(parsed['Unusable_Both'], 'YES')
        self.assertIn('⛔', parsed['Primary_Reason'])
        self.assertEqual(parsed['P14_Status'], '❌ Too old')
        self.assertEqual(parsed['P56_Status'], '❌ Too old')

    def test_parse_unschedulable_too_late_p14(self):
        reason = (
            '❌ TOO LATE FOR P14 — P14 date is today (2026-02-09) — '
            'harvest must be scheduled in advance'
        )
        parsed = parse_unschedulable_reason(reason)
        self.assertEqual(parsed['Too_Old_For_P14'], 'YES')
        self.assertEqual(parsed['Too_Old_For_P56'], 'NO')
        self.assertIn('Too Late', parsed['Primary_Reason'])

    def test_parse_unschedulable_too_old_p14_only(self):
        reason = '❌ TOO OLD FOR P14 — P14 date was 2025-01-01 (300 days ago)'
        parsed = parse_unschedulable_reason(reason)
        self.assertEqual(parsed['Too_Old_For_P14'], 'YES')
        self.assertEqual(parsed['Too_Old_For_P56'], 'NO')
        self.assertEqual(parsed['Unusable_Both'], 'NO')

    def test_parse_unschedulable_too_old_p56_only(self):
        reason = (
            '❌ TOO OLD FOR P56 — P56 behavior window ended 2025-01-20 '
            '(200 days ago). Unusable for P56.'
        )
        parsed = parse_unschedulable_reason(reason)
        self.assertEqual(parsed['Too_Old_For_P56'], 'YES')
        self.assertEqual(parsed['Too_Old_For_P14'], 'NO')
        self.assertEqual(parsed['Unusable_Both'], 'NO')
        self.assertIn('TOO OLD', parsed['Primary_Reason'].upper())

    def test_parse_unschedulable_quota_filled(self):
        reason = 'P56 quota filled for strain — reassigned to P14 (also filled)'
        parsed = parse_unschedulable_reason(reason)
        self.assertIn('Quota', parsed['Primary_Reason'])
        self.assertEqual(parsed['Too_Old_For_P14'], 'NO')
        self.assertEqual(parsed['Unusable_Both'], 'NO')

    def test_parse_unschedulable_capacity(self):
        reason = 'P56 date at capacity (2025-11-19); P14 unavailable: P14 date has passed'
        parsed = parse_unschedulable_reason(reason)
        self.assertIn('Capacity', parsed['Primary_Reason'])
        self.assertEqual(parsed['P56_Status'], '🔴 Date full')
        self.assertEqual(parsed['P14_Status'], '❌ Unavailable')

    def test_parse_unschedulable_incomplete_group(self):
        reason = 'Incomplete P56 group; P14 unavailable: P14 falls on Saturday'
        parsed = parse_unschedulable_reason(reason)
        self.assertIn('Incomplete', parsed['Primary_Reason'])
        self.assertIn('< 3', parsed['P56_Status'])

    def test_parse_unschedulable_toe_clip(self):
        reason = 'Has Toe Clip marker — not allowed for P56 behavior'
        parsed = parse_unschedulable_reason(reason)
        self.assertIn('Toe Clip', parsed['Primary_Reason'])
        self.assertIn('🚫', parsed['P56_Status'])

    def test_parse_unschedulable_invalid_day(self):
        reason = 'P14 falls on Saturday (2025-11-22) — not a valid harvest day'
        parsed = parse_unschedulable_reason(reason)
        self.assertIn('Invalid Day', parsed['Primary_Reason'])
        self.assertIn('⚠️', parsed['P14_Status'])

    def test_parse_unschedulable_no_birth_date(self):
        reason = 'No birth date'
        parsed = parse_unschedulable_reason(reason)
        self.assertIn('No Birth Date', parsed['Primary_Reason'])
        self.assertEqual(parsed['Too_Old_For_P14'], 'NO')
        self.assertEqual(parsed['Unusable_Both'], 'NO')

    def test_parse_unschedulable_empty(self):
        parsed = parse_unschedulable_reason('')
        self.assertIsNotNone(parsed)
        for key in [
            'Primary_Reason', 'P14_Status', 'P56_Status',
            'Too_Old_For_P14', 'Too_Old_For_P56', 'Unusable_Both', 'Detail'
        ]:
            self.assertIn(key, parsed, f"Missing key: {key}")

    def test_parse_unschedulable_none(self):
        parsed = parse_unschedulable_reason(None)
        self.assertIsNotNone(parsed)
        self.assertEqual(parsed['Primary_Reason'], 'Unknown')

    def test_parse_unschedulable_wild_genotype(self):
        reason = 'Wild genotype — not usable for harvest'
        parsed = parse_unschedulable_reason(reason)
        self.assertIn('Wild', parsed['Primary_Reason'])
        self.assertNotIn('Genotype Deadline', parsed['Primary_Reason'])
        self.assertNotIn('Blank', parsed['Primary_Reason'])

    def test_parse_unschedulable_blank_genotype(self):
        parsed = parse_unschedulable_reason('Blank')
        self.assertIn('Blank', parsed['Primary_Reason'])
        self.assertNotIn('Wild', parsed['Primary_Reason'])

    # ── Blank genotype misclassification tests ────────────────────────────────

    def test_blank_genotype_not_classified_as_wild(self):
        """Blank genotype must never appear as Wild Genotype in unschedulable report."""
        geno_excluded = pd.DataFrame([{
            'Animal_Name': 'BlankAnimal1',
            'Birth_ID': 'B099',
            'Strain': 'SHANK3',
            'Genotype': GENOTYPE_BLANK,
            'Sex': 'Male',
            'Birth_Date': '2025-10-01',
            'Reason': (
                "'Half' STRAIN — 2 blank genotype(s) from birth 2025-10-01. "
                "~1 of 2 expected Het (Het×WT cross). "
                "RECOMMEND: Genotype by 2025-11-12 (7 days) for P14 on 2025-11-15"
            ),
        }])

        report = create_unschedulable_report(
            pd.DataFrame(),
            pd.DataFrame(),
            geno_excluded
        )
        self.assertEqual(len(report), 1)
        row = report.iloc[0]

        self.assertEqual(row['Genotype'], GENOTYPE_BLANK,
            f"Blank genotype shown as '{row['Genotype']}' — should be '{GENOTYPE_BLANK}'")

        self.assertNotIn('Wild', row['Primary_Reason'],
            f"Primary_Reason incorrectly says '{row['Primary_Reason']}' "
            f"for a blank-genotype animal")

        self.assertIn('🧬', row['Primary_Reason'])
        self.assertIn('Blank', row['Primary_Reason'])

    def test_blank_genotype_nan_not_wild(self):
        """NaN genotype must resolve to Blank, not Wild."""
        geno_excluded = pd.DataFrame([{
            'Animal_Name': 'NaNGenoAnimal',
            'Birth_ID': 'B100',
            'Strain': 'CHD8',
            'Genotype': float('nan'),
            'Sex': 'Female',
            'Birth_Date': '2025-10-01',
            'Reason': 'blank genotype — scheduling analysis pending',
        }])

        report = create_unschedulable_report(
            pd.DataFrame(), pd.DataFrame(), geno_excluded
        )
        self.assertEqual(len(report), 1)
        self.assertEqual(report.iloc[0]['Genotype'], GENOTYPE_BLANK)
        self.assertNotEqual(report.iloc[0]['Genotype'], GENOTYPE_WILD)

    def test_blank_genotype_empty_string_not_wild(self):
        """Empty string genotype must resolve to Blank."""
        geno_excluded = pd.DataFrame([{
            'Animal_Name': 'EmptyGenoAnimal',
            'Birth_ID': 'B101',
            'Strain': 'FMR1',
            'Genotype': '',
            'Sex': 'Male',
            'Birth_Date': '2025-10-01',
            'Reason': "'All' STRAIN — 3 blank genotype(s) from birth 2025-10-01.",
        }])

        report = create_unschedulable_report(
            pd.DataFrame(), pd.DataFrame(), geno_excluded
        )
        self.assertEqual(len(report), 1)
        self.assertEqual(report.iloc[0]['Genotype'], GENOTYPE_BLANK)

    def test_parse_reason_blank_scheduling_analysis_not_wild(self):
        """Scheduling analysis strings for blank animals must not be parsed as Wild."""
        blank_scheduling_reasons = [
            "'Half' STRAIN — 2 blank genotype(s) from birth 2025-10-01. ~1 Het expected.",
            "'All' STRAIN — 4 blank genotype(s) from birth 2025-10-01. ALL 4 usable.",
            "UNKNOWN STRAIN — 1 blank genotype(s) from birth 2025-10-01.",
            "blank genotype analysis: not schedulable",
            "Blank",
            GENOTYPE_BLANK,
        ]
        for reason in blank_scheduling_reasons:
            parsed = parse_unschedulable_reason(reason)
            self.assertNotIn('Wild', parsed['Primary_Reason'],
                f"Reason '{reason[:60]}' was misclassified as Wild: "
                f"'{parsed['Primary_Reason']}'")
            self.assertIn('Blank', parsed['Primary_Reason'],
                f"Reason '{reason[:60]}' should be Blank but got: "
                f"'{parsed['Primary_Reason']}'")

    def test_parse_reason_inconclusive_not_blank(self):
        """Inconclusive must be reported distinctly from Blank."""
        inconclusive_reasons = [
            'Inconclusive',
            GENOTYPE_INCONC,
            'genotype inconclusive \u2014 released to available',
        ]
        for reason in inconclusive_reasons:
            parsed = parse_unschedulable_reason(reason)
            self.assertIn('Inconclusive', parsed['Primary_Reason'],
                f"Reason '{reason[:60]}' should be Inconclusive but got: "
                f"'{parsed['Primary_Reason']}'")
            self.assertNotIn('Wild', parsed['Primary_Reason'])

    def test_parse_reason_wild_genotype_exact_string(self):
        """Exact wild-genotype exclusion reason must still be classified as Wild."""
        wild_reasons = [
            'Wild genotype — not usable for harvest',
            GENOTYPE_WILD,
        ]
        for reason in wild_reasons:
            parsed = parse_unschedulable_reason(reason)
            self.assertIn('Wild', parsed['Primary_Reason'],
                f"Reason '{reason}' should be Wild but got: '{parsed['Primary_Reason']}'")
            self.assertNotIn('Blank', parsed['Primary_Reason'])

    def test_resolve_genotype_blank_canonical_stays_blank(self):
        """The canonical string 'Blank' must resolve to Blank."""
        geno_excluded = pd.DataFrame([{
            'Animal_Name': 'Test',
            'Birth_ID': 'B001',
            'Strain': 'SHANK3',
            'Genotype': 'Blank',
            'Sex': 'Male',
            'Birth_Date': '2025-10-01',
            'Reason': 'Blank',
        }])
        report = create_unschedulable_report(
            pd.DataFrame(), pd.DataFrame(), geno_excluded
        )
        self.assertEqual(report.iloc[0]['Genotype'], GENOTYPE_BLANK)
        self.assertNotEqual(report.iloc[0]['Genotype'], GENOTYPE_WILD)

    def test_wild_genotype_excluded_animals_show_wild(self):
        """Animals excluded in first pass as Wild must show Wild genotype in report."""
        wild_excluded = pd.DataFrame([{
            'Animal_Name': 'WildAnimal1',
            'Birth_ID': 'B200',
            'Strain': 'SHANK3',
            'Genotype': GENOTYPE_WILD,
            'Sex': 'Male',
            'Birth_Date': '2025-10-01',
            'Reason': 'Wild genotype — not usable for harvest',
        }])

        report = create_unschedulable_report(
            pd.DataFrame(), pd.DataFrame(), wild_excluded
        )
        self.assertEqual(len(report), 1)
        row = report.iloc[0]
        self.assertEqual(row['Genotype'], GENOTYPE_WILD,
            f"Wild animal shown as '{row['Genotype']}' — should be '{GENOTYPE_WILD}'")
        self.assertIn('Wild', row['Primary_Reason'],
            f"Primary_Reason should mention Wild but got: '{row['Primary_Reason']}'")
        self.assertNotIn('Blank', row['Primary_Reason'])

    # ── _assess_genotype_worth_it tests ───────────────────────────────────────

    def test_assess_genotype_worth_it_all_strain_both_available(self):
        worth = _assess_genotype_worth_it(
            num_blanks=4, breeding_type='All',
            p14_available=True, p56_available=True,
            is_schedulable=True,
            p14_date=date(2025, 12, 15),
            p56_date=date(2025, 12, 17),
            today=date(2025, 12, 1),
            remaining_needs={},
            strain='CNTNAP2',
            p56_group_size=0,
        )
        self.assertIn('✅ YES', worth['P14_Worth_Genotyping'])
        self.assertIn('✅ YES', worth['P56_Worth_Genotyping'])

    def test_assess_genotype_worth_it_half_strain_good_yield(self):
        worth = _assess_genotype_worth_it(
            num_blanks=6, breeding_type='Half',
            p14_available=True, p56_available=True,
            is_schedulable=True,
            p14_date=date(2025, 12, 15),
            p56_date=date(2025, 12, 17),
            today=date(2025, 12, 1),
            remaining_needs={},
            strain='SHANK3',
            p56_group_size=0,
        )
        self.assertIn('✅ YES', worth['P14_Worth_Genotyping'])
        self.assertIn('✅ YES', worth['P56_Worth_Genotyping'])

    def test_assess_genotype_worth_it_half_strain_low_yield(self):
        worth = _assess_genotype_worth_it(
            num_blanks=1, breeding_type='Half',
            p14_available=True, p56_available=False,
            is_schedulable=True,
            p14_date=date(2025, 12, 15),
            p56_date=None,
            today=date(2025, 12, 1),
            remaining_needs={},
            strain='SHANK3',
            p56_group_size=0,
        )
        self.assertIn('❌ UNLIKELY', worth['P14_Worth_Genotyping'])
        self.assertIn('❌ NO', worth['P56_Worth_Genotyping'])

    def test_assess_genotype_worth_it_not_schedulable(self):
        worth = _assess_genotype_worth_it(
            num_blanks=6, breeding_type='Half',
            p14_available=False, p56_available=False,
            is_schedulable=False,
            p14_date=date(2025, 1, 1),
            p56_date=date(2025, 1, 15),
            today=date(2025, 12, 1),
            remaining_needs={},
            strain='SHANK3',
            p56_group_size=0,
        )
        self.assertIn('❌ NO', worth['P14_Worth_Genotyping'])
        self.assertIn('❌ NO', worth['P56_Worth_Genotyping'])

    def test_assess_genotype_worth_it_quota_met(self):
        strain_key = 'SHANK3'
        remaining = {
            strain_key: {
                'P14': {
                    'Male':   {'Perfusion': {'needed': 0}, 'MERFISH': {'needed': 0}, 'RNAseq': {'needed': 0}},
                    'Female': {'Perfusion': {'needed': 0}, 'MERFISH': {'needed': 0}, 'RNAseq': {'needed': 0}},
                },
                'P56': {
                    'Male':   {'Perfusion': {'needed': 0}, 'MERFISH': {'needed': 0}, 'RNAseq': {'needed': 0}},
                    'Female': {'Perfusion': {'needed': 0}, 'MERFISH': {'needed': 0}, 'RNAseq': {'needed': 0}},
                },
            }
        }
        worth = _assess_genotype_worth_it(
            num_blanks=6, breeding_type='Half',
            p14_available=True, p56_available=True,
            is_schedulable=True,
            p14_date=date(2025, 12, 15),
            p56_date=date(2025, 12, 17),
            today=date(2025, 12, 1),
            remaining_needs=remaining,
            strain='SHANK3',
            p56_group_size=0,
        )
        self.assertIn('QUOTA MET', worth['P14_Worth_Genotyping'])
        self.assertIn('QUOTA MET', worth['P56_Worth_Genotyping'])

    def test_assess_genotype_worth_it_b6_never_quota_met(self):
        worth = _assess_genotype_worth_it(
            num_blanks=4, breeding_type='All',
            p14_available=True, p56_available=True,
            is_schedulable=True,
            p14_date=date(2025, 12, 15),
            p56_date=date(2025, 12, 17),
            today=date(2025, 12, 1),
            remaining_needs={},
            strain='B6J',
            p56_group_size=0,
        )
        self.assertNotIn('QUOTA MET', worth['P14_Worth_Genotyping'])
        self.assertNotIn('QUOTA MET', worth['P56_Worth_Genotyping'])

    def test_assess_genotype_worth_it_past_deadline(self):
        worth = _assess_genotype_worth_it(
            num_blanks=6, breeding_type='Half',
            p14_available=False, p56_available=False,
            is_schedulable=False,
            p14_date=date(2024, 1, 1),
            p56_date=date(2024, 1, 15),
            today=date(2025, 12, 1),
            remaining_needs={},
            strain='SHANK3',
            p56_group_size=0,
        )
        self.assertIn('❌ NO', worth['P14_Worth_Genotyping'])
        self.assertIn('❌ NO', worth['P56_Worth_Genotyping'])

    # ── NEW: P56 Wednesday-level grouping tests ───────────────────────────────

    def test_assess_genotype_worth_it_p56_group_size_18_half_strain(self):
        """18 females sharing one Wednesday → 9 expected Hets → full cage → YES."""
        worth = _assess_genotype_worth_it(
            num_blanks=3,
            breeding_type='Half',
            p14_available=True,
            p56_available=True,
            is_schedulable=True,
            p14_date=date(2026, 3, 31),
            p56_date=date(2026, 5, 6),
            today=date(2026, 3, 20),
            remaining_needs={},
            strain='SHANK3',
            p56_group_size=18,
        )
        # 18 * 0.5 = 9 expected Hets >= CAGE_SIZE(3) → YES
        self.assertIn('✅ YES', worth['P56_Worth_Genotyping'])
        self.assertIn('9', worth['P56_Worth_Genotyping'])

    def test_assess_genotype_worth_it_p56_group_size_4_half_strain(self):
        """4 animals in Wednesday window → 2 expected Hets → MAYBE (< cage size of 3)."""
        worth = _assess_genotype_worth_it(
            num_blanks=4,
            breeding_type='Half',
            p14_available=True,
            p56_available=True,
            is_schedulable=True,
            p14_date=date(2026, 3, 31),
            p56_date=date(2026, 5, 6),
            today=date(2026, 3, 20),
            remaining_needs={},
            strain='SHANK3',
            p56_group_size=4,
        )
        # 4 * 0.5 = 2 expected Hets < CAGE_SIZE(3) → MAYBE
        self.assertIn('🟡 MAYBE', worth['P56_Worth_Genotyping'])

    def test_assess_genotype_worth_it_p56_group_size_1_half_strain(self):
        """1 animal in Wednesday window → 0.5 expected Hets → UNLIKELY."""
        worth = _assess_genotype_worth_it(
            num_blanks=1,
            breeding_type='Half',
            p14_available=True,
            p56_available=True,
            is_schedulable=True,
            p14_date=date(2026, 3, 31),
            p56_date=date(2026, 5, 6),
            today=date(2026, 3, 20),
            remaining_needs={},
            strain='SHANK3',
            p56_group_size=1,
        )
        # 1 * 0.5 = 0.5 < 1 → UNLIKELY
        self.assertIn('❌ UNLIKELY', worth['P56_Worth_Genotyping'])

    def test_assess_genotype_worth_it_p56_group_all_strain_large_group(self):
        """'All' strain: 6 animals in window → all 6 usable → YES."""
        worth = _assess_genotype_worth_it(
            num_blanks=2,
            breeding_type='All',
            p14_available=True,
            p56_available=True,
            is_schedulable=True,
            p14_date=date(2026, 3, 31),
            p56_date=date(2026, 5, 6),
            today=date(2026, 3, 20),
            remaining_needs={},
            strain='CNTNAP2',
            p56_group_size=6,
        )
        # 6 * 1.0 = 6 >= CAGE_SIZE(3) → YES
        self.assertIn('✅ YES', worth['P56_Worth_Genotyping'])
        self.assertIn('6', worth['P56_Worth_Genotyping'])

    def test_assess_genotype_worth_it_p56_group_size_zero_falls_back(self):
        """p56_group_size=0 falls back to num_blanks."""
        worth_explicit = _assess_genotype_worth_it(
            num_blanks=6, breeding_type='Half',
            p14_available=True, p56_available=True,
            is_schedulable=True,
            p14_date=date(2026, 3, 31),
            p56_date=date(2026, 5, 6),
            today=date(2026, 3, 20),
            remaining_needs={}, strain='SHANK3',
            p56_group_size=6,
        )
        worth_fallback = _assess_genotype_worth_it(
            num_blanks=6, breeding_type='Half',
            p14_available=True, p56_available=True,
            is_schedulable=True,
            p14_date=date(2026, 3, 31),
            p56_date=date(2026, 5, 6),
            today=date(2026, 3, 20),
            remaining_needs={}, strain='SHANK3',
            p56_group_size=0,   # should use num_blanks=6
        )
        self.assertEqual(
            worth_explicit['P56_Worth_Genotyping'],
            worth_fallback['P56_Worth_Genotyping'],
        )

    def test_get_p56_behavior_wednesday_valid_birth(self):
        """Birth date of 2025-10-01 should map to a Wednesday in the P42-P49 window."""
        birth = date(2025, 10, 1)
        result = get_p56_behavior_wednesday(birth)
        self.assertIsNotNone(result)
        self.assertEqual(result.weekday(), 2)   # Wednesday
        dates = calculate_schedule_dates(birth)
        self.assertGreaterEqual(result, dates['p56_behavior_window_start'])
        self.assertLessEqual(result,   dates['p56_behavior_window_end'])

    def test_get_p56_behavior_wednesday_none_input(self):
        self.assertIsNone(get_p56_behavior_wednesday(None))

    def test_nearby_birth_dates_share_wednesday(self):
        """Birth dates 3/17/26-3/24/26 should all map to the same Wednesday."""
        births = [date(2026, 3, 17) + timedelta(days=i) for i in range(8)]
        wednesdays = [get_p56_behavior_wednesday(b) for b in births]
        valid = [w for w in wednesdays if w is not None]
        self.assertGreater(len(valid), 0)
        # All valid results should be the same Wednesday
        self.assertEqual(len(set(valid)), 1,
                         f"Expected one unique Wednesday, got: {set(valid)}")
        # That Wednesday should be 2026-05-06
        self.assertEqual(valid[0], date(2026, 5, 6))

    # ── B6 monthly summary tests ──────────────────────────────────────────────

    def test_b6_monthly_summary_empty_input(self):
        result = create_b6_monthly_summary(pd.DataFrame())
        self.assertIsInstance(result, pd.DataFrame)
        self.assertEqual(len(result), 0)

    def test_b6_monthly_summary_meets_minimum(self):
        today = datetime.now().date()
        target_date = today + timedelta(days=35)
        min_count = CONFIG.get('B6_MIN_PER_MONTH', 3)
        rows = []
        for i in range(min_count):
            rows.append({
                'Animal_Name': f'B6J_{i}',
                'Strain': 'B6J',
                'Assigned_Timepoint': 'P14',
                'P14_Date': target_date,
                'P56_Harvest_Date': None,
                'Priority': 'B6_CONTROL',
            })
        df = pd.DataFrame(rows)
        summary = create_b6_monthly_summary(df)
        if len(summary) > 0:
            row = summary.iloc[0]
            self.assertEqual(row['B6_B6N_Count'], min_count)
            self.assertIn('✅', row['Status'])
            self.assertEqual(row['Shortfall'], 0)

    def test_b6_monthly_summary_below_minimum(self):
        today = datetime.now().date()
        target_date = today + timedelta(days=35)
        min_needed = CONFIG.get('B6_MIN_PER_MONTH', 3)
        rows = [{
            'Animal_Name': 'B6J_0',
            'Strain': 'B6J',
            'Assigned_Timepoint': 'P14',
            'P14_Date': target_date,
            'P56_Harvest_Date': None,
            'Priority': 'B6_CONTROL',
        }]
        df = pd.DataFrame(rows)
        summary = create_b6_monthly_summary(df)
        if len(summary) > 0:
            row = summary.iloc[0]
            self.assertEqual(row['B6_B6N_Count'], 1)
            self.assertIn('⚠️', row['Status'])
            self.assertEqual(row['Shortfall'], min_needed - 1)

    def test_b6_monthly_summary_non_b6_not_counted(self):
        today = datetime.now().date()
        target_date = today + timedelta(days=35)
        rows = [
            {
                'Animal_Name': 'SHANK3_0',
                'Strain': 'SHANK3',
                'Assigned_Timepoint': 'P14',
                'P14_Date': target_date,
                'P56_Harvest_Date': None,
                'Priority': 'HIGH',
            },
            {
                'Animal_Name': 'SHANK3_1',
                'Strain': 'SHANK3',
                'Assigned_Timepoint': 'P14',
                'P14_Date': target_date,
                'P56_Harvest_Date': None,
                'Priority': 'HIGH',
            },
        ]
        df = pd.DataFrame(rows)
        summary = create_b6_monthly_summary(df)
        if len(summary) > 0:
            row = summary.iloc[0]
            self.assertEqual(row['B6_B6N_Count'], 0)
            self.assertIn('⚠️', row['Status'])

    # ── Births sexing schedule tests ──────────────────────────────────────────

    def test_build_births_sexing_schedule_empty(self):
        result = build_births_sexing_schedule(pd.DataFrame())
        self.assertIsInstance(result, pd.DataFrame)
        self.assertEqual(len(result), 0)

    def test_build_births_sexing_schedule_none(self):
        result = build_births_sexing_schedule(None)
        self.assertIsInstance(result, pd.DataFrame)
        self.assertEqual(len(result), 0)

    def test_build_births_sexing_schedule_columns(self):
        births = pd.DataFrame({
            'Birth ID': ['B001'],
            'Status': ['Sing Inventory'],
            'Birth Date': [pd.Timestamp(datetime.now().date() - timedelta(days=5))],
            'Live Count': [6],
            '# of Pups': [6],
            'Line (Short)': ['SHANK3'],
            'Dam': ['Dam1'],
            'Sire': ['Sire1'],
        })
        result = build_births_sexing_schedule(births)
        self.assertGreater(len(result), 0)
        for col in [
            'Birth_ID', 'Strain', 'Birth_Date', 'Num_Pups',
            'Sexing_Date', 'Day_of_Week', 'Days_Until_Sexing',
            'Sexing_Status', 'P14_Expected_Date', 'P14_Day_of_Week',
        ]:
            self.assertIn(col, result.columns, f"Missing column: {col}")

    def test_build_births_sexing_schedule_correct_date(self):
        birth_date = datetime.now().date() - timedelta(days=3)
        expected_sexing = birth_date + timedelta(days=9)
        births = pd.DataFrame({
            'Birth ID': ['B001'],
            'Status': ['Sing Inventory'],
            'Birth Date': [pd.Timestamp(birth_date)],
            'Live Count': [6],
            '# of Pups': [6],
            'Line (Short)': ['SHANK3'],
            'Dam': ['Dam1'],
            'Sire': ['Sire1'],
        })
        result = build_births_sexing_schedule(births)
        self.assertEqual(len(result), 1)
        self.assertEqual(
            result.iloc[0]['Sexing_Date'],
            expected_sexing.strftime('%Y-%m-%d')
        )

    def test_build_births_sexing_schedule_correct_p14_date(self):
        birth_date = datetime.now().date() - timedelta(days=3)
        expected_p14 = birth_date + timedelta(days=14)
        births = pd.DataFrame({
            'Birth ID': ['B001'],
            'Status': ['Sing Inventory'],
            'Birth Date': [pd.Timestamp(birth_date)],
            'Live Count': [6],
            '# of Pups': [6],
            'Line (Short)': ['SHANK3'],
            'Dam': ['Dam1'],
            'Sire': ['Sire1'],
        })
        result = build_births_sexing_schedule(births)
        self.assertEqual(
            result.iloc[0]['P14_Expected_Date'],
            expected_p14.strftime('%Y-%m-%d')
        )

    def test_build_births_sexing_schedule_filters_non_sing(self):
        births = pd.DataFrame({
            'Birth ID': ['B001', 'B002'],
            'Status': ['Sing Inventory', 'Other Status'],
            'Birth Date': [
                pd.Timestamp(datetime.now().date() - timedelta(days=3)),
                pd.Timestamp(datetime.now().date() - timedelta(days=5)),
            ],
            'Live Count': [6, 4],
            '# of Pups': [6, 4],
            'Line (Short)': ['SHANK3', 'CHD8'],
            'Dam': ['Dam1', 'Dam2'],
            'Sire': ['Sire1', 'Sire2'],
        })
        result = build_births_sexing_schedule(births)
        self.assertEqual(len(result), 1)
        self.assertEqual(result.iloc[0]['Birth_ID'], 'B001')

    def test_build_births_sexing_schedule_excludes_already_sexed(self):
        birth_date = datetime.now().date() - timedelta(days=5)
        births = pd.DataFrame({
            'Birth ID': ['B001', 'B002'],
            'Status': ['Sing Inventory', 'Sing Inventory'],
            'Birth Date': [pd.Timestamp(birth_date), pd.Timestamp(birth_date)],
            'Live Count': [6, 4],
            '# of Pups': [6, 4],
            'Line (Short)': ['SHANK3', 'CHD8'],
            'Dam': ['Dam1', 'Dam2'],
            'Sire': ['Sire1', 'Sire2'],
        })
        animals = pd.DataFrame({
            'Birth ID': ['B001', 'B001', 'B001'],
            'Name': ['A1', 'A2', 'A3'],
            'Status': ['Alive', 'Alive', 'Alive'],
        })
        result = build_births_sexing_schedule(births, animals_df=animals)
        self.assertEqual(len(result), 1)
        self.assertEqual(result.iloc[0]['Birth_ID'], 'B002')

    def test_build_births_sexing_schedule_no_animals_df(self):
        birth_date = datetime.now().date() - timedelta(days=5)
        births = pd.DataFrame({
            'Birth ID': ['B001', 'B002'],
            'Status': ['Sing Inventory', 'Sing Inventory'],
            'Birth Date': [pd.Timestamp(birth_date), pd.Timestamp(birth_date)],
            'Live Count': [6, 4],
            '# of Pups': [6, 4],
            'Line (Short)': ['SHANK3', 'CHD8'],
            'Dam': ['Dam1', 'Dam2'],
            'Sire': ['Sire1', 'Sire2'],
        })
        result = build_births_sexing_schedule(births, animals_df=None)
        self.assertEqual(len(result), 2)

    def test_build_births_sexing_schedule_urgency_today(self):
        birth_date = datetime.now().date() - timedelta(days=9)
        births = pd.DataFrame({
            'Birth ID': ['B001'],
            'Status': ['Sing Inventory'],
            'Birth Date': [pd.Timestamp(birth_date)],
            'Live Count': [6],
            '# of Pups': [6],
            'Line (Short)': ['SHANK3'],
            'Dam': ['Dam1'],
            'Sire': ['Sire1'],
        })
        result = build_births_sexing_schedule(births)
        self.assertEqual(len(result), 1)
        self.assertIn('TODAY', result.iloc[0]['Sexing_Status'])

    def test_build_births_sexing_schedule_urgency_done(self):
        birth_date = datetime.now().date() - timedelta(days=15)
        births = pd.DataFrame({
            'Birth ID': ['B001'],
            'Status': ['Sing Inventory'],
            'Birth Date': [pd.Timestamp(birth_date)],
            'Live Count': [6],
            '# of Pups': [6],
            'Line (Short)': ['SHANK3'],
            'Dam': ['Dam1'],
            'Sire': ['Sire1'],
        })
        result = build_births_sexing_schedule(births)
        self.assertEqual(len(result), 1)
        self.assertIn('Done', result.iloc[0]['Sexing_Status'])

    def test_build_births_sexing_schedule_no_birth_date(self):
        births = pd.DataFrame({
            'Birth ID': ['B001'],
            'Status': ['Sing Inventory'],
            'Birth Date': [pd.NaT],
            'Live Count': [6],
            '# of Pups': [6],
            'Line (Short)': ['SHANK3'],
            'Dam': ['Dam1'],
            'Sire': ['Sire1'],
        })
        result = build_births_sexing_schedule(births)
        self.assertEqual(len(result), 1)
        self.assertEqual(result.iloc[0]['Sexing_Date'], 'N/A')
        self.assertEqual(result.iloc[0]['Days_Until_Sexing'], 'N/A')

    # ── Misc utility tests ────────────────────────────────────────────────────

    def test_normalize_genotype_basic(self):
        self.assertEqual(normalize_genotype('-/+'), '-/+')
        self.assertEqual(normalize_genotype('  -/+  '), '-/+')

    def test_normalize_genotype_none(self):
        result = normalize_genotype(None)
        self.assertIsNone(result)

    def test_group_has_quota_b6j_always_true(self):
        self.assertTrue(group_has_quota('B6J', 'Male', 'P14', {}))
        self.assertTrue(group_has_quota('B6J', 'Female', 'P56', {}))
        self.assertTrue(group_has_quota('B6NJ', 'Male', 'P14', {}))

    def test_group_has_quota_empty_needs_true(self):
        self.assertTrue(group_has_quota('SHANK3', 'Male', 'P14', {}))

    def test_group_has_quota_with_needs(self):
        remaining = {
            'SHANK3': {
                'P14': {
                    'Male':   {'Perfusion': {'needed': 3}, 'MERFISH': {'needed': 1}, 'RNAseq': {'needed': 1}},
                    'Female': {'Perfusion': {'needed': 0}, 'MERFISH': {'needed': 0}, 'RNAseq': {'needed': 0}},
                },
                'P56': {
                    'Male':   {'Perfusion': {'needed': 0}, 'MERFISH': {'needed': 0}, 'RNAseq': {'needed': 0}},
                    'Female': {'Perfusion': {'needed': 0}, 'MERFISH': {'needed': 0}, 'RNAseq': {'needed': 0}},
                },
            }
        }
        self.assertTrue(group_has_quota('SHANK3', 'Male', 'P14', remaining))
        self.assertFalse(group_has_quota('SHANK3', 'Female', 'P14', remaining))
        self.assertFalse(group_has_quota('SHANK3', 'Male', 'P56', remaining))

    def test_diagnose_animal_file_runs_without_error(self):
        test_data = pd.DataFrame({
            'Name': ['Animal1', 'Animal2'],
            'Birth Date': [pd.Timestamp('2025-10-01'), pd.Timestamp('2025-09-15')],
            'Sex': ['Male', 'Female'],
            'Line (Short)': ['SHANK3', 'CHD8'],
            'Genotype': ['Het', 'Wild'],
            'Use': ['Sing Inventory', 'Other'],
            'Status': ['Alive', 'Alive'],
            'Birth ID': ['B001', 'B002'],
            'Marker Type': ['Ear Punch', 'Ear Punch'],
        })
        try:
            diagnose_animal_file(test_data)
        except Exception as e:
            self.fail(f"diagnose_animal_file raised an exception: {e}")

    def test_filter_animals_by_use_no_use_column(self):
        test_data = pd.DataFrame({
            'Name': ['Animal1'],
            'Birth Date': [pd.Timestamp('2025-10-01')],
            'Sex': ['Male'],
            'Line (Short)': ['SHANK3'],
            'Genotype': ['Het'],
            'Status': ['Alive'],
            'Birth ID': ['B001'],
            'Marker Type': ['Ear Punch'],
        })
        filtered, excluded = filter_animals_by_use(test_data)
        self.assertEqual(len(filtered), 1)
        self.assertEqual(len(excluded), 0)

    def test_filter_animals_by_use_sing_inventory(self):
        test_data = pd.DataFrame({
            'Name': ['Animal1', 'Animal2', 'Animal3'],
            'Birth Date': [
                pd.Timestamp('2025-10-01'),
                pd.Timestamp('2025-10-01'),
                pd.Timestamp('2025-10-01'),
            ],
            'Sex': ['Male', 'Female', 'Male'],
            'Line (Short)': ['SHANK3', 'SHANK3', 'CHD8'],
            'Genotype': ['Het', 'Het', 'Wild'],
            'Use': ['Sing Inventory', 'Other Use', 'Sing Inventory'],
            'Status': ['Alive', 'Alive', 'Alive'],
            'Birth ID': ['B001', 'B002', 'B003'],
            'Marker Type': ['Ear Punch', 'Ear Punch', 'Ear Punch'],
        })
        filtered, excluded = filter_animals_by_use(test_data)
        self.assertEqual(len(filtered), 2)
        self.assertEqual(len(excluded), 1)
        self.assertIn('Animal2', excluded['Animal_Name'].values)

    def test_filter_animals_by_genotype_excludes_wild(self):
        test_data = pd.DataFrame({
            'Name': ['A1', 'A2', 'A3'],
            'Birth Date': [pd.Timestamp('2025-10-01')] * 3,
            'Sex': ['Male', 'Female', 'Male'],
            'Line (Short)': ['SHANK3', 'SHANK3', 'SHANK3'],
            'Genotype': [GENOTYPE_HET, GENOTYPE_WILD, GENOTYPE_BLANK],
            'Use': ['Sing Inventory'] * 3,
            'Status': ['Alive'] * 3,
            'Birth ID': ['B001', 'B002', 'B003'],
            'Marker Type': ['Ear Punch'] * 3,
        })
        filtered, excluded, blanks = filter_animals_by_genotype_first_pass(test_data)
        self.assertEqual(len(filtered), 1)
        self.assertEqual(len(excluded), 1)
        self.assertEqual(len(blanks), 1)
        self.assertEqual(filtered.iloc[0]['Name'], 'A1')
        self.assertEqual(excluded.iloc[0]['Animal_Name'], 'A2')
        self.assertEqual(blanks.iloc[0]['Name'], 'A3')

    def test_age_at_p14_harvest_days(self):
        birth = date(2025, 11, 3)
        dates = calculate_schedule_dates(birth)
        self.assertIsNotNone(dates)
        age_days = (dates['p14_harvest'] - birth).days
        self.assertEqual(age_days, 14)

    def test_age_at_p56_behavior_range(self):
        birth = date(2025, 10, 1)
        dates = calculate_schedule_dates(birth)
        self.assertIsNotNone(dates)
        first_wed = next_wednesday(dates['p56_behavior_window_start'])
        self.assertIsNotNone(first_wed)
        self.assertLessEqual(first_wed, dates['p56_behavior_window_end'])
        age_days = (first_wed - birth).days
        self.assertGreaterEqual(age_days, 42)
        self.assertLessEqual(age_days, 49)

    def test_age_at_p56_harvest_is_behavior_plus_14(self):
        birth = date(2025, 10, 1)
        dates = calculate_schedule_dates(birth)
        first_wed = next_wednesday(dates['p56_behavior_window_start'])
        harvest = first_wed + timedelta(days=P56_HARVEST_OFFSET_FROM_BEHAVIOR)
        behavior_age = (first_wed - birth).days
        harvest_age = (harvest - birth).days
        self.assertEqual(harvest_age - behavior_age, 14)

    def test_age_months_rounding_boundaries(self):
        self.assertEqual(round(14 / 30.44), 0)
        self.assertEqual(round(42 / 30.44), 1)
        self.assertEqual(round(45 / 30.44), 1)
        self.assertEqual(round(49 / 30.44), 2)
        self.assertEqual(round(56 / 30.44), 2)
        self.assertEqual(round(63 / 30.44), 2)




# ============================================================================
# HARVEST PIPELINE — CONFIGURATION
# ============================================================================

SCHEDULE_FILE_PREFIX = "Complete_Schedule"

LABELS_ACROSS = 5
LABELS_DOWN = 17
LABELS_PER_PAGE = LABELS_ACROSS * LABELS_DOWN  # 85

ADD_SAMPLE_COLUMNS = [
    'Sample Name', 'Type', 'Status', 'Preservation Method', 'Date Harvest',
    'Date Expiration', 'Description', 'Source AnimalID', 'Source SampleID',
    'Volume', 'Volume Units', 'Project', 'Notes'
]

ENVISION_TEMPLATE_COLUMNS = [
    'Group', 'Cage', 'Animal ID', 'Envision Ear Tag', 'Strain',
    'Coat Color', 'Genotype', 'Additional Detail', 'Sex',
    'Birth Date', 'Ear notch', 'Metal ear tag', 'Other ID',
    'RapID code', 'RapID tag color', 'RFID', 'Tail Tattoo'
]

HARVEST_SHEET_COLUMNS = [
    'Name', 'Sample Number', 'Line', 'BD', 'Housing', 'Identification',
    'Sex', 'Age (Days)', 'Envision Date', 'Harvest Date', 'Harvested by',
    'Protocol', 'Time Pickup', 'Time Start', 'Pickup to Harvest Time',
    'Weight g', '4% Tribro mL 10-14', '4% Tribro Units P14-10%', 'Dye',
    '4% PFA per mouse', 'Time Complete', 'Round Duration', 'Perfusion Quality',
    '4% PFA Total',
    'Distilled Water', '2xPBS', '16% PFA', 'Notes'
]

PROTOCOL_SORT_ORDER = {
    '8 Weeks - 20mL PBS 25mL 4%PFA (Plus 20mL Storage) - 6mL/min': 0,
    'P14 - 15mL PBS 20mL 4%PFA (Plus 20mL Storage) - 4mL/min': 1,
    'MERFISH - OCT': 2,
    'RNA-Seq': 3,
    'Extra - Sex & Timepoint Full': 4
}



# ============================================================================
# HARVEST PIPELINE — UTILITIES & PIPELINE FUNCTIONS
# ============================================================================

def format_date_only(val):
    """Convert to date only string (no time component)."""
    try:
        dt = pd.to_datetime(val)
        if pd.notna(dt):
            return dt.strftime('%m/%d/%Y')
    except Exception:
        pass
    if pd.notna(val) and str(val).strip() != '':
        return str(val)
    return ''


def combine_sample_numbers(sample_list):
    """Combine sample numbers into range format."""
    if not sample_list:
        return ""
    base_numbers = []
    for sample in sample_list:
        sample_str = str(sample)
        if '-' in sample_str:
            base_num = sample_str.split('-')[0]
        else:
            base_num = sample_str
        try:
            base_numbers.append(int(base_num))
        except (ValueError, TypeError):
            continue
    if not base_numbers:
        return ""
    if len(base_numbers) == 1:
        return str(base_numbers[0])
    else:
        return f"{min(base_numbers)}-{max(base_numbers)}"


def clean_genotype_base(genotype, strain):
    """Remove <content>, Probe, Generic LacZ tg/0, and zygosity markers."""
    if pd.isna(genotype):
        return ""
    if pd.notna(strain):
        strain_str = str(strain).strip()
        if strain_str == 'C57BL/6NJ':
            return 'B6NJ'
        elif strain_str == 'C57BL/6J':
            return 'B6J'
    result = str(genotype)
    result = re.sub(r'<[^>]*>', '', result)
    result = re.sub(r'‹[^›]*›', '', result)
    result = re.sub(r'â€¹[^â€º]*â€º', '', result)
    result = re.sub(r'\[[^\]]*\]', '', result)
    result = re.sub(r'\([^\)]*\)', '', result)
    for ch in ['<', '>', '‹', '›', 'â€¹', 'â€º', '[', ']', '(', ')']:
        result = result.replace(ch, '')
    result = re.sub(r'Probe\s*', '', result)
    result = re.sub(r'Generic LacZ tg/0,\s*', '', result)
    for zyg in ['-/-', '-/+', '+/-', '-/Y', '+/Y']:
        result = result.replace(zyg, '')
    result = result.replace('Inbred', '')
    result = re.sub(r'\s+', ' ', result)
    return result.strip()


def clean_genotype(genotype):
    """Convert any Climb genotype string to standard symbol for output."""
    return genotype_to_symbol(genotype)


def genotype_to_symbol(genotype, strain: str = '') -> str:
    """
    Convert any raw Climb genotype string to a standard display symbol:
      +/-   Het / Carrier
      -/-   Hom / Knockout
      -/Y   Hemi / X-linked
      +/+   Wild-type
      Inbred  B6/B6NJ inbred
      Blank   Unknown / ungenotyped
    """
    canon = canonicalize_genotype(genotype, strain)
    return {
        GENOTYPE_HET:    '+/-',
        GENOTYPE_HOM:    '-/-',
        GENOTYPE_HEMI:   '-/Y',
        GENOTYPE_WILD:   '+/+',
        GENOTYPE_INBRED: 'Inbred',
        GENOTYPE_BLANK:  'Blank',
    }.get(canon, 'Blank')


def clean_genotype_labels(genotype):
    """Clean genotype specifically for label formatting — returns standard symbol."""
    return genotype_to_symbol(genotype)


def natural_sort_key(name):
    """Create a sort key that handles numbers naturally."""
    if pd.isna(name):
        return []
    parts = re.split(r'(\d+)', str(name))
    return [int(part) if part.isdigit() else part.lower() for part in parts]


def translate_protocol(harvest_type, timepoint):
    """Translate Harvest_Type + Assigned_Timepoint to full protocol name."""
    harvest_type = str(harvest_type).strip()
    timepoint = str(timepoint).strip()
    # Strip NB suffix — same protocol as regular type
    base = harvest_type.replace(' NB', '').strip()

    if base == 'Perfusion':
        if timepoint == 'P14':
            return 'P14 - 15mL PBS 20mL 4%PFA (Plus 20mL Storage) - 4mL/min'
        else:
            return '8 Weeks - 20mL PBS 25mL 4%PFA (Plus 20mL Storage) - 6mL/min'
    elif base == 'MERFISH':
        return 'MERFISH - OCT'
    elif base == 'RNAseq':
        return 'RNA-Seq'
    elif base in ('COMPLETE (Quota Filled)', 'Extra'):
        return 'Extra - Sex & Timepoint Full'
    else:
        return 'Extra - Sex & Timepoint Full'


def get_harvest_date(row):
    """Get the appropriate harvest date based on timepoint."""
    timepoint = str(row.get('Assigned_Timepoint', '')).strip()
    if timepoint == 'P14':
        return row.get('P14_Date', '')
    elif timepoint == 'P56':
        return row.get('P56_Harvest_Date', '')
    return ''


def get_age_days(row):
    """Get age at harvest in days based on timepoint."""
    timepoint = str(row.get('Assigned_Timepoint', '')).strip()
    if timepoint == 'P14':
        return row.get('P14_Age_At_Harvest_Days', '')
    elif timepoint == 'P56':
        return row.get('P56_Age_At_Harvest_Days', '')
    return ''


def get_envision_date(row):
    """Get envision (behavior) date — only for P56. NB animals return 'NB'."""
    harvest_type = str(row.get('Harvest_Type', '')).strip()
    if harvest_type.endswith(' NB'):
        return 'NB'
    timepoint = str(row.get('Assigned_Timepoint', '')).strip()
    if timepoint == 'P56':
        return row.get('P56_Behavior_Date', '')
    return ''


def auto_width_worksheet(ws):
    """Auto-adjust column widths based on content."""
    for column in ws.columns:
        max_length = 0
        col_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                cell_len = len(str(cell.value)) if cell.value is not None else 0
                if cell_len > max_length:
                    max_length = cell_len
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = max(max_length + 3, 8)


def save_df_to_excel(df, filepath, sheet_name='Sheet1'):
    """Save DataFrame to Excel with auto-width columns."""
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    for col_num, header in enumerate(df.columns, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = Font(bold=True)

    for row_num, row_data in enumerate(df.values, 2):
        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num)
            if pd.isna(value):
                cell.value = ''
            else:
                cell.value = value

    auto_width_worksheet(ws)
    wb.save(filepath)


def get_preservation_method(protocol):
    """Determine preservation method based on protocol."""
    protocol = str(protocol).strip()
    if "MERFISH - OCT" in protocol:
        return "OCT Block"
    elif "RNA-Seq" in protocol:
        return "Flash Frozen"
    elif "PFA" in protocol:
        return "4% PFA Fixed"
    else:
        return ""


def get_sample_count(protocol):
    """Determine how many samples to generate based on protocol."""
    protocol = str(protocol).strip()
    if protocol == "8 Weeks - 20mL PBS 25mL 4%PFA (Plus 20mL Storage) - 6mL/min":
        return (1, [""])
    elif protocol == "P14 - 15mL PBS 20mL 4%PFA (Plus 20mL Storage) - 4mL/min":
        return (1, [""])
    elif protocol == "MERFISH - OCT":
        return (2, ["", ""])
    elif protocol == "RNA-Seq":
        return (8, ["-0", "-1", "-2", "-3", "-4", "-5", "-6", "-C"])
    elif protocol == "Extra - Sex & Timepoint Full":
        return (0, [])
    elif protocol in ("WT Envison Controls", "WT Envision Controls", "Found Dead"):
        # Behavior-only (WT Envision controls) or nothing to collect (Found Dead).
        # Expected to produce no sample — not an error, so no alarm.
        return (0, [])
    else:
        print(f"  Warning: Unknown protocol '{protocol}'. Defaulting to 1.")
        return (1, [""])


def sort_working_df(df):
    """Sort by Protocol order -> Line (Short) -> Animal_Name (natural sort)."""
    sort_keys = []
    for idx, row in df.iterrows():
        sort_keys.append((
            idx,
            row.get('Protocol_Sort', 99),
            str(row.get('Line (Short)', '')),
            natural_sort_key(row.get('Animal_Name', ''))
        ))
    sort_keys.sort(key=lambda x: (x[1], x[2], x[3]))
    sorted_indices = [item[0] for item in sort_keys]
    return df.loc[sorted_indices].reset_index(drop=True)


# ============================================================
# LOAD DATA
# ============================================================



def build_working_data(all_animals_df):
    """Build working dataset. Filter Unschedulable. Translate protocols."""
    print("\n  Building working data...")
    df = all_animals_df.copy()

    before = len(df)
    df = df[df['Assigned_Timepoint'] != 'Unschedulable'].copy()
    print(f"  Filtered: {before - len(df)} Unschedulable removed")
    print(f"  Remaining: {len(df)} animals")

    # Flag NB animals — no behavior, so no Envision tagging needed
    df['_is_nb'] = df['Harvest_Type'].apply(
        lambda h: str(h).strip().endswith(' NB')
    )
    nb_count = df['_is_nb'].sum()
    if nb_count > 0:
        print(f"  NB animals (no behavior, no Envision): {nb_count}")

    df['Protocol'] = df.apply(
        lambda row: translate_protocol(
            row.get('Harvest_Type', ''), row.get('Assigned_Timepoint', '')),
        axis=1)
    df['Harvest_Date'] = df.apply(get_harvest_date, axis=1)
    df['Age_Days'] = df.apply(get_age_days, axis=1)
    df['Envision_Date'] = df.apply(get_envision_date, axis=1)
    df['Protocol_Sort'] = df['Protocol'].map(PROTOCOL_SORT_ORDER).fillna(99)
    df['Preservation'] = df['Protocol'].apply(get_preservation_method)

    # Calculate age in weeks
    df['Age_Weeks'] = ''
    for idx, row in df.iterrows():
        try:
            bd = pd.to_datetime(row.get('Birth_Date'))
            hd = pd.to_datetime(row.get('Harvest_Date'))
            if pd.notna(bd) and pd.notna(hd):
                df.at[idx, 'Age_Weeks'] = round((hd - bd).days / 7, 1)
        except Exception:
            pass

    print(f"\n  Protocol breakdown:")
    for protocol, count in df['Protocol'].value_counts().items():
        print(f"    {protocol}: {count}")

    print(f"\n  Columns available in working data ({len(df.columns)}):")
    print(f"    {list(df.columns)}")

    return df


# ============================================================
# STEPS 0+1: BUILD HARVEST WORKSHEET & CREATE SAMPLES
# ============================================================

def _next_sample_from_harvest(harvest_path: str) -> int:
    """Read Harvest Worksheet and return max valid sample number + 1.

    Handles all known formats: plain ints, ranges (1753-1754),
    NB suffix (1224 NB), old A-suffix (295-295A), comma-separated (346, 347).
    Skips Fail, QC Fail, Extra NB, Floxed, Found Dead, Extra, and blanks.
    """
    SKIP = {'fail', 'qc fail', 'extra nb', 'floxed', 'found dead', 'extra'}

    def _parse_max(val: str) -> int:
        nums = [int(n) for n in re.findall(r'\d+', val)]
        return max(nums) if nums else 0

    df = pd.read_excel(harvest_path, sheet_name='Harvest Worksheet', dtype=str).fillna('')
    max_num = 0
    for raw in df['Sample Number']:
        s = str(raw).strip()
        if not s or s.lower() in SKIP:
            continue
        n = _parse_max(s)
        if n > max_num:
            max_num = n
    return max_num + 1


# ── TEST MODE ────────────────────────────────────────────────────────────────
# When on, NOTHING is written to Climb. Every live PUT/POST is intercepted and
# logged instead of sent, and the Climb import CSV is renamed so it cannot be
# uploaded by accident. Read-only GETs still run so the pipeline behaves
# normally and all output files are still produced for inspection.
TEST_MODE = False


class _FakeResponse:
    """Stand-in for a requests Response so callers' `if not r.ok` still works."""
    ok = True
    status_code = 200
    text = '[TEST MODE] not sent'
    def json(self):
        return {}


def _api_put(url, **kw):
    if TEST_MODE:
        print(f"    [TEST MODE] SKIPPED PUT  {url}")
        return _FakeResponse()
    return requests.put(url, **kw)


def _api_post(url, **kw):
    if TEST_MODE:
        print(f"    [TEST MODE] SKIPPED POST {url}")
        return _FakeResponse()
    return requests.post(url, **kw)


def _load_sing_climb():
    """Load sing_climb from lib\\, falling back to the script folder.

    Handles both sing_climb.py and date-suffixed sing_climb_YYYYMMDD.py.
    Returns the loaded module.
    """
    import importlib.util as _ilu

    # Try bare import first — lib\ and the script folder are already on sys.path
    try:
        import sing_climb as _sc
        return _sc
    except ModuleNotFoundError:
        pass

    # Fall back: find sing_climb*.py, most recent by filename, lib\ first
    candidates = []
    for _d in (_LIB_DIR, _SCRIPT_DIR):
        if os.path.isdir(_d):
            candidates += sorted(glob.glob(os.path.join(_d, 'sing_climb*.py')),
                                 reverse=True)
    if not candidates:
        raise ModuleNotFoundError(
            'sing_climb not found. Searched:\n'
            f'  {_LIB_DIR}\n'
            f'  {_SCRIPT_DIR}\n'
            'Place sing_climb.py (or sing_climb_YYYYMMDD.py) in the lib folder.'
        )
    spec = _ilu.spec_from_file_location('sing_climb', candidates[0])
    mod  = _ilu.module_from_spec(spec)
    spec.loader.exec_module(mod)
    return mod


def get_starting_sample_number():
    """Fetch next sample number from Climb (globally sequential across all projects)."""
    try:
        _sc = _load_sing_climb()
        return _sc.get_next_sample_number(verbose=False)
    except Exception as ex:
        logger.error(f'Could not fetch next sample number from Climb: {ex}')
        raise


def _is_perfusion_protocol(protocol) -> bool:
    """
    True for the protocols that actually perfuse — the ones that consume PFA.

    MERFISH, RNA-Seq and Extra rows get 0, matching how the sheet is filled in
    by hand.
    """
    p = str(protocol).strip().lower()
    if not p:
        return False
    if 'merfish' in p or 'rna-seq' in p or 'rnaseq' in p:
        return False
    if p.startswith('extra'):
        return False
    return 'pbs' in p and 'pfa' in p


def add_harvest_reagent_totals(harvest_df):
    """
    Fill the batch reagent columns on the first row of each harvest date.

        4% PFA Total    = PFA per mouse x number of perfusions that day
        Distilled Water = Total x 0.25
        2xPBS           = Total x 0.50
        16% PFA         = Total x 0.25

    Matches how the sheet is filled in by hand: one set of numbers per harvest
    day, sitting on the first row, since the solution is mixed as a single
    batch.
    """
    if harvest_df.empty or 'Harvest Date' not in harvest_df.columns:
        return harvest_df

    ratios = CONFIG.get('PFA_MIX_RATIOS',
                        {'Distilled Water': 0.25, '2xPBS': 0.50, '16% PFA': 0.25})

    # object dtype — pandas 3 infers str from '' and then refuses numbers
    for col in ('4% PFA Total', *ratios):
        if col in harvest_df.columns:
            harvest_df[col] = pd.Series([''] * len(harvest_df),
                                        index=harvest_df.index, dtype=object)

    for date_val, group in harvest_df.groupby('Harvest Date', sort=False):
        n_perf = sum(1 for _, r in group.iterrows()
                     if _is_perfusion_protocol(r.get('Protocol')))
        if n_perf == 0:
            continue

        total = n_perf * CONFIG['PFA_PER_MOUSE_ML']
        first = group.index[0]
        harvest_df.at[first, '4% PFA Total'] = total
        for col, frac in ratios.items():
            if col in harvest_df.columns:
                harvest_df.at[first, col] = round(total * frac, 2)

        print(f'    {date_val}: {n_perf} perfusion(s) \u2192 {total} mL 4% PFA')

    return harvest_df


def run_harvest_and_samples(working_df, timestamp):
    """
    STEPS 0+1: Build harvest worksheet AND create samples together.
    Returns:
        harvest_df: Complete harvest worksheet with sample numbers
        samples_for_chain: DataFrame for Steps 2 and 4 (uses Animal_Name as Source)
        climb_import_df: DataFrame for Climb import (uses Animal ID)
    """
    print("\n" + "=" * 80)
    print("STEPS 0+1: BUILD HARVEST WORKSHEET & CREATE SAMPLES")
    print("=" * 80)

    if working_df.empty:
        print("  ✗ No data. Skipping.")
        return pd.DataFrame(), pd.DataFrame(), pd.DataFrame()

    # Sort working data
    sorted_df = sort_working_df(working_df)

    # Get starting sample number
    next_sample_num = get_starting_sample_number()
    print(f"  Starting with sample number: {next_sample_num}")

    # Build animal ID lookup: Animal_Name → JCMS ID (integer)
    animal_lookup = {}
    if 'ID' in sorted_df.columns:
        for _, row in sorted_df.iterrows():
            aname = str(row.get('Animal_Name', '')).strip()
            aid   = row.get('ID', '')
            if aname and pd.notna(aid) and str(aid).strip() not in ('', 'nan'):
                try:
                    animal_lookup[aname] = str(int(float(aid)))
                except (ValueError, TypeError):
                    animal_lookup[aname] = str(aid).strip()
        print(f"  Built Animal ID lookup: {len(animal_lookup)} entries")
    else:
        print("  ⚠ No 'ID' column in working data — Source AnimalID will fall back to animal name")

    # Process each animal
    harvest_rows = []
    climb_import_rows = []
    chain_rows = []
    samples_added = 0
    preservation_failures = []   # (animal, protocol) — no Preservation, no sample

    for idx, row in sorted_df.iterrows():
        animal_name = str(row.get('Animal_Name', '')).strip()
        protocol = row.get('Protocol', '')
        harvest_date = row.get('Harvest_Date', '')
        envision_date = row.get('Envision_Date', '')
        preservation = get_preservation_method(protocol)
        count, suffixes = get_sample_count(protocol)

        # TAILS guarantee: a sample is never created without a Preservation.
        # A blank Preservation in Climb silently mis-routes tube labels and
        # drops the sample from the Deliverables trackers (which filter on the
        # exact string). If the protocol yields no preservation, create nothing
        # and record it for the alarm below.
        if count > 0 and str(preservation).strip() == '':
            preservation_failures.append((animal_name, protocol))
            count, suffixes = 0, []

        animal_id = animal_lookup.get(animal_name, animal_name)

        # Format dates without time
        harvest_date_str = format_date_only(harvest_date)
        envision_date_str = format_date_only(envision_date)
        birth_date_str = format_date_only(row.get('Birth_Date', ''))

        # Generate samples for this animal
        animal_samples = []
        for i in range(count):
            sample_name = f"{next_sample_num}{suffixes[i]}"

            # Climb import row (uses Animal ID)
            climb_import_rows.append({
                'Sample Name': sample_name,
                'Type': 'Brain',
                'Status': 'Available',
                'Preservation Method': preservation,
                'Date Harvest': harvest_date_str,
                'Date Expiration': '',
                'Description': '',
                'Source AnimalID': animal_id,
                'Source SampleID': '',
                'Volume': '',
                'Volume Units': '',
                'Project': '',
                'Notes': ''
            })

            # Chain row (uses Animal_Name for merging in Steps 2 and 4)
            chain_rows.append({
                'Name': sample_name,
                'Source': animal_name,
                'Preservation': preservation,
                'Harvest Date': harvest_date_str,
            })

            animal_samples.append(sample_name)
            next_sample_num += 1
            samples_added += 1

        # Combined sample number for harvest sheet.
        # Extras produce no samples, so combine_sample_numbers() returns '' —
        # label them 'Extra' instead of leaving the cell blank to be typed by hand.
        # Every NB animal gets ' NB' appended, real samples and Extras alike.
        # This column is what gets read when loading Envision cages, so the flag
        # has to be visible there and not only in the Envision Date column.
        _htype = str(row.get('Harvest_Type', '')).strip()
        _is_nb = _htype.endswith(' NB')
        _base_htype = _htype.replace(' NB', '').strip()

        combined_sample = combine_sample_numbers(animal_samples)
        if not combined_sample and _base_htype in ('Extra', 'COMPLETE (Quota Filled)'):
            combined_sample = 'Extra'
        if combined_sample and _is_nb:
            combined_sample = f'{combined_sample} NB'

        # Harvest sheet row
        harvest_rows.append({
            'Name': animal_name,
            'Sample Number': combined_sample,
            'Line': row.get('Line (Short)', ''),
            'BD': birth_date_str,
            'Housing': row.get('Housing ID', ''),
            'Identification': row.get('Marker', row.get('Marker_Type', '')),
            'Sex': row.get('Sex', ''),
            'Age (Days)': f"P{int(row.get('Age_Days', ''))}" if row.get('Age_Days', '') != '' else '',
            'Envision Date': envision_date_str,
            'Harvest Date': harvest_date_str,
            'Harvested by': '',
            'Protocol': protocol,
            'Time Pickup': '',
            'Time Start': '',
            'Pickup to Harvest Time': '',
            'Weight g': '',
            '4% Tribro mL 10-14': 0,
            '4% Tribro Units P14-10%': 0,
            'Dye': 0,
            '4% PFA per mouse': (CONFIG['PFA_PER_MOUSE_ML']
                                 if _is_perfusion_protocol(protocol) else 0),
            'Time Complete': '',
            'Round Duration': '',
            # Blank on perfusions — filled in by hand after the harvest.
            'Perfusion Quality': ('' if _is_perfusion_protocol(protocol)
                                  else 'Not Perfusion'),
            '4% PFA Total': '',
            'Distilled Water': '',
            '2xPBS': '',
            '16% PFA': '',
            'Notes': ''
        })

    # Build DataFrames
    harvest_df = pd.DataFrame(harvest_rows, columns=HARVEST_SHEET_COLUMNS)
    print('  Reagent totals per harvest date:')
    harvest_df = add_harvest_reagent_totals(harvest_df)
    climb_import_df = pd.DataFrame(climb_import_rows, columns=ADD_SAMPLE_COLUMNS)
    samples_for_chain = pd.DataFrame(chain_rows)

    # Debug info for chain
    print(f"\n  DEBUG chain data:")
    print(f"    samples_for_chain columns: {list(samples_for_chain.columns)}")
    print(f"    samples_for_chain rows: {len(samples_for_chain)}")
    if len(samples_for_chain) > 0:
        print(f"    First row: {samples_for_chain.iloc[0].to_dict()}")
        print(f"    Unique Source values (first 5): {samples_for_chain['Source'].unique()[:5].tolist()}")

    # ── ALARM: protocols that produced no Preservation ────────────────────────
    # These animals got NO sample. Nothing with a blank Preservation is ever
    # written to Climb. Fix the Protocol on the harvest sheet and re-run.
    if preservation_failures:
        print("")
        print("  " + "!" * 70)
        print(f"  !!  ALARM — {len(preservation_failures)} animal(s) have a Protocol that maps to NO")
        print("  !!  Preservation method. NO SAMPLE was created for them.")
        print("  " + "!" * 70)
        for a_name, proto in preservation_failures:
            print(f"  !!    Animal {a_name}   Protocol={proto!r}")
        print("  !!")
        print("  !!  A sample with a blank Preservation mis-routes its tube label and")
        print("  !!  drops out of the Deliverables trackers. So none was made.")
        print("  !!")
        print("  !!  Recognised protocols:")
        print("  !!    '8 Weeks - ...4%PFA...'  /  'P14 - ...4%PFA...'  -> 4% PFA Fixed")
        print("  !!    'MERFISH - OCT'                                  -> OCT Block")
        print("  !!    'RNA-Seq'                                        -> Flash Frozen")
        print("  !!    'Extra - Sex & Timepoint Full'                   -> no sample (expected)")
        print("  !!")
        print("  !!  Fix the Protocol on the harvest sheet, then re-run this step.")
        print("  " + "!" * 70)
        print("")

    # Save Harvest Sheet Import
    harvest_file = f"Harvest_Sheet_Import_{timestamp}.xlsx"
    save_df_to_excel(harvest_df, harvest_file, sheet_name='Harvest Worksheet')
    print(f"\n  📄 Saved: {harvest_file}")

    # Save Climb Sample Import
    # Backstop: nothing with a blank Preservation Method may reach this file.
    # The guard in the loop above should make this impossible; this catches any
    # future path that bypasses it rather than letting a bad row through.
    if len(climb_import_df) > 0:
        _blank_pres = climb_import_df[
            climb_import_df['Preservation Method'].fillna('').astype(str).str.strip() == ''
        ]
        if len(_blank_pres) > 0:
            print("")
            print("  " + "!" * 70)
            print(f"  !!  ALARM — {len(_blank_pres)} Climb import row(s) have a blank Preservation")
            print("  !!  Method. These rows were REMOVED from the import file.")
            for _, _r in _blank_pres.iterrows():
                print(f"  !!    Sample {_r.get('Sample Name')}  Animal {_r.get('Source AnimalID')}")
            print("  !!  Do not hand-add these to Climb without a Preservation.")
            print("  " + "!" * 70)
            print("")
            climb_import_df = climb_import_df.drop(_blank_pres.index)

    climb_file = (f"TESTONLY_DO_NOT_IMPORT_Climb_Sample_Import_{timestamp}.csv"
                  if TEST_MODE else f"Climb_Sample_Import_{timestamp}.csv")
    climb_import_df.to_csv(climb_file, index=False)
    print(f"  📄 Saved: {climb_file}")

    print(f"\n  ✓ Steps 0+1 complete:")
    print(f"    {len(harvest_df)} animals on harvest worksheet")
    print(f"    {samples_added} samples created")
    print(f"    {len(climb_import_df)} rows in Climb import")

    return harvest_df, samples_for_chain, climb_import_df


# ============================================================
# STEP 2: DELIVERABLES
# ============================================================

class MultiSheetExporter:

    # Full-nomenclature translations for strains whose Line name in the animals
    # export doesn't match the canonical JAX nomenclature.
    # Key   = exact string as it appears in the 'Line' column of the animals CSV
    # Value = correct full nomenclature to use in Line_subject
    LINE_TRANSLATIONS = {
        'B6J-Shank3 (Gfn) use hets':                     'B6.129-Shank3<tm2Gfng>/J',
        'B6J-Cntnap2-/-':                                 'B6.129(Cg)-Cntnap2<tm1Pele>/J',
        'B6J-Fmr1 -/- (X chr)':                          'B6.129P2-Fmr1\u2039tm1Cgr\u203a/J',
        'B6NJ-Bcl11b Cyfip2-S968F\u2039J\u203a H Lethal':           'B6N(Cg)-Cyfip2em2Kumr Bcl11btm1.1(KOMP)Vlcg/Kumr',
        'B6NJ-Kcnd3-/- Cyfip2-S968F\u2039J\u203a Hom Breed Well':   'C57BL/6NJ-Kcnd3em1(IMPC)J Cyfip2em2Kumr/Kumr',
        'B6NJ-Kdm5b Cyfip2-S968F\u2039J\u203a HSubVi':              'C57BL/6N-Cyfip2em2Kumr Kdm5bem1(IMPC)Wtsi/Kumr',
    }

    def _translate_line(self, raw: str) -> str:
        """Return the canonical line name, applying LINE_TRANSLATIONS if needed."""
        return self.LINE_TRANSLATIONS.get(str(raw).strip(), raw)

    def __init__(self, working_df, samples_df, output_filename):
        """Initialize using in-memory DataFrames."""
        self.output_filename = output_filename
        self.workbook = Workbook()

        self.working_df = working_df.copy()
        self.samples_df = samples_df.copy()

        print(f"\n  DEBUG Deliverables init:")
        print(f"    samples_df columns: {list(self.samples_df.columns)}")
        print(f"    samples_df rows: {len(self.samples_df)}")
        print(f"    working_df columns (first 15): {list(self.working_df.columns)[:15]}")
        print(f"    working_df rows: {len(self.working_df)}")

        # Parse dates in working data
        for col in ['Birth_Date', 'Wean Date', 'Harvest_Date', 'Envision_Date']:
            if col in self.working_df.columns:
                self.working_df[col] = pd.to_datetime(
                    self.working_df[col], errors='coerce')

        # Parse dates in samples
        if 'Harvest Date' in self.samples_df.columns:
            self.samples_df['Harvest Date'] = pd.to_datetime(
                self.samples_df['Harvest Date'], errors='coerce')

        # Normalise Harvest Worksheet column names to internal format
        _harvest_col_map = {
            'Name':           'Animal_Name',
            'Sample Number':  'Sample_Name',
            'BD':             'Birth_Date',
            'Harvest Date':   'Harvest Date',   # keep as-is
        }
        for _src, _dst in _harvest_col_map.items():
            if _src in self.samples_df.columns and _dst not in self.samples_df.columns:
                self.samples_df = self.samples_df.rename(columns={_src: _dst})
            if _src in self.working_df.columns and _dst not in self.working_df.columns:
                self.working_df = self.working_df.rename(columns={_src: _dst})

        if 'Animal_Name' in self.samples_df.columns:
            self.samples_df['Animal_Name'] = self.samples_df['Animal_Name'].astype(str).str.strip()
        if 'Animal_Name' in self.working_df.columns:
            self.working_df['Animal_Name'] = self.working_df['Animal_Name'].astype(str).str.strip()

        # Debug merge values
        if 'Animal_Name' in self.samples_df.columns and 'Animal_Name' in self.working_df.columns:
            sample_names = set(self.samples_df['Animal_Name'].unique())
            working_names = set(self.working_df['Animal_Name'].unique())
            common = sample_names.intersection(working_names)
            print(f"    Sample unique Animal_Names: {len(sample_names)}")
            print(f"    Working unique Animal_Names: {len(working_names)}")
            print(f"    Common (will match): {len(common)}")
            if len(common) == 0:
                print(f"    ⚠ NO MATCHES! Sample names: {list(sample_names)[:3]}")
                print(f"    ⚠ Working names: {list(working_names)[:3]}")

        # Merge samples with working data.
        # INNER join — only animals present in animals.csv are exported. The
        # Harvest Worksheet holds the full project history; a left join would
        # carry every historical row through with blank animal details.
        if 'Animal_Name' in self.samples_df.columns and 'Animal_Name' in self.working_df.columns:
            before = len(self.samples_df)
            self.merged_df = pd.merge(
                self.samples_df, self.working_df,
                on='Animal_Name', how='inner',
                suffixes=('_sample', '_animal')
            )
            print(f"    Merged result: {len(self.merged_df)} rows "
                  f"(dropped {before - len(self.merged_df)} not in animals.csv)")
        else:
            self.merged_df = self.samples_df.copy()
            print(f"    No merge possible — using {len(self.merged_df)} sample rows")

        # ── Route each row to a tracker tab by Protocol ──────────────────────
        # Same logic as Sing Sanity:
        #   Protocol contains 'rna-seq' or 'rnaseq'  -> RNA-Seq
        #   Protocol contains 'merfish'              -> MERFISH
        #   anything else                            -> LSFM/MRI
        prot_col = next((c for c in ['Protocol', 'Protocol_sample', 'Protocol_animal',
                                     'Harvest Type', 'Harvest_Type', 'Assigned_Harvest_Type']
                         if c in self.merged_df.columns), None)

        def _route(protocol) -> str:
            pl = str(protocol).strip().lower()
            if 'rna-seq' in pl or 'rnaseq' in pl:
                return 'RNA-Seq'
            if 'merfish' in pl:
                return 'MERFISH'
            return 'LSFM/MRI'

        if prot_col:
            self.merged_df['_tracker'] = self.merged_df[prot_col].apply(_route)
            print(f"    Routed on '{prot_col}': "
                  f"{self.merged_df['_tracker'].value_counts().to_dict()}")
        else:
            self.merged_df['_tracker'] = 'LSFM/MRI'
            print(f"    \u26a0 No Protocol column \u2014 all rows routed to LSFM/MRI")
            print(f"      Looked for: Protocol, Harvest Type, Harvest_Type")
            print(f"      Available: {list(self.merged_df.columns)}")

        if 'Sheet' in self.workbook.sheetnames:
            del self.workbook['Sheet']

    def _calculate_age_weeks(self, birth_date, harvest_date):
        try:
            if pd.notna(birth_date) and pd.notna(harvest_date):
                birth = pd.to_datetime(birth_date)
                harvest = pd.to_datetime(harvest_date)
                return round((harvest - birth).days / 7, 1)
        except Exception:
            pass
        return ''

    @staticmethod
    def _expand_sample_numbers(value) -> list:
        """
        Expand a combined sample-number string into individual numbers.

            '1796'            -> ['1796']
            '1805-1806'       -> ['1805', '1806']
            '1807-1814'       -> ['1807', ..., '1814']
            '1796, 1798-1800' -> ['1796', '1798', '1799', '1800']

        Anything that does not start with a digit is dropped, so the row never
        reaches the tracker:

            'Extra'           -> []
            'QC Fail'         -> []
            'Found Dead'      -> []
            ''                -> []

        Returning an empty list means the caller adds no row at all.
        """
        raw = str(value).strip() if value is not None else ''
        if raw == '' or raw.lower() in ('nan', 'none'):
            return []

        out = []
        for chunk in raw.split(','):
            seg = chunk.strip()
            if not seg or not seg[0].isdigit():
                continue                        # Extra, QC Fail, etc.
            parts = seg.split('-')
            if len(parts) == 2:
                lo, hi = parts[0].strip(), parts[1].strip()
                if lo.isdigit() and hi.isdigit():
                    lo_i, hi_i = int(lo), int(hi)
                    if 0 <= hi_i - lo_i < 1000:      # sanity guard
                        width = len(lo)              # preserve zero padding
                        out.extend(str(n).zfill(width)
                                   for n in range(lo_i, hi_i + 1))
                        continue
            out.append(seg)
        return out

    # RNA-Seq tube suffixes, applied in order to each animal's samples.
    # Eight tubes per animal: seven numbered, then C.
    RNA_SUFFIXES = ['0', '1', '2', '3', '4', '5', '6', 'C']
    RNA_PAD      = 5      # zero-pad sample numbers to this width

    @classmethod
    def _rna_sample_name(cls, sample_no: str, position: int) -> str:
        """
        Format an RNA-Seq sample name: zero-padded number + positional suffix.

            (1608, 0) -> '01608-0'
            (1615, 7) -> '01615-C'

        Positions beyond the suffix list fall back to the index number so
        nothing is silently dropped or duplicated.
        """
        num = str(sample_no).strip()
        if num.isdigit():
            num = num.zfill(cls.RNA_PAD)
        suffix = (cls.RNA_SUFFIXES[position]
                  if position < len(cls.RNA_SUFFIXES) else str(position))
        return f'{num}-{suffix}'

    def _safe_get(self, row, *columns, default=''):
        """Try multiple column names, return first non-null value."""
        for col in columns:
            if col in row.index:
                val = row[col]
                if pd.notna(val) and str(val).lower() != 'nan':
                    return val
        return default

    def create_sing_harvest_sheet(self):
        ws = self.workbook.create_sheet("Sing Harvest Sheet")
        print("\n  Creating Sing Harvest Sheet...")

        grouped_data = {}
        for idx, row in self.merged_df.iterrows():
            animal_name = self._safe_get(row, 'Animal_Name')
            sample_name = self._safe_get(row, 'Sample_Name')
            if not animal_name or animal_name == '':
                continue
            if animal_name not in grouped_data:
                grouped_data[animal_name] = {
                    'samples': [],
                    'data': {
                        'Name': animal_name,
                        'Line': self._safe_get(row, 'Line (Short)', 'Line', 'Strain'),
                        'BD': self._safe_get(row, 'Birth_Date'),
                        'Housing': self._safe_get(row, 'Housing ID'),
                        'Identification': self._safe_get(row, 'Marker_Type', 'Marker'),
                        'Sex': self._safe_get(row, 'Sex', 'Sex_animal', 'Sex_sample'),
                        'Age (Days)': ('P' + str(int(float(v))) if (v := self._safe_get(row, 'Age_Days', 'P56_Age_At_Harvest_Days', 'P14_Age_At_Harvest_Days')) not in (None, '', 'None') else '')
                    }
                }
            if sample_name:
                grouped_data[animal_name]['samples'].append(sample_name)

        harvest_data = []
        for animal_name, group in grouped_data.items():
            row_data = group['data'].copy()
            row_data['Sample Number'] = combine_sample_numbers(group['samples'])
            harvest_data.append(row_data)

        df = pd.DataFrame(harvest_data)
        column_order = ['Name', 'Sample Number', 'Line', 'BD', 'Housing',
                        'Identification', 'Sex', 'Age (Days)']
        if df.empty:
            df = pd.DataFrame(columns=column_order)
        else:
            for col in column_order:
                if col not in df.columns:
                    df[col] = ''
            df = df[column_order]

        self._apply_sheet_styling(ws, df, column_order)
        print(f"    ✓ {len(df)} rows")
        return ws

    def create_animal_sample_tracking_sheet(self):
        ws = self.workbook.create_sheet("Animal and Sample Tracking")
        print("\n  Creating Animal and Sample Tracking sheet...")

        filtered_df = self.merged_df[self.merged_df['_tracker'] == 'LSFM/MRI'].copy()
        print(f"    {len(filtered_df)} LSFM/MRI rows")

        tracking_data = []
        for idx, row in filtered_df.iterrows():
            age_weeks = self._calculate_age_weeks(
                self._safe_get(row, 'Birth_Date'),
                self._safe_get(row, 'Harvest Date', 'Harvest_Date')
            )
            timepoint = self._safe_get(row, 'Assigned_Timepoint', 'Harvest Timepoint',
                                       'Age (Days)', 'Age (Days)_sample') or ''
            harvest_date = self._safe_get(row, 'Harvest Date', 'Harvest_Date')
            wean_date = harvest_date if str(timepoint).strip() == 'P14' else self._safe_get(row, 'Wean Date')
            for _sample_no in self._expand_sample_numbers(
                    self._safe_get(row, 'Sample_Name')):
                tracking_data.append({
                    'Name_sample': _sample_no,
                    'Harvest Date': harvest_date,
                    'Age (weeks)_sample': age_weeks,
                    'Name_subject': self._safe_get(row, 'Animal_Name'),
                    'Sex': self._safe_get(row, 'Sex', 'Sex_animal', 'Sex_sample'),
                    'Line_subject': self._translate_line(self._safe_get(row, 'Line_animal', 'Line', 'Strain')),
                    'Line (Short)': self._safe_get(row, 'Line (Short)'),
                    'Line (Stock)': self._safe_get(row, 'Line (Stock)'),
                    'Species_subject': 'Mouse',
                    'Genotype': self._safe_get(row, 'Raw_Genotype', 'Raw_Genotype_animal', 'Genotype', 'Genotype_animal', 'Genotype_sample'),
                    'Birth Date': self._safe_get(row, 'Birth_Date'),
                    'Wean Date': wean_date,
                    'Harvest Timepoint': timepoint
                })

        column_order = [
            'Name_sample', 'Harvest Date', 'Age (weeks)_sample', 'Name_subject',
            'Sex', 'Line_subject', 'Line (Short)', 'Line (Stock)',
            'Species_subject', 'Genotype', 'Birth Date', 'Wean Date',
            'Harvest Timepoint'
        ]
        df = pd.DataFrame(tracking_data)
        if df.empty:
            df = pd.DataFrame(columns=column_order)
        else:
            df = df[column_order]

        self._apply_sheet_styling(ws, df, column_order)
        print(f"    ✓ {len(df)} rows (4% PFA Fixed)")
        return ws

    def create_merfish_sample_tracker_sheet(self):
        ws = self.workbook.create_sheet("MERFISH Sample Tracker")
        print("\n  Creating MERFISH Sample Tracker sheet...")

        filtered_df = self.merged_df[self.merged_df['_tracker'] == 'MERFISH'].copy()
        print(f"    {len(filtered_df)} MERFISH rows")

        tracker_data = []
        for idx, row in filtered_df.iterrows():
            age_weeks = self._calculate_age_weeks(
                self._safe_get(row, 'Birth_Date'),
                self._safe_get(row, 'Harvest Date', 'Harvest_Date')
            )
            timepoint = self._safe_get(row, 'Assigned_Timepoint', 'Harvest Timepoint',
                                       'Age (Days)', 'Age (Days)_sample') or ''
            harvest_date = self._safe_get(row, 'Harvest Date', 'Harvest_Date')
            wean_date = harvest_date if str(timepoint).strip() == 'P14' else self._safe_get(row, 'Wean Date')
            for _sample_no in self._expand_sample_numbers(
                    self._safe_get(row, 'Sample_Name')):
                tracker_data.append({
                    'Name_sample': _sample_no,
                    'Age (weeks)_sample': age_weeks,
                    'Name_subject': self._safe_get(row, 'Animal_Name'),
                    'Sex': self._safe_get(row, 'Sex', 'Sex_animal', 'Sex_sample'),
                    'Line_subject': self._translate_line(self._safe_get(row, 'Line_animal', 'Line', 'Strain')),
                    'Line (Short)': self._safe_get(row, 'Line (Short)'),
                    'Line (Stock)': self._safe_get(row, 'Line (Stock)'),
                    'Species_subject': 'Mouse',
                    'Genotype': self._safe_get(row, 'Raw_Genotype', 'Raw_Genotype_animal', 'Genotype', 'Genotype_animal', 'Genotype_sample'),
                    'Birth Date': self._safe_get(row, 'Birth_Date'),
                    'Wean Date': wean_date,
                    'Dissect Date': harvest_date
                })

        column_order = [
            'Name_sample', 'Line (Short)', 'Age (weeks)_sample', 'Sex',
            'Name_subject', 'Line_subject', 'Line (Stock)', 'Species_subject',
            'Genotype', 'Birth Date', 'Wean Date', 'Dissect Date'
        ]
        df = pd.DataFrame(tracker_data)
        if df.empty:
            df = pd.DataFrame(columns=column_order)
        else:
            df = df[column_order]

        self._apply_sheet_styling(ws, df, column_order)
        print(f"    ✓ {len(df)} rows (OCT Block)")
        return ws

    def create_rnaseq_sample_tracker_sheet(self):
        ws = self.workbook.create_sheet("RNASeq Sample Tracker")
        print("\n  Creating RNASeq Sample Tracker sheet...")

        filtered_df = self.merged_df[self.merged_df['_tracker'] == 'RNA-Seq'].copy()
        print(f"    {len(filtered_df)} RNA-Seq rows")

        tracker_data = []
        for idx, row in filtered_df.iterrows():
            age_weeks = self._calculate_age_weeks(
                self._safe_get(row, 'Birth_Date'),
                self._safe_get(row, 'Harvest Date', 'Harvest_Date')
            )
            timepoint = self._safe_get(row, 'Assigned_Timepoint', 'Harvest Timepoint',
                                       'Age (Days)', 'Age (Days)_sample') or ''
            harvest_date = self._safe_get(row, 'Harvest Date', 'Harvest_Date')
            wean_date = harvest_date if str(timepoint).strip() == 'P14' else self._safe_get(row, 'Wean Date')
            for _pos, _sample_no in enumerate(self._expand_sample_numbers(
                    self._safe_get(row, 'Sample_Name'))):
                tracker_data.append({
                    'Name_sample': self._rna_sample_name(_sample_no, _pos),
                    'Age (weeks)_sample': age_weeks,
                    'Name_subject': self._safe_get(row, 'Animal_Name'),
                    'Sex': self._safe_get(row, 'Sex', 'Sex_animal', 'Sex_sample'),
                    'Line_subject': self._translate_line(self._safe_get(row, 'Line_animal', 'Line', 'Strain')),
                    'Line (Short)': self._safe_get(row, 'Line (Short)'),
                    'Line (Stock)': self._safe_get(row, 'Line (Stock)'),
                    'Species_subject': 'Mouse',
                    'Genotype': self._safe_get(row, 'Raw_Genotype', 'Raw_Genotype_animal', 'Genotype', 'Genotype_animal', 'Genotype_sample'),
                    'Birth Date': self._safe_get(row, 'Birth_Date'),
                    'Wean Date': wean_date,
                    'Dissect Date': harvest_date
                })

        column_order = [
            'Name_sample', 'Age (weeks)_sample', 'Name_subject', 'Sex',
            'Line_subject', 'Line (Short)', 'Line (Stock)', 'Species_subject',
            'Genotype', 'Birth Date', 'Wean Date', 'Dissect Date'
        ]
        df = pd.DataFrame(tracker_data)
        if df.empty:
            df = pd.DataFrame(columns=column_order)
        else:
            df = df[column_order]

        self._apply_sheet_styling(ws, df, column_order)
        print(f"    ✓ {len(df)} rows (Flash Frozen)")
        return ws

    def _apply_sheet_styling(self, ws, df, column_order):
        """Apply consistent styling to a worksheet."""
        for col_num, header in enumerate(column_order, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = Font(bold=True, size=12, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092",
                                    fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin'))

        for row_num, row_data in enumerate(df.values, 2):
            for col_num, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col_num)
                if pd.isna(value):
                    cell.value = ''
                else:
                    cell.value = value
                cell.alignment = Alignment(horizontal="left", vertical="center")
                cell.border = Border(
                    left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'), bottom=Side(style='thin'))
                if 'Date' in column_order[col_num - 1] and cell.value:
                    try:
                        if pd.notna(cell.value) and cell.value != '':
                            cell.number_format = 'MM/DD/YYYY'
                    except Exception:
                        pass

        auto_width_worksheet(ws)

    def create_all_sheets(self):
        # Sing Harvest Sheet tab removed — use Harvest_Sheet_Import file instead
        self.create_animal_sample_tracking_sheet()
        self.create_merfish_sample_tracker_sheet()
        self.create_rnaseq_sample_tracker_sheet()

    def save(self):
        self.workbook.save(self.output_filename)
        print(f"\n  📄 Saved: {self.output_filename}")
        return self.output_filename


def run_deliverables(working_df, samples_df, timestamp, output_dir=None):
    """STEP 2: Create multi-sheet deliverables Excel file."""
    print("\n" + "=" * 80)
    print("STEP 2: DELIVERABLES")
    print("=" * 80)

    if samples_df is None or samples_df.empty:
        print("  \u2717 No sample data. Skipping.")
        return None

    _out_dir = output_dir or _SCRIPT_DIR

    # Filename carries every harvest date present in the export:
    #   one date    -> Lab_Data_Export_2026_09_09.xlsx
    #   a few dates -> Lab_Data_Export_2026_09_02_2026_09_09.xlsx
    #   many dates  -> Lab_Data_Export_2026_09_02_to_2026_10_07.xlsx
    _dates = []
    for _col in ['Harvest Date', 'Harvest_Date', 'P56_Harvest_Date']:
        if _col in samples_df.columns:
            _parsed = pd.to_datetime(samples_df[_col], errors='coerce').dropna()
            if not _parsed.empty:
                _dates = sorted({d.strftime('%Y_%m_%d') for d in _parsed})
                break

    if not _dates:
        _date_str = ''
    elif len(_dates) <= 4:
        _date_str = '_'.join(_dates)
    else:
        _date_str = f'{_dates[0]}_to_{_dates[-1]}'

    if _dates:
        print(f"  {len(_dates)} harvest date(s): {', '.join(_dates)}")

    if _date_str:
        output_filename = _make_dated_path(_out_dir, 'Lab_Data_Export', _date_str)
    else:
        output_filename = os.path.join(_out_dir, f'Lab_Data_Export_{timestamp}.xlsx')

    try:
        exporter = MultiSheetExporter(
            working_df=working_df,
            samples_df=samples_df,
            output_filename=output_filename
        )
        exporter.create_all_sheets()
        saved_file = exporter.save()
        print(f"\n  \u2713 Step 2 complete: 4 sheets created")
        return saved_file
    except Exception as e:
        print(f"  \u2717 Error: {e}")
        traceback.print_exc()
        return None


# ============================================================
# STEP 3: CLIMB TO ENVISION
# ============================================================

def assign_ear_tags_by_strain_sex(df):
    """Assign S4, S3, S2 in repeating pattern after sorting."""
    df_sorted = df.copy()
    sort_data = []
    for idx, row in df_sorted.iterrows():
        sort_data.append((
            idx, row['Line'], row['Sex'],
            natural_sort_key(row['Animal_Name'])
        ))
    sort_data.sort(key=lambda x: (x[1], x[2], x[3]))
    sorted_indices = [item[0] for item in sort_data]
    df_sorted = df_sorted.loc[sorted_indices].reset_index(drop=True)

    tags = []
    current_strain = None
    current_sex = None
    counter = 0

    for idx, row in df_sorted.iterrows():
        strain = row['Line']
        sex = row['Sex']
        if strain != current_strain or sex != current_sex:
            current_strain = strain
            current_sex = sex
            counter = 0
        position = (counter % 3) + 1
        tags.append('S4' if position == 1 else 'S3' if position == 2 else 'S2')
        counter += 1

    df_sorted['Envision Ear Tag'] = tags
    return df_sorted


def group_animals_by_housing(df):
    """Group animals with same Group ID, numbering groups of 3."""
    group_housing_counts = defaultdict(lambda: defaultdict(list))
    for idx, row in df.iterrows():
        group_housing_counts[row['Group_base']][row['Housing ID']].append(idx)

    group_suffixes = {}
    for group_name, housing_dict in group_housing_counts.items():
        total = sum(len(v) for v in housing_dict.values())
        if total <= 3:
            for indices in housing_dict.values():
                for idx in indices:
                    group_suffixes[idx] = group_name
        else:
            assigned = 0
            for housing_id, indices in sorted(housing_dict.items()):
                for idx in indices:
                    group_suffixes[idx] = f"{group_name}{(assigned // 3) + 1}"
                    assigned += 1
    return group_suffixes


def _make_dated_path(folder: str, stem: str, date_str: str, ext: str = '.xlsx') -> str:
    """Return a collision-safe path: stem_date.xlsx → stem_date (1).xlsx → ..."""
    base = f"{stem}_{date_str}{ext}"
    path = os.path.join(folder, base)
    if not os.path.exists(path):
        return path
    n = 1
    while True:
        path = os.path.join(folder, f"{stem}_{date_str} ({n}){ext}")
        if not os.path.exists(path):
            return path
        n += 1


def _is_p14(timepoint) -> bool:
    """True if a timepoint value represents P14."""
    t = str(timepoint).strip().upper()
    return t in ('P14', '14') or t.startswith('P14')


def _file_date_for_row(row) -> str:
    """
    Pick the date that names the output file for one animal.

        P14   -> harvest date
        Adult -> behavior date (the Envision Date column)

    Falls back to harvest date if no behavior date is recorded, so a file is
    always named something meaningful rather than 'unknown_date'.
    """
    def _first(*names):
        for n in names:
            try:
                v = row.get(n)
            except AttributeError:
                v = row[n] if n in row else None
            if v is not None and str(v).strip() not in ('', 'nan', 'NaT', 'None'):
                return v
        return None

    timepoint = _first('Age (Days)', 'Assigned_Timepoint', 'Harvest Timepoint',
                       'Timepoint')
    harvest   = _first('Harvest Date', 'Sample Harvest Date', 'Harvest_Date',
                       'P14_Date', 'P56_Harvest_Date')
    behavior  = _first('Envision Date', 'P56_Behavior_Date', 'Behavior Date',
                       'Envision_Date')

    if _is_p14(timepoint):
        return harvest or behavior or ''
    return behavior or harvest or ''


def _harvest_date_str(val) -> str:
    """Convert any date-like value to YYYY_MM_DD string for filenames."""
    if val is None or str(val).strip() in ('', 'None', 'NaT', 'nan'):
        return ''
    try:
        return pd.to_datetime(val).strftime('%Y_%m_%d')
    except Exception:
        return str(val).strip().replace('-', '_').replace('/', '_')[:10]


def _norm_assay(name: str) -> str:
    """Strip < > and spaces so TGS and Climb assay names compare equal.
    'Shank3tm2Gfng Probe' == 'Shank3<tm2Gfng> Probe'"""
    return (str(name).replace('<', '').replace('>', '')
            .replace(' ', '').strip().lower())


# TGS call column -> Climb genotype symbol
_TGS_CALL_COLUMNS = {
    'wild':         '+/+',
    'het':          '-/+',
    'hom':          '-/-',
    'inconclusive': 'Inconclusive',
}


def parse_tgs_report(path):
    """
    Read the genotype calls out of one TGS typing report.

    TGS reports are HTML tables saved with an .xls extension, so they're read
    with pd.read_html rather than an Excel engine.

    Animals are matched on 'Pedigree #', NOT 'Mouse Id' — Mouse Id is the ear
    notch (R, 2R1L) which repeats across litters.

    Returns (records, meta).
    """
    tables = pd.read_html(path)

    meta = {'file': os.path.basename(path), 'strain': '',
            'sampled': '', 'completed': '', 'assays': []}

    try:
        info = tables[1]
        for _, r in info.iterrows():
            label = str(r[0]).strip().rstrip(':').lower()
            if label == 'strain':
                meta['strain'] = str(r[1]).strip()
            elif label == 'sampler' and len(r) > 3:
                meta['sampled'] = str(r[3]).strip()
    except Exception:
        pass

    try:
        log  = tables[9]
        done = log[log[0].astype(str).str.strip() == 'Completed']
        if not done.empty:
            meta['completed'] = str(done.iloc[-1][1]).strip()
    except Exception:
        pass

    t          = tables[5]
    assay_row  = t.iloc[0]
    header_row = t.iloc[1]
    body       = t.iloc[2:]

    headers = {}
    for i, h in enumerate(header_row):
        key = str(h).strip().lower()
        if key and key != 'nan':
            headers.setdefault(key, []).append(i)

    def col(name):
        idx = headers.get(name)
        return idx[0] if idx else None

    c_animal = col('pedigree #')
    c_notch  = col('mouse id')
    if c_animal is None:
        raise ValueError(f"{meta['file']}: no 'Pedigree #' column")

    call_cols = []
    for label, symbol in _TGS_CALL_COLUMNS.items():
        for i in headers.get(label, []):
            assay = str(assay_row.iloc[i]).strip()
            if assay and assay.lower() != 'nan':
                call_cols.append((i, symbol, assay))
    if not call_cols:
        raise ValueError(f"{meta['file']}: no call columns found")

    meta['assays'] = sorted({a for _, _, a in call_cols})

    records = []
    for _, row in body.iterrows():
        animal = str(row.iloc[c_animal]).strip()
        if not animal or animal.lower() == 'nan':
            continue
        notch = str(row.iloc[c_notch]).strip() if c_notch is not None else ''
        for i, symbol, assay in call_cols:
            v = row.iloc[i]
            if pd.notna(v) and str(v).strip() != '':
                records.append({'animal': animal, 'assay': assay,
                                'symbol': symbol, 'notch': notch})
    return records, meta


def run_tgs_genotypes(script_dir):
    """
    Upload TGS genotype calls to Climb before scheduling.

    Looks for TGS_Typing_*.xls in a 'genotypes' subfolder (or the script
    folder), and posts any call the animal does not already have. Animals that
    already carry that assay and result are skipped — Climb does not
    deduplicate genotypes.

    Runs before the animals pull so the fresh CSV reflects the new calls, which
    is what stops those animals scheduling as Blank.

    Never raises — a failure here must not stop the run.
    """
    import requests, time, glob as _glob, json as _json_tgs

    try:
        folders = [os.path.join(script_dir, 'genotypes'), script_dir]
        found = []
        for f in folders:
            if os.path.isdir(f):
                found += sorted(_glob.glob(os.path.join(f, 'TGS_Typing_*.xls')))

        if not found:
            return

        # Skip reports already uploaded. Every successful run writes a receipt
        # (tgs_genotype_upload_*.json) listing the reports it processed, so a
        # report named in one of those has already been done.
        #
        # This replaces the old age-based filter, which skipped anything over
        # TGS_MAX_AGE_DAYS whether or not it had actually uploaded — so an old
        # report that never went through was silently ignored forever.
        #
        # A report is only treated as done if its receipt recorded no failures
        # for it. A partial upload gets retried rather than being written off.
        done_files = set()
        for f in folders:
            if not os.path.isdir(f):
                continue
            for receipt in sorted(_glob.glob(os.path.join(
                    f, 'tgs_genotype_upload_*.json'))):
                try:
                    with open(receipt, 'r', encoding='utf-8') as fh:
                        data = _json_tgs.load(fh)
                except Exception:
                    continue                      # unreadable receipt — ignore
                bad = set()
                for key in ('failed', 'problems', 'parse_errors'):
                    for item in (data.get(key) or []):
                        if isinstance(item, dict) and item.get('file'):
                            bad.add(str(item['file']).strip())
                for rep in (data.get('reports') or []):
                    name = str(rep.get('file', '')).strip()
                    if name and name not in bad:
                        done_files.add(name)

        reports, already = [], []
        for p in found:
            (already if os.path.basename(p) in done_files else reports).append(p)

        print('\n' + '=' * 80)
        print('TGS GENOTYPES')
        print('=' * 80)
        print(f'  {len(found)} report(s) found')
        if already:
            print(f'  {len(already)} already uploaded (receipt on file) \u2014 skipped')
            for a in already:
                print(f'    {os.path.basename(a)}')

        if not reports:
            print('  Nothing new to upload.')
            return
        print(f'  {len(reports)} to process')

        try:
            import lxml  # noqa: F401
        except ImportError:
            print('  \u26a0 Cannot read TGS reports \u2014 lxml is not installed.')
            print('    Run:  pip install lxml')
            return

        records, metas = [], []
        for p in reports:
            try:
                recs, meta = parse_tgs_report(p)
                try:
                    d = pd.to_datetime(meta['completed']).strftime('%Y%m%d')
                except Exception:
                    d = datetime.now().strftime('%Y%m%d')
                for r in recs:
                    r['date'] = d
                records += recs
                metas.append(meta)
                print(f"    {meta['file']}: {len(recs)} call(s), "
                      f"{', '.join(meta['assays'])}")
            except Exception as e:
                print(f'    {os.path.basename(p)}: could not read \u2014 {e}')

        if not records:
            print('  No calls found.')
            return

        sc = _load_sing_climb()

        def _hdr():
            return {'Authorization':   f'Bearer {sc._get_token()}',
                    'X-Workgroup-Key': sc._WORKGROUP_KEY,
                    'Content-Type':    'application/json'}

        def _vocab(endpoint):
            out, page = {}, 1
            while True:
                time.sleep(0.12)
                r = requests.get(f'{sc._API_BASE}{endpoint}', headers=_hdr(),
                                 params={'pageNumber': page, 'pageSize': 100},
                                 timeout=30)
                r.raise_for_status()
                body = r.json().get('data', {})
                for item in body.get('items', []):
                    out[item['name']] = int(item['key'])
                if page >= body.get('pageCount', 1):
                    break
                page += 1
            return out

        animals = sc._get_all('/api/animals')
        by_name = {str(a.get('animalName', '')).strip(): a for a in animals
                   if str(a.get('animalName', '')).strip()}

        assay_keys  = _vocab('/api/vocabulary/genotypeAssay')
        symbol_keys = _vocab('/api/vocabulary/genotypeSymbol')
        assay_lookup = {_norm_assay(k): (k, v) for k, v in assay_keys.items()}

        # Animals that already have this assay — skip regardless of the call,
        # so a re-genotype never silently adds a second conflicting record.
        have_assay = set()
        for g in sc._get_all('/api/genotypes'):
            nm = str(g.get('animalName', '')).strip()
            if nm:
                have_assay.add((nm, _norm_assay(g.get('assay', ''))))

        jobs, skipped, problems = [], [], []
        for r in records:
            climb = by_name.get(r['animal'])
            if not climb:
                problems.append(f"{r['animal']}: not in Climb")
                continue
            hit = assay_lookup.get(_norm_assay(r['assay']))
            if not hit:
                problems.append(f"{r['animal']}: assay '{r['assay']}' not in Climb")
                continue
            climb_assay, assay_key = hit
            if r['symbol'] not in symbol_keys:
                problems.append(f"{r['animal']}: symbol '{r['symbol']}' not in Climb")
                continue
            if (r['animal'], _norm_assay(climb_assay)) in have_assay:
                skipped.append(r['animal'])
                continue
            jobs.append({
                'animal':    r['animal'],
                'animal_id': climb.get('animalId') or climb.get('animalID'),
                'assay':     climb_assay,
                'assay_key': assay_key,
                'symbol':    r['symbol'],
                'symbol_key': symbol_keys[r['symbol']],
                'date':      r['date'],
            })

        print(f'\n  To upload        : {len(jobs)}')
        print(f'  Already genotyped: {len(skipped)}  (skipped)')
        if problems:
            print(f'  Problems         : {len(problems)}')
            for p in problems[:10]:
                print(f'    {p}')
            if len(problems) > 10:
                print(f'    ... and {len(problems) - 10} more')

        if not jobs:
            print('  Nothing to upload.')
            return

        ok, failed = 0, []
        for j in jobs:
            payload = {'genotypeRequestDtos': [{
                'animalID': j['animal_id'],
                'genotypes': [{
                    'date':              j['date'],
                    'genotypeAssayKey':  j['assay_key'],
                    'genotypeSymbolKey': j['symbol_key'],
                }]
            }]}
            try:
                time.sleep(0.12)
                r = _api_post(f'{sc._API_BASE}/api/genotypes',
                                  headers=_hdr(), json=payload, timeout=30)
                if not r.ok:
                    raise requests.HTTPError(f'{r.status_code} {r.text[:200]}')
                ok += 1
                print(f"    {j['animal']}  {j['assay']}  {j['symbol']}")
            except Exception as e:
                failed.append((j['animal'], str(e)))
                print(f"    {j['animal']}: FAILED \u2014 {e}")

        print(f'\n  \u2713 {ok} genotype(s) uploaded')
        if failed:
            print(f'  \u26a0 {len(failed)} failed \u2014 re-run to retry')

    except Exception as e:
        print(f'  \u26a0 TGS genotype upload failed: {e}')
        print('    Continuing with the run.')


def _push_rapid_markers_to_climb(output_df, source_df, marker_type='RapID',
                                 separator=', '):
    """
    Write Envision RapID tags into Climb.

        Marker      ->  "<original marker>, <RapID tag>"    e.g. "R, S4"
        Marker Type ->  RapID

    Only runs on a COMPLETE list — if any animal in the export is missing a
    tag, nothing is pushed. A partial export must not leave Climb half-updated.

    Climb's PUT /api/animals/{id} replaces the whole record, so each animal is
    read first and every other field carried back unchanged.

    Never raises: a failure here must not lose the Envision export that was
    already written.
    """
    import requests, time

    try:
        # ── Completeness gate ────────────────────────────────────────────────
        tags = output_df['Envision Ear Tag'].astype(str).str.strip()
        missing = int((tags == '').sum())
        if missing:
            print(f"  \u26a0 Climb marker push skipped \u2014 {missing} of "
                  f"{len(output_df)} animals have no tag.")
            print("    Climb is only updated from a complete list.")
            return
        if output_df.empty:
            print("  \u26a0 Climb marker push skipped \u2014 nothing in the export.")
            return

        sc = _load_sing_climb()

        def _hdr():
            return {'Authorization':   f'Bearer {sc._get_token()}',
                    'X-Workgroup-Key': sc._WORKGROUP_KEY,
                    'Content-Type':    'application/json'}

        # ── Look up animals and the marker-type key ──────────────────────────
        animals   = sc._get_all('/api/animals')
        by_name   = {}
        type_keys = {}
        for a in animals:
            n = str(a.get('animalName', '')).strip()
            if n:
                by_name[n] = a
            if a.get('physicalMarkerTypeKey') and a.get('markerType'):
                type_keys[a['markerType']] = a['physicalMarkerTypeKey']

        if marker_type not in type_keys:
            print(f"  \u26a0 Climb marker push skipped \u2014 marker type "
                  f"'{marker_type}' not found in Climb.")
            return
        type_key = type_keys[marker_type]

        # Original markers from the source data (animals.csv), keyed by name
        orig_markers = {}
        if 'Marker' in source_df.columns and 'Animal_Name' in source_df.columns:
            for _, r in source_df.iterrows():
                orig_markers[str(r['Animal_Name']).strip()] = str(r.get('Marker') or '').strip()

        PRESERVE = [
            'alternatePhysicalID', 'heldFor', 'citesNumber', 'lineKey', 'sexKey',
            'generationKey', 'breedingStatusKey', 'dietKey', 'animalStatusKey',
            'exitReasonKey', 'animalName', 'dateBorn', 'dateExit', 'comments',
            'commentStatus', 'owner', 'arrivalDate', 'animalUseKey',
            'iacucprotocolKey', 'materialOriginKey', 'externalIdentifier',
            'microchipIdentifier',
        ]
        # Dates are passed straight back as Climb returned them. Reformatting
        # to date-only sets the stored time to 00:00 UTC, which renders as the
        # PREVIOUS day in Eastern. Never touch a date you are only carrying
        # through.

        # ── Build the work list — must resolve fully before writing ──────────
        jobs, unknown, already = [], [], []
        for _, row in output_df.iterrows():
            name = str(row['Animal ID']).strip()
            tag  = str(row['Envision Ear Tag']).strip()

            climb = by_name.get(name)
            if not climb:
                unknown.append(name)
                continue

            current = str(climb.get('physicalMarker') or '').strip()
            if current.endswith(tag):
                already.append(name)
                continue

            orig = orig_markers.get(name)
            if orig is None:
                orig = current          # fall back to what Climb already holds
            new_marker = tag if orig == '' else f'{orig}{separator}{tag}'

            jobs.append((name, climb, new_marker, current))

        if unknown:
            print(f"  \u26a0 Climb marker push skipped \u2014 {len(unknown)} animal(s) "
                  f"not found in Climb: {unknown[:5]}")
            print("    Climb is only updated from a complete list.")
            return

        if not jobs:
            print(f"  Climb markers already up to date ({len(already)} animals).")
            return

        # ── Write ────────────────────────────────────────────────────────────
        print(f"  Updating {len(jobs)} markers in Climb "
              f"({len(already)} already tagged)...")
        ok, failed = 0, []
        for name, climb, new_marker, was in jobs:
            payload = {}
            for f in PRESERVE:
                v = climb.get(f)
                if v is not None:
                    payload[f] = v
            payload['physicalMarker']        = new_marker
            payload['physicalMarkerTypeKey'] = type_key
            # Climb rejects the PUT without these four arrays. cohortKeys must
            # carry the real cohorts — [] would remove the animal from them.
            payload['cohortKeys'] = [c.get('cohortKey')
                                     for c in (climb.get('cohorts') or [])
                                     if c.get('cohortKey') is not None]
            payload['jobKeys']               = []
            payload['housings']              = []
            payload['animalCharacteristics'] = []

            animal_id = climb.get('animalId') or climb.get('animalID')
            try:
                time.sleep(0.12)
                r = _api_put(f'{sc._API_BASE}/api/animals/{animal_id}',
                                 headers=_hdr(), json=payload, timeout=30)
                if not r.ok:
                    raise requests.HTTPError(f'{r.status_code} {r.text[:200]}')
                ok += 1
                print(f'    {name}: {was or "(blank)"} \u2192 {new_marker}')
            except Exception as e:
                failed.append((name, str(e)))
                print(f'    {name}: FAILED \u2014 {e}')

        print(f"  \u2713 Climb markers updated: {ok} of {len(jobs)}")
        if failed:
            print(f"  \u26a0 {len(failed)} failed \u2014 re-run to retry "
                  f"(animals already updated are skipped).")

    except Exception as e:
        print(f'  \u26a0 Climb marker push failed: {e}')
        print('    The Envision export was still written.')


def release_unusable_to_available(animals_df, use_available='Available'):
    """
    Move Wild and Inconclusive animals out of the Sing pool.

    Neither can be scheduled, so they sit in Sing Inventory indefinitely.
    Setting Use to 'Available' frees them for other projects.

    Excluded lines (CONFIG['RELEASE_EXCLUDE_LINES']) are left alone — those
    colonies keep their wild-types as controls or breeders.

    Never raises.
    """
    import requests, time

    print('\n' + '=' * 80)
    print('RELEASE UNUSABLE ANIMALS')
    print('=' * 80)

    try:
        exclude  = [s.strip().lower()
                    for s in CONFIG.get('RELEASE_EXCLUDE_LINES', [])]
        statuses = {s.strip().lower()
                    for s in CONFIG.get('RELEASE_GENOTYPES',
                                        [GENOTYPE_WILD, GENOTYPE_INCONC])}

        line_col = next((c for c in ('Line (Short)', 'Strain', 'Line_Short')
                         if c in animals_df.columns), None)
        geno_col = next((c for c in ('Genotype', 'Genotype_clean')
                         if c in animals_df.columns), None)
        name_col = next((c for c in ('Animal_Name', 'Name')
                         if c in animals_df.columns), None)

        if not (geno_col and name_col):
            print('  \u2717 Need Genotype and Name columns \u2014 skipping.')
            return
        if not line_col:
            print('  \u2717 No line column \u2014 cannot honour the exclusions, skipping.')
            return

        print(f"  Releasing   : {', '.join(sorted(statuses))}")
        print(f"  Excluding   : {', '.join(CONFIG.get('RELEASE_EXCLUDE_LINES', []))}")

        # -- Pick candidates --------------------------------------------------
        candidates, excluded_counts = [], {}
        for _, row in animals_df.iterrows():
            geno = canonicalize_genotype(row.get(geno_col))
            if geno.lower() not in statuses:
                continue

            line = str(row.get(line_col) or '').strip()
            if any(x and x in line.lower() for x in exclude):
                excluded_counts[line] = excluded_counts.get(line, 0) + 1
                continue

            name = str(row.get(name_col) or '').strip()
            if name:
                candidates.append((name, line, geno))

        if excluded_counts:
            print('\n  Left alone (excluded lines):')
            for line, n in sorted(excluded_counts.items()):
                print(f'    {line:34} {n}')

        if not candidates:
            print('\n  Nothing to release.')
            return

        # -- Climb ------------------------------------------------------------
        sc = _load_sing_climb()

        def _hdr():
            return {'Authorization':   f'Bearer {sc._get_token()}',
                    'X-Workgroup-Key': sc._WORKGROUP_KEY,
                    'Content-Type':    'application/json'}

        animals  = sc._get_all('/api/animals')
        by_name  = {}
        use_keys = {}
        for a in animals:
            n = str(a.get('animalName', '')).strip()
            if n:
                by_name[n] = a
            if a.get('animalUseKey') and a.get('use'):
                use_keys[a['use']] = a['animalUseKey']

        if use_available not in use_keys:
            print(f"\n  \u2717 Use '{use_available}' not found in Climb.")
            print(f'    Available: {sorted(use_keys)}')
            return
        target_key = use_keys[use_available]

        PRESERVE = [
            'alternatePhysicalID', 'heldFor', 'citesNumber', 'lineKey', 'sexKey',
            'generationKey', 'breedingStatusKey', 'dietKey', 'animalStatusKey',
            'exitReasonKey', 'animalName', 'physicalMarker', 'dateBorn',
            'dateExit', 'comments', 'commentStatus', 'owner', 'arrivalDate',
            'iacucprotocolKey', 'physicalMarkerTypeKey', 'materialOriginKey',
            'externalIdentifier', 'microchipIdentifier',
        ]

        jobs, already, missing = [], [], []
        by_geno = {}
        for name, line, geno in candidates:
            climb = by_name.get(name)
            if not climb:
                missing.append(name)
                continue
            if climb.get('animalUseKey') == target_key:
                already.append(name)
                continue
            jobs.append({'name': name, 'line': line, 'geno': geno,
                         'climb': climb, 'was': climb.get('use', '')})
            by_geno[geno] = by_geno.get(geno, 0) + 1

        print(f'\n  To release   : {len(jobs)}')
        for g, n in sorted(by_geno.items()):
            print(f'    {g:16} {n}')
        print(f'  Already set  : {len(already)}')
        if missing:
            print(f'  Not in Climb : {len(missing)}')

        if not jobs:
            print('  Nothing to change.')
            return

        ok, failed = 0, []
        for j in jobs:
            climb   = j['climb']
            payload = {f: climb[f] for f in PRESERVE if climb.get(f) is not None}
            payload['animalUseKey'] = target_key
            payload['cohortKeys']   = [c.get('cohortKey')
                                       for c in (climb.get('cohorts') or [])
                                       if c.get('cohortKey') is not None]
            payload['jobKeys']               = []
            payload['housings']              = []
            payload['animalCharacteristics'] = []

            animal_id = climb.get('animalId') or climb.get('animalID')
            try:
                time.sleep(0.12)
                r = _api_put(f'{sc._API_BASE}/api/animals/{animal_id}',
                                 headers=_hdr(), json=payload, timeout=30)
                if not r.ok:
                    raise requests.HTTPError(f'{r.status_code} {r.text[:200]}')
                ok += 1
                print(f"    {j['name']:10} {j['line'][:18]:18} {j['geno']:14} "
                      f"{j['was'] or '(blank)'} \u2192 {use_available}")
            except Exception as e:
                failed.append((j['name'], str(e)))
                print(f"    {j['name']}: FAILED \u2014 {e}")

        print(f'\n  \u2713 Released {ok} of {len(jobs)} animals')
        if failed:
            print(f'  \u26a0 {len(failed)} failed \u2014 re-run to retry')

    except Exception as e:
        print(f'  \u26a0 Release failed: {e}')
        print('    The schedule is unaffected.')


def update_animal_use(working_df, use_p14='Sing - P14', use_p56='Sing - P56'):
    """
    Set the Use field on scheduled animals according to their timepoint.

        P14   -> 'Sing - P14'
        Adult -> 'Sing - P56'

    Runs after harvest assignments are confirmed. Animals already carrying the
    right Use are skipped, so it is safe to re-run.

    Uses the same read-modify-write as every other write path: PUT replaces the
    whole animal record, so everything else is carried back untouched and the
    four array fields are required.

    Never raises — a failure here must not lose the schedule.
    """
    import requests, time

    print('\n' + '=' * 80)
    print('ANIMAL USE UPDATE')
    print('=' * 80)

    try:
        sc = _load_sing_climb()

        def _hdr():
            return {'Authorization':   f'Bearer {sc._get_token()}',
                    'X-Workgroup-Key': sc._WORKGROUP_KEY,
                    'Content-Type':    'application/json'}

        animals   = sc._get_all('/api/animals')
        by_name   = {}
        use_keys  = {}
        for a in animals:
            n = str(a.get('animalName', '')).strip()
            if n:
                by_name[n] = a
            if a.get('animalUseKey') and a.get('use'):
                use_keys[a['use']] = a['animalUseKey']

        missing = [u for u in (use_p14, use_p56) if u not in use_keys]
        if missing:
            print(f'  \u2717 Use value(s) not found in Climb: {missing}')
            print(f'    Available: {sorted(use_keys)}')
            return

        key_p14, key_p56 = use_keys[use_p14], use_keys[use_p56]

        PRESERVE = [
            'alternatePhysicalID', 'heldFor', 'citesNumber', 'lineKey', 'sexKey',
            'generationKey', 'breedingStatusKey', 'dietKey', 'animalStatusKey',
            'exitReasonKey', 'animalName', 'physicalMarker', 'dateBorn',
            'dateExit', 'comments', 'commentStatus', 'owner', 'arrivalDate',
            'iacucprotocolKey', 'physicalMarkerTypeKey', 'materialOriginKey',
            'externalIdentifier', 'microchipIdentifier',
        ]

        jobs, already, not_found = [], [], []

        for _, row in working_df.iterrows():
            name = str(row.get('Animal_Name') or row.get('Name') or '').strip()
            if not name:
                continue

            timepoint = (row.get('Assigned_Timepoint')
                         or row.get('Harvest Timepoint')
                         or row.get('Age (Days)') or '')
            want_key, want_name = ((key_p14, use_p14) if _is_p14(timepoint)
                                   else (key_p56, use_p56))

            climb = by_name.get(name)
            if not climb:
                not_found.append(name)
                continue

            if climb.get('animalUseKey') == want_key:
                already.append(name)
                continue

            jobs.append({'name': name, 'climb': climb,
                         'was': climb.get('use', ''),
                         'want_key': want_key, 'want_name': want_name})

        print(f'  To update  : {len(jobs)}')
        print(f'  Already set: {len(already)}')
        if not_found:
            print(f'  Not in Climb: {len(not_found)}  {not_found[:5]}')

        if not jobs:
            print('  Nothing to change.')
            return

        counts = {}
        for j in jobs:
            counts[j['want_name']] = counts.get(j['want_name'], 0) + 1
        for k, v in sorted(counts.items()):
            print(f'    {k:16} {v}')

        ok, failed = 0, []
        for j in jobs:
            climb   = j['climb']
            payload = {f: climb[f] for f in PRESERVE if climb.get(f) is not None}
            payload['animalUseKey'] = j['want_key']
            payload['cohortKeys']   = [c.get('cohortKey')
                                       for c in (climb.get('cohorts') or [])
                                       if c.get('cohortKey') is not None]
            payload['jobKeys']               = []
            payload['housings']              = []
            payload['animalCharacteristics'] = []

            animal_id = climb.get('animalId') or climb.get('animalID')
            try:
                time.sleep(0.12)
                r = _api_put(f'{sc._API_BASE}/api/animals/{animal_id}',
                                 headers=_hdr(), json=payload, timeout=30)
                if not r.ok:
                    raise requests.HTTPError(f'{r.status_code} {r.text[:200]}')
                ok += 1
                print(f"    {j['name']}: {j['was'] or '(blank)'} \u2192 {j['want_name']}")
            except Exception as e:
                failed.append((j['name'], str(e)))
                print(f"    {j['name']}: FAILED \u2014 {e}")

        print(f'\n  \u2713 Use updated on {ok} of {len(jobs)} animals')
        if failed:
            print(f'  \u26a0 {len(failed)} failed \u2014 re-run to retry')

    except Exception as e:
        print(f'  \u26a0 Use update failed: {e}')
        print('    The schedule is unaffected.')


def _cohort_name_for_row(row) -> str:
    """
    Cohort a scheduled animal belongs in.

        P14   -> 'P14 <harvest date>'
        Adult -> 'P56 <behavior date>'

    Returns '' if no usable date, so the animal is reported rather than
    silently assigned to the wrong cohort.
    """
    date_val = _file_date_for_row(row)
    if not date_val:
        return ''
    date_str = _harvest_date_str(date_val)
    if not date_str:
        return ''

    def _get(name):
        try:
            return row.get(name)
        except AttributeError:
            return row[name] if name in row else None

    timepoint = (_get('Age (Days)') or _get('Assigned_Timepoint')
                 or _get('Harvest Timepoint') or '')
    prefix = 'P14' if _is_p14(timepoint) else 'P56'
    return f'{prefix} {date_str}'


def plan_cohorts(working_df):
    """
    Work out which cohorts the scheduled animals need.

    Returns (plan, unassignable) where plan is:
        { cohort_name: {'animals': [names], 'birth_dates': [sorted unique]} }
    and unassignable is a list of animal names with no usable date.
    """
    plan = {}
    unassignable = []

    for _, row in working_df.iterrows():
        try:
            name = str(row.get('Animal_Name') or row.get('Name') or '').strip()
        except AttributeError:
            name = ''
        if not name:
            continue

        cohort = _cohort_name_for_row(row)
        if not cohort:
            unassignable.append(name)
            continue

        entry = plan.setdefault(cohort, {'animals': [], 'birth_dates': set()})
        entry['animals'].append(name)

        bd = None
        for col in ('Birth_Date', 'Birth Date', 'BD'):
            try:
                v = row.get(col)
            except AttributeError:
                v = row[col] if col in row else None
            if v is not None and str(v).strip() not in ('', 'nan', 'NaT', 'None'):
                bd = v
                break
        if bd is not None:
            try:
                entry['birth_dates'].add(pd.to_datetime(bd).strftime('%m/%d/%Y'))
            except Exception:
                entry['birth_dates'].add(str(bd).strip())

    for c in plan.values():
        c['birth_dates'] = sorted(c['birth_dates'])

    return plan, unassignable


def fetch_existing_cohorts():
    """Return {cohort_name: cohortKey} for every cohort in Climb."""
    sc = _load_sing_climb()
    out = {}
    for c in sc._get_all('/api/cohorts'):
        name = str(c.get('name', '') or '').strip()
        key  = c.get('cohortKey') or c.get('key')
        if name and key is not None:
            out[name] = int(key)
    return out


def _cohort_description(cohort_name: str, birth_dates: list) -> str:
    """
    Description text for a cohort, matching how they're written in Climb.

        P14  ->  'BD 08/26/2026'
                 One date. P14 harvest is a fixed 14 days after birth, so every
                 animal in the cohort shares a birth date.

        P56  ->  'BD 06/24/2026 - 07/01/2026'
                 A range. Behavior happens on day 42-49, both multiples of 7,
                 so the window runs Wednesday to Wednesday.

    The dates are derived from the cohort's own date, not from the animals, so
    the range is the full eligible window even if the animals in it don't span
    all of it. Falls back to the observed birth dates if the name won't parse.
    """
    try:
        prefix, date_part = cohort_name.split(' ', 1)
        anchor = datetime.strptime(date_part.strip(), '%Y_%m_%d')
    except Exception:
        return 'BD ' + ', '.join(birth_dates) if birth_dates else ''

    if prefix.upper() == 'P14':
        born = anchor - timedelta(days=CONFIG['P14_HARVEST_AGE_DAYS'])
        return f'BD {born.strftime("%m/%d/%Y")}'

    # P56 — behavior date minus the far and near ends of the window
    start = anchor - timedelta(days=CONFIG['P56_BEHAVIOR_END_DAY'])
    end   = anchor - timedelta(days=CONFIG['P56_BEHAVIOR_START_DAY'])
    return f'BD {start.strftime("%m/%d/%Y")} - {end.strftime("%m/%d/%Y")}'


def write_cohort_report(plan, existing, unassignable, output_dir, timestamp):
    """
    Write an XLSX listing the cohorts to create and which animals go in each.

    Sheet 'Cohorts to Create' — one row per missing cohort, with the born dates
                                that belong in its description.
    Sheet 'Animal Assignments' — every animal and its cohort.
    """
    missing = {n: d for n, d in plan.items() if n not in existing}

    cohort_rows = []
    for name in sorted(plan):
        info = plan[name]
        cohort_rows.append({
            'Cohort Name':   name,
            'Status':        'EXISTS' if name in existing else 'CREATE',
            'Cohort Key':    existing.get(name, ''),
            'Animals':       len(info['animals']),
            'Description':   _cohort_description(name, info['birth_dates']),
            'Born Dates':    ', '.join(info['birth_dates']),
        })

    animal_rows = []
    for name in sorted(plan):
        for a in sorted(plan[name]['animals'], key=natural_sort_key):
            animal_rows.append({
                'Animal':      a,
                'Cohort Name': name,
                'Status':      'EXISTS' if name in existing else 'CREATE',
            })
    for a in sorted(unassignable, key=natural_sort_key):
        animal_rows.append({
            'Animal': a, 'Cohort Name': '', 'Status': 'NO DATE — cannot assign',
        })

    dates = sorted({n.split(' ', 1)[1] for n in plan if ' ' in n})
    if not dates:
        date_str = timestamp
    elif len(dates) <= 4:
        date_str = '_'.join(dates)
    else:
        date_str = f'{dates[0]}_to_{dates[-1]}'

    out_path = _make_dated_path(output_dir, 'Cohorts_To_Create', date_str)
    with pd.ExcelWriter(out_path, engine='openpyxl') as w:
        pd.DataFrame(cohort_rows).to_excel(w, sheet_name='Cohorts to Create',
                                          index=False)
        pd.DataFrame(animal_rows).to_excel(w, sheet_name='Animal Assignments',
                                          index=False)
    try:
        from openpyxl import load_workbook
        wb = load_workbook(out_path)
        for ws in wb.worksheets:
            auto_width_worksheet(ws)
        wb.save(out_path)
    except Exception:
        pass

    print(f'  \U0001f4c4 Saved: {os.path.basename(out_path)}  '
          f'({len(missing)} to create, {len(plan) - len(missing)} existing)')
    return out_path


def assign_animals_to_cohorts(plan, existing):
    """
    Add scheduled animals to their cohorts via PUT /api/animals/{animalID}.

    Read-modify-write: the PUT replaces the whole record, so every other field
    is carried back unchanged and existing cohort membership is preserved.

    Returns (updated, skipped, failed).
    """
    import requests, time

    sc = _load_sing_climb()

    def _hdr():
        return {'Authorization':   f'Bearer {sc._get_token()}',
                'X-Workgroup-Key': sc._WORKGROUP_KEY,
                'Content-Type':    'application/json'}

    animals = sc._get_all('/api/animals')
    by_name = {str(a.get('animalName', '')).strip(): a for a in animals
               if str(a.get('animalName', '')).strip()}

    PRESERVE = [
        'alternatePhysicalID', 'heldFor', 'citesNumber', 'lineKey', 'sexKey',
        'generationKey', 'breedingStatusKey', 'dietKey', 'animalStatusKey',
        'exitReasonKey', 'animalName', 'physicalMarker', 'dateBorn', 'dateExit',
        'comments', 'commentStatus', 'owner', 'arrivalDate', 'animalUseKey',
        'iacucprotocolKey', 'physicalMarkerTypeKey', 'materialOriginKey',
        'externalIdentifier', 'microchipIdentifier',
    ]
    # Dates go back exactly as received — see the marker push for why.

    updated, skipped, failed = [], [], []

    for cohort_name in sorted(plan):
        key = existing.get(cohort_name)
        if key is None:
            continue                      # not created — caller handles this
        for name in plan[cohort_name]['animals']:
            climb = by_name.get(name)
            if not climb:
                failed.append((name, 'not found in Climb'))
                continue

            current_keys = [c.get('cohortKey') for c in (climb.get('cohorts') or [])
                            if c.get('cohortKey') is not None]
            if key in current_keys:
                skipped.append(name)
                continue

            payload = {}
            for f in PRESERVE:
                v = climb.get(f)
                if v is not None:
                    payload[f] = v
            payload['cohortKeys'] = sorted(set(current_keys) | {key})
            # Required by Climb or the PUT 400s
            payload['jobKeys']               = []
            payload['housings']              = []
            payload['animalCharacteristics'] = []

            animal_id = climb.get('animalId') or climb.get('animalID')
            try:
                time.sleep(0.12)
                r = _api_put(f'{sc._API_BASE}/api/animals/{animal_id}',
                                 headers=_hdr(), json=payload, timeout=30)
                if not r.ok:
                    raise requests.HTTPError(f'{r.status_code} {r.text[:200]}')
                updated.append((name, cohort_name))
                print(f'    {name} \u2192 {cohort_name}')
            except Exception as e:
                failed.append((name, str(e)))
                print(f'    {name}: FAILED \u2014 {e}')

    return updated, skipped, failed


def run_cohorts(working_df, timestamp, output_dir=None):
    """
    Assign scheduled animals to cohorts.

    Cohorts cannot be created through the API (see CLIMB_API_LIMITATIONS.md), so
    if any are missing this pauses, lists what to create, and waits. On resume
    it re-checks; anything still missing is listed again. Skippable at any point.
    """
    print('\n' + '=' * 80)
    print('COHORT ASSIGNMENT')
    print('=' * 80)

    out_dir = output_dir or _SCRIPT_DIR

    plan, unassignable = plan_cohorts(working_df)
    if not plan:
        print('  No cohorts needed \u2014 no animals with usable dates.')
        return None

    print(f'  {len(plan)} cohort(s) needed for '
          f'{sum(len(v["animals"]) for v in plan.values())} animals')
    if unassignable:
        print(f'  \u26a0 {len(unassignable)} animal(s) have no date and cannot be '
              f'assigned: {unassignable[:5]}')

    try:
        existing = fetch_existing_cohorts()
    except Exception as e:
        print(f'  \u2717 Could not read cohorts from Climb: {e}')
        return None

    report = write_cohort_report(plan, existing, unassignable, out_dir, timestamp)

    # ── Wait for missing cohorts to be created ───────────────────────────────
    attempt = 0
    while True:
        missing = {n: v for n, v in plan.items() if n not in existing}
        if not missing:
            break

        attempt += 1
        print(f'\n  {len(missing)} cohort(s) do not exist in Climb yet:')
        for n in sorted(missing):
            info = missing[n]
            bd = _cohort_description(n, info['birth_dates'])
            print(f'    {n}   ({len(info["animals"])} animals)   {bd}')

        answer = _gui_ask('cohort_create',
                          missing=[{'name': n,
                                    'count': len(v['animals']),
                                    'description': _cohort_description(n, v['birth_dates'])}
                                   for n, v in sorted(missing.items())],
                          attempt=attempt,
                          report=os.path.basename(report))

        if answer == 'skip':
            print('\n  Cohort assignment skipped.')
            print(f'  See {os.path.basename(report)} for what still needs creating.')
            return report

        # 'recheck' — pull the list again
        print('\n  Re-checking Climb...')
        try:
            existing = fetch_existing_cohorts()
        except Exception as e:
            print(f'  \u2717 Could not re-read cohorts: {e}')
            return report
        still = [n for n in plan if n not in existing]
        if still:
            print(f'  {len(still)} still missing.')
        else:
            print('  All cohorts found.')
            write_cohort_report(plan, existing, unassignable, out_dir, timestamp)

    # ── Assign ───────────────────────────────────────────────────────────────
    print(f'\n  Adding animals to {len(plan)} cohort(s)...')
    updated, skipped, failed = assign_animals_to_cohorts(plan, existing)

    print(f'\n  \u2713 Cohort assignment complete')
    print(f'    Added        : {len(updated)}')
    print(f'    Already in   : {len(skipped)}')
    print(f'    Failed       : {len(failed)}')
    for name, err in failed[:10]:
        print(f'      {name}: {err}')

    return report


def run_climb_to_envision(working_df, timestamp, output_dir=None):
    """STEP 3: Create Envision translation."""
    print("\n" + "=" * 80)
    print("STEP 3: CLIMB TO ENVISION")
    print("=" * 80)

    if working_df.empty:
        print("  ✗ No data. Skipping.")
        return None

    df = working_df.copy()

    # Normalise column names from CSV variants to internal names
    _col_map = {
        'Name':       'Animal_Name',
        'Birth Date': 'Birth_Date',
    }
    for _src, _dst in _col_map.items():
        if _src in df.columns and _dst not in df.columns:
            df = df.rename(columns={_src: _dst})

    required = ['Genotype', 'Sex', 'Housing ID', 'Animal_Name', 'Line', 'Birth_Date']
    missing = [col for col in required if col not in df.columns]
    if missing:
        print(f"  \u274c Envision: missing columns: {missing}")
        print(f"  Available: {list(df.columns)}")
        raise RuntimeError(
            f"Envision translation cannot run \u2014 missing columns: {missing}"
        )

    print(f"  Processing {len(df)} animals...")

    df['genotype_base'] = df.apply(
        lambda row: clean_genotype_base(row['Genotype'], row['Line']), axis=1)
    df['sex_initial'] = df['Sex'].str[0].str.upper()

    # Use Line (Short) for the group label so it shows e.g. "Dll1-F" not "Het-F".
    # Fall back to genotype_base if Line (Short) is missing or empty.
    if 'Line (Short)' in df.columns:
        line_short = df['Line (Short)'].str.strip()
    elif 'Strain' in df.columns:
        line_short = df['Strain'].str.strip()
    else:
        line_short = pd.Series('', index=df.index)
    line_short = line_short.where(line_short != '', df['genotype_base'])
    df['Group_base'] = line_short + '-' + df['sex_initial']

    group_suffixes = group_animals_by_housing(df)
    df['Group'] = df.index.map(group_suffixes)

    df = assign_ear_tags_by_strain_sex(df)
    df['Genotype_clean'] = df['Genotype'].apply(clean_genotype)

    output_df = pd.DataFrame({
        'Group': df['Group'],
        'Cage': df['Housing ID'],
        'Animal ID': df['Animal_Name'],
        'Envision Ear Tag': df['Envision Ear Tag'],
        'Strain': df['Line'],
        'Coat Color': '',
        'Genotype': df['Genotype_clean'],
        'Additional Detail': '',
        'Sex': df['Sex'],
        'Birth Date': df['Birth_Date'],
        'Ear notch': '',
        'Metal ear tag': '',
        'Other ID': '',
        'RapID code': '',
        'RapID tag color': '',
        'RFID': '',
        'Tail Tattoo': ''
    })
    output_df = output_df[ENVISION_TEMPLATE_COLUMNS]

    # File date: harvest date for P14, behavior date for adults
    output_df['_harvest_date'] = df.apply(_file_date_for_row, axis=1).values

    _out_dir = output_dir if output_dir else _SCRIPT_DIR
    saved_paths = []

    for date_val, group in output_df.groupby('_harvest_date', sort=True):
        date_str = _harvest_date_str(date_val) if date_val else 'unknown_date'
        out_path = _make_dated_path(_out_dir, 'Envision', date_str)
        export_df = group.drop(columns=['_harvest_date'])

        wb = Workbook()
        ws = wb.active
        ws.title = 'template_csv_v1.0'
        ws.append(list(export_df.columns))
        for _, row in export_df.iterrows():
            ws.append([row[c] for c in export_df.columns])
        auto_width_worksheet(ws)
        wb.save(out_path)
        saved_paths.append(out_path)
        print(f"  \U0001f4c4 Saved: {os.path.basename(out_path)}  ({len(export_df)} animals)")

    print(f"  \u2713 Envision complete: {len(saved_paths)} file(s)")

    # Push the assigned tags into Climb — only from a complete list.
    if saved_paths and CONFIG.get('PUSH_RAPID_MARKERS_TO_CLIMB', True):
        _push_rapid_markers_to_climb(output_df, df)

    return saved_paths if saved_paths else None


# ============================================================
# STEP 4: LABELS
# ============================================================

def safe_date_format(date_value, date_name='Date'):
    try:
        return pd.to_datetime(date_value).strftime('%m/%d/%y')
    except Exception:
        if pd.notna(date_value):
            return str(date_value)
        return 'N/A'


def safe_int_format(value, default='N/A'):
    try:
        if pd.notna(value):
            return int(float(value))
        return default
    except Exception:
        return default


def safe_get_label(row, *keys, default='N/A'):
    """Try multiple keys, return first non-null value."""
    for key in keys:
        if key in row:
            value = row[key]
            if isinstance(value, pd.Series):
                value = value.dropna()
                if not value.empty:
                    return value.iloc[0]
            elif pd.notna(value) and str(value).lower() != 'nan':
                return value
    return default


def determine_label_type(preservation):
    preservation_str = str(preservation).strip().lower()
    if 'oct' in preservation_str and 'block' in preservation_str:
        return 'skip', 0
    elif 'frozen' in preservation_str:
        return 'rna', 1
    elif 'pfa' in preservation_str or 'fixed' in preservation_str:
        return 'perfusion', 2
    else:
        # Unrecognised / blank Preservation. Previously defaulted to 'rna',
        # which silently gave perfusion animals RNA tube labels (e.g. the
        # 'Extra - Sex & Timepoint Full' protocol yields Preservation = '').
        # Now produces no label and is reported at the end of the run.
        return 'unknown', 0


def format_sample_number(sample_name, pad=True):
    """
    Format sample name for RNA tube labels.
    Sample names are in the form '765-0', '765-C', etc.
    (numeric part + hyphen + tube suffix 0-6 or C)

    Sides tab (pad=True):  zero-pad numeric part to 4 digits  -> '0765-0'
    Tops  tab (pad=False): strip leading zero from numeric part -> '765-0'
    """
    try:
        s = str(sample_name).strip()
        if '-' in s:
            parts    = s.rsplit('-', 1)
            num_part = parts[0]
            suffix   = parts[1]
        else:
            num_part = s
            suffix   = None
        digits = ''.join(filter(str.isdigit, num_part))
        if not digits:
            return s
        formatted_num = digits.zfill(4) if pad else str(int(digits))
        return f"{formatted_num}-{suffix}" if suffix is not None else formatted_num
    except Exception:
        return str(sample_name)


def create_rna_excel(rna_labels, output_folder, timestamp):
    """
    Create RNA tube labeler Excel file with two tabs: Sides and Tops.
    One row per label — no grid/page logic.
    """
    if not rna_labels:
        _pipeline_queue.put({'kind': _MSG_LOG, 'text': '  ⚠ No RNA labels to create.'})
        return None

    _pipeline_queue.put({'kind': _MSG_LOG, 'text': '  Creating RNA Tube Labeler file...'})

    # Error check: label numbers must match between Sides and Tops
    mismatches = [i + 1 for i, lbl in enumerate(rna_labels)
                  if lbl['Sides_Label_Num'] != lbl['Tops_Label_Num']]
    if mismatches:
        _pipeline_queue.put({'kind': _MSG_LOG,
                             'text': f'  ❌ RNA label number mismatch at positions: {mismatches}'})
        raise ValueError(f'RNA label number mismatch at rows: {mismatches}')

    output_files = []
    from itertools import groupby as _groupby
    labels_by_date = {}
    for lbl in rna_labels:
        d = lbl.get('_harvest_date', '') or 'unknown_date'
        labels_by_date.setdefault(d, []).append(lbl)

    for date_str, date_labels in sorted(labels_by_date.items()):
        # Convert display date (MM/DD/YY) to filename format (YYYY_MM_DD)
        try:
            fname_date = pd.to_datetime(date_str).strftime('%Y_%m_%d')
        except Exception:
            fname_date = date_str.replace('/', '_').replace('-', '_')

        output_file = _make_dated_path(output_folder, 'Tube_Labeler_RNA', fname_date)

        sides_df = pd.DataFrame({
            'Label Number':  [l['Sides_Label_Num'] for l in date_labels],
            'Sample_Date':   [l['Sides_B']          for l in date_labels],
            'Animal_Strain': [l['Sides_C']          for l in date_labels],
        })
        tops_df = pd.DataFrame({
            'Label Number':  [l['Tops_Label_Num'] for l in date_labels],
            'Sample Number': [l['Tops_B']          for l in date_labels],
            'Animal Number': [l['Tops_C']          for l in date_labels],
        })

        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            sides_df.to_excel(writer, sheet_name='Sides', index=False, header=False)
            tops_df.to_excel(writer,  sheet_name='Tops',  index=False, header=False)

        _pipeline_queue.put({'kind': _MSG_LOG,
                             'text': f'  \u2713 RNA Tube Labeler saved: {_os.path.basename(output_file)}'
                                     f'  (Sides={len(sides_df)}, Tops={len(tops_df)})'})
        output_files.append(output_file)

    return output_files if output_files else None


def format_label_rows(row, label_type):
    """Create the 4 rows of text for each label."""
    harvest_date = safe_date_format(
        safe_get_label(row, 'Sample Harvest Date', 'Harvest Date', 'Harvest_Date'))
    born_date = safe_date_format(
        safe_get_label(row, 'Birth_Date', 'Birth Date'))

    sex_val = safe_get_label(row, 'Sex', 'Sex_animal', 'Sex_sample')
    sex = str(sex_val).upper()[0] if pd.notna(sex_val) and str(sex_val) != 'N/A' else 'U'

    line_stock_val = safe_get_label(row, 'Line (Stock)')
    line_stock = str(line_stock_val).lstrip('0') if pd.notna(line_stock_val) and str(line_stock_val) != 'N/A' else ''

    time_point = str(safe_get_label(row, 'Assigned_Timepoint')).strip()
    if time_point and time_point not in ('N/A', 'nan', ''):
        days = ''.join(filter(str.isdigit, time_point))
        if days:
            time_point = f"P{days}"
    else:
        time_point = "N/A"

    genotype = clean_genotype_labels(
        safe_get_label(row, 'Genotype', 'Genotype_animal', 'Genotype_sample'))

    # Calculate age in weeks
    age_weeks = 'N/A'
    try:
        bd = pd.to_datetime(safe_get_label(row, 'Birth_Date', 'Birth Date'))
        hd = pd.to_datetime(safe_get_label(row, 'Sample Harvest Date', 'Harvest Date', 'Harvest_Date'))
        if pd.notna(bd) and pd.notna(hd):
            age_weeks = int((hd - bd).days / 7)
    except Exception:
        pass

    sample_name = safe_get_label(row, 'Sample Name', 'Sample_Name')
    animal_name = safe_get_label(row, 'Animal Name', 'Animal_Name')
    line_short = safe_get_label(row, 'Line (Short)')

    # Calculate age in days at harvest for label (P prefix)
    age_days_label = 'N/A'
    try:
        bd = pd.to_datetime(safe_get_label(row, 'Birth_Date', 'Birth Date'))
        hd = pd.to_datetime(safe_get_label(row, 'Sample Harvest Date', 'Harvest Date', 'Harvest_Date'))
        if pd.notna(bd) and pd.notna(hd):
            age_days_label = f"P{int((hd - bd).days)}"
    except Exception:
        pass

    row1 = f"{sample_name}_{harvest_date}_{animal_name}"
    row2 = f"{age_weeks}Wks_{sex}_{line_short}_{line_stock}"
    row3 = f"{genotype}_{born_date}_{age_days_label}"
    row4 = "Mouse_Perfused Brain" if label_type.lower() == 'perfusion' else "Mouse_Frozen Brain"

    return row1, row2, row3, row4


def generate_all_labels(merged_df):
    """Generate all label data from merged dataframe.

    Returns:
        perfusion_labels  – list of {'Row 1'…'Row 4'} dicts for mail-merge sheets
        rna_labels        – list of Sides/Tops dicts for the RNA Tube Labeler file
        perfusion_count, rna_count, oct_count
    """
    perfusion_labels = []
    rna_labels       = []
    perfusion_count  = 0
    rna_count        = 0
    oct_count        = 0
    unknown_rows     = []   # (sample, animal, preservation) — reported below

    for _, data_row in merged_df.iterrows():
        preservation = safe_get_label(data_row, 'Preservation', 'Preservation_sample', 'Preservation_animal')
        label_type, copies = determine_label_type(preservation)
        sample_name  = safe_get_label(data_row, 'Sample Name', 'Sample_Name')
        animal_name  = safe_get_label(data_row, 'Animal Name', 'Animal_Name')

        if label_type == 'skip':
            oct_count += 1
            continue

        if label_type == 'unknown':
            unknown_rows.append((sample_name, animal_name, preservation))
            continue

        if label_type == 'perfusion':
            perfusion_count += 1
            try:
                row1, row2, row3, row4 = format_label_rows(data_row, label_type)
            except Exception as e:
                print(f"    ✗ Error formatting perfusion '{sample_name}': {e}")
                traceback.print_exc()
                continue
            for _ in range(copies):
                perfusion_labels.append({
                    'Row 1': row1, 'Row 2': row2, 'Row 3': row3, 'Row 4': row4,
                    '_harvest_date': safe_date_format(_file_date_for_row(data_row))
                })

        else:  # rna
            rna_count += 1
            try:
                harvest_date   = safe_date_format(
                    safe_get_label(data_row, 'Sample Harvest Date', 'Harvest Date', 'Harvest_Date'))
                line_short     = safe_get_label(data_row, 'Line (Short)')
                sample_padded  = format_sample_number(sample_name, pad=True)
                sample_raw     = format_sample_number(sample_name, pad=False)
                animal_str     = str(animal_name).strip()

                rna_labels.append({
                    'Sides_Label_Num': rna_count,
                    'Sides_B':         f"{sample_padded}_{harvest_date}",
                    'Sides_C':         f"{animal_str}_{line_short}",
                    'Tops_Label_Num':  rna_count,
                    'Tops_B':          sample_raw,
                    'Tops_C':          animal_str,
                    '_harvest_date':   safe_date_format(_file_date_for_row(data_row)),
                })
            except Exception as e:
                print(f"    ✗ Error formatting RNA '{sample_name}': {e}")
                traceback.print_exc()
                continue

    print(f"    Perfusion: {perfusion_count} × 2 = {perfusion_count * 2} labels")
    print(f"    RNA:       {rna_count} × 1 = {rna_count} labels")
    if oct_count > 0:
        print(f"    OCT Block: {oct_count} × 0 = skipped")
    print(f"    Total perfusion labels: {len(perfusion_labels)}")
    print(f"    Total RNA labels:       {rna_count}")

    if unknown_rows:
        print("")
        print(f"    !! {len(unknown_rows)} sample(s) had an unrecognised Preservation "
              f"value — NO LABEL was made for these:")
        for s_name, a_name, pres in unknown_rows:
            print(f"       Sample {s_name}  Animal {a_name}  Preservation={pres!r}")
        print("       Check the Protocol on these animals in the harvest sheet.")
        print("")

    return perfusion_labels, rna_labels, perfusion_count, rna_count, oct_count


def create_label_sheets(all_labels, output_folder, timestamp):
    """Create Excel label sheets — uses GUI dialogs for per-sheet offset input."""
    return _create_label_sheets_gui(all_labels, output_folder, timestamp)

def run_labels(samples_df, working_df, timestamp):
    """STEP 4: Generate label files."""
    print("\n" + "=" * 80)
    print("STEP 4: LABELS")
    print("=" * 80)

    if samples_df is None:
        print("  ✗ No sample data — the Climb Samples module did not run.")
        print("    Labels are built from sample records, so tick 'Climb Samples'")
        print("    as well as 'Labels', or place a samples.csv in the script folder.")
        return None
    if samples_df.empty:
        print("  ✗ No sample data. Skipping.")
        return None
    if working_df.empty:
        print("  ✗ No animal data. Skipping.")
        return None

    # Prepare samples — rename for merge
    s_df = samples_df.copy()
    s_rename = {}
    if 'Name' in s_df.columns:
        s_rename['Name'] = 'Sample Name'
    if 'Source' in s_df.columns:
        s_rename['Source'] = 'Animal Name'
    if 'Harvest Date' in s_df.columns:
        s_rename['Harvest Date'] = 'Sample Harvest Date'
    s_df = s_df.rename(columns=s_rename)

    # Prepare animals — rename Animal_Name for merge.
    # Scheduler output uses 'Animal_Name'; a raw Climb animals.csv uses 'Name'.
    # Accept either so Labels can run standalone against a Climb export.
    a_df = working_df.copy()
    if 'Animal_Name' in a_df.columns:
        a_df = a_df.rename(columns={'Animal_Name': 'Animal Name'})
    elif 'Name' in a_df.columns:
        a_df = a_df.rename(columns={'Name': 'Animal Name'})

    if 'Animal Name' not in s_df.columns:
        print("  ✗ 'Animal Name' not found in samples after rename")
        return None
    if 'Animal Name' not in a_df.columns:
        print("  ✗ 'Animal Name' not found in animal data after rename")
        return None

    s_df['Animal Name'] = s_df['Animal Name'].astype(str).str.strip()
    a_df['Animal Name'] = a_df['Animal Name'].astype(str).str.strip()

    common = set(s_df['Animal Name'].unique()).intersection(set(a_df['Animal Name'].unique()))
    if len(common) == 0:
        print("  ✗ No matching animal names between samples and animals — cannot create labels.")
        return None

    merged_df = pd.merge(s_df, a_df, on='Animal Name', how='inner',
                         suffixes=('_sample', '_animal'))

    unmatched = len(s_df) - len(merged_df)
    if unmatched > 0:
        print(f"  ⚠ {unmatched} samples did not match")
    print(f"  Matched {len(merged_df)} samples with animal data")

    if len(merged_df) == 0:
        print("  ✗ No matches — cannot create labels.")
        return None

    # Sort by animal number then sample number before generating labels
    def _animal_sort_key(name):
        parts = re.split(r'(\d+)', str(name))
        return ''.join(p.zfill(10) if p.isdigit() else p.lower() for p in parts)

    def _sample_sort_key(name):
        s = str(name).strip()
        base = s.split('-')[0] if '-' in s else s
        digits = ''.join(filter(str.isdigit, base))
        return int(digits) if digits else 0

    merged_df['_animal_sort'] = merged_df['Animal Name'].apply(_animal_sort_key)
    merged_df['_sample_sort'] = merged_df['Sample Name'].apply(_sample_sort_key)
    merged_df = merged_df.sort_values(['_animal_sort', '_sample_sort']).drop(
        ['_animal_sort', '_sample_sort'], axis=1
    ).reset_index(drop=True)

    print("\n  Generating labels...")
    perfusion_labels, rna_labels, perf_count, rna_count, oct_count = generate_all_labels(merged_df)

    if not perfusion_labels and not rna_labels:
        if oct_count > 0:
            print("  ⚠ All samples are OCT Block — no labels needed.")
        else:
            print("  ✗ No labels generated.")
        return None

    script_dir = _SCRIPT_DIR
    created_files = []

    # --- RNA Tube Labeler (already grouped by date inside create_rna_excel) ---
    if rna_labels:
        rna_files = create_rna_excel(rna_labels, script_dir, timestamp)
        if rna_files:
            if isinstance(rna_files, list):
                created_files.extend(rna_files)
            else:
                created_files.append(rna_files)

    # --- Perfusion Mail-Merge sheets — grouped by harvest date ---
    if perfusion_labels:
        labels_by_date = {}
        for lbl in perfusion_labels:
            d = lbl.get('_harvest_date', '') or 'unknown_date'
            labels_by_date.setdefault(d, []).append(lbl)

        for date_disp, date_labels in sorted(labels_by_date.items()):
            try:
                date_str = pd.to_datetime(date_disp).strftime('%Y_%m_%d')
            except Exception:
                date_str = date_disp.replace('/', '_').replace('-', '_')
            num_sheets, perf_files = _create_label_sheets_gui(
                date_labels, script_dir, timestamp, date_str=date_str
            )
            created_files.extend(perf_files)

    total = len(created_files)
    if total > 0:
        print(f"\n  \u2713 Step 4 complete: {total} label file(s)")
        if oct_count > 0:
            print(f"    Note: {oct_count} OCT Block sample(s) skipped")

    return created_files if created_files else None


# ============================================================


# ============================================================================
# UNIFIED MAIN — Scheduler → Harvest Pipeline in one shot
# ============================================================================


# ============================================================================
# PIPELINE GUI LAUNCHER
# ============================================================================

# ============================================================================
# PIPELINE GUI LAUNCHER  (replaces all terminal interaction)
# ============================================================================
import queue      as _queue
import threading  as _threading
import tkinter    as tk
from tkinter import ttk, filedialog, messagebox, scrolledtext
import os as _os

# ---------------------------------------------------------------------------
# Thread ↔ GUI messaging
# ---------------------------------------------------------------------------
_MSG_LOG      = 'log'        # pipeline → GUI: a line of text
_MSG_REQUEST  = 'request'    # pipeline → GUI: needs input
_MSG_DONE     = 'done'       # pipeline → GUI: finished (ok or error)

_pipeline_queue  = _queue.Queue()   # pipeline → GUI
_response_queue  = _queue.Queue()   # GUI → pipeline (responses to requests)


def _gui_ask(request_type, **kwargs):
    """
    Called from the pipeline thread to pause and request GUI input.
    Blocks the pipeline thread until the GUI responds.
    """
    _pipeline_queue.put({'kind': _MSG_REQUEST, 'type': request_type, **kwargs})
    return _response_queue.get()   # blocks until GUI calls _gui_respond()


def _gui_respond(value):
    _response_queue.put(value)


# ---------------------------------------------------------------------------
# Redirect stdout → GUI log (so all print() calls appear in the log widget)
# ---------------------------------------------------------------------------
class _QueueWriter:
    def __init__(self):
        self.encoding = 'utf-8'
    def write(self, text):
        if text and text != '\n':
            for line in text.splitlines():
                if line.strip():
                    _pipeline_queue.put({'kind': _MSG_LOG, 'text': line})
    def flush(self):
        pass


# ---------------------------------------------------------------------------
# Replacements for terminal input() calls inside the pipeline
# ---------------------------------------------------------------------------
def get_starting_sample_number():
    """Fetch next sample number from Climb (globally sequential across all projects)."""
    try:
        _sc = _load_sing_climb()
        fetched = _sc.get_next_sample_number(verbose=False)
    except Exception as ex:
        _pipeline_queue.put({'kind': _MSG_LOG,
                             'text': f'ERROR: Could not fetch sample number from Climb: {ex}'})
        raise
    _pipeline_queue.put({'kind': _MSG_LOG,
                         'text': f'  Next sample number: {fetched}'})
    return fetched


def _create_label_sheets_gui(all_labels, output_folder, timestamp, date_str=''):
    """GUI version of create_label_sheets — always starts from label position 1."""
    if not all_labels:
        _pipeline_queue.put({'kind': _MSG_LOG, 'text': '    \u2717 No labels to create!'})
        return 0, []

    created_files       = []
    current_label_index = 0
    sheet_num           = 1

    while current_label_index < len(all_labels):
        labels_remaining = len(all_labels) - current_label_index
        _pipeline_queue.put({'kind': _MSG_LOG,
                              'text': f'    \U0001f4c4 Label sheet {sheet_num}  ({labels_remaining} labels remaining)'})

        # Always start from position 1 — no skipped labels
        sheet_labels = []
        ci = current_label_index
        while len(sheet_labels) < LABELS_PER_PAGE and ci < len(all_labels):
            lbl = all_labels[ci]
            sheet_labels.append({k: v for k, v in lbl.items() if not k.startswith('_')})
            ci += 1

        labels_placed = ci - current_label_index

        import pandas as _pd
        df = _pd.DataFrame(sheet_labels)
        if date_str:
            out_path = _make_dated_path(output_folder, f'Labels_Mailmerge_sheet{sheet_num}', date_str)
        else:
            out_path = _os.path.join(output_folder, f'Labels_Mailmerge_{timestamp}_sheet{sheet_num}.xlsx')
        save_df_to_excel(df, out_path, sheet_name='Labels')

        _pipeline_queue.put({'kind': _MSG_LOG,
                              'text': f'    \U0001f4c4 Saved: {_os.path.basename(out_path)}  '
                                      f'(placed={labels_placed})'})

        created_files.append(out_path)
        current_label_index += labels_placed
        sheet_num += 1

    return len(created_files), created_files


# ---------------------------------------------------------------------------
# Wednesday capacity GUI  (replaces prompt_wednesday_capacity)
# ---------------------------------------------------------------------------
def prompt_wednesday_capacity_gui(parent=None):
    """
    Show a standalone window for Wednesday capacity entry.
    Returns (wednesdays, full_dates_or_None) — same contract as the original.
    """
    wednesdays = get_next_wednesdays(6)
    capacity   = CONFIG['WEDNESDAY_CAPACITY']
    result     = {'value': None}

    win = tk.Toplevel(parent) if parent else tk.Tk()
    win.title('Wednesday Capacity')
    win.configure(bg='#f0f0f0')
    win.resizable(False, False)
    win.grab_set()

    # Header
    hdr = tk.Frame(win, bg='#2c3e50', pady=10)
    hdr.pack(fill='x')
    tk.Label(hdr, text='Wednesday Behavior Capacity',
             font=('Helvetica', 14, 'bold'),
             bg='#2c3e50', fg='white').pack()
    tk.Label(hdr,
             text=f'Maximum capacity: {capacity} animals per Wednesday\n'
                  'Enter how many slots are already booked for each date.',
             font=('Helvetica', 9), bg='#2c3e50', fg='#bdc3c7').pack(pady=(2, 6))

    body = tk.Frame(win, bg='#f0f0f0', padx=20, pady=14)
    body.pack()

    tk.Label(body, text='Wednesday', width=22, anchor='w',
             font=('Helvetica', 10, 'bold'), bg='#f0f0f0').grid(row=0, column=0, padx=(0, 8))
    tk.Label(body, text='Already Booked', width=14, anchor='center',
             font=('Helvetica', 10, 'bold'), bg='#f0f0f0').grid(row=0, column=1)
    tk.Label(body, text='Remaining', width=10, anchor='center',
             font=('Helvetica', 10, 'bold'), bg='#f0f0f0').grid(row=0, column=2)
    tk.Label(body, text='Status', width=12, anchor='center',
             font=('Helvetica', 10, 'bold'), bg='#f0f0f0').grid(row=0, column=3)

    entries     = {}
    status_vars = {}
    remain_vars = {}

    def _update_row(wed, var, status_lbl, remain_lbl):
        try:
            booked = int(var.get()) if var.get().strip() else 0
            booked = max(0, booked)
        except ValueError:
            booked = 0
        rem = capacity - booked
        remain_lbl.configure(text=str(rem))
        if rem <= 0:
            status_lbl.configure(text='🔴 FULL',  fg='#c0392b')
        elif rem <= 3:
            status_lbl.configure(text='🟡 LOW',   fg='#e67e22')
        else:
            status_lbl.configure(text='🟢 Open',  fg='#27ae60')

    for i, wed in enumerate(wednesdays, 1):
        r = i
        label = wed.strftime('%A, %Y-%m-%d')
        bg = '#ffffff' if i % 2 == 0 else '#f7f7f7'

        tk.Label(body, text=label, width=22, anchor='w',
                 font=('Helvetica', 9), bg='#f0f0f0').grid(row=r, column=0, pady=4, padx=(0, 8))

        var = tk.StringVar(value='0')
        e   = ttk.Spinbox(body, from_=0, to=capacity, textvariable=var, width=6)
        e.grid(row=r, column=1, pady=4)
        entries[wed] = var

        remain_lbl = tk.Label(body, text=str(capacity), width=10, anchor='center',
                              font=('Helvetica', 9), bg='#f0f0f0', fg='#27ae60')
        remain_lbl.grid(row=r, column=2)

        status_lbl = tk.Label(body, text='🟢 Open', width=12, anchor='center',
                              font=('Helvetica', 9), bg='#f0f0f0', fg='#27ae60')
        status_lbl.grid(row=r, column=3)

        var.trace_add('write', lambda *_, w=wed, v=var, sl=status_lbl, rl=remain_lbl:
                      _update_row(w, v, sl, rl))

    def _confirm():
        full_dates = []
        for wed in wednesdays:
            try:
                booked = int(entries[wed].get()) if entries[wed].get().strip() else 0
            except ValueError:
                booked = 0
            if capacity - booked <= 0:
                full_dates.append(wed)
        result['value'] = (wednesdays, full_dates if full_dates else None)
        win.destroy()

    foot = tk.Frame(win, bg='#ecf0f1', pady=8)
    foot.pack(fill='x', padx=20)
    tk.Button(foot, text='Continue  →', command=_confirm,
              font=('Helvetica', 11, 'bold'), bg='#27ae60', fg='white',
              relief='flat', padx=16, pady=6, cursor='hand2').pack(side='right')

    # center on parent
    win.update_idletasks()
    pw = win.winfo_screenwidth();  ph = win.winfo_screenheight()
    ww = win.winfo_width();        wh = win.winfo_height()
    win.geometry(f'+{(pw-ww)//2}+{(ph-wh)//2}')

    if parent:
        parent.wait_window(win)
    else:
        win.mainloop()

    if result['value'] is None:
        # Window closed without confirming — use empty
        return wednesdays, None
    return result['value']


# ---------------------------------------------------------------------------
# Main GUI launcher
# ---------------------------------------------------------------------------

# ── Design tokens ────────────────────────────────────────────────────────────
_T = {
    # Surfaces
    'bg':          '#1a1f2e',   # main body
    'bg_subtle':   '#13161f',   # header / footer strips
    'bg_inset':    '#242836',   # cards, inset sections
    # Text
    'text':        '#e2e8f0',   # primary — high contrast on dark
    'text_muted':  '#a0aec0',   # descriptions, secondary
    'text_faint':  '#718096',   # hints, disabled-ish
    # Borders
    'border':      '#2f3550',   # default hairline
    'border_mid':  '#3d4a6b',   # stronger divider
    # Accent (teal)
    'accent':      '#4fd1c5',   # teal — main action color
    'accent_lt':   '#1a3535',   # teal tint background
    'accent_text': '#4fd1c5',   # teal text on dark surfaces
    # Status
    'red':         '#fc8181',   # bright red — visible on dark
    'red_lt':      '#2d1515',   # red tint background
    'amber':       '#f6ad55',   # bright amber — visible on dark
    'amber_lt':    '#2d1e0a',   # amber tint background
    # Header strip
    'hdr_bg':      '#13161f',
    'hdr_border':  '#2f3550',
}

def _make_styled_button(parent, text, command, style='primary', **kwargs):
    """Return a flat styled button. style = 'primary' | 'secondary' | 'ghost'."""
    styles = {
        'primary':   {'bg': _T['accent'],    'fg': '#0f2929',         'ab': '#38b2ac'},
        'secondary': {'bg': _T['bg_inset'],  'fg': _T['text'],        'ab': _T['border']},
        'ghost':     {'bg': _T['bg'],        'fg': _T['text_muted'],  'ab': _T['bg_inset']},
    }
    s = styles.get(style, styles['secondary'])
    btn = tk.Button(
        parent, text=text, command=command,
        bg=s['bg'], fg=s['fg'], activebackground=s['ab'], activeforeground=s['fg'],
        font=('Helvetica', 10, 'bold' if style == 'primary' else 'normal'),
        relief='flat', bd=0, padx=16, pady=7, cursor='hand2',
        **kwargs
    )
    return btn


def _run_sing_sanity(script_dir: str, timestamp: str) -> None:
    """Run the Sing Sanity tracker comparison and write a report to script_dir."""
    # TODO: integrate Sing_Sanity logic here when Sing Sanity is ported in
    print('  Sing Sanity not yet integrated into the pipeline — run Sing_Sanity.py directly.')


def run_pipeline_gui():
    """Entry point — shows the full GUI pipeline."""

    root = tk.Tk()
    root.title('TAILS')
    root.configure(bg=_T['bg'])
    # ── Module definitions ────────────────────────────────────────────────────
    MODULES = [
        {
            'key':   'schedule',
            'label': 'Schedule Harvest',
            'desc':  'Pull animals from Climb, assign to harvest dates, generate calendar events.',
            'needs': ['climb'],
        },
        {
            'key':   'labels',
            'label': 'Generate Labels',
            'desc':  'Create MRI, MERFISH, and RNA-Seq label sheets.',
            'needs': ['csv', 'harvest_xlsx'],
        },
        {
            'key':   'climb_samples',
            'label': 'Create Climb Samples',
            'desc':  'Register new samples in Climb for the scheduled animals.',
            'needs': ['csv'],
        },
        {
            'key':   'envision',
            'label': 'Climb to Envision Translation',
            'desc':  'Generate Envision-formatted output for tag attachment.',
            'needs': ['csv'],
        },
        {
            'key':   'deliverables',
            'label': 'Export Deliverables Sheet',
            'desc':  'Write confirmed harvest data to the collaborator deliverables sheet.',
            'needs': ['csv', 'harvest_xlsx'],
        },
        {
            'key':   'sanity',
            'label': 'Sing Sanity',
            'desc':  'Compare Harvest Worksheet against LSFM, MERFISH, and RNA-Seq trackers.',
            'needs': ['harvest_xlsx', 'tracker_xlsx'],
        },
    ]

    AUTO_FILES = {
        'harvest_xlsx': ('Sing Harvest Sheet.xlsx',           'Harvest Sheet'),
        'tracker_lsfm': ('Animal and sample tracking.xlsx',   'Animal & Sample Tracking'),
        'tracker_mrf':  ('MERFISH-RNASeq_SampleTracker.xlsx', 'MERFISH / RNA-Seq Tracker'),
    }
    REQUIRED_ANIMAL_COLS = CONFIG.get('REQUIRED_ANIMAL_COLUMNS', [])

    # ── Screen 1: Module selector ─────────────────────────────────────────────
    def screen_module_select():
        root.title('TAILS')

        # Header
        hdr = tk.Frame(root, bg=_T['hdr_bg'], pady=14)
        hdr.pack(fill='x', side='top')
        tk.Label(hdr, text='TAILS',
                 font=('Helvetica', 18, 'bold'),
                 bg=_T['hdr_bg'], fg=_T['text']).pack()
        tk.Label(hdr, text='Tracking Animal Inventory, Logging Shipments',
                 font=('Helvetica', 9),
                 bg=_T['hdr_bg'], fg=_T['text_faint']).pack(pady=(1, 0))
        badge_var = tk.StringVar(value='\u25c9  Full Pipeline')
        badge_lbl = tk.Label(hdr, textvariable=badge_var,
                             font=('Helvetica', 10),
                             bg=_T['hdr_bg'], fg='#63e6be')
        badge_lbl.pack(pady=(2, 0))

        # Footer packed before body so it is always visible
        tk.Frame(root, bg=_T['border'], height=1).pack(fill='x', side='bottom')
        foot = tk.Frame(root, bg=_T['bg_subtle'], pady=10, padx=20)
        foot.pack(fill='x', side='bottom')
        err_lbl = tk.Label(foot, text='', font=('Helvetica', 9, 'italic'),
                           bg=_T['bg_subtle'], fg=_T['red'])
        err_lbl.pack(anchor='w')
        tk.Label(foot, text=f'v{PIPELINE_VERSION}', font=('Helvetica', 8),
                 bg=_T['bg_subtle'], fg=_T['text_muted']).pack(anchor='w')

        module_vars = {m['key']: tk.BooleanVar(value=True) for m in MODULES}

        def _update_badge(*_):
            all_on = all(v.get() for v in module_vars.values())
            n      = sum(1 for v in module_vars.values() if v.get())
            if all_on:
                badge_var.set('\u25c9  Full Pipeline')
                badge_lbl.configure(fg='#63e6be')
            elif n == 0:
                badge_var.set('\u25ce  No modules selected')
                badge_lbl.configure(fg=_T['red'])
            else:
                badge_var.set(f'\u25ce  {n} of {len(MODULES)} modules')
                badge_lbl.configure(fg=_T['text_muted'])

        def _proceed():
            selected = {m['key']: module_vars[m['key']].get() for m in MODULES}
            if not any(selected.values()):
                err_lbl.configure(text='\u26a0  Select at least one module.')
                return
            err_lbl.configure(text='')
            state['modules'] = selected
            if selected.get('schedule'):
                # Schedule pulls from Climb — go to preflight
                _switch(screen_preflight)
            else:
                # Labels/Samples/Envision/Deliverables — use CSV from script folder
                _switch(screen_file_check)

        _make_styled_button(foot, 'Run  \u2192', _proceed, style='primary').pack(side='right')

        # Body
        body = tk.Frame(root, bg=_T['bg'])
        body.pack(fill='both', expand=True, side='top')

        tk.Label(body, text='What do you need?',
                 font=('Helvetica', 11, 'bold'),
                 bg=_T['bg'], fg=_T['text']).pack(anchor='w', padx=18, pady=(12, 6))

        for m in MODULES:
            card = tk.Frame(body, bg=_T['bg_inset'], relief='solid', bd=1, padx=12, pady=6)
            card.pack(fill='x', padx=18, pady=3)
            row = tk.Frame(card, bg=_T['bg_inset'])
            row.pack(fill='x')
            tk.Checkbutton(row, variable=module_vars[m['key']],
                           bg=_T['bg_inset'], activebackground=_T['bg_inset'],
                           selectcolor=_T['accent'],
                           cursor='hand2', command=_update_badge).pack(side='left')
            tf = tk.Frame(row, bg=_T['bg_inset'])
            tf.pack(side='left', fill='x', expand=True, padx=(4, 0))
            tk.Label(tf, text=m['label'],
                     font=('Helvetica', 10, 'bold'),
                     bg=_T['bg_inset'], fg=_T['text'], anchor='w').pack(fill='x')
            tk.Label(tf, text=m['desc'],
                     font=('Helvetica', 9),
                     bg=_T['bg_inset'], fg=_T['text_muted'], anchor='w').pack(fill='x')

        # ── TEST MODE toggle ─────────────────────────────────────────────
        # Blocks every write to Climb. Output files are still produced so the
        # run can be checked, but the Climb import CSV is renamed so it can't
        # be uploaded by mistake.
        test_var = tk.BooleanVar(value=bool(globals().get('TEST_MODE', False)))

        def _toggle_test():
            globals()['TEST_MODE'] = bool(test_var.get())
            _test_lbl.config(
                text=('TEST MODE ON \u2014 nothing will be written to Climb'
                      if test_var.get() else
                      'Test mode (no Climb writes)'),
                fg=(_T['amber'] if test_var.get() else _T['text']))

        test_card = tk.Frame(body, bg=_T['bg_inset'], relief='solid', bd=1,
                             padx=12, pady=6)
        test_card.pack(fill='x', padx=18, pady=(10, 3))
        trow = tk.Frame(test_card, bg=_T['bg_inset'])
        trow.pack(fill='x')
        tk.Checkbutton(trow, variable=test_var,
                       bg=_T['bg_inset'], activebackground=_T['bg_inset'],
                       selectcolor=_T['accent'],
                       cursor='hand2', command=_toggle_test).pack(side='left')
        ttf = tk.Frame(trow, bg=_T['bg_inset'])
        ttf.pack(side='left', fill='x', expand=True, padx=(4, 0))
        _test_lbl = tk.Label(ttf, text='Test mode (no Climb writes)',
                             font=('Helvetica', 10, 'bold'),
                             bg=_T['bg_inset'], fg=_T['text'], anchor='w')
        _test_lbl.pack(fill='x')
        tk.Label(ttf, text=('Runs everything and makes all files, but blocks '
                            'every write to Climb. Import CSV is renamed so it '
                            'cannot be uploaded by accident.'),
                 font=('Helvetica', 9), wraplength=520, justify='left',
                 bg=_T['bg_inset'], fg=_T['text_muted'], anchor='w').pack(fill='x')

        ctrl = tk.Frame(body, bg=_T['bg'])
        ctrl.pack(fill='x', padx=18, pady=(8, 12))

        def _select_all():
            for v in module_vars.values(): v.set(True)
            _update_badge()

        def _clear_all():
            for v in module_vars.values(): v.set(False)
            _update_badge()

        tk.Button(ctrl, text='Select All', command=_select_all,
                  bg=_T['bg'], fg=_T['text_muted'],
                  relief='flat', font=('Helvetica', 9),
                  cursor='hand2').pack(side='left')
        tk.Label(ctrl, text='\u00b7',
                 bg=_T['bg'], fg=_T['text_faint']).pack(side='left', padx=6)
        tk.Button(ctrl, text='Clear All', command=_clear_all,
                  bg=_T['bg'], fg=_T['text_muted'],
                  relief='flat', font=('Helvetica', 9),
                  cursor='hand2').pack(side='left')

        # Size window to fit content exactly, capped at screen height - 80px
        root.update_idletasks()
        w = 660
        h = min(root.winfo_reqheight(), root.winfo_screenheight() - 80)
        x = (root.winfo_screenwidth()  - w) // 2
        y = (root.winfo_screenheight() - h) // 2
        root.geometry(f'{w}x{h}+{x}+{y}')

    # ── Screen 2 (partial runs): file check ───────────────────────────────────
    def screen_file_check():
        import datetime as _dt
        import pandas as _pd

        root.title('TAILS \u2014 Files')
        root.geometry('660x520')

        selected_modules = state.get('modules', {})
        needed = set()
        for m in MODULES:
            if selected_modules.get(m['key']):
                needed.update(m.get('needs', []))

        needs_csv     = 'csv'          in needed
        needs_harvest = 'harvest_xlsx' in needed
        needs_tracker = 'tracker_xlsx' in needed

        hdr = tk.Frame(root, bg=_T['hdr_bg'], pady=14)
        hdr.pack(fill='x')
        tk.Label(hdr, text='Files Needed',
                 font=('Helvetica', 16, 'bold'),
                 bg=_T['hdr_bg'], fg=_T['text']).pack()
        tk.Label(hdr, text='Confirm the files below before running.',
                 font=('Helvetica', 9),
                 bg=_T['hdr_bg'], fg=_T['text_muted']).pack(pady=(2, 0))

        body = tk.Frame(root, bg=_T['bg_subtle'], padx=20, pady=14)
        body.pack(fill='both', expand=True)

        def _check_auto(filename):
            path = _os.path.join(_SCRIPT_DIR, filename)
            if not _os.path.exists(path):
                return False, False
            mtime = _dt.date.fromtimestamp(_os.path.getmtime(path))
            return True, mtime == _dt.date.today()

        if needs_harvest or needs_tracker:
            tk.Label(body, text='Found automatically',
                     font=('Helvetica', 10, 'bold'),
                     bg=_T['bg_subtle'], fg=_T['text']).pack(anchor='w', pady=(0, 4))
            check_keys = []
            if needs_harvest: check_keys.append('harvest_xlsx')
            if needs_tracker: check_keys.extend(['tracker_lsfm', 'tracker_mrf'])
            for key in check_keys:
                fname, display = AUTO_FILES[key]
                found, today = _check_auto(fname)
                if found:
                    icon  = '\u2713' if today else '\u26a0'
                    age   = 'today'  if today else 'not from today'
                    color = _T['accent'] if today else _T['amber']
                else:
                    icon, age, color = '\u2717', 'not found', _T['red']
                row = tk.Frame(body, bg=_T['bg_subtle'])
                row.pack(fill='x', pady=1)
                tk.Label(row, text=f'{icon}  {display}',
                         font=('Helvetica', 10),
                         bg=_T['bg_subtle'], fg=color).pack(side='left')
                tk.Label(row, text=f'  ({fname}  \u00b7  {age})',
                         font=('Helvetica', 9),
                         bg=_T['bg_subtle'], fg=_T['text_faint']).pack(side='left')
            tk.Frame(body, bg=_T['border'], height=1).pack(fill='x', pady=10)

        csv_path    = _os.path.join(_SCRIPT_DIR, 'animals.csv')
        confirm_btn = [None]

        def _validate_csv():
            if not _os.path.exists(csv_path):
                return False, 'Not found  \u2014  drop animals.csv in the script folder'
            try:
                test = _pd.read_csv(csv_path, nrows=2)
                missing = [c for c in REQUIRED_ANIMAL_COLS if c not in test.columns]
                if missing:
                    return False, f'Missing columns: {missing}'
            except Exception as ex:
                return False, f'Cannot read file: {ex}'
            return True, 'animals.csv  \u00b7  columns OK'

        if needs_csv:
            tk.Label(body, text='Animal list  (you provide)',
                     font=('Helvetica', 10, 'bold'),
                     bg=_T['bg_subtle'], fg=_T['text']).pack(anchor='w', pady=(0, 6))

            csv_row = tk.Frame(body, bg=_T['bg_subtle'])
            csv_row.pack(fill='x')
            csv_icon_lbl = tk.Label(csv_row, text='',
                                    font=('Helvetica', 10),
                                    bg=_T['bg_subtle'], width=2)
            csv_icon_lbl.pack(side='left')
            csv_msg_lbl = tk.Label(csv_row, text='',
                                   font=('Helvetica', 10),
                                   bg=_T['bg_subtle'], fg=_T['text_muted'], anchor='w')
            csv_msg_lbl.pack(side='left', fill='x', expand=True)

            tk.Label(body,
                     text='\u26a0  This should be a specific list for this run,\n'
                          '    not the full Climb inventory.',
                     font=('Helvetica', 9, 'italic'),
                     bg=_T['amber_lt'], fg=_T['amber'],
                     padx=10, pady=6, justify='left', anchor='w').pack(fill='x', pady=(8, 0))

            tk.Label(body,
                     text=f'Export from Climb \u2192 save as animals.csv \u2192 drop in:\n{_SCRIPT_DIR}',
                     font=('Helvetica', 9),
                     bg=_T['bg_subtle'], fg=_T['text_muted'],
                     justify='left', anchor='w').pack(fill='x', pady=(8, 0))

            def _refresh():
                ok, msg = _validate_csv()
                csv_icon_lbl.configure(text='\u2713' if ok else '\u2026',
                                       fg=_T['accent'] if ok else _T['text_faint'])
                csv_msg_lbl.configure(text=msg,
                                      fg=_T['accent'] if ok else _T['text_muted'])
                if confirm_btn[0]:
                    confirm_btn[0].configure(state='normal' if ok else 'disabled')
                if not ok:
                    root.after(2000, _refresh)

            root.after(50, _refresh)

        tk.Frame(root, bg=_T['border'], height=1).pack(fill='x')
        foot = tk.Frame(root, bg=_T['bg_subtle'], pady=12, padx=20)
        foot.pack(fill='x')
        err_lbl = tk.Label(foot, text='', font=('Helvetica', 9, 'italic'),
                           bg=_T['bg_subtle'], fg=_T['red'])
        err_lbl.pack(anchor='w')
        _make_styled_button(foot, '\u2190 Back',
                            lambda: _switch(screen_module_select),
                            style='ghost').pack(side='left')

        def _confirm():
            if needs_csv:
                ok, msg = _validate_csv()
                if not ok:
                    err_lbl.configure(text=f'\u26a0  {msg}')
                    return
                state['animal_file'] = csv_path
            err_lbl.configure(text='')
            # Ensure keys needed by _run_pipeline are present for partial runs
            state.setdefault('wednesday_dates',     get_next_wednesdays(6))
            state.setdefault('full_behavior_dates', None)
            state.setdefault('births_file',         None)
            import traceback as _tb, pathlib as _pl
            try:
                _switch(screen_progress)
            except Exception as _ex:
                log_path = _pl.Path(script_dir) / 'pipeline_error.log'
                log_path.write_text(
                    f'Confirm button error:\n{_tb.format_exc()}', encoding='utf-8'
                )
                messagebox.showerror(
                    'Error',
                    f'Failed to start pipeline:\n{_ex}\n\nSee pipeline_error.log for details.'
                )

        btn = _make_styled_button(foot, 'Confirm  \u2192', _confirm, style='primary')
        btn.configure(state='normal' if not needs_csv else 'disabled')
        btn.pack(side='right')
        confirm_btn[0] = btn

    # ── Screen 2 (full pipeline): pre-flight file check ───────────────────────
    def screen_preflight():
        import datetime as _dt
        import pandas as _pd

        root.title('TAILS \u2014 Getting Ready')

        # Header
        hdr = tk.Frame(root, bg=_T['hdr_bg'], pady=14)
        hdr.pack(fill='x', side='top')
        tk.Label(hdr, text='Getting Ready',
                 font=('Helvetica', 18, 'bold'),
                 bg=_T['hdr_bg'], fg=_T['text']).pack()
        tk.Label(hdr, text='Checking for required files\u2026',
                 font=('Helvetica', 10),
                 bg=_T['hdr_bg'], fg=_T['text_muted']).pack(pady=(2, 0))

        # Footer pinned at bottom
        tk.Frame(root, bg=_T['border'], height=1).pack(fill='x', side='bottom')
        foot = tk.Frame(root, bg=_T['bg_subtle'], pady=10, padx=20)
        foot.pack(fill='x', side='bottom')
        _make_styled_button(foot, '\u2190 Back',
                            lambda: _switch(screen_module_select),
                            style='ghost').pack(side='left')
        proceed_btn = _make_styled_button(foot, 'Continue \u2192',
                                          lambda: _switch(screen_wednesday),
                                          style='primary')
        proceed_btn.configure(state='disabled')
        proceed_btn.pack(side='right')

        # Body
        body = tk.Frame(root, bg=_T['bg'], padx=24, pady=16)
        body.pack(fill='both', expand=True, side='top')

        # Build the file list based on which modules are selected.
        # Animals and births come from Climb API — never manual at this screen.
        selected_modules = state.get('modules', {})
        FILE_DEFS = [
            ('Harvest Sheet', 'Sing Harvest Sheet.xlsx', True, None),
        ]
        if selected_modules.get('sanity'):
            FILE_DEFS += [
                ('Animal & Sample Tracking', 'Animal and sample tracking.xlsx',    True, None),
                ('MERFISH / RNA-Seq Tracker', 'MERFISH-RNASeq_SampleTracker.xlsx', True, None),
            ]

        REQUIRED_COLS = {}

        def _file_status(fname, auto):
            """Return (ok, today, msg) for a file in the script folder."""
            path = _os.path.join(_SCRIPT_DIR, fname)
            if not _os.path.exists(path):
                return False, False, 'Not found'
            mtime = _dt.date.fromtimestamp(_os.path.getmtime(path))
            today = mtime == _dt.date.today()
            if not auto and not today:
                return False, False, f'Found but not from today ({mtime})'
            if fname in REQUIRED_COLS:
                try:
                    test = _pd.read_csv(path, nrows=2)
                    missing = [c for c in REQUIRED_COLS[fname] if c not in test.columns]
                    if missing:
                        return False, today, f'Wrong columns: {missing}'
                except Exception as ex:
                    return False, today, f'Cannot read: {ex}'
            return True, today, 'Ready'

        # Build status rows
        status_labels = {}
        for display, fname, auto, instructions in FILE_DEFS:
            row = tk.Frame(body, bg=_T['bg'])
            row.pack(fill='x', pady=4)

            icon_lbl = tk.Label(row, text='\u231b', width=3,
                                font=('Helvetica', 12),
                                bg=_T['bg'], fg=_T['text_faint'], anchor='w')
            icon_lbl.pack(side='left')

            right = tk.Frame(row, bg=_T['bg'])
            right.pack(side='left', fill='x', expand=True)

            name_lbl = tk.Label(right, text=display,
                                font=('Helvetica', 10, 'bold'),
                                bg=_T['bg'], fg=_T['text'], anchor='w')
            name_lbl.pack(fill='x')

            msg_lbl = tk.Label(right, text='Checking\u2026',
                               font=('Helvetica', 9),
                               bg=_T['bg'], fg=_T['text_muted'], anchor='w')
            msg_lbl.pack(fill='x')

            if instructions:
                tk.Label(right,
                         text=f'\u2192  {instructions}\n'
                              f'   Drop in:  {_SCRIPT_DIR}',
                         font=('Helvetica', 8, 'italic'),
                         bg=_T['bg'], fg=_T['text_faint'],
                         justify='left', anchor='w').pack(fill='x', pady=(2, 0))

            status_labels[fname] = (icon_lbl, msg_lbl)

        tk.Frame(body, bg=_T['border'], height=1).pack(fill='x', pady=(12, 0))

        ready_lbl = tk.Label(body, text='',
                             font=('Helvetica', 10, 'bold'),
                             bg=_T['bg'], fg=_T['text_muted'])
        ready_lbl.pack(anchor='w', pady=(8, 0))

        def _refresh():
            all_ready = True
            n_ready = 0
            for display, fname, auto, instructions in FILE_DEFS:
                ok, today, msg = _file_status(fname, auto)
                icon_lbl, msg_lbl = status_labels[fname]
                if ok:
                    icon_lbl.configure(text='\u2713', fg=_T['accent'])
                    msg_lbl.configure(text=msg if not auto else fname,
                                      fg=_T['accent'])
                    n_ready += 1
                elif not auto:
                    icon_lbl.configure(text='\u25cb', fg=_T['text_faint'])
                    msg_lbl.configure(text=msg, fg=_T['text_muted'])
                    all_ready = False
                else:
                    icon_lbl.configure(text='\u2717', fg=_T['red'])
                    msg_lbl.configure(text=msg, fg=_T['red'])
                    all_ready = False

            total = len(FILE_DEFS)
            if all_ready:
                ready_lbl.configure(
                    text=f'\u2713  All {total} files ready \u2014 good to go.',
                    fg=_T['accent'])
                proceed_btn.configure(state='normal')
            else:
                waiting = total - n_ready
                ready_lbl.configure(
                    text=f'{n_ready} of {total} ready \u2014 waiting on {waiting} file{"s" if waiting != 1 else ""}\u2026',
                    fg=_T['text_muted'])
                proceed_btn.configure(state='disabled')
                root.after(2000, _refresh)

        _refresh()

        root.update_idletasks()
        w = 680
        h = min(root.winfo_reqheight() + 20, root.winfo_screenheight() - 80)
        x = (root.winfo_screenwidth()  - w) // 2
        y = (root.winfo_screenheight() - h) // 2
        root.geometry(f'{w}x{h}+{x}+{y}')

    script_dir = _os.path.dirname(_os.path.abspath(__file__))

    # ── Shared state ─────────────────────────────────────────────────────────
    state = {
        'animal_file':   _os.path.join(script_dir, CONFIG['INPUT_ANIMAL_FILE']),
        'tracking_file': _os.path.join(script_dir, CONFIG['INPUT_TRACKING_FILE']),
        'births_file':   _os.path.join(script_dir, CONFIG['INPUT_BIRTHS_FILE']),
    }

    REQUIRED_COLOR = _T['red']
    OPTIONAL_COLOR = _T['text_muted']
    OK_COLOR       = _T['accent']

    # ── Helper: clear root and show a new screen ──────────────────────────────
    def _switch(frame_fn):
        for w in root.winfo_children():
            w.destroy()
        frame_fn()

    # ── Shared header builder ─────────────────────────────────────────────────
    def _make_header(eyebrow: str, title: str, subtitle: str = '') -> tk.Frame:
        """Render the consistent light-grey header used on every screen."""
        hdr = tk.Frame(root, bg=_T['hdr_bg'], pady=16)
        hdr.pack(fill='x')
        # Bottom border line
        sep = tk.Frame(root, bg=_T['hdr_border'], height=1)
        sep.pack(fill='x')
        if eyebrow:
            tk.Label(hdr, text=eyebrow.upper(),
                     font=('Helvetica', 9), bg=_T['hdr_bg'],
                     fg=_T['text_faint']).pack()
        tk.Label(hdr, text=title,
                 font=('Helvetica', 16, 'bold'),
                 bg=_T['hdr_bg'], fg=_T['text']).pack()
        if subtitle:
            tk.Label(hdr, text=subtitle,
                     font=('Helvetica', 10), bg=_T['hdr_bg'],
                     fg=_T['text_muted']).pack(pady=(2, 0))
        return hdr

    def _make_footer() -> tk.Frame:
        """Render the consistent light footer strip."""
        sep = tk.Frame(root, bg=_T['border'], height=1)
        sep.pack(fill='x')
        foot = tk.Frame(root, bg=_T['bg_subtle'], pady=10)
        foot.pack(fill='x')
        return foot

    # ─────────────────────────────────────────────────────────────────────────
    # SCREEN 1: File Setup
    # ─────────────────────────────────────────────────────────────────────────
    def screen_file_setup():
        root.title('TAILS')
        root.geometry('700x560')

        # ── Header ────────────────────────────────────────────────────────────
        hdr = tk.Frame(root, bg=_T['hdr_bg'], pady=18)
        hdr.pack(fill='x')
        tk.Label(hdr, text='TAILS',
                 font=('Helvetica', 20, 'bold'),
                 bg=_T['hdr_bg'], fg=_T['text']).pack()
        tk.Label(hdr, text='Tracking Animal Inventory, Logging Shipments',
                 font=('Helvetica', 9),
                 bg=_T['hdr_bg'], fg=_T['text_faint']).pack(pady=(1, 0))
        tk.Label(hdr, text='Which input files do you have ready?',
                 font=('Helvetica', 10), bg=_T['hdr_bg'], fg=_T['text_muted']).pack(pady=(4, 0))

        body = tk.Frame(root, bg=_T['bg'], padx=24, pady=18)
        body.pack(fill='both', expand=True)

        # ── File card definitions ─────────────────────────────────────────────
        FILE_DEFS = [
            {
                'key':      'animal_file',
                'default':  CONFIG['INPUT_ANIMAL_FILE'],
                'required': True,
                'label':    'Animal Inventory',
                'hint':     f'Required  •  usually "{CONFIG['INPUT_ANIMAL_FILE']}"',
                'desc':     'The main list of all animals currently in the colony.',
            },
            {
                'key':      'tracking_file',
                'default':  CONFIG['INPUT_TRACKING_FILE'],
                'required': False,
                'label':    'Harvest Tracking Sheet',
                'hint':     'Optional  •  check the box to include in this run',
                'desc':     'Tracks how many of each strain/type have already been harvested.',
            },
            {
                'key':      'births_file',
                'default':  CONFIG['INPUT_BIRTHS_FILE'],
                'required': False,
                'label':    'Births Record',
                'hint':     'Optional  •  check the box to include in this run',
                'desc':     'Log of recent births used to identify new P14 animals.',
            },
        ]

        path_vars   = {}
        status_lbls = {}
        toggle_vars = {}   # BooleanVar: True = file is available
        detail_frames = {}

        err_lbl = tk.Label(body, text='', font=('Helvetica', 9, 'italic'),
                           bg=_T['bg'], fg=_T['red'])

        def _update_status(key, path, lbl):
            if not path.strip():
                lbl.configure(text='', fg=_T['text_faint'])
            elif _os.path.exists(path):
                lbl.configure(text='✓ Found', fg=_T['accent'])
            else:
                lbl.configure(text='✗ File not found at this path', fg=_T['red'])

        def _browse(key, var, lbl, title):
            path = filedialog.askopenfilename(
                parent=root, title=title,
                initialdir=_os.path.dirname(var.get()) or script_dir,
                filetypes=[('CSV files', '*.csv'), ('All files', '*.*')]
            )
            if path:
                var.set(path)
                _update_status(key, path, lbl)

        def _toggle_card(key, toggle_var, detail_frame):
            # Detail frame is always visible — toggle only controls whether the
            # file is included in the run (handled in _proceed).
            pass

        for fd in FILE_DEFS:
            key      = fd['key']
            required = fd['required']
            default  = _os.path.join(script_dir, fd['default'])

            # Auto-detect: pre-tick if file exists in script_dir
            exists = _os.path.exists(default)
            initial_path = default if exists else state.get(key, default)

            # ── Card frame ────────────────────────────────────────────────────
            card = tk.Frame(body, bg=_T['bg_inset'], relief='solid', bd=1, padx=12, pady=10)
            card.pack(fill='x', pady=6)

            # Top row: toggle + label
            top_row = tk.Frame(card, bg=_T['bg_inset'])
            top_row.pack(fill='x')

            tvar = tk.BooleanVar(value=exists or required)
            toggle_vars[key] = tvar

            # Checkbox (disabled for required file)
            chk = tk.Checkbutton(
                top_row, variable=tvar,
                bg=_T['bg_inset'], activebackground=_T['bg_inset'],
                cursor='hand2' if not required else 'arrow',
                state='normal' if not required else 'disabled',
            )
            chk.pack(side='left', padx=(0, 6))

            name_color = _T['text'] if required else _T['text_muted']
            tk.Label(top_row, text=fd['label'],
                     font=('Helvetica', 11, 'bold'),
                     bg=_T['bg_inset'], fg=name_color).pack(side='left')

            badge_text  = '  Required  ' if required else '  Optional  '
            badge_color = '#f87171' if required else _T['text_faint']
            tk.Label(top_row, text=badge_text,
                     font=('Helvetica', 8, 'bold'),
                     bg=badge_color, fg='white', padx=4).pack(side='left', padx=8)

            tk.Label(card, text=fd['desc'],
                     font=('Helvetica', 9), bg=_T['bg_inset'], fg=_T['text_muted'],
                     anchor='w').pack(fill='x')

            # Expandable detail section (path + browse)
            detail = tk.Frame(card, bg=_T['bg'], padx=8, pady=6, relief='groove', bd=1)
            detail_frames[key] = detail

            path_row = tk.Frame(detail, bg=_T['bg'])
            path_row.pack(fill='x')

            pvar = tk.StringVar(value=initial_path)
            path_vars[key] = pvar
            state[key] = initial_path

            tk.Label(path_row, text=fd['hint'],
                     font=('Helvetica', 8), bg=_T['bg'], fg=_T['text_faint'],
                     anchor='w').pack(fill='x', pady=(0, 4))

            entry_row = tk.Frame(detail, bg=_T['bg'])
            entry_row.pack(fill='x')

            # Create status label first, then Browse (pack order = right-to-left)
            slbl = tk.Label(entry_row, text='', font=('Helvetica', 9),
                            bg=_T['bg'], width=14, anchor='w')
            status_lbls[key] = slbl

            tk.Button(entry_row, text='Browse…',
                      command=lambda k=key, v=pvar, l=slbl, t=fd['label']:
                          _browse(k, v, l, f'Select {t}'),
                      font=('Helvetica', 9), bg=_T['accent'], fg='#0f2929',
                      relief='flat', padx=8, pady=2, cursor='hand2').pack(side='right', padx=(4, 0))

            slbl.pack(side='right', padx=(6, 0))

            entry = tk.Entry(entry_row, textvariable=pvar, font=('Helvetica', 9))
            entry.pack(side='left', fill='x', expand=True)

            pvar.trace_add('write', lambda *_, k=key, v=pvar, l=slbl:
                           _update_status(k, v.get(), l))
            _update_status(key, initial_path, slbl)

            # Wire toggle
            tvar.trace_add('write', lambda *_, k=key, tv=tvar, df=detail:
                           _toggle_card(k, tv, df))

            # Always show the path row so the user can always browse
            detail.pack(fill='x', pady=(6, 0))

        err_lbl.pack(fill='x', pady=(4, 0))

        # ── Footer ────────────────────────────────────────────────────────────
        foot = tk.Frame(root, bg=_T['bg_subtle'], pady=10)
        foot.pack(fill='x', padx=24)

        def _proceed():
            err_lbl.configure(text='')

            # Required file
            animal = path_vars['animal_file'].get().strip()
            if not toggle_vars['animal_file'].get() or not animal:
                err_lbl.configure(text='⚠  The Animal Inventory file is required to continue.')
                return
            if not _os.path.exists(animal):
                err_lbl.configure(text=f'⚠  Animal Inventory not found: {animal}')
                return

            try:
                import pandas as _pd
                test_df = _pd.read_csv(animal, nrows=2)
                missing = [c for c in CONFIG.get('REQUIRED_ANIMAL_COLUMNS', [])
                           if c not in test_df.columns]
                if missing:
                    err_lbl.configure(text=f'⚠  Animal file missing columns: {missing}')
                    return
            except Exception as ex:
                err_lbl.configure(text=f'⚠  Cannot read Animal Inventory: {ex}')
                return

            state['animal_file'] = animal

            for key in ('tracking_file', 'births_file'):
                if toggle_vars[key].get():
                    p = path_vars[key].get().strip()
                    state[key] = p if _os.path.exists(p) else None
                    if toggle_vars[key].get() and not _os.path.exists(p):
                        err_lbl.configure(
                            text=f'⚠  You indicated the file is available but it was not found:\n{p}\n'
                                 f'Please browse to it or uncheck the box.'
                        )
                        return
                else:
                    state[key] = None

            _switch(screen_wednesday)

        tk.Button(foot, text='Next: Wednesday Capacity  →',
                  command=_proceed,
                  font=('Helvetica', 11, 'bold'), bg=_T['accent'], fg='#0f2929',
                  relief='flat', padx=16, pady=7, cursor='hand2').pack(side='right')

        # Fit window height to content
        def _fit_window():
            root.update_idletasks()
            w = root.winfo_width()
            h = root.winfo_reqheight()
            screen_h = root.winfo_screenheight()
            h = min(h + 20, screen_h - 80)
            x = (root.winfo_screenwidth()  - w) // 2
            y = (root.winfo_screenheight() - h) // 2
            root.geometry(f'{w}x{h}+{x}+{y}')
        root.after(10, _fit_window)


    # ─────────────────────────────────────────────────────────────────────────
    # SCREEN 2: Wednesday Capacity
    # ─────────────────────────────────────────────────────────────────────────
    def screen_wednesday():
        import datetime as _dt
        import pandas as _pd
        import re as _re

        root.title('TAILS \u2014 Scheduled Harvests')

        SKIP = {'fail', 'qc fail', 'extra nb', 'floxed', 'found dead'}

        def _classify(age_str):
            s = str(age_str).strip().upper()
            if s == 'P14':
                return 'P14'
            num = _re.sub(r'[^0-9]', '', s)
            if num and int(num) <= 21:
                return 'P14'
            return 'Adult'

        def _norm_date(raw):
            for fmt in ('%Y-%m-%d %H:%M:%S', '%Y-%m-%d', '%m/%d/%Y', '%m/%d/%y'):
                try:
                    return _dt.datetime.strptime(str(raw).strip(), fmt).date()
                except ValueError:
                    continue
            return None

        # Read Harvest Worksheet and group future scheduled animals
        harvest_path = _os.path.join(script_dir, 'Sing Harvest Sheet.xlsx')
        rows_by_date = {}   # {date: {'P14': int, 'Adult': int}}
        load_error   = None
        today        = _dt.date.today()

        try:
            df = _pd.read_excel(harvest_path, sheet_name='Harvest Worksheet', dtype=str).fillna('')
            for _, row in df.iterrows():
                sample = str(row.get('Sample Number', '')).strip()
                if sample.lower() in SKIP:
                    continue
                d = _norm_date(row.get('Harvest Date', ''))
                if d is None or d < today:
                    continue
                tp = _classify(row.get('Age (Days)', ''))
                if d not in rows_by_date:
                    rows_by_date[d] = {'P14': 0, 'Adult': 0}
                rows_by_date[d][tp] += 1
        except Exception as ex:
            load_error = str(ex)

        sorted_dates = sorted(rows_by_date.keys())

        # Header
        hdr = tk.Frame(root, bg=_T['hdr_bg'], pady=14)
        hdr.pack(fill='x', side='top')
        tk.Label(hdr, text='Scheduled Harvests',
                 font=('Helvetica', 18, 'bold'),
                 bg=_T['hdr_bg'], fg=_T['text']).pack()
        sub = f'{len(sorted_dates)} dates with animals scheduled' if sorted_dates else 'No upcoming harvests found'
        tk.Label(hdr, text=sub,
                 font=('Helvetica', 10),
                 bg=_T['hdr_bg'], fg=_T['text_muted']).pack(pady=(2, 0))

        # Footer
        tk.Frame(root, bg=_T['border'], height=1).pack(fill='x', side='bottom')
        foot = tk.Frame(root, bg=_T['bg_subtle'], pady=10, padx=20)
        foot.pack(fill='x', side='bottom')
        _make_styled_button(foot, '\u2190 Back',
                            lambda: _switch(screen_preflight),
                            style='ghost').pack(side='left')

        def _proceed():
            state['wednesday_dates']     = get_next_wednesdays(6)
            state['full_behavior_dates'] = None
            _switch(screen_progress)

        _make_styled_button(foot, 'Run pipeline  \u2192',
                            command=_proceed, style='primary').pack(side='right')

        # Body
        body = tk.Frame(root, bg=_T['bg'], padx=24, pady=16)
        body.pack(fill='both', expand=True, side='top')

        if load_error:
            tk.Label(body,
                     text=f'\u26a0  Could not read Harvest Sheet:\n{load_error}',
                     font=('Helvetica', 9), bg=_T['bg'], fg=_T['amber'],
                     justify='left', anchor='w').pack(fill='x', pady=(0, 12))

        if not sorted_dates:
            tk.Label(body,
                     text='No incomplete scheduled harvests found from today forward.',
                     font=('Helvetica', 10), bg=_T['bg'], fg=_T['text_muted'],
                     anchor='w').pack(fill='x', pady=12)
        else:
            tbl = tk.Frame(body, bg=_T['bg'])
            tbl.pack(fill='x')

            # Fixed column widths in pixels via minsize
            COL_W = [160, 60, 70, 70, 70]
            for i, w in enumerate(COL_W):
                tbl.columnconfigure(i, minsize=w)

            # Headers
            for col, txt in enumerate(['Date', 'Day', 'P14', 'Adult', 'Total']):
                anchor = 'w' if col < 2 else 'e'
                tk.Label(tbl, text=txt.upper(),
                         font=('Helvetica', 8), fg=_T['text_faint'],
                         bg=_T['bg'], anchor=anchor).grid(
                    row=0, column=col, sticky='ew', pady=(0, 4))

            # Divider
            div = tk.Frame(tbl, bg=_T['border'], height=1)
            div.grid(row=1, column=0, columnspan=5, sticky='ew', pady=(0, 6))

            # Data rows
            for r, d in enumerate(sorted_dates, start=2):
                counts = rows_by_date[d]
                p14    = counts['P14']
                adult  = counts['Adult']
                total  = p14 + adult

                vals = [
                    (d.strftime('%Y-%m-%d'), 'w', _T['text'],       False),
                    (d.strftime('%a'),        'w', _T['text_muted'], False),
                    (str(p14)   if p14   else '\u2014', 'e',
                     _T['accent'] if p14   else _T['text_faint'], False),
                    (str(adult) if adult else '\u2014', 'e',
                     _T['accent'] if adult else _T['text_faint'], False),
                    (str(total), 'e', _T['text'], True),
                ]
                for col, (txt, anchor, color, bold) in enumerate(vals):
                    tk.Label(tbl, text=txt,
                             font=('Helvetica', 10, 'bold' if bold else 'normal'),
                             bg=_T['bg'], fg=color, anchor=anchor).grid(
                        row=r, column=col, sticky='ew', pady=3)

        root.update_idletasks()
        w = 560
        h = min(root.winfo_reqheight() + 10, root.winfo_screenheight() - 80)
        x = (root.winfo_screenwidth()  - w) // 2
        y = (root.winfo_screenheight() - h) // 2
        root.geometry(f'{w}x{h}+{x}+{y}')

    # ─────────────────────────────────────────────────────────────────────────
    # SCREEN 3: Progress + mid-run dialogs
    # ─────────────────────────────────────────────────────────────────────────
    def screen_progress():
        root.title('TAILS — Running…')
        root.geometry('760x560')

        hdr = tk.Frame(root, bg=_T['hdr_bg'], pady=14)
        hdr.pack(fill='x')
        hdr_inner = tk.Frame(hdr, bg=_T['hdr_bg'])
        hdr_inner.pack(fill='x', padx=24)
        tk.Label(hdr_inner, text='Pipeline running',
                 font=('Helvetica', 14, 'bold'),
                 bg=_T['hdr_bg'], fg=_T['text']).pack(side='left')
        status_var = tk.StringVar(value='Starting up…')
        tk.Label(hdr_inner, textvariable=status_var,
                 font=('Helvetica', 9), bg=_T['hdr_bg'],
                 fg=_T['text_muted']).pack(side='right')
        tk.Frame(root, bg=_T['border'], height=1).pack(fill='x')

        log_widget = scrolledtext.ScrolledText(
            root, font=('Courier', 9), bg='#0f1117', fg='#7ec8a4',
            insertbackground='white', wrap='word', state='disabled',
            relief='flat', bd=0,
        )
        log_widget.pack(fill='both', expand=True, padx=16, pady=(12, 4))
        log_widget.tag_config('err', foreground='#f87171')

        foot = _make_footer()

        def _append_log(text):
            log_widget.configure(state='normal')
            if 'ERROR' in text or '✗' in text:
                log_widget.insert('end', text + '\n', 'err')
            else:
                log_widget.insert('end', text + '\n')
            log_widget.see('end')
            log_widget.configure(state='disabled')

        # ── Mid-run dialog helper ─────────────────────────────────────────────
        def _make_dialog(title_text, body_text):
            """Create a clean modal dialog, returning (dlg, frame_for_inputs)."""
            dlg = tk.Toplevel(root)
            dlg.title(title_text)
            dlg.configure(bg=_T['bg'])
            dlg.grab_set()
            dlg.resizable(False, False)
            tk.Frame(dlg, bg=_T['border'], height=1).pack(fill='x')
            tk.Label(dlg, text=title_text,
                     font=('Helvetica', 12, 'bold'),
                     bg=_T['bg'], fg=_T['text']).pack(pady=(16, 4), padx=24)
            tk.Label(dlg, text=body_text,
                     font=('Helvetica', 9), bg=_T['bg'],
                     fg=_T['text_muted'], justify='center').pack(padx=24)
            inner = tk.Frame(dlg, bg=_T['bg'])
            inner.pack(pady=12, padx=24)
            return dlg, inner

        # ── Mid-run dialog: sample number verify ──────────────────────────────
        def _ask_cohort_create(missing, attempt, report):
            """
            Pause while the user creates missing cohorts in Climb.
            Responds 'recheck' or 'skip'.
            """
            again = ' (still missing)' if attempt > 1 else ''
            dlg, inner = _make_dialog(
                f'Cohorts need creating{again}',
                f'{len(missing)} cohort(s) do not exist in Climb yet.\n'
                'Cohorts cannot be created through the API — make them in Climb,\n'
                'then choose "I\'ve created them" to continue.'
            )

            # Scrollable list of cohorts to create
            box = tk.Frame(inner, bg=_T['bg_inset'], highlightthickness=1,
                           highlightbackground=_T['border'])
            box.grid(row=0, column=0, sticky='nsew', pady=(0, 6))

            hdr = tk.Frame(box, bg=_T['hdr_bg'])
            hdr.pack(fill='x')
            for txt, w in (('Cohort name', 22), ('Animals', 8), ('Description', 34)):
                tk.Label(hdr, text=txt, width=w, anchor='w',
                         font=('Helvetica', 9, 'bold'),
                         bg=_T['hdr_bg'], fg='white').pack(side='left', padx=4, pady=4)

            canvas = tk.Canvas(box, bg=_T['bg_inset'], highlightthickness=0,
                               height=min(220, 26 * max(len(missing), 1) + 10))
            rows_f = tk.Frame(canvas, bg=_T['bg_inset'])
            vsb    = ttk.Scrollbar(box, orient='vertical', command=canvas.yview)
            canvas.configure(yscrollcommand=vsb.set)
            canvas.pack(side='left', fill='both', expand=True)
            vsb.pack(side='right', fill='y')
            canvas.create_window((0, 0), window=rows_f, anchor='nw')

            for m in missing:
                r = tk.Frame(rows_f, bg=_T['bg_inset'])
                r.pack(fill='x')
                tk.Label(r, text=m['name'], width=22, anchor='w',
                         font=('Courier New', 10, 'bold'),
                         bg=_T['bg_inset'], fg=_T['accent']).pack(side='left', padx=4, pady=3)
                tk.Label(r, text=str(m['count']), width=8, anchor='w',
                         font=('Helvetica', 9),
                         bg=_T['bg_inset'], fg=_T['text']).pack(side='left', padx=4)
                tk.Label(r, text=m['description'], width=34, anchor='w',
                         font=('Helvetica', 9),
                         bg=_T['bg_inset'], fg=_T['text_muted']).pack(side='left', padx=4)

            rows_f.update_idletasks()
            canvas.configure(scrollregion=canvas.bbox('all'))

            tk.Label(inner,
                     text=f'Full list with every animal:  {report}',
                     font=('Helvetica', 8, 'italic'),
                     bg=_T['bg'], fg=_T['text_muted']).grid(row=1, column=0,
                                                            sticky='w')

            btns = tk.Frame(dlg, bg=_T['bg'])
            btns.pack(pady=(10, 16))

            def _respond(value):
                _gui_respond(value)
                dlg.destroy()

            _make_styled_button(btns, "I've created them  \u2192",
                                lambda: _respond('recheck'),
                                style='primary').pack(side='left', padx=6)
            _make_styled_button(btns, 'Skip cohort assignment',
                                lambda: _respond('skip'),
                                style='ghost').pack(side='left', padx=6)

            dlg.protocol('WM_DELETE_WINDOW', lambda: _respond('skip'))

            dlg.update_idletasks()
            x = root.winfo_x() + (root.winfo_width()  - dlg.winfo_width())  // 2
            y = root.winfo_y() + (root.winfo_height() - dlg.winfo_height()) // 2
            dlg.geometry(f'+{x}+{y}')
            root.wait_window(dlg)

        def _ask_sample_number_verify(fetched_num):
            dlg, inner = _make_dialog(
                'Verify sample number',
                f'Climb returned {fetched_num} as the next sample number.\n'
                'Confirm or enter a different number to override.'
            )
            tk.Label(inner, text='Next sample number:',
                     font=('Helvetica', 9), bg=_T['bg'],
                     fg=_T['text_muted']).grid(row=0, column=0, padx=(0, 8))
            var = tk.StringVar(value=str(fetched_num))
            e = ttk.Entry(inner, textvariable=var, width=10)
            e.grid(row=0, column=1)
            e.select_range(0, 'end')
            e.focus()

            err_lbl = tk.Label(dlg, text='', font=('Helvetica', 9),
                               bg=_T['bg'], fg=_T['red'])
            err_lbl.pack()

            def _ok():
                try:
                    val = int(var.get().strip())
                    if val <= 0:
                        raise ValueError
                    _gui_respond(val)
                    dlg.destroy()
                except ValueError:
                    err_lbl.configure(text='Enter a valid whole number.')

            _make_styled_button(dlg, 'Confirm', _ok,
                                style='primary').pack(pady=(6, 16))
            dlg.bind('<Return>', lambda e: _ok())

            dlg.update_idletasks()
            x = root.winfo_x() + (root.winfo_width()  - dlg.winfo_width())  // 2
            y = root.winfo_y() + (root.winfo_height() - dlg.winfo_height()) // 2
            dlg.geometry(f'+{x}+{y}')
            root.wait_window(dlg)
        def _ask_sample_number():
            dlg, inner = _make_dialog(
                'Sample number setup',
                'Enter the last sample number used.\nThe next run will start from that number + 1.'
            )
            tk.Label(inner, text='Last sample number used:',
                     font=('Helvetica', 9), bg=_T['bg'],
                     fg=_T['text_muted']).grid(row=0, column=0, padx=(0, 8))
            var = tk.StringVar()
            e = ttk.Entry(inner, textvariable=var, width=10)
            e.grid(row=0, column=1)
            e.focus()

            preview = tk.Label(dlg, text='', font=('Helvetica', 9, 'italic'),
                               bg=_T['bg'], fg=_T['accent'])
            preview.pack()
            err_lbl = tk.Label(dlg, text='', font=('Helvetica', 9),
                               bg=_T['bg'], fg=_T['red'])
            err_lbl.pack()

            def _update_preview(*_):
                try:
                    nxt = int(var.get()) + 1
                    preview.configure(text=f'Next sample will start at: {nxt}')
                    err_lbl.configure(text='')
                except ValueError:
                    preview.configure(text='')

            var.trace_add('write', _update_preview)

            def _ok():
                try:
                    nxt = int(var.get()) + 1
                    _gui_respond(nxt)
                    dlg.destroy()
                except ValueError:
                    err_lbl.configure(text='Please enter a valid whole number.')

            _make_styled_button(dlg, 'Confirm', _ok,
                                style='primary').pack(pady=(6, 16))
            dlg.bind('<Return>', lambda e: _ok())

            dlg.update_idletasks()
            x = root.winfo_x() + (root.winfo_width()  - dlg.winfo_width())  // 2
            y = root.winfo_y() + (root.winfo_height() - dlg.winfo_height()) // 2
            dlg.geometry(f'+{x}+{y}')
            root.wait_window(dlg)

        # ── Mid-run dialog: label offset ──────────────────────────────────────
        def _ask_label_offset(sheet_num, labels_remaining):
            dlg, inner = _make_dialog(
                f'Label sheet {sheet_num}',
                f'{labels_remaining} labels remaining to place.\n'
                f'How many slots are already used on this sheet?\n'
                f'(Enter 0 if the sheet is blank.)'
            )
            tk.Label(inner, text='Labels already used:',
                     font=('Helvetica', 9), bg=_T['bg'],
                     fg=_T['text_muted']).grid(row=0, column=0, padx=(0, 8))
            var = tk.StringVar(value='0')
            ttk.Spinbox(inner, from_=0, to=LABELS_PER_PAGE - 1,
                        textvariable=var, width=6).grid(row=0, column=1)

            err_lbl = tk.Label(dlg, text='', font=('Helvetica', 9),
                               bg=_T['bg'], fg=_T['red'])
            err_lbl.pack()

            def _ok():
                try:
                    n = int(var.get())
                    if 0 <= n < LABELS_PER_PAGE:
                        _gui_respond(n)
                        dlg.destroy()
                    else:
                        err_lbl.configure(text=f'Enter 0–{LABELS_PER_PAGE - 1}')
                except ValueError:
                    err_lbl.configure(text='Please enter a valid number.')

            _make_styled_button(dlg, 'Confirm', _ok,
                                style='primary').pack(pady=(6, 16))
            dlg.bind('<Return>', lambda e: _ok())

            dlg.update_idletasks()
            x = root.winfo_x() + (root.winfo_width()  - dlg.winfo_width())  // 2
            y = root.winfo_y() + (root.winfo_height() - dlg.winfo_height()) // 2
            dlg.geometry(f'+{x}+{y}')
            root.wait_window(dlg)

        # ── Mid-run dialog: label continue ────────────────────────────────────
        def _ask_label_continue(sheet_num):
            dlg, _ = _make_dialog(
                f'Ready for sheet {sheet_num}?',
                'Load the next label sheet into your printer, then click Continue.'
            )
            def _ok():
                _gui_respond(True)
                dlg.destroy()

            _make_styled_button(dlg, 'Continue  →', _ok,
                                style='primary').pack(pady=(12, 16))
            dlg.bind('<Return>', lambda e: _ok())

            dlg.update_idletasks()
            x = root.winfo_x() + (root.winfo_width()  - dlg.winfo_width())  // 2
            y = root.winfo_y() + (root.winfo_height() - dlg.winfo_height()) // 2
            dlg.geometry(f'+{x}+{y}')
            root.wait_window(dlg)

        # ── Poll queue from GUI thread ─────────────────────────────────────────
        pipeline_done = {'ok': False, 'error': None, 'result': None}

        def _poll():
            try:
                while True:
                    msg = _pipeline_queue.get_nowait()
                    if msg['kind'] == _MSG_LOG:
                        _append_log(msg['text'])
                    elif msg['kind'] == _MSG_DONE:
                        pipeline_done['ok']     = msg.get('ok', False)
                        pipeline_done['error']  = msg.get('error')
                        pipeline_done['result'] = msg.get('result')
                        if pipeline_done['ok']:
                            status_var.set('✓ Complete')
                            _switch(lambda: screen_summary(pipeline_done['result']))
                        else:
                            status_var.set('✗ Error — see log')
                            _append_log(f'\n✗ ERROR: {pipeline_done["error"]}')
                            _add_close_button()
                        return
                    elif msg['kind'] == _MSG_REQUEST:
                        rtype = msg['type']
                        if rtype == 'sample_number_verify':
                            _ask_sample_number_verify(msg.get('fetched', 1))
                        elif rtype == 'label_offset':
                            _ask_label_offset(msg['sheet_num'], msg['labels_remaining'])
                        elif rtype == 'label_continue':
                            _ask_label_continue(msg['sheet_num'])
                        elif rtype == 'cohort_create':
                            _ask_cohort_create(msg['missing'], msg['attempt'],
                                               msg['report'])
            except _queue.Empty:
                pass
            root.after(120, _poll)

        def _add_close_button():
            _make_styled_button(foot, 'Close', root.destroy,
                                style='secondary').pack(side='right', padx=8)

        # ── Pipeline thread ───────────────────────────────────────────────────
        def _run_pipeline():
            import sys as _sys
            old_stdout = _sys.stdout
            _sys.stdout = _QueueWriter()
            try:
                setup_logging(script_dir, CONFIG['LOG_LEVEL'])

                # ── Fetch data ────────────────────────────────────────────────
                selected_modules = state.get('modules', {})

                if selected_modules.get('schedule', False):
                    # Genotypes first — the animals pull below must reflect
                    # them, or newly typed animals still schedule as Blank.
                    if CONFIG.get('UPLOAD_TGS_GENOTYPES', True):
                        run_tgs_genotypes(script_dir)

                    # Schedule selected — pull animals and births from Climb
                    print('Fetching animals from Climb...')
                    try:
                        _sc = _load_sing_climb()
                        animals_climb_df = _sc.get_animals_df(verbose=False)
                        animals_path = _os.path.join(script_dir, CONFIG['INPUT_ANIMAL_FILE'])
                        animals_climb_df.to_csv(animals_path, index=False)
                        state['animal_file'] = animals_path
                        print(f'  \u2713 {len(animals_climb_df):,} animals fetched from Climb')
                    except Exception as _ex:
                        raise RuntimeError(
                            f'\u274c  Could not fetch animals from Climb: {_ex}\n'
                            'Check your Climb credentials and connection, then try again.'
                        )

                    print('Fetching births from Climb...')
                    try:
                        births_climb_df = _sc.get_births_df(verbose=False)
                        births_path = _os.path.join(script_dir, CONFIG['INPUT_BIRTHS_FILE'])
                        births_climb_df.to_csv(births_path, index=False)
                        state['births_file'] = births_path
                        print(f'  \u2713 {len(births_climb_df):,} births fetched from Climb')
                    except Exception as _ex:
                        raise RuntimeError(
                            f'\u274c  Could not fetch births from Climb: {_ex}\n'
                            'Check your Climb credentials and connection, then try again.'
                        )
                else:
                    # Partial run — use CSV from script folder
                    csv_path = state.get('animal_file',
                                         _os.path.join(script_dir, CONFIG['INPUT_ANIMAL_FILE']))
                    if not _os.path.exists(csv_path):
                        raise RuntimeError(
                            f'\u274c  Animal CSV not found:\n  {csv_path}\n'
                            f'Drop {CONFIG["INPUT_ANIMAL_FILE"]} in the script folder and try again.'
                        )
                    print(f'Using animal CSV: {_os.path.basename(csv_path)}')
                schedule_selected = selected_modules.get('schedule', False)

                if schedule_selected:
                    # ── Full pipeline ─────────────────────────────────────────
                    schedule_file, assignments_df = create_complete_schedule(
                        animal_file         = state['animal_file'],
                        tracking_file       = state.get('tracking_file'),
                        births_file         = state.get('births_file'),
                        output_dir          = script_dir,
                        birth_date_start    = None,
                        birth_date_end      = None,
                        behavior_date_start = None,
                        behavior_date_end   = None,
                        full_behavior_dates = state.get('full_behavior_dates'),
                    )

                    timestamp = __import__('datetime').datetime.now().strftime('%Y%m%d_%H%M%S')
                    output_files = [schedule_file]

                    if assignments_df is not None and not assignments_df.empty:
                        if 'Line' not in assignments_df.columns:
                            if 'Strain' in assignments_df.columns:
                                assignments_df['Line'] = assignments_df['Strain']
                            elif 'Line (Short)' in assignments_df.columns:
                                assignments_df['Line'] = assignments_df['Line (Short)']

                        working_df = build_working_data(assignments_df)

                        if not working_df.empty:
                            # Use first — quick and unattended. Cohorts can
                            # pause for input, so it goes second.
                            if CONFIG.get('UPDATE_ANIMAL_USE', True):
                                update_animal_use(working_df)

                            # Wild / Inconclusive animals can never be
                            # scheduled — release them from the Sing pool.
                            if CONFIG.get('RELEASE_UNUSABLE', True):
                                try:
                                    _pool = pd.read_csv(state['animal_file'])
                                    _pool, _ = filter_animals_by_use(_pool)
                                    release_unusable_to_available(_pool)
                                except Exception as _re:
                                    print(f'  \u26a0 Release skipped: {_re}')

                            # Cohorts — this pauses for input, and the other
                            # modules don't depend on it.
                            if CONFIG.get('ASSIGN_COHORTS', True):
                                try:
                                    run_cohorts(working_df, timestamp,
                                                output_dir=script_dir)
                                except Exception as _ce:
                                    print(f'  \u26a0 Cohort assignment failed: {_ce}')
                                    print('    Continuing with the rest of the run.')

                            if selected_modules.get('climb_samples', True):
                                harvest_df, samples_df, climb_import_df = run_harvest_and_samples(
                                    working_df, timestamp)
                            else:
                                samples_df = None

                            if selected_modules.get('deliverables', True) and samples_df is not None:
                                run_deliverables(working_df, samples_df, timestamp)

                            if selected_modules.get('envision', True):
                                envision_df = working_df[
                                    ~working_df.get('_is_nb', pd.Series(False, index=working_df.index)) &
                                    (working_df.get('Assigned_Timepoint', '') == 'P56')
                                ].copy()
                                if not envision_df.empty:
                                    run_climb_to_envision(envision_df, timestamp, output_dir=script_dir)
                                else:
                                    print('  \u24d8 No Envision output \u2014 no P56 non-NB animals scheduled.')

                            if selected_modules.get('labels', True) and samples_df is not None:
                                run_labels(samples_df, working_df, timestamp)

                    schedule_file_out = schedule_file

                else:
                    # ── Partial run — no scheduling ───────────────────────────
                    import datetime as _dt
                    timestamp = _dt.datetime.now().strftime('%Y%m%d_%H%M%S')
                    schedule_file_out = None
                    output_files = []

                    print(f'Loading CSV: {_os.path.basename(csv_path)}')
                    try:
                        animals_df = pd.read_csv(csv_path)
                        print(f'  {len(animals_df):,} rows loaded')
                    except Exception as _csv_ex:
                        raise RuntimeError(f'\u274c Could not read CSV: {_csv_ex}')

                    # Rename Name → Animal_Name if needed
                    if 'Name' in animals_df.columns and 'Animal_Name' not in animals_df.columns:
                        animals_df = animals_df.rename(columns={'Name': 'Animal_Name'})

                    if selected_modules.get('envision'):
                        print('\nRunning Climb \u2192 Envision translation...')
                        print(f'  animals_df shape: {animals_df.shape}')
                        print(f'  animals_df columns: {list(animals_df.columns)}')
                        print(f'  animals_df empty: {animals_df.empty}')
                        try:
                            envision_out = run_climb_to_envision(animals_df, timestamp, output_dir=script_dir)
                            print(f'  envision_out = {envision_out!r}')
                            if envision_out:
                                output_files.append(envision_out)
                                print(f'  Added to output_files: {envision_out}')
                            else:
                                print('  WARNING: run_climb_to_envision returned None or empty')
                        except Exception as _env_ex:
                            import traceback as _tb2
                            print(f'  ERROR in Envision: {_env_ex}')
                            print(_tb2.format_exc())

                    if selected_modules.get('climb_samples'):
                        print('\nCreating Climb samples...')
                        harvest_df, samples_df, climb_import_df = run_harvest_and_samples(
                            animals_df, timestamp)

                    if selected_modules.get('labels'):
                        print('\nGenerating labels...')
                        samples_df = locals().get('samples_df')
                        # Labels are built from sample records. If the Climb Samples
                        # module was not selected this pass, fall back to a samples.csv
                        # in the script folder rather than crashing on None.
                        if samples_df is None:
                            _s_path = _os.path.join(_SCRIPT_DIR, 'samples.csv')
                            if _os.path.exists(_s_path):
                                print(f'  Climb Samples did not run — reading {_os.path.basename(_s_path)}')
                                try:
                                    samples_df = pd.read_csv(_s_path, dtype=str)
                                    print(f'  {len(samples_df)} sample rows loaded')
                                except Exception as _sx:
                                    print(f'  Could not read samples.csv: {_sx}')
                                    samples_df = None
                            else:
                                print('  No samples.csv found in the script folder.')
                        run_labels(samples_df, animals_df, timestamp)

                    if selected_modules.get('deliverables'):
                        print('\nExporting deliverables sheet...')
                        import glob as _glob, datetime as _dt2
                        # Find the most recent Harvest_Sheet_Import file in the folder
                        imports = sorted(
                            _glob.glob(_os.path.join(script_dir, 'Harvest_Sheet_Import_*.xlsx')),
                            reverse=True
                        )
                        if imports:
                            print(f'  Using: {_os.path.basename(imports[0])}')
                            samples_df = pd.read_excel(imports[0], dtype=str).fillna('')
                        else:
                            # Fall back to Harvest Worksheet from Sing Harvest Sheet
                            harvest_xlsx = _os.path.join(script_dir, 'Sing Harvest Sheet.xlsx')
                            if _os.path.exists(harvest_xlsx):
                                print('  Using Harvest Worksheet from Sing Harvest Sheet.xlsx')
                                samples_df = pd.read_excel(
                                    harvest_xlsx, sheet_name='Harvest Worksheet', dtype=str
                                ).fillna('')
                            else:
                                samples_df = animals_df

                        # ── Scope to animals.csv ──────────────────────────────
                        # The Harvest Worksheet holds the whole project history.
                        # Only the animals in animals.csv belong in this export.
                        _name_col = next((c for c in ('Animal_Name', 'Name')
                                          if c in animals_df.columns), None)
                        _samp_col = next((c for c in ('Animal_Name', 'Name')
                                          if c in samples_df.columns), None)
                        if _name_col and _samp_col:
                            _wanted = set(
                                animals_df[_name_col].astype(str).str.strip()
                            )
                            _before = len(samples_df)
                            samples_df = samples_df[
                                samples_df[_samp_col].astype(str).str.strip().isin(_wanted)
                            ].copy()
                            print(f'  Scoped to animals.csv: {len(samples_df)} of '
                                  f'{_before} worksheet rows '
                                  f'({len(_wanted)} animals in CSV)')
                            if samples_df.empty:
                                print('  \u26a0 No worksheet rows matched animals.csv \u2014 '
                                      'nothing to export.')
                        else:
                            print(f'  \u26a0 Cannot scope to animals.csv \u2014 no name column. '
                                  f'CSV has: {list(animals_df.columns)[:8]} | '
                                  f'Worksheet has: {list(samples_df.columns)[:8]}')

                        run_deliverables(animals_df, samples_df, timestamp,
                                         output_dir=script_dir)

                    if selected_modules.get('sanity'):
                        print('\nRunning Sing Sanity...')
                        _run_sing_sanity(script_dir, timestamp)

                    # Collect all output files written during this run
                    new_files = sorted([
                        _os.path.join(script_dir, f)
                        for f in _os.listdir(script_dir)
                        if timestamp in f and _os.path.isfile(_os.path.join(script_dir, f))
                    ])

                new_files = sorted([
                    _os.path.join(script_dir, f)
                    for f in _os.listdir(script_dir)
                    if timestamp in f
                ])
                output_files.extend(new_files)

                _pipeline_queue.put({
                    'kind': _MSG_DONE, 'ok': True,
                    'result': {'schedule_file': schedule_file_out, 'output_files': output_files}
                })

            except Exception as ex:
                import traceback as _tb
                _tb.print_exc()
                _pipeline_queue.put({'kind': _MSG_DONE, 'ok': False, 'error': traceback.format_exc()})
            finally:
                _sys.stdout = old_stdout

        _append_log(f'TAILS v{PIPELINE_VERSION}\n')
        _append_log('Starting pipeline…\n')
        status_var.set('Running — please wait…')
        t = _threading.Thread(target=_run_pipeline, daemon=True)
        t.start()
        root.after(120, _poll)

    # ─────────────────────────────────────────────────────────────────────────
    # SCREEN 4: Summary
    # ─────────────────────────────────────────────────────────────────────────
    def screen_summary(result):
        root.title('TAILS — Complete')
        root.geometry('640x440')

        # Header with green accent on title
        hdr = tk.Frame(root, bg=_T['hdr_bg'], pady=16)
        hdr.pack(fill='x')
        tk.Frame(root, bg=_T['border'], height=1).pack(fill='x')
        title_row = tk.Frame(hdr, bg=_T['hdr_bg'])
        title_row.pack()
        tk.Label(title_row, text='✓',
                 font=('Helvetica', 14), bg=_T['hdr_bg'],
                 fg=_T['accent']).pack(side='left', padx=(0, 6))
        tk.Label(title_row, text='Pipeline complete',
                 font=('Helvetica', 16, 'bold'),
                 bg=_T['hdr_bg'], fg=_T['text']).pack(side='left')
        tk.Label(hdr, text='Output files saved to the script folder.',
                 font=('Helvetica', 10), bg=_T['hdr_bg'],
                 fg=_T['text_muted']).pack(pady=(2, 0))

        body = tk.Frame(root, bg=_T['bg'], padx=24, pady=14)
        body.pack(fill='both', expand=True)

        # Folder path strip
        path_strip = tk.Frame(body, bg=_T['bg_inset'], padx=10, pady=6)
        path_strip.pack(fill='x', pady=(0, 12))
        tk.Label(path_strip, text=script_dir,
                 font=('Courier', 8), bg=_T['bg_inset'],
                 fg=_T['text_muted'], anchor='w').pack(fill='x')

        # File list
        files = result.get('output_files', [])
        for fpath in files:
            name = _os.path.basename(fpath)
            try:
                size  = _os.path.getsize(fpath)
                size_str = f'{size:,} bytes'
            except Exception:
                size_str = ''

            row = tk.Frame(body, bg=_T['bg'], pady=5)
            row.pack(fill='x')
            tk.Frame(row, bg=_T['border'], height=1).pack(fill='x', pady=(0, 5))

            inner = tk.Frame(row, bg=_T['bg'])
            inner.pack(fill='x')

            # Checkmark
            tk.Label(inner, text='✓', font=('Helvetica', 10),
                     bg=_T['bg'], fg=_T['accent'],
                     width=2).pack(side='left')
            tk.Label(inner, text=name, font=('Helvetica', 9),
                     bg=_T['bg'], fg=_T['text'], anchor='w').pack(side='left', fill='x', expand=True)
            if size_str:
                tk.Label(inner, text=size_str, font=('Helvetica', 8),
                         bg=_T['bg'], fg=_T['text_faint']).pack(side='right')

        foot = _make_footer()

        def _run_again():
            for q in (_pipeline_queue, _response_queue):
                while not q.empty():
                    try:
                        q.get_nowait()
                    except _queue.Empty:
                        pass
            _switch(screen_file_setup)

        _make_styled_button(foot, 'Run again', _run_again,
                            style='secondary').pack(side='left', padx=8)
        _make_styled_button(foot, 'Close', root.destroy,
                            style='primary').pack(side='right', padx=8)

    # ── Start on screen 1 ────────────────────────────────────────────────────
    w = min(root.winfo_screenwidth() - 100, 720)
    h = min(root.winfo_screenheight() - 100, 660)
    x = (root.winfo_screenwidth()  - w) // 2
    y = (root.winfo_screenheight() - h) // 2
    root.geometry(f'{w}x{h}+{x}+{y}')
    root.minsize(560, 600)

    screen_module_select()
    root.mainloop()


if __name__ == "__main__":
    run_pipeline_gui()
