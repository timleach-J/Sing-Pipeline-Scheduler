# ============================================================
# Sing Pipeline
# Version 1.2
# ============================================================

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Side
import sys
import traceback
import glob
from datetime import datetime, timedelta, timezone, date
import warnings
import os
import re
import copy
from collections import defaultdict
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font
import unicodedata
import logging
from logging.handlers import RotatingFileHandler
from typing import Dict, List, Optional, Tuple, Union
from datetime import date as date_type
import unittest

warnings.filterwarnings('ignore')

# Try to import tqdm for progress bars
try:
    from tqdm import tqdm
    HAS_TQDM = True
except ImportError:
    HAS_TQDM = False
    class tqdm:
        def __init__(self, iterable=None, total=None, desc=None, **kwargs):
            self.iterable = iterable
            self.total = total or (len(iterable) if hasattr(iterable, '__len__') else None)
            self.desc = desc
            self.current = 0
            if desc and self.total:
                print(f"{desc}... (0/{self.total})")

        def __iter__(self):
            for item in self.iterable:
                self.current += 1
                if self.desc and self.total and self.total > 0 and self.current % max(1, self.total // 10) == 0:
                    print(f"{self.desc}... ({self.current}/{self.total})")
                yield item

        def __enter__(self):
            return self

        def __exit__(self, *args):
            if self.desc:
                print(f"{self.desc}... Complete!")

        def update(self, n=1):
            self.current += n

# ============================================================================
# CONFIGURATION
# ============================================================================
CONFIG = {
    'INPUT_ANIMAL_FILE': 'animals.csv',
    'INPUT_TRACKING_FILE': 'Sing Harvest Sheet - Summary Sheet.csv',
    'INPUT_BIRTHS_FILE': 'births.csv',
    'INPUT_OVERRIDES_FILE': 'harvest_overrides.csv',  # optional — leave blank rows to auto-assign

    'WEDNESDAY_CAPACITY': 18,
    'CAGE_SIZE': 3,
    'P14_VALID_DAYS': [0, 1, 2, 3, 4],  # Monday=0 through Friday=4
    'P56_BEHAVIOR_START_DAY': 42,
    'P56_BEHAVIOR_END_DAY': 49,
    'P56_BEHAVIOR_DAY_OF_WEEK': 2,  # Wednesday=2
    'P56_HARVEST_DAYS_AFTER_BEHAVIOR': 14,

    'SEXING_OFFSET_DAYS': 9,

    'B6_MIN_PER_MONTH': 3,
    'B6_STRAINS': ['B6J', 'B6NJ'],

    'DATE_VALIDATION': {
        'MAX_FUTURE_DAYS': 365,
        'MAX_PAST_DAYS': 730
    },

    'CHUNK_SIZE': 10000,
    'ENABLE_PROGRESS_BARS': True,

    'LOG_LEVEL': 'INFO',
    'LOG_TO_FILE': True,
    'LOG_TO_CONSOLE': True,

    'DEBUG_MODE': False,
    'RUN_TESTS': False,

    'HARVEST_TARGETS': {
        'Perfusion': 5,   # Base — actual target is 5 or 6 per sex (11 total per timepoint)
        'MERFISH': 1,
        'RNAseq': 1
    },

    'REQUIRED_ANIMAL_COLUMNS': [
        'Name', 'Birth Date', 'Sex', 'Line (Short)',
        'Genotype', 'Use', 'Status', 'Birth ID', 'Marker Type'
    ],

    'REQUIRED_BIRTHS_COLUMNS': [
        'Birth ID', 'Status', 'Birth Date', 'Live Count'
    ],

    'SUPER_PRIORITY_STRAINS': [
        'ARID1B', 'CACNA1G', 'CHD8', 'CNTNAP2', 'CTCF',
        'CTNNB1', 'DLL1', 'FMR1', 'GABRA1', 'KMT2C',
        'SCN2A', 'SHANK3', 'SMARCC2'
    ],

    'PRIORITY_STRAINS': {
        'AFF3': 'All', 'AP2S1': 'Half', 'ARID1B': 'Half', 'ASXL3': 'All',
        'ATP6V0A1': 'Half', 'AUTS2': 'Half', 'B6J': 'All', 'B6NJ': 'All',
        'BAP1': 'Half', 'BCL11B': 'Half', 'C3': 'All', 'CACNA1A': 'Half',
        'CACNA1C': 'Half', 'CACNA1G': 'Half', 'CAMK2B': 'Half', 'CASKIN1': 'All',
        'CDKL5': 'All', 'CERT1': 'Half', 'CHAMP1': 'Half', 'CHD2': 'All',
        'CHD8': 'Half', 'CNTNAP2': 'All', 'CTCF': 'All', 'CTNNB1': 'All',
        'CYFIP2': 'All', 'DDX23': 'Half', 'DEAF1': 'All', 'DHDDS': 'Half',
        'DLG4': 'All', 'DLL1': 'Half', 'DNMT3A': 'Half', 'DYRK1A': 'All',
        'EBF3': 'Half', 'EHMT1': 'Half', 'EIF5A': 'All', 'EP300': 'Half',
        'ERF': 'Half', 'FAM120A': 'All', 'FBN1': 'Half', 'FMR1': 'All',
        'FOXP1': 'Half', 'GABRA1': 'All', 'GABRG2': 'Half', 'GRIA2': 'Half',
        'GRIN2A': 'Half', 'GRIN2B': 'Half', 'GRN': 'All', 'HECW2': 'Half',
        'HERC1': 'All', 'IQSEC2': 'All', 'ITPR1': 'All', 'KAT6B': 'Half',
        'KBTBD7': 'All', 'KCNB1': 'All', 'KCND3': 'All', 'KCNMA1': 'All',
        'KCNT1': 'All', 'KCNT2': 'All', 'KDM5B': 'Half', 'KDM6B': 'Half',
        'KMT2C': 'All', 'KMT2E': 'Half', 'MAGEL2': 'Half', 'MECP2': 'Half',
        'MED13L': 'All', 'MED23': 'Half', 'MTOR': 'Half', 'MYT1L': 'Half',
        'NAA10': 'All', 'NALCN': 'Half', 'NFIX': 'All', 'NRXN1': 'All',
        'PACS2': 'All', 'PAH': 'All', 'PAX5': 'Half', 'POGZ': 'All',
        'POLR3B': 'Half', 'PREP': 'Half', 'PTEN': 'Half', 'PTPRD': 'Half',
        'RAC1': 'Half', 'RALA': 'Half', 'RB1CC1': 'Half', 'RBOBTB2': 'All',
        'RYR2': 'Half', 'SATB1': 'Half', 'SATB2': 'Half', 'SCN1A': 'Half',
        'SCN2A': 'Half', 'SETD1A': 'Half', 'SETD2': 'All', 'SETD5': 'Half',
        'SHANK3': 'Half', 'SLC6A1': 'Half', 'SMARCC2': 'Half', 'SMARCE1': 'Half',
        'SOX2': 'Half', 'SPAST': 'All', 'STXBP1': 'Half', 'SYNCRIP': 'Half',
        'SYNGAP1': 'Half', 'TAOK1': 'Half', 'TBR1': 'Half', 'TCF20': 'All',
        'TCF4': 'Half', 'TCF7L2': 'Half', 'TFAP4': 'All', 'TOP2B': 'Half',
        'TRIO': 'Half', 'U2AF2': 'Half', 'UBE3A': 'Half', 'VPS13B': 'All',
        'WAC': 'Half', 'XPO1': 'Half', 'ZBTB10': 'Half', 'ZBTB21': 'All',
        'ZFHX4': 'Half', 'ZMYM2': 'Half', 'ZNF292': 'All'
    }
}

DAYS_IN_WEEK = 7
P14_OFFSET_DAYS = 14
P56_HARVEST_OFFSET_FROM_BEHAVIOR = 14

# ============================================================================
# CUSTOM EXCEPTIONS
# ============================================================================

class SchedulerError(Exception):
    pass

class DataValidationError(SchedulerError):
    pass

class SchedulingError(SchedulerError):
    pass

# ============================================================================
# LOGGING SETUP
# ============================================================================

def setup_logging(output_dir: str, level: str = 'INFO') -> logging.Logger:
    log_dir = os.path.join(output_dir, 'logs')
    os.makedirs(log_dir, exist_ok=True)

    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    log_file = os.path.join(log_dir, f'scheduler_{timestamp}.log')

    formatter = logging.Formatter(
        '%(asctime)s - %(levelname)s - %(funcName)s:%(lineno)d - %(message)s',
        datefmt='%Y-%m-%d %H:%M:%S'
    )

    logger_instance = logging.getLogger()
    logger_instance.setLevel(getattr(logging, level.upper()))
    logger_instance.handlers.clear()

    if CONFIG['LOG_TO_FILE']:
        file_handler = RotatingFileHandler(log_file, maxBytes=10*1024*1024, backupCount=5)
        file_handler.setLevel(getattr(logging, level.upper()))
        file_handler.setFormatter(formatter)
        logger_instance.addHandler(file_handler)

    if CONFIG['LOG_TO_CONSOLE']:
        console_handler = logging.StreamHandler()
        console_handler.setLevel(logging.INFO)
        console_handler.setFormatter(formatter)
        logger_instance.addHandler(console_handler)

    logger_instance.info(f"Logging initialized: {log_file}")
    return logger_instance

logger = logging.getLogger(__name__)

# ============================================================================
# CONFIGURATION VALIDATION
# ============================================================================

def validate_config_advanced() -> bool:
    errors = []
    warnings_list = []

    if CONFIG['CAGE_SIZE'] <= 0:
        errors.append(f"CAGE_SIZE must be positive, got {CONFIG['CAGE_SIZE']}")
    elif CONFIG['CAGE_SIZE'] > 10:
        warnings_list.append(f"CAGE_SIZE ({CONFIG['CAGE_SIZE']}) is unusually large")

    if CONFIG['WEDNESDAY_CAPACITY'] <= 0:
        errors.append("WEDNESDAY_CAPACITY must be positive")
    elif CONFIG['WEDNESDAY_CAPACITY'] % CONFIG['CAGE_SIZE'] != 0:
        warnings_list.append(
            f"WEDNESDAY_CAPACITY ({CONFIG['WEDNESDAY_CAPACITY']}) is not a multiple of "
            f"CAGE_SIZE ({CONFIG['CAGE_SIZE']})"
        )

    for harvest_type, target in CONFIG['HARVEST_TARGETS'].items():
        if target < 0:
            errors.append(f"HARVEST_TARGETS['{harvest_type}'] must be non-negative")

    if not CONFIG['P14_VALID_DAYS']:
        errors.append("P14_VALID_DAYS cannot be empty")
    elif not all(0 <= day <= 6 for day in CONFIG['P14_VALID_DAYS']):
        errors.append("P14_VALID_DAYS must contain values 0-6")

    if CONFIG.get('SEXING_OFFSET_DAYS', 9) <= 0:
        errors.append("SEXING_OFFSET_DAYS must be positive")

    if CONFIG.get('B6_MIN_PER_MONTH', 3) < 0:
        errors.append("B6_MIN_PER_MONTH must be non-negative")

    priority_strains = set(s.upper() for s in CONFIG['PRIORITY_STRAINS'].keys())
    super_priority_strains = set(s.upper() for s in CONFIG['SUPER_PRIORITY_STRAINS'])
    missing_from_priority = super_priority_strains - priority_strains
    if missing_from_priority:
        warnings_list.append(
            f"SUPER_PRIORITY_STRAINS contains strains not in PRIORITY_STRAINS: "
            f"{', '.join(sorted(missing_from_priority))}"
        )

    if errors:
        error_msg = "Configuration errors:\n  - " + "\n  - ".join(errors)
        raise ValueError(error_msg)

    return True

try:
    validate_config_advanced()
except ValueError as e:
    print(f"❌ Configuration Error: {e}")
    raise

_PRIORITY_STRAINS_UPPER = {s.upper(): v for s, v in CONFIG['PRIORITY_STRAINS'].items()}
_SUPER_PRIORITY_STRAINS_UPPER = frozenset(s.upper() for s in CONFIG['SUPER_PRIORITY_STRAINS'])
_B6_STRAINS_UPPER = frozenset(s.upper() for s in CONFIG.get('B6_STRAINS', ['B6J', 'B6NJ']))

# ============================================================================
# CANONICAL GENOTYPE LABELS
# ============================================================================

GENOTYPE_WILD   = 'Wild'
GENOTYPE_HET    = 'Het'
GENOTYPE_HOM    = 'Hom'
GENOTYPE_HEMI   = 'Hemi'
GENOTYPE_INBRED = 'Inbred'
GENOTYPE_BLANK  = 'Blank'

_CANONICAL_GENOTYPES = frozenset([
    GENOTYPE_WILD, GENOTYPE_HET, GENOTYPE_HOM,
    GENOTYPE_HEMI, GENOTYPE_INBRED, GENOTYPE_BLANK
])

# ============================================================================
# UTILITY FUNCTIONS
# ============================================================================

def to_date(date_obj: Union[date_type, datetime, pd.Timestamp, None]) -> Optional[date_type]:
    if date_obj is None:
        return None
    if pd.isna(date_obj):
        return None
    if isinstance(date_obj, pd.Timestamp):
        return date_obj.date()
    if isinstance(date_obj, datetime):
        return date_obj.date()
    if isinstance(date_obj, date_type):
        return date_obj
    try:
        ts = pd.to_datetime(date_obj, errors='coerce')
        if pd.isna(ts):
            return None
        return ts.date()
    except Exception:
        return None


def normalize_genotype(genotype: str) -> str:
    if pd.isna(genotype):
        return genotype
    geno_str = str(genotype)
    replacements = {
        '‹': '<', '›': '>', '«': '<', '»': '>',
        '⟨': '<', '⟩': '>', '〈': '<', '〉': '>'
    }
    for old, new in replacements.items():
        geno_str = geno_str.replace(old, new)
    geno_str = unicodedata.normalize('NFKC', geno_str)
    geno_str = ' '.join(geno_str.split())
    return geno_str


def canonicalize_genotype(genotype, strain: str = '') -> str:
    """
    Normalize any genotype string to one of six canonical labels:
      Wild, Het, Hom, Hemi, Inbred, Blank
    """
    if isinstance(genotype, str) and genotype in _CANONICAL_GENOTYPES:
        return genotype

    if genotype is None:
        return GENOTYPE_BLANK
    try:
        if pd.isna(genotype):
            return GENOTYPE_BLANK
    except (TypeError, ValueError):
        pass
    geno_str = str(genotype).strip()
    if geno_str == '' or geno_str.lower() in ('nan', 'none', 'n/a', 'na', '-'):
        return GENOTYPE_BLANK

    if strain:
        strain_upper = str(strain).strip().upper()
        if strain_upper in _B6_STRAINS_UPPER:
            return GENOTYPE_INBRED

    geno_norm = normalize_genotype(geno_str)
    gl = geno_norm.lower()

    if any(kw in gl for kw in ('inconclusive', 'pending', 'regenotype',
                                'failed', 'no call', 'tbd')):
        return GENOTYPE_BLANK

    hemi_patterns = [
        r'hem[i]?', r'tg/\+', r'\+/tg', r'tg/-', r'-/y', r'[a-z]/y',
    ]
    if any(re.search(p, gl) for p in hemi_patterns):
        return GENOTYPE_HEMI

    wild_patterns = [
        r'\+/\+', r'\+/y', r'\bwt\b', r'\bwild.?type\b', r'\bwildtype\b',
        r'cre.ncar', r'generic.cre', r'cre \+/\+', r'cre \+/y',
    ]
    if any(re.search(p, gl) for p in wild_patterns):
        if not (re.search(r'-/\+', geno_norm) or re.search(r'\+/-', geno_norm)):
            return GENOTYPE_WILD

    hom_patterns = [
        r'-/-', r'\bhom\b', r'\bhomozygous\b', r'mut/mut', r'ko/ko',
    ]
    if any(re.search(p, gl) for p in hom_patterns):
        return GENOTYPE_HOM

    het_patterns = [
        r'-/\+', r'\+/-', r'\bhet\b', r'\bheterozygous\b', r'\bcarrier\b',
    ]
    if any(re.search(p, gl) for p in het_patterns):
        return GENOTYPE_HET

    if is_wildtype_cre_only(geno_norm):
        return GENOTYPE_WILD

    logger.debug(f"canonicalize_genotype: unrecognised genotype '{geno_str}' — returning Blank")
    return GENOTYPE_BLANK


def is_heterozygous(genotype: str) -> bool:
    if pd.isna(genotype):
        return False
    geno_str = str(genotype).strip()
    if geno_str == GENOTYPE_HET:
        return True
    if geno_str in _CANONICAL_GENOTYPES:
        return False
    if '-/+' in geno_str or '+/-' in geno_str:
        return True
    if 'HET' in geno_str.upper():
        return True
    return False


def is_wildtype_cre_only(genotype: str) -> bool:
    if pd.isna(genotype):
        return False
    geno_str = str(genotype).strip()
    if geno_str in _CANONICAL_GENOTYPES:
        return geno_str == GENOTYPE_WILD
    if geno_str == '':
        return False
    geno_lower = geno_str.lower()
    cre_patterns = ['cre ncar', 'cre-ncar', 'generic cre', 'cre +/+', 'cre +/y']
    has_cre_pattern = any(p in geno_lower for p in cre_patterns)
    if not has_cre_pattern:
        return False
    if is_heterozygous(geno_str):
        return False
    if '-/-' in geno_str:
        return False
    return True


def is_priority_strain(strain: str) -> bool:
    if pd.isna(strain):
        return False
    return str(strain).strip().upper() in _PRIORITY_STRAINS_UPPER


def is_super_priority_strain(strain: str) -> bool:
    if pd.isna(strain):
        return False
    return str(strain).strip().upper() in _SUPER_PRIORITY_STRAINS_UPPER


def is_b6_strain(strain: str) -> bool:
    if pd.isna(strain):
        return False
    return str(strain).strip().upper() in _B6_STRAINS_UPPER


def get_strain_breeding_type(strain: str, sex: Optional[str] = None) -> str:
    if pd.isna(strain):
        return 'Unknown'
    strain_upper = str(strain).strip().upper()
    return _PRIORITY_STRAINS_UPPER.get(strain_upper, 'Half')


def has_toe_clip(marker_type: str) -> bool:
    if pd.isna(marker_type):
        return False
    return 'Toe Clip' in str(marker_type)


def parse_date(date_str: str) -> Optional[date_type]:
    if not date_str or date_str.strip() == '':
        return None
    try:
        return datetime.strptime(date_str.strip(), '%Y-%m-%d').date()
    except ValueError as e:
        logger.error(f"Invalid date format '{date_str}': {e}")
        return None


def parse_multiple_dates(input_str: str) -> List[date_type]:
    if not input_str or input_str.strip() == '':
        return []
    dates = []
    for date_str in input_str.split(','):
        date_obj = parse_date(date_str.strip())
        if date_obj:
            dates.append(date_obj)
    return dates


def is_valid_p14_day(date_obj: date_type) -> bool:
    date_obj = to_date(date_obj)
    if date_obj is None:
        return False
    return date_obj.weekday() in CONFIG['P14_VALID_DAYS']


def next_wednesday(target_date: date_type) -> Optional[date_type]:
    target_date = to_date(target_date)
    if target_date is None:
        return None
    days_ahead = CONFIG['P56_BEHAVIOR_DAY_OF_WEEK'] - target_date.weekday()
    if days_ahead < 0:
        days_ahead += DAYS_IN_WEEK
    return target_date + timedelta(days=days_ahead)


def calculate_schedule_dates(birth_date: Union[date_type, datetime, pd.Timestamp]) -> Optional[Dict[str, date_type]]:
    birth_date = to_date(birth_date)
    if birth_date is None:
        return None

    today = datetime.now().date()
    max_future = today + timedelta(days=CONFIG['DATE_VALIDATION']['MAX_FUTURE_DAYS'])

    if birth_date > max_future:
        return None

    try:
        p14_harvest = birth_date + timedelta(days=P14_OFFSET_DAYS)
        behavior_start_min = birth_date + timedelta(days=CONFIG['P56_BEHAVIOR_START_DAY'])
        behavior_start_max = birth_date + timedelta(days=CONFIG['P56_BEHAVIOR_END_DAY'])
        sexing_date = birth_date + timedelta(days=CONFIG['SEXING_OFFSET_DAYS'])

        return {
            'birth_date': birth_date,
            'sexing_date': sexing_date,
            'p14_harvest': p14_harvest,
            'p56_behavior_window_start': behavior_start_min,
            'p56_behavior_window_end': behavior_start_max,
        }
    except (OverflowError, ValueError) as e:
        logger.warning(f"Error calculating dates for birth {birth_date}: {e}")
        return None


def get_next_wednesdays(n: int = 6, from_date: Optional[date_type] = None) -> List[date_type]:
    if from_date is None:
        from_date = datetime.now().date()

    wednesdays = []
    current = next_wednesday(from_date)

    if current == from_date:
        wednesdays.append(current)
        current = current + timedelta(days=DAYS_IN_WEEK)
    else:
        wednesdays.append(current)
        current = current + timedelta(days=DAYS_IN_WEEK)

    while len(wednesdays) < n:
        wednesdays.append(current)
        current = current + timedelta(days=DAYS_IN_WEEK)

    return wednesdays


def get_p56_behavior_wednesday(birth_date_obj: Optional[date_type]) -> Optional[date_type]:
    """
    Return the first Wednesday that falls in the P42-P49 window for a given
    birth date, or None if the birth date is invalid / window has no Wednesday.

    This is used to group blank-genotype animals that share the same behavior
    session for P56 worth-it assessment.
    """
    if birth_date_obj is None:
        return None
    dates = calculate_schedule_dates(birth_date_obj)
    if dates is None:
        return None
    first_wed = next_wednesday(dates['p56_behavior_window_start'])
    if first_wed is None:
        return None
    if first_wed > dates['p56_behavior_window_end']:
        return None  # No Wednesday falls inside P42-P49
    return first_wed


def prompt_wednesday_capacity() -> Tuple[List[date_type], Optional[List[date_type]]]:
    wednesdays = get_next_wednesdays(6)
    capacity = CONFIG['WEDNESDAY_CAPACITY']

    print("\n" + "=" * 70)
    print("WEDNESDAY P56 BEHAVIOR CAPACITY CHECK")
    print("=" * 70)
    print(f"Maximum capacity per Wednesday: {capacity} animals")
    print(f"Enter how many animals are ALREADY SCHEDULED for each Wednesday.")
    print(f"Press Enter to skip (assumes 0 scheduled).\n")

    scheduled_counts = {}
    full_dates = []

    for i, wed in enumerate(wednesdays, 1):
        day_label = wed.strftime('%A, %Y-%m-%d')
        while True:
            raw = input(f"  {i}. {day_label} — Already scheduled: ").strip()
            if raw == '':
                scheduled_counts[wed] = 0
                break
            try:
                count = int(raw)
                if count < 0:
                    print(f"     ⚠️  Please enter a number >= 0")
                    continue
                if count > capacity:
                    print(f"     ⚠️  That's already over capacity ({capacity})!")
                scheduled_counts[wed] = count
                break
            except ValueError:
                print(f"     ⚠️  Please enter a whole number (e.g., 0, 6, 18)")

    print(f"\n  {'Wednesday':<28} {'Scheduled':>10} {'Remaining':>10} {'Status':>12}")
    print(f"  {'-'*28} {'-'*10} {'-'*10} {'-'*12}")

    for wed in wednesdays:
        count = scheduled_counts[wed]
        remaining = capacity - count
        if remaining <= 0:
            status = '🔴 FULL'
            full_dates.append(wed)
        elif remaining <= 3:
            status = '🟡 LOW'
        else:
            status = '🟢 OPEN'
        print(f"  {wed.strftime('%A, %Y-%m-%d'):<28} {count:>10} {remaining:>10} {status:>12}")

    if full_dates:
        print(f"\n  ⚠️  {len(full_dates)} Wednesday(s) at capacity")
    else:
        print(f"\n  ✓ All Wednesdays have available capacity")

    return wednesdays, full_dates if full_dates else None


def auto_size_columns(worksheet):
    for column in worksheet.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if cell.value:
                    cell_length = len(str(cell.value))
                    if cell_length > max_length:
                        max_length = cell_length
            except Exception:
                pass
        adjusted_width = min(max_length + 2, 60)
        worksheet.column_dimensions[column_letter].width = adjusted_width


def truncate_sheet_name(name: str, max_length: int = 31) -> str:
    if len(name) <= max_length:
        return name
    return name[:max_length - 3] + '...'


# ============================================================================
# FILE READING AND VALIDATION
# ============================================================================

def validate_animal_file(df: pd.DataFrame) -> bool:
    missing = [col for col in CONFIG['REQUIRED_ANIMAL_COLUMNS'] if col not in df.columns]
    if missing:
        raise DataValidationError(f"Missing required columns in animal file: {missing}")
    return True


def validate_births_file(df: pd.DataFrame) -> bool:
    core_required = ['Birth ID', 'Status', 'Birth Date']
    missing_core = [col for col in core_required if col not in df.columns]
    if missing_core:
        raise DataValidationError(f"Missing required columns in births file: {missing_core}")
    return True


def process_large_dataset(animal_file: str, chunk_size: int = None) -> pd.DataFrame:
    if chunk_size is None:
        chunk_size = CONFIG['CHUNK_SIZE']
    try:
        total_rows = sum(1 for _ in open(animal_file)) - 1
    except Exception:
        return pd.read_csv(animal_file)

    if total_rows < chunk_size:
        return pd.read_csv(animal_file)

    chunks = []
    for chunk in pd.read_csv(animal_file, chunksize=chunk_size):
        chunks.append(chunk[chunk['Status'] == 'Alive'].copy())
    return pd.concat(chunks, ignore_index=True)


def read_animal_data(filename: str) -> pd.DataFrame:
    if not os.path.exists(filename):
        raise FileNotFoundError(f"Animal file not found: {filename}")
    df = process_large_dataset(filename)
    validate_animal_file(df)
    df['Birth Date'] = pd.to_datetime(df['Birth Date'], errors='coerce')
    df = df[df['Status'] == 'Alive'].copy()
    df['Birth ID'] = df['Birth ID'].astype(str)

    df['Genotype'] = df['Genotype'].apply(normalize_genotype)
    df['Genotype'] = df.apply(
        lambda row: canonicalize_genotype(row['Genotype'], row.get('Line (Short)', '')),
        axis=1
    )

    logger.info(f"Loaded {len(df)} alive animals")
    logger.info(f"Genotype breakdown:\n{df['Genotype'].value_counts().to_string()}")
    return df


def read_births_data(filename: str) -> Optional[pd.DataFrame]:
    if filename is None or not os.path.exists(filename):
        logger.warning(f"Births file not found: {filename}")
        return None
    try:
        df = pd.read_csv(filename)
    except Exception as e:
        logger.warning(f"Error reading births file: {e}")
        return None
    try:
        validate_births_file(df)
    except DataValidationError as e:
        logger.warning(f"{e}")
        return None
    df['Birth Date'] = pd.to_datetime(df['Birth Date'], errors='coerce')
    df['Birth ID'] = df['Birth ID'].astype(str)
    # Births CSV has 'Line' (full name) but not 'Line (Short)' — use 'Line' as fallback
    if 'Line (Short)' not in df.columns and 'Line' in df.columns:
        df['Line (Short)'] = df['Line']
        logger.info("Births: 'Line (Short)' not found — using 'Line' as fallback")
    logger.info(f"Loaded {len(df)} birth records")
    return df


def read_tracking_data(filename: str) -> Optional[pd.DataFrame]:
    if filename is None or not os.path.exists(filename):
        logger.warning(f"Tracking file not found: {filename}")
        return None
    try:
        df = pd.read_csv(filename)
        logger.info(f"Loaded tracking file: {len(df)} rows")
        return df
    except Exception as e:
        logger.warning(f"Error reading tracking file: {e}")
        return None


# ============================================================================
# DIAGNOSTIC HELPER
# ============================================================================

def diagnose_animal_file(df: pd.DataFrame) -> None:
    print("\n" + "=" * 70)
    print("DIAGNOSTIC: ANIMAL FILE CONTENTS")
    print("=" * 70)
    print(f"  Total rows loaded:     {len(df):,}")
    print(f"  Total columns:         {len(df.columns)}")
    print(f"\n  Column names:")
    for col in df.columns.tolist():
        print(f"    - {repr(col)}")

    if 'Status' in df.columns:
        print(f"\n  'Status' value counts:")
        for val, cnt in df['Status'].value_counts(dropna=False).items():
            print(f"    {repr(val)}: {cnt}")
        alive_count = len(df[df['Status'] == 'Alive'])
        print(f"\n  Animals with Status == 'Alive': {alive_count}")
    else:
        print("\n  ⚠️  'Status' column NOT FOUND")

    if 'Use' in df.columns:
        print(f"\n  'Use' value counts (top 10):")
        for val, cnt in df['Use'].value_counts(dropna=False).head(10).items():
            print(f"    {repr(val)}: {cnt}")
        sing_mask = df['Use'].str.contains('Sing Inventory', na=False, case=False)
        print(f"\n  Animals matching 'Sing Inventory' in Use: {sing_mask.sum()}")
    else:
        print("\n  ⚠️  'Use' column NOT FOUND")

    if 'Genotype' in df.columns:
        print(f"\n  'Genotype' canonical value counts:")
        for val, cnt in df['Genotype'].value_counts(dropna=False).items():
            print(f"    {repr(val)}: {cnt}")
    else:
        print("\n  ⚠️  'Genotype' column NOT FOUND")

    if 'Line (Short)' in df.columns:
        print(f"\n  'Line (Short)' (strain) value counts (top 15):")
        for val, cnt in df['Line (Short)'].value_counts(dropna=False).head(15).items():
            print(f"    {repr(val)}: {cnt}")
    else:
        print("\n  ⚠️  'Line (Short)' column NOT FOUND")

    print("=" * 70 + "\n")


# ============================================================================
# BIRTHS ANALYSIS
# ============================================================================

def calculate_sexing_date(birth_date: Union[date_type, datetime, pd.Timestamp]) -> Optional[date_type]:
    bd = to_date(birth_date)
    if bd is None:
        return None
    return bd + timedelta(days=CONFIG['SEXING_OFFSET_DAYS'])


def build_births_sexing_schedule(
    births_df: pd.DataFrame,
    animals_df: Optional[pd.DataFrame] = None
) -> pd.DataFrame:
    """
    Build a sexing schedule for births that have not yet been sexed.
    Any Birth ID that already has animals in animals_df is excluded.
    """
    if births_df is None or len(births_df) == 0:
        return pd.DataFrame()

    today = datetime.now().date()

    sing = births_df[
        births_df['Status'].str.contains('Sing Inventory', na=False, case=False)
    ].copy()

    if len(sing) == 0:
        return pd.DataFrame()

    already_sexed_birth_ids = set()
    if animals_df is not None and len(animals_df) > 0 and 'Birth ID' in animals_df.columns:
        already_sexed_birth_ids = set(
            animals_df['Birth ID'].astype(str).unique()
        )
        logger.info(
            f"build_births_sexing_schedule: {len(already_sexed_birth_ids)} "
            f"Birth IDs already have animals entered (already sexed)"
        )

    rows = []
    skipped_already_sexed = 0

    for _, birth in sing.iterrows():
        birth_id = str(birth.get('Birth ID', 'N/A'))

        if birth_id in already_sexed_birth_ids:
            skipped_already_sexed += 1
            continue

        birth_date_obj = to_date(birth['Birth Date'])
        strain = birth.get('Line (Short)', 'N/A')
        dam = birth.get('Dam', 'N/A')
        sire = birth.get('Sire', 'N/A')
        num_pups = birth.get('# of Pups', birth.get('Live Count', 'N/A'))

        if birth_date_obj is None:
            rows.append({
                'Birth_ID': birth_id,
                'Strain': strain if pd.notna(strain) else 'N/A',
                'Dam': dam if pd.notna(dam) else 'N/A',
                'Sire': sire if pd.notna(sire) else 'N/A',
                'Birth_Date': 'N/A',
                'Num_Pups': num_pups if pd.notna(num_pups) else 'N/A',
                'Sexing_Date': 'N/A',
                'Day_of_Week': 'N/A',
                'Days_Until_Sexing': 'N/A',
                'Sexing_Status': '❓ Unknown — No birth date',
                'P14_Expected_Date': 'N/A',
                'P14_Day_of_Week': 'N/A',
            })
            continue

        sexing_date = birth_date_obj + timedelta(days=CONFIG['SEXING_OFFSET_DAYS'])
        p14_date = birth_date_obj + timedelta(days=P14_OFFSET_DAYS)
        days_until = (sexing_date - today).days

        if days_until < 0:
            status = f'✅ Done (was {sexing_date.strftime("%Y-%m-%d")})'
        elif days_until == 0:
            status = '🔴 TODAY — Sex pups now!'
        elif days_until == 1:
            status = '🟠 TOMORROW — Prepare'
        elif days_until <= 3:
            status = f'🟡 SOON — {days_until} days'
        else:
            status = f'🟢 Upcoming — {days_until} days'

        rows.append({
            'Birth_ID': birth_id,
            'Strain': strain if pd.notna(strain) else 'N/A',
            'Dam': dam if pd.notna(dam) else 'N/A',
            'Sire': sire if pd.notna(sire) else 'N/A',
            'Birth_Date': birth_date_obj.strftime('%Y-%m-%d'),
            'Num_Pups': int(num_pups) if pd.notna(num_pups) else 'N/A',
            'Sexing_Date': sexing_date.strftime('%Y-%m-%d'),
            'Day_of_Week': sexing_date.strftime('%A'),
            'Days_Until_Sexing': days_until,
            'Sexing_Status': status,
            'P14_Expected_Date': p14_date.strftime('%Y-%m-%d'),
            'P14_Day_of_Week': p14_date.strftime('%A'),
        })

    if skipped_already_sexed > 0:
        logger.info(
            f"build_births_sexing_schedule: skipped {skipped_already_sexed} "
            f"births already sexed (animals entered in system)"
        )

    df = pd.DataFrame(rows)
    if len(df) == 0:
        return df

    def sort_key(row):
        val = row['Days_Until_Sexing']
        if isinstance(val, int):
            return (0 if val >= 0 else 1, val if val >= 0 else -val)
        return (2, 0)

    df = df.iloc[pd.Series(range(len(df))).apply(lambda i: sort_key(df.iloc[i])).argsort()]
    df = df.reset_index(drop=True)
    return df


def analyze_birth_scheduling_potential(birth: pd.Series, requirements: Dict,
                                       remaining_needs: Dict, today: date_type) -> Dict:
    birth_date = birth['Birth Date']
    if pd.isna(birth_date):
        return {
            'P14_Potential': 'Unknown', 'P14_Reason': 'No birth date',
            'P14_Expected_Date': 'N/A', 'P14_Day_of_Week': 'N/A',
            'P56_Potential': 'Unknown', 'P56_Reason': 'No birth date',
            'P56_Expected_Behavior_Date': 'N/A', 'P56_Expected_Harvest_Date': 'N/A',
            'Quota_Status': 'Unknown', 'Priority_Strain': 'Unknown', 'Age_Today_Days': 'N/A',
            'Sexing_Date': 'N/A', 'Sexing_Day_of_Week': 'N/A',
        }

    birth_date_obj = to_date(birth_date)
    strain = birth.get('Line (Short)', '')

    if birth_date_obj is None:
        return {
            'P14_Potential': 'Unknown', 'P14_Reason': 'Invalid birth date',
            'P14_Expected_Date': 'N/A', 'P14_Day_of_Week': 'N/A',
            'P56_Potential': 'Unknown', 'P56_Reason': 'Invalid birth date',
            'P56_Expected_Behavior_Date': 'N/A', 'P56_Expected_Harvest_Date': 'N/A',
            'Quota_Status': 'Unknown',
            'Priority_Strain': 'YES' if is_priority_strain(strain) else 'No',
            'Age_Today_Days': 'N/A', 'Sexing_Date': 'N/A', 'Sexing_Day_of_Week': 'N/A',
        }

    dates = calculate_schedule_dates(birth_date_obj)

    if dates is None:
        return {
            'P14_Potential': 'Unknown', 'P14_Reason': 'Invalid birth date',
            'P14_Expected_Date': 'N/A', 'P14_Day_of_Week': 'N/A',
            'P56_Potential': 'Unknown', 'P56_Reason': 'Invalid birth date',
            'P56_Expected_Behavior_Date': 'N/A', 'P56_Expected_Harvest_Date': 'N/A',
            'Quota_Status': 'Unknown',
            'Priority_Strain': 'YES' if is_priority_strain(strain) else 'No',
            'Age_Today_Days': 'N/A', 'Sexing_Date': 'N/A', 'Sexing_Day_of_Week': 'N/A',
        }

    p14_harvest = dates['p14_harvest']
    behavior_window_start = dates['p56_behavior_window_start']
    behavior_window_end = dates['p56_behavior_window_end']
    sexing_date = dates['sexing_date']

    p14_valid = is_valid_p14_day(p14_harvest)
    p14_in_future = p14_harvest > today

    if not p14_in_future:
        p14_potential = 'Past'
        if p14_harvest == today:
            p14_reason = f'P14 date is today ({p14_harvest}) — too late to schedule'
        else:
            p14_reason = f'P14 date ({p14_harvest}) has passed'
    elif not p14_valid:
        p14_potential = 'No'
        p14_reason = f'P14 falls on {p14_harvest.strftime("%A")} (invalid day)'
    else:
        p14_potential = 'Yes'
        p14_reason = f'Could schedule on {p14_harvest.strftime("%A, %Y-%m-%d")}'

    first_wednesday = next_wednesday(behavior_window_start)
    p56_harvest_date = None

    if first_wednesday is None:
        p56_potential = 'No'
        p56_reason = 'Cannot calculate P56 behavior date'
    elif first_wednesday > behavior_window_end:
        p56_potential = 'No'
        p56_reason = 'No Wednesday in P42-49 window'
    elif first_wednesday < today:
        p56_potential = 'Past'
        p56_reason = f'P56 window ({first_wednesday}) has passed'
    else:
        p56_potential = 'Yes'
        p56_reason = f'Could schedule behavior on {first_wednesday.strftime("%A, %Y-%m-%d")}'
        p56_harvest_date = first_wednesday + timedelta(days=P56_HARVEST_OFFSET_FROM_BEHAVIOR)

    quota_status = 'Unknown'
    quota_details = []
    is_priority = is_priority_strain(strain)

    if remaining_needs and pd.notna(strain):
        strain_key = str(strain).strip().upper()
        if strain_key in remaining_needs:
            p14_needs = remaining_needs[strain_key]['P14']
            p56_needs = remaining_needs[strain_key]['P56']
            for timepoint, needs_dict in [('P14', p14_needs), ('P56', p56_needs)]:
                for sex in ['Male', 'Female']:
                    total = sum(needs_dict[sex][ht]['needed'] for ht in ['Perfusion', 'MERFISH', 'RNAseq'])
                    if total > 0:
                        quota_details.append(f"{timepoint} {sex}: {total} needed")
            quota_status = 'NEEDED - ' + '; '.join(quota_details) if quota_details else 'Quota Complete'
        else:
            quota_status = 'Not tracked in requirements'

    age_days = (today - birth_date_obj).days

    return {
        'P14_Potential': p14_potential,
        'P14_Reason': p14_reason,
        'P14_Expected_Date': p14_harvest.strftime('%Y-%m-%d'),
        'P14_Day_of_Week': p14_harvest.strftime('%A'),
        'P56_Potential': p56_potential,
        'P56_Reason': p56_reason,
        'P56_Expected_Behavior_Date': first_wednesday.strftime('%Y-%m-%d') if p56_potential != 'No' and first_wednesday else 'N/A',
        'P56_Expected_Harvest_Date': p56_harvest_date.strftime('%Y-%m-%d') if p56_harvest_date else 'N/A',
        'Quota_Status': quota_status,
        'Priority_Strain': 'YES' if is_priority else 'No',
        'Age_Today_Days': age_days,
        'Sexing_Date': sexing_date.strftime('%Y-%m-%d') if sexing_date else 'N/A',
        'Sexing_Day_of_Week': sexing_date.strftime('%A') if sexing_date else 'N/A',
    }


def estimate_expected_animals(birth: pd.Series) -> Dict:
    num_pups = None
    count_source = 'No count field'

    if 'Live Count' in birth.index:
        num_pups = birth.get('Live Count', None)
        if pd.notna(num_pups):
            count_source = 'Live Count'

    if num_pups is None and '# of Pups' in birth.index:
        num_pups = birth.get('# of Pups', None)
        if pd.notna(num_pups):
            count_source = '# of Pups'

    strain = birth.get('Line (Short)', '')
    breeding_type = get_strain_breeding_type(strain)

    if pd.isna(num_pups):
        return {
            'Expected_Total_Born': 'Unknown', 'Expected_Usable': 'Unknown',
            'Expected_Usable_Males': 'Unknown', 'Expected_Usable_Females': 'Unknown',
            'Breeding_Type': breeding_type,
            'Estimation_Note': 'No pup count in birth record'
        }

    try:
        total_pups = int(num_pups)
    except (ValueError, TypeError):
        return {
            'Expected_Total_Born': 'Unknown', 'Expected_Usable': 'Unknown',
            'Expected_Usable_Males': 'Unknown', 'Expected_Usable_Females': 'Unknown',
            'Breeding_Type': breeding_type,
            'Estimation_Note': f'Invalid pup count: {num_pups}'
        }

    if breeding_type == 'Half':
        expected_usable = total_pups // 2
        expected_usable_males = expected_usable // 2
        expected_usable_females = expected_usable - expected_usable_males
        note = f'Het x WT: ~50% usable ({expected_usable} of {total_pups}) [from {count_source}]'
    elif breeding_type == 'All':
        expected_usable = total_pups
        expected_usable_males = total_pups // 2
        expected_usable_females = total_pups - expected_usable_males
        note = f'Hom x Hom/Inbred: All usable ({expected_usable} of {total_pups}) [from {count_source}]'
    else:
        expected_usable = total_pups // 2
        expected_usable_males = expected_usable // 2
        expected_usable_females = expected_usable - expected_usable_males
        note = f'Unknown strain: Assuming Het x WT (~50%) [from {count_source}]'

    return {
        'Expected_Total_Born': total_pups,
        'Expected_Usable': expected_usable,
        'Expected_Usable_Males': f'~{expected_usable_males}',
        'Expected_Usable_Females': f'~{expected_usable_females}',
        'Breeding_Type': breeding_type,
        'Estimation_Note': note
    }


def determine_action_required(potential: Dict, expectations: Dict, age_days) -> str:
    actions = []
    if potential['P14_Potential'] == 'Yes' or potential['P56_Potential'] == 'Yes':
        actions.append('🔍 VERIFY animals exist and have correct Birth ID')
        if potential['Quota_Status'].startswith('NEEDED'):
            actions.append('⚠️ URGENT: Quota needs exist - locate animals immediately')
    if age_days is not None and age_days != 'N/A':
        if age_days > 56:
            actions.append('❌ Too old for P56 - consider P14 retrospective or exclude')
        elif age_days > 49:
            actions.append('⏰ P56 window closing - urgent genotyping needed')
        elif age_days >= 42:
            actions.append('📋 P56 window open - genotype and schedule behavior')
        elif age_days > 14:
            actions.append('⏰ P14 window passed - plan for P56')
        elif age_days >= 10:
            actions.append('📋 Genotype for P14 scheduling')
        else:
            actions.append('⏳ Monitor - too young for scheduling')
    if expectations.get('Expected_Total_Born') == 0:
        actions.append('ℹ️ Birth shows 0 pups - verify and update status')
    if not actions:
        actions.append('📧 Contact lab manager for clarification')
    return ' | '.join(actions)


def find_unmatched_births_enhanced(births_df: Optional[pd.DataFrame], animals_df: pd.DataFrame,
                                    requirements: Dict, remaining_needs: Dict) -> pd.DataFrame:
    if births_df is None or len(births_df) == 0:
        return pd.DataFrame()

    today = datetime.now().date()
    logger.info("Analyzing unmatched births...")

    sing_inventory_births = births_df[
        births_df['Status'].str.contains('Sing Inventory', na=False, case=False)
    ].copy()

    if len(sing_inventory_births) == 0:
        return pd.DataFrame()

    animal_birth_ids = set(animals_df['Birth ID'].astype(str).unique())
    unmatched_births = []

    for idx, birth in sing_inventory_births.iterrows():
        birth_id = str(birth['Birth ID'])
        if birth_id == 'nan' or birth_id.strip() == '':
            continue
        if birth_id not in animal_birth_ids:
            birth_date = birth['Birth Date']
            birth_date_str = to_date(birth_date).strftime('%Y-%m-%d') if pd.notna(birth_date) else 'N/A'
            strain = birth.get('Line (Short)', 'N/A')
            dam = birth.get('Dam', 'N/A')
            sire = birth.get('Sire', 'N/A')
            num_pups = birth.get('# of Pups', 'N/A')

            potential = analyze_birth_scheduling_potential(birth, requirements, remaining_needs, today)
            expectations = estimate_expected_animals(birth)
            age_days = potential.get('Age_Today_Days', 'N/A')

            if age_days != 'N/A':
                if age_days > 56:
                    urgency = '🔴 URGENT - Past P56'
                elif age_days > 42:
                    urgency = '🟡 HIGH - In P56 window'
                elif age_days > 14:
                    urgency = '🟢 MEDIUM - Past P14'
                elif age_days >= 10:
                    urgency = '🟢 LOW - Approaching P14'
                else:
                    urgency = '⚪ INFO - Too young'
            else:
                urgency = '❓ UNKNOWN - No birth date'

            possible_reasons = []
            if pd.notna(num_pups) and num_pups == 0:
                possible_reasons.append('Birth record shows 0 pups')
            elif pd.notna(birth_date) and age_days != 'N/A' and age_days < 5:
                possible_reasons.append('Birth too recent - animals may not be entered yet')
            else:
                possible_reasons.append('Animals not found/entered in Climb')
                possible_reasons.append('Animals may have been culled')
                possible_reasons.append('Birth ID mismatch possible')

            unmatched_births.append({
                'Urgency': urgency,
                'Birth_ID': birth_id,
                'Birth_Date': birth_date_str,
                'Age_Days': age_days,
                'Strain': strain if pd.notna(strain) else 'N/A',
                'Priority_Strain': potential.get('Priority_Strain', 'Unknown'),
                'Dam': dam if pd.notna(dam) else 'N/A',
                'Sire': sire if pd.notna(sire) else 'N/A',
                'Num_Pups_Recorded': num_pups if pd.notna(num_pups) else 'N/A',
                'Status': birth['Status'],
                **expectations,
                'Sexing_Date': potential.get('Sexing_Date', 'N/A'),
                'Sexing_Day_of_Week': potential.get('Sexing_Day_of_Week', 'N/A'),
                'P14_Potential': potential['P14_Potential'],
                'P14_Expected_Date': potential['P14_Expected_Date'],
                'P14_Day_of_Week': potential['P14_Day_of_Week'],
                'P14_Analysis': potential['P14_Reason'],
                'P56_Potential': potential['P56_Potential'],
                'P56_Expected_Behavior_Date': potential['P56_Expected_Behavior_Date'],
                'P56_Expected_Harvest_Date': potential['P56_Expected_Harvest_Date'],
                'P56_Analysis': potential['P56_Reason'],
                'Quota_Status': potential['Quota_Status'],
                'Possible_Reasons': ' | '.join(possible_reasons),
                'Action_Required': determine_action_required(potential, expectations, age_days)
            })

    unmatched_df = pd.DataFrame(unmatched_births)

    if len(unmatched_df) > 0:
        urgency_order = {
            '🔴 URGENT - Past P56': 0, '🟡 HIGH - In P56 window': 1,
            '🟢 MEDIUM - Past P14': 2, '🟢 LOW - Approaching P14': 3,
            '⚪ INFO - Too young': 4, '❓ UNKNOWN - No birth date': 5
        }
        unmatched_df['_urgency_sort'] = unmatched_df['Urgency'].map(urgency_order)
        unmatched_df = unmatched_df.sort_values(['_urgency_sort', 'Birth_Date'])
        unmatched_df = unmatched_df.drop(columns=['_urgency_sort'])

    return unmatched_df


def create_unmatched_births_summary(unmatched_df: pd.DataFrame) -> pd.DataFrame:
    if len(unmatched_df) == 0:
        return pd.DataFrame()

    summary_data = []
    summary_data.append({'Category': 'Total Unmatched Births', 'Count': len(unmatched_df), 'Details': ''})

    if 'Urgency' in unmatched_df.columns:
        for urgency_val in unmatched_df['Urgency'].unique():
            count = len(unmatched_df[unmatched_df['Urgency'] == urgency_val])
            summary_data.append({'Category': 'By Urgency', 'Count': count, 'Details': urgency_val})

    if 'P14_Potential' in unmatched_df.columns:
        p14_yes = len(unmatched_df[unmatched_df['P14_Potential'] == 'Yes'])
        summary_data.append({'Category': 'P14 Schedulable', 'Count': p14_yes, 'Details': 'Could be scheduled for P14 if animals found'})

    if 'P56_Potential' in unmatched_df.columns:
        p56_yes = len(unmatched_df[unmatched_df['P56_Potential'] == 'Yes'])
        summary_data.append({'Category': 'P56 Schedulable', 'Count': p56_yes, 'Details': 'Could be scheduled for P56 if animals found'})

    if 'Priority_Strain' in unmatched_df.columns:
        priority_count = len(unmatched_df[unmatched_df['Priority_Strain'] == 'YES'])
        summary_data.append({'Category': 'Priority Strains', 'Count': priority_count, 'Details': 'High-priority strains needing immediate attention'})

    if 'Quota_Status' in unmatched_df.columns:
        quota_needed = len(unmatched_df[unmatched_df['Quota_Status'].str.contains('NEEDED', na=False)])
        summary_data.append({'Category': 'Has Quota Needs', 'Count': quota_needed, 'Details': 'Strains where quotas are not yet filled'})

    return pd.DataFrame(summary_data)


# ============================================================================
# REQUIREMENTS PARSING
# ============================================================================

def parse_requirements(tracking_df: Optional[pd.DataFrame]) -> Dict:
    if tracking_df is None or len(tracking_df) == 0:
        return {}

    logger.info("Parsing tracking file")

    column_indices = {
        'P14': {'Male': {'Perfusion': 3, 'MERFISH': 7, 'RNAseq': 11},
                'Female': {'Perfusion': 4, 'MERFISH': 8, 'RNAseq': 12}},
        'P56': {'Male': {'Perfusion': 5, 'MERFISH': 9, 'RNAseq': 13},
                'Female': {'Perfusion': 6, 'MERFISH': 10, 'RNAseq': 14}}
    }

    if len(tracking_df.columns) < 15:
        logger.warning(f"Expected at least 15 columns, found {len(tracking_df.columns)}")
        return {}

    # Filter to real strain rows only — real rows always have 'Yes' or 'No'
    # in column 1 (P14 Decisions). Ghost/percentage rows have decimals or NaN.
    col1_upper = tracking_df.iloc[:, 1].astype(str).str.strip().str.upper()
    tracking_df = tracking_df[col1_upper.isin(['YES', 'NO'])].reset_index(drop=True)
    logger.info(f"After filtering to Yes/No rows: {len(tracking_df)} strain rows")

    requirements = {}
    for idx, row in tracking_df.iterrows():
        strain = row.iloc[0]
        if pd.isna(strain) or str(strain).strip() in ['Lines', 'Line', '']:
            continue

        strain_str = str(strain).strip()
        strain_key = strain_str.upper()

        completed = {
            'P14': {'Male': {'Perfusion': 0, 'MERFISH': 0, 'RNAseq': 0},
                    'Female': {'Perfusion': 0, 'MERFISH': 0, 'RNAseq': 0}},
            'P56': {'Male': {'Perfusion': 0, 'MERFISH': 0, 'RNAseq': 0},
                    'Female': {'Perfusion': 0, 'MERFISH': 0, 'RNAseq': 0}}
        }

        try:
            for timepoint in ['P14', 'P56']:
                for sex in ['Male', 'Female']:
                    for harvest_type in ['Perfusion', 'MERFISH', 'RNAseq']:
                        col_idx = column_indices[timepoint][sex][harvest_type]
                        if col_idx < len(row):
                            value = row.iloc[col_idx]
                            completed[timepoint][sex][harvest_type] = int(value) if pd.notna(value) and str(value).strip() != '' else 0
        except Exception as e:
            logger.warning(f"Could not parse row for strain '{strain}': {e}")
            continue

        # Read P14/P56 completion flags from tracking sheet (cols 17 and 18).
        # These flags are maintained in the sheet and are the authoritative source.
        def _is_complete_flag(val):
            s = str(val).strip().upper()
            return s in ('TRUE', 'YES', '1')

        p14_complete = _is_complete_flag(row.iloc[17]) if len(row) > 17 else False
        p56_complete = _is_complete_flag(row.iloc[18]) if len(row) > 18 else False

        # Perfusion target is 11 per timepoint (one sex gets 6, other gets 5).
        # Assign the higher target (6) to whichever sex has fewer completions.
        # If equal, Male gets 5 and Female gets 6.
        targets = {}
        for _tp, _complete in [('P14', p14_complete), ('P56', p56_complete)]:
            if _complete:
                # Timepoint done — set targets equal to completed so needed=0
                targets[_tp] = {
                    'Male':   {ht: completed[_tp]['Male'][ht] for ht in ['Perfusion', 'MERFISH', 'RNAseq']},
                    'Female': {ht: completed[_tp]['Female'][ht] for ht in ['Perfusion', 'MERFISH', 'RNAseq']},
                }
            else:
                m_perf = completed[_tp]['Male']['Perfusion']
                f_perf = completed[_tp]['Female']['Perfusion']
                if m_perf < f_perf:
                    male_perf_target, female_perf_target = 6, 5
                else:
                    male_perf_target, female_perf_target = 5, 6
                targets[_tp] = {
                    'Male':   {**dict(CONFIG['HARVEST_TARGETS']), 'Perfusion': male_perf_target},
                    'Female': {**dict(CONFIG['HARVEST_TARGETS']), 'Perfusion': female_perf_target},
                }

        requirements[strain_key] = {
            'original_name': strain_str,
            'completed': completed,
            'targets': targets,
            'is_priority': is_priority_strain(strain_str)
        }

    logger.info(f"Parsed {len(requirements)} strains")
    return requirements


def calculate_remaining_needs(requirements: Dict) -> Dict:
    if not requirements:
        return {}

    remaining = {}
    for strain_key, data in requirements.items():
        remaining[strain_key] = {}
        for timepoint in ['P14', 'P56']:
            remaining[strain_key][timepoint] = {}
            for sex in ['Male', 'Female']:
                remaining[strain_key][timepoint][sex] = {}
                for harvest_type in ['Perfusion', 'MERFISH', 'RNAseq']:
                    completed = data['completed'][timepoint][sex][harvest_type]
                    target = data['targets'][timepoint][sex][harvest_type]
                    needed = max(0, target - completed)
                    remaining[strain_key][timepoint][sex][harvest_type] = {
                        'completed': completed, 'target': target, 'needed': needed
                    }
    return remaining




def check_extra_perfusion_status(requirements: Dict) -> Dict:
    """
    Check extra perfusion quota status per strain per timepoint.

    Rule: For each strain + timepoint, one sex must have 6 perfusions
    and the other must have at least 5 (11 total). The quota is fulfilled
    when one sex hits 6. Once both timepoints are fulfilled the strain
    is complete and should not receive any more perfusion scheduling.

    Returns a dict keyed by strain_key:
    {
        'P14': {'male_completed': int, 'female_completed': int,
                'fulfilled': bool, 'status': str},
        'P56': { ... },
        'strain_complete': bool   # True only if BOTH timepoints fulfilled
    }
    """
    EXTRA_PERF_TARGET_HIGH = 6
    EXTRA_PERF_TARGET_LOW  = 5

    result = {}
    for strain_key, data in requirements.items():
        strain_result = {}
        both_fulfilled = True

        for timepoint in ['P14', 'P56']:
            male_done   = data['completed'][timepoint]['Male']['Perfusion']
            female_done = data['completed'][timepoint]['Female']['Perfusion']

            # Fulfilled when either sex hits 6 AND the other has at least 5
            fulfilled = (
                (male_done >= EXTRA_PERF_TARGET_HIGH and female_done >= EXTRA_PERF_TARGET_LOW) or
                (female_done >= EXTRA_PERF_TARGET_HIGH and male_done >= EXTRA_PERF_TARGET_LOW)
            )

            if not fulfilled:
                both_fulfilled = False

            # Build a human-readable status
            if fulfilled:
                status = f'✅ Complete ({male_done}M / {female_done}F)'
            else:
                male_needed   = max(0, EXTRA_PERF_TARGET_LOW  - male_done)
                female_needed = max(0, EXTRA_PERF_TARGET_LOW  - female_done)
                # One of them needs to reach 6
                if male_done >= female_done:
                    male_needed   = max(0, EXTRA_PERF_TARGET_HIGH - male_done)
                    female_needed = max(0, EXTRA_PERF_TARGET_LOW  - female_done)
                else:
                    female_needed = max(0, EXTRA_PERF_TARGET_HIGH - female_done)
                    male_needed   = max(0, EXTRA_PERF_TARGET_LOW  - male_done)
                status = (f'⚠ In Progress ({male_done}M / {female_done}F) — '
                          f'need {male_needed} more M, {female_needed} more F')

            strain_result[timepoint] = {
                'male_completed':   male_done,
                'female_completed': female_done,
                'fulfilled':        fulfilled,
                'status':           status,
            }

        strain_result['strain_complete'] = both_fulfilled
        result[strain_key] = strain_result

    return result


def is_extra_perfusion_complete(strain: str, timepoint: str,
                                extra_perf_status: Dict) -> bool:
    """
    Returns True if the extra perfusion quota for this strain + timepoint
    is already fulfilled — meaning no more perfusions should be scheduled.
    """
    if not extra_perf_status:
        return False
    strain_key = str(strain).strip().upper()
    if strain_key not in extra_perf_status:
        return False
    return extra_perf_status[strain_key].get(timepoint, {}).get('fulfilled', False)

def group_has_quota(strain: str, sex: str, timepoint: str, remaining_needs: Dict) -> bool:
    strain_upper = str(strain).strip().upper()
    if strain_upper in _B6_STRAINS_UPPER:
        return True
    if not remaining_needs:
        return True

    strain_key = strain_upper
    if strain_key not in remaining_needs:
        return True

    needs = remaining_needs[strain_key][timepoint][sex]
    total_needed = needs['MERFISH']['needed'] + needs['RNAseq']['needed'] + needs['Perfusion']['needed']
    return total_needed >= 1


def create_requirements_status(remaining_needs: Dict, requirements: Dict, extra_perf_status: Dict = None) -> pd.DataFrame:
    if not remaining_needs or not requirements:
        return pd.DataFrame()

    status_rows = []
    for strain_key, timepoints in remaining_needs.items():
        original_strain = requirements[strain_key]['original_name']
        is_priority = requirements[strain_key]['is_priority']

        for timepoint, sexes in timepoints.items():
            for sex, harvest_types in sexes.items():
                for harvest_type, counts in harvest_types.items():
                    # Extra perfusion status for this strain/timepoint
                    extra_perf_col = ''
                    if extra_perf_status:
                        strain_key_ep = original_strain.strip().upper()
                        if strain_key_ep in extra_perf_status:
                            tp_data = extra_perf_status[strain_key_ep].get(timepoint, {})
                            extra_perf_col = tp_data.get('status', '')

                    status_rows.append({
                        'Strain': original_strain,
                        'Strain_Priority': 'PRIORITY' if is_priority else 'Standard',
                        'Timepoint': timepoint,
                        'Sex': sex,
                        'Harvest_Type': harvest_type,
                        'Target': counts['target'],
                        'Completed': counts['completed'],
                        'Remaining': counts['needed'],
                        'Progress': f"{counts['completed']}/{counts['target']}",
                        'Status': '✓ Complete' if counts['needed'] == 0 else f'Need {counts["needed"]} more',
                        'Extra_Perfusion': extra_perf_col,
                    })

    status_df = pd.DataFrame(status_rows)
    status_df = status_df.sort_values(
        ['Strain_Priority', 'Strain', 'Timepoint', 'Sex', 'Harvest_Type'],
        ascending=[False, True, True, True, True]
    )
    return status_df




# ============================================================================
# ANIMAL FILTERING
# ============================================================================

def filter_animals_by_use(animals_df: pd.DataFrame) -> Tuple[pd.DataFrame, pd.DataFrame]:
    """Filter animals to only those with 'Sing Inventory' in Use column."""
    if 'Use' not in animals_df.columns:
        return animals_df, pd.DataFrame()

    mask = animals_df['Use'].str.contains('Sing Inventory', na=False, case=False)
    filtered = animals_df[mask].copy()

    excluded_records = []
    for _, row in animals_df[~mask].iterrows():
        excluded_records.append({
            'Animal_Name': row.get('Name', 'Unknown'),
            'Birth_ID':    row.get('Birth ID', 'N/A'),
            'Strain':      row.get('Line (Short)', 'N/A'),
            'Genotype':    row.get('Genotype', 'N/A'),
            'Sex':         row.get('Sex', 'N/A'),
            'Reason':      f"Use = '{row.get('Use', 'N/A')}' — not 'Sing Inventory'",
        })

    return filtered, pd.DataFrame(excluded_records)


def filter_animals_by_genotype_first_pass(
    animals_df: pd.DataFrame
) -> Tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    """
    First-pass genotype filter.

    Returns
    -------
    filtered_df   : animals with usable genotypes (Het, Hom, Hemi, Inbred)
    excluded_df   : animals definitively excluded (Wild, Cre-only Wild)
    blank_df      : animals with Blank genotype — held for second pass
    """
    excluded_records = []
    blank_records    = []
    keep_mask        = []

    for _, row in animals_df.iterrows():
        geno   = row.get('Genotype', GENOTYPE_BLANK)
        strain = row.get('Line (Short)', '')
        name   = row.get('Name', 'Unknown')

        if geno == GENOTYPE_BLANK:
            blank_records.append(row)
            keep_mask.append(False)

        elif geno == GENOTYPE_WILD:
            excluded_records.append({
                'Animal_Name': name,
                'Birth_ID':    row.get('Birth ID', 'N/A'),
                'Strain':      strain,
                'Genotype':    geno,
                'Sex':         row.get('Sex', 'N/A'),
                'Birth_Date':  (
                    to_date(row.get('Birth Date')).strftime('%Y-%m-%d')
                    if to_date(row.get('Birth Date')) else 'N/A'
                ),
                'Reason': 'Wild genotype — not usable for harvest',
            })
            keep_mask.append(False)

        elif geno in (GENOTYPE_HET, GENOTYPE_HOM, GENOTYPE_HEMI, GENOTYPE_INBRED):
            keep_mask.append(True)

        else:
            # Unrecognised canonical value — treat as blank
            blank_records.append(row)
            keep_mask.append(False)

    filtered_df = animals_df[keep_mask].copy()
    excluded_df = pd.DataFrame(excluded_records)
    blank_df    = (
        pd.DataFrame(blank_records)
        if blank_records
        else pd.DataFrame(columns=animals_df.columns)
    )

    logger.info(
        f"Genotype first pass: {len(filtered_df)} kept, "
        f"{len(excluded_df)} excluded, {len(blank_df)} blanks"
    )
    return filtered_df, excluded_df, blank_df


def filter_animals_by_dates(animals_df: pd.DataFrame,
                            birth_date_start: Optional[date_type] = None,
                            birth_date_end: Optional[date_type] = None,
                            behavior_date_start: Optional[date_type] = None,
                            behavior_date_end: Optional[date_type] = None) -> Tuple[pd.DataFrame, pd.DataFrame]:
    if not any([birth_date_start, birth_date_end, behavior_date_start, behavior_date_end]):
        return animals_df, pd.DataFrame()

    df = animals_df.copy()
    df['_birth_date_obj'] = df['Birth Date'].apply(to_date)
    mask = pd.Series(True, index=df.index)
    reasons = pd.Series('', index=df.index)

    if birth_date_start:
        too_early = df['_birth_date_obj'].apply(lambda d: d is not None and d < birth_date_start)
        mask &= ~too_early
        reasons = reasons.where(~too_early, reasons + f'Birth date before {birth_date_start} | ')

    if birth_date_end:
        too_late = df['_birth_date_obj'].apply(lambda d: d is not None and d > birth_date_end)
        mask &= ~too_late
        reasons = reasons.where(~too_late, reasons + f'Birth date after {birth_date_end} | ')

    if behavior_date_start or behavior_date_end:
        def calc_first_wednesday(birth_d):
            if birth_d is None:
                return None
            sched = calculate_schedule_dates(birth_d)
            if sched is None:
                return None
            return next_wednesday(sched['p56_behavior_window_start'])

        df['_first_wed'] = df['_birth_date_obj'].apply(calc_first_wednesday)

        if behavior_date_start:
            too_early_b = df['_first_wed'].apply(lambda d: d is not None and d < behavior_date_start)
            mask &= ~too_early_b
            reasons = reasons.where(~too_early_b, reasons + f'P56 behavior before {behavior_date_start} | ')

        if behavior_date_end:
            too_late_b = df['_first_wed'].apply(lambda d: d is not None and d > behavior_date_end)
            mask &= ~too_late_b
            reasons = reasons.where(~too_late_b, reasons + f'P56 behavior after {behavior_date_end} | ')

    filtered = animals_df.loc[mask].copy()

    excluded_indices = df.index[~mask]
    if len(excluded_indices) > 0:
        excluded_records = []
        for idx in excluded_indices:
            row = animals_df.loc[idx]
            birth_d = to_date(row.get('Birth Date'))
            excluded_records.append({
                'Animal_Name': row.get('Name', 'Unknown'),
                'Birth_ID': row.get('Birth ID', 'N/A'),
                'Strain': row.get('Line (Short)', 'N/A'),
                'Sex': row.get('Sex', 'N/A'),
                'Birth_Date': birth_d.strftime('%Y-%m-%d') if birth_d else 'N/A',
                'Reason': reasons.loc[idx].rstrip(' | ')
            })
        excluded_df = pd.DataFrame(excluded_records)
    else:
        excluded_df = pd.DataFrame()

    return filtered, excluded_df


# ============================================================================
# ELIGIBILITY CHECKING
# ============================================================================

def check_eligibility(animals_df: pd.DataFrame,
                      full_behavior_dates: Optional[List[date_type]] = None) -> pd.DataFrame:
    today = datetime.now().date()
    full_dates_set = set(full_behavior_dates) if full_behavior_dates else set()

    logger.info("Checking animal eligibility...")

    if len(animals_df) == 0:
        logger.warning("check_eligibility received empty DataFrame — returning empty result")
        return pd.DataFrame()

    original_columns = animals_df.columns.tolist()
    eligibility = []

    iterator = animals_df.iterrows()
    if CONFIG['ENABLE_PROGRESS_BARS']:
        iterator = tqdm(list(animals_df.iterrows()), total=len(animals_df), desc="Checking eligibility")

    for idx, row in iterator:
        animal_name = row.get('Name', 'Unknown')
        birth_date = to_date(row.get('Birth Date'))
        strain = row.get('Line (Short)', 'N/A')
        genotype = row.get('Genotype')
        sex = row.get('Sex')
        marker_type = row.get('Marker Type', '')
        birth_id = row.get('Birth ID', 'N/A')

        original_data = {col: row.get(col) for col in original_columns}

        base_record = {
            **original_data,
            'Animal_Name': animal_name,
            'Birth_ID': str(birth_id),
            'Strain': strain,
            'Genotype': genotype if pd.notna(genotype) else GENOTYPE_BLANK,
            'Sex': sex,
            'Marker_Type': marker_type,
        }

        if birth_date is None:
            eligibility.append({
                **base_record,
                'Birth_Date': 'N/A',
                'Age_Today_Days': None,
                'P14_Eligible': False,
                'P14_Reason': 'No birth date',
                'P14_Too_Old': False,
                'P14_Date': None,
                'P14_Age_At_Harvest_Days': None,
                'P14_Age_At_Harvest_Months': None,
                'P56_Eligible': False,
                'P56_Reason': 'No birth date',
                'P56_Too_Old': False,
                'P56_Behavior_Date': None,
                'P56_Harvest_Date': None,
                'P56_Age_At_Behavior_Days': None,
                'P56_Age_At_Behavior_Months': None,
                'P56_Age_At_Harvest_Days': None,
                'P56_Age_At_Harvest_Months': None,
                'Unusable_Note': '',
            })
            continue

        dates = calculate_schedule_dates(birth_date)

        if dates is None:
            eligibility.append({
                **base_record,
                'Birth_Date': birth_date.strftime('%Y-%m-%d'),
                'Age_Today_Days': (today - birth_date).days,
                'P14_Eligible': False,
                'P14_Reason': 'Invalid birth date',
                'P14_Too_Old': False,
                'P14_Date': None,
                'P14_Age_At_Harvest_Days': None,
                'P14_Age_At_Harvest_Months': None,
                'P56_Eligible': False,
                'P56_Reason': 'Invalid birth date',
                'P56_Too_Old': False,
                'P56_Behavior_Date': None,
                'P56_Harvest_Date': None,
                'P56_Age_At_Behavior_Days': None,
                'P56_Age_At_Behavior_Months': None,
                'P56_Age_At_Harvest_Days': None,
                'P56_Age_At_Harvest_Months': None,
                'Unusable_Note': '',
            })
            continue

        p14_harvest = dates['p14_harvest']
        behavior_window_start = dates['p56_behavior_window_start']
        behavior_window_end = dates['p56_behavior_window_end']
        age_today = (today - birth_date).days

        # P14 eligibility
        p14_age_at_harvest_days = (p14_harvest - birth_date).days
        p14_age_at_harvest_months = round(p14_age_at_harvest_days / 30.44)
        p14_too_old = p14_harvest <= today

        if p14_too_old:
            p14_eligible = False
            days_past = (today - p14_harvest).days
            if days_past == 0:
                p14_reason = (
                    f'❌ TOO LATE FOR P14 — P14 date is today '
                    f'({p14_harvest.strftime("%Y-%m-%d")}) — '
                    f'harvest must be scheduled in advance'
                )
            else:
                p14_reason = (
                    f'❌ TOO OLD FOR P14 — P14 date was '
                    f'{p14_harvest.strftime("%Y-%m-%d")} '
                    f'({days_past} days ago, animal is {age_today}d old)'
                )
        elif not is_valid_p14_day(p14_harvest):
            p14_eligible = False
            p14_reason = (
                f'P14 falls on {p14_harvest.strftime("%A")} '
                f'({p14_harvest.strftime("%Y-%m-%d")}) — not a valid harvest day'
            )
        else:
            p14_eligible = True
            p14_reason = f'Eligible: {p14_harvest.strftime("%A, %Y-%m-%d")}'

        # P56 eligibility
        p56_eligible = False
        p56_reason = ''
        p56_too_old = False
        behavior_suggested = None
        p56_age_at_behavior_days = None
        p56_age_at_behavior_months = None
        p56_age_at_harvest_days = None
        p56_age_at_harvest_months = None
        p56_window_passed = behavior_window_end < today

        if has_toe_clip(marker_type):
            p56_reason = 'Has Toe Clip marker — not allowed for P56 behavior'
        elif p56_window_passed:
            p56_too_old = True
            days_past_p56 = (today - behavior_window_end).days
            p56_reason = (
                f'❌ TOO OLD FOR P56 — P56 behavior window ended '
                f'{behavior_window_end.strftime("%Y-%m-%d")} '
                f'({days_past_p56} days ago, animal is {age_today}d old). '
                f'Unusable for P56.'
            )
        else:
            first_wednesday = next_wednesday(behavior_window_start)

            if first_wednesday is None:
                p56_reason = 'Cannot calculate P56 behavior date'
            elif first_wednesday > behavior_window_end:
                p56_reason = 'No Wednesday falls within the P42–P49 window'
            elif first_wednesday < today:
                p56_too_old = True
                p56_reason = (
                    f'❌ TOO OLD FOR P56 — P56 window '
                    f'({first_wednesday.strftime("%Y-%m-%d")}) has passed '
                    f'(animal is {age_today}d old). Unusable for P56.'
                )
            elif first_wednesday in full_dates_set:
                p56_reason = (
                    f'Wednesday {first_wednesday.strftime("%Y-%m-%d")} '
                    f'is at capacity — cannot schedule P56'
                )
            else:
                p56_eligible = True
                p56_reason = f'Eligible: {first_wednesday.strftime("%A, %Y-%m-%d")}'
                behavior_suggested = first_wednesday

            if first_wednesday is not None and first_wednesday <= behavior_window_end:
                p56_age_at_behavior_days = (first_wednesday - birth_date).days
                p56_age_at_behavior_months = round(p56_age_at_behavior_days / 30.44)
                p56_harvest_calc = first_wednesday + timedelta(days=P56_HARVEST_OFFSET_FROM_BEHAVIOR)
                p56_age_at_harvest_days = (p56_harvest_calc - birth_date).days
                p56_age_at_harvest_months = round(p56_age_at_harvest_days / 30.44)

        # Combined unusable note
        if p14_too_old and p56_too_old:
            combined_unusable_note = (
                f'⛔ UNUSABLE FOR BOTH TIMEPOINTS — '
                f'Too old for P14 (was due {p14_harvest.strftime("%Y-%m-%d")}) '
                f'AND too old for P56 (window ended {behavior_window_end.strftime("%Y-%m-%d")}). '
                f'Animal is {age_today} days old.'
            )
        elif p14_too_old and not p56_too_old:
            combined_unusable_note = 'Too old for P14 only — P56 may still be viable'
        elif p56_too_old and not p14_too_old:
            combined_unusable_note = 'Too old for P56 only — P14 still viable'
        else:
            combined_unusable_note = ''

        eligibility.append({
            **base_record,
            'Birth_Date': birth_date.strftime('%Y-%m-%d'),
            'Age_Today_Days': age_today,
            'P14_Eligible': p14_eligible,
            'P14_Too_Old': p14_too_old,
            'P14_Reason': p14_reason,
            'P14_Date': p14_harvest if p14_eligible else None,
            'P14_Age_At_Harvest_Days': p14_age_at_harvest_days,
            'P14_Age_At_Harvest_Months': p14_age_at_harvest_months,
            'P56_Eligible': p56_eligible,
            'P56_Too_Old': p56_too_old,
            'P56_Reason': p56_reason,
            'P56_Behavior_Date': behavior_suggested if p56_eligible else None,
            'P56_Harvest_Date': (
                behavior_suggested + timedelta(days=P56_HARVEST_OFFSET_FROM_BEHAVIOR)
                if p56_eligible and behavior_suggested else None
            ),
            'P56_Age_At_Behavior_Days': p56_age_at_behavior_days,
            'P56_Age_At_Behavior_Months': p56_age_at_behavior_months,
            'P56_Age_At_Harvest_Days': p56_age_at_harvest_days,
            'P56_Age_At_Harvest_Months': p56_age_at_harvest_months,
            'Unusable_Note': combined_unusable_note,
        })

    logger.info(f"Eligibility check complete: {len(eligibility)} animals processed")
    return pd.DataFrame(eligibility)


# ============================================================================
# ANIMAL ASSIGNMENT
# ============================================================================

def assign_animals_smart(eligibility_df: pd.DataFrame, remaining_needs: Dict, extra_perf_status: Dict = None) -> pd.DataFrame:
    logger.info("Assigning animals to timepoints...")

    if len(eligibility_df) == 0:
        logger.warning("assign_animals_smart: eligibility_df is empty — no animals to assign")
        print("\n⚠️  WARNING: No animals passed eligibility checks.")
        print("    Check the diagnostic output above to see where animals were filtered.")
        return pd.DataFrame()

    required_cols = ['Strain', 'Sex', 'Genotype', 'Birth_Date', 'Animal_Name']
    missing_cols = [c for c in required_cols if c not in eligibility_df.columns]
    if missing_cols:
        logger.error(
            f"assign_animals_smart: eligibility_df missing columns: {missing_cols}. "
            f"Available: {eligibility_df.columns.tolist()}"
        )
        raise KeyError(
            f"eligibility_df is missing required columns: {missing_cols}. "
            f"Available columns: {eligibility_df.columns.tolist()}"
        )

    eligibility_df = eligibility_df.sort_values(
        ['Strain', 'Sex', 'Genotype', 'Birth_Date', 'Animal_Name']
    ).reset_index(drop=True)

    eligibility_df['breeding_type'] = eligibility_df['Strain'].apply(get_strain_breeding_type)

    is_super = eligibility_df['Strain'].apply(is_super_priority_strain)
    is_prio = eligibility_df['Strain'].apply(is_priority_strain)
    bt = eligibility_df['breeding_type']

    tier0a = eligibility_df[is_super & (bt == 'Half')].copy()
    tier0b = eligibility_df[is_super & (bt == 'All')].copy()
    tier1 = eligibility_df[is_prio & ~is_super & (bt == 'Half')].copy()
    tier2 = eligibility_df[is_prio & ~is_super & (bt == 'All')].copy()
    tier3 = eligibility_df[~is_prio & (bt == 'Half')].copy()
    tier4 = eligibility_df[~is_prio & (bt == 'All')].copy()
    tier5 = eligibility_df[~is_prio & (bt == 'Unknown')].copy()

    all_assignments = []

    tier_names = [
        "🔴 SUPER PRIORITY - Half (Het×WT) - HIGHEST",
        "🔴 SUPER PRIORITY - All (Hom×Hom)",
        "Priority - Half (Het×WT)",
        "Priority - All (Hom×Hom)",
        "Standard - Half (Het×WT)",
        "Standard - All (Hom×Hom)",
        "Standard - Unknown"
    ]

    for tier_num, (tier_name, animals_batch) in enumerate(
        zip(tier_names, [tier0a, tier0b, tier1, tier2, tier3, tier4, tier5])
    ):
        if len(animals_batch) == 0:
            continue

        logger.info(f"Tier {tier_num}: {tier_name} — {len(animals_batch)} animals")

        animals_batch = animals_batch.copy()
        animals_batch['is_het'] = animals_batch['Genotype'].apply(is_heterozygous)
        animals_batch = animals_batch.sort_values('is_het', ascending=False)

        p56_candidates = animals_batch[animals_batch['P56_Eligible']].copy()
        p56_blocked_by_full_date = animals_batch[
            (~animals_batch['P56_Eligible']) &
            (animals_batch['P56_Reason'].str.contains('capacity', na=False, case=False))
        ].copy()

        p56_assignments = []
        p56_fallback = []

        if len(p56_candidates) > 0:
            p56_candidates['Genotype'] = p56_candidates['Genotype'].apply(
                lambda g: g if g in _CANONICAL_GENOTYPES else canonicalize_genotype(g)
            )

        grouped = (
            p56_candidates.groupby(['Strain', 'Genotype', 'Sex', 'P56_Behavior_Date'])
            if len(p56_candidates) > 0 else []
        )

        sorted_groups = sorted(
            grouped,
            key=lambda x: (
                not is_heterozygous(x[0][1]),
                str(x[0][0]) if x[0][0] is not None else '',
                str(x[0][2]) if x[0][2] is not None else '',
                to_date(x[0][3]) if x[0][3] is not None else date.min,
                str(x[0][1]) if x[0][1] is not None else ''
            )
        ) if len(p56_candidates) > 0 else []

        unschedulable = []

        for (strain, genotype, sex, behavior_date), group in sorted_groups:
            group_sorted = group.sort_values('Animal_Name').copy()
            animals = group_sorted.to_dict('records')

            if is_extra_perfusion_complete(strain, 'P56', extra_perf_status):
                for animal in animals:
                    animal['_quota_limited_complete_group'] = True
                    animal['_incomplete_group'] = False
                    animal['_extra_perf_complete'] = True
                    p56_fallback.append(animal)
                continue

            if not group_has_quota(strain, sex, 'P56', remaining_needs):
                for animal in animals:
                    animal['_quota_limited_complete_group'] = True
                    animal['_incomplete_group'] = False
                    p56_fallback.append(animal)
                continue

            num_complete_groups = len(animals) // CONFIG['CAGE_SIZE']

            strain_key = str(strain).strip().upper()
            if remaining_needs and strain_key in remaining_needs and strain_key not in _B6_STRAINS_UPPER:
                needs = remaining_needs[strain_key]['P56'][sex]
                total_needed = needs['MERFISH']['needed'] + needs['RNAseq']['needed'] + needs['Perfusion']['needed']
                max_groups_by_quota = (total_needed + CONFIG['CAGE_SIZE'] - 1) // CONFIG['CAGE_SIZE']
                num_complete_groups = min(num_complete_groups, max_groups_by_quota)

            for i in range(num_complete_groups * CONFIG['CAGE_SIZE']):
                p56_assignments.append({
                    **animals[i],
                    'Assigned_Timepoint': 'P56',
                    'Assignment_Reason': 'Complete cage group of 3',
                })

            leftover = animals[num_complete_groups * CONFIG['CAGE_SIZE']:]
            if leftover:
                birth_groups = defaultdict(list)
                for animal in leftover:
                    birth_groups[animal.get('Birth_ID', 'Unknown')].append(animal)
                for bid, ba in birth_groups.items():
                    for animal in ba:
                        animal['_quota_limited_complete_group'] = len(ba) >= CONFIG['CAGE_SIZE']
                        animal['_incomplete_group'] = len(ba) < CONFIG['CAGE_SIZE']
                        p56_fallback.append(animal)

        # Handle P56 blocked by full date
        if len(p56_blocked_by_full_date) > 0:
            p56_blocked_by_full_date = p56_blocked_by_full_date.copy()
            p56_blocked_by_full_date['Genotype'] = p56_blocked_by_full_date['Genotype'].apply(
                lambda g: g if g in _CANONICAL_GENOTYPES else canonicalize_genotype(g)
            )

            def get_p56_behavior_date(row):
                birth_d = to_date(row['Birth_Date'])
                if birth_d is None:
                    return None
                sched_dates = calculate_schedule_dates(birth_d)
                if sched_dates is None:
                    return None
                return next_wednesday(sched_dates['p56_behavior_window_start'])

            p56_blocked_by_full_date['P56_Behavior_Date_Calc'] = p56_blocked_by_full_date.apply(
                get_p56_behavior_date, axis=1
            )

            for (strain, genotype, sex, behavior_date), group in p56_blocked_by_full_date.groupby(
                ['Strain', 'Genotype', 'Sex', 'P56_Behavior_Date_Calc']
            ):
                animals = group.sort_values('Animal_Name').to_dict('records')
                num_complete_groups = len(animals) // CONFIG['CAGE_SIZE']

                if num_complete_groups > 0:
                    for animal in animals:
                        if animal['P14_Eligible']:
                            animal['_full_date_complete'] = True
                            animal['_incomplete_group'] = False
                            animal['_quota_limited_complete_group'] = False
                            p56_fallback.append(animal)
                        else:
                            unschedulable.append({
                                **animal,
                                'Assigned_Timepoint': 'Unschedulable',
                                'Assignment_Reason': (
                                    f'P56 date at capacity '
                                    f'({behavior_date.strftime("%Y-%m-%d") if behavior_date else "?"}). '
                                    f'P14 unavailable: {animal["P14_Reason"]}'
                                ),
                            })
                else:
                    for animal in animals:
                        if animal['P14_Eligible']:
                            animal['_incomplete_group'] = True
                            animal['_full_date_complete'] = False
                            animal['_quota_limited_complete_group'] = False
                            p56_fallback.append(animal)
                        else:
                            unschedulable.append({
                                **animal,
                                'Assigned_Timepoint': 'Unschedulable',
                                'Assignment_Reason': (
                                    f'Incomplete P56 group; P14 unavailable: {animal["P14_Reason"]}'
                                ),
                            })

        # P14 fallback
        p14_assignments = []
        for animal in sorted(p56_fallback, key=lambda x: not is_heterozygous(x.get('Genotype', ''))):
            if not animal.get('P14_Eligible', False):
                if animal.get('_quota_limited_complete_group'):
                    reason_prefix = 'P56 quota filled (complete cage not needed)'
                elif animal.get('_full_date_complete'):
                    reason_prefix = 'P56 date at capacity'
                else:
                    reason_prefix = 'Incomplete P56 group'
                unschedulable.append({
                    **animal,
                    'Assigned_Timepoint': 'Unschedulable',
                    'Assignment_Reason': f'{reason_prefix}; P14 unavailable: {animal.get("P14_Reason", "Unknown")}',
                })
                continue

            strain = animal.get('Strain', '')
            sex = animal.get('Sex', '')

            if is_extra_perfusion_complete(strain, 'P14', extra_perf_status):
                for animal in animals_p14:
                    animal['_extra_perf_complete'] = True
                    unschedulable_p14.append(animal)
                continue

            if group_has_quota(strain, sex, 'P14', remaining_needs):
                if animal.get('_quota_limited_complete_group'):
                    reason = 'P56 quota filled for strain — reassigned to P14'
                elif animal.get('_full_date_complete'):
                    reason = 'P56 date at capacity — reassigned to P14'
                else:
                    reason = 'Incomplete P56 group — reassigned to P14'
                p14_assignments.append({
                    **animal,
                    'Assigned_Timepoint': 'P14',
                    'Assignment_Reason': reason,
                })
            else:
                if animal.get('_quota_limited_complete_group'):
                    unsched = 'P56 quota filled; P14 quota also filled'
                elif animal.get('_full_date_complete'):
                    unsched = 'P56 date at capacity; P14 quota also filled'
                else:
                    unsched = 'Incomplete P56 group; P14 quota also filled'
                unschedulable.append({
                    **animal,
                    'Assigned_Timepoint': 'Unschedulable',
                    'Assignment_Reason': unsched,
                })

        # P14-only animals
        p14_only = animals_batch[
            animals_batch['P14_Eligible'] &
            ~animals_batch['P56_Eligible'] &
            ~animals_batch['P56_Reason'].str.contains('capacity', na=False, case=False)
        ].copy().sort_values('is_het', ascending=False)

        for idx2, animal in p14_only.iterrows():
            strain = animal['Strain']
            sex = animal['Sex']
            if group_has_quota(strain, sex, 'P14', remaining_needs):
                p14_assignments.append({
                    **animal.to_dict(),
                    'Assigned_Timepoint': 'P14',
                    'Assignment_Reason': f'P14 only (P56: {animal["P56_Reason"]})',
                })
            else:
                unschedulable.append({
                    **animal.to_dict(),
                    'Assigned_Timepoint': 'Unschedulable',
                    'Assignment_Reason': f'P14 quota filled for {strain} {sex}',
                })

        # Neither eligible
        neither = animals_batch[
            ~animals_batch['P14_Eligible'] &
            ~animals_batch['P56_Eligible'] &
            ~animals_batch['P56_Reason'].str.contains('capacity', na=False, case=False)
        ].copy()

        for idx2, animal in neither.iterrows():
            unschedulable.append({
                **animal.to_dict(),
                'Assigned_Timepoint': 'Unschedulable',
                'Assignment_Reason': f'P14: {animal["P14_Reason"]}; P56: {animal["P56_Reason"]}',
            })

        all_assignments.extend(p56_assignments + p14_assignments + unschedulable)

    logger.info(f"Assignment complete: {len(all_assignments)} animals")
    return pd.DataFrame(all_assignments)


# ============================================================================
# B6/B6N MONTHLY MINIMUM ENFORCEMENT
# ============================================================================

def enforce_b6_monthly_minimum(assignments_df: pd.DataFrame,
                                eligibility_df: pd.DataFrame,
                                remaining_needs: Dict) -> pd.DataFrame:
    min_per_month = CONFIG.get('B6_MIN_PER_MONTH', 3)
    if min_per_month <= 0:
        return assignments_df

    if len(assignments_df) == 0:
        return assignments_df

    logger.info(f"Enforcing B6/B6N minimum of {min_per_month}/month after quota...")

    scheduled = assignments_df[
        assignments_df['Assigned_Timepoint'].isin(['P14', 'P56'])
    ].copy()

    def harvest_month(row):
        tp = row.get('Assigned_Timepoint', '')
        if tp == 'P14':
            d = to_date(row.get('P14_Date'))
        elif tp == 'P56':
            d = to_date(row.get('P56_Harvest_Date'))
        else:
            d = None
        return (d.year, d.month) if d else None

    scheduled['_harvest_month'] = scheduled.apply(harvest_month, axis=1)
    scheduled = scheduled[scheduled['_harvest_month'].notna()]

    all_harvest_months = sorted(scheduled['_harvest_month'].unique())
    if not all_harvest_months:
        logger.info("No scheduled harvest months found — skipping B6/B6N minimum check")
        return assignments_df

    b6_scheduled = scheduled[scheduled['Strain'].apply(is_b6_strain)].copy()
    b6_per_month: Dict[Tuple, int] = {}
    for month in all_harvest_months:
        b6_per_month[month] = int((b6_scheduled['_harvest_month'] == month).sum())

    logger.info("B6/B6N current scheduled counts by month:")
    for month, count in sorted(b6_per_month.items()):
        logger.info(f"  {month[0]}-{month[1]:02d}: {count} (min required: {min_per_month})")

    already_scheduled_names = set(
        assignments_df[
            assignments_df['Assigned_Timepoint'].isin(['P14', 'P56'])
        ]['Animal_Name'].tolist()
    )

    if len(eligibility_df) > 0:
        b6_pool = eligibility_df[
            eligibility_df['Strain'].apply(is_b6_strain)
        ].copy()
        b6_pool = b6_pool[~b6_pool['Animal_Name'].isin(already_scheduled_names)].copy()
    else:
        b6_pool = pd.DataFrame()

    b6_unschedulable = assignments_df[
        (assignments_df['Assigned_Timepoint'] == 'Unschedulable') &
        assignments_df['Strain'].apply(is_b6_strain)
    ].copy()

    logger.info(f"B6/B6N pool: {len(b6_pool)} eligible not yet scheduled, "
                f"{len(b6_unschedulable)} currently unschedulable")

    new_rows: List[Dict] = []
    added_names: set = set()

    for month in all_harvest_months:
        current_count = b6_per_month.get(month, 0)
        shortfall = min_per_month - current_count

        if shortfall <= 0:
            logger.info(f"  {month[0]}-{month[1]:02d}: already has {current_count} >= {min_per_month} — OK")
            continue

        logger.info(f"  {month[0]}-{month[1]:02d}: needs {shortfall} more B6/B6N (has {current_count})")
        added_this_month = 0

        if len(b6_pool) > 0:
            p14_candidates = b6_pool[b6_pool['P14_Eligible'] == True].copy()

            for _, candidate in p14_candidates.iterrows():
                if added_this_month >= shortfall:
                    break
                name = candidate['Animal_Name']
                if name in added_names:
                    continue

                p14_date = to_date(candidate.get('P14_Date'))
                if p14_date is None:
                    continue

                candidate_month = (p14_date.year, p14_date.month)
                if candidate_month != month:
                    continue

                new_rows.append({
                    **candidate.to_dict(),
                    'Assigned_Timepoint': 'P14',
                    'Assignment_Reason': (
                        f'B6/B6N monthly minimum top-up '
                        f'(month {month[0]}-{month[1]:02d} had only {current_count}, '
                        f'min={min_per_month})'
                    ),
                    'Harvest_Type': 'Perfusion',
                    'Priority': 'B6_MIN',
                    'Strain_Priority': 'B6/B6N Control',
                    'Genotype_Priority': 'B6/B6N',
                })
                added_names.add(name)
                added_this_month += 1

            if added_this_month < shortfall:
                p56_candidates = b6_pool[b6_pool['P56_Eligible'] == True].copy()

                for _, candidate in p56_candidates.iterrows():
                    if added_this_month >= shortfall:
                        break
                    name = candidate['Animal_Name']
                    if name in added_names:
                        continue

                    p56_harvest = to_date(candidate.get('P56_Harvest_Date'))
                    if p56_harvest is None:
                        bhv = to_date(candidate.get('P56_Behavior_Date'))
                        if bhv:
                            p56_harvest = bhv + timedelta(days=P56_HARVEST_OFFSET_FROM_BEHAVIOR)

                    if p56_harvest is None:
                        continue

                    candidate_month = (p56_harvest.year, p56_harvest.month)
                    if candidate_month != month:
                        continue

                    bhv_date = to_date(candidate.get('P56_Behavior_Date'))

                    new_rows.append({
                        **candidate.to_dict(),
                        'Assigned_Timepoint': 'P56',
                        'Assignment_Reason': (
                            f'B6/B6N monthly minimum top-up '
                            f'(month {month[0]}-{month[1]:02d} had only {current_count}, '
                            f'min={min_per_month})'
                        ),
                        'Harvest_Type': 'Perfusion',
                        'Priority': 'B6_MIN',
                        'Strain_Priority': 'B6/B6N Control',
                        'Genotype_Priority': 'B6/B6N',
                        'P56_Behavior_Date': bhv_date,
                        'P56_Harvest_Date': p56_harvest,
                    })
                    added_names.add(name)
                    added_this_month += 1

        if added_this_month < shortfall and len(b6_unschedulable) > 0:
            for _, candidate in b6_unschedulable.iterrows():
                if added_this_month >= shortfall:
                    break
                name = candidate['Animal_Name']
                if name in added_names:
                    continue

                if candidate.get('P14_Eligible', False):
                    p14_date = to_date(candidate.get('P14_Date'))
                    if p14_date and (p14_date.year, p14_date.month) == month:
                        new_rows.append({
                            **candidate.to_dict(),
                            'Assigned_Timepoint': 'P14',
                            'Assignment_Reason': (
                                f'B6/B6N monthly minimum top-up from unschedulable pool '
                                f'(month {month[0]}-{month[1]:02d}, min={min_per_month})'
                            ),
                            'Harvest_Type': 'Perfusion',
                            'Priority': 'B6_MIN',
                            'Strain_Priority': 'B6/B6N Control',
                            'Genotype_Priority': 'B6/B6N',
                        })
                        added_names.add(name)
                        added_this_month += 1
                        continue

                if candidate.get('P56_Eligible', False):
                    p56_harvest = to_date(candidate.get('P56_Harvest_Date'))
                    if p56_harvest and (p56_harvest.year, p56_harvest.month) == month:
                        new_rows.append({
                            **candidate.to_dict(),
                            'Assigned_Timepoint': 'P56',
                            'Assignment_Reason': (
                                f'B6/B6N monthly minimum top-up from unschedulable pool '
                                f'(month {month[0]}-{month[1]:02d}, min={min_per_month})'
                            ),
                            'Harvest_Type': 'Perfusion',
                            'Priority': 'B6_MIN',
                            'Strain_Priority': 'B6/B6N Control',
                            'Genotype_Priority': 'B6/B6N',
                        })
                        added_names.add(name)
                        added_this_month += 1

        if added_this_month > 0:
            logger.info(f"  → Added {added_this_month} B6/B6N top-up animals for {month[0]}-{month[1]:02d}")
        else:
            logger.warning(
                f"  ⚠️ Could not find enough B6/B6N animals for {month[0]}-{month[1]:02d} "
                f"(added {added_this_month}/{shortfall})"
            )

    if new_rows:
        top_up_df = pd.DataFrame(new_rows)
        updated_assignments = assignments_df[
            ~(
                (assignments_df['Animal_Name'].isin(added_names)) &
                (assignments_df['Assigned_Timepoint'] == 'Unschedulable')
            )
        ].copy()
        all_cols = updated_assignments.columns.tolist()
        for col in all_cols:
            if col not in top_up_df.columns:
                top_up_df[col] = None
        updated_assignments = pd.concat(
            [updated_assignments, top_up_df[all_cols]], ignore_index=True
        )
        logger.info(f"B6/B6N minimum enforcement: added {len(new_rows)} top-up animals")
        return updated_assignments

    logger.info("B6/B6N minimum enforcement: no top-up needed or no animals available")
    return assignments_df


# ============================================================================
# HARVEST ASSIGNMENT GUI
# ============================================================================

# ============================================================================
# HARVEST ASSIGNMENT GUI
# ============================================================================

import tkinter as tk
from tkinter import ttk, messagebox
import copy as _copy


def _compute_auto_types(schedulable_df, remaining_needs):
    """
    Run the same quota logic as assign_harvest_types but just return a
    name → type dict without modifying the real assignments.
    """
    working = _copy.deepcopy(remaining_needs)
    result = {}

    sorted_df = schedulable_df.sort_values(
        ['Strain', 'Sex', 'Assigned_Timepoint', 'Animal_Name']
    ).reset_index(drop=True)

    for _, row in sorted_df.iterrows():
        name      = str(row.get('Animal_Name', '')).strip()
        strain    = row.get('Strain', '')
        sex       = row.get('Sex', '')
        timepoint = row.get('Assigned_Timepoint', '')
        strain_key = str(strain).strip().upper()

        if strain_key in _B6_STRAINS_UPPER:
            result[name] = 'Perfusion'
            continue

        if strain_key not in working:
            result[name] = 'Perfusion'
            continue

        needs = working[strain_key][timepoint][sex]

        if needs['MERFISH']['needed'] > 0:
            result[name] = 'MERFISH'
            needs['MERFISH']['needed'] -= 1
        elif needs['RNAseq']['needed'] > 0:
            result[name] = 'RNAseq'
            needs['RNAseq']['needed'] -= 1
        elif needs['Perfusion']['needed'] > 0:
            result[name] = 'Perfusion'
            needs['Perfusion']['needed'] -= 1
        else:
            result[name] = 'Perfusion'   # over-quota default

    return result


def _compute_quota_status(selections, schedulable_df, remaining_needs):
    """
    Compare current GUI selections against remaining_needs.
    Returns a list of (strain, timepoint, sex, harvest_type, needed, selected, status_str).
    """
    if not remaining_needs:
        return []

    # Count selected per (strain_key, timepoint, sex, harvest_type)
    counts = {}
    df_map = {
        str(r.get('Animal_Name', '')).strip(): r
        for _, r in schedulable_df.iterrows()
    }

    for name, htype in selections.items():
        if htype in ('Do Not Schedule', 'Extra'):
            continue
        row = df_map.get(name)
        if row is None:
            continue
        strain_key = str(row.get('Strain', '')).strip().upper()
        timepoint  = str(row.get('Assigned_Timepoint', '')).strip()
        sex        = str(row.get('Sex', '')).strip()
        key = (strain_key, timepoint, sex, htype)
        counts[key] = counts.get(key, 0) + 1
        logger.debug(f"QUOTA COUNT: {name} → {strain_key} {timepoint} {sex} {htype} (running total: {counts[key]})")

    # Build the set of (strain, timepoint, sex) combos that have at least
    # one animal actually being harvested (not Extra or Do Not Schedule).
    # This ensures we only flag quota mismatches for groups with real
    # harvest animals in this run.
    # Build set of (strain, timepoint, sex, harvest_type) combos that
    # actually have animals selected in this run (excluding Extra/DNS).
    # Only these specific combos should be checked for mismatches.
    present_combos = set()
    for name, htype in selections.items():
        if htype in ('Do Not Schedule', 'Extra'):
            continue
        row = df_map.get(name)
        if row is None:
            continue
        sk = str(row.get('Strain', '')).strip().upper()
        tp = str(row.get('Assigned_Timepoint', '')).strip()
        sx = str(row.get('Sex', '')).strip()
        if tp in ('P14', 'P56'):
            present_combos.add((sk, tp, sx, htype))

    rows = []
    for strain_key, timepoints in remaining_needs.items():
        for timepoint, sexes in timepoints.items():
            for sex, htypes in sexes.items():
                for htype, info in htypes.items():
                    selected = counts.get((strain_key, timepoint, sex, htype), 0)
                    needed   = info['needed']
                    # Only flag if this specific harvest type was actually selected
                    # in this run — avoids false mismatches for types not scheduled
                    if (strain_key, timepoint, sex, htype) not in present_combos:
                        continue
                    if needed == 0 and selected == 0:
                        continue
                    if selected == needed:
                        status = '✓ Match'
                    elif selected > needed:
                        status = f'↑ {selected - needed} over'
                    else:
                        status = f'↓ {needed - selected} short'
                    rows.append((
                        strain_key, timepoint, sex, htype,
                        needed, selected, status
                    ))
    return rows


def prompt_harvest_assignments_gui(assignments_df, remaining_needs):
    """
    Block the pipeline and show a GUI letting the user review and confirm
    harvest type assignments for every scheduled animal.

    Returns a dict { animal_name: harvest_type_or_None }
    where None / missing means 'auto-assign as normal'.
    'Do Not Schedule' is returned as the string 'DO_NOT_SCHEDULE'
    so the caller can act on it.

    If tkinter is unavailable the function returns {} and the pipeline
    continues with auto-assignment.
    """
    try:
        import tkinter as _tk_test
        _tk_test.Tk().destroy()
    except Exception:
        print("  ⚠ tkinter not available — skipping harvest assignment GUI.")
        return {}

    HARVEST_OPTIONS = ['Perfusion', 'MERFISH', 'RNAseq', 'Extra', 'Do Not Schedule']
    OPTION_COLORS   = {
        'Perfusion':        '#d4edda',
        'MERFISH':          '#cce5ff',
        'RNAseq':           '#fff3cd',
        'Extra':            '#e8d5f5',
        'Do Not Schedule':  '#f8d7da',
    }
    STATUS_COLORS = {
        '✓ Match':  '#c3e6cb',
        '↑':        '#ffeeba',
        '↓':        '#f5c6cb',
    }

    # Only show schedulable animals (P14 / P56), skip Unschedulable
    schedulable = assignments_df[
        assignments_df['Assigned_Timepoint'].isin(['P14', 'P56'])
    ].copy()

    if len(schedulable) == 0:
        return {}

    # Compute auto-suggested types
    auto_types = _compute_auto_types(schedulable, remaining_needs)

    # ── Build the window ──────────────────────────────────────────────────────
    root = tk.Tk()
    root.title("Harvest Assignment Review")
    root.configure(bg='#f0f0f0')
    root.resizable(True, True)

    # Keep track of the user's final decision
    confirmed = {'result': None}

    # ── Header ────────────────────────────────────────────────────────────────
    header_frame = tk.Frame(root, bg='#2c3e50', pady=12)
    header_frame.pack(fill='x')
    tk.Label(
        header_frame,
        text="Harvest Assignment Review",
        font=('Helvetica', 16, 'bold'),
        bg='#2c3e50', fg='white'
    ).pack()
    tk.Label(
        header_frame,
        text=(f"{len(schedulable)} animals ready to schedule  •  "
              f"Review assignments below, make any changes, then confirm."),
        font=('Helvetica', 10),
        bg='#2c3e50', fg='#bdc3c7'
    ).pack()

    # ── Main pane (left = animal table, right = quota panel) ──────────────────
    main_frame = tk.Frame(root, bg='#f0f0f0')
    main_frame.pack(fill='both', expand=True, padx=12, pady=8)

    # ── LEFT: Animal table ────────────────────────────────────────────────────
    left_frame = tk.LabelFrame(
        main_frame, text=" Animals to Schedule ",
        font=('Helvetica', 11, 'bold'),
        bg='#f0f0f0', fg='#2c3e50', padx=6, pady=6
    )
    left_frame.pack(side='left', fill='both', expand=True, padx=(0, 6))

    # Column headers
    headers = ['Animal Name', 'Strain', 'Sex', 'Timepoint', 'Date', 'Harvest Type']
    col_widths = [18, 12, 8, 10, 12, 16]

    hdr_row = tk.Frame(left_frame, bg='#2c3e50')
    hdr_row.pack(fill='x')
    for h, w in zip(headers, col_widths):
        tk.Label(
            hdr_row, text=h, width=w, anchor='w',
            font=('Helvetica', 9, 'bold'),
            bg='#2c3e50', fg='white', padx=4, pady=4
        ).pack(side='left')

    # ── Color key ─────────────────────────────────────────────────────────────
    key_frame = tk.Frame(left_frame, bg='#e8e8e8', pady=3)
    key_frame.pack(fill='x')
    tk.Label(
        key_frame, text='Row color = selected harvest type:',
        font=('Helvetica', 8, 'italic'), bg='#e8e8e8', fg='#555555', padx=6
    ).pack(side='left')
    for label, color in OPTION_COLORS.items():
        swatch = tk.Frame(key_frame, bg=color, width=12, height=12,
                          relief='solid', bd=1)
        swatch.pack(side='left', padx=(4, 1), pady=2)
        swatch.pack_propagate(False)
        tk.Label(
            key_frame, text=label,
            font=('Helvetica', 8), bg='#e8e8e8', fg='#333333', padx=2
        ).pack(side='left')

    # Scrollable rows
    canvas = tk.Canvas(left_frame, bg='#f0f0f0', highlightthickness=0)
    scrollbar = ttk.Scrollbar(left_frame, orient='vertical', command=canvas.yview)
    rows_frame = tk.Frame(canvas, bg='#f0f0f0')

    rows_frame.bind(
        '<Configure>',
        lambda e: canvas.configure(scrollregion=canvas.bbox('all'))
    )
    canvas.create_window((0, 0), window=rows_frame, anchor='nw')
    canvas.configure(yscrollcommand=scrollbar.set)

    canvas.pack(side='left', fill='both', expand=True)
    scrollbar.pack(side='right', fill='y')

    # Mouse-wheel scroll
    def _on_mousewheel(event):
        canvas.yview_scroll(int(-1 * (event.delta / 120)), 'units')
    canvas.bind_all('<MouseWheel>', _on_mousewheel)

    # Sort: by timepoint then harvest date then strain then name
    def _harvest_sort_key(row):
        tp = str(row.get('Assigned_Timepoint', ''))
        d  = str(row.get('P14_Date' if tp == 'P14' else 'P56_Harvest_Date', '') or '')
        return (tp, d, str(row.get('Strain', '')), str(row.get('Animal_Name', '')))

    sorted_rows = sorted(schedulable.to_dict('records'), key=_harvest_sort_key)

    # Store StringVars so we can read them later
    selection_vars   = {}   # name → StringVar
    selection_values = {}   # name → current value (always in sync, avoids tkinter canvas StringVar decouple bug)
    row_frames       = {}   # name → tk.Frame (for recoloring)

    def _on_type_change_cb(name, combobox, frame):
        """Called on <<ComboboxSelected>> — reads directly from combobox widget."""
        val = combobox.get()
        selection_values[name] = val   # store in plain dict (reliable)
        color = OPTION_COLORS.get(val, '#ffffff')
        frame.configure(bg=color)
        for w in frame.winfo_children():
            try:
                w.configure(bg=color)
            except Exception:
                pass
        _refresh_quota_panel()

    def _on_type_change(name, var, frame):
        val = var.get()
        selection_values[name] = val
        color = OPTION_COLORS.get(val, '#ffffff')
        frame.configure(bg=color)
        for w in frame.winfo_children():
            try:
                w.configure(bg=color)
            except Exception:
                pass
        _refresh_quota_panel()

    for i, row in enumerate(sorted_rows):
        name      = str(row.get('Animal_Name', '')).strip()
        strain    = str(row.get('Strain', '')).strip()
        sex       = str(row.get('Sex', '')).strip()
        timepoint = str(row.get('Assigned_Timepoint', '')).strip()
        default   = auto_types.get(name, 'Perfusion')

        var = tk.StringVar(value=default)
        selection_vars[name] = var
        selection_values[name] = default   # seed plain-dict copy

        bg = '#ffffff' if i % 2 == 0 else '#f7f7f7'
        frame = tk.Frame(rows_frame, bg=bg)
        frame.pack(fill='x')
        row_frames[name] = frame

        # Pick the relevant date: P14 -> harvest date, P56 -> behavior date
        # A missing date here means a scheduling logic error — flag it clearly
        if timepoint == 'P14':
            raw_date = str(row.get('P14_Date', '') or '')
        else:
            raw_date = str(row.get('P56_Behavior_Date', '') or '')
        try:
            from datetime import datetime as _dt
            display_date = _dt.strptime(raw_date, '%Y-%m-%d').strftime('%m/%d/%y')
        except Exception:
            display_date = '⚠ NO DATE'
            logger.warning(f"Animal {name} ({timepoint}) is scheduled but has no date — check scheduling logic")

        for val, w in zip([name, strain, sex, timepoint, display_date], col_widths[:5]):
            tk.Label(
                frame, text=val, width=w, anchor='w',
                font=('Helvetica', 9), bg=bg, padx=4, pady=3
            ).pack(side='left')

        menu = ttk.Combobox(
            frame, textvariable=var,
            values=HARVEST_OPTIONS,
            state='readonly', width=col_widths[4] - 2
        )
        menu.pack(side='left', padx=2, pady=2)
        menu.bind('<<ComboboxSelected>>',
                  lambda e, n=name, f=frame, m=menu: _on_type_change_cb(n, m, f))

        # Apply initial color
        c = OPTION_COLORS.get(default, bg)
        frame.configure(bg=c)
        for w in frame.winfo_children():
            try:
                w.configure(bg=c)
            except Exception:
                pass

    # ── RIGHT: Quota comparison panel ─────────────────────────────────────────
    right_frame = tk.LabelFrame(
        main_frame, text=" Quota Comparison ",
        font=('Helvetica', 11, 'bold'),
        bg='#f0f0f0', fg='#2c3e50', padx=6, pady=6
    )
    right_frame.pack(side='right', fill='y', padx=(6, 0))
    right_frame.pack_propagate(False)
    right_frame.configure(width=340)

    quota_inner = tk.Frame(right_frame, bg='#f0f0f0')
    quota_inner.pack(fill='both', expand=True)

    quota_header_cols = ['Strain', 'TP', 'Sex', 'Type', 'Need', 'Sel', 'Status']
    quota_col_widths  = [10, 4,  5,  8, 5, 5, 10]

    qhdr = tk.Frame(quota_inner, bg='#2c3e50')
    qhdr.pack(fill='x')
    for h, w in zip(quota_header_cols, quota_col_widths):
        tk.Label(
            qhdr, text=h, width=w, anchor='w',
            font=('Helvetica', 8, 'bold'),
            bg='#2c3e50', fg='white', padx=2, pady=3
        ).pack(side='left')

    quota_rows_frame = tk.Frame(quota_inner, bg='#f0f0f0')
    quota_rows_frame.pack(fill='both', expand=True)

    def _refresh_quota_panel():
        for w in quota_rows_frame.winfo_children():
            w.destroy()

        current = dict(selection_values)  # use plain-dict copy, not StringVar (avoids canvas decouple bug)
        quota_data = _compute_quota_status(current, schedulable, remaining_needs)

        if not quota_data:
            tk.Label(
                quota_rows_frame,
                text="No quota tracking data\navailable.",
                font=('Helvetica', 9), bg='#f0f0f0', fg='#7f8c8d',
                justify='center'
            ).pack(pady=20)
            return

        all_match = all(r[6] == '✓ Match' for r in quota_data)
        summary_color = '#c3e6cb' if all_match else '#ffeeba'
        summary_text  = '✓ All quotas satisfied' if all_match else '⚠ Some quotas need attention'
        tk.Label(
            quota_rows_frame,
            text=summary_text,
            font=('Helvetica', 9, 'bold'),
            bg=summary_color, fg='#155724' if all_match else '#856404',
            pady=4
        ).pack(fill='x', pady=(0, 4))

        for j, (strain_k, tp, sex, htype, needed, selected, status) in enumerate(quota_data):
            bg = '#ffffff' if j % 2 == 0 else '#f7f7f7'
            # Tint by status
            if status == '✓ Match':
                bg = '#eafaf1'
            elif status.startswith('↑'):
                bg = '#fef9e7'
            elif status.startswith('↓'):
                bg = '#fdf0ef'

            qrow = tk.Frame(quota_rows_frame, bg=bg)
            qrow.pack(fill='x')
            for val, w in zip(
                [strain_k, tp, sex, htype, str(needed), str(selected), status],
                quota_col_widths
            ):
                tk.Label(
                    qrow, text=val, width=w, anchor='w',
                    font=('Helvetica', 8), bg=bg, padx=2, pady=2
                ).pack(side='left')

    _refresh_quota_panel()

    # ── Footer buttons ────────────────────────────────────────────────────────
    footer = tk.Frame(root, bg='#ecf0f1', pady=8)
    footer.pack(fill='x', padx=12)

    def _reset_to_auto():
        for name, var in selection_vars.items():
            val = auto_types.get(name, 'Perfusion')
            var.set(val)
            selection_values[name] = val  # keep plain-dict in sync
            frame = row_frames[name]
            c = OPTION_COLORS.get(var.get(), '#ffffff')
            frame.configure(bg=c)
            for w in frame.winfo_children():
                try:
                    w.configure(bg=c)
                except Exception:
                    pass
        _refresh_quota_panel()

    def _confirm():
        current = dict(selection_values)  # use plain-dict copy, not StringVar (avoids canvas decouple bug)
        # Debug: log all non-Perfusion selections so we can verify they're captured
        for _n, _h in sorted(current.items()):
            if _h != 'Perfusion':
                logger.info(f"CONFIRM selection: {_n} → {_h}")

        # ── Quota check (harvest types only, not Extra or Do Not Schedule) ────
        quota_data = _compute_quota_status(current, schedulable, remaining_needs)
        mismatches = [r for r in quota_data if r[6] != '✓ Match']

        if mismatches:
            lines = '\n'.join(
                f"  {r[0]} {r[1]} {r[2]} {r[3]}: need {r[4]}, selected {r[5]} ({r[6]})"
                for r in mismatches[:8]
            )
            if len(mismatches) > 8:
                lines += f"\n  ... and {len(mismatches) - 8} more"
            proceed = messagebox.askyesno(
                "Quota Mismatch",
                f"The following assignments don't match the tracking sheet:\n\n"
                f"{lines}\n\n"
                f"Proceed anyway?",
                icon='warning'
            )
            if not proceed:
                return

        # ── Group-of-3 check (Harvest + Extra count per Birth_ID, P56 only) ──
        # Build a map of animal name -> row for quick lookup
        df_map = {
            str(r.get('Animal_Name', '')).strip(): r
            for _, r in schedulable.iterrows()
        }
        # Group animals by Birth_ID, P56 only
        birth_groups = {}
        for name, htype in current.items():
            row = df_map.get(name)
            if row is None:
                continue
            if str(row.get('Assigned_Timepoint', '')).strip() != 'P56':
                continue
            birth_id = str(row.get('Birth_ID', 'Unknown')).strip()
            if birth_id not in birth_groups:
                birth_groups[birth_id] = []
            birth_groups[birth_id].append((name, htype))

        incomplete_groups = []
        for birth_id, animals in birth_groups.items():
            # Count animals that are Harvest or Extra (not Do Not Schedule)
            active = [a for a in animals if a[1] != 'Do Not Schedule']
            if 0 < len(active) < CONFIG['CAGE_SIZE']:
                strain = str(df_map.get(animals[0][0], {}).get('Strain', '')).strip()
                incomplete_groups.append(
                    f"  Birth {birth_id} ({strain}): {len(active)} of {CONFIG['CAGE_SIZE']} animals active"
                )

        if incomplete_groups:
            lines = '\n'.join(incomplete_groups[:8])
            if len(incomplete_groups) > 8:
                lines += f"\n  ... and {len(incomplete_groups) - 8} more"
            proceed = messagebox.askyesno(
                "Incomplete Cage Groups",
                f"The following P56 cage groups have fewer than {CONFIG['CAGE_SIZE']} animals\n"
                f"(counting Harvest + Extra, excluding Do Not Schedule):\n\n"
                f"{lines}\n\n"
                f"Proceed anyway?",
                icon='warning'
            )
            if not proceed:
                return

        confirmed['result'] = current
        root.destroy()

    def _cancel():
        if messagebox.askyesno(
            "Skip Review",
            "Skip the harvest review and use auto-assignments for all animals?",
            icon='question'
        ):
            confirmed['result'] = {}
            root.destroy()

    tk.Button(
        footer, text="↺  Reset to Suggested",
        command=_reset_to_auto,
        font=('Helvetica', 10), bg='#95a5a6', fg='white',
        relief='flat', padx=12, pady=6, cursor='hand2'
    ).pack(side='left', padx=(0, 8))

    tk.Button(
        footer, text="Skip / Use Auto-Assignments",
        command=_cancel,
        font=('Helvetica', 10), bg='#bdc3c7', fg='#2c3e50',
        relief='flat', padx=12, pady=6, cursor='hand2'
    ).pack(side='left')

    tk.Button(
        footer, text="Confirm Assignments  →",
        command=_confirm,
        font=('Helvetica', 11, 'bold'), bg='#27ae60', fg='white',
        relief='flat', padx=16, pady=6, cursor='hand2'
    ).pack(side='right')

    # Size and center
    root.update_idletasks()
    w = min(root.winfo_screenwidth() - 80, 1100)
    h = min(root.winfo_screenheight() - 80, 700)
    x = (root.winfo_screenwidth()  - w) // 2
    y = (root.winfo_screenheight() - h) // 2
    root.geometry(f"{w}x{h}+{x}+{y}")
    root.minsize(700, 400)

    root.mainloop()

    # ── Process result ────────────────────────────────────────────────────────
    if confirmed['result'] is None:
        # Window was closed without confirming — treat as auto
        return {}

    # Convert 'Do Not Schedule' to sentinel; strip unchanged auto-assignments
    final = {}
    for name, htype in confirmed['result'].items():
        if htype == 'Do Not Schedule':
            final[name] = 'DO_NOT_SCHEDULE'
        else:
            final[name] = htype

    dns_count  = sum(1 for v in final.values() if v == 'DO_NOT_SCHEDULE')
    over_count = sum(1 for v in final.values() if v != 'DO_NOT_SCHEDULE')
    print(f"\n  ✓ Harvest review confirmed: {over_count} scheduled, {dns_count} skipped")
    return final



def write_harvest_overrides_template(assignments_df: pd.DataFrame, overrides_file: str) -> None:
    """
    Write a pre-filled harvest_overrides.csv from the current auto-assignments.

    Only written when the file does not already exist — so editing and re-running
    is safe; your changes won't be stomped.

    The CSV has one row per schedulable animal with the auto-assigned Harvest_Type
    already filled in.  Edit only the rows you want to change, then re-run.
    Rows for Unschedulable animals are included but commented out so you have
    the full picture without them cluttering the active list.
    """
    if os.path.exists(overrides_file):
        logger.info(f"harvest_overrides.csv already exists — not overwriting: {overrides_file}")
        return

    if assignments_df is None or len(assignments_df) == 0:
        return

    rows = []

    # Scheduled animals first — sorted by harvest date then name
    scheduled = assignments_df[
        assignments_df['Assigned_Timepoint'].isin(['P14', 'P56'])
    ].copy()

    # Determine the relevant harvest date for sorting
    def _harvest_date_for_sort(row):
        if row.get('Assigned_Timepoint') == 'P14':
            return str(row.get('P14_Date', '') or '')
        return str(row.get('P56_Harvest_Date', '') or '')

    if len(scheduled) > 0:
        scheduled = scheduled.copy()
        scheduled['_sort_date'] = scheduled.apply(_harvest_date_for_sort, axis=1)
        scheduled = scheduled.sort_values(['_sort_date', 'Strain', 'Animal_Name'])

        for _, row in scheduled.iterrows():
            name      = str(row.get('Animal_Name', '')).strip()
            htype     = str(row.get('Harvest_Type', '')).strip()
            timepoint = str(row.get('Assigned_Timepoint', '')).strip()
            strain    = str(row.get('Strain', '')).strip()
            sex       = str(row.get('Sex', '')).strip()
            priority  = str(row.get('Priority', '')).strip()

            # Skip auto quota-filled — not real harvests; keep Extra (user-assigned)
            if htype == 'COMPLETE (Quota Filled)':
                continue

            rows.append({
                'Animal_Name':        name,
                'Harvest_Type':       htype,
                'Assigned_Timepoint': timepoint,
                'Strain':             strain,
                'Sex':                sex,
                'Auto_Priority':      priority,
                'Notes':              '',
            })

    try:
        import csv as _csv
        fieldnames = ['Animal_Name', 'Harvest_Type', 'Assigned_Timepoint',
                      'Strain', 'Sex', 'Auto_Priority', 'Notes']

        with open(overrides_file, 'w', newline='', encoding='utf-8') as f:
            f.write(
                "# harvest_overrides.csv — edit Harvest_Type for any animal, then re-run.\n"
                "# Valid values: Perfusion   MERFISH   RNAseq\n"
                "# Leave Harvest_Type blank to keep the auto-assignment.\n"
                "# DO NOT change Animal_Name — it must match exactly.\n"
                "# Assigned_Timepoint, Strain, Sex, Auto_Priority, Notes are for reference only.\n"
                "#\n"
            )
            writer = _csv.DictWriter(f, fieldnames=fieldnames)
            writer.writeheader()
            writer.writerows(rows)

        print(f"\n  📋 harvest_overrides.csv written ({len(rows)} animals)")
        print(f"     → Edit Harvest_Type for any animal you want to change, then re-run.")
        print(f"     → File: {overrides_file}")
        logger.info(f"harvest_overrides.csv written: {overrides_file} ({len(rows)} rows)")

    except Exception as e:
        logger.error(f"Could not write harvest_overrides.csv: {e}")
        print(f"  ⚠ Could not write harvest_overrides.csv: {e}")


def load_harvest_overrides(overrides_file: Optional[str]) -> Dict[str, str]:
    """
    Load manual harvest-type overrides from a CSV file.

    Expected CSV columns (case-insensitive):
        Animal_Name   — the exact animal name as it appears in the schedule
        Harvest_Type  — one of: Perfusion, MERFISH, RNAseq
                        Leave blank / omit to let the scheduler auto-assign.

    Optional columns (both must be present to use):
        Assigned_Timepoint  — P14 or P56 (if omitted, override applies regardless)

    Returns a dict:  { 'AnimalName': 'HarvestType', ... }
    Only rows with a valid non-blank Harvest_Type are included.
    """
    VALID_TYPES = {'Perfusion', 'MERFISH', 'RNAseq', 'Extra'}

    if not overrides_file:
        return {}
    if not os.path.exists(overrides_file):
        logger.info(f"No harvest overrides file found at: {overrides_file} — auto-assigning all.")
        return {}

    try:
        df = pd.read_csv(overrides_file)
        # Normalise column names to lowercase for flexible matching
        df.columns = [c.strip().lower().replace(' ', '_') for c in df.columns]

        name_col = next((c for c in df.columns if 'animal' in c and 'name' in c), None)
        type_col = next((c for c in df.columns if 'harvest' in c and 'type' in c), None)

        if name_col is None or type_col is None:
            logger.warning(
                f"harvest_overrides.csv must have 'Animal_Name' and 'Harvest_Type' columns. "
                f"Found: {list(df.columns)}"
            )
            return {}

        overrides: Dict[str, str] = {}
        skipped = 0
        for _, row in df.iterrows():
            name = str(row[name_col]).strip() if pd.notna(row[name_col]) else ''
            htype = str(row[type_col]).strip() if pd.notna(row[type_col]) else ''

            if not name or name.lower() in ('nan', ''):
                continue
            if not htype or htype.lower() in ('nan', ''):
                continue  # blank = let scheduler decide

            # Case-insensitive match to valid types
            matched = next((v for v in VALID_TYPES if v.lower() == htype.lower()), None)
            if matched is None:
                logger.warning(
                    f"  Override for '{name}': '{htype}' is not a valid Harvest_Type "
                    f"(use Perfusion, MERFISH, or RNAseq). Skipping."
                )
                skipped += 1
                continue

            overrides[name] = matched

        loaded = len(overrides)
        print(f"  ✓ Loaded {loaded} harvest override(s) from {os.path.basename(overrides_file)}")
        if skipped:
            print(f"  ⚠ {skipped} override row(s) skipped — invalid Harvest_Type value")
        if loaded:
            for aname, htype in list(overrides.items())[:5]:
                print(f"    {aname!r:40s} → {htype}")
            if loaded > 5:
                print(f"    ... and {loaded - 5} more")
        return overrides

    except Exception as e:
        logger.error(f"Could not read harvest overrides file '{overrides_file}': {e}")
        print(f"  ⚠ Could not read harvest_overrides.csv: {e} — auto-assigning all.")
        return {}


def assign_harvest_types(assignments_df: pd.DataFrame,
                         remaining_needs: Dict,
                         requirements: Dict,
                         harvest_overrides: Optional[Dict[str, str]] = None) -> pd.DataFrame:
    if len(assignments_df) == 0:
        return assignments_df

    if harvest_overrides is None:
        harvest_overrides = {}

    if not remaining_needs:
        assignments_df = assignments_df.copy()
        assignments_df['Harvest_Type'] = assignments_df['Animal_Name'].map(harvest_overrides).fillna('Not Assigned')
        assignments_df['Priority'] = assignments_df['Animal_Name'].apply(
            lambda n: 'MANUAL' if n in harvest_overrides else 'Unknown'
        )
        assignments_df['Strain_Priority'] = 'Unknown'
        assignments_df['Genotype_Priority'] = 'Unknown'
        return assignments_df

    logger.info("Assigning harvest types...")
    working_needs = copy.deepcopy(remaining_needs)
    assignments_with_types = []

    sorted_df = assignments_df.sort_values(
        ['Strain', 'Sex', 'Assigned_Timepoint', 'Animal_Name']
    ).reset_index(drop=True)

    for idx, row in sorted_df.iterrows():
        strain    = row.get('Strain')
        sex       = row.get('Sex')
        genotype  = row.get('Genotype')
        timepoint = row.get('Assigned_Timepoint')
        name      = str(row.get('Animal_Name', '')).strip()

        is_prio = is_priority_strain(strain)
        strain_priority  = 'PRIORITY STRAIN' if is_prio else 'Standard'
        is_het           = is_heterozygous(genotype)
        genotype_priority = 'Het' if is_het else str(genotype) if genotype else 'Other'

        # ── Manual override — always wins, no quota consumed ─────────────────
        if name in harvest_overrides:
            manual_type = harvest_overrides[name]
            assignments_with_types.append({
                **row.to_dict(),
                'Harvest_Type':      manual_type,
                'Priority':          'MANUAL',
                'Strain_Priority':   strain_priority,
                'Genotype_Priority': genotype_priority,
            })
            logger.debug(f"Manual override applied: {name!r} → {manual_type}")
            continue

        if row.get('Priority') == 'B6_MIN':
            assignments_with_types.append({
                **row.to_dict(),
                'Strain_Priority':   row.get('Strain_Priority', 'B6/B6N Control'),
                'Genotype_Priority': row.get('Genotype_Priority', 'B6/B6N'),
            })
            continue

        if timepoint == 'Unschedulable':
            assignments_with_types.append({
                **row.to_dict(),
                'Harvest_Type':      'N/A',
                'Priority':          'N/A',
                'Strain_Priority':   strain_priority,
                'Genotype_Priority': genotype_priority,
            })
            continue

        strain_key = str(strain).strip().upper()

        if strain_key in _B6_STRAINS_UPPER:
            assignments_with_types.append({
                **row.to_dict(),
                'Harvest_Type':      'Perfusion',
                'Priority':          'B6_CONTROL',
                'Strain_Priority':   'B6/B6N Control',
                'Genotype_Priority': genotype_priority,
            })
            continue

        if strain_key not in working_needs:
            assignments_with_types.append({
                **row.to_dict(),
                'Harvest_Type':      'Not Tracked',
                'Priority':          'Unknown',
                'Strain_Priority':   strain_priority,
                'Genotype_Priority': genotype_priority,
            })
            continue

        needs = working_needs[strain_key][timepoint][sex]

        if needs['MERFISH']['needed'] > 0:
            harvest_type = 'MERFISH'
            priority = 'HIGH'
            needs['MERFISH']['needed'] -= 1
        elif needs['RNAseq']['needed'] > 0:
            harvest_type = 'RNAseq'
            priority = 'HIGH'
            needs['RNAseq']['needed'] -= 1
        elif needs['Perfusion']['needed'] > 0:
            harvest_type = 'Perfusion'
            priority = 'MEDIUM'
            needs['Perfusion']['needed'] -= 1
        else:
            harvest_type = 'COMPLETE (Quota Filled)'
            priority = 'NONE'

        assignments_with_types.append({
            **row.to_dict(),
            'Harvest_Type':      harvest_type,
            'Priority':          priority,
            'Strain_Priority':   strain_priority,
            'Genotype_Priority': genotype_priority,
        })

    overridden_count = sum(1 for r in assignments_with_types if r.get('Priority') == 'MANUAL')
    logger.info(f"Harvest types assigned: {len(assignments_with_types)} animals "
                f"({overridden_count} manually overridden)")
    return pd.DataFrame(assignments_with_types)


def check_capacity_and_reassign(assignments_df: pd.DataFrame,
                                remaining_needs: Dict) -> pd.DataFrame:
    if len(assignments_df) == 0:
        return assignments_df

    logger.info("Checking Wednesday capacity...")

    p56_assigned = assignments_df[assignments_df['Assigned_Timepoint'] == 'P56'].copy()
    if len(p56_assigned) == 0:
        return assignments_df

    p56_assigned = p56_assigned[p56_assigned['P56_Behavior_Date'].notna()].copy()
    try:
        p56_assigned['P56_Behavior_Date'] = p56_assigned['P56_Behavior_Date'].apply(to_date)
    except Exception as e:
        logger.warning(f"Error converting P56_Behavior_Date: {e}")
        return assignments_df

    p56_assigned = p56_assigned[p56_assigned['P56_Behavior_Date'].notna()].copy()
    if len(p56_assigned) == 0:
        return assignments_df

    wednesday_counts = p56_assigned.groupby('P56_Behavior_Date').size().to_dict()
    overflow_animals = []
    kept_animals = []

    for behavior_date, group in p56_assigned.groupby('P56_Behavior_Date'):
        animals = group.to_dict('records')
        count = wednesday_counts[behavior_date]

        if count <= CONFIG['WEDNESDAY_CAPACITY']:
            kept_animals.extend(animals)
        else:
            logger.warning(f"Wednesday {behavior_date} over capacity: {count} > {CONFIG['WEDNESDAY_CAPACITY']}")
            animals_sorted = sorted(animals, key=lambda x: not is_heterozygous(x.get('Genotype', '')))
            kept_animals.extend(animals_sorted[:CONFIG['WEDNESDAY_CAPACITY']])
            overflow_animals.extend(animals_sorted[CONFIG['WEDNESDAY_CAPACITY']:])

    reassigned = []
    still_unschedulable = []

    for animal in overflow_animals:
        if not animal.get('P14_Eligible'):
            animal['Assigned_Timepoint'] = 'Unschedulable'
            animal['Assignment_Reason'] = (
                f'Wednesday over capacity (>{CONFIG["WEDNESDAY_CAPACITY"]}); '
                f'P14 unavailable: {animal.get("P14_Reason", "Unknown")}'
            )
            still_unschedulable.append(animal)
            continue

        strain = animal.get('Strain')
        sex = animal.get('Sex')
        if group_has_quota(strain, sex, 'P14', remaining_needs):
            animal['Assigned_Timepoint'] = 'P14'
            animal['Assignment_Reason'] = (
                f'Wednesday over capacity (>{CONFIG["WEDNESDAY_CAPACITY"]}) — reassigned to P14'
            )
            reassigned.append(animal)
        else:
            animal['Assigned_Timepoint'] = 'Unschedulable'
            animal['Assignment_Reason'] = (
                f'Wednesday over capacity (>{CONFIG["WEDNESDAY_CAPACITY"]}); '
                f'P14 quota also filled'
            )
            still_unschedulable.append(animal)

    other_assignments = assignments_df[assignments_df['Assigned_Timepoint'] != 'P56']
    all_cols = assignments_df.columns.tolist()

    frames = [pd.DataFrame(kept_animals), other_assignments,
              pd.DataFrame(reassigned), pd.DataFrame(still_unschedulable)]
    for i, frame in enumerate(frames):
        if len(frame) > 0:
            for col in all_cols:
                if col not in frame.columns:
                    frame[col] = None

    final = pd.concat(frames, ignore_index=True)
    logger.info(f"Capacity check: {len(reassigned)} reassigned to P14, {len(still_unschedulable)} unschedulable")
    return final


# ============================================================================
# GENOTYPE BLANK ANALYSIS
# ============================================================================

def _assess_genotype_worth_it(
    num_blanks: int,
    breeding_type: str,
    p14_available: bool,
    p56_available: bool,
    is_schedulable: bool,
    p14_date,
    p56_date,
    today: date_type,
    remaining_needs: Dict,
    strain: str,
    # ── new parameter ────────────────────────────────────────────────────────
    p56_group_size: int = 0,    # total blank animals sharing this P56 Wednesday
) -> Dict:
    """
    Assess whether genotyping blank animals is worth doing for P14 / P56.

    P14 worth  — uses num_blanks (the animals in *this* exact-birth-date group).
    P56 worth  — uses p56_group_size (all blanks that share the same behavior
                 Wednesday, which may span several birth dates).  Falls back to
                 num_blanks when p56_group_size is not supplied (0).

    Returns a dict with keys:
        P14_Worth_Genotyping : str
        P56_Worth_Genotyping : str
    """
    min_cage = CONFIG['CAGE_SIZE']  # typically 3

    # ── quota / B6 helpers ────────────────────────────────────────────────────
    def _quota_met(timepoint: str) -> bool:
        """Return True when *all* needs for this strain/timepoint are zero."""
        if is_b6_strain(strain):
            return False            # B6/B6N never considered quota-met
        strain_upper = str(strain).strip().upper()
        if not remaining_needs or strain_upper not in remaining_needs:
            return False
        tp_needs = remaining_needs[strain_upper][timepoint]
        return all(
            tp_needs[sex][ht]['needed'] == 0
            for sex in ['Male', 'Female']
            for ht in ['Perfusion', 'MERFISH', 'RNAseq']
        )

    # ── generic worth evaluator ───────────────────────────────────────────────
    def _worth(available: bool, timepoint: str, group_n: int) -> str:
        """
        Evaluate whether genotyping is worth it for one timepoint.

        Parameters
        ----------
        available : bool   — is the scheduling window still open?
        timepoint : str    — 'P14' or 'P56'
        group_n   : int    — number of blank animals in the relevant group
                             (birth-date group for P14, Wednesday group for P56)
        """
        if not available:
            return '❌ NO — window not available'

        if is_b6_strain(strain):
            if breeding_type == 'All':
                return '✅ YES — All usable (B6/B6N control)'
            return '✅ YES — B6/B6N control'

        if _quota_met(timepoint):
            return '⚠️ QUOTA MET — genotyping low priority'

        # Expected usable animals from Mendelian ratios
        if breeding_type == 'All':
            expected = group_n
        elif breeding_type == 'Half':
            expected = group_n * 0.5
        else:
            expected = group_n * 0.5     # Unknown → assume Het×WT

        # For P56 the cage-size threshold matters: need ≥ CAGE_SIZE usable
        if timepoint == 'P56':
            if expected >= min_cage:
                return (
                    f'✅ YES — ~{int(round(expected))} usable expected '
                    f'from {group_n} animals in window '
                    f'(≥{min_cage} needed for a full cage)'
                )
            elif expected >= 1.0:
                return (
                    f'🟡 MAYBE — ~{int(round(expected))} usable expected '
                    f'from {group_n} animals in window '
                    f'(need {min_cage} for a full cage)'
                )
            else:
                return (
                    f'❌ UNLIKELY — <1 usable expected from {group_n} animals '
                    f'in window ({breeding_type} cross, need {min_cage} for a cage)'
                )
        else:
            # P14: simpler threshold — at least 1 usable is sufficient
            if expected >= 2.0:
                return f'✅ YES — ~{int(round(expected))} usable expected'
            elif expected >= 1.0:
                return f'🟡 MAYBE — ~{int(round(expected))} usable expected'
            else:
                return (
                    f'❌ UNLIKELY — <1 usable expected '
                    f'({group_n} blanks, {breeding_type} cross)'
                )

    # ── evaluate each timepoint with its own group size ───────────────────────
    effective_p56_n = p56_group_size if p56_group_size > 0 else num_blanks

    p14_worth = _worth(p14_available, 'P14', num_blanks)
    p56_worth = _worth(p56_available, 'P56', effective_p56_n)

    return {
        'P14_Worth_Genotyping': p14_worth,
        'P56_Worth_Genotyping': p56_worth,
    }


def analyze_blank_genotype_for_scheduling(
    blank_animals_group: List[Dict],
    strain: str,
    sex: str,
    birth_date: str,
    full_behavior_dates: Optional[List[date_type]] = None,
    remaining_needs: Optional[Dict] = None,
    p56_group_size: int = 0,    # total blank animals sharing this P56 Wednesday window
) -> Dict:
    """
    Analyse a group of blank-genotype animals from the same birth/strain/sex
    and predict scheduling viability.

    p56_group_size — when supplied, overrides num_blanks for the P56 worth-it
                     assessment so that animals with different birth dates that
                     map to the same Wednesday are evaluated together.
    """
    today               = datetime.now().date()
    num_blanks          = len(blank_animals_group)
    breeding_type       = get_strain_breeding_type(strain)
    birth_date_obj      = None
    scheduling_window   = 'Unknown'
    genotype_needed_by  = None
    p14_date            = None
    p56_behavior_date   = None
    is_schedulable      = False
    p14_available       = False
    p56_available       = False
    p14_valid           = False
    p56_valid           = False

    if blank_animals_group and 'Birth Date' in blank_animals_group[0]:
        birth_date_raw = blank_animals_group[0]['Birth Date']
        birth_date_obj = to_date(birth_date_raw)

    if birth_date_obj:
        dates = calculate_schedule_dates(birth_date_obj)
        if dates is None:
            scheduling_window = "Invalid birth date - cannot calculate scheduling windows"
            is_schedulable = False
        else:
            p14_date              = dates['p14_harvest']
            behavior_window_start = dates['p56_behavior_window_start']
            behavior_window_end   = dates['p56_behavior_window_end']
            p56_behavior_date     = next_wednesday(behavior_window_start)

            p14_valid     = is_valid_p14_day(p14_date)
            p14_in_future = p14_date > today
            p14_available = p14_valid and p14_in_future

            if p56_behavior_date is not None:
                p56_valid     = p56_behavior_date <= behavior_window_end
                p56_in_future = p56_behavior_date >= today
                p56_not_full  = True
                if full_behavior_dates and p56_behavior_date in full_behavior_dates:
                    p56_not_full = False
                p56_available = p56_valid and p56_in_future and p56_not_full
            else:
                p56_available = False

            possible_dates = []
            if p14_available:
                possible_dates.append(('P14', p14_date))
            if p56_available:
                possible_dates.append(('P56', p56_behavior_date))

            if possible_dates:
                earliest_type, earliest_date = min(possible_dates, key=lambda x: x[1])
                genotype_needed_by = earliest_date - timedelta(days=1)
                if len(possible_dates) == 2:
                    scheduling_window = (
                        f"P14 on {p14_date.strftime('%Y-%m-%d')} "
                        f"or P56 on {p56_behavior_date.strftime('%Y-%m-%d')}"
                    )
                else:
                    scheduling_window = (
                        f"{earliest_type} on {earliest_date.strftime('%Y-%m-%d')}"
                    )
            else:
                reasons = []
                if p14_date and p14_date <= today:
                    reasons.append(f"P14 window passed ({p14_date.strftime('%Y-%m-%d')})")
                elif p14_date and not p14_valid:
                    reasons.append(f"P14 on invalid day ({p14_date.strftime('%A')})")
                if p56_behavior_date and p56_behavior_date < today:
                    reasons.append(
                        f"P56 window passed ({p56_behavior_date.strftime('%Y-%m-%d')})"
                    )
                elif p56_behavior_date and not p56_valid:
                    reasons.append("P56 no valid Wednesday in window")
                elif (
                    p56_behavior_date and full_behavior_dates
                    and p56_behavior_date in full_behavior_dates
                ):
                    reasons.append(
                        f"P56 date full ({p56_behavior_date.strftime('%Y-%m-%d')})"
                    )
                elif p56_behavior_date is None:
                    reasons.append("Cannot calculate P56 behavior date")
                scheduling_window = (
                    "; ".join(reasons) if reasons else "No scheduling windows available"
                )
                genotype_needed_by = None

            is_schedulable = p14_available or p56_available
    else:
        is_schedulable = False

    worth_it = _assess_genotype_worth_it(
        num_blanks      = num_blanks,
        breeding_type   = breeding_type,
        p14_available   = p14_available,
        p56_available   = p56_available,
        is_schedulable  = is_schedulable,
        p14_date        = p14_date,
        p56_date        = p56_behavior_date,
        today           = today,
        remaining_needs = remaining_needs or {},
        strain          = strain,
        p56_group_size  = p56_group_size,   # ← Wednesday-level group size
    )

    prediction      = 'UNKNOWN'
    reason          = ''
    expected_usable = 0

    if breeding_type == 'All':
        expected_usable = num_blanks
        if not is_schedulable:
            prediction = 'NOT SCHEDULABLE'
            reason = (
                f"'All' STRAIN — {num_blanks} blank genotype(s) from birth {birth_date}. "
                f"ALL {num_blanks} expected usable (Hom×Hom/Inbred cross) "
                f"BUT NOT SCHEDULABLE. Reason: {scheduling_window}"
            )
        elif genotype_needed_by and genotype_needed_by >= today:
            days_until = (genotype_needed_by - today).days
            urgency    = "URGENT" if days_until <= 7 else "HIGH PRIORITY"
            prediction = 'LIKELY USABLE'
            reason = (
                f"'All' STRAIN — {num_blanks} blank genotype(s) from birth {birth_date}. "
                f"ALL {num_blanks} expected usable (Hom×Hom/Inbred). "
                f"{urgency}: Genotype by "
                f"{genotype_needed_by.strftime('%Y-%m-%d')} ({days_until} days) "
                f"for {scheduling_window}"
            )
        elif genotype_needed_by and genotype_needed_by < today:
            prediction = 'DEADLINE PASSED'
            reason = (
                f"[GENOTYPE DEADLINE PASSED] "
                f"'All' STRAIN — {num_blanks} blank genotype(s) from birth {birth_date}. "
                f"ALL should be usable but genotyping deadline passed "
                f"({genotype_needed_by.strftime('%Y-%m-%d')}). {scheduling_window}"
            )
        else:
            prediction = 'LIKELY USABLE'
            reason = (
                f"'All' STRAIN — {num_blanks} blank genotype(s) from birth {birth_date}. "
                f"ALL {num_blanks} expected usable (Hom×Hom/Inbred). "
                f"Genotype ASAP! {scheduling_window}"
            )

    elif breeding_type == 'Half':
        expected_hets = num_blanks * 0.5

        if not is_schedulable:
            prediction      = 'NOT SCHEDULABLE'
            expected_usable = 0
            reason = (
                f"'Half' STRAIN — {num_blanks} blank genotype(s) from birth {birth_date}. "
                f"~{int(round(expected_hets))} Hets expected (50% from Het×WT) "
                f"BUT NOT SCHEDULABLE. Reason: {scheduling_window}"
            )
        elif expected_hets >= 2.0:
            prediction      = 'LIKELY USABLE'
            expected_usable = int(round(expected_hets))
            if genotype_needed_by and genotype_needed_by >= today:
                days_until = (genotype_needed_by - today).days
                urgency    = "URGENT" if days_until <= 7 else "RECOMMEND"
                reason = (
                    f"'Half' STRAIN — {num_blanks} blank genotype(s) from birth {birth_date}. "
                    f"~{int(round(expected_hets))} of {num_blanks} expected Het "
                    f"(50% from Het×WT). {urgency}: Genotype by "
                    f"{genotype_needed_by.strftime('%Y-%m-%d')} ({days_until} days) "
                    f"for {scheduling_window}"
                )
            elif genotype_needed_by and genotype_needed_by < today:
                prediction = 'DEADLINE PASSED'
                reason = (
                    f"[GENOTYPE DEADLINE PASSED] "
                    f"'Half' STRAIN — {num_blanks} blank genotype(s) from birth {birth_date}. "
                    f"~{int(round(expected_hets))} Hets expected but genotyping deadline "
                    f"passed ({genotype_needed_by.strftime('%Y-%m-%d')}). {scheduling_window}"
                )
            else:
                reason = (
                    f"'Half' STRAIN — {num_blanks} blank genotype(s) from birth {birth_date}. "
                    f"~{int(round(expected_hets))} of {num_blanks} expected Het "
                    f"(50% from Het×WT). Genotype for scheduling! {scheduling_window}"
                )
        elif expected_hets >= 1.0:
            prediction      = 'POSSIBLY USABLE' if is_schedulable else 'NOT SCHEDULABLE'
            expected_usable = int(round(expected_hets)) if is_schedulable else 0
            if genotype_needed_by and genotype_needed_by >= today:
                days_until = (genotype_needed_by - today).days
                reason = (
                    f"'Half' STRAIN — {num_blanks} blank genotype(s) from birth {birth_date}. "
                    f"May contain ~{int(round(expected_hets))} Het "
                    f"(50% expected from Het×WT). Consider genotyping by "
                    f"{genotype_needed_by.strftime('%Y-%m-%d')} ({days_until} days). "
                    f"{scheduling_window}"
                )
            elif genotype_needed_by and genotype_needed_by < today:
                prediction = 'DEADLINE PASSED'
                reason = (
                    f"[GENOTYPE DEADLINE PASSED] "
                    f"'Half' STRAIN — {num_blanks} blank genotype(s) from birth {birth_date}. "
                    f"May have had ~{int(round(expected_hets))} Het but genotyping deadline "
                    f"passed ({genotype_needed_by.strftime('%Y-%m-%d')}). {scheduling_window}"
                )
            else:
                reason = (
                    f"'Half' STRAIN — {num_blanks} blank genotype(s) from birth {birth_date}. "
                    f"May contain ~{int(round(expected_hets))} Het. "
                    f"Consider genotyping. {scheduling_window}"
                )
        else:
            prediction      = 'LIKELY WILD' if is_schedulable else 'NOT SCHEDULABLE'
            expected_usable = 0
            if genotype_needed_by and genotype_needed_by >= today:
                days_until = (genotype_needed_by - today).days
                reason = (
                    f"'Half' STRAIN — {num_blanks} blank genotype(s) from birth {birth_date}. "
                    f"50% chance Het, 50% chance Wild (Het×WT cross). "
                    f"Low statistical likelihood of usable animals. "
                    f"Deadline: {genotype_needed_by.strftime('%Y-%m-%d')} ({days_until} days). "
                    f"{scheduling_window}"
                )
            elif genotype_needed_by and genotype_needed_by < today:
                prediction = 'DEADLINE PASSED'
                reason = (
                    f"[GENOTYPE DEADLINE PASSED] "
                    f"'Half' STRAIN — {num_blanks} blank genotype(s) from birth {birth_date}. "
                    f"50%/50% Het/Wild expected. Genotyping deadline passed "
                    f"({genotype_needed_by.strftime('%Y-%m-%d')}). {scheduling_window}"
                )
            else:
                reason = (
                    f"'Half' STRAIN — {num_blanks} blank genotype(s) from birth "
                    f"{birth_date}. 50% chance Het, 50% chance Wild. "
                    f"Low likelihood. {scheduling_window}"
                )
    else:
        expected_hets = num_blanks * 0.5
        prediction    = 'UNKNOWN' if is_schedulable else 'NOT SCHEDULABLE'
        if not is_schedulable:
            expected_usable = 0
            reason = (
                f"UNKNOWN STRAIN — {num_blanks} blank genotype(s) from birth {birth_date}. "
                f"Not in priority list. NOT SCHEDULABLE. {scheduling_window}"
            )
        elif genotype_needed_by and genotype_needed_by >= today:
            days_until      = (genotype_needed_by - today).days
            expected_usable = int(round(expected_hets))
            reason = (
                f"UNKNOWN STRAIN — {num_blanks} blank genotype(s) from birth {birth_date}. "
                f"Assuming Het×WT (~{int(round(expected_hets))} Hets). "
                f"Genotype by {genotype_needed_by.strftime('%Y-%m-%d')} ({days_until} days). "
                f"{scheduling_window}"
            )
        elif genotype_needed_by and genotype_needed_by < today:
            prediction      = 'DEADLINE PASSED'
            expected_usable = 0
            reason = (
                f"[GENOTYPE DEADLINE PASSED] "
                f"UNKNOWN STRAIN — {num_blanks} blank genotype(s) from birth {birth_date}. "
                f"Genotyping deadline passed "
                f"({genotype_needed_by.strftime('%Y-%m-%d')}). {scheduling_window}"
            )
        else:
            expected_usable = int(round(expected_hets))
            reason = (
                f"UNKNOWN STRAIN — {num_blanks} blank genotype(s) from birth {birth_date}. "
                f"Assuming ~{int(round(expected_hets))} usable. {scheduling_window}"
            )

    return {
        'prediction':           prediction,
        'reason':               reason,
        'expected_usable':      expected_usable,
        'breeding_type':        breeding_type,
        'genotype_needed_by':   genotype_needed_by.strftime('%Y-%m-%d') if genotype_needed_by else 'N/A',
        'days_until_deadline':  (
            (genotype_needed_by - today).days
            if genotype_needed_by and genotype_needed_by >= today else None
        ),
        'scheduling_window':    scheduling_window,
        'p14_date':             p14_date.strftime('%Y-%m-%d') if p14_date else 'N/A',
        'p56_date':             p56_behavior_date.strftime('%Y-%m-%d') if p56_behavior_date else 'N/A',
        'p14_available':        p14_available,
        'p56_available':        p56_available,
        'is_schedulable':       is_schedulable,
        'P14_Worth_Genotyping': worth_it['P14_Worth_Genotyping'],
        'P56_Worth_Genotyping': worth_it['P56_Worth_Genotyping'],
    }


def analyze_blank_genotypes_second_pass(
    blank_genotypes_df: pd.DataFrame,
    full_behavior_dates: List[date_type],
    remaining_needs: Optional[Dict] = None,
) -> pd.DataFrame:
    """
    Second pass: analyse blank-genotype animals with capacity info.

    Grouping strategy
    -----------------
    * P14 worth-it  — assessed per exact (birth_date, strain, sex) group,
                      because P14 scheduling is birth-date-specific.
    * P56 worth-it  — assessed per (p56_wednesday, strain, sex) group,
                      because all animals whose birth dates map to the same
                      Wednesday behavior session compete for the same cages.
                      Example: 18 females born 3/17-3/24 all map to the
                      5/6/26 Wednesday → evaluated as one group of 18,
                      yielding ~9 expected Hets which fills 3 cages → YES.

    NOTE: This function operates exclusively on blank-genotype animals.
          Genotyped animals (Het, Hom, Hemi, Inbred) are never passed here
          and are completely unaffected by this logic.
    """
    if len(blank_genotypes_df) == 0:
        return pd.DataFrame()

    logger.info("Second pass: Analyzing blank genotypes with capacity info...")

    # ── Step 1: compute P56 Wednesday for every animal ───────────────────────
    def _p56_wednesday(birth_date_val) -> Optional[date_type]:
        bd = to_date(birth_date_val) if pd.notna(birth_date_val) else None
        return get_p56_behavior_wednesday(bd)

    blank_genotypes_df = blank_genotypes_df.copy()
    blank_genotypes_df['_p56_wed'] = blank_genotypes_df['Birth Date'].apply(_p56_wednesday)

    # ── Step 2: pre-compute P56 group sizes (wednesday, strain, sex) ─────────
    # A "P56 group" = all blank animals of the same strain & sex whose birth
    # dates map to the same behavior Wednesday.
    p56_group_sizes: Dict[Tuple, int] = {}
    for (p56_wed, strain, sex), grp in blank_genotypes_df.groupby(
        ['_p56_wed', 'Line (Short)', 'Sex'], dropna=False
    ):
        key = (p56_wed, str(strain), str(sex))
        p56_group_sizes[key] = len(grp)

    # ── Step 3: iterate over exact (birth_date, strain, sex) sub-groups ──────
    excluded = []
    grouped = list(
        blank_genotypes_df.groupby(['Birth Date', 'Line (Short)', 'Sex'], dropna=False)
    )
    if CONFIG['ENABLE_PROGRESS_BARS']:
        grouped = tqdm(grouped, desc="Analyzing blank genotype groups")

    for (birth_date_val, strain, sex), group in grouped:
        group_animals = group.to_dict('records')

        birth_date_str = (
            to_date(birth_date_val).strftime('%Y-%m-%d')
            if pd.notna(birth_date_val) and to_date(birth_date_val) is not None
            else 'Unknown Date'
        )

        # P56 group size: look up by the Wednesday this birth date maps to
        p56_wed = _p56_wednesday(birth_date_val)
        p56_key = (p56_wed, str(strain), str(sex))
        p56_group_n = p56_group_sizes.get(p56_key, len(group_animals))

        analysis = analyze_blank_genotype_for_scheduling(
            group_animals,
            strain,
            sex,
            birth_date_str,
            full_behavior_dates=full_behavior_dates,
            remaining_needs=remaining_needs,
            p56_group_size=p56_group_n,     # ← Wednesday-level group size
        )

        # ── build one output row per animal ───────────────────────────────────
        for animal in group_animals:
            excluded.append({
                'Animal_Name':              animal['Name'],
                'Birth_ID':                 animal.get('Birth ID', 'N/A'),
                'Birth_Date':               birth_date_str,
                'Strain':                   strain,
                'Sex':                      sex,
                'Genotype':                 GENOTYPE_BLANK,
                'Breeding_Type':            analysis['breeding_type'],
                'Prediction':               analysis['prediction'],
                'Expected_Usable_In_Group': analysis['expected_usable'],
                # ── P56 window group info (new columns) ────────────────────
                'P56_Window_Group_Size':    p56_group_n,
                'P56_Window_Wednesday':     (
                    p56_wed.strftime('%Y-%m-%d') if p56_wed else 'N/A'
                ),
                # ── deadline / scheduling ──────────────────────────────────
                'Genotype_Needed_By':       analysis['genotype_needed_by'],
                'Days_Until_Deadline':      (
                    analysis['days_until_deadline']
                    if analysis['days_until_deadline'] is not None
                    else 'N/A'
                ),
                'P14_Date':                 analysis['p14_date'],
                'P14_Worth_Genotyping':     analysis['P14_Worth_Genotyping'],
                'P56_Date':                 analysis['p56_date'],
                'P56_Worth_Genotyping':     analysis['P56_Worth_Genotyping'],
                'Scheduling_Window':        analysis['scheduling_window'],
                'Reason':                   analysis['reason'],
            })

    logger.info(f"Analyzed {len(excluded)} animals with blank genotypes")
    return pd.DataFrame(excluded)


def summarize_genotype_exclusions(genotype_excluded_df: pd.DataFrame) -> pd.DataFrame:
    if len(genotype_excluded_df) == 0:
        return pd.DataFrame()

    summary_data = []

    summary_data.append({
        'Category': 'Total Excluded',
        'Count': len(genotype_excluded_df),
        'Details': 'All animals excluded due to genotype issues'
    })

    if 'Prediction' in genotype_excluded_df.columns:
        for prediction_val in genotype_excluded_df['Prediction'].value_counts().index:
            count = len(genotype_excluded_df[genotype_excluded_df['Prediction'] == prediction_val])
            summary_data.append({
                'Category': f'Prediction: {prediction_val}',
                'Count': count,
                'Details': f'{count} animals with this prediction'
            })

    if 'Expected_Usable_In_Group' in genotype_excluded_df.columns:
        total_expected = 0
        for val in genotype_excluded_df['Expected_Usable_In_Group']:
            try:
                total_expected += int(val)
            except (ValueError, TypeError):
                pass
        if total_expected > 0:
            summary_data.append({
                'Category': '📊 STATISTICAL PREDICTION',
                'Count': int(total_expected),
                'Details': f'~{int(total_expected)} usable animals expected among blanks (Mendelian ratios)'
            })

    if 'Prediction' in genotype_excluded_df.columns:
        likely = genotype_excluded_df[genotype_excluded_df['Prediction'] == 'LIKELY USABLE']
        possibly = genotype_excluded_df[genotype_excluded_df['Prediction'] == 'POSSIBLY USABLE']
        not_sched = genotype_excluded_df[genotype_excluded_df['Prediction'] == 'NOT SCHEDULABLE']

        if len(likely) > 0:
            summary_data.append({
                'Category': '🔍 LIKELY USABLE (blank genotype)',
                'Count': len(likely),
                'Details': f'{len(likely)} animals — RECOMMEND GENOTYPING for P14/P56'
            })
        if len(possibly) > 0:
            summary_data.append({
                'Category': '🔍 POSSIBLY USABLE (blank genotype)',
                'Count': len(possibly),
                'Details': f'{len(possibly)} animals — Consider genotyping'
            })
        if len(not_sched) > 0:
            summary_data.append({
                'Category': '🚫 NOT SCHEDULABLE (blank genotype)',
                'Count': len(not_sched),
                'Details': f'{len(not_sched)} animals — No available P14/P56 windows'
            })

    if 'Reason' in genotype_excluded_df.columns:
        wild_count = len(genotype_excluded_df[
            genotype_excluded_df['Reason'].str.contains('Wild genotype', na=False)])
        cre_wild_count = len(genotype_excluded_df[
            genotype_excluded_df['Reason'].str.contains('Cre-only', na=False, case=False)])
        critical_count = len(genotype_excluded_df[
            genotype_excluded_df['Reason'].str.contains('⚠️', na=False)])

        summary_data.append({
            'Category': 'Wild genotype',
            'Count': wild_count,
            'Details': 'Animals with Wild genotype'
        })
        summary_data.append({
            'Category': 'Cre-only Wild',
            'Count': cre_wild_count,
            'Details': 'Generic Cre, no mutation of interest'
        })
        summary_data.append({
            'Category': '⚠️ CRITICAL ISSUES',
            'Count': critical_count,
            'Details': 'Possible breeding pair errors'
        })

    return pd.DataFrame(summary_data)


# ============================================================================
# UNSCHEDULABLE REASON PARSING
# ============================================================================

def parse_unschedulable_reason(reason) -> Dict:
    """
    Parse a raw assignment reason string into structured fields for the
    Unschedulable report.
    """
    if reason is None:
        return {
            'Primary_Reason': 'Unknown',
            'P14_Status':     'Unknown',
            'P56_Status':     'Unknown',
            'Too_Old_For_P14': 'NO',
            'Too_Old_For_P56': 'NO',
            'Unusable_Both':   'NO',
            'Detail':          '',
        }

    r  = str(reason).strip()
    rl = r.lower()

    too_old_p14   = 'NO'
    too_old_p56   = 'NO'
    unusable_both = 'NO'
    primary       = 'Unknown'
    p14_status    = 'See detail'
    p56_status    = 'See detail'

    # ── [GENOTYPE DEADLINE PASSED] sentinel — must be checked FIRST ──────────
    if '[GENOTYPE DEADLINE PASSED]' in r:
        primary    = '🧬 Genotype Deadline Passed'
        p14_status = '🧬 Deadline passed'
        p56_status = '🧬 Deadline passed'
        return {
            'Primary_Reason':  primary,
            'P14_Status':      p14_status,
            'P56_Status':      p56_status,
            'Too_Old_For_P14': too_old_p14,
            'Too_Old_For_P56': too_old_p56,
            'Unusable_Both':   unusable_both,
            'Detail':          r,
        }

    # ── Blank / inconclusive genotype — check BEFORE wild ────────────────────
    _blank_reason_indicators = (
        r == GENOTYPE_BLANK,
        'inconclusive' in rl,
        'pending' in rl,
        rl == 'blank',
        r.startswith('Blank'),
        'regenotype' in rl,
        'no genotype' in rl,
        'genotype not available' in rl,
        "'half' strain" in rl and 'blank genotype' in rl,
        "'all' strain" in rl and 'blank genotype' in rl,
        'unknown strain' in rl and 'blank genotype' in rl,
        'blank genotype' in rl,
    )
    if any(_blank_reason_indicators):
        primary    = '🧬 Blank / Inconclusive Genotype — Regenotype Needed'
        p14_status = '🧬 Regenotype'
        p56_status = '🧬 Regenotype'
        return {
            'Primary_Reason':  primary,
            'P14_Status':      p14_status,
            'P56_Status':      p56_status,
            'Too_Old_For_P14': too_old_p14,
            'Too_Old_For_P56': too_old_p56,
            'Unusable_Both':   unusable_both,
            'Detail':          r,
        }

    # ── Unusable for BOTH ─────────────────────────────────────────────────────
    if '⛔' in r or 'unusable for both' in rl:
        too_old_p14   = 'YES'
        too_old_p56   = 'YES'
        unusable_both = 'YES'
        primary       = '⛔ Unusable for BOTH Timepoints (Too Old)'
        p14_status    = '❌ Too old'
        p56_status    = '❌ Too old'

    # ── Too late / too old for P14 ────────────────────────────────────────────
    elif 'too late for p14' in rl or ('too late' in rl and 'p14' in rl):
        too_old_p14 = 'YES'
        primary     = '❌ Too Late For P14 (Today)'
        p14_status  = '❌ Too late (today)'
        p56_status  = 'See detail'

    elif 'too old for p14' in rl:
        too_old_p14 = 'YES'
        primary     = '❌ Too Old For P14'
        p14_status  = '❌ Too old'
        if 'too old for p56' in rl:
            too_old_p56   = 'YES'
            unusable_both = 'YES'
            primary       = '⛔ Unusable for BOTH Timepoints (Too Old)'
            p56_status    = '❌ Too old'
        else:
            p56_status = 'See detail'

    # ── Too old for P56 ───────────────────────────────────────────────────────
    elif 'too old for p56' in rl:
        too_old_p56 = 'YES'
        primary     = '❌ Too Old For P56'
        p56_status  = '❌ Too old'
        p14_status  = 'See detail'

    # ── Quota filled ──────────────────────────────────────────────────────────
    elif 'quota' in rl and ('filled' in rl or 'complete' in rl or 'met' in rl):
        primary    = '✅ Quota Filled / Complete'
        p14_status = '✅ Quota met'
        p56_status = '✅ Quota met'

    # ── Incomplete cage group ─────────────────────────────────────────────────
    elif 'incomplete' in rl and ('group' in rl or 'cage' in rl):
        primary    = f'⚠️ Incomplete P56 Group (< {CONFIG["CAGE_SIZE"]} animals)'
        p56_status = f'⚠️ Incomplete group < {CONFIG["CAGE_SIZE"]}'
        if 'p14 unavailable' in rl:
            p14_status = '❌ Unavailable'
        elif 'p14' in rl and ('passed' in rl or 'too old' in rl):
            too_old_p14 = 'YES'
            p14_status  = '❌ Too old / passed'
        elif 'quota' in rl:
            p14_status  = '✅ Quota met'
        else:
            p14_status  = 'See detail'

    # ── Capacity / overflow ───────────────────────────────────────────────────
    elif 'capacity' in rl or 'over capacity' in rl:
        primary    = '🔴 P56 Date at Capacity / Overflow'
        p56_status = '🔴 Date full'
        if 'p14 unavailable' in rl:
            p14_status = '❌ Unavailable'
        elif 'p14 quota' in rl:
            p14_status = '✅ Quota met'
        else:
            p14_status = 'See detail'

    # ── Invalid harvest day ───────────────────────────────────────────────────
    elif 'invalid day' in rl or 'falls on' in rl or 'valid harvest day' in rl:
        primary    = '⚠️ P14 Falls on Invalid Day (Weekend)'
        p14_status = '⚠️ Invalid day'
        if 'p56' in rl:
            if 'passed' in rl:
                p56_status = '❌ Window passed'
            elif 'no wednesday' in rl:
                p56_status = '❌ No Wed in window'
            elif 'toe clip' in rl:
                p56_status = '❌ Toe clip'
            else:
                p56_status = 'See detail'
        else:
            p56_status = 'N/A'

    # ── Both windows passed ───────────────────────────────────────────────────
    elif 'p14' in rl and 'passed' in rl and 'p56' in rl and 'passed' in rl:
        primary    = '❌ Both Windows Have Passed'
        p14_status = '❌ Window passed'
        p56_status = '❌ Window passed'

    # ── P14 window passed only ────────────────────────────────────────────────
    elif 'p14' in rl and 'passed' in rl and 'p56' not in rl:
        primary    = '❌ P14 Window Passed'
        p14_status = '❌ Window passed'
        p56_status = 'N/A'

    # ── P56 window passed only ────────────────────────────────────────────────
    elif 'p56' in rl and 'passed' in rl and 'p14' not in rl:
        primary    = '❌ P56 Window Passed'
        p14_status = 'N/A'
        p56_status = '❌ Window passed'

    # ── No Wednesday in P56 window ────────────────────────────────────────────
    elif 'no wednesday' in rl or 'p42' in rl or 'p42–p49' in rl:
        primary    = '❌ No Wednesday in P56 Window (P42–P49)'
        p56_status = '❌ No Wed in window'
        if 'p14' in rl:
            p14_status = 'See detail'

    # ── Toe clip ──────────────────────────────────────────────────────────────
    elif 'toe clip' in rl:
        primary    = '🚫 Toe Clip — Not Allowed for P56 Behavior'
        p56_status = '🚫 Toe clip'
        if 'p14' in rl:
            if 'passed' in rl or 'too old' in rl or 'too late' in rl:
                too_old_p14 = 'YES'
                p14_status  = '❌ Too old / passed'
            elif 'quota' in rl:
                p14_status = '✅ Quota met'
            elif 'invalid' in rl or 'falls on' in rl:
                p14_status = '⚠️ Invalid day'
            else:
                p14_status = 'See detail'
        else:
            p14_status = 'N/A'

    # ── Sing inventory ────────────────────────────────────────────────────────
    elif 'sing inventory' in rl:
        primary    = '🔒 Not Assigned to Sing Inventory'
        p14_status = 'N/A'
        p56_status = 'N/A'

    # ── No birth date ─────────────────────────────────────────────────────────
    elif 'no birth date' in rl:
        primary    = '❓ No Birth Date Recorded'
        p14_status = '❓ No birth date'
        p56_status = '❓ No birth date'

    # ── Invalid birth date ────────────────────────────────────────────────────
    elif 'invalid birth date' in rl:
        primary    = '❓ Invalid Birth Date'
        p14_status = '❓ Invalid date'
        p56_status = '❓ Invalid date'

    # ── Wild genotype — STRICT matching only ─────────────────────────────────
    elif (
        r == GENOTYPE_WILD
        or rl == 'wild genotype — not usable for harvest'
        or rl == 'wild genotype - not usable for harvest'
        or rl.startswith('wild genotype')
        or ('+/+' in r and len(r) < 20)
    ):
        primary    = '🧬 Wild Genotype — Not Usable'
        p14_status = '🧬 Wild excluded'
        p56_status = '🧬 Wild excluded'

    # ── Cre-only wild ─────────────────────────────────────────────────────────
    elif 'cre' in rl and ('wildtype' in rl or 'ncar' in rl or 'cre-only' in rl):
        primary    = '🧬 Cre-Only Wild — No Mutation of Interest'
        p14_status = '🧬 Cre-only Wild'
        p56_status = '🧬 Cre-only Wild'

    # ── Wednesday over capacity (overflow) ───────────────────────────────────
    elif 'wednesday over capacity' in rl:
        primary    = '🔴 Wednesday Over Capacity — Overflow'
        p56_status = '🔴 Over capacity'
        if 'p14 unavailable' in rl:
            p14_status = '❌ Unavailable'
        elif 'p14 quota' in rl:
            p14_status = '✅ Quota met'
        else:
            p14_status = 'See detail'

    # ── Fallback ──────────────────────────────────────────────────────────────
    else:
        primary    = r[:80] + ('…' if len(r) > 80 else '')
        p14_status = 'See detail'
        p56_status = 'See detail'

    return {
        'Primary_Reason':  primary,
        'P14_Status':      p14_status,
        'P56_Status':      p56_status,
        'Too_Old_For_P14': too_old_p14,
        'Too_Old_For_P56': too_old_p56,
        'Unusable_Both':   unusable_both,
        'Detail':          r,
    }


# ============================================================================
# UNSCHEDULABLE REPORT
# ============================================================================

def create_unschedulable_report(assignments_df: pd.DataFrame,
                                use_excluded_df: pd.DataFrame,
                                genotype_excluded_df: pd.DataFrame,
                                date_excluded_df: Optional[pd.DataFrame] = None) -> pd.DataFrame:
    logger.info("Creating unschedulable report...")
    report_rows = []

    def _resolve_genotype(row: pd.Series) -> str:
        """
        Safely resolve the Genotype field to a canonical label.
        Blank/NaN/empty → GENOTYPE_BLANK. Never coerces blank to Wild.
        """
        geno = row.get('Genotype', None)

        if geno is None:
            return GENOTYPE_BLANK
        try:
            if pd.isna(geno):
                return GENOTYPE_BLANK
        except (TypeError, ValueError):
            pass

        geno_str = str(geno).strip()

        if geno_str == '' or geno_str.lower() in ('nan', 'none', 'n/a', 'na', '-'):
            return GENOTYPE_BLANK

        if geno_str in _CANONICAL_GENOTYPES:
            return geno_str

        strain = row.get('Strain', row.get('Line (Short)', ''))
        return canonicalize_genotype(geno_str, strain)

    # ── From unschedulable assignments ────────────────────────────────────────
    if len(assignments_df) > 0:
        unschedulable = assignments_df[
            assignments_df['Assigned_Timepoint'] == 'Unschedulable'
        ].copy()

        for _, row in unschedulable.iterrows():
            raw_reason    = row.get('Assignment_Reason', '')
            unusable_note = row.get('Unusable_Note', '')

            combined_reason = raw_reason
            if unusable_note and str(unusable_note) not in str(raw_reason):
                combined_reason = f"{unusable_note} | {raw_reason}"

            parsed = parse_unschedulable_reason(combined_reason)

            report_rows.append({
                'Animal_Name':              row.get('Animal_Name', 'Unknown'),
                'Birth_ID':                 row.get('Birth_ID', 'N/A'),
                'Strain':                   row.get('Strain', 'N/A'),
                'Genotype':                 _resolve_genotype(row),
                'Sex':                      row.get('Sex', 'N/A'),
                'Marker_Type':              row.get('Marker_Type', 'N/A'),
                'Birth_Date':               row.get('Birth_Date', 'N/A'),
                'Age_Today_Days':           row.get('Age_Today_Days', 'N/A'),
                'Primary_Reason':           parsed['Primary_Reason'],
                'P14_Status':               parsed['P14_Status'],
                'P56_Status':               parsed['P56_Status'],
                'Too_Old_For_P14':          parsed['Too_Old_For_P14'],
                'Too_Old_For_P56':          parsed['Too_Old_For_P56'],
                'Unusable_Both_Timepoints': parsed['Unusable_Both'],
                'Full_Detail':              parsed['Detail'],
            })

    # ── From use exclusions ───────────────────────────────────────────────────
    if len(use_excluded_df) > 0:
        for _, row in use_excluded_df.iterrows():
            report_rows.append({
                'Animal_Name':              row.get('Animal_Name', 'Unknown'),
                'Birth_ID':                 row.get('Birth_ID', 'N/A'),
                'Strain':                   row.get('Strain', 'N/A'),
                'Genotype':                 _resolve_genotype(row),
                'Sex':                      row.get('Sex', 'N/A'),
                'Marker_Type':              'N/A',
                'Birth_Date':               'N/A',
                'Age_Today_Days':           'N/A',
                'Primary_Reason':           '🔒 Not Assigned to Sing Inventory',
                'P14_Status':               'N/A',
                'P56_Status':               'N/A',
                'Too_Old_For_P14':          '',
                'Too_Old_For_P56':          '',
                'Unusable_Both_Timepoints': '',
                'Full_Detail':              row.get('Reason', ''),
            })

    # ── From genotype exclusions ──────────────────────────────────────────────
    if len(genotype_excluded_df) > 0:
        for _, row in genotype_excluded_df.iterrows():
            raw_reason = row.get('Reason', '')
            parsed = parse_unschedulable_reason(raw_reason)
            report_rows.append({
                'Animal_Name':              row.get('Animal_Name', 'Unknown'),
                'Birth_ID':                 row.get('Birth_ID', 'N/A'),
                'Strain':                   row.get('Strain', 'N/A'),
                'Genotype':                 _resolve_genotype(row),
                'Sex':                      row.get('Sex', 'N/A'),
                'Marker_Type':              'N/A',
                'Birth_Date':               row.get('Birth_Date', 'N/A'),
                'Age_Today_Days':           'N/A',
                'Primary_Reason':           parsed['Primary_Reason'],
                'P14_Status':               parsed['P14_Status'],
                'P56_Status':               parsed['P56_Status'],
                'Too_Old_For_P14':          parsed['Too_Old_For_P14'],
                'Too_Old_For_P56':          parsed['Too_Old_For_P56'],
                'Unusable_Both_Timepoints': parsed['Unusable_Both'],
                'Full_Detail':              raw_reason,
            })

    # ── From date exclusions ──────────────────────────────────────────────────
    if date_excluded_df is not None and len(date_excluded_df) > 0:
        for _, row in date_excluded_df.iterrows():
            raw_reason = row.get('Reason', '')
            report_rows.append({
                'Animal_Name':              row.get('Animal_Name', 'Unknown'),
                'Birth_ID':                 row.get('Birth_ID', 'N/A'),
                'Strain':                   row.get('Strain', 'N/A'),
                'Genotype':                 _resolve_genotype(row),
                'Sex':                      row.get('Sex', 'N/A'),
                'Marker_Type':              'N/A',
                'Birth_Date':               row.get('Birth_Date', 'N/A'),
                'Age_Today_Days':           'N/A',
                'Primary_Reason':           '📅 Outside Date Filter Range',
                'P14_Status':               'Filtered out',
                'P56_Status':               'Filtered out',
                'Too_Old_For_P14':          '',
                'Too_Old_For_P56':          '',
                'Unusable_Both_Timepoints': '',
                'Full_Detail':              raw_reason,
            })

    report = pd.DataFrame(report_rows)

    if len(report) > 0:
        priority_map = {
            '⛔': 0, '❌': 1, '🔴': 2, '⚠️': 3,
            '✅': 4, '🔒': 5, '🧬': 6, '📅': 7, '❓': 8,
        }

        def sort_key(val):
            for emoji, rank in priority_map.items():
                if str(val).startswith(emoji):
                    return rank
            return 9

        report['_sort'] = report['Primary_Reason'].apply(sort_key)
        report = report.sort_values(['_sort', 'Strain', 'Animal_Name'])
        report = report.drop(columns=['_sort'])

    logger.info(f"Unschedulable report: {len(report)} animals")
    return report


# ============================================================================
# SCHEDULE CREATION
# ============================================================================

def create_p14_schedule(assignments_df: pd.DataFrame) -> pd.DataFrame:
    if len(assignments_df) == 0:
        return pd.DataFrame()

    p14 = assignments_df[assignments_df['Assigned_Timepoint'] == 'P14'].copy()
    if len(p14) == 0:
        return pd.DataFrame()

    p14 = p14[~p14['Harvest_Type'].isin(['COMPLETE (Quota Filled)', 'Extra'])].copy()
    if len(p14) == 0:
        return pd.DataFrame()

    p14['Day_of_Week'] = pd.to_datetime(p14['P14_Date']).dt.day_name()

    p14 = p14.sort_values(
        ['P14_Date', 'Genotype_Priority', 'Sex', 'Genotype', 'Animal_Name'],
        ascending=[True, False, True, True, True]
    )

    desired_cols = [
        'P14_Date', 'Day_of_Week',
        'P14_Age_At_Harvest_Days', 'P14_Age_At_Harvest_Months',
        'Animal_Name', 'Strain', 'Strain_Priority',
        'Genotype', 'Genotype_Priority',
        'Sex', 'Marker_Type', 'Harvest_Type', 'Priority',
        'Birth_Date', 'Birth_ID', 'Assignment_Reason'
    ]
    available_cols = [c for c in desired_cols if c in p14.columns]

    logger.info(f"P14 schedule: {len(p14)} animals")
    return p14[available_cols]


def create_p56_schedule(assignments_df: pd.DataFrame) -> pd.DataFrame:
    if len(assignments_df) == 0:
        return pd.DataFrame()

    p56 = assignments_df[assignments_df['Assigned_Timepoint'] == 'P56'].copy()
    if len(p56) == 0:
        return pd.DataFrame()

    kept_animals = []
    for (behavior_date, strain, genotype, sex), group in p56.groupby(
        ['P56_Behavior_Date', 'Strain', 'Genotype', 'Sex']
    ):
        animals = group.to_dict('records')
        for i in range(0, len(animals), CONFIG['CAGE_SIZE']):
            cage_group = animals[i:i + CONFIG['CAGE_SIZE']]
            if len(cage_group) == CONFIG['CAGE_SIZE']:
                all_filled = all(
                    a.get('Harvest_Type') == 'COMPLETE (Quota Filled)'
                    for a in cage_group
                )
                if not all_filled:
                    kept_animals.extend(cage_group)
            else:
                # Partial cage: keep non-COMPLETE; always keep Extra (user-assigned)
                kept_animals.extend([
                    a for a in cage_group
                    if a.get('Harvest_Type') not in ('COMPLETE (Quota Filled)',)
                ])

    if len(kept_animals) == 0:
        return pd.DataFrame()

    p56_filtered = pd.DataFrame(kept_animals)
    p56_filtered['Day_of_Week'] = 'Wednesday'

    p56_filtered = p56_filtered.sort_values(
        ['P56_Harvest_Date', 'Genotype_Priority', 'Sex', 'Genotype', 'Animal_Name'],
        ascending=[True, False, True, True, True]
    )

    desired_cols = [
        'P56_Behavior_Date', 'P56_Harvest_Date', 'Day_of_Week',
        'P56_Age_At_Behavior_Days', 'P56_Age_At_Behavior_Months',
        'P56_Age_At_Harvest_Days', 'P56_Age_At_Harvest_Months',
        'Animal_Name', 'Strain', 'Strain_Priority',
        'Genotype', 'Genotype_Priority',
        'Sex', 'Marker_Type', 'Harvest_Type', 'Priority',
        'Birth_Date', 'Birth_ID', 'Assignment_Reason'
    ]
    available_cols = [c for c in desired_cols if c in p56_filtered.columns]

    logger.info(f"P56 schedule: {len(p56_filtered)} animals")
    return p56_filtered[available_cols]


def create_capacity_summary(p56_schedule_df: pd.DataFrame) -> pd.DataFrame:
    if len(p56_schedule_df) == 0:
        return pd.DataFrame()
    if CONFIG['CAGE_SIZE'] <= 0:
        return pd.DataFrame()

    capacity = p56_schedule_df.groupby('P56_Behavior_Date').size().reset_index()
    capacity.columns = ['Behavior_Start', 'Animals_Scheduled']
    capacity['Capacity'] = CONFIG['WEDNESDAY_CAPACITY']
    capacity['Available_Slots'] = CONFIG['WEDNESDAY_CAPACITY'] - capacity['Animals_Scheduled']
    capacity['Cages_Scheduled'] = capacity['Animals_Scheduled'] / CONFIG['CAGE_SIZE']
    capacity['Status'] = capacity['Available_Slots'].apply(
        lambda x: '✓ OK' if x >= 0 else '✗ OVER CAPACITY'
    )
    return capacity.sort_values('Behavior_Start')


def create_strain_summary(assignments_df: pd.DataFrame) -> pd.DataFrame:
    if len(assignments_df) == 0:
        return pd.DataFrame()

    summary = assignments_df.groupby(
        ['Strain', 'Genotype', 'Sex', 'Assigned_Timepoint']
    ).size().reset_index()
    summary.columns = ['Strain', 'Genotype', 'Sex', 'Timepoint', 'Count']

    pivot = summary.pivot_table(
        index=['Strain', 'Genotype', 'Sex'],
        columns='Timepoint',
        values='Count',
        fill_value=0
    ).reset_index()

    if 'P56' in pivot.columns and CONFIG['CAGE_SIZE'] > 0:
        pivot['P56_Complete_Cages'] = pivot['P56'] // CONFIG['CAGE_SIZE']
        pivot['P56_Animals_in_Cages'] = pivot['P56_Complete_Cages'] * CONFIG['CAGE_SIZE']
        pivot['P56_Leftover'] = pivot['P56'] % CONFIG['CAGE_SIZE']

    return pivot


def create_b6_monthly_summary(assignments_df: pd.DataFrame) -> pd.DataFrame:
    if len(assignments_df) == 0:
        return pd.DataFrame()

    min_per_month = CONFIG.get('B6_MIN_PER_MONTH', 3)

    scheduled = assignments_df[
        assignments_df['Assigned_Timepoint'].isin(['P14', 'P56'])
    ].copy()

    def harvest_month_label(row):
        tp = row.get('Assigned_Timepoint', '')
        if tp == 'P14':
            d = to_date(row.get('P14_Date'))
        elif tp == 'P56':
            d = to_date(row.get('P56_Harvest_Date'))
        else:
            d = None
        return d.strftime('%Y-%m') if d else None

    scheduled['Harvest_Month'] = scheduled.apply(harvest_month_label, axis=1)
    scheduled = scheduled[scheduled['Harvest_Month'].notna()]

    if len(scheduled) == 0:
        return pd.DataFrame()

    all_months = sorted(scheduled['Harvest_Month'].unique())

    rows = []
    for month in all_months:
        month_data = scheduled[scheduled['Harvest_Month'] == month]
        b6_data = month_data[month_data['Strain'].apply(is_b6_strain)]

        total_in_month = len(month_data)
        b6_count = len(b6_data)
        b6_topup = (
            len(b6_data[b6_data['Priority'] == 'B6_MIN'])
            if 'Priority' in b6_data.columns else 0
        )

        meets_min = b6_count >= min_per_month
        status = (
            '✅ Meets Minimum' if meets_min
            else f'⚠️ Below Minimum (need {min_per_month - b6_count} more)'
        )

        rows.append({
            'Harvest_Month':           month,
            'Total_Animals_Scheduled': total_in_month,
            'B6_B6N_Count':            b6_count,
            'B6_B6N_TopUp_Count':      b6_topup,
            'Minimum_Required':        min_per_month,
            'Shortfall':               max(0, min_per_month - b6_count),
            'Status':                  status,
        })

    return pd.DataFrame(rows)


def save_backup_csvs(output_dir: str, timestamp: str, **dataframes) -> str:
    backup_dir = os.path.join(output_dir, f'backup_{timestamp}')
    os.makedirs(backup_dir, exist_ok=True)
    saved = []
    for name, df in dataframes.items():
        if df is not None and isinstance(df, pd.DataFrame) and len(df) > 0:
            filepath = os.path.join(backup_dir, f'{name}.csv')
            try:
                df.to_csv(filepath, index=False)
                saved.append(filepath)
            except Exception as e:
                logger.warning(f"Could not save backup CSV {name}: {e}")
    if saved:
        print(f"\n✓ Backup CSVs saved to: {backup_dir} ({len(saved)} files)")
    return backup_dir


# ============================================================================
# MAIN SCHEDULING FUNCTION
# ============================================================================

def create_complete_schedule(animal_file: str, tracking_file: str, births_file: str,
                             output_dir: Optional[str] = None,
                             birth_date_start: Optional[date_type] = None,
                             birth_date_end: Optional[date_type] = None,
                             behavior_date_start: Optional[date_type] = None,
                             behavior_date_end: Optional[date_type] = None,
                             full_behavior_dates: Optional[List[date_type]] = None) -> str:
    logger.info("=" * 70)
    logger.info("COMPREHENSIVE ANIMAL SCHEDULER")
    logger.info("=" * 70)
    print("=" * 70)
    print("COMPREHENSIVE ANIMAL SCHEDULER")
    print("=" * 70)

    if output_dir is None:
        output_dir = os.path.dirname(os.path.abspath(animal_file))

    if full_behavior_dates:
        full_behavior_dates = [to_date(d) for d in full_behavior_dates if d is not None]
        full_behavior_dates = [d for d in full_behavior_dates if d is not None]

    # Read data
    animals_df = read_animal_data(animal_file)
    total_alive_count = len(animals_df)
    animals_df_raw = animals_df.copy()

    tracking_df = read_tracking_data(tracking_file) if tracking_file else None
    births_df = read_births_data(births_file) if births_file else None

    print(f"\nTotal alive animals loaded: {total_alive_count:,}")

    diagnose_animal_file(animals_df)

    # Parse requirements
    requirements = parse_requirements(tracking_df)
    remaining_needs = calculate_remaining_needs(requirements)
    extra_perf_status = check_extra_perfusion_status(requirements) if requirements else {}

    # Births analysis
    print("\n" + "=" * 70)
    print("BIRTHS ANALYSIS")
    print("=" * 70)

    sexing_schedule_df = pd.DataFrame()
    if births_df is not None:
        sexing_schedule_df = build_births_sexing_schedule(births_df, animals_df_raw)
        upcoming = (
            sexing_schedule_df[
                sexing_schedule_df['Days_Until_Sexing'].apply(
                    lambda x: isinstance(x, int) and 0 <= x <= 7
                )
            ] if len(sexing_schedule_df) > 0 else pd.DataFrame()
        )
        print(f"  Births needing sexing (not yet entered): {len(sexing_schedule_df)}")
        if len(upcoming) > 0:
            print(f"  ⚠️  {len(upcoming)} litter(s) need sexing within the next 7 days!")

    unmatched_births_df = find_unmatched_births_enhanced(
        births_df, animals_df, requirements, remaining_needs
    )
    unmatched_births_summary = create_unmatched_births_summary(unmatched_births_df)

    # Animal filtering
    print("\n" + "=" * 70)
    print("ANIMAL FILTERING")
    print("=" * 70)

    animals_df, use_excluded = filter_animals_by_use(animals_df)
    print(f"After 'Sing Inventory' filter: {len(animals_df):,} animals remain")
    if len(animals_df) == 0:
        print("  ⚠️  ALL animals were excluded by the Use filter.")
        print("  Check the 'Use' column values in your CSV.")

    animals_df, genotype_excluded_pass1, blank_genotypes = filter_animals_by_genotype_first_pass(animals_df)
    print(f"After genotype first pass:     {len(animals_df):,} animals remain")
    print(f"  Excluded (Wild, Cre-only Wild, Inconclusive): {len(genotype_excluded_pass1)}")
    print(f"  Blank genotypes (pending 2nd pass):           {len(blank_genotypes)}")
    if len(animals_df) == 0:
        print("  ⚠️  ALL animals were excluded by genotype filtering.")

    animals_df, date_excluded = filter_animals_by_dates(
        animals_df, birth_date_start, birth_date_end,
        behavior_date_start, behavior_date_end
    )
    if len(date_excluded) > 0:
        print(f"After date filtering:          {len(animals_df):,} animals remain "
              f"(excluded {len(date_excluded)})")

    print(f"\nAnimals entering eligibility check: {len(animals_df):,}")

    # Eligibility
    print("\nChecking eligibility...")
    eligibility = check_eligibility(animals_df, full_behavior_dates)

    print(f"Eligibility results: {len(eligibility):,} animals processed")
    if len(eligibility) > 0:
        p14_elig_count = eligibility['P14_Eligible'].sum() if 'P14_Eligible' in eligibility.columns else 0
        p56_elig_count = eligibility['P56_Eligible'].sum() if 'P56_Eligible' in eligibility.columns else 0
        print(f"  P14 eligible: {p14_elig_count}")
        print(f"  P56 eligible: {p56_elig_count}")

    het_count_df = (
        eligibility[eligibility['Genotype'].apply(is_heterozygous)]
        if len(eligibility) > 0 and 'Genotype' in eligibility.columns
        else pd.DataFrame()
    )

    # Assignment
    print("\nAssigning animals to timepoints...")
    assignments = assign_animals_smart(eligibility, remaining_needs, extra_perf_status)

    if len(assignments) > 0:
        assignments = check_capacity_and_reassign(assignments, remaining_needs)
    else:
        print("  ⚠️  No assignments to process.")

    # Determine actually-full dates
    if len(assignments) > 0:
        p56_assigned = assignments[assignments['Assigned_Timepoint'] == 'P56'].copy()
        if len(p56_assigned) > 0:
            p56_assigned = p56_assigned[p56_assigned['P56_Behavior_Date'].notna()].copy()
            p56_assigned['P56_Behavior_Date'] = p56_assigned['P56_Behavior_Date'].apply(to_date)
            p56_assigned = p56_assigned[p56_assigned['P56_Behavior_Date'].notna()]
            if len(p56_assigned) > 0:
                wed_counts = p56_assigned.groupby('P56_Behavior_Date').size()
                actual_full = wed_counts[wed_counts >= CONFIG['WEDNESDAY_CAPACITY']].index.tolist()
                all_full_dates = list(set((full_behavior_dates or []) + actual_full))
            else:
                all_full_dates = full_behavior_dates or []
        else:
            all_full_dates = full_behavior_dates or []
    else:
        all_full_dates = full_behavior_dates or []

    # Second pass: blank genotypes
    print("Second pass: Analyzing blank genotypes...")
    genotype_excluded_pass2 = analyze_blank_genotypes_second_pass(
        blank_genotypes, all_full_dates, remaining_needs
    )
    genotype_excluded = pd.concat(
        [genotype_excluded_pass1, genotype_excluded_pass2], ignore_index=True
    )

    # Harvest type assignment
    print("\n" + "=" * 70)
    print("HARVEST ASSIGNMENT REVIEW")
    print("=" * 70)
    overrides_file = os.path.join(output_dir, CONFIG.get('INPUT_OVERRIDES_FILE', 'harvest_overrides.csv'))

    if len(assignments) > 0:
        # Show the GUI — user reviews and confirms (or skips for auto)
        gui_selections = prompt_harvest_assignments_gui(assignments, remaining_needs)

        # Separate out any "Do Not Schedule" animals
        do_not_schedule = {
            name for name, htype in gui_selections.items()
            if htype == 'DO_NOT_SCHEDULE'
        }
        if do_not_schedule:
            print(f"  ⚠ {len(do_not_schedule)} animal(s) marked 'Do Not Schedule' — removed from assignments.")
            logger.info(f"Do Not Schedule: {sorted(do_not_schedule)}")
            assignments = assignments[
                ~assignments['Animal_Name'].isin(do_not_schedule)
            ].copy()

        # Build final override dict (exclude DO_NOT_SCHEDULE sentinels)
        harvest_overrides = {
            name: htype
            for name, htype in gui_selections.items()
            if htype != 'DO_NOT_SCHEDULE'
        }

        assignments = assign_harvest_types(
            assignments, remaining_needs, requirements, harvest_overrides
        )
    else:
        assignments = pd.DataFrame()
        harvest_overrides = {}

    # Persist confirmed assignments as the override file for reference / next run
    write_harvest_overrides_template(assignments, overrides_file)

    # B6/B6N monthly minimum
    print(f"Enforcing B6/B6N minimum ({CONFIG['B6_MIN_PER_MONTH']}/month)...")
    if len(assignments) > 0:
        assignments = enforce_b6_monthly_minimum(assignments, eligibility, remaining_needs)

    # Build output sheets
    print("Creating schedule sheets...")
    # Debug: log Harvest_Type for each animal going into P56 schedule
    p56_debug = assignments[assignments['Assigned_Timepoint'] == 'P56'] if len(assignments) > 0 else pd.DataFrame()
    if len(p56_debug) > 0:
        logger.info("P56 animals entering create_p56_schedule:")
        for _, _row in p56_debug.iterrows():
            logger.info(f"  {_row.get('Animal_Name')} | {_row.get('Strain')} | {_row.get('Sex')} | {_row.get('Harvest_Type')} | Priority={_row.get('Priority')}")
    p14_schedule = create_p14_schedule(assignments)
    p56_schedule = create_p56_schedule(assignments)
    unschedulable = create_unschedulable_report(
        assignments if len(assignments) > 0 else pd.DataFrame(),
        use_excluded,
        genotype_excluded,
        date_excluded
    )
    capacity = create_capacity_summary(p56_schedule)
    strain_summary = create_strain_summary(assignments)
    requirements_status = create_requirements_status(remaining_needs, requirements, extra_perf_status)
    genotype_summary = summarize_genotype_exclusions(genotype_excluded)
    b6_monthly_summary = create_b6_monthly_summary(assignments)

    # Counts for summary
    p14_count = len(p14_schedule) if len(p14_schedule) > 0 else 0
    p56_count = len(p56_schedule) if len(p56_schedule) > 0 else 0
    p56_cages = p56_count // CONFIG['CAGE_SIZE'] if CONFIG['CAGE_SIZE'] > 0 else 0

    genotype_critical_count = (
        len(genotype_excluded[genotype_excluded['Reason'].str.contains('⚠️', na=False)])
        if len(genotype_excluded) > 0 and 'Reason' in genotype_excluded.columns
        else 0
    )
    high_priority_count = (
        len(assignments[assignments['Priority'] == 'HIGH'])
        if len(assignments) > 0 and 'Priority' in assignments.columns else 0
    )
    b6_topup_count = (
        len(assignments[assignments['Priority'] == 'B6_MIN'])
        if len(assignments) > 0 and 'Priority' in assignments.columns else 0
    )

    toe_clip_excluded = (
        eligibility[eligibility['P56_Reason'].str.contains('Toe Clip', na=False)]
        if len(eligibility) > 0 and 'P56_Reason' in eligibility.columns
        else pd.DataFrame()
    )
    full_date_excluded = (
        eligibility[eligibility['P56_Reason'].str.contains('capacity', na=False, case=False)]
        if len(eligibility) > 0 and 'P56_Reason' in eligibility.columns
        else pd.DataFrame()
    )
    unusable_both = (
        len(unschedulable[unschedulable['Unusable_Both_Timepoints'] == 'YES'])
        if len(unschedulable) > 0 and 'Unusable_Both_Timepoints' in unschedulable.columns
        else 0
    )

    unmatched_p14_count = (
        len(unmatched_births_df[unmatched_births_df['P14_Potential'] == 'Yes'])
        if len(unmatched_births_df) > 0 and 'P14_Potential' in unmatched_births_df.columns else 0
    )
    unmatched_p56_count = (
        len(unmatched_births_df[unmatched_births_df['P56_Potential'] == 'Yes'])
        if len(unmatched_births_df) > 0 and 'P56_Potential' in unmatched_births_df.columns else 0
    )
    unmatched_priority_count = (
        len(unmatched_births_df[unmatched_births_df['Priority_Strain'] == 'YES'])
        if len(unmatched_births_df) > 0 and 'Priority_Strain' in unmatched_births_df.columns else 0
    )
    unmatched_quota_count = (
        len(unmatched_births_df[unmatched_births_df['Quota_Status'].str.contains('NEEDED', na=False)])
        if len(unmatched_births_df) > 0 and 'Quota_Status' in unmatched_births_df.columns else 0
    )

    upcoming_sexing_count = 0
    if len(sexing_schedule_df) > 0 and 'Days_Until_Sexing' in sexing_schedule_df.columns:
        upcoming_sexing_count = len(sexing_schedule_df[
            sexing_schedule_df['Days_Until_Sexing'].apply(
                lambda x: isinstance(x, int) and 0 <= x <= 7
            )
        ])

    summary_data = {
        'Metric': [
            '── ANIMAL COUNTS ──',
            'Total Alive Animals',
            'Excluded (Not Sing Inventory)',
            'Excluded (Genotype)',
            'Excluded (Genotype - CRITICAL)',
            'Excluded (Date Filters)',
            'Excluded (Toe Clip for P56)',
            'Excluded (Full P56 Dates)',
            'Unusable for BOTH Timepoints',
            'Heterozygous (Het) Animals',
            'Animals Processed',
            '── SCHEDULE ──',
            'P14 Assigned',
            'P56 Assigned',
            'P56 Complete Cages',
            'Unschedulable',
            'HIGH Priority Animals',
            '── EXTRA PERFUSIONS ──',
            'Extra Perfusion Strains Complete (Both Timepoints)',
            'Extra Perfusion Strains In Progress',
            '── B6/B6N ──',
            'B6/B6N Monthly Minimum Required',
            'B6/B6N Top-Up Animals Added',
            '── BIRTHS / SEXING ──',
            'Births Needing Sexing (not yet entered)',
            'Sexing Due Within 7 Days',
            'Unmatched Births (Sing Inventory)',
            'Unmatched - Can Schedule P14',
            'Unmatched - Can Schedule P56',
            'Unmatched - Priority Strains',
            'Unmatched - With Quota Needs',
            '── SETTINGS ──',
            'Wednesday Capacity',
            'Sexing Day Offset (days)',
            'Birth Date Filter Start',
            'Birth Date Filter End',
            'Behavior Date Filter Start',
            'Behavior Date Filter End',
            'Full P56 Behavior Dates',
            'Generated On',
            'Animal File',
            'Tracking File',
            'Births File',
        ],
        'Value': [
            '',
            total_alive_count,
            len(use_excluded),
            len(genotype_excluded),
            genotype_critical_count,
            len(date_excluded) if len(date_excluded) > 0 else 0,
            len(toe_clip_excluded),
            len(full_date_excluded),
            unusable_both,
            len(het_count_df),
            len(assignments) if len(assignments) > 0 else 0,
            '',
            p14_count,
            p56_count,
            p56_cages,
            len(unschedulable),
            high_priority_count,
            '',
            sum(1 for v in extra_perf_status.values() if v.get('strain_complete', False)),
            sum(1 for v in extra_perf_status.values() if not v.get('strain_complete', False) and requirements),
            '',
            CONFIG.get('B6_MIN_PER_MONTH', 3),
            b6_topup_count,
            '',
            len(sexing_schedule_df),
            upcoming_sexing_count,
            len(unmatched_births_df),
            unmatched_p14_count,
            unmatched_p56_count,
            unmatched_priority_count,
            unmatched_quota_count,
            '',
            CONFIG['WEDNESDAY_CAPACITY'],
            CONFIG.get('SEXING_OFFSET_DAYS', 9),
            str(birth_date_start) if birth_date_start else 'None',
            str(birth_date_end) if birth_date_end else 'None',
            str(behavior_date_start) if behavior_date_start else 'None',
            str(behavior_date_end) if behavior_date_end else 'None',
            ', '.join([str(d) for d in full_behavior_dates]) if full_behavior_dates else 'None',
            datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            os.path.basename(animal_file),
            os.path.basename(tracking_file) if tracking_file else 'N/A',
            os.path.basename(births_file) if births_file else 'N/A',
        ]
    }

    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    output_filename = f'Complete_Schedule_{timestamp}.xlsx'
    output_path = os.path.join(output_dir, output_filename)

    save_backup_csvs(
        output_dir, timestamp,
        p14_schedule=p14_schedule,
        p56_schedule=p56_schedule,
        unschedulable=unschedulable,
        capacity=capacity,
        strain_summary=strain_summary,
        requirements_status=requirements_status,
        unmatched_births=unmatched_births_df,
        genotype_excluded=genotype_excluded,
        sexing_schedule=sexing_schedule_df,
        b6_monthly_summary=b6_monthly_summary,
        all_animals=assignments,
    )

    print(f"\nWriting Excel: {output_filename}")

    try:
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:

            pd.DataFrame(summary_data).to_excel(writer, sheet_name='Summary', index=False)

            if len(requirements_status) > 0:
                requirements_status.to_excel(writer, sheet_name='Requirements Status', index=False)


            if len(p14_schedule) > 0:
                p14_schedule.to_excel(writer, sheet_name='P14 Schedule', index=False)

            if len(p56_schedule) > 0:
                p56_schedule.to_excel(writer, sheet_name='P56 Schedule', index=False)

            if len(capacity) > 0:
                capacity.to_excel(writer, sheet_name='Wednesday Capacity', index=False)

            if len(sexing_schedule_df) > 0:
                sexing_schedule_df.to_excel(writer, sheet_name='Sexing Schedule', index=False)

            if len(genotype_excluded) > 0:
                genotype_excluded.to_excel(
                    writer,
                    sheet_name=truncate_sheet_name('Genotype Excluded Details'),
                    index=False
                )

            if len(strain_summary) > 0:
                strain_summary.to_excel(writer, sheet_name='Strain Summary', index=False)

            # All Animals tab
            if len(assignments) > 0:
                _INTERNAL_COLUMNS = {
                    '_quota_limited_complete_group', '_incomplete_group',
                    '_full_date_complete', '_urgency_sort', '_birth_date_obj',
                    '_first_wed', 'is_het', 'breeding_type'
                }
                _EXCLUDED_ORIGINAL_AGE_COLUMNS = {'Age (days)', 'Age (weeks)', 'Age (months)'}
                _REDUNDANT_WITH_COMPUTED = {
                    'Name': 'Animal_Name', 'Birth ID': 'Birth_ID',
                    'Line (Short)': 'Strain', 'Birth Date': 'Birth_Date',
                    'Marker Type': 'Marker_Type',
                }

                computed_cols_front = [
                    'Animal_Name', 'Birth_ID', 'Strain', 'Strain_Priority',
                    'Genotype', 'Genotype_Priority', 'Sex', 'Marker_Type',
                    'Birth_Date', 'Age_Today_Days',
                    'Assigned_Timepoint', 'Harvest_Type', 'Priority',
                    'P14_Eligible', 'P14_Too_Old', 'P14_Date', 'P14_Reason',
                    'P14_Age_At_Harvest_Days', 'P14_Age_At_Harvest_Months',
                    'P56_Eligible', 'P56_Too_Old', 'P56_Behavior_Date',
                    'P56_Harvest_Date', 'P56_Reason',
                    'P56_Age_At_Behavior_Days', 'P56_Age_At_Behavior_Months',
                    'P56_Age_At_Harvest_Days', 'P56_Age_At_Harvest_Months',
                    'Unusable_Note', 'Assignment_Reason',
                ]

                all_available = assignments.columns.tolist()
                ordered_cols = []
                seen = set()

                for col in computed_cols_front:
                    if col in all_available and col not in seen:
                        ordered_cols.append(col)
                        seen.add(col)

                for col in all_available:
                    if col in seen:
                        continue
                    if col in _INTERNAL_COLUMNS:
                        continue
                    if col in _EXCLUDED_ORIGINAL_AGE_COLUMNS:
                        continue
                    if col in _REDUNDANT_WITH_COMPUTED:
                        if _REDUNDANT_WITH_COMPUTED[col] in seen:
                            continue
                    ordered_cols.append(col)
                    seen.add(col)

                assignments[ordered_cols].to_excel(writer, sheet_name='All Animals', index=False)

            # ── Formatting ────────────────────────────────────────────────────
            wb = writer.book

            if 'Sexing Schedule' in wb.sheetnames:
                ws = wb['Sexing Schedule']
                headers = [cell.value for cell in ws[1]]
                status_col = headers.index('Sexing_Status') + 1 if 'Sexing_Status' in headers else None

                for row_idx in range(2, ws.max_row + 1):
                    if status_col:
                        cell = ws.cell(row=row_idx, column=status_col)
                        val = str(cell.value) if cell.value else ''
                        color = None
                        if 'TODAY' in val:
                            color = 'FF0000'
                            cell.font = Font(bold=True, color='FFFFFF')
                        elif 'TOMORROW' in val:
                            color = 'FF8C00'
                            cell.font = Font(bold=True)
                        elif 'SOON' in val:
                            color = 'FFD700'
                        elif 'Upcoming' in val:
                            color = 'A8E6CF'
                        elif 'Done' in val:
                            color = 'D3D3D3'
                        if color:
                            cell.fill = PatternFill(
                                start_color=color, end_color=color, fill_type='solid'
                            )



            geno_sheet = truncate_sheet_name('Genotype Excluded Details')
            if geno_sheet in wb.sheetnames:
                ws = wb[geno_sheet]
                headers = [cell.value for cell in ws[1]]
                pred_col      = headers.index('Prediction')           + 1 if 'Prediction'           in headers else None
                days_col      = headers.index('Days_Until_Deadline')  + 1 if 'Days_Until_Deadline'  in headers else None
                p14_worth_col = headers.index('P14_Worth_Genotyping') + 1 if 'P14_Worth_Genotyping' in headers else None
                p56_worth_col = headers.index('P56_Worth_Genotyping') + 1 if 'P56_Worth_Genotyping' in headers else None

                for row_idx in range(2, ws.max_row + 1):
                    if pred_col:
                        cell = ws.cell(row=row_idx, column=pred_col)
                        val = str(cell.value) if cell.value else ''
                        if 'LIKELY USABLE' in val:
                            cell.fill = PatternFill(start_color='C6E0B4', end_color='C6E0B4', fill_type='solid')
                            cell.font = Font(bold=True, color='00AA00')
                        elif 'POSSIBLY USABLE' in val:
                            cell.fill = PatternFill(start_color='FFE699', end_color='FFE699', fill_type='solid')
                        elif 'NOT SCHEDULABLE' in val:
                            cell.fill = PatternFill(start_color='FFB6B6', end_color='FFB6B6', fill_type='solid')
                        elif 'DEADLINE PASSED' in val:
                            cell.fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')

                    if days_col:
                        cell = ws.cell(row=row_idx, column=days_col)
                        try:
                            val = cell.value
                            if val != 'N/A' and isinstance(val, (int, float)):
                                if val <= 7:
                                    cell.fill = PatternFill(start_color='FF6B6B', end_color='FF6B6B', fill_type='solid')
                                    cell.font = Font(bold=True, color='FFFFFF')
                                elif val <= 14:
                                    cell.fill = PatternFill(start_color='FFD93D', end_color='FFD93D', fill_type='solid')
                        except Exception:
                            pass

                    for worth_col in [p14_worth_col, p56_worth_col]:
                        if worth_col:
                            cell = ws.cell(row=row_idx, column=worth_col)
                            val = str(cell.value) if cell.value else ''
                            if '✅ YES' in val:
                                cell.fill = PatternFill(start_color='C6E0B4', end_color='C6E0B4', fill_type='solid')
                                cell.font = Font(bold=True, color='006400')
                            elif '🟡' in val:
                                cell.fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
                            elif '❌' in val:
                                cell.fill = PatternFill(start_color='FFB6B6', end_color='FFB6B6', fill_type='solid')
                            elif '⚠️ QUOTA' in val:
                                cell.fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')

            for sheet_name in wb.sheetnames:
                auto_size_columns(wb[sheet_name])

        logger.info("✓ Excel written successfully")
        print("✓ Excel written successfully")

    except Exception as e:
        logger.error(f"Excel write failed: {e}", exc_info=True)
        print(f"\n⚠️ Excel write failed: {e}")
        print("✓ Backup CSVs were saved — you can open those directly.")
        raise

    # Console summary
    print("\n" + "=" * 70)
    print("ASSIGNMENT SUMMARY")
    print("=" * 70)
    print(f"  P14 assigned:          {p14_count:>6}")
    print(f"  P56 assigned:          {p56_count:>6}  ({p56_cages} complete cages)")
    print(f"  Unschedulable:         {len(unschedulable):>6}")
    if unusable_both > 0:
        print(f"  ⛔ Unusable for both:  {unusable_both:>6}  (too old for P14 AND P56)")
    if b6_topup_count > 0:
        print(f"  B6/B6N top-up added:  {b6_topup_count:>6}  (to meet {CONFIG['B6_MIN_PER_MONTH']}/month minimum)")
    if len(sexing_schedule_df) > 0:
        print(f"\n  Births needing sexing:    {len(sexing_schedule_df)}")
        if upcoming_sexing_count > 0:
            print(f"  ⚠️  Sexing due ≤7 days:    {upcoming_sexing_count}")
    if len(unmatched_births_df) > 0:
        print(f"\n  ⚠️  Unmatched births:       {len(unmatched_births_df)}")
    if genotype_critical_count > 0:
        print(f"\n  ⚠️  CRITICAL genotype issues: {genotype_critical_count}")

    print("\n" + "=" * 70)
    print(f"✓ Schedule saved to:\n  {output_path}")
    print("=" * 70)

    return output_path, assignments


# ============================================================================
# UNIT TESTS
# ============================================================================

class TestSchedulerFunctions(unittest.TestCase):

    def setUp(self):
        self.test_date = date(2025, 11, 15)

    def test_to_date_conversion(self):
        self.assertEqual(to_date(self.test_date), self.test_date)
        self.assertEqual(to_date(datetime(2025, 11, 15, 10, 30)), self.test_date)
        self.assertEqual(to_date(pd.Timestamp('2025-11-15')), self.test_date)
        self.assertIsNone(to_date(None))
        self.assertIsNone(to_date(pd.NaT))

    def test_is_heterozygous(self):
        self.assertTrue(is_heterozygous('-/+'))
        self.assertTrue(is_heterozygous('+/-'))
        self.assertTrue(is_heterozygous('HET'))
        self.assertTrue(is_heterozygous('Het1'))
        self.assertFalse(is_heterozygous('+/+'))
        self.assertFalse(is_heterozygous(''))
        self.assertFalse(is_heterozygous(None))
        self.assertTrue(is_heterozygous('Het'))
        self.assertFalse(is_heterozygous('Wild'))
        self.assertFalse(is_heterozygous('Hom'))
        self.assertFalse(is_heterozygous('Blank'))
        self.assertFalse(is_heterozygous('Inbred'))
        self.assertFalse(is_heterozygous('Hemi'))

    def test_canonicalize_genotype_het(self):
        for raw in ['-/+', '+/-', 'HET', 'het', 'Heterozygous', 'carrier']:
            self.assertEqual(
                canonicalize_genotype(raw), GENOTYPE_HET,
                f"Expected Het for '{raw}'"
            )

    def test_canonicalize_genotype_wild(self):
        for raw in ['+/+', '+/Y', 'WT', 'wildtype', 'wild-type',
                    'Cre ncar', 'Generic Cre', 'cre +/+']:
            self.assertEqual(
                canonicalize_genotype(raw), GENOTYPE_WILD,
                f"Expected Wild for '{raw}'"
            )

    def test_canonicalize_genotype_hom(self):
        for raw in ['-/-', 'HOM', 'homozygous', 'mut/mut', 'KO/KO']:
            self.assertEqual(
                canonicalize_genotype(raw), GENOTYPE_HOM,
                f"Expected Hom for '{raw}'"
            )

    def test_canonicalize_genotype_hemi(self):
        for raw in ['hemi', 'hemizygous', 'tg/+', '+/tg', '-/Y']:
            self.assertEqual(
                canonicalize_genotype(raw), GENOTYPE_HEMI,
                f"Expected Hemi for '{raw}'"
            )

    def test_canonicalize_genotype_inbred(self):
        self.assertEqual(
            canonicalize_genotype('+/+', strain='B6J'), GENOTYPE_INBRED
        )
        self.assertEqual(
            canonicalize_genotype('', strain='B6NJ'), GENOTYPE_INBRED
        )

    def test_canonicalize_genotype_blank(self):
        for raw in [None, '', 'nan', 'N/A', 'Inconclusive', 'Pending']:
            self.assertEqual(
                canonicalize_genotype(raw), GENOTYPE_BLANK,
                f"Expected Blank for '{raw}'"
            )

    def test_canonicalize_already_canonical(self):
        for label in ['Wild', 'Het', 'Hom', 'Hemi', 'Inbred', 'Blank']:
            self.assertEqual(canonicalize_genotype(label), label)

    def test_is_wildtype_cre_only(self):
        self.assertTrue(is_wildtype_cre_only('Cre ncar'))
        self.assertTrue(is_wildtype_cre_only('Generic Cre'))
        self.assertTrue(is_wildtype_cre_only('Cre-ncar'))
        self.assertFalse(is_wildtype_cre_only('Cre ncar; -/+'))
        self.assertFalse(is_wildtype_cre_only('Cre ncar HET'))
        self.assertFalse(is_wildtype_cre_only('Cre ncar; -/-'))
        self.assertFalse(is_wildtype_cre_only('+/+'))
        self.assertFalse(is_wildtype_cre_only('-/+'))
        self.assertFalse(is_wildtype_cre_only(''))
        self.assertFalse(is_wildtype_cre_only(None))
        self.assertTrue(is_wildtype_cre_only(GENOTYPE_WILD))
        self.assertFalse(is_wildtype_cre_only(GENOTYPE_HET))

    def test_is_b6_strain(self):
        self.assertTrue(is_b6_strain('B6J'))
        self.assertTrue(is_b6_strain('b6j'))
        self.assertTrue(is_b6_strain('B6NJ'))
        self.assertTrue(is_b6_strain('b6nj'))
        self.assertFalse(is_b6_strain('SHANK3'))
        self.assertFalse(is_b6_strain('CHD8'))
        self.assertFalse(is_b6_strain(None))
        self.assertFalse(is_b6_strain(''))

    def test_calculate_sexing_date(self):
        bd = date(2025, 11, 1)
        expected = date(2025, 11, 10)
        self.assertEqual(calculate_sexing_date(bd), expected)
        self.assertIsNone(calculate_sexing_date(None))
        self.assertIsNone(calculate_sexing_date(pd.NaT))

    def test_calculate_sexing_date_pd_timestamp(self):
        ts = pd.Timestamp('2025-11-01')
        expected = date(2025, 11, 10)
        self.assertEqual(calculate_sexing_date(ts), expected)

    def test_sexing_date_in_schedule_dates(self):
        bd = date(2025, 11, 1)
        dates = calculate_schedule_dates(bd)
        self.assertIsNotNone(dates)
        self.assertIn('sexing_date', dates)
        self.assertEqual(dates['sexing_date'], date(2025, 11, 10))

    def test_sexing_date_offset_configurable(self):
        original = CONFIG['SEXING_OFFSET_DAYS']
        try:
            CONFIG['SEXING_OFFSET_DAYS'] = 7
            bd = date(2025, 11, 1)
            result = calculate_sexing_date(bd)
            self.assertEqual(result, date(2025, 11, 8))
        finally:
            CONFIG['SEXING_OFFSET_DAYS'] = original

    def test_next_wednesday(self):
        self.assertEqual(next_wednesday(date(2025, 11, 14)), date(2025, 11, 19))
        self.assertEqual(next_wednesday(date(2025, 11, 19)), date(2025, 11, 19))
        self.assertEqual(next_wednesday(date(2025, 11, 20)), date(2025, 11, 26))
        self.assertEqual(next_wednesday(date(2025, 11, 17)), date(2025, 11, 19))

    def test_calculate_schedule_dates(self):
        bd = date(2025, 11, 1)
        dates = calculate_schedule_dates(bd)
        self.assertIsNotNone(dates)
        self.assertEqual(dates['birth_date'], bd)
        self.assertEqual(dates['p14_harvest'], date(2025, 11, 15))
        self.assertEqual(dates['p56_behavior_window_start'], date(2025, 12, 13))
        self.assertEqual(dates['p56_behavior_window_end'], date(2025, 12, 20))
        self.assertEqual(dates['sexing_date'], date(2025, 11, 10))

    def test_is_valid_p14_day(self):
        self.assertTrue(is_valid_p14_day(date(2025, 11, 17)))
        self.assertTrue(is_valid_p14_day(date(2025, 11, 18)))
        self.assertTrue(is_valid_p14_day(date(2025, 11, 19)))
        self.assertTrue(is_valid_p14_day(date(2025, 11, 20)))
        self.assertTrue(is_valid_p14_day(date(2025, 11, 21)))
        self.assertFalse(is_valid_p14_day(date(2025, 11, 22)))
        self.assertFalse(is_valid_p14_day(date(2025, 11, 23)))
        self.assertFalse(is_valid_p14_day(None))

    def test_is_priority_strain(self):
        self.assertTrue(is_priority_strain('SHANK3'))
        self.assertTrue(is_priority_strain('shank3'))
        self.assertTrue(is_priority_strain('B6J'))
        self.assertTrue(is_priority_strain('CHD8'))
        self.assertFalse(is_priority_strain('NONEXISTENT_STRAIN'))
        self.assertFalse(is_priority_strain(None))
        self.assertFalse(is_priority_strain(''))

    def test_is_super_priority_strain(self):
        self.assertTrue(is_super_priority_strain('SHANK3'))
        self.assertTrue(is_super_priority_strain('shank3'))
        self.assertTrue(is_super_priority_strain('CHD8'))
        self.assertTrue(is_super_priority_strain('FMR1'))
        self.assertFalse(is_super_priority_strain('B6J'))
        self.assertFalse(is_super_priority_strain(None))
        self.assertFalse(is_super_priority_strain('NONEXISTENT'))

    def test_get_next_wednesdays_count(self):
        for n in [1, 3, 6, 10]:
            result = get_next_wednesdays(n, from_date=date(2025, 11, 17))
            self.assertEqual(len(result), n)

    def test_get_next_wednesdays_all_on_wednesday(self):
        wednesdays = get_next_wednesdays(6, from_date=date(2025, 11, 15))
        for wed in wednesdays:
            self.assertEqual(wed.weekday(), 2)

    def test_get_next_wednesdays_spacing(self):
        wednesdays = get_next_wednesdays(6, from_date=date(2025, 11, 15))
        for i in range(1, len(wednesdays)):
            delta = (wednesdays[i] - wednesdays[i - 1]).days
            self.assertEqual(delta, 7)

    def test_get_next_wednesdays_from_wednesday_includes_today(self):
        start = date(2025, 11, 19)
        wednesdays = get_next_wednesdays(6, from_date=start)
        self.assertEqual(wednesdays[0], start)

    def test_get_next_wednesdays_from_saturday(self):
        start = date(2025, 11, 15)
        wednesdays = get_next_wednesdays(6, from_date=start)
        self.assertEqual(wednesdays[0], date(2025, 11, 19))
        self.assertEqual(wednesdays[5], date(2025, 12, 24))

    # ── P14 today-is-too-late tests ───────────────────────────────────────────

    def test_p14_today_is_too_late(self):
        """P14 date == today should be ineligible."""
        today = datetime.now().date()
        birth = today - timedelta(days=14)
        test_data = pd.DataFrame({
            'Name': ['Animal1'],
            'Birth Date': [pd.Timestamp(birth)],
            'Sex': ['Male'],
            'Line (Short)': ['SHANK3'],
            'Genotype': ['Het'],
            'Use': ['Sing Inventory'],
            'Status': ['Alive'],
            'Birth ID': ['B001'],
            'Marker Type': ['Ear Punch'],
        })
        result = check_eligibility(test_data, full_behavior_dates=[])
        row = result.iloc[0]
        self.assertFalse(row['P14_Eligible'], "P14 date == today should NOT be eligible")
        self.assertTrue(row['P14_Too_Old'],    "P14 date == today should set Too_Old flag")
        self.assertIn('TOO LATE', row['P14_Reason'].upper())

    def test_p14_tomorrow_is_eligible(self):
        """P14 date == tomorrow should be eligible if a valid weekday."""
        today = datetime.now().date()
        tomorrow = today + timedelta(days=1)
        birth = tomorrow - timedelta(days=14)
        test_data = pd.DataFrame({
            'Name': ['Animal1'],
            'Birth Date': [pd.Timestamp(birth)],
            'Sex': ['Male'],
            'Line (Short)': ['SHANK3'],
            'Genotype': ['Het'],
            'Use': ['Sing Inventory'],
            'Status': ['Alive'],
            'Birth ID': ['B001'],
            'Marker Type': ['Ear Punch'],
        })
        result = check_eligibility(test_data, full_behavior_dates=[])
        row = result.iloc[0]
        if tomorrow.weekday() in CONFIG['P14_VALID_DAYS']:
            self.assertTrue(row['P14_Eligible'])
            self.assertFalse(row['P14_Too_Old'])
        else:
            self.assertFalse(row['P14_Eligible'])
            self.assertIn('valid harvest day', row['P14_Reason'].lower())

    def test_p14_yesterday_is_too_old(self):
        """P14 date == yesterday should be ineligible."""
        today = datetime.now().date()
        birth = today - timedelta(days=15)
        test_data = pd.DataFrame({
            'Name': ['Animal1'],
            'Birth Date': [pd.Timestamp(birth)],
            'Sex': ['Male'],
            'Line (Short)': ['SHANK3'],
            'Genotype': ['Het'],
            'Use': ['Sing Inventory'],
            'Status': ['Alive'],
            'Birth ID': ['B001'],
            'Marker Type': ['Ear Punch'],
        })
        result = check_eligibility(test_data, full_behavior_dates=[])
        row = result.iloc[0]
        self.assertFalse(row['P14_Eligible'])
        self.assertTrue(row['P14_Too_Old'])
        self.assertIn('TOO OLD', row['P14_Reason'].upper())

    # ── Age column tests ──────────────────────────────────────────────────────

    def test_eligibility_p56_age_values(self):
        test_data = pd.DataFrame({
            'Name': ['Animal1'],
            'Birth Date': [pd.Timestamp('2025-10-01')],
            'Sex': ['Male'],
            'Line (Short)': ['SHANK3'],
            'Genotype': ['Het'],
            'Use': ['Sing Inventory'],
            'Status': ['Alive'],
            'Birth ID': ['B001'],
            'Marker Type': ['Ear Punch'],
        })
        result = check_eligibility(test_data, full_behavior_dates=[])
        row = result.iloc[0]
        if row['P56_Age_At_Behavior_Days'] is not None:
            self.assertGreaterEqual(row['P56_Age_At_Behavior_Days'], 42)
            self.assertLessEqual(row['P56_Age_At_Behavior_Days'], 49)
            self.assertIsInstance(row['P56_Age_At_Behavior_Months'], int)
            self.assertEqual(
                row['P56_Age_At_Harvest_Days'],
                row['P56_Age_At_Behavior_Days'] + 14
            )
            self.assertIsInstance(row['P56_Age_At_Harvest_Months'], int)

    def test_eligibility_too_old_both_flags(self):
        old_birth = date(2020, 1, 1)
        test_data = pd.DataFrame({
            'Name': ['OldAnimal'],
            'Birth Date': [pd.Timestamp(old_birth)],
            'Sex': ['Male'],
            'Line (Short)': ['SHANK3'],
            'Genotype': ['Het'],
            'Use': ['Sing Inventory'],
            'Status': ['Alive'],
            'Birth ID': ['B001'],
            'Marker Type': ['Ear Punch'],
        })
        result = check_eligibility(test_data, full_behavior_dates=[])
        row = result.iloc[0]
        self.assertTrue(row['P14_Too_Old'], "Should be flagged too old for P14")
        self.assertTrue(row['P56_Too_Old'], "Should be flagged too old for P56")
        self.assertIn('TOO OLD', row['P14_Reason'].upper())
        self.assertIn('TOO OLD', row['P56_Reason'].upper())
        self.assertIn('Unusable_Note', result.columns)
        self.assertIn('UNUSABLE FOR BOTH', row['Unusable_Note'].upper())

    def test_eligibility_not_too_old_fresh_animal(self):
        recent_birth = datetime.now().date() - timedelta(days=3)
        test_data = pd.DataFrame({
            'Name': ['YoungAnimal'],
            'Birth Date': [pd.Timestamp(recent_birth)],
            'Sex': ['Female'],
            'Line (Short)': ['CHD8'],
            'Genotype': ['Het'],
            'Use': ['Sing Inventory'],
            'Status': ['Alive'],
            'Birth ID': ['B002'],
            'Marker Type': ['Ear Punch'],
        })
        result = check_eligibility(test_data, full_behavior_dates=[])
        row = result.iloc[0]
        self.assertFalse(row['P14_Too_Old'])
        self.assertFalse(row['P56_Too_Old'])

    def test_eligibility_age_today_days_present(self):
        today = datetime.now().date()
        birth = today - timedelta(days=20)
        test_data = pd.DataFrame({
            'Name': ['Animal1'],
            'Birth Date': [pd.Timestamp(birth)],
            'Sex': ['Male'],
            'Line (Short)': ['SHANK3'],
            'Genotype': ['Het'],
            'Use': ['Sing Inventory'],
            'Status': ['Alive'],
            'Birth ID': ['B003'],
            'Marker Type': ['Ear Punch'],
        })
        result = check_eligibility(test_data, full_behavior_dates=[])
        row = result.iloc[0]
        self.assertIn('Age_Today_Days', result.columns)
        self.assertEqual(row['Age_Today_Days'], 20)

    def test_eligibility_strain_column_present(self):
        test_data = pd.DataFrame({
            'Name': ['Animal1'],
            'Birth Date': [pd.Timestamp('2025-10-01')],
            'Sex': ['Male'],
            'Line (Short)': ['SHANK3'],
            'Genotype': ['Het'],
            'Use': ['Sing Inventory'],
            'Status': ['Alive'],
            'Birth ID': ['B001'],
            'Marker Type': ['Ear Punch'],
        })
        result = check_eligibility(test_data, full_behavior_dates=[])
        self.assertIn('Strain', result.columns)
        self.assertEqual(result.iloc[0]['Strain'], 'SHANK3')

    def test_eligibility_birth_date_column_present(self):
        test_data = pd.DataFrame({
            'Name': ['Animal1'],
            'Birth Date': [pd.Timestamp('2025-10-01')],
            'Sex': ['Male'],
            'Line (Short)': ['SHANK3'],
            'Genotype': ['Het'],
            'Use': ['Sing Inventory'],
            'Status': ['Alive'],
            'Birth ID': ['B001'],
            'Marker Type': ['Ear Punch'],
        })
        result = check_eligibility(test_data, full_behavior_dates=[])
        self.assertIn('Birth_Date', result.columns)

    def test_eligibility_animal_name_column_present(self):
        test_data = pd.DataFrame({
            'Name': ['MyAnimal'],
            'Birth Date': [pd.Timestamp('2025-10-01')],
            'Sex': ['Male'],
            'Line (Short)': ['SHANK3'],
            'Genotype': ['Het'],
            'Use': ['Sing Inventory'],
            'Status': ['Alive'],
            'Birth ID': ['B001'],
            'Marker Type': ['Ear Punch'],
        })
        result = check_eligibility(test_data, full_behavior_dates=[])
        self.assertIn('Animal_Name', result.columns)
        self.assertEqual(result.iloc[0]['Animal_Name'], 'MyAnimal')

    def test_eligibility_genotype_is_canonical(self):
        test_data = pd.DataFrame({
            'Name': ['Animal1'],
            'Birth Date': [pd.Timestamp('2025-10-01')],
            'Sex': ['Male'],
            'Line (Short)': ['SHANK3'],
            'Genotype': ['Het'],
            'Use': ['Sing Inventory'],
            'Status': ['Alive'],
            'Birth ID': ['B001'],
            'Marker Type': ['Ear Punch'],
        })
        result = check_eligibility(test_data, full_behavior_dates=[])
        geno = result.iloc[0]['Genotype']
        self.assertIn(geno, _CANONICAL_GENOTYPES,
                      f"'{geno}' is not a canonical genotype label")

    # ── Assignment tests ──────────────────────────────────────────────────────

    def test_assign_animals_smart_empty_input(self):
        result = assign_animals_smart(pd.DataFrame(), {})
        self.assertIsInstance(result, pd.DataFrame)
        self.assertEqual(len(result), 0)

    def test_assign_animals_smart_missing_strain_column(self):
        bad_df = pd.DataFrame({
            'Animal_Name': ['A1'],
            'Sex': ['Male'],
            'Genotype': ['Het'],
            'Birth_Date': ['2025-10-01'],
        })
        with self.assertRaises(KeyError):
            assign_animals_smart(bad_df, {})

    def test_assign_animals_smart_required_columns_present(self):
        today = datetime.now().date()
        birth = today - timedelta(days=20)
        test_data = pd.DataFrame({
            'Name': ['Animal1'],
            'Birth Date': [pd.Timestamp(birth)],
            'Sex': ['Male'],
            'Line (Short)': ['SHANK3'],
            'Genotype': ['Het'],
            'Use': ['Sing Inventory'],
            'Status': ['Alive'],
            'Birth ID': ['B001'],
            'Marker Type': ['Ear Punch'],
        })
        eligibility = check_eligibility(test_data, full_behavior_dates=[])
        result = assign_animals_smart(eligibility, {})
        self.assertIsInstance(result, pd.DataFrame)

    # ── parse_unschedulable_reason tests ─────────────────────────────────────

    def test_parse_unschedulable_genotype_deadline_passed(self):
        reason = (
            "[GENOTYPE DEADLINE PASSED] "
            "'Half' STRAIN — 2 blank genotype(s) from birth 2026-02-09. "
            "May have had ~1 Het but genotyping deadline passed (2026-02-22). "
            "P14 on 2026-02-23 or P56 on 2026-03-25"
        )
        parsed = parse_unschedulable_reason(reason)
        self.assertIn('Genotype Deadline', parsed['Primary_Reason'])
        self.assertEqual(parsed['Too_Old_For_P14'], 'NO')
        self.assertEqual(parsed['Too_Old_For_P56'], 'NO')
        self.assertEqual(parsed['Unusable_Both'],   'NO')
        self.assertNotIn('Window passed', parsed['P14_Status'])
        self.assertNotIn('Window passed', parsed['P56_Status'])
        self.assertIn('🧬', parsed['Primary_Reason'])

    def test_parse_unschedulable_windows_actually_passed(self):
        reason = "P14: TOO OLD FOR P14 — P14 date was 2025-01-01; P56: TOO OLD FOR P56"
        parsed = parse_unschedulable_reason(reason)
        self.assertNotIn('Genotype Deadline', parsed['Primary_Reason'])

    def test_parse_unschedulable_unusable_both(self):
        reason = '⛔ UNUSABLE FOR BOTH TIMEPOINTS — Too old for P14 AND P56. Animal is 1200d old.'
        parsed = parse_unschedulable_reason(reason)
        self.assertEqual(parsed['Too_Old_For_P14'], 'YES')
        self.assertEqual(parsed['Too_Old_For_P56'], 'YES')
        self.assertEqual(parsed['Unusable_Both'], 'YES')
        self.assertIn('⛔', parsed['Primary_Reason'])
        self.assertEqual(parsed['P14_Status'], '❌ Too old')
        self.assertEqual(parsed['P56_Status'], '❌ Too old')

    def test_parse_unschedulable_too_late_p14(self):
        reason = (
            '❌ TOO LATE FOR P14 — P14 date is today (2026-02-09) — '
            'harvest must be scheduled in advance'
        )
        parsed = parse_unschedulable_reason(reason)
        self.assertEqual(parsed['Too_Old_For_P14'], 'YES')
        self.assertEqual(parsed['Too_Old_For_P56'], 'NO')
        self.assertIn('Too Late', parsed['Primary_Reason'])

    def test_parse_unschedulable_too_old_p14_only(self):
        reason = '❌ TOO OLD FOR P14 — P14 date was 2025-01-01 (300 days ago)'
        parsed = parse_unschedulable_reason(reason)
        self.assertEqual(parsed['Too_Old_For_P14'], 'YES')
        self.assertEqual(parsed['Too_Old_For_P56'], 'NO')
        self.assertEqual(parsed['Unusable_Both'], 'NO')

    def test_parse_unschedulable_too_old_p56_only(self):
        reason = (
            '❌ TOO OLD FOR P56 — P56 behavior window ended 2025-01-20 '
            '(200 days ago). Unusable for P56.'
        )
        parsed = parse_unschedulable_reason(reason)
        self.assertEqual(parsed['Too_Old_For_P56'], 'YES')
        self.assertEqual(parsed['Too_Old_For_P14'], 'NO')
        self.assertEqual(parsed['Unusable_Both'], 'NO')
        self.assertIn('TOO OLD', parsed['Primary_Reason'].upper())

    def test_parse_unschedulable_quota_filled(self):
        reason = 'P56 quota filled for strain — reassigned to P14 (also filled)'
        parsed = parse_unschedulable_reason(reason)
        self.assertIn('Quota', parsed['Primary_Reason'])
        self.assertEqual(parsed['Too_Old_For_P14'], 'NO')
        self.assertEqual(parsed['Unusable_Both'], 'NO')

    def test_parse_unschedulable_capacity(self):
        reason = 'P56 date at capacity (2025-11-19); P14 unavailable: P14 date has passed'
        parsed = parse_unschedulable_reason(reason)
        self.assertIn('Capacity', parsed['Primary_Reason'])
        self.assertEqual(parsed['P56_Status'], '🔴 Date full')
        self.assertEqual(parsed['P14_Status'], '❌ Unavailable')

    def test_parse_unschedulable_incomplete_group(self):
        reason = 'Incomplete P56 group; P14 unavailable: P14 falls on Saturday'
        parsed = parse_unschedulable_reason(reason)
        self.assertIn('Incomplete', parsed['Primary_Reason'])
        self.assertIn('< 3', parsed['P56_Status'])

    def test_parse_unschedulable_toe_clip(self):
        reason = 'Has Toe Clip marker — not allowed for P56 behavior'
        parsed = parse_unschedulable_reason(reason)
        self.assertIn('Toe Clip', parsed['Primary_Reason'])
        self.assertIn('🚫', parsed['P56_Status'])

    def test_parse_unschedulable_invalid_day(self):
        reason = 'P14 falls on Saturday (2025-11-22) — not a valid harvest day'
        parsed = parse_unschedulable_reason(reason)
        self.assertIn('Invalid Day', parsed['Primary_Reason'])
        self.assertIn('⚠️', parsed['P14_Status'])

    def test_parse_unschedulable_no_birth_date(self):
        reason = 'No birth date'
        parsed = parse_unschedulable_reason(reason)
        self.assertIn('No Birth Date', parsed['Primary_Reason'])
        self.assertEqual(parsed['Too_Old_For_P14'], 'NO')
        self.assertEqual(parsed['Unusable_Both'], 'NO')

    def test_parse_unschedulable_empty(self):
        parsed = parse_unschedulable_reason('')
        self.assertIsNotNone(parsed)
        for key in [
            'Primary_Reason', 'P14_Status', 'P56_Status',
            'Too_Old_For_P14', 'Too_Old_For_P56', 'Unusable_Both', 'Detail'
        ]:
            self.assertIn(key, parsed, f"Missing key: {key}")

    def test_parse_unschedulable_none(self):
        parsed = parse_unschedulable_reason(None)
        self.assertIsNotNone(parsed)
        self.assertEqual(parsed['Primary_Reason'], 'Unknown')

    def test_parse_unschedulable_wild_genotype(self):
        reason = 'Wild genotype — not usable for harvest'
        parsed = parse_unschedulable_reason(reason)
        self.assertIn('Wild', parsed['Primary_Reason'])
        self.assertNotIn('Genotype Deadline', parsed['Primary_Reason'])
        self.assertNotIn('Blank', parsed['Primary_Reason'])

    def test_parse_unschedulable_blank_genotype(self):
        parsed = parse_unschedulable_reason('Blank')
        self.assertIn('Blank', parsed['Primary_Reason'])
        self.assertNotIn('Wild', parsed['Primary_Reason'])

    # ── Blank genotype misclassification tests ────────────────────────────────

    def test_blank_genotype_not_classified_as_wild(self):
        """Blank genotype must never appear as Wild Genotype in unschedulable report."""
        geno_excluded = pd.DataFrame([{
            'Animal_Name': 'BlankAnimal1',
            'Birth_ID': 'B099',
            'Strain': 'SHANK3',
            'Genotype': GENOTYPE_BLANK,
            'Sex': 'Male',
            'Birth_Date': '2025-10-01',
            'Reason': (
                "'Half' STRAIN — 2 blank genotype(s) from birth 2025-10-01. "
                "~1 of 2 expected Het (50% from Het×WT). "
                "RECOMMEND: Genotype by 2025-11-12 (7 days) for P14 on 2025-11-15"
            ),
        }])

        report = create_unschedulable_report(
            pd.DataFrame(),
            pd.DataFrame(),
            geno_excluded
        )
        self.assertEqual(len(report), 1)
        row = report.iloc[0]

        self.assertEqual(row['Genotype'], GENOTYPE_BLANK,
            f"Blank genotype shown as '{row['Genotype']}' — should be '{GENOTYPE_BLANK}'")

        self.assertNotIn('Wild', row['Primary_Reason'],
            f"Primary_Reason incorrectly says '{row['Primary_Reason']}' "
            f"for a blank-genotype animal")

        self.assertIn('🧬', row['Primary_Reason'])
        self.assertIn('Blank', row['Primary_Reason'])

    def test_blank_genotype_nan_not_wild(self):
        """NaN genotype must resolve to Blank, not Wild."""
        geno_excluded = pd.DataFrame([{
            'Animal_Name': 'NaNGenoAnimal',
            'Birth_ID': 'B100',
            'Strain': 'CHD8',
            'Genotype': float('nan'),
            'Sex': 'Female',
            'Birth_Date': '2025-10-01',
            'Reason': 'blank genotype — scheduling analysis pending',
        }])

        report = create_unschedulable_report(
            pd.DataFrame(), pd.DataFrame(), geno_excluded
        )
        self.assertEqual(len(report), 1)
        self.assertEqual(report.iloc[0]['Genotype'], GENOTYPE_BLANK)
        self.assertNotEqual(report.iloc[0]['Genotype'], GENOTYPE_WILD)

    def test_blank_genotype_empty_string_not_wild(self):
        """Empty string genotype must resolve to Blank."""
        geno_excluded = pd.DataFrame([{
            'Animal_Name': 'EmptyGenoAnimal',
            'Birth_ID': 'B101',
            'Strain': 'FMR1',
            'Genotype': '',
            'Sex': 'Male',
            'Birth_Date': '2025-10-01',
            'Reason': "'All' STRAIN — 3 blank genotype(s) from birth 2025-10-01.",
        }])

        report = create_unschedulable_report(
            pd.DataFrame(), pd.DataFrame(), geno_excluded
        )
        self.assertEqual(len(report), 1)
        self.assertEqual(report.iloc[0]['Genotype'], GENOTYPE_BLANK)

    def test_parse_reason_blank_scheduling_analysis_not_wild(self):
        """Scheduling analysis strings for blank animals must not be parsed as Wild."""
        blank_scheduling_reasons = [
            "'Half' STRAIN — 2 blank genotype(s) from birth 2025-10-01. ~1 Het expected.",
            "'All' STRAIN — 4 blank genotype(s) from birth 2025-10-01. ALL 4 usable.",
            "UNKNOWN STRAIN — 1 blank genotype(s) from birth 2025-10-01.",
            "blank genotype analysis: not schedulable",
            "Blank",
            GENOTYPE_BLANK,
        ]
        for reason in blank_scheduling_reasons:
            parsed = parse_unschedulable_reason(reason)
            self.assertNotIn('Wild', parsed['Primary_Reason'],
                f"Reason '{reason[:60]}' was misclassified as Wild: "
                f"'{parsed['Primary_Reason']}'")
            self.assertIn('Blank', parsed['Primary_Reason'],
                f"Reason '{reason[:60]}' should be Blank but got: "
                f"'{parsed['Primary_Reason']}'")

    def test_parse_reason_wild_genotype_exact_string(self):
        """Exact wild-genotype exclusion reason must still be classified as Wild."""
        wild_reasons = [
            'Wild genotype — not usable for harvest',
            GENOTYPE_WILD,
        ]
        for reason in wild_reasons:
            parsed = parse_unschedulable_reason(reason)
            self.assertIn('Wild', parsed['Primary_Reason'],
                f"Reason '{reason}' should be Wild but got: '{parsed['Primary_Reason']}'")
            self.assertNotIn('Blank', parsed['Primary_Reason'])

    def test_resolve_genotype_blank_canonical_stays_blank(self):
        """The canonical string 'Blank' must resolve to Blank."""
        geno_excluded = pd.DataFrame([{
            'Animal_Name': 'Test',
            'Birth_ID': 'B001',
            'Strain': 'SHANK3',
            'Genotype': 'Blank',
            'Sex': 'Male',
            'Birth_Date': '2025-10-01',
            'Reason': 'Blank',
        }])
        report = create_unschedulable_report(
            pd.DataFrame(), pd.DataFrame(), geno_excluded
        )
        self.assertEqual(report.iloc[0]['Genotype'], GENOTYPE_BLANK)
        self.assertNotEqual(report.iloc[0]['Genotype'], GENOTYPE_WILD)

    def test_wild_genotype_excluded_animals_show_wild(self):
        """Animals excluded in first pass as Wild must show Wild genotype in report."""
        wild_excluded = pd.DataFrame([{
            'Animal_Name': 'WildAnimal1',
            'Birth_ID': 'B200',
            'Strain': 'SHANK3',
            'Genotype': GENOTYPE_WILD,
            'Sex': 'Male',
            'Birth_Date': '2025-10-01',
            'Reason': 'Wild genotype — not usable for harvest',
        }])

        report = create_unschedulable_report(
            pd.DataFrame(), pd.DataFrame(), wild_excluded
        )
        self.assertEqual(len(report), 1)
        row = report.iloc[0]
        self.assertEqual(row['Genotype'], GENOTYPE_WILD,
            f"Wild animal shown as '{row['Genotype']}' — should be '{GENOTYPE_WILD}'")
        self.assertIn('Wild', row['Primary_Reason'],
            f"Primary_Reason should mention Wild but got: '{row['Primary_Reason']}'")
        self.assertNotIn('Blank', row['Primary_Reason'])

    # ── _assess_genotype_worth_it tests ───────────────────────────────────────

    def test_assess_genotype_worth_it_all_strain_both_available(self):
        worth = _assess_genotype_worth_it(
            num_blanks=4, breeding_type='All',
            p14_available=True, p56_available=True,
            is_schedulable=True,
            p14_date=date(2025, 12, 15),
            p56_date=date(2025, 12, 17),
            today=date(2025, 12, 1),
            remaining_needs={},
            strain='CNTNAP2',
            p56_group_size=0,
        )
        self.assertIn('✅ YES', worth['P14_Worth_Genotyping'])
        self.assertIn('✅ YES', worth['P56_Worth_Genotyping'])

    def test_assess_genotype_worth_it_half_strain_good_yield(self):
        worth = _assess_genotype_worth_it(
            num_blanks=6, breeding_type='Half',
            p14_available=True, p56_available=True,
            is_schedulable=True,
            p14_date=date(2025, 12, 15),
            p56_date=date(2025, 12, 17),
            today=date(2025, 12, 1),
            remaining_needs={},
            strain='SHANK3',
            p56_group_size=0,
        )
        self.assertIn('✅ YES', worth['P14_Worth_Genotyping'])
        self.assertIn('✅ YES', worth['P56_Worth_Genotyping'])

    def test_assess_genotype_worth_it_half_strain_low_yield(self):
        worth = _assess_genotype_worth_it(
            num_blanks=1, breeding_type='Half',
            p14_available=True, p56_available=False,
            is_schedulable=True,
            p14_date=date(2025, 12, 15),
            p56_date=None,
            today=date(2025, 12, 1),
            remaining_needs={},
            strain='SHANK3',
            p56_group_size=0,
        )
        self.assertIn('❌ UNLIKELY', worth['P14_Worth_Genotyping'])
        self.assertIn('❌ NO', worth['P56_Worth_Genotyping'])

    def test_assess_genotype_worth_it_not_schedulable(self):
        worth = _assess_genotype_worth_it(
            num_blanks=6, breeding_type='Half',
            p14_available=False, p56_available=False,
            is_schedulable=False,
            p14_date=date(2025, 1, 1),
            p56_date=date(2025, 1, 15),
            today=date(2025, 12, 1),
            remaining_needs={},
            strain='SHANK3',
            p56_group_size=0,
        )
        self.assertIn('❌ NO', worth['P14_Worth_Genotyping'])
        self.assertIn('❌ NO', worth['P56_Worth_Genotyping'])

    def test_assess_genotype_worth_it_quota_met(self):
        strain_key = 'SHANK3'
        remaining = {
            strain_key: {
                'P14': {
                    'Male':   {'Perfusion': {'needed': 0}, 'MERFISH': {'needed': 0}, 'RNAseq': {'needed': 0}},
                    'Female': {'Perfusion': {'needed': 0}, 'MERFISH': {'needed': 0}, 'RNAseq': {'needed': 0}},
                },
                'P56': {
                    'Male':   {'Perfusion': {'needed': 0}, 'MERFISH': {'needed': 0}, 'RNAseq': {'needed': 0}},
                    'Female': {'Perfusion': {'needed': 0}, 'MERFISH': {'needed': 0}, 'RNAseq': {'needed': 0}},
                },
            }
        }
        worth = _assess_genotype_worth_it(
            num_blanks=6, breeding_type='Half',
            p14_available=True, p56_available=True,
            is_schedulable=True,
            p14_date=date(2025, 12, 15),
            p56_date=date(2025, 12, 17),
            today=date(2025, 12, 1),
            remaining_needs=remaining,
            strain='SHANK3',
            p56_group_size=0,
        )
        self.assertIn('QUOTA MET', worth['P14_Worth_Genotyping'])
        self.assertIn('QUOTA MET', worth['P56_Worth_Genotyping'])

    def test_assess_genotype_worth_it_b6_never_quota_met(self):
        worth = _assess_genotype_worth_it(
            num_blanks=4, breeding_type='All',
            p14_available=True, p56_available=True,
            is_schedulable=True,
            p14_date=date(2025, 12, 15),
            p56_date=date(2025, 12, 17),
            today=date(2025, 12, 1),
            remaining_needs={},
            strain='B6J',
            p56_group_size=0,
        )
        self.assertNotIn('QUOTA MET', worth['P14_Worth_Genotyping'])
        self.assertNotIn('QUOTA MET', worth['P56_Worth_Genotyping'])

    def test_assess_genotype_worth_it_past_deadline(self):
        worth = _assess_genotype_worth_it(
            num_blanks=6, breeding_type='Half',
            p14_available=False, p56_available=False,
            is_schedulable=False,
            p14_date=date(2024, 1, 1),
            p56_date=date(2024, 1, 15),
            today=date(2025, 12, 1),
            remaining_needs={},
            strain='SHANK3',
            p56_group_size=0,
        )
        self.assertIn('❌ NO', worth['P14_Worth_Genotyping'])
        self.assertIn('❌ NO', worth['P56_Worth_Genotyping'])

    # ── NEW: P56 Wednesday-level grouping tests ───────────────────────────────

    def test_assess_genotype_worth_it_p56_group_size_18_half_strain(self):
        """18 females sharing one Wednesday → 9 expected Hets → full cage → YES."""
        worth = _assess_genotype_worth_it(
            num_blanks=3,
            breeding_type='Half',
            p14_available=True,
            p56_available=True,
            is_schedulable=True,
            p14_date=date(2026, 3, 31),
            p56_date=date(2026, 5, 6),
            today=date(2026, 3, 20),
            remaining_needs={},
            strain='SHANK3',
            p56_group_size=18,
        )
        # 18 * 0.5 = 9 expected Hets >= CAGE_SIZE(3) → YES
        self.assertIn('✅ YES', worth['P56_Worth_Genotyping'])
        self.assertIn('9', worth['P56_Worth_Genotyping'])

    def test_assess_genotype_worth_it_p56_group_size_4_half_strain(self):
        """4 animals in Wednesday window → 2 expected Hets → MAYBE (< cage size of 3)."""
        worth = _assess_genotype_worth_it(
            num_blanks=4,
            breeding_type='Half',
            p14_available=True,
            p56_available=True,
            is_schedulable=True,
            p14_date=date(2026, 3, 31),
            p56_date=date(2026, 5, 6),
            today=date(2026, 3, 20),
            remaining_needs={},
            strain='SHANK3',
            p56_group_size=4,
        )
        # 4 * 0.5 = 2 expected Hets < CAGE_SIZE(3) → MAYBE
        self.assertIn('🟡 MAYBE', worth['P56_Worth_Genotyping'])

    def test_assess_genotype_worth_it_p56_group_size_1_half_strain(self):
        """1 animal in Wednesday window → 0.5 expected Hets → UNLIKELY."""
        worth = _assess_genotype_worth_it(
            num_blanks=1,
            breeding_type='Half',
            p14_available=True,
            p56_available=True,
            is_schedulable=True,
            p14_date=date(2026, 3, 31),
            p56_date=date(2026, 5, 6),
            today=date(2026, 3, 20),
            remaining_needs={},
            strain='SHANK3',
            p56_group_size=1,
        )
        # 1 * 0.5 = 0.5 < 1 → UNLIKELY
        self.assertIn('❌ UNLIKELY', worth['P56_Worth_Genotyping'])

    def test_assess_genotype_worth_it_p56_group_all_strain_large_group(self):
        """'All' strain: 6 animals in window → all 6 usable → YES."""
        worth = _assess_genotype_worth_it(
            num_blanks=2,
            breeding_type='All',
            p14_available=True,
            p56_available=True,
            is_schedulable=True,
            p14_date=date(2026, 3, 31),
            p56_date=date(2026, 5, 6),
            today=date(2026, 3, 20),
            remaining_needs={},
            strain='CNTNAP2',
            p56_group_size=6,
        )
        # 6 * 1.0 = 6 >= CAGE_SIZE(3) → YES
        self.assertIn('✅ YES', worth['P56_Worth_Genotyping'])
        self.assertIn('6', worth['P56_Worth_Genotyping'])

    def test_assess_genotype_worth_it_p56_group_size_zero_falls_back(self):
        """p56_group_size=0 falls back to num_blanks."""
        worth_explicit = _assess_genotype_worth_it(
            num_blanks=6, breeding_type='Half',
            p14_available=True, p56_available=True,
            is_schedulable=True,
            p14_date=date(2026, 3, 31),
            p56_date=date(2026, 5, 6),
            today=date(2026, 3, 20),
            remaining_needs={}, strain='SHANK3',
            p56_group_size=6,
        )
        worth_fallback = _assess_genotype_worth_it(
            num_blanks=6, breeding_type='Half',
            p14_available=True, p56_available=True,
            is_schedulable=True,
            p14_date=date(2026, 3, 31),
            p56_date=date(2026, 5, 6),
            today=date(2026, 3, 20),
            remaining_needs={}, strain='SHANK3',
            p56_group_size=0,   # should use num_blanks=6
        )
        self.assertEqual(
            worth_explicit['P56_Worth_Genotyping'],
            worth_fallback['P56_Worth_Genotyping'],
        )

    def test_get_p56_behavior_wednesday_valid_birth(self):
        """Birth date of 2025-10-01 should map to a Wednesday in the P42-P49 window."""
        birth = date(2025, 10, 1)
        result = get_p56_behavior_wednesday(birth)
        self.assertIsNotNone(result)
        self.assertEqual(result.weekday(), 2)   # Wednesday
        dates = calculate_schedule_dates(birth)
        self.assertGreaterEqual(result, dates['p56_behavior_window_start'])
        self.assertLessEqual(result,   dates['p56_behavior_window_end'])

    def test_get_p56_behavior_wednesday_none_input(self):
        self.assertIsNone(get_p56_behavior_wednesday(None))

    def test_nearby_birth_dates_share_wednesday(self):
        """Birth dates 3/17/26-3/24/26 should all map to the same Wednesday."""
        births = [date(2026, 3, 17) + timedelta(days=i) for i in range(8)]
        wednesdays = [get_p56_behavior_wednesday(b) for b in births]
        valid = [w for w in wednesdays if w is not None]
        self.assertGreater(len(valid), 0)
        # All valid results should be the same Wednesday
        self.assertEqual(len(set(valid)), 1,
                         f"Expected one unique Wednesday, got: {set(valid)}")
        # That Wednesday should be 2026-05-06
        self.assertEqual(valid[0], date(2026, 5, 6))

    # ── B6 monthly summary tests ──────────────────────────────────────────────

    def test_b6_monthly_summary_empty_input(self):
        result = create_b6_monthly_summary(pd.DataFrame())
        self.assertIsInstance(result, pd.DataFrame)
        self.assertEqual(len(result), 0)

    def test_b6_monthly_summary_meets_minimum(self):
        today = datetime.now().date()
        target_date = today + timedelta(days=35)
        min_count = CONFIG.get('B6_MIN_PER_MONTH', 3)
        rows = []
        for i in range(min_count):
            rows.append({
                'Animal_Name': f'B6J_{i}',
                'Strain': 'B6J',
                'Assigned_Timepoint': 'P14',
                'P14_Date': target_date,
                'P56_Harvest_Date': None,
                'Priority': 'B6_CONTROL',
            })
        df = pd.DataFrame(rows)
        summary = create_b6_monthly_summary(df)
        if len(summary) > 0:
            row = summary.iloc[0]
            self.assertEqual(row['B6_B6N_Count'], min_count)
            self.assertIn('✅', row['Status'])
            self.assertEqual(row['Shortfall'], 0)

    def test_b6_monthly_summary_below_minimum(self):
        today = datetime.now().date()
        target_date = today + timedelta(days=35)
        min_needed = CONFIG.get('B6_MIN_PER_MONTH', 3)
        rows = [{
            'Animal_Name': 'B6J_0',
            'Strain': 'B6J',
            'Assigned_Timepoint': 'P14',
            'P14_Date': target_date,
            'P56_Harvest_Date': None,
            'Priority': 'B6_CONTROL',
        }]
        df = pd.DataFrame(rows)
        summary = create_b6_monthly_summary(df)
        if len(summary) > 0:
            row = summary.iloc[0]
            self.assertEqual(row['B6_B6N_Count'], 1)
            self.assertIn('⚠️', row['Status'])
            self.assertEqual(row['Shortfall'], min_needed - 1)

    def test_b6_monthly_summary_non_b6_not_counted(self):
        today = datetime.now().date()
        target_date = today + timedelta(days=35)
        rows = [
            {
                'Animal_Name': 'SHANK3_0',
                'Strain': 'SHANK3',
                'Assigned_Timepoint': 'P14',
                'P14_Date': target_date,
                'P56_Harvest_Date': None,
                'Priority': 'HIGH',
            },
            {
                'Animal_Name': 'SHANK3_1',
                'Strain': 'SHANK3',
                'Assigned_Timepoint': 'P14',
                'P14_Date': target_date,
                'P56_Harvest_Date': None,
                'Priority': 'HIGH',
            },
        ]
        df = pd.DataFrame(rows)
        summary = create_b6_monthly_summary(df)
        if len(summary) > 0:
            row = summary.iloc[0]
            self.assertEqual(row['B6_B6N_Count'], 0)
            self.assertIn('⚠️', row['Status'])

    # ── Births sexing schedule tests ──────────────────────────────────────────

    def test_build_births_sexing_schedule_empty(self):
        result = build_births_sexing_schedule(pd.DataFrame())
        self.assertIsInstance(result, pd.DataFrame)
        self.assertEqual(len(result), 0)

    def test_build_births_sexing_schedule_none(self):
        result = build_births_sexing_schedule(None)
        self.assertIsInstance(result, pd.DataFrame)
        self.assertEqual(len(result), 0)

    def test_build_births_sexing_schedule_columns(self):
        births = pd.DataFrame({
            'Birth ID': ['B001'],
            'Status': ['Sing Inventory'],
            'Birth Date': [pd.Timestamp(datetime.now().date() - timedelta(days=5))],
            'Live Count': [6],
            '# of Pups': [6],
            'Line (Short)': ['SHANK3'],
            'Dam': ['Dam1'],
            'Sire': ['Sire1'],
        })
        result = build_births_sexing_schedule(births)
        self.assertGreater(len(result), 0)
        for col in [
            'Birth_ID', 'Strain', 'Birth_Date', 'Num_Pups',
            'Sexing_Date', 'Day_of_Week', 'Days_Until_Sexing',
            'Sexing_Status', 'P14_Expected_Date', 'P14_Day_of_Week',
        ]:
            self.assertIn(col, result.columns, f"Missing column: {col}")

    def test_build_births_sexing_schedule_correct_date(self):
        birth_date = datetime.now().date() - timedelta(days=3)
        expected_sexing = birth_date + timedelta(days=9)
        births = pd.DataFrame({
            'Birth ID': ['B001'],
            'Status': ['Sing Inventory'],
            'Birth Date': [pd.Timestamp(birth_date)],
            'Live Count': [6],
            '# of Pups': [6],
            'Line (Short)': ['SHANK3'],
            'Dam': ['Dam1'],
            'Sire': ['Sire1'],
        })
        result = build_births_sexing_schedule(births)
        self.assertEqual(len(result), 1)
        self.assertEqual(
            result.iloc[0]['Sexing_Date'],
            expected_sexing.strftime('%Y-%m-%d')
        )

    def test_build_births_sexing_schedule_correct_p14_date(self):
        birth_date = datetime.now().date() - timedelta(days=3)
        expected_p14 = birth_date + timedelta(days=14)
        births = pd.DataFrame({
            'Birth ID': ['B001'],
            'Status': ['Sing Inventory'],
            'Birth Date': [pd.Timestamp(birth_date)],
            'Live Count': [6],
            '# of Pups': [6],
            'Line (Short)': ['SHANK3'],
            'Dam': ['Dam1'],
            'Sire': ['Sire1'],
        })
        result = build_births_sexing_schedule(births)
        self.assertEqual(
            result.iloc[0]['P14_Expected_Date'],
            expected_p14.strftime('%Y-%m-%d')
        )

    def test_build_births_sexing_schedule_filters_non_sing(self):
        births = pd.DataFrame({
            'Birth ID': ['B001', 'B002'],
            'Status': ['Sing Inventory', 'Other Status'],
            'Birth Date': [
                pd.Timestamp(datetime.now().date() - timedelta(days=3)),
                pd.Timestamp(datetime.now().date() - timedelta(days=5)),
            ],
            'Live Count': [6, 4],
            '# of Pups': [6, 4],
            'Line (Short)': ['SHANK3', 'CHD8'],
            'Dam': ['Dam1', 'Dam2'],
            'Sire': ['Sire1', 'Sire2'],
        })
        result = build_births_sexing_schedule(births)
        self.assertEqual(len(result), 1)
        self.assertEqual(result.iloc[0]['Birth_ID'], 'B001')

    def test_build_births_sexing_schedule_excludes_already_sexed(self):
        birth_date = datetime.now().date() - timedelta(days=5)
        births = pd.DataFrame({
            'Birth ID': ['B001', 'B002'],
            'Status': ['Sing Inventory', 'Sing Inventory'],
            'Birth Date': [pd.Timestamp(birth_date), pd.Timestamp(birth_date)],
            'Live Count': [6, 4],
            '# of Pups': [6, 4],
            'Line (Short)': ['SHANK3', 'CHD8'],
            'Dam': ['Dam1', 'Dam2'],
            'Sire': ['Sire1', 'Sire2'],
        })
        animals = pd.DataFrame({
            'Birth ID': ['B001', 'B001', 'B001'],
            'Name': ['A1', 'A2', 'A3'],
            'Status': ['Alive', 'Alive', 'Alive'],
        })
        result = build_births_sexing_schedule(births, animals_df=animals)
        self.assertEqual(len(result), 1)
        self.assertEqual(result.iloc[0]['Birth_ID'], 'B002')

    def test_build_births_sexing_schedule_no_animals_df(self):
        birth_date = datetime.now().date() - timedelta(days=5)
        births = pd.DataFrame({
            'Birth ID': ['B001', 'B002'],
            'Status': ['Sing Inventory', 'Sing Inventory'],
            'Birth Date': [pd.Timestamp(birth_date), pd.Timestamp(birth_date)],
            'Live Count': [6, 4],
            '# of Pups': [6, 4],
            'Line (Short)': ['SHANK3', 'CHD8'],
            'Dam': ['Dam1', 'Dam2'],
            'Sire': ['Sire1', 'Sire2'],
        })
        result = build_births_sexing_schedule(births, animals_df=None)
        self.assertEqual(len(result), 2)

    def test_build_births_sexing_schedule_urgency_today(self):
        birth_date = datetime.now().date() - timedelta(days=9)
        births = pd.DataFrame({
            'Birth ID': ['B001'],
            'Status': ['Sing Inventory'],
            'Birth Date': [pd.Timestamp(birth_date)],
            'Live Count': [6],
            '# of Pups': [6],
            'Line (Short)': ['SHANK3'],
            'Dam': ['Dam1'],
            'Sire': ['Sire1'],
        })
        result = build_births_sexing_schedule(births)
        self.assertEqual(len(result), 1)
        self.assertIn('TODAY', result.iloc[0]['Sexing_Status'])

    def test_build_births_sexing_schedule_urgency_done(self):
        birth_date = datetime.now().date() - timedelta(days=15)
        births = pd.DataFrame({
            'Birth ID': ['B001'],
            'Status': ['Sing Inventory'],
            'Birth Date': [pd.Timestamp(birth_date)],
            'Live Count': [6],
            '# of Pups': [6],
            'Line (Short)': ['SHANK3'],
            'Dam': ['Dam1'],
            'Sire': ['Sire1'],
        })
        result = build_births_sexing_schedule(births)
        self.assertEqual(len(result), 1)
        self.assertIn('Done', result.iloc[0]['Sexing_Status'])

    def test_build_births_sexing_schedule_no_birth_date(self):
        births = pd.DataFrame({
            'Birth ID': ['B001'],
            'Status': ['Sing Inventory'],
            'Birth Date': [pd.NaT],
            'Live Count': [6],
            '# of Pups': [6],
            'Line (Short)': ['SHANK3'],
            'Dam': ['Dam1'],
            'Sire': ['Sire1'],
        })
        result = build_births_sexing_schedule(births)
        self.assertEqual(len(result), 1)
        self.assertEqual(result.iloc[0]['Sexing_Date'], 'N/A')
        self.assertEqual(result.iloc[0]['Days_Until_Sexing'], 'N/A')

    # ── Misc utility tests ────────────────────────────────────────────────────

    def test_normalize_genotype_basic(self):
        self.assertEqual(normalize_genotype('-/+'), '-/+')
        self.assertEqual(normalize_genotype('  -/+  '), '-/+')

    def test_normalize_genotype_none(self):
        result = normalize_genotype(None)
        self.assertIsNone(result)

    def test_group_has_quota_b6j_always_true(self):
        self.assertTrue(group_has_quota('B6J', 'Male', 'P14', {}))
        self.assertTrue(group_has_quota('B6J', 'Female', 'P56', {}))
        self.assertTrue(group_has_quota('B6NJ', 'Male', 'P14', {}))

    def test_group_has_quota_empty_needs_true(self):
        self.assertTrue(group_has_quota('SHANK3', 'Male', 'P14', {}))

    def test_group_has_quota_with_needs(self):
        remaining = {
            'SHANK3': {
                'P14': {
                    'Male':   {'Perfusion': {'needed': 3}, 'MERFISH': {'needed': 1}, 'RNAseq': {'needed': 1}},
                    'Female': {'Perfusion': {'needed': 0}, 'MERFISH': {'needed': 0}, 'RNAseq': {'needed': 0}},
                },
                'P56': {
                    'Male':   {'Perfusion': {'needed': 0}, 'MERFISH': {'needed': 0}, 'RNAseq': {'needed': 0}},
                    'Female': {'Perfusion': {'needed': 0}, 'MERFISH': {'needed': 0}, 'RNAseq': {'needed': 0}},
                },
            }
        }
        self.assertTrue(group_has_quota('SHANK3', 'Male', 'P14', remaining))
        self.assertFalse(group_has_quota('SHANK3', 'Female', 'P14', remaining))
        self.assertFalse(group_has_quota('SHANK3', 'Male', 'P56', remaining))

    def test_diagnose_animal_file_runs_without_error(self):
        test_data = pd.DataFrame({
            'Name': ['Animal1', 'Animal2'],
            'Birth Date': [pd.Timestamp('2025-10-01'), pd.Timestamp('2025-09-15')],
            'Sex': ['Male', 'Female'],
            'Line (Short)': ['SHANK3', 'CHD8'],
            'Genotype': ['Het', 'Wild'],
            'Use': ['Sing Inventory', 'Other'],
            'Status': ['Alive', 'Alive'],
            'Birth ID': ['B001', 'B002'],
            'Marker Type': ['Ear Punch', 'Ear Punch'],
        })
        try:
            diagnose_animal_file(test_data)
        except Exception as e:
            self.fail(f"diagnose_animal_file raised an exception: {e}")

    def test_filter_animals_by_use_no_use_column(self):
        test_data = pd.DataFrame({
            'Name': ['Animal1'],
            'Birth Date': [pd.Timestamp('2025-10-01')],
            'Sex': ['Male'],
            'Line (Short)': ['SHANK3'],
            'Genotype': ['Het'],
            'Status': ['Alive'],
            'Birth ID': ['B001'],
            'Marker Type': ['Ear Punch'],
        })
        filtered, excluded = filter_animals_by_use(test_data)
        self.assertEqual(len(filtered), 1)
        self.assertEqual(len(excluded), 0)

    def test_filter_animals_by_use_sing_inventory(self):
        test_data = pd.DataFrame({
            'Name': ['Animal1', 'Animal2', 'Animal3'],
            'Birth Date': [
                pd.Timestamp('2025-10-01'),
                pd.Timestamp('2025-10-01'),
                pd.Timestamp('2025-10-01'),
            ],
            'Sex': ['Male', 'Female', 'Male'],
            'Line (Short)': ['SHANK3', 'SHANK3', 'CHD8'],
            'Genotype': ['Het', 'Het', 'Wild'],
            'Use': ['Sing Inventory', 'Other Use', 'Sing Inventory'],
            'Status': ['Alive', 'Alive', 'Alive'],
            'Birth ID': ['B001', 'B002', 'B003'],
            'Marker Type': ['Ear Punch', 'Ear Punch', 'Ear Punch'],
        })
        filtered, excluded = filter_animals_by_use(test_data)
        self.assertEqual(len(filtered), 2)
        self.assertEqual(len(excluded), 1)
        self.assertIn('Animal2', excluded['Animal_Name'].values)

    def test_filter_animals_by_genotype_excludes_wild(self):
        test_data = pd.DataFrame({
            'Name': ['A1', 'A2', 'A3'],
            'Birth Date': [pd.Timestamp('2025-10-01')] * 3,
            'Sex': ['Male', 'Female', 'Male'],
            'Line (Short)': ['SHANK3', 'SHANK3', 'SHANK3'],
            'Genotype': [GENOTYPE_HET, GENOTYPE_WILD, GENOTYPE_BLANK],
            'Use': ['Sing Inventory'] * 3,
            'Status': ['Alive'] * 3,
            'Birth ID': ['B001', 'B002', 'B003'],
            'Marker Type': ['Ear Punch'] * 3,
        })
        filtered, excluded, blanks = filter_animals_by_genotype_first_pass(test_data)
        self.assertEqual(len(filtered), 1)
        self.assertEqual(len(excluded), 1)
        self.assertEqual(len(blanks), 1)
        self.assertEqual(filtered.iloc[0]['Name'], 'A1')
        self.assertEqual(excluded.iloc[0]['Animal_Name'], 'A2')
        self.assertEqual(blanks.iloc[0]['Name'], 'A3')

    def test_age_at_p14_harvest_days(self):
        birth = date(2025, 11, 3)
        dates = calculate_schedule_dates(birth)
        self.assertIsNotNone(dates)
        age_days = (dates['p14_harvest'] - birth).days
        self.assertEqual(age_days, 14)

    def test_age_at_p56_behavior_range(self):
        birth = date(2025, 10, 1)
        dates = calculate_schedule_dates(birth)
        self.assertIsNotNone(dates)
        first_wed = next_wednesday(dates['p56_behavior_window_start'])
        self.assertIsNotNone(first_wed)
        self.assertLessEqual(first_wed, dates['p56_behavior_window_end'])
        age_days = (first_wed - birth).days
        self.assertGreaterEqual(age_days, 42)
        self.assertLessEqual(age_days, 49)

    def test_age_at_p56_harvest_is_behavior_plus_14(self):
        birth = date(2025, 10, 1)
        dates = calculate_schedule_dates(birth)
        first_wed = next_wednesday(dates['p56_behavior_window_start'])
        harvest = first_wed + timedelta(days=P56_HARVEST_OFFSET_FROM_BEHAVIOR)
        behavior_age = (first_wed - birth).days
        harvest_age = (harvest - birth).days
        self.assertEqual(harvest_age - behavior_age, 14)

    def test_age_months_rounding_boundaries(self):
        self.assertEqual(round(14 / 30.44), 0)
        self.assertEqual(round(42 / 30.44), 1)
        self.assertEqual(round(45 / 30.44), 1)
        self.assertEqual(round(49 / 30.44), 2)
        self.assertEqual(round(56 / 30.44), 2)
        self.assertEqual(round(63 / 30.44), 2)




# ============================================================================
# HARVEST PIPELINE — CONFIGURATION
# ============================================================================

SCHEDULE_FILE_PREFIX = "Complete_Schedule"

LABELS_ACROSS = 5
LABELS_DOWN = 17
LABELS_PER_PAGE = LABELS_ACROSS * LABELS_DOWN  # 85

ADD_SAMPLE_COLUMNS = [
    'Sample Name', 'Type', 'Status', 'Preservation Method', 'Date Harvest',
    'Date Expiration', 'Description', 'Source AnimalID', 'Source SampleID',
    'Volume', 'Volume Units', 'Project', 'Notes'
]

ENVISION_TEMPLATE_COLUMNS = [
    'Group', 'Cage', 'Animal ID', 'Envision Ear Tag', 'Strain',
    'Coat Color', 'Genotype', 'Additional Detail', 'Sex',
    'Birth Date', 'Ear notch', 'Metal ear tag', 'Other ID',
    'RapID code', 'RapID tag color', 'RFID', 'Tail Tattoo'
]

HARVEST_SHEET_COLUMNS = [
    'Name', 'Sample Number', 'Line', 'BD', 'Housing', 'Identification',
    'Sex', 'Age (Days)', 'Envision Date', 'Harvest Date', 'Harvested by',
    'Protocol', 'Time Pickup', 'Time Start', 'Pickup to Harvest Time',
    'Weight g', '4% Tribro mL 10-14', '4% Tribro Units P14-10%', 'Dye',
    '4% PFA per mouse', 'Time Complete', 'Round Duration', '4% PFA Total',
    'Distilled Water', '2xPBS', '16% PFA', 'Notes'
]

PROTOCOL_SORT_ORDER = {
    '8 Weeks - 20mL PBS 25mL 4%PFA (Plus 20mL Storage) - 6mL/min': 0,
    'P14 - 15mL PBS 20mL 4%PFA (Plus 20mL Storage) - 4mL/min': 1,
    'MERFISH - OCT': 2,
    'RNA-Seq': 3,
    'Extra - Sex & Timepoint Full': 4
}



# ============================================================================
# HARVEST PIPELINE — UTILITIES & PIPELINE FUNCTIONS
# ============================================================================

def format_date_only(val):
    """Convert to date only string (no time component)."""
    try:
        dt = pd.to_datetime(val)
        if pd.notna(dt):
            return dt.strftime('%m/%d/%Y')
    except:
        pass
    if pd.notna(val) and str(val).strip() != '':
        return str(val)
    return ''


def combine_sample_numbers(sample_list):
    """Combine sample numbers into range format."""
    if not sample_list:
        return ""
    base_numbers = []
    for sample in sample_list:
        sample_str = str(sample)
        if '-' in sample_str:
            base_num = sample_str.split('-')[0]
        else:
            base_num = sample_str
        try:
            base_numbers.append(int(base_num))
        except (ValueError, TypeError):
            continue
    if not base_numbers:
        return ""
    if len(base_numbers) == 1:
        return str(base_numbers[0])
    else:
        return f"{min(base_numbers)}-{max(base_numbers)}"


def clean_genotype_base(genotype, strain):
    """Remove <content>, Probe, Generic LacZ tg/0, and zygosity markers."""
    if pd.isna(genotype):
        return ""
    if pd.notna(strain):
        strain_str = str(strain).strip()
        if strain_str == 'C57BL/6NJ':
            return 'B6NJ'
        elif strain_str == 'C57BL/6J':
            return 'B6J'
    result = str(genotype)
    result = re.sub(r'<[^>]*>', '', result)
    result = re.sub(r'‹[^›]*›', '', result)
    result = re.sub(r'â€¹[^â€º]*â€º', '', result)
    result = re.sub(r'\[[^\]]*\]', '', result)
    result = re.sub(r'\([^\)]*\)', '', result)
    for ch in ['<', '>', '‹', '›', 'â€¹', 'â€º', '[', ']', '(', ')']:
        result = result.replace(ch, '')
    result = re.sub(r'Probe\s*', '', result)
    result = re.sub(r'Generic LacZ tg/0,\s*', '', result)
    for zyg in ['-/-', '-/+', '+/-', '-/Y', '+/Y']:
        result = result.replace(zyg, '')
    result = result.replace('Inbred', '')
    result = re.sub(r'\s+', ' ', result)
    return result.strip()


def clean_genotype(genotype):
    """Remove <content> and Probe markers only, keep zygosity."""
    if pd.isna(genotype):
        return ""
    result = str(genotype)
    result = re.sub(r'<[^>]*>', '', result)
    result = re.sub(r'‹[^›]*›', '', result)
    result = re.sub(r'â€¹[^â€º]*â€º', '', result)
    result = re.sub(r'\[[^\]]*\]', '', result)
    for ch in ['<', '>', '‹', '›', 'â€¹', 'â€º', '[', ']']:
        result = result.replace(ch, '')
    result = re.sub(r'Probe\s*', '', result)
    result = re.sub(r'Generic LacZ tg/0,\s*', '', result)
    result = re.sub(r'\s+', ' ', result)
    return result.strip()


def clean_genotype_labels(genotype):
    """Clean genotype specifically for label formatting."""
    if pd.isna(genotype):
        return 'N/A'
    result = str(genotype)
    if result in ('', 'nan', 'N/A'):
        return 'N/A'
    result = re.sub(r'‹[^›]*›', '', result)
    result = re.sub(r'<[^>]*>', '', result)
    result = re.sub(r'Generic\s+LacZ\s+tg/0,?\s*', '', result, flags=re.IGNORECASE)
    result = re.sub(r'\bprobe\b', '', result, flags=re.IGNORECASE)
    result = ' '.join(result.split())
    if not result.strip():
        return 'N/A'
    return result


def natural_sort_key(name):
    """Create a sort key that handles numbers naturally."""
    if pd.isna(name):
        return []
    parts = re.split(r'(\d+)', str(name))
    return [int(part) if part.isdigit() else part.lower() for part in parts]


def translate_protocol(harvest_type, timepoint):
    """Translate Harvest_Type + Assigned_Timepoint to full protocol name."""
    harvest_type = str(harvest_type).strip()
    timepoint = str(timepoint).strip()
    if harvest_type == 'Perfusion':
        if timepoint == 'P14':
            return 'P14 - 15mL PBS 20mL 4%PFA (Plus 20mL Storage) - 4mL/min'
        else:
            return '8 Weeks - 20mL PBS 25mL 4%PFA (Plus 20mL Storage) - 6mL/min'
    elif harvest_type == 'MERFISH':
        return 'MERFISH - OCT'
    elif harvest_type == 'RNAseq':
        return 'RNA-Seq'
    elif harvest_type in ('COMPLETE (Quota Filled)', 'Extra'):
        return 'Extra - Sex & Timepoint Full'
    else:
        return 'Extra - Sex & Timepoint Full'


def get_harvest_date(row):
    """Get the appropriate harvest date based on timepoint."""
    timepoint = str(row.get('Assigned_Timepoint', '')).strip()
    if timepoint == 'P14':
        return row.get('P14_Date', '')
    elif timepoint == 'P56':
        return row.get('P56_Harvest_Date', '')
    return ''


def get_age_days(row):
    """Get age at harvest in days based on timepoint."""
    timepoint = str(row.get('Assigned_Timepoint', '')).strip()
    if timepoint == 'P14':
        return row.get('P14_Age_At_Harvest_Days', '')
    elif timepoint == 'P56':
        return row.get('P56_Age_At_Harvest_Days', '')
    return ''


def get_envision_date(row):
    """Get envision (behavior) date — only for P56."""
    timepoint = str(row.get('Assigned_Timepoint', '')).strip()
    if timepoint == 'P56':
        return row.get('P56_Behavior_Date', '')
    return ''


def auto_width_worksheet(ws):
    """Auto-adjust column widths based on content."""
    for column in ws.columns:
        max_length = 0
        col_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                cell_len = len(str(cell.value)) if cell.value is not None else 0
                if cell_len > max_length:
                    max_length = cell_len
            except:
                pass
        ws.column_dimensions[col_letter].width = max(max_length + 3, 8)


def save_df_to_excel(df, filepath, sheet_name='Sheet1'):
    """Save DataFrame to Excel with auto-width columns."""
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    for col_num, header in enumerate(df.columns, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = Font(bold=True)

    for row_num, row_data in enumerate(df.values, 2):
        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num)
            if pd.isna(value):
                cell.value = ''
            else:
                cell.value = value

    auto_width_worksheet(ws)
    wb.save(filepath)


def get_preservation_method(protocol):
    """Determine preservation method based on protocol."""
    protocol = str(protocol).strip()
    if "MERFISH - OCT" in protocol:
        return "OCT Block"
    elif "RNA-Seq" in protocol:
        return "Flash Frozen"
    elif "PFA" in protocol:
        return "4% PFA Fixed"
    else:
        return ""


def get_sample_count(protocol):
    """Determine how many samples to generate based on protocol."""
    protocol = str(protocol).strip()
    if protocol == "8 Weeks - 20mL PBS 25mL 4%PFA (Plus 20mL Storage) - 6mL/min":
        return (1, [""])
    elif protocol == "P14 - 15mL PBS 20mL 4%PFA (Plus 20mL Storage) - 4mL/min":
        return (1, [""])
    elif protocol == "MERFISH - OCT":
        return (2, ["", ""])
    elif protocol == "RNA-Seq":
        return (8, ["-0", "-1", "-2", "-3", "-4", "-5", "-6", "-C"])
    elif protocol == "Extra - Sex & Timepoint Full":
        return (0, [])
    else:
        print(f"  Warning: Unknown protocol '{protocol}'. Defaulting to 1.")
        return (1, [""])


def sort_working_df(df):
    """Sort by Protocol order -> Line (Short) -> Animal_Name (natural sort)."""
    sort_keys = []
    for idx, row in df.iterrows():
        sort_keys.append((
            idx,
            row.get('Protocol_Sort', 99),
            str(row.get('Line (Short)', '')),
            natural_sort_key(row.get('Animal_Name', ''))
        ))
    sort_keys.sort(key=lambda x: (x[1], x[2], x[3]))
    sorted_indices = [item[0] for item in sort_keys]
    return df.loc[sorted_indices].reset_index(drop=True)


# ============================================================
# LOAD DATA
# ============================================================



def build_working_data(all_animals_df):
    """Build working dataset. Filter Unschedulable. Translate protocols."""
    print("\n  Building working data...")
    df = all_animals_df.copy()

    before = len(df)
    df = df[df['Assigned_Timepoint'] != 'Unschedulable'].copy()
    print(f"  Filtered: {before - len(df)} Unschedulable removed")

    # Remove Do Not Schedule animals only — Extras still appear on the
    # harvest worksheet for day-of paperwork, but generate no samples or labels
    before2 = len(df)
    df = df[~df['Harvest_Type'].isin(['DO_NOT_SCHEDULE', 'Do Not Schedule'])].copy()
    filtered2 = before2 - len(df)
    if filtered2 > 0:
        print(f"  Filtered: {filtered2} Do Not Schedule removed from harvest pipeline")
    print(f"  Remaining: {len(df)} animals")

    df['Protocol'] = df.apply(
        lambda row: translate_protocol(
            row.get('Harvest_Type', ''), row.get('Assigned_Timepoint', '')),
        axis=1)
    df['Harvest_Date'] = df.apply(get_harvest_date, axis=1)
    df['Age_Days'] = df.apply(get_age_days, axis=1)
    df['Envision_Date'] = df.apply(get_envision_date, axis=1)
    df['Protocol_Sort'] = df['Protocol'].map(PROTOCOL_SORT_ORDER).fillna(99)
    df['Preservation'] = df['Protocol'].apply(get_preservation_method)

    # Calculate age in weeks
    df['Age_Weeks'] = ''
    for idx, row in df.iterrows():
        try:
            bd = pd.to_datetime(row.get('Birth_Date'))
            hd = pd.to_datetime(row.get('Harvest_Date'))
            if pd.notna(bd) and pd.notna(hd):
                df.at[idx, 'Age_Weeks'] = round((hd - bd).days / 7, 1)
        except:
            pass

    print(f"\n  Protocol breakdown:")
    for protocol, count in df['Protocol'].value_counts().items():
        print(f"    {protocol}: {count}")

    print(f"\n  Columns available in working data ({len(df.columns)}):")
    print(f"    {list(df.columns)}")

    return df


# ============================================================
# STEPS 0+1: BUILD HARVEST WORKSHEET & CREATE SAMPLES
# ============================================================

def get_starting_sample_number():
    """Ask user for the last sample number used — via GUI dialog."""
    return _gui_ask('sample_number')


def run_harvest_and_samples(working_df, timestamp):
    """
    STEPS 0+1: Build harvest worksheet AND create samples together.
    Returns:
        harvest_df: Complete harvest worksheet with sample numbers
        samples_for_chain: DataFrame for Steps 2 and 4 (uses Animal_Name as Source)
        climb_import_df: DataFrame for Climb import (uses Animal ID)
    """
    print("\n" + "=" * 80)
    print("STEPS 0+1: BUILD HARVEST WORKSHEET & CREATE SAMPLES")
    print("=" * 80)

    if working_df.empty:
        print("  ✗ No data. Skipping.")
        return pd.DataFrame(), pd.DataFrame(), pd.DataFrame()

    # Sort working data
    sorted_df = sort_working_df(working_df)

    # Get starting sample number
    next_sample_num = get_starting_sample_number()
    print(f"  Starting with sample number: {next_sample_num}")

    # Build animal ID lookup
    animal_lookup = {}
    if 'ID' in sorted_df.columns:
        for _, row in sorted_df.iterrows():
            aname = str(row.get('Animal_Name', '')).strip()
            aid = row.get('ID', '')
            if aname and pd.notna(aid) and str(aid).strip() != '':
                animal_lookup[aname] = str(aid).strip()
        print(f"  Built Animal ID lookup: {len(animal_lookup)} entries")

    # Process each animal
    harvest_rows = []
    climb_import_rows = []
    chain_rows = []
    samples_added = 0

    for idx, row in sorted_df.iterrows():
        animal_name = str(row.get('Animal_Name', '')).strip()
        protocol = row.get('Protocol', '')
        harvest_date = row.get('Harvest_Date', '')
        envision_date = row.get('Envision_Date', '')
        preservation = get_preservation_method(protocol)
        count, suffixes = get_sample_count(protocol)

        animal_id = animal_lookup.get(animal_name, animal_name)

        # Format dates without time
        harvest_date_str = format_date_only(harvest_date)
        envision_date_str = format_date_only(envision_date)
        birth_date_str = format_date_only(row.get('Birth_Date', ''))

        # Generate samples for this animal
        animal_samples = []
        for i in range(count):
            sample_name = f"{next_sample_num}{suffixes[i]}"

            # Climb import row (uses Animal ID)
            climb_import_rows.append({
                'Sample Name': sample_name,
                'Type': 'Brain',
                'Status': 'Available',
                'Preservation Method': preservation,
                'Date Harvest': harvest_date_str,
                'Date Expiration': '',
                'Description': '',
                'Source AnimalID': animal_id,
                'Source SampleID': '',
                'Volume': '',
                'Volume Units': '',
                'Project': '',
                'Notes': ''
            })

            # Chain row (uses Animal_Name for merging in Steps 2 and 4)
            chain_rows.append({
                'Name': sample_name,
                'Source': animal_name,
                'Preservation': preservation,
                'Harvest Date': harvest_date_str,
            })

            animal_samples.append(sample_name)
            next_sample_num += 1
            samples_added += 1

        # Combined sample number for harvest sheet
        combined_sample = combine_sample_numbers(animal_samples)

        # Harvest sheet row
        harvest_rows.append({
            'Name': animal_name,
            'Sample Number': combined_sample,
            'Line': row.get('Line (Short)', ''),
            'BD': birth_date_str,
            'Housing': row.get('Housing ID', ''),
            'Identification': row.get('Marker_Type', ''),
            'Sex': row.get('Sex', ''),
            'Age (Days)': row.get('Age_Days', ''),
            'Envision Date': envision_date_str,
            'Harvest Date': harvest_date_str,
            'Harvested by': '',
            'Protocol': protocol,
            'Time Pickup': '',
            'Time Start': '',
            'Pickup to Harvest Time': '',
            'Weight g': '',
            '4% Tribro mL 10-14': '',
            '4% Tribro Units P14-10%': '',
            'Dye': '',
            '4% PFA per mouse': '',
            'Time Complete': '',
            'Round Duration': '',
            '4% PFA Total': '',
            'Distilled Water': '',
            '2xPBS': '',
            '16% PFA': '',
            'Notes': ''
        })

    # Build DataFrames
    harvest_df = pd.DataFrame(harvest_rows, columns=HARVEST_SHEET_COLUMNS)
    climb_import_df = pd.DataFrame(climb_import_rows, columns=ADD_SAMPLE_COLUMNS)
    samples_for_chain = pd.DataFrame(chain_rows)

    # Debug info for chain
    print(f"\n  DEBUG chain data:")
    print(f"    samples_for_chain columns: {list(samples_for_chain.columns)}")
    print(f"    samples_for_chain rows: {len(samples_for_chain)}")
    if len(samples_for_chain) > 0:
        print(f"    First row: {samples_for_chain.iloc[0].to_dict()}")
        print(f"    Unique Source values (first 5): {samples_for_chain['Source'].unique()[:5].tolist()}")

    # Save Harvest Sheet Import
    harvest_file = f"Harvest_Sheet_Import_{timestamp}.xlsx"
    save_df_to_excel(harvest_df, harvest_file, sheet_name='Harvest Worksheet')
    print(f"\n  📄 Saved: {harvest_file}")

    # Save Climb Sample Import
    climb_file = f"Climb_Sample_Import_{timestamp}.xlsx"
    save_df_to_excel(climb_import_df, climb_file, sheet_name='Samples')
    print(f"  📄 Saved: {climb_file}")

    print(f"\n  ✓ Steps 0+1 complete:")
    print(f"    {len(harvest_df)} animals on harvest worksheet")
    print(f"    {samples_added} samples created")
    print(f"    {len(climb_import_df)} rows in Climb import")

    return harvest_df, samples_for_chain, climb_import_df


# ============================================================
# STEP 2: DELIVERABLES
# ============================================================

class MultiSheetExporter:
    def __init__(self, working_df, samples_df, output_filename):
        """Initialize using in-memory DataFrames."""
        self.output_filename = output_filename
        self.workbook = Workbook()

        self.working_df = working_df.copy()
        self.samples_df = samples_df.copy()

        print(f"\n  DEBUG Deliverables init:")
        print(f"    samples_df columns: {list(self.samples_df.columns)}")
        print(f"    samples_df rows: {len(self.samples_df)}")
        print(f"    working_df columns (first 15): {list(self.working_df.columns)[:15]}")
        print(f"    working_df rows: {len(self.working_df)}")

        # Parse dates in working data
        for col in ['Birth_Date', 'Wean Date', 'Harvest_Date', 'Envision_Date']:
            if col in self.working_df.columns:
                self.working_df[col] = pd.to_datetime(
                    self.working_df[col], errors='coerce')

        # Parse dates in samples
        if 'Harvest Date' in self.samples_df.columns:
            self.samples_df['Harvest Date'] = pd.to_datetime(
                self.samples_df['Harvest Date'], errors='coerce')

        # Rename sample columns for merge
        # Source contains Animal_Name (not ID) for chain purposes
        if 'Source' in self.samples_df.columns:
            self.samples_df = self.samples_df.rename(columns={'Source': 'Animal_Name'})
        if 'Name' in self.samples_df.columns:
            self.samples_df = self.samples_df.rename(columns={'Name': 'Sample_Name'})

        if 'Animal_Name' in self.samples_df.columns:
            self.samples_df['Animal_Name'] = self.samples_df['Animal_Name'].astype(str).str.strip()
        if 'Animal_Name' in self.working_df.columns:
            self.working_df['Animal_Name'] = self.working_df['Animal_Name'].astype(str).str.strip()

        # Debug merge values
        if 'Animal_Name' in self.samples_df.columns and 'Animal_Name' in self.working_df.columns:
            sample_names = set(self.samples_df['Animal_Name'].unique())
            working_names = set(self.working_df['Animal_Name'].unique())
            common = sample_names.intersection(working_names)
            print(f"    Sample unique Animal_Names: {len(sample_names)}")
            print(f"    Working unique Animal_Names: {len(working_names)}")
            print(f"    Common (will match): {len(common)}")
            if len(common) == 0:
                print(f"    ⚠ NO MATCHES! Sample names: {list(sample_names)[:3]}")
                print(f"    ⚠ Working names: {list(working_names)[:3]}")

        # Merge samples with working data
        if 'Animal_Name' in self.samples_df.columns and 'Animal_Name' in self.working_df.columns:
            self.merged_df = pd.merge(
                self.samples_df, self.working_df,
                on='Animal_Name', how='left',
                suffixes=('_sample', '_animal')
            )
            print(f"    Merged result: {len(self.merged_df)} rows")
            print(f"    Merged columns: {list(self.merged_df.columns)}")
        else:
            self.merged_df = self.samples_df.copy()
            print(f"    No merge possible — using {len(self.merged_df)} sample rows")

        if 'Sheet' in self.workbook.sheetnames:
            del self.workbook['Sheet']

    def _calculate_age_weeks(self, birth_date, harvest_date):
        try:
            if pd.notna(birth_date) and pd.notna(harvest_date):
                birth = pd.to_datetime(birth_date)
                harvest = pd.to_datetime(harvest_date)
                return round((harvest - birth).days / 7, 1)
        except:
            pass
        return ''

    def _safe_get(self, row, *columns, default=''):
        """Try multiple column names, return first non-null value."""
        for col in columns:
            if col in row.index:
                val = row[col]
                if pd.notna(val) and str(val).lower() != 'nan':
                    return val
        return default

    def create_sing_harvest_sheet(self):
        ws = self.workbook.create_sheet("Sing Harvest Sheet")
        print("\n  Creating Sing Harvest Sheet...")

        grouped_data = {}
        for idx, row in self.merged_df.iterrows():
            animal_name = self._safe_get(row, 'Animal_Name')
            sample_name = self._safe_get(row, 'Sample_Name')
            if not animal_name or animal_name == '':
                continue
            if animal_name not in grouped_data:
                grouped_data[animal_name] = {
                    'samples': [],
                    'data': {
                        'Name': animal_name,
                        'Line': self._safe_get(row, 'Line (Short)', 'Line', 'Strain'),
                        'BD': self._safe_get(row, 'Birth_Date'),
                        'Housing': self._safe_get(row, 'Housing ID'),
                        'Identification': self._safe_get(row, 'Marker_Type', 'Marker'),
                        'Sex': self._safe_get(row, 'Sex', 'Sex_animal', 'Sex_sample'),
                        'Age (Days)': self._safe_get(row, 'Age_Days',
                                                      'P56_Age_At_Harvest_Days',
                                                      'P14_Age_At_Harvest_Days')
                    }
                }
            if sample_name:
                grouped_data[animal_name]['samples'].append(sample_name)

        harvest_data = []
        for animal_name, group in grouped_data.items():
            row_data = group['data'].copy()
            row_data['Sample Number'] = combine_sample_numbers(group['samples'])
            harvest_data.append(row_data)

        df = pd.DataFrame(harvest_data)
        column_order = ['Name', 'Sample Number', 'Line', 'BD', 'Housing',
                        'Identification', 'Sex', 'Age (Days)']
        if df.empty:
            df = pd.DataFrame(columns=column_order)
        else:
            for col in column_order:
                if col not in df.columns:
                    df[col] = ''
            df = df[column_order]

        self._apply_sheet_styling(ws, df, column_order)
        print(f"    ✓ {len(df)} rows")
        return ws

    def create_animal_sample_tracking_sheet(self):
        ws = self.workbook.create_sheet("Animal and Sample Tracking")
        print("\n  Creating Animal and Sample Tracking sheet...")

        filtered_df = self.merged_df.copy()
        pres_col = None
        for col_name in ['Preservation', 'Preservation_sample', 'Preservation_animal']:
            if col_name in filtered_df.columns:
                pres_col = col_name
                break

        if pres_col:
            filtered_df = filtered_df[filtered_df[pres_col] == '4% PFA Fixed']
            print(f"    Filtered on '{pres_col}' to {len(filtered_df)} PFA Fixed samples")
        else:
            print(f"    ⚠ No Preservation column found")
            print(f"    Available columns: {list(filtered_df.columns)}")

        tracking_data = []
        for idx, row in filtered_df.iterrows():
            age_weeks = self._calculate_age_weeks(
                self._safe_get(row, 'Birth_Date'),
                self._safe_get(row, 'Harvest Date', 'Harvest_Date')
            )
            tracking_data.append({
                'Name_sample': self._safe_get(row, 'Sample_Name'),
                'Harvest Date': self._safe_get(row, 'Harvest Date', 'Harvest_Date'),
                'Age (weeks)_sample': age_weeks,
                'Name_subject': self._safe_get(row, 'Animal_Name'),
                'Sex': self._safe_get(row, 'Sex', 'Sex_animal', 'Sex_sample'),
                'Line_subject': self._safe_get(row, 'Line', 'Line_animal', 'Strain'),
                'Line (Short)': self._safe_get(row, 'Line (Short)'),
                'Line (Stock)': self._safe_get(row, 'Line (Stock)'),
                'Species_subject': 'Mouse',
                'Genotype': self._safe_get(row, 'Genotype', 'Genotype_animal', 'Genotype_sample'),
                'Birth Date': self._safe_get(row, 'Birth_Date'),
                'Wean Date': self._safe_get(row, 'Wean Date'),
                'Harvest Timepoint': self._safe_get(row, 'Assigned_Timepoint')
            })

        column_order = [
            'Name_sample', 'Harvest Date', 'Age (weeks)_sample', 'Name_subject',
            'Sex', 'Line_subject', 'Line (Short)', 'Line (Stock)',
            'Species_subject', 'Genotype', 'Birth Date', 'Wean Date',
            'Harvest Timepoint'
        ]
        df = pd.DataFrame(tracking_data)
        if df.empty:
            df = pd.DataFrame(columns=column_order)
        else:
            df = df[column_order]

        self._apply_sheet_styling(ws, df, column_order)
        print(f"    ✓ {len(df)} rows (4% PFA Fixed)")
        return ws

    def create_merfish_sample_tracker_sheet(self):
        ws = self.workbook.create_sheet("MERFISH Sample Tracker")
        print("\n  Creating MERFISH Sample Tracker sheet...")

        filtered_df = self.merged_df.copy()
        pres_col = None
        for col_name in ['Preservation', 'Preservation_sample', 'Preservation_animal']:
            if col_name in filtered_df.columns:
                pres_col = col_name
                break

        if pres_col:
            filtered_df = filtered_df[filtered_df[pres_col] == 'OCT Block']
            print(f"    Filtered on '{pres_col}' to {len(filtered_df)} OCT Block samples")
        else:
            print(f"    ⚠ No Preservation column found")

        tracker_data = []
        for idx, row in filtered_df.iterrows():
            age_weeks = self._calculate_age_weeks(
                self._safe_get(row, 'Birth_Date'),
                self._safe_get(row, 'Harvest Date', 'Harvest_Date')
            )
            tracker_data.append({
                'Name_sample': self._safe_get(row, 'Sample_Name'),
                'Age (weeks)_sample': age_weeks,
                'Name_subject': self._safe_get(row, 'Animal_Name'),
                'Sex': self._safe_get(row, 'Sex', 'Sex_animal', 'Sex_sample'),
                'Line_subject': self._safe_get(row, 'Line', 'Line_animal', 'Strain'),
                'Line (Short)': self._safe_get(row, 'Line (Short)'),
                'Line (Stock)': self._safe_get(row, 'Line (Stock)'),
                'Species_subject': 'Mouse',
                'Genotype': self._safe_get(row, 'Genotype', 'Genotype_animal', 'Genotype_sample'),
                'Birth Date': self._safe_get(row, 'Birth_Date'),
                'Wean Date': self._safe_get(row, 'Wean Date'),
                'Dissect Date': self._safe_get(row, 'Harvest Date', 'Harvest_Date')
            })

        column_order = [
            'Name_sample', 'Age (weeks)_sample', 'Name_subject', 'Sex',
            'Line_subject', 'Line (Short)', 'Line (Stock)', 'Species_subject',
            'Genotype', 'Birth Date', 'Wean Date', 'Dissect Date'
        ]
        df = pd.DataFrame(tracker_data)
        if df.empty:
            df = pd.DataFrame(columns=column_order)
        else:
            df = df[column_order]

        self._apply_sheet_styling(ws, df, column_order)
        print(f"    ✓ {len(df)} rows (OCT Block)")
        return ws

    def create_rnaseq_sample_tracker_sheet(self):
        ws = self.workbook.create_sheet("RNASeq Sample Tracker")
        print("\n  Creating RNASeq Sample Tracker sheet...")

        filtered_df = self.merged_df.copy()
        pres_col = None
        for col_name in ['Preservation', 'Preservation_sample', 'Preservation_animal']:
            if col_name in filtered_df.columns:
                pres_col = col_name
                break

        if pres_col:
            filtered_df = filtered_df[
                filtered_df[pres_col].isin(['Flash Frozen', 'Frozen'])]
            print(f"    Filtered on '{pres_col}' to {len(filtered_df)} Flash Frozen samples")
        else:
            print(f"    ⚠ No Preservation column found")

        tracker_data = []
        for idx, row in filtered_df.iterrows():
            age_weeks = self._calculate_age_weeks(
                self._safe_get(row, 'Birth_Date'),
                self._safe_get(row, 'Harvest Date', 'Harvest_Date')
            )
            tracker_data.append({
                'Name_sample': self._safe_get(row, 'Sample_Name'),
                'Age (weeks)_sample': age_weeks,
                'Name_subject': self._safe_get(row, 'Animal_Name'),
                'Sex': self._safe_get(row, 'Sex', 'Sex_animal', 'Sex_sample'),
                'Line_subject': self._safe_get(row, 'Line', 'Line_animal', 'Strain'),
                'Line (Short)': self._safe_get(row, 'Line (Short)'),
                'Line (Stock)': self._safe_get(row, 'Line (Stock)'),
                'Species_subject': 'Mouse',
                'Genotype': self._safe_get(row, 'Genotype', 'Genotype_animal', 'Genotype_sample'),
                'Birth Date': self._safe_get(row, 'Birth_Date'),
                'Wean Date': self._safe_get(row, 'Wean Date'),
                'Dissect Date': self._safe_get(row, 'Harvest Date', 'Harvest_Date')
            })

        column_order = [
            'Name_sample', 'Age (weeks)_sample', 'Name_subject', 'Sex',
            'Line_subject', 'Line (Short)', 'Line (Stock)', 'Species_subject',
            'Genotype', 'Birth Date', 'Wean Date', 'Dissect Date'
        ]
        df = pd.DataFrame(tracker_data)
        if df.empty:
            df = pd.DataFrame(columns=column_order)
        else:
            df = df[column_order]

        self._apply_sheet_styling(ws, df, column_order)
        print(f"    ✓ {len(df)} rows (Flash Frozen)")
        return ws

    def _apply_sheet_styling(self, ws, df, column_order):
        """Apply consistent styling to a worksheet."""
        for col_num, header in enumerate(column_order, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = Font(bold=True, size=12, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092",
                                    fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin'))

        for row_num, row_data in enumerate(df.values, 2):
            for col_num, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col_num)
                if pd.isna(value):
                    cell.value = ''
                else:
                    cell.value = value
                cell.alignment = Alignment(horizontal="left", vertical="center")
                cell.border = Border(
                    left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'), bottom=Side(style='thin'))
                if 'Date' in column_order[col_num - 1] and cell.value:
                    try:
                        if pd.notna(cell.value) and cell.value != '':
                            cell.number_format = 'MM/DD/YYYY'
                    except:
                        pass

        auto_width_worksheet(ws)

    def create_all_sheets(self):
        self.create_sing_harvest_sheet()
        self.create_animal_sample_tracking_sheet()
        self.create_merfish_sample_tracker_sheet()
        self.create_rnaseq_sample_tracker_sheet()

    def save(self):
        self.workbook.save(self.output_filename)
        print(f"\n  📄 Saved: {self.output_filename}")
        return self.output_filename


def run_deliverables(working_df, samples_df, timestamp):
    """STEP 2: Create multi-sheet deliverables Excel file."""
    print("\n" + "=" * 80)
    print("STEP 2: DELIVERABLES")
    print("=" * 80)

    if samples_df.empty:
        print("  ✗ No sample data. Skipping.")
        return None

    output_filename = f"Lab_Data_Export_{timestamp}.xlsx"
    try:
        exporter = MultiSheetExporter(
            working_df=working_df,
            samples_df=samples_df,
            output_filename=output_filename
        )
        exporter.create_all_sheets()
        saved_file = exporter.save()
        print(f"\n  ✓ Step 2 complete: 4 sheets created")
        return saved_file
    except Exception as e:
        print(f"  ✗ Error: {e}")
        traceback.print_exc()
        return None


# ============================================================
# STEP 3: CLIMB TO ENVISION
# ============================================================

def assign_ear_tags_by_strain_sex(df):
    """Assign S4, S3, S2 in repeating pattern after sorting."""
    df_sorted = df.copy()
    sort_data = []
    for idx, row in df_sorted.iterrows():
        sort_data.append((
            idx, row['Line'], row['Sex'],
            natural_sort_key(row['Animal_Name'])
        ))
    sort_data.sort(key=lambda x: (x[1], x[2], x[3]))
    sorted_indices = [item[0] for item in sort_data]
    df_sorted = df_sorted.loc[sorted_indices].reset_index(drop=True)

    tags = []
    current_strain = None
    current_sex = None
    counter = 0

    for idx, row in df_sorted.iterrows():
        strain = row['Line']
        sex = row['Sex']
        if strain != current_strain or sex != current_sex:
            current_strain = strain
            current_sex = sex
            counter = 0
        position = (counter % 3) + 1
        tags.append('S4' if position == 1 else 'S3' if position == 2 else 'S2')
        counter += 1

    df_sorted['Envision Ear Tag'] = tags
    return df_sorted


def group_animals_by_housing(df):
    """Group animals with same Group ID, numbering groups of 3."""
    group_housing_counts = defaultdict(lambda: defaultdict(list))
    for idx, row in df.iterrows():
        group_housing_counts[row['Group_base']][row['Housing ID']].append(idx)

    group_suffixes = {}
    for group_name, housing_dict in group_housing_counts.items():
        total = sum(len(v) for v in housing_dict.values())
        if total <= 3:
            for indices in housing_dict.values():
                for idx in indices:
                    group_suffixes[idx] = group_name
        else:
            assigned = 0
            for housing_id, indices in sorted(housing_dict.items()):
                for idx in indices:
                    group_suffixes[idx] = f"{group_name}{(assigned // 3) + 1}"
                    assigned += 1
    return group_suffixes


def run_climb_to_envision(working_df, timestamp):
    """STEP 3: Create Envision translation."""
    print("\n" + "=" * 80)
    print("STEP 3: CLIMB TO ENVISION")
    print("=" * 80)

    if working_df.empty:
        print("  ✗ No data. Skipping.")
        return None

    df = working_df.copy()

    required = ['Genotype', 'Sex', 'Housing ID', 'Animal_Name', 'Line', 'Birth_Date']
    missing = [col for col in required if col not in df.columns]
    if missing:
        print(f"  ✗ Missing columns: {missing}")
        return None

    print(f"  Processing {len(df)} animals...")

    df['genotype_base'] = df.apply(
        lambda row: clean_genotype_base(row['Genotype'], row['Line']), axis=1)
    df['sex_initial'] = df['Sex'].str[0].str.upper()
    df['Group_base'] = df['genotype_base'] + '-' + df['sex_initial']

    group_suffixes = group_animals_by_housing(df)
    df['Group'] = df.index.map(group_suffixes)

    df = assign_ear_tags_by_strain_sex(df)
    df['Genotype_clean'] = df['Genotype'].apply(clean_genotype)

    output_df = pd.DataFrame({
        'Group': df['Group'],
        'Cage': df['Housing ID'],
        'Animal ID': df['Animal_Name'],
        'Envision Ear Tag': df['Envision Ear Tag'],
        'Strain': df['Line'],
        'Coat Color': '',
        'Genotype': df['Genotype_clean'],
        'Additional Detail': '',
        'Sex': df['Sex'],
        'Birth Date': df['Birth_Date'],
        'Ear notch': '',
        'Metal ear tag': '',
        'Other ID': '',
        'RapID code': '',
        'RapID tag color': '',
        'RFID': '',
        'Tail Tattoo': ''
    })
    output_df = output_df[ENVISION_TEMPLATE_COLUMNS]

    output_filename = f"Envision_{timestamp}.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = 'template_csv_v1.0'

    for col_num, header in enumerate(ENVISION_TEMPLATE_COLUMNS, 1):
        ws.cell(row=1, column=col_num, value=header)

    for row_num, row_data in enumerate(output_df.values, 2):
        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num)
            if pd.isna(value):
                cell.value = ''
            else:
                cell.value = value

    auto_width_worksheet(ws)
    wb.save(output_filename)

    print(f"  📄 Saved: {output_filename}")
    print(f"  ✓ Step 3 complete: {len(output_df)} animals")

    group_counts = output_df.groupby(['Group', 'Cage']).size().reset_index(name='Count')
    print(f"\n  Group Summary:")
    for _, row in group_counts.iterrows():
        print(f"    {row['Group']} | Cage {row['Cage']} | {row['Count']} animals")

    return output_filename


# ============================================================
# STEP 4: LABELS
# ============================================================

def safe_date_format(date_value, date_name='Date'):
    try:
        return pd.to_datetime(date_value).strftime('%m/%d/%y')
    except:
        if pd.notna(date_value):
            return str(date_value)
        return 'N/A'


def safe_int_format(value, default='N/A'):
    try:
        if pd.notna(value):
            return int(float(value))
        return default
    except:
        return default


def safe_get_label(row, *keys, default='N/A'):
    """Try multiple keys, return first non-null value."""
    for key in keys:
        if key in row:
            value = row[key]
            if isinstance(value, pd.Series):
                value = value.dropna()
                if not value.empty:
                    return value.iloc[0]
            elif pd.notna(value) and str(value).lower() != 'nan':
                return value
    return default


def determine_label_type(preservation):
    preservation_str = str(preservation).strip().lower()
    if 'oct' in preservation_str and 'block' in preservation_str:
        return 'skip', 0
    elif 'frozen' in preservation_str:
        return 'rna', 1
    elif 'pfa' in preservation_str or 'fixed' in preservation_str:
        return 'perfusion', 2
    else:
        return 'rna', 1


def format_sample_number(sample_name, pad=True):
    """
    Format sample name for RNA tube labels.
    Sample names are in the form '765-0', '765-C', etc.
    (numeric part + hyphen + tube suffix 0-6 or C)

    Sides tab (pad=True):  zero-pad numeric part to 4 digits  -> '0765-0'
    Tops  tab (pad=False): strip leading zero from numeric part -> '765-0'
    """
    try:
        s = str(sample_name).strip()
        if '-' in s:
            parts    = s.rsplit('-', 1)
            num_part = parts[0]
            suffix   = parts[1]
        else:
            num_part = s
            suffix   = None
        digits = ''.join(filter(str.isdigit, num_part))
        if not digits:
            return s
        formatted_num = digits.zfill(4) if pad else str(int(digits))
        return f"{formatted_num}-{suffix}" if suffix is not None else formatted_num
    except:
        return str(sample_name)


def create_rna_excel(rna_labels, output_folder, timestamp):
    """
    Create RNA tube labeler Excel file with two tabs: Sides and Tops.
    One row per label — no grid/page logic.
    """
    if not rna_labels:
        _pipeline_queue.put({'kind': _MSG_LOG, 'text': '  ⚠ No RNA labels to create.'})
        return None

    _pipeline_queue.put({'kind': _MSG_LOG, 'text': '  Creating RNA Tube Labeler file...'})

    # Error check: label numbers must match between Sides and Tops
    mismatches = [i + 1 for i, lbl in enumerate(rna_labels)
                  if lbl['Sides_Label_Num'] != lbl['Tops_Label_Num']]
    if mismatches:
        _pipeline_queue.put({'kind': _MSG_LOG,
                             'text': f'  ❌ RNA label number mismatch at positions: {mismatches}'})
        raise ValueError(f'RNA label number mismatch at rows: {mismatches}')

    sides_df = pd.DataFrame({
        'Label Number':  [l['Sides_Label_Num'] for l in rna_labels],
        'Sample_Date':   [l['Sides_B']         for l in rna_labels],
        'Animal_Strain': [l['Sides_C']         for l in rna_labels],
    })
    tops_df = pd.DataFrame({
        'Label Number':  [l['Tops_Label_Num'] for l in rna_labels],
        'Sample Number': [l['Tops_B']         for l in rna_labels],
        'Animal Number': [l['Tops_C']         for l in rna_labels],
    })

    output_file = _os.path.join(output_folder, f'Tube_Labeler_RNA_{timestamp}.xlsx')
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        sides_df.to_excel(writer, sheet_name='Sides', index=False, header=False)
        tops_df.to_excel(writer,  sheet_name='Tops',  index=False, header=False)

    _pipeline_queue.put({'kind': _MSG_LOG,
                         'text': f'  ✓ RNA Tube Labeler saved: {_os.path.basename(output_file)}'
                                 f'  (Sides={len(sides_df)}, Tops={len(tops_df)})'})
    return output_file


def format_label_rows(row, label_type):
    """Create the 4 rows of text for each label."""
    harvest_date = safe_date_format(
        safe_get_label(row, 'Sample Harvest Date', 'Harvest Date', 'Harvest_Date'))
    born_date = safe_date_format(
        safe_get_label(row, 'Birth_Date', 'Birth Date'))

    sex_val = safe_get_label(row, 'Sex', 'Sex_animal', 'Sex_sample')
    sex = str(sex_val).upper()[0] if pd.notna(sex_val) and str(sex_val) != 'N/A' else 'U'

    line_stock_val = safe_get_label(row, 'Line (Stock)')
    line_stock = str(line_stock_val).lstrip('0') if pd.notna(line_stock_val) and str(line_stock_val) != 'N/A' else ''

    time_point = str(safe_get_label(row, 'Assigned_Timepoint')).strip()
    if time_point and time_point not in ('N/A', 'nan', ''):
        days = ''.join(filter(str.isdigit, time_point))
        if days:
            time_point = f"P{days}"
    else:
        time_point = "N/A"

    genotype = clean_genotype_labels(
        safe_get_label(row, 'Genotype', 'Genotype_animal', 'Genotype_sample'))

    # Calculate age in weeks
    age_weeks = 'N/A'
    try:
        bd = pd.to_datetime(safe_get_label(row, 'Birth_Date', 'Birth Date'))
        hd = pd.to_datetime(safe_get_label(row, 'Sample Harvest Date', 'Harvest Date', 'Harvest_Date'))
        if pd.notna(bd) and pd.notna(hd):
            age_weeks = int((hd - bd).days / 7)
    except:
        pass

    sample_name = safe_get_label(row, 'Sample Name', 'Sample_Name')
    animal_name = safe_get_label(row, 'Animal Name', 'Animal_Name')
    line_short = safe_get_label(row, 'Line (Short)')

    row1 = f"{sample_name}_{harvest_date}_{animal_name}"
    row2 = f"{age_weeks}Wks_{sex}_{line_short}_{line_stock}"
    row3 = f"{genotype}_{born_date}_{time_point}"
    row4 = "Mouse_Perfused Brain" if label_type.lower() == 'perfusion' else "Mouse_Frozen Brain"

    return row1, row2, row3, row4


def generate_all_labels(merged_df):
    """Generate all label data from merged dataframe.

    Returns:
        perfusion_labels  – list of {'Row 1'…'Row 4'} dicts for mail-merge sheets
        rna_labels        – list of Sides/Tops dicts for the RNA Tube Labeler file
        perfusion_count, rna_count, oct_count
    """
    perfusion_labels = []
    rna_labels       = []
    perfusion_count  = 0
    rna_count        = 0
    oct_count        = 0

    for _, data_row in merged_df.iterrows():
        preservation = safe_get_label(data_row, 'Preservation', 'Preservation_sample', 'Preservation_animal')
        label_type, copies = determine_label_type(preservation)
        sample_name  = safe_get_label(data_row, 'Sample Name', 'Sample_Name')
        animal_name  = safe_get_label(data_row, 'Animal Name', 'Animal_Name')

        if label_type == 'skip':
            oct_count += 1
            continue

        if label_type == 'perfusion':
            perfusion_count += 1
            try:
                row1, row2, row3, row4 = format_label_rows(data_row, label_type)
            except Exception as e:
                print(f"    ✗ Error formatting perfusion '{sample_name}': {e}")
                traceback.print_exc()
                continue
            for _ in range(copies):
                perfusion_labels.append(
                    {'Row 1': row1, 'Row 2': row2, 'Row 3': row3, 'Row 4': row4}
                )

        else:  # rna
            rna_count += 1
            try:
                harvest_date   = safe_date_format(
                    safe_get_label(data_row, 'Sample Harvest Date', 'Harvest Date', 'Harvest_Date'))
                line_short     = safe_get_label(data_row, 'Line (Short)')
                sample_padded  = format_sample_number(sample_name, pad=True)
                sample_raw     = format_sample_number(sample_name, pad=False)
                animal_str     = str(animal_name).strip()

                rna_labels.append({
                    'Sides_Label_Num': rna_count,
                    'Sides_B':         f"{sample_padded}_{harvest_date}",
                    'Sides_C':         f"{animal_str}_{line_short}",
                    'Tops_Label_Num':  rna_count,
                    'Tops_B':          sample_raw,
                    'Tops_C':          animal_str,
                })
            except Exception as e:
                print(f"    ✗ Error formatting RNA '{sample_name}': {e}")
                traceback.print_exc()
                continue

    print(f"    Perfusion: {perfusion_count} × 2 = {perfusion_count * 2} labels")
    print(f"    RNA:       {rna_count} × 1 = {rna_count} labels")
    if oct_count > 0:
        print(f"    OCT Block: {oct_count} × 0 = skipped")
    print(f"    Total perfusion labels: {len(perfusion_labels)}")
    print(f"    Total RNA labels:       {rna_count}")

    return perfusion_labels, rna_labels, perfusion_count, rna_count, oct_count


def create_label_sheets(all_labels, output_folder, timestamp):
    """Create Excel label sheets — uses GUI dialogs for per-sheet offset input."""
    return _create_label_sheets_gui(all_labels, output_folder, timestamp)

def run_labels(samples_df, working_df, timestamp):
    """STEP 4: Generate label files."""
    print("\n" + "=" * 80)
    print("STEP 4: LABELS")
    print("=" * 80)

    if samples_df.empty:
        print("  ✗ No sample data. Skipping.")
        return None
    if working_df.empty:
        print("  ✗ No animal data. Skipping.")
        return None

    # Debug input
    print(f"\n  DEBUG Labels input:")
    print(f"    samples_df columns: {list(samples_df.columns)}")
    print(f"    samples_df rows: {len(samples_df)}")
    if len(samples_df) > 0:
        print(f"    samples_df first row: {samples_df.iloc[0].to_dict()}")
    print(f"    working_df columns (first 10): {list(working_df.columns)[:10]}")
    print(f"    working_df rows: {len(working_df)}")

    # Prepare samples — rename for merge
    s_df = samples_df.copy()
    s_rename = {}
    if 'Name' in s_df.columns:
        s_rename['Name'] = 'Sample Name'
    if 'Source' in s_df.columns:
        s_rename['Source'] = 'Animal Name'
    if 'Harvest Date' in s_df.columns:
        s_rename['Harvest Date'] = 'Sample Harvest Date'
    s_df = s_df.rename(columns=s_rename)

    print(f"    After rename - s_df columns: {list(s_df.columns)}")

    # Prepare animals — rename Animal_Name for merge
    a_df = working_df.copy()
    a_df = a_df.rename(columns={'Animal_Name': 'Animal Name'})

    print(f"    After rename - a_df 'Animal Name' present: {'Animal Name' in a_df.columns}")

    if 'Animal Name' not in s_df.columns:
        print("  ✗ 'Animal Name' not found in samples after rename")
        print(f"    s_df columns: {list(s_df.columns)}")
        return None
    if 'Animal Name' not in a_df.columns:
        print("  ✗ 'Animal Name' not found in animal data after rename")
        print(f"    a_df columns: {list(a_df.columns)}")
        return None

    # Debug merge values
    s_names = set(s_df['Animal Name'].astype(str).str.strip().unique())
    a_names = set(a_df['Animal Name'].astype(str).str.strip().unique())
    common = s_names.intersection(a_names)
    print(f"    Sample Animal Names (first 3): {list(s_names)[:3]}")
    print(f"    Working Animal Names (first 3): {list(a_names)[:3]}")
    print(f"    Common names: {len(common)}")

    if len(common) == 0:
        print("  ✗ NO MATCHING NAMES between samples and animals!")
        print(f"    This means the merge will produce 0 rows.")
        print(f"    Sample names sample: {list(s_names)[:5]}")
        print(f"    Animal names sample: {list(a_names)[:5]}")
        return None

    # Ensure string types match
    s_df['Animal Name'] = s_df['Animal Name'].astype(str).str.strip()
    a_df['Animal Name'] = a_df['Animal Name'].astype(str).str.strip()

    merged_df = pd.merge(s_df, a_df, on='Animal Name', how='inner',
                         suffixes=('_sample', '_animal'))

    unmatched = len(s_df) - len(merged_df)
    if unmatched > 0:
        print(f"  ⚠ {unmatched} samples did not match")
    print(f"  Matched {len(merged_df)} samples with animal data")
    print(f"  Merged columns: {list(merged_df.columns)}")

    if len(merged_df) == 0:
        print("  ✗ No matches — cannot create labels.")
        return None

    print("\n  Generating labels...")
    perfusion_labels, rna_labels, perf_count, rna_count, oct_count = generate_all_labels(merged_df)

    if not perfusion_labels and not rna_labels:
        if oct_count > 0:
            print("  ⚠ All samples are OCT Block — no labels needed.")
        else:
            print("  ✗ No labels generated.")
        return None

    script_dir = os.path.dirname(os.path.abspath(__file__))
    created_files = []

    # --- RNA Tube Labeler ---
    if rna_labels:
        rna_file = create_rna_excel(rna_labels, script_dir, timestamp)
        if rna_file:
            created_files.append(rna_file)

    # --- Perfusion Mail-Merge sheets ---
    if perfusion_labels:
        num_sheets, perf_files = create_label_sheets(perfusion_labels, script_dir, timestamp)
        created_files.extend(perf_files)

    total = len(created_files)
    if total > 0:
        print(f"\n  ✓ Step 4 complete: {total} label file(s)")
        if oct_count > 0:
            print(f"    Note: {oct_count} OCT Block sample(s) skipped")

    return created_files if created_files else None


# ============================================================


# ============================================================================
# UNIFIED MAIN — Scheduler → Harvest Pipeline in one shot
# ============================================================================


# ============================================================================
# PIPELINE GUI LAUNCHER
# ============================================================================

# ============================================================================
# PIPELINE GUI LAUNCHER  (replaces all terminal interaction)
# ============================================================================
import queue      as _queue
import threading  as _threading
import tkinter    as tk
from tkinter import ttk, filedialog, messagebox, scrolledtext
import os as _os

# ---------------------------------------------------------------------------
# Thread ↔ GUI messaging
# ---------------------------------------------------------------------------
_MSG_LOG      = 'log'        # pipeline → GUI: a line of text
_MSG_REQUEST  = 'request'    # pipeline → GUI: needs input
_MSG_DONE     = 'done'       # pipeline → GUI: finished (ok or error)

_pipeline_queue  = _queue.Queue()   # pipeline → GUI
_response_queue  = _queue.Queue()   # GUI → pipeline (responses to requests)


def _gui_ask(request_type, **kwargs):
    """
    Called from the pipeline thread to pause and request GUI input.
    Blocks the pipeline thread until the GUI responds.
    """
    _pipeline_queue.put({'kind': _MSG_REQUEST, 'type': request_type, **kwargs})
    return _response_queue.get()   # blocks until GUI calls _gui_respond()


def _gui_respond(value):
    _response_queue.put(value)


# ---------------------------------------------------------------------------
# Redirect stdout → GUI log (so all print() calls appear in the log widget)
# ---------------------------------------------------------------------------
class _QueueWriter:
    def __init__(self):
        self.encoding = 'utf-8'
    def write(self, text):
        if text and text != '\n':
            for line in text.splitlines():
                if line.strip():
                    _pipeline_queue.put({'kind': _MSG_LOG, 'text': line})
    def flush(self):
        pass


# ---------------------------------------------------------------------------
# Replacements for terminal input() calls inside the pipeline
# ---------------------------------------------------------------------------
def get_starting_sample_number():
    """GUI version — pauses pipeline, shows dialog, returns int."""
    return _gui_ask('sample_number')


def _create_label_sheets_gui(all_labels, output_folder, timestamp):
    """GUI version of create_label_sheets — asks offsets via dialog."""
    if not all_labels:
        _pipeline_queue.put({'kind': _MSG_LOG, 'text': '    ✗ No labels to create!'})
        return 0, []

    created_files      = []
    current_label_index = 0
    sheet_num           = 1

    while current_label_index < len(all_labels):
        labels_remaining = len(all_labels) - current_label_index
        _pipeline_queue.put({'kind': _MSG_LOG,
                              'text': f'    📄 Label sheet {sheet_num}  ({labels_remaining} labels remaining)'})

        labels_used = _gui_ask('label_offset',
                                sheet_num=sheet_num,
                                labels_remaining=labels_remaining)

        sheet_labels = [{'Row 1': '', 'Row 2': '', 'Row 3': '', 'Row 4': ''}
                        for _ in range(labels_used)]

        labels_placed = 0
        ci = current_label_index
        while len(sheet_labels) < LABELS_PER_PAGE and ci < len(all_labels):
            sheet_labels.append(all_labels[ci])
            ci += 1
            labels_placed += 1

        import pandas as _pd
        df = _pd.DataFrame(sheet_labels)
        output_file = _os.path.join(output_folder,
                                    f'Labels_Mailmerge_{timestamp}_sheet{sheet_num}.xlsx')
        save_df_to_excel(df, output_file, sheet_name='Labels')

        _pipeline_queue.put({'kind': _MSG_LOG,
                              'text': f'    📄 Saved: {_os.path.basename(output_file)}  '
                                      f'(empty={labels_used}, placed={labels_placed})'})

        created_files.append(output_file)
        current_label_index += labels_placed
        sheet_num += 1

        if current_label_index < len(all_labels):
            remaining = len(all_labels) - current_label_index
            _pipeline_queue.put({'kind': _MSG_LOG,
                                  'text': f'    ⚠  {remaining} labels still to place.'})
            _gui_ask('label_continue', sheet_num=sheet_num)

    return len(created_files), created_files


# ---------------------------------------------------------------------------
# Wednesday capacity GUI  (replaces prompt_wednesday_capacity)
# ---------------------------------------------------------------------------
def prompt_wednesday_capacity_gui(parent=None):
    """
    Show a standalone window for Wednesday capacity entry.
    Returns (wednesdays, full_dates_or_None) — same contract as the original.
    """
    wednesdays = get_next_wednesdays(6)
    capacity   = CONFIG['WEDNESDAY_CAPACITY']
    result     = {'value': None}

    win = tk.Toplevel(parent) if parent else tk.Tk()
    win.title('Wednesday Capacity')
    win.configure(bg='#f0f0f0')
    win.resizable(False, False)
    win.grab_set()

    # Header
    hdr = tk.Frame(win, bg='#2c3e50', pady=10)
    hdr.pack(fill='x')
    tk.Label(hdr, text='Wednesday Behavior Capacity',
             font=('Helvetica', 14, 'bold'),
             bg='#2c3e50', fg='white').pack()
    tk.Label(hdr,
             text=f'Maximum capacity: {capacity} animals per Wednesday\n'
                  'Enter how many slots are already booked for each date.',
             font=('Helvetica', 9), bg='#2c3e50', fg='#bdc3c7').pack(pady=(2, 6))

    body = tk.Frame(win, bg='#f0f0f0', padx=20, pady=14)
    body.pack()

    tk.Label(body, text='Wednesday', width=22, anchor='w',
             font=('Helvetica', 10, 'bold'), bg='#f0f0f0').grid(row=0, column=0, padx=(0, 8))
    tk.Label(body, text='Already Booked', width=14, anchor='center',
             font=('Helvetica', 10, 'bold'), bg='#f0f0f0').grid(row=0, column=1)
    tk.Label(body, text='Remaining', width=10, anchor='center',
             font=('Helvetica', 10, 'bold'), bg='#f0f0f0').grid(row=0, column=2)
    tk.Label(body, text='Status', width=12, anchor='center',
             font=('Helvetica', 10, 'bold'), bg='#f0f0f0').grid(row=0, column=3)

    entries     = {}
    status_vars = {}
    remain_vars = {}

    def _update_row(wed, var, status_lbl, remain_lbl):
        try:
            booked = int(var.get()) if var.get().strip() else 0
            booked = max(0, booked)
        except ValueError:
            booked = 0
        rem = capacity - booked
        remain_lbl.configure(text=str(rem))
        if rem <= 0:
            status_lbl.configure(text='🔴 FULL',  fg='#c0392b')
        elif rem <= 3:
            status_lbl.configure(text='🟡 LOW',   fg='#e67e22')
        else:
            status_lbl.configure(text='🟢 Open',  fg='#27ae60')

    for i, wed in enumerate(wednesdays, 1):
        r = i
        label = wed.strftime('%A, %Y-%m-%d')
        bg = '#ffffff' if i % 2 == 0 else '#f7f7f7'

        tk.Label(body, text=label, width=22, anchor='w',
                 font=('Helvetica', 9), bg='#f0f0f0').grid(row=r, column=0, pady=4, padx=(0, 8))

        var = tk.StringVar(value='0')
        e   = ttk.Spinbox(body, from_=0, to=capacity, textvariable=var, width=6)
        e.grid(row=r, column=1, pady=4)
        entries[wed] = var

        remain_lbl = tk.Label(body, text=str(capacity), width=10, anchor='center',
                              font=('Helvetica', 9), bg='#f0f0f0', fg='#27ae60')
        remain_lbl.grid(row=r, column=2)

        status_lbl = tk.Label(body, text='🟢 Open', width=12, anchor='center',
                              font=('Helvetica', 9), bg='#f0f0f0', fg='#27ae60')
        status_lbl.grid(row=r, column=3)

        var.trace_add('write', lambda *_, w=wed, v=var, sl=status_lbl, rl=remain_lbl:
                      _update_row(w, v, sl, rl))

    def _confirm():
        full_dates = []
        for wed in wednesdays:
            try:
                booked = int(entries[wed].get()) if entries[wed].get().strip() else 0
            except ValueError:
                booked = 0
            if capacity - booked <= 0:
                full_dates.append(wed)
        result['value'] = (wednesdays, full_dates if full_dates else None)
        win.destroy()

    foot = tk.Frame(win, bg='#ecf0f1', pady=8)
    foot.pack(fill='x', padx=20)
    tk.Button(foot, text='Continue  →', command=_confirm,
              font=('Helvetica', 11, 'bold'), bg='#27ae60', fg='white',
              relief='flat', padx=16, pady=6, cursor='hand2').pack(side='right')

    # center on parent
    win.update_idletasks()
    pw = win.winfo_screenwidth();  ph = win.winfo_screenheight()
    ww = win.winfo_width();        wh = win.winfo_height()
    win.geometry(f'+{(pw-ww)//2}+{(ph-wh)//2}')

    if parent:
        parent.wait_window(win)
    else:
        win.mainloop()

    if result['value'] is None:
        # Window closed without confirming — use empty
        return wednesdays, None
    return result['value']


# ---------------------------------------------------------------------------
# Main GUI launcher
# ---------------------------------------------------------------------------
def run_pipeline_gui():
    """Entry point — shows the full GUI pipeline."""

    root = tk.Tk()
    root.title('Sing Lab Scheduler')
    root.configure(bg='#f0f0f0')
    root.resizable(True, True)

    script_dir = _os.path.dirname(_os.path.abspath(__file__))

    # ── Shared state ─────────────────────────────────────────────────────────
    state = {
        'animal_file':   _os.path.join(script_dir, CONFIG['INPUT_ANIMAL_FILE']),
        'tracking_file': _os.path.join(script_dir, CONFIG['INPUT_TRACKING_FILE']),
        'births_file':   _os.path.join(script_dir, CONFIG['INPUT_BIRTHS_FILE']),
    }

    REQUIRED_COLOR = '#c0392b'
    OPTIONAL_COLOR = '#7f8c8d'
    OK_COLOR       = '#27ae60'

    # ── Helper: clear root and show a new screen ──────────────────────────────
    def _switch(frame_fn):
        for w in root.winfo_children():
            w.destroy()
        frame_fn()

    # ─────────────────────────────────────────────────────────────────────────
    # SCREEN 1: File Setup
    # ─────────────────────────────────────────────────────────────────────────
    def screen_file_setup():
        root.title('Sing Lab Scheduler')
        root.geometry('700x560')

        # ── Header ────────────────────────────────────────────────────────────
        hdr = tk.Frame(root, bg='#2c3e50', pady=18)
        hdr.pack(fill='x')
        tk.Label(hdr, text='Sing Lab Scheduler',
                 font=('Helvetica', 20, 'bold'),
                 bg='#2c3e50', fg='white').pack()
        tk.Label(hdr, text='Which input files do you have ready?',
                 font=('Helvetica', 10), bg='#2c3e50', fg='#bdc3c7').pack(pady=(4, 0))

        body = tk.Frame(root, bg='#f0f0f0', padx=24, pady=18)
        body.pack(fill='both', expand=True)

        # ── File card definitions ─────────────────────────────────────────────
        FILE_DEFS = [
            {
                'key':      'animal_file',
                'default':  CONFIG['INPUT_ANIMAL_FILE'],
                'required': True,
                'label':    'Animal Inventory',
                'hint':     f'Required  •  usually "{CONFIG['INPUT_ANIMAL_FILE']}"',
                'desc':     'The main list of all animals currently in the colony.',
            },
            {
                'key':      'tracking_file',
                'default':  CONFIG['INPUT_TRACKING_FILE'],
                'required': False,
                'label':    'Harvest Tracking Sheet',
                'hint':     'Optional  •  check the box to include in this run',
                'desc':     'Tracks how many of each strain/type have already been harvested.',
            },
            {
                'key':      'births_file',
                'default':  CONFIG['INPUT_BIRTHS_FILE'],
                'required': False,
                'label':    'Births Record',
                'hint':     'Optional  •  check the box to include in this run',
                'desc':     'Log of recent births used to identify new P14 animals.',
            },
        ]

        path_vars   = {}
        status_lbls = {}
        toggle_vars = {}   # BooleanVar: True = file is available
        detail_frames = {}

        err_lbl = tk.Label(body, text='', font=('Helvetica', 9, 'italic'),
                           bg='#f0f0f0', fg='#c0392b')

        def _update_status(key, path, lbl):
            if not path.strip():
                lbl.configure(text='', fg='#7f8c8d')
            elif _os.path.exists(path):
                lbl.configure(text='✓ Found', fg='#27ae60')
            else:
                lbl.configure(text='✗ File not found at this path', fg='#c0392b')

        def _browse(key, var, lbl, title):
            path = filedialog.askopenfilename(
                parent=root, title=title,
                initialdir=_os.path.dirname(var.get()) or script_dir,
                filetypes=[('CSV files', '*.csv'), ('All files', '*.*')]
            )
            if path:
                var.set(path)
                _update_status(key, path, lbl)

        def _toggle_card(key, toggle_var, detail_frame):
            # Detail frame is always visible — toggle only controls whether the
            # file is included in the run (handled in _proceed).
            pass

        for fd in FILE_DEFS:
            key      = fd['key']
            required = fd['required']
            default  = _os.path.join(script_dir, fd['default'])

            # Auto-detect: pre-tick if file exists in script_dir
            exists = _os.path.exists(default)
            initial_path = default if exists else state.get(key, default)

            # ── Card frame ────────────────────────────────────────────────────
            card = tk.Frame(body, bg='#ffffff', relief='solid', bd=1, padx=12, pady=10)
            card.pack(fill='x', pady=6)

            # Top row: toggle + label
            top_row = tk.Frame(card, bg='#ffffff')
            top_row.pack(fill='x')

            tvar = tk.BooleanVar(value=exists or required)
            toggle_vars[key] = tvar

            # Checkbox (disabled for required file)
            chk = tk.Checkbutton(
                top_row, variable=tvar,
                bg='#ffffff', activebackground='#ffffff',
                cursor='hand2' if not required else 'arrow',
                state='normal' if not required else 'disabled',
            )
            chk.pack(side='left', padx=(0, 6))

            name_color = '#2c3e50' if required else '#34495e'
            tk.Label(top_row, text=fd['label'],
                     font=('Helvetica', 11, 'bold'),
                     bg='#ffffff', fg=name_color).pack(side='left')

            badge_text  = '  Required  ' if required else '  Optional  '
            badge_color = '#e74c3c'       if required else '#95a5a6'
            tk.Label(top_row, text=badge_text,
                     font=('Helvetica', 8, 'bold'),
                     bg=badge_color, fg='white', padx=4).pack(side='left', padx=8)

            tk.Label(card, text=fd['desc'],
                     font=('Helvetica', 9), bg='#ffffff', fg='#7f8c8d',
                     anchor='w').pack(fill='x')

            # Expandable detail section (path + browse)
            detail = tk.Frame(card, bg='#f8f9fa', padx=8, pady=6, relief='groove', bd=1)
            detail_frames[key] = detail

            path_row = tk.Frame(detail, bg='#f8f9fa')
            path_row.pack(fill='x')

            pvar = tk.StringVar(value=initial_path)
            path_vars[key] = pvar
            state[key] = initial_path

            tk.Label(path_row, text=fd['hint'],
                     font=('Helvetica', 8), bg='#f8f9fa', fg='#7f8c8d',
                     anchor='w').pack(fill='x', pady=(0, 4))

            entry_row = tk.Frame(detail, bg='#f8f9fa')
            entry_row.pack(fill='x')

            # Create status label first, then Browse (pack order = right-to-left)
            slbl = tk.Label(entry_row, text='', font=('Helvetica', 9),
                            bg='#f8f9fa', width=14, anchor='w')
            status_lbls[key] = slbl

            tk.Button(entry_row, text='Browse…',
                      command=lambda k=key, v=pvar, l=slbl, t=fd['label']:
                          _browse(k, v, l, f'Select {t}'),
                      font=('Helvetica', 9), bg='#3498db', fg='white',
                      relief='flat', padx=8, pady=2, cursor='hand2').pack(side='right', padx=(4, 0))

            slbl.pack(side='right', padx=(6, 0))

            entry = tk.Entry(entry_row, textvariable=pvar, font=('Helvetica', 9))
            entry.pack(side='left', fill='x', expand=True)

            pvar.trace_add('write', lambda *_, k=key, v=pvar, l=slbl:
                           _update_status(k, v.get(), l))
            _update_status(key, initial_path, slbl)

            # Wire toggle
            tvar.trace_add('write', lambda *_, k=key, tv=tvar, df=detail:
                           _toggle_card(k, tv, df))

            # Always show the path row so the user can always browse
            detail.pack(fill='x', pady=(6, 0))

        err_lbl.pack(fill='x', pady=(4, 0))

        # ── Footer ────────────────────────────────────────────────────────────
        foot = tk.Frame(root, bg='#ecf0f1', pady=10)
        foot.pack(fill='x', padx=24)

        def _proceed():
            err_lbl.configure(text='')

            # Required file
            animal = path_vars['animal_file'].get().strip()
            if not toggle_vars['animal_file'].get() or not animal:
                err_lbl.configure(text='⚠  The Animal Inventory file is required to continue.')
                return
            if not _os.path.exists(animal):
                err_lbl.configure(text=f'⚠  Animal Inventory not found: {animal}')
                return

            try:
                import pandas as _pd
                test_df = _pd.read_csv(animal, nrows=2)
                missing = [c for c in CONFIG.get('REQUIRED_ANIMAL_COLUMNS', [])
                           if c not in test_df.columns]
                if missing:
                    err_lbl.configure(text=f'⚠  Animal file missing columns: {missing}')
                    return
            except Exception as ex:
                err_lbl.configure(text=f'⚠  Cannot read Animal Inventory: {ex}')
                return

            state['animal_file'] = animal

            for key in ('tracking_file', 'births_file'):
                if toggle_vars[key].get():
                    p = path_vars[key].get().strip()
                    state[key] = p if _os.path.exists(p) else None
                    if toggle_vars[key].get() and not _os.path.exists(p):
                        err_lbl.configure(
                            text=f'⚠  You indicated the file is available but it was not found:\n{p}\n'
                                 f'Please browse to it or uncheck the box.'
                        )
                        return
                else:
                    state[key] = None

            _switch(screen_wednesday)

        tk.Button(foot, text='Next: Wednesday Capacity  →',
                  command=_proceed,
                  font=('Helvetica', 11, 'bold'), bg='#27ae60', fg='white',
                  relief='flat', padx=16, pady=7, cursor='hand2').pack(side='right')

        # Fit window height to content
        def _fit_window():
            root.update_idletasks()
            w = root.winfo_width()
            h = root.winfo_reqheight()
            screen_h = root.winfo_screenheight()
            h = min(h + 20, screen_h - 80)
            x = (root.winfo_screenwidth()  - w) // 2
            y = (root.winfo_screenheight() - h) // 2
            root.geometry(f'{w}x{h}+{x}+{y}')
        root.after(10, _fit_window)


    # ─────────────────────────────────────────────────────────────────────────
    # SCREEN 2: Wednesday Capacity
    # ─────────────────────────────────────────────────────────────────────────
    def screen_wednesday():
        root.title('Sing Lab Scheduler — Wednesday Capacity')
        root.geometry('620x460')

        wednesdays = get_next_wednesdays(6)
        capacity   = CONFIG['WEDNESDAY_CAPACITY']

        hdr = tk.Frame(root, bg='#2c3e50', pady=10)
        hdr.pack(fill='x')
        tk.Label(hdr, text='Wednesday Behavior Capacity',
                 font=('Helvetica', 14, 'bold'),
                 bg='#2c3e50', fg='white').pack()
        tk.Label(hdr,
                 text=f'Max {capacity} animals per Wednesday.\n'
                      'Enter how many slots are already booked.',
                 font=('Helvetica', 9), bg='#2c3e50', fg='#bdc3c7').pack(pady=(2, 4))

        body = tk.Frame(root, bg='#f0f0f0', padx=24, pady=14)
        body.pack(fill='both', expand=True)

        # Column headers
        for col, (txt, w) in enumerate([('Wednesday', 24), ('Already Booked', 15),
                                         ('Remaining', 12), ('Status', 14)]):
            tk.Label(body, text=txt, width=w, anchor='w' if col == 0 else 'center',
                     font=('Helvetica', 10, 'bold'), bg='#f0f0f0'
                     ).grid(row=0, column=col, pady=(0, 6))

        entries = {}

        def _update_row(wed, var, rl, sl):
            try:
                booked = int(var.get()) if var.get().strip() else 0
                booked = max(0, booked)
            except ValueError:
                booked = 0
            rem = capacity - booked
            rl.configure(text=str(rem))
            if rem <= 0:
                sl.configure(text='🔴 FULL',  fg='#c0392b')
            elif rem <= 3:
                sl.configure(text='🟡 LOW',   fg='#e67e22')
            else:
                sl.configure(text='🟢 Open',  fg='#27ae60')

        for i, wed in enumerate(wednesdays, 1):
            label = wed.strftime('%A, %Y-%m-%d')
            tk.Label(body, text=label, width=24, anchor='w',
                     font=('Helvetica', 9), bg='#f0f0f0'
                     ).grid(row=i, column=0, pady=5)

            var = tk.StringVar(value='0')
            ttk.Spinbox(body, from_=0, to=capacity, textvariable=var, width=7
                        ).grid(row=i, column=1, pady=5)
            entries[wed] = var

            rl = tk.Label(body, text=str(capacity), width=12, anchor='center',
                          font=('Helvetica', 9), bg='#f0f0f0', fg='#27ae60')
            rl.grid(row=i, column=2)

            sl = tk.Label(body, text='🟢 Open', width=14, anchor='center',
                          font=('Helvetica', 9), bg='#f0f0f0', fg='#27ae60')
            sl.grid(row=i, column=3)

            var.trace_add('write', lambda *_, w=wed, v=var, r=rl, s=sl:
                          _update_row(w, v, r, s))

        def _proceed():
            full_dates = []
            for wed in wednesdays:
                try:
                    booked = int(entries[wed].get()) if entries[wed].get().strip() else 0
                except ValueError:
                    booked = 0
                if capacity - booked <= 0:
                    full_dates.append(wed)
            state['wednesday_dates']     = wednesdays
            state['full_behavior_dates'] = full_dates if full_dates else None
            _switch(screen_progress)

        foot = tk.Frame(root, bg='#ecf0f1', pady=10)
        foot.pack(fill='x', padx=24)
        tk.Button(foot, text='← Back',
                  command=lambda: _switch(screen_file_setup),
                  font=('Helvetica', 10), bg='#95a5a6', fg='white',
                  relief='flat', padx=12, pady=6, cursor='hand2').pack(side='left')
        tk.Button(foot, text='Run Pipeline  →',
                  command=_proceed,
                  font=('Helvetica', 11, 'bold'), bg='#27ae60', fg='white',
                  relief='flat', padx=16, pady=7, cursor='hand2').pack(side='right')

    # ─────────────────────────────────────────────────────────────────────────
    # SCREEN 3: Progress + mid-run dialogs
    # ─────────────────────────────────────────────────────────────────────────
    def screen_progress():
        root.title('Sing Lab Scheduler — Running…')
        root.geometry('760x560')

        hdr = tk.Frame(root, bg='#2c3e50', pady=10)
        hdr.pack(fill='x')
        tk.Label(hdr, text='Pipeline Running',
                 font=('Helvetica', 14, 'bold'),
                 bg='#2c3e50', fg='white').pack()
        status_var = tk.StringVar(value='Starting up…')
        tk.Label(hdr, textvariable=status_var,
                 font=('Helvetica', 9), bg='#2c3e50', fg='#bdc3c7').pack()

        log_widget = scrolledtext.ScrolledText(
            root, font=('Courier', 9), bg='#1e1e1e', fg='#d4d4d4',
            insertbackground='white', wrap='word', state='disabled'
        )
        log_widget.pack(fill='both', expand=True, padx=12, pady=(8, 4))

        foot = tk.Frame(root, bg='#ecf0f1', pady=8)
        foot.pack(fill='x', padx=12)

        def _append_log(text):
            log_widget.configure(state='normal')
            log_widget.insert('end', text + '\n')
            log_widget.see('end')
            log_widget.configure(state='disabled')

        # ── Mid-run dialog: sample number ─────────────────────────────────────
        def _ask_sample_number():
            dlg = tk.Toplevel(root)
            dlg.title('Starting Sample Number')
            dlg.configure(bg='#f0f0f0')
            dlg.grab_set()
            dlg.resizable(False, False)

            tk.Label(dlg, text='Sample Number Setup',
                     font=('Helvetica', 12, 'bold'),
                     bg='#f0f0f0').pack(pady=(16, 4), padx=20)
            tk.Label(dlg,
                     text='Enter the last sample number used.\nThe next run will start from that number + 1.',
                     font=('Helvetica', 9), bg='#f0f0f0', justify='center').pack(padx=20)

            frame = tk.Frame(dlg, bg='#f0f0f0')
            frame.pack(pady=12, padx=20)
            tk.Label(frame, text='Last sample number used:',
                     font=('Helvetica', 9), bg='#f0f0f0').grid(row=0, column=0, padx=(0, 8))
            var = tk.StringVar()
            e = ttk.Entry(frame, textvariable=var, width=10)
            e.grid(row=0, column=1)
            e.focus()

            preview = tk.Label(dlg, text='', font=('Helvetica', 9, 'italic'),
                               bg='#f0f0f0', fg='#27ae60')
            preview.pack()
            err_lbl = tk.Label(dlg, text='', font=('Helvetica', 9),
                                bg='#f0f0f0', fg='#c0392b')
            err_lbl.pack()

            def _update_preview(*_):
                try:
                    nxt = int(var.get()) + 1
                    preview.configure(text=f'Next sample will start at: {nxt}')
                    err_lbl.configure(text='')
                except ValueError:
                    preview.configure(text='')

            var.trace_add('write', _update_preview)

            def _ok():
                try:
                    nxt = int(var.get()) + 1
                    _gui_respond(nxt)
                    dlg.destroy()
                except ValueError:
                    err_lbl.configure(text='Please enter a valid whole number.')

            tk.Button(dlg, text='Confirm', command=_ok,
                      font=('Helvetica', 10, 'bold'), bg='#27ae60', fg='white',
                      relief='flat', padx=14, pady=5, cursor='hand2').pack(pady=(6, 16))
            dlg.bind('<Return>', lambda e: _ok())

            dlg.update_idletasks()
            x = root.winfo_x() + (root.winfo_width()  - dlg.winfo_width())  // 2
            y = root.winfo_y() + (root.winfo_height() - dlg.winfo_height()) // 2
            dlg.geometry(f'+{x}+{y}')
            root.wait_window(dlg)

        # ── Mid-run dialog: label offset ──────────────────────────────────────
        def _ask_label_offset(sheet_num, labels_remaining):
            dlg = tk.Toplevel(root)
            dlg.title(f'Label Sheet {sheet_num}')
            dlg.configure(bg='#f0f0f0')
            dlg.grab_set()
            dlg.resizable(False, False)

            tk.Label(dlg, text=f'Label Sheet {sheet_num}',
                     font=('Helvetica', 12, 'bold'),
                     bg='#f0f0f0').pack(pady=(16, 4), padx=20)
            tk.Label(dlg,
                     text=f'{labels_remaining} labels remaining to place.\n'
                          f'How many label slots are already used on this sheet?\n'
                          f'(Enter 0 if the sheet is blank.)',
                     font=('Helvetica', 9), bg='#f0f0f0', justify='center').pack(padx=20)

            frame = tk.Frame(dlg, bg='#f0f0f0')
            frame.pack(pady=12, padx=20)
            tk.Label(frame, text='Labels already used:',
                     font=('Helvetica', 9), bg='#f0f0f0').grid(row=0, column=0, padx=(0, 8))
            var = tk.StringVar(value='0')
            ttk.Spinbox(frame, from_=0, to=LABELS_PER_PAGE - 1,
                        textvariable=var, width=6).grid(row=0, column=1)

            err_lbl = tk.Label(dlg, text='', font=('Helvetica', 9),
                                bg='#f0f0f0', fg='#c0392b')
            err_lbl.pack()

            def _ok():
                try:
                    n = int(var.get())
                    if 0 <= n < LABELS_PER_PAGE:
                        _gui_respond(n)
                        dlg.destroy()
                    else:
                        err_lbl.configure(text=f'Enter 0–{LABELS_PER_PAGE - 1}')
                except ValueError:
                    err_lbl.configure(text='Please enter a valid number.')

            tk.Button(dlg, text='Confirm', command=_ok,
                      font=('Helvetica', 10, 'bold'), bg='#27ae60', fg='white',
                      relief='flat', padx=14, pady=5, cursor='hand2').pack(pady=(6, 16))
            dlg.bind('<Return>', lambda e: _ok())

            dlg.update_idletasks()
            x = root.winfo_x() + (root.winfo_width()  - dlg.winfo_width())  // 2
            y = root.winfo_y() + (root.winfo_height() - dlg.winfo_height()) // 2
            dlg.geometry(f'+{x}+{y}')
            root.wait_window(dlg)

        # ── Mid-run dialog: label continue ────────────────────────────────────
        def _ask_label_continue(sheet_num):
            dlg = tk.Toplevel(root)
            dlg.title('Load Next Sheet')
            dlg.configure(bg='#f0f0f0')
            dlg.grab_set()
            dlg.resizable(False, False)

            tk.Label(dlg, text=f'Ready for Sheet {sheet_num}?',
                     font=('Helvetica', 12, 'bold'),
                     bg='#f0f0f0').pack(pady=(16, 4), padx=20)
            tk.Label(dlg,
                     text='Load the next label sheet into your printer, then click Continue.',
                     font=('Helvetica', 9), bg='#f0f0f0').pack(padx=20, pady=4)

            def _ok():
                _gui_respond(True)
                dlg.destroy()

            tk.Button(dlg, text='Continue  →', command=_ok,
                      font=('Helvetica', 10, 'bold'), bg='#27ae60', fg='white',
                      relief='flat', padx=14, pady=5, cursor='hand2').pack(pady=(12, 16))
            dlg.bind('<Return>', lambda e: _ok())

            dlg.update_idletasks()
            x = root.winfo_x() + (root.winfo_width()  - dlg.winfo_width())  // 2
            y = root.winfo_y() + (root.winfo_height() - dlg.winfo_height()) // 2
            dlg.geometry(f'+{x}+{y}')
            root.wait_window(dlg)

        # ── Poll queue from GUI thread ─────────────────────────────────────────
        pipeline_done = {'ok': False, 'error': None, 'result': None}

        def _poll():
            try:
                while True:
                    msg = _pipeline_queue.get_nowait()
                    if msg['kind'] == _MSG_LOG:
                        _append_log(msg['text'])
                    elif msg['kind'] == _MSG_DONE:
                        pipeline_done['ok']     = msg.get('ok', False)
                        pipeline_done['error']  = msg.get('error')
                        pipeline_done['result'] = msg.get('result')
                        if pipeline_done['ok']:
                            status_var.set('✓ Complete')
                            _switch(lambda: screen_summary(pipeline_done['result']))
                        else:
                            status_var.set('✗ Error — see log')
                            _append_log(f'\n✗ ERROR: {pipeline_done["error"]}')
                            _add_close_button()
                        return   # stop polling
                    elif msg['kind'] == _MSG_REQUEST:
                        rtype = msg['type']
                        if rtype == 'sample_number':
                            _ask_sample_number()
                        elif rtype == 'label_offset':
                            _ask_label_offset(msg['sheet_num'], msg['labels_remaining'])
                        elif rtype == 'label_continue':
                            _ask_label_continue(msg['sheet_num'])
            except _queue.Empty:
                pass
            root.after(120, _poll)

        def _add_close_button():
            tk.Button(foot, text='Close', command=root.destroy,
                      font=('Helvetica', 10), bg='#e74c3c', fg='white',
                      relief='flat', padx=12, pady=5).pack(side='right')

        # ── Pipeline thread ───────────────────────────────────────────────────
        def _run_pipeline():
            import sys as _sys
            old_stdout = _sys.stdout
            _sys.stdout = _QueueWriter()
            try:
                setup_logging(script_dir, CONFIG['LOG_LEVEL'])

                schedule_file, assignments_df = create_complete_schedule(
                    animal_file       = state['animal_file'],
                    tracking_file     = state.get('tracking_file'),
                    births_file       = state.get('births_file'),
                    output_dir        = script_dir,
                    birth_date_start  = None,
                    birth_date_end    = None,
                    behavior_date_start = None,
                    behavior_date_end   = None,
                    full_behavior_dates = state.get('full_behavior_dates'),
                )

                timestamp = __import__('datetime').datetime.now().strftime('%Y%m%d_%H%M%S')

                output_files = [schedule_file]

                if assignments_df is not None and len(assignments_df) > 0:
                    if 'Line' not in assignments_df.columns:
                        if 'Strain' in assignments_df.columns:
                            assignments_df['Line'] = assignments_df['Strain']
                        elif 'Line (Short)' in assignments_df.columns:
                            assignments_df['Line'] = assignments_df['Line (Short)']

                    working_df = build_working_data(assignments_df)

                    if not working_df.empty:
                        harvest_df, samples_df, climb_import_df = run_harvest_and_samples(
                            working_df, timestamp)
                        run_deliverables(working_df, samples_df, timestamp)
                        run_climb_to_envision(working_df, timestamp)
                        run_labels(samples_df, working_df, timestamp)

                        new_files = sorted([
                            _os.path.join(script_dir, f)
                            for f in _os.listdir(script_dir)
                            if timestamp in f
                        ])
                        output_files.extend(new_files)

                _pipeline_queue.put({
                    'kind': _MSG_DONE, 'ok': True,
                    'result': {'schedule_file': schedule_file, 'output_files': output_files}
                })

            except Exception as ex:
                import traceback as _tb
                _tb.print_exc()
                _pipeline_queue.put({'kind': _MSG_DONE, 'ok': False, 'error': str(ex)})
            finally:
                _sys.stdout = old_stdout

        _append_log('Starting pipeline…\n')
        status_var.set('Running — please wait…')
        t = _threading.Thread(target=_run_pipeline, daemon=True)
        t.start()
        root.after(120, _poll)

    # ─────────────────────────────────────────────────────────────────────────
    # SCREEN 4: Summary
    # ─────────────────────────────────────────────────────────────────────────
    def screen_summary(result):
        root.title('Sing Lab Scheduler — Complete')
        root.geometry('640x420')

        hdr = tk.Frame(root, bg='#27ae60', pady=14)
        hdr.pack(fill='x')
        tk.Label(hdr, text='✓  Pipeline Complete',
                 font=('Helvetica', 16, 'bold'),
                 bg='#27ae60', fg='white').pack()

        body = tk.Frame(root, bg='#f0f0f0', padx=24, pady=16)
        body.pack(fill='both', expand=True)

        tk.Label(body, text='Output files saved to:',
                 font=('Helvetica', 10, 'bold'),
                 bg='#f0f0f0', fg='#2c3e50').pack(anchor='w')
        tk.Label(body, text=script_dir,
                 font=('Helvetica', 9), bg='#f0f0f0', fg='#7f8c8d').pack(anchor='w', pady=(0, 12))

        files = result.get('output_files', [])
        for fpath in files:
            name = _os.path.basename(fpath)
            try:
                size = _os.path.getsize(fpath)
                size_str = f'{size:,} bytes'
            except Exception:
                size_str = ''
            row = tk.Frame(body, bg='#ffffff', padx=8, pady=4,
                           relief='solid', bd=1)
            row.pack(fill='x', pady=2)
            tk.Label(row, text=f'📄 {name}', font=('Helvetica', 9),
                     bg='#ffffff', fg='#2c3e50', anchor='w').pack(side='left')
            if size_str:
                tk.Label(row, text=size_str, font=('Helvetica', 8),
                         bg='#ffffff', fg='#95a5a6').pack(side='right')

        foot = tk.Frame(root, bg='#ecf0f1', pady=10)
        foot.pack(fill='x', padx=24)

        def _run_again():
            # Clear queues
            for q in (_pipeline_queue, _response_queue):
                while not q.empty():
                    try: q.get_nowait()
                    except: pass
            _switch(screen_file_setup)

        tk.Button(foot, text='Run Again',
                  command=_run_again,
                  font=('Helvetica', 10), bg='#3498db', fg='white',
                  relief='flat', padx=12, pady=6, cursor='hand2').pack(side='left')
        tk.Button(foot, text='Close',
                  command=root.destroy,
                  font=('Helvetica', 11, 'bold'), bg='#27ae60', fg='white',
                  relief='flat', padx=16, pady=6, cursor='hand2').pack(side='right')

    # ── Start on screen 1 ────────────────────────────────────────────────────
    w = min(root.winfo_screenwidth() - 100, 760)
    h = min(root.winfo_screenheight() - 100, 560)
    x = (root.winfo_screenwidth()  - w) // 2
    y = (root.winfo_screenheight() - h) // 2
    root.geometry(f'{w}x{h}+{x}+{y}')
    root.minsize(560, 380)

    screen_file_setup()
    root.mainloop()


if __name__ == "__main__":
    run_pipeline_gui()
