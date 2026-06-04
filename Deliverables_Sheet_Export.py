import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime
import os
import sys
import traceback
import re

def genotype_to_symbol(raw) -> str:
    """Convert any Climb genotype string to standard display symbol."""
    if raw is None or (isinstance(raw, float) and raw != raw):
        return 'Blank'
    s = str(raw).strip().lower()
    if not s or s in ('nan', 'none', 'n/a', '-', ''):
        return 'Blank'
    s = re.sub(r'[‹<][^›>]*[›>]', '', s)
    s = re.sub(r'\[[^\]]*\]', '', s)
    s = re.sub(r'probe\s*', '', s)
    s = ' '.join(s.split())
    if any(k in s for k in ('inconclusive', 'pending', 'failed', 'no call')):
        return 'Blank'
    if re.search(r'\bhom\d*\b|-/-', s):   return '-/-'
    if re.search(r'\bhet\d*\b|-/\+|\+/-', s): return '+/-'
    if re.search(r'hem[i]?|tg/\+|\+/tg|-/y', s): return '-/Y'
    if re.search(r'\+/\+|\bwt\b|wild.?type', s):  return '+/+'
    if 'inbred' in s:                      return 'Inbred'
    return 'Blank'

def combine_sample_numbers(sample_list):
    """
    Combine sample numbers into range format.
    Example: [571, 572, 573] -> "571-573"
    Example: [100] -> "100"
    """
    if not sample_list:
        return ""
    
    # Extract base numbers (remove any suffixes if present)
    base_numbers = []
    for sample in sample_list:
        sample_str = str(sample)
        # Remove suffix (everything after and including the dash)
        if '-' in sample_str:
            base_num = sample_str.split('-')[0]
        else:
            base_num = sample_str
        try:
            # Always convert to int for comparison
            base_numbers.append(int(base_num))
        except (ValueError, TypeError):
            # If conversion fails, skip this sample
            continue
    
    # If no valid numbers were found, return empty string
    if not base_numbers:
        return ""
    
    if len(base_numbers) == 1:
        return str(base_numbers[0])
    else:
        # Return as range
        first = min(base_numbers)
        last = max(base_numbers)
        return f"{first}-{last}"

class MultiSheetExporter:
    def __init__(self, animals_csv, samples_csv, output_filename=None):
        """
        Initialize the multi-sheet exporter
        
        Args:
            animals_csv: Path to animals CSV file
            samples_csv: Path to samples CSV file
            output_filename: Name of the output Excel file (if None, will auto-generate with timestamp)
        """
        # Generate filename with timestamp if not provided
        if output_filename is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_filename = f"lab_data_export_{timestamp}.xlsx"
        
        self.output_filename = output_filename
        self.workbook = Workbook()
        
        # Load CSV files
        print("Loading CSV files...")
        
        # Load animals
        animals_df = pd.DataFrame()
        if animals_csv and os.path.exists(animals_csv):
            try:
                # Load as standard CSV (comma-separated with quotes)
                animals_df = pd.read_csv(animals_csv)
                
                # Parse dates only if columns exist
                date_cols = ['Birth Date', 'Death/Exit Date', 'Wean Date', 'Arrival Date']
                for col in date_cols:
                    if col in animals_df.columns:
                        animals_df[col] = pd.to_datetime(animals_df[col], errors='coerce')
                
                # Check if file has any data (not just headers)
                if len(animals_df) > 0:
                    print(f"  ✓ Loaded {len(animals_df)} animals")
                else:
                    print(f"  ⚠ {animals_csv} exists but is empty (no data rows)")
                    animals_df = pd.DataFrame()
            except Exception as e:
                print(f"  ⚠ Warning loading {animals_csv}: {e}")
                traceback.print_exc()
        else:
            print(f"  ⚠ Animals file not found: {animals_csv}")
        
        self.animals_df = animals_df
        
        if self.animals_df.empty:
            print("\n  ⚠ No animal data available - will export samples without animal information")
        
        # Load samples (required)
        try:
            # Load as standard CSV (comma-separated with quotes)
            self.samples_df = pd.read_csv(samples_csv)
            
            # Parse dates only if columns exist
            date_cols = ['Harvest Date', 'Expiration Date']
            for col in date_cols:
                if col in self.samples_df.columns:
                    self.samples_df[col] = pd.to_datetime(self.samples_df[col], errors='coerce')
            
            if len(self.samples_df) == 0:
                print(f"  ⚠ Warning: {samples_csv} exists but is empty!")
            else:
                print(f"  ✓ Loaded {len(self.samples_df)} samples")
        except Exception as e:
            print(f"  ✗ Error loading samples.csv: {e}")
            raise
        
        # Rename key columns to match between files
        print("\nPreparing data for merge...")
        
        # Check if required columns exist in samples
        if 'Source' not in self.samples_df.columns:
            print(f"  ✗ Error: 'Source' column not found in samples.csv")
            print(f"  Available columns: {list(self.samples_df.columns)}")
            raise ValueError("Missing 'Source' column in samples.csv")
        
        # Rename columns in samples
        self.samples_df = self.samples_df.rename(columns={'Source': 'Animal_Name'})
        
        # Also rename 'Name' in samples to avoid confusion
        if 'Name' in self.samples_df.columns:
            self.samples_df = self.samples_df.rename(columns={'Name': 'Sample_Name'})
        
        # Convert sample Animal_Name to string
        self.samples_df['Animal_Name'] = self.samples_df['Animal_Name'].astype(str).str.strip()
        
        # Process animals if we have any
        if not self.animals_df.empty:
            # Check if Name column exists in animals
            if 'Name' not in self.animals_df.columns:
                print(f"  ✗ Error: 'Name' column not found in animals CSV file")
                print(f"  Available columns: {list(self.animals_df.columns)}")
                print(f"  ⚠ Continuing without animal data due to missing 'Name' column")
                self.animals_df = pd.DataFrame()  # Clear it out
            else:
                # Rename columns in animals
                self.animals_df = self.animals_df.rename(columns={'Name': 'Animal_Name'})
                
                # Convert animals Animal_Name to string
                self.animals_df['Animal_Name'] = self.animals_df['Animal_Name'].astype(str).str.strip()
                
                print(f"  ✓ Sample linking column: 'Animal_Name' (was 'Source')")
                print(f"  ✓ Animal linking column: 'Animal_Name' (was 'Name')")

                # Check if any match
                sample_set = set(self.samples_df['Animal_Name'].unique())
                animal_set = set(self.animals_df['Animal_Name'].unique())
                common = sample_set.intersection(animal_set)
                print(f"  {len(common)} of {len(sample_set)} samples matched to animals")
                if len(common) == 0:
                    print(f"  ⚠ No common Animal_Name values found — check source data")
        
        # Perform merge if we have animal data
        if not self.animals_df.empty:
            print("\n🔗 Merging samples with animals...")
            try:
                self.merged_df = pd.merge(
                    self.samples_df,
                    self.animals_df,
                    on='Animal_Name',
                    how='left',  # Keep all samples, even if no animal match
                    suffixes=('_sample', '_animal')
                )
                print(f"  ✓ Merge complete: {len(self.merged_df)} rows")
                
                # Check matching results
                animal_cols = [col for col in self.merged_df.columns if col.endswith('_animal')]
                if animal_cols:
                    matched = self.merged_df[animal_cols[0]].notna().sum()
                    total = len(self.merged_df)
                    print(f"  ✓ {matched}/{total} samples matched with animals")
                    
                    if matched < total:
                        unmatched_count = total - matched
                        print(f"  ⚠ Warning: {unmatched_count} samples did not match")
                
            except Exception as e:
                print(f"  ✗ Error during merge: {e}")
                traceback.print_exc()
                print(f"  Continuing with samples only...")
                self.merged_df = self.samples_df.copy()
        else:
            # No animal data - just use samples as-is
            print("\n⚠ No animal data to merge - using samples only")
            self.merged_df = self.samples_df.copy()
        
        
        # Remove default sheet
        if 'Sheet' in self.workbook.sheetnames:
            del self.workbook['Sheet']
    
    def _calculate_age_weeks(self, birth_date, harvest_date):
        """Calculate age in weeks from birth date to harvest date."""
        try:
            if pd.notna(birth_date) and pd.notna(harvest_date):
                days_diff = (pd.to_datetime(harvest_date) - pd.to_datetime(birth_date)).days
                return round(days_diff / 7, 1)
        except:
            pass
        return ''

    def _calculate_age_days_p(self, birth_date, harvest_date):
        """Calculate age in days from birth to harvest, returned as P-prefixed string (e.g. P60)."""
        try:
            if pd.notna(birth_date) and pd.notna(harvest_date):
                days = (pd.to_datetime(harvest_date) - pd.to_datetime(birth_date)).days
                return f'P{days}'
        except:
            pass
        return ''
    
    def _safe_get(self, row, *columns, default=''):
        """
        Safely get a value from a row, trying multiple column names
        """
        for col in columns:
            if col in row.index:
                val = row[col]
                if pd.notna(val):
                    # Don't return 'nan' string
                    if str(val).lower() != 'nan':
                        return val
        return default
    
    def create_sing_harvest_sheet(self):
        """
        Create the Sing Harvest Sheet by combining animals and samples data
        Groups samples by animal and combines sample numbers into ranges
        INCLUDES ALL SAMPLES (no filtering)
        """
        sheet_name = "Sing Harvest Sheet"
        ws = self.workbook.create_sheet(sheet_name)
        
        print("\nCreating Sing Harvest Sheet...")
        
        # Group samples by Animal_Name to combine sample numbers
        grouped_data = {}
        
        for idx, row in self.merged_df.iterrows():
            animal_name = self._safe_get(row, 'Animal_Name')
            sample_name = self._safe_get(row, 'Sample_Name', 'Name_sample')
            
            if not animal_name or animal_name == '':
                continue
            
            # Initialize group if first time seeing this animal
            if animal_name not in grouped_data:
                grouped_data[animal_name] = {
                    'samples': [],
                    'data': {
                        'Name': animal_name,
                        'Line': self._safe_get(row, 'Line_animal', 'Line_sample', 'Line'),
                        'BD': self._safe_get(row, 'Birth Date'),
                        'Housing': self._safe_get(row, 'Housing ID', 'Prior Housing ID'),
                        'Identification': self._safe_get(row, 'Marker'),
                        'Sex': self._safe_get(row, 'Sex'),
                        'Age (Days)': self._calculate_age_days_p(self._safe_get(row, 'Birth Date'), self._safe_get(row, 'Harvest Date'))
                    }
                }
            
            # Add sample to this animal's list
            if sample_name:
                grouped_data[animal_name]['samples'].append(sample_name)
        
        # Build the final data with combined sample numbers
        harvest_data = []
        
        for animal_name, group in grouped_data.items():
            combined_samples = combine_sample_numbers(group['samples'])
            row_data = group['data'].copy()
            row_data['Sample Number'] = combined_samples
            harvest_data.append(row_data)
        
        df = pd.DataFrame(harvest_data)
        
        if df.empty:
            df = pd.DataFrame(columns=[
                'Name', 'Sample Number', 'Line', 'BD', 'Housing', 
                'Identification', 'Sex', 'Age (Days)'
            ])
        
        column_order = [
            'Name', 'Sample Number', 'Line', 'BD', 'Housing', 
            'Identification', 'Sex', 'Age (Days)'
        ]
        df = df[column_order]
        
        self._apply_sheet_styling(ws, df, column_order)
        
        print(f"  ✓ Created Sing Harvest Sheet with {len(df)} rows (grouped by animal, all samples)")
        return ws
    
    def create_animal_sample_tracking_sheet(self):
        """
        Create the Animal and Sample Tracking sheet
        FILTERS: Only samples with Preservation = "4% PFA Fixed"
        """
        sheet_name = "Animal and Sample Tracking"
        ws = self.workbook.create_sheet(sheet_name)
        
        print("\nCreating Animal and Sample Tracking sheet...")
        
        # Filter for only "4% PFA Fixed" samples
        filtered_df = self.merged_df.copy()
        if 'Preservation' in filtered_df.columns:
            filtered_df = filtered_df[filtered_df['Preservation'] == '4% PFA Fixed']
            print(f"  Filtered to {len(filtered_df)} samples with Preservation = '4% PFA Fixed'")
        else:
            print(f"  ⚠ Warning: 'Preservation' column not found, including all samples")
        
        tracking_data = []
        
        for idx, row in filtered_df.iterrows():
            age_weeks = self._calculate_age_weeks(
                self._safe_get(row, 'Birth Date'), 
                self._safe_get(row, 'Harvest Date')
            )
            
            data = {
                'Name_sample': self._safe_get(row, 'Sample_Name', 'Name_sample'),
                'Harvest Date': self._safe_get(row, 'Harvest Date'),
                'Age (weeks)_sample': age_weeks,
                'Name_subject': self._safe_get(row, 'Animal_Name'),
                'Sex': self._safe_get(row, 'Sex'),
                'Line_subject': self._safe_get(row, 'Line_animal', 'Line_sample', 'Line'),
                'Line (Short)': self._safe_get(row, 'Line (Short)'),
                'Line (Stock)': self._safe_get(row, 'Line (Stock)'),
                'Species_subject': 'Mouse' if self._safe_get(row, 'Sex') != '' else '',
                'Genotype': genotype_to_symbol(self._safe_get(row, 'Genotype')),
                'Birth Date': self._safe_get(row, 'Birth Date'),
                'Wean Date': self._safe_get(row, 'Wean Date'),
                'Harvest Timepoint': self._safe_get(row, 'Time Point')
            }
            
            tracking_data.append(data)
        
        df = pd.DataFrame(tracking_data)
        
        if df.empty:
            df = pd.DataFrame(columns=[
                'Name_sample', 'Harvest Date', 'Age (weeks)_sample', 'Name_subject', 
                'Sex', 'Line_subject', 'Line (Short)', 'Line (Stock)', 
                'Species_subject', 'Genotype', 'Birth Date', 'Wean Date', 
                'Harvest Timepoint'
            ])
        
        column_order = [
            'Name_sample', 'Harvest Date', 'Age (weeks)_sample', 'Name_subject', 
            'Sex', 'Line_subject', 'Line (Short)', 'Line (Stock)', 
            'Species_subject', 'Genotype', 'Birth Date', 'Wean Date', 
            'Harvest Timepoint'
        ]
        df = df[column_order]
        
        self._apply_sheet_styling(ws, df, column_order)
        
        print(f"  ✓ Created Animal and Sample Tracking sheet with {len(df)} rows (4% PFA Fixed only)")
        return ws
    
    def create_merfish_sample_tracker_sheet(self):
        """
        Create the MERFISH Sample Tracker sheet
        FILTERS: Only samples with Preservation = "OCT Block"
        """
        sheet_name = "MERFISH Sample Tracker"  # Added space between Sample and Tracker
        ws = self.workbook.create_sheet(sheet_name)
        
        print("\nCreating MERFISH Sample Tracker sheet...")
        
        # Filter for only "OCT Block" samples
        filtered_df = self.merged_df.copy()
        if 'Preservation' in filtered_df.columns:
            filtered_df = filtered_df[filtered_df['Preservation'] == 'OCT Block']
            print(f"  Filtered to {len(filtered_df)} samples with Preservation = 'OCT Block'")
        else:
            print(f"  ⚠ Warning: 'Preservation' column not found, including all samples")
        
        tracker_data = []
        
        for idx, row in filtered_df.iterrows():
            age_weeks = self._calculate_age_weeks(
                self._safe_get(row, 'Birth Date'), 
                self._safe_get(row, 'Harvest Date')
            )
            
            data = {
                'Name_sample': self._safe_get(row, 'Sample_Name', 'Name_sample'),
                'Age (weeks)_sample': age_weeks,
                'Name_subject': self._safe_get(row, 'Animal_Name'),
                'Sex': self._safe_get(row, 'Sex'),
                'Line_subject': self._safe_get(row, 'Line_animal', 'Line_sample', 'Line'),
                'Line (Short)': self._safe_get(row, 'Line (Short)'),
                'Line (Stock)': self._safe_get(row, 'Line (Stock)'),
                'Species_subject': 'Mouse' if self._safe_get(row, 'Sex') != '' else '',
                'Genotype': genotype_to_symbol(self._safe_get(row, 'Genotype')),
                'Birth Date': self._safe_get(row, 'Birth Date'),
                'Wean Date': self._safe_get(row, 'Wean Date'),
                'Dissect Date': self._safe_get(row, 'Harvest Date')
            }
            
            tracker_data.append(data)
        
        df = pd.DataFrame(tracker_data)
        
        if df.empty:
            df = pd.DataFrame(columns=[
                'Name_sample', 'Age (weeks)_sample', 'Name_subject', 'Sex', 
                'Line_subject', 'Line (Short)', 'Line (Stock)', 'Species_subject', 
                'Genotype', 'Birth Date', 'Wean Date', 'Dissect Date'
            ])
        
        column_order = [
            'Name_sample', 'Age (weeks)_sample', 'Name_subject', 'Sex', 
            'Line_subject', 'Line (Short)', 'Line (Stock)', 'Species_subject', 
            'Genotype', 'Birth Date', 'Wean Date', 'Dissect Date'
        ]
        df = df[column_order]
        
        self._apply_sheet_styling(ws, df, column_order)
        
        print(f"  ✓ Created MERFISH Sample Tracker sheet with {len(df)} rows (OCT Block only)")
        return ws
    
    def create_rnaseq_sample_tracker_sheet(self):
        """
        Create the RNASeq Sample Tracker sheet
        FILTERS: Only samples with Preservation = "Flash Frozen" or "Frozen"
        """
        sheet_name = "RNASeq Sample Tracker"  # Added space between Sample and Tracker
        ws = self.workbook.create_sheet(sheet_name)
        
        print("\nCreating RNASeq Sample Tracker sheet...")
        
        # Filter for only "Flash Frozen" or "Frozen" samples
        filtered_df = self.merged_df.copy()
        if 'Preservation' in filtered_df.columns:
            filtered_df = filtered_df[filtered_df['Preservation'].isin(['Flash Frozen', 'Frozen'])]
            print(f"  Filtered to {len(filtered_df)} samples with Preservation = 'Flash Frozen' or 'Frozen'")
        else:
            print(f"  ⚠ Warning: 'Preservation' column not found, including all samples")
        
        tracker_data = []
        
        for idx, row in filtered_df.iterrows():
            age_weeks = self._calculate_age_weeks(
                self._safe_get(row, 'Birth Date'), 
                self._safe_get(row, 'Harvest Date')
            )
            
            data = {
                'Name_sample': self._safe_get(row, 'Sample_Name', 'Name_sample'),
                'Age (weeks)_sample': age_weeks,
                'Name_subject': self._safe_get(row, 'Animal_Name'),
                'Sex': self._safe_get(row, 'Sex'),
                'Line_subject': self._safe_get(row, 'Line_animal', 'Line_sample', 'Line'),
                'Line (Short)': self._safe_get(row, 'Line (Short)'),
                'Line (Stock)': self._safe_get(row, 'Line (Stock)'),
                'Species_subject': 'Mouse' if self._safe_get(row, 'Sex') != '' else '',
                'Genotype': genotype_to_symbol(self._safe_get(row, 'Genotype')),
                'Birth Date': self._safe_get(row, 'Birth Date'),
                'Wean Date': self._safe_get(row, 'Wean Date'),
                'Dissect Date': self._safe_get(row, 'Harvest Date')
            }
            
            tracker_data.append(data)
        
        df = pd.DataFrame(tracker_data)
        
        if df.empty:
            df = pd.DataFrame(columns=[
                'Name_sample', 'Age (weeks)_sample', 'Name_subject', 'Sex', 
                'Line_subject', 'Line (Short)', 'Line (Stock)', 'Species_subject', 
                'Genotype', 'Birth Date', 'Wean Date', 'Dissect Date'
            ])
        
        column_order = [
            'Name_sample', 'Age (weeks)_sample', 'Name_subject', 'Sex', 
            'Line_subject', 'Line (Short)', 'Line (Stock)', 'Species_subject', 
            'Genotype', 'Birth Date', 'Wean Date', 'Dissect Date'
        ]
        df = df[column_order]
        
        self._apply_sheet_styling(ws, df, column_order)
        
        print(f"  ✓ Created RNASeq Sample Tracker sheet with {len(df)} rows (Flash Frozen/Frozen only)")
        return ws
    
    def _apply_sheet_styling(self, ws, df, column_order):
        """Apply consistent styling to a worksheet"""
        # Write headers with styling
        for col_num, header in enumerate(column_order, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = Font(bold=True, size=12, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        
        # Write data
        for row_num, row_data in enumerate(df.values, 2):
            for col_num, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col_num, value=value)
                cell.alignment = Alignment(horizontal="left", vertical="center")
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                
                # Format dates
                if 'Date' in column_order[col_num-1] and value:
                    try:
                        if pd.notna(value) and value != '':
                            cell.number_format = 'MM/DD/YYYY'
                    except:
                        pass
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def create_all_sheets(self):
        """Create all sheets at once"""
        print("\n" + "="*80)
        print("Creating Excel Export with Multiple Sheets")
        print("="*80)
        
        self.create_sing_harvest_sheet()
        self.create_animal_sample_tracking_sheet()
        self.create_merfish_sample_tracker_sheet()
        self.create_rnaseq_sample_tracker_sheet()
        
        print("\n" + "="*80)
    
    def save(self):
        """Save the workbook to file"""
        self.workbook.save(self.output_filename)
        print(f"\n✅ Excel file saved as: {self.output_filename}")
        return self.output_filename


def main():
    """Main function to run the exporter"""
    animals_csv = "animals.csv"
    samples_csv = "samples.csv"
    
    print("="*80)
    print("LAB DATA EXPORT - MULTI-SHEET EXCEL GENERATOR")
    print("="*80)
    print(f"\nWorking directory: {os.getcwd()}\n")
    
    if not os.path.exists(samples_csv):
        print(f"❌ Error: {samples_csv} not found!")
        input("\nPress Enter to exit...")
        return False
    
    try:
        exporter = MultiSheetExporter(
            animals_csv=animals_csv if os.path.exists(animals_csv) else None,
            samples_csv=samples_csv
        )
        
        exporter.create_all_sheets()
        saved_file = exporter.save()
        
        print("\n" + "="*80)
        print("📊 EXPORT COMPLETE!")
        print("="*80)
        print("\nSheets created:")
        print("  1. Sing Harvest Sheet (all samples)")
        print("  2. Animal and Sample Tracking (4% PFA Fixed only)")
        print("  3. MERFISH Sample Tracker (OCT Block only)")
        print("  4. RNASeq Sample Tracker (Flash Frozen/Frozen only)")
        print(f"\n📁 File location: {os.path.abspath(saved_file)}")
        print("="*80 + "\n")
        
        return True
        
    except KeyboardInterrupt:
        print("\n\n❌ Process cancelled by user (Ctrl+C)")
        input("\nPress Enter to exit...")
        return False
        
    except Exception as e:
        print(f"\n❌ ERROR OCCURRED:")
        print(f"Error type: {type(e).__name__}")
        print(f"Error message: {str(e)}")
        print("\nFull traceback:")
        traceback.print_exc()
        input("\nPress Enter to exit...")
        return False


if __name__ == "__main__":
    success = main()
    
    if success:
        input("\n✓ Press Enter to exit...")
    
    sys.exit(0 if success else 1)